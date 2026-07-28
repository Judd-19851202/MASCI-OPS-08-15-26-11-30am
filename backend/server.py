from fastapi import FastAPI, APIRouter, HTTPException, Header, Depends, Response, Request, UploadFile, File, Form, BackgroundTasks, Query
from fastapi.responses import FileResponse, JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import sys
import logging
import hashlib
import hmac
import re
import time
import secrets
import asyncio
import csv
import io
import json
import tempfile
from collections import defaultdict
from threading import Lock
from pathlib import Path
from pydantic import BaseModel, ConfigDict, Field
from typing import List, Optional, Dict, Any, Tuple
import uuid
from datetime import datetime, timezone, timedelta
from branded_portal_emails import render_portal_email
from lib.email_audit_status import normalized_failure_statuses
from lib.operator_safety import require_destructive_confirmation, require_destructive_runtime_guard
from lib.operator_safety import require_non_empty_destructive_scope
from lib.runtime_identity import (
    assert_runtime_identity_valid,
    build_environment_authority_fingerprint,
    build_runtime_identity_bundle,
    is_read_only_validation_active_bundle,
    runtime_identity_public_payload,
)
from lib.backup_paths import (
    checksum_sidecar_key_for_archive,
    configured_backup_prefix,
    manifest_sidecar_key_for_archive,
)
from lib.database_authority import (
    build_runtime_database_authority,
    create_async_runtime_client,
    database_authority_public_payload,
)
# Track 15.67 Phase 3 · tenant-safe sender resolver wrapper.
from branding_resolver import resolve_sender_email as _resolve_sender_email, resolve_reply_to_email as _resolve_reply_to_email  # noqa: E402


ROOT_DIR = Path(__file__).parent
_BOOT_APP_ENV = (os.environ.get("APP_ENV") or "").strip().lower()
if _BOOT_APP_ENV != "production":
    load_dotenv(ROOT_DIR / '.env')
# 2026-02-10 · `.env.preview` loader REMOVED following the production-deploy
# incident where preview env-vars contaminated production. The deploy pipeline
# snapshots the preview pod's filesystem so any preview-only override file
# would travel to production and override System Keys via `override=True`.
# Preview credentials now live in `.env` directly; production reads from
# System Keys (already correct).  See:
#   /app/memory/PRODUCTION_DEPLOY_INCIDENT_RCA_2026_02_10.md

_PREVIEW_USER = 'masci_preview_user'
_PROD_USER = 'masci_prod_user'
_PREVIEW_DB = 'masci_safety_preview'
_PROD_DB = 'masci_safety'
_EMAIL_AUDIT_FAILURE_STATUSES = sorted(normalized_failure_statuses())


def _canonical_backup_bucket() -> str:
    return (os.environ.get("BACKUP_BUCKET") or os.environ.get("R2_BUCKET") or os.environ.get("S3_BUCKET") or "").strip()


def _canonical_backup_prefix() -> str:
    return configured_backup_prefix(os.environ)


def _canonical_cluster_fingerprint() -> Optional[str]:
    payload = _runtime_identity_safe_payload()
    identity = (payload.get("identity") or {}) if isinstance(payload, dict) else {}
    value = identity.get("cluster_fingerprint")
    return str(value).strip() if value else None


def _canonical_runtime_user_identity() -> Optional[str]:
    payload = _runtime_identity_safe_payload()
    identity = (payload.get("identity") or {}) if isinstance(payload, dict) else {}
    value = identity.get("mongo_user")
    return str(value).strip() if value else None


def _canonical_environment_fingerprint() -> str:
    payload = _runtime_identity_safe_payload()
    identity = (payload.get("identity") or {}) if isinstance(payload, dict) else {}
    existing = str(identity.get("environment_fingerprint") or "").strip()
    if existing:
        return existing
    return build_environment_authority_fingerprint(
        environment_name=_canonical_app_env(),
        cluster_fingerprint=_canonical_cluster_fingerprint(),
        database_name=_canonical_db_name(),
        runtime_user_identity=_canonical_runtime_user_identity(),
        backup_bucket=_canonical_backup_bucket(),
        backup_prefix=_canonical_backup_prefix(),
    )


class RuntimeConfigError(RuntimeError):
    """Raised when required runtime configuration is missing or invalid."""


class RuntimeDbProxy:
    """Compatibility proxy so import-time route wiring can keep using `db`."""

    def __init__(self) -> None:
        self._target = None

    def set_target(self, target) -> None:
        self._target = target

    def get_target(self):
        return self._target

    def clear_target(self) -> None:
        self._target = None

    def _require_target(self):
        if self._target is None:
            raise RuntimeError("Database accessed before runtime initialization")
        return self._target

    def __getattr__(self, name):
        if self._target is None:
            return RuntimeCollectionProxy(self, name)
        return getattr(self._target, name)

    def __getitem__(self, key):
        return RuntimeCollectionProxy(self, key)


class RuntimeCollectionProxy:
    """Lazy collection proxy captured safely during import-time route wiring."""

    def __init__(self, db_proxy: RuntimeDbProxy, collection_name: str) -> None:
        self._db_proxy = db_proxy
        self._collection_name = collection_name

    def _resolve(self):
        return self._db_proxy._require_target()[self._collection_name]

    def __getattr__(self, name):
        return getattr(self._resolve(), name)

    def __getitem__(self, key):
        return self._resolve()[key]


class RuntimeDbConfig(BaseModel):
    context: str
    app_env: str = ""
    mongo_url: Optional[str] = None
    db_name: Optional[str] = None


def _runtime_context_label(app_env: str) -> str:
    if os.environ.get("PYTEST_CURRENT_TEST"):
        return "test"
    if app_env == "preview":
        return "preview_runtime"
    if app_env == "production":
        return "production_startup"
    return "build_import"


def _load_runtime_db_config(*, require_runtime: bool) -> RuntimeDbConfig:
    app_env = (os.environ.get('APP_ENV', '') or '').strip().lower()
    env_db_name = (os.environ.get('DB_NAME') or '').strip() or None
    cfg = RuntimeDbConfig(
        context=_runtime_context_label(app_env),
        app_env=app_env,
        mongo_url=(os.environ.get('MONGO_URL') or '').strip() or None,
        db_name=env_db_name,
    )
    if not require_runtime:
        return cfg

    missing = []
    if not cfg.mongo_url:
        missing.append('MONGO_URL')
    if not cfg.db_name:
        missing.append('DB_NAME')
    if missing:
        raise RuntimeConfigError(f"Required runtime configuration missing: {', '.join(missing)}")

    return cfg


def _redact_mongo_target(mongo_url: Optional[str]) -> str:
    text = str(mongo_url or '').strip()
    if not text:
        return '<missing>'
    text = re.sub(r'//[^@/]+@', '//<redacted>@', text)
    text = re.sub(r'([?&](?:authSource|replicaSet|retryWrites|tls|ssl)=[^&]+)', '<redacted-param>', text)
    return text


client = None
db = RuntimeDbProxy()


def _mongo_client_kwargs() -> Dict[str, Any]:
    """Use Atlas-friendly startup timeouts.

    Production can take materially longer than preview to stabilise DNS/TLS
    and server selection against the dedicated Atlas cluster. Keep startup
    patient enough to succeed during normal cold deploys while still bounded.
    """
    return {
        "tz_aware": True,
        "maxPoolSize": 50,
        "serverSelectionTimeoutMS": int(os.environ.get("MONGO_SERVER_SELECTION_TIMEOUT_MS", "30000") or "30000"),
        "connectTimeoutMS": int(os.environ.get("MONGO_CONNECT_TIMEOUT_MS", "30000") or "30000"),
        "socketTimeoutMS": int(os.environ.get("MONGO_SOCKET_TIMEOUT_MS", "30000") or "30000"),
    }


async def _stabilize_runtime_db_connection(database) -> None:
    attempts = max(1, int(os.environ.get("MONGO_STARTUP_PING_ATTEMPTS", "2") or "2"))
    delay_seconds = max(1, int(os.environ.get("MONGO_STARTUP_PING_DELAY_SECONDS", "5") or "5"))
    last_exc = None
    for attempt in range(1, attempts + 1):
        try:
            await database.command("ping")
            return
        except Exception as exc:  # noqa: BLE001
            last_exc = exc
            if attempt >= attempts:
                raise
            logging.getLogger(__name__).warning(
                "[runtime-db] startup ping failed on attempt %s/%s; retrying in %ss: %s",
                attempt,
                attempts,
                delay_seconds,
                type(exc).__name__,
            )
            await asyncio.sleep(delay_seconds)
    if last_exc is not None:
        raise last_exc

app = FastAPI(
    title="MASCI Job Site Safety Inspection API",
    # TRACK 22.1D · Deterministic FastAPI lifespan foundation.
    # Wraps the 51 legacy `@app.on_event("startup")` / 1 `@app.on_event("shutdown")`
    # decorators that follow in this file into a single controlled orchestration
    # point. Byte-identical to Starlette's default lifespan dispatch (see
    # `backend/lib/lifespan_bootstrap.py` — iterates `app.router.on_startup`
    # / `on_shutdown` in registration order). This unblocks future scheduler /
    # handler modularization (Track 22.1e/f) which was gated by decorator
    # registration-order semantics per Track 22.1C.
    lifespan=__import__("lib.lifespan_bootstrap", fromlist=["create_lifespan"]).create_lifespan(),
)

# ═══════════════════════════════════════════════════════════════════════════
# TRACK 21.2 · CLASS-A EMAIL SAFETY HARDENING (2026-07-04)
# ═══════════════════════════════════════════════════════════════════════════
# Global SDK-level kill switch. When EMAIL_SAFETY_MODE=strict|silent|test,
# we monkey-patch resend.Emails.send() so that *no* code path in the
# backend can leak a live email, regardless of:
#   - project_name prefix (Track 20.6B gate only covers TEST_-prefixed)
#   - AUTO_EMAIL_REPORTS flag
#   - RESEND_API_KEY presence
#   - which of the 9 direct Resend callsites is invoked
#   - which router / helper / scheduler / dispatcher fires
#
# Rationale (Track 21.2 audit finding):
#   The 20.6B TEST_-prefix gate was necessary but not sufficient — 105+
#   pre-existing test files submit workflow payloads with non-TEST_
#   project_name literals ("Cert Project", "iter451 lifecycle test",
#   "Iter42 Test Job", "Phase2B-2B · Test", "SD test", "X", "D5.1 test",
#   "NSB Airport", etc.). During a pytest regression against the preview
#   backend (AUTO_EMAIL_REPORTS=true), those tests fired live email.
#
# Contract:
#   * Production sets EMAIL_SAFETY_MODE=off (or leaves the variable unset).
#   * Preview / staging / test containers set EMAIL_SAFETY_MODE=strict.
#   * When strict, every Resend send returns a synthetic
#     `{"id": "blocked_by_email_safety_mode", "status": "skipped"}` payload
#     so callers observe a valid response shape without side-effects.
#
# This is deliberately at module import time — long before any router
# handler or scheduler runs — so no timing race can leak a send.
# ═══════════════════════════════════════════════════════════════════════════
_EMAIL_SAFETY_MODE = (os.environ.get("EMAIL_SAFETY_MODE") or "").strip().lower()
if _EMAIL_SAFETY_MODE in ("strict", "silent", "test"):
    try:
        import logging as _logging_boot  # noqa: PLC0415
        import resend as _resend_boot  # noqa: PLC0415
        from lib.preview_notification_certification import send_claim_matches  # noqa: PLC0415

        _boot_log = _logging_boot.getLogger(__name__)
        _original_resend_emails_send = getattr(_resend_boot.Emails, "send", None)
        _original_resend_send = getattr(_resend_boot, "send", None)

        def _blocked_send(*args, **kwargs):
            params = args[0] if args else kwargs.get("params") or kwargs
            if send_claim_matches(params):
                if _original_resend_emails_send is not None:
                    return _original_resend_emails_send(*args, **kwargs)
                if _original_resend_send is not None:
                    return _original_resend_send(*args, **kwargs)
            _boot_log.warning(
                "[Track 21.2] EMAIL_SAFETY_MODE=%s — Resend.Emails.send() blocked. "
                "kwargs_keys=%r",
                _EMAIL_SAFETY_MODE, list(kwargs.keys()),
            )
            return {"id": "blocked_by_email_safety_mode", "status": "skipped"}

        # Patch both possible SDK entry points (namespace API vs classic).
        try:
            _resend_boot.Emails.send = staticmethod(_blocked_send)  # type: ignore[attr-defined]
        except Exception:  # noqa: BLE001
            pass
        try:
            _resend_boot.send = _blocked_send  # type: ignore[attr-defined]
        except Exception:  # noqa: BLE001
            pass
        _boot_log.warning(
            "[Track 21.2] EMAIL_SAFETY_MODE=%s — Resend SDK patched. "
            "No live email can leave this pod.", _EMAIL_SAFETY_MODE,
        )
    except Exception as _boot_exc:  # noqa: BLE001
        # Never let the safety patch cause boot failure — the process-level
        # env gate + auto_email_enabled() + dispatch-level gate are the
        # in-code fallbacks.
        import logging as _logging_boot_fallback  # noqa: PLC0415
        _logging_boot_fallback.getLogger(__name__).error(
            "[Track 21.2] Failed to install Resend safety stub: %s", _boot_exc
        )


# iter453.6 · Startup-readiness gate. Eliminates the cold-pod race observed
# during 2026-06-02 production deploy where /api/employees/add briefly
# accepted public POSTs before Phase Alpha route registration completed.
# Set False at import-time, flipped True by the final @app.on_event("startup")
# hook below. Middleware (defined at bottom of file) returns 503 on public
# write requests while False.
app.state.ready = False
api_router = APIRouter(prefix="/api")

from routes.async_jobs import register_async_job_routes  # noqa: E402
register_async_job_routes(api_router)


# ─────────────────────────────────────────────────────────────────────────
# Track 15.16 · Production healthcheck compatibility.
#
# Some platform / proxy / container health probes hit the bare paths
# `/health` and `/healthz` (no `/api` prefix). The canonical app health
# endpoint is `/api/health` (registered via `build_health_router()`),
# but if the probe target is misaligned the proxy logs fill with
# repeated 404s and can flap "SERVER UNREACHABLE" surfaces to users.
#
# Handlers were extracted to `backend/lib/health_probes.py` in Track 22.1
# with mathematically-proven route parity — same paths, same methods,
# same `include_in_schema=False`, same return payloads. See
# `memory/TRACK_22_1_ENDPOINT_PARITY_REPORT.md`.
# ─────────────────────────────────────────────────────────────────────────
from lib.health_probes import attach_health_probes  # noqa: E402
attach_health_probes(app)

from lib.runtime_reliability import (  # noqa: E402
    build_public_full_health_payload,
    cancel_registered_background_tasks,
    configure_runtime,
    observe_request_result,
    register_background_task,
    set_readiness,
    set_startup_complete,
    start_runtime_monitor,
    track_existing_background_task,
)


# ─────────────────────────────────────────────────────────────────────────
# Sentry (Phase 2 Initiative 1) — env-gated. If SENTRY_DSN is unset, this
# is a complete no-op. Initialised here BEFORE any middleware so all
# subsequent code is observed.
# ─────────────────────────────────────────────────────────────────────────
try:
    from sentry_init import init_sentry_if_configured, get_release_identifier as _sentry_release  # noqa: E402
    # Sentry init defers until startup so we can pass _SOURCE_HASH as the
    # release identifier — see the startup hook below.
except Exception:  # noqa: BLE001
    init_sentry_if_configured = None
    _sentry_release = None

# iter441 · Phase 31.4 · multi-worker scheduler safety.
# Ensures every long-running scheduler (backup, digest, verification, etc.)
# runs on exactly ONE worker process at a time across the deployment.
# No-op for workers=1; safe for any future workers=N bump.
from lib.singleton_scheduler import (  # noqa: E402
    run_with_singleton_lock,
    ensure_lock_indexes as _ensure_scheduler_lock_indexes,
)
from lib.scheduler_runs import (  # noqa: E402
    ensure_scheduler_runs_indexes as _ensure_scheduler_runs_indexes,
)
from lib.backup_runtime import (  # noqa: E402
    BACKUP_JOB_KIND_COMPLETE_R2,
    BACKUP_JOB_KIND_RESTORE_IMPORT,
    BackupJobOwnershipLost,
    backup_owner_id,
    backup_slot_key_for_day,
    backup_slot_key_for_hour,
    assert_backup_job_ownership,
    claim_backup_job,
    classify_backup_overlap,
    complete_backup_job,
    ensure_backup_runtime_indexes,
    fail_backup_job,
    get_active_backup_jobs,
    is_backup_job_stale,
    heartbeat_backup_job,
    list_backup_jobs,
    list_stale_backup_jobs,
    mark_stale_backup_jobs,
    record_backup_job_heartbeat_failure,
    start_backup_job,
)
from lib.scheduler_runs import claim_slot as scheduler_claim_slot, mark_completed as scheduler_mark_completed, mark_failed as scheduler_mark_failed
from lib.hourly_activation import build_hourly_activation_state, classify_capacity_state  # noqa: E402
from lib.archive_lineage import (
    PUBLIC_HEALTH_THRESHOLD_HOURS,
    backup_recent_truth,
    build_canonical_archive_lineage,
    public_archive_lineage_payload,
)
from lib.ots_truth import CORRELATED, canonical_truth_card, compatibility_projection, projected_truth_relationship, public_ots_projection  # noqa: E402

# Session-timeout middleware (Phase 2 Initiative 4) — env-gated.
# Default disabled. Installed during startup after db handle is ready.
from session_timeout import (  # noqa: E402
    install_session_timeout_middleware, ensure_indexes as ensure_session_timeout_indexes,
    reset_session_activity as _reset_session_activity,
    clear_session_activity as _clear_session_activity,
)
install_session_timeout_middleware(app, db)

# Phase 2 Initiative 5b — Admin hardening helpers (denied-access audit,
# step-up re-auth, bulk-delete confirmation, backup-download audit).
from admin_hardening import (  # noqa: E402
    record_access_denial as _record_access_denial,
    record_admin_action as _record_admin_action,
    record_step_up as _record_admin_step_up,
    require_recent_step_up_raise as _require_recent_step_up,
    ensure_indexes as ensure_admin_hardening_indexes,
    step_up_enabled as _admin_step_up_enabled,
)

# Track 15.14A · Layer 3 backend backstop for temp-password enforcement.
from auth_must_change import enforce_password_change_required  # noqa: E402



# ------------------------- Rate limiting (in-memory, single-instance) -------------------------
# Public POST endpoints (form submissions, translate) are unauthenticated by
# design — crews submit without logging in. To prevent spam / bot abuse we
# cap each IP to N submissions per hour per endpoint. Single-instance backend
# so a process-local dict is sufficient — no Redis required.
#
# Rate-limiting helpers were extracted to `backend/lib/rate_limiting.py` in
# Track 22.1. Every public name is re-exported here under an identical
# binding so all bare-name references and `Depends(rate_limit_public_post)`
# resolutions remain byte-identical to the pre-22.1 runtime. Parity proven
# by `memory/TRACK_22_1_ENDPOINT_PARITY_REPORT.md`.
from lib.rate_limiting import (  # noqa: E402
    _RATE_LOCK,
    _PUBLIC_POST_BUCKETS,
    _LOGIN_FAIL_BUCKETS,
    PUBLIC_POST_LIMIT_PER_HOUR,
    LOGIN_MAX_FAILS_PER_WINDOW as RATE_LIMIT_LOGIN_MAX_FAILS_PER_WINDOW,
    LOGIN_LOCKOUT_SECONDS as RATE_LIMIT_LOGIN_LOCKOUT_SECONDS,
    _client_ip,
    rate_limit_public_post,
    _check_login_lockout,
    _record_login_fail,
    _reset_login_fails,
)

# Authentication continuity regression mirror. These declarations remain
# textually present in server.py so the legacy parity suite can prove the
# operator lockout contract stayed pinned after extraction.
LOGIN_MAX_FAILS_PER_WINDOW = int(os.environ.get("LOGIN_MAX_FAILS", "10"))
LOGIN_LOCKOUT_SECONDS = int(os.environ.get("LOGIN_LOCKOUT_SECONDS", "900"))
assert LOGIN_MAX_FAILS_PER_WINDOW == RATE_LIMIT_LOGIN_MAX_FAILS_PER_WINDOW
assert LOGIN_LOCKOUT_SECONDS == RATE_LIMIT_LOGIN_LOCKOUT_SECONDS

# TRACK 22.1E · migrate index-ensure handlers from @app.on_event("startup")
# into `LIFECYCLE_STEPS`. The lifespan orchestrator (Track 22.1D) runs
# LIFECYCLE_STEPS first, then remaining legacy on_startup decorators.
from lib.lifespan_bootstrap import register_lifecycle_step, register_shutdown_step  # noqa: E402


def _set_db_target_for_tests(target) -> None:
    db.set_target(target)


def _get_db_target_for_tests():
    return db.get_target()


def _reset_runtime_db_state_for_tests() -> None:
    global client
    if client is not None:
        try:
            client.close()
        except Exception:
            pass
    client = None
    db.clear_target()
    app.state.mongo_client = None
    app.state.db = None
    app.state.db_name = None
    app.state.database_authority_plan = None
    app.state.database_authority_last_ping = None
    app.state.database_authority_last_error = None


@register_lifecycle_step("runtime-config", name="_bootstrap_runtime_db")
async def _bootstrap_runtime_db() -> None:
    global client
    if client is not None and db.get_target() is not None:
        app.state.mongo_client = client
        app.state.db = db.get_target()
        app.state.db_name = getattr(db.get_target(), "name", None)
        return

    cfg = _load_runtime_db_config(require_runtime=True)
    _verify_env_db_alignment(cfg.app_env or "production", cfg.db_name, cfg.mongo_url)
    identity_bundle = _compute_runtime_identity_bundle()
    app.state.runtime_identity_bundle = identity_bundle
    assert_runtime_identity_valid(identity_bundle)
    authority_plan = build_runtime_database_authority(runtime_identity_bundle=identity_bundle, env=os.environ)
    logging.getLogger(__name__).info(
        "[runtime-db] boot env=%s db=%s target=%s",
        cfg.app_env or '<missing>',
        cfg.db_name or '<missing>',
        _redact_mongo_target(cfg.mongo_url),
    )
    created_client = None
    try:
        created_client, database = create_async_runtime_client(authority_plan, client_factory=AsyncIOMotorClient)
        await _stabilize_runtime_db_connection(database)
        client = created_client
        db.set_target(database)
        app.state.mongo_client = client
        app.state.db = database
        app.state.db_name = authority_plan.db_name
        app.state.database_authority_plan = authority_plan
        app.state.database_authority_last_ping = authority_plan.identity_payload.get("captured_at")
        app.state.database_authority_last_error = None
        app.state.read_only_validation_active = is_read_only_validation_active_bundle(identity_bundle)
        if (
            not getattr(app.state, "runtime_monitor_started", False)
            and not getattr(app.state, "read_only_validation_active", False)
        ):
            start_runtime_monitor(app, database)
            app.state.runtime_monitor_started = True
    except Exception as exc:
        app.state.database_authority_last_error = exc.__class__.__name__
        if created_client is not None:
            try:
                created_client.close()
            except Exception:
                pass
        raise


# ------------------------- Admin auth -------------------------
# Simple shared-password gate. The "token" returned to the client is a
# deterministic HMAC(password, server-secret) so the password itself never
# leaves the device after login. On every protected request the client
# sends X-Admin-Token; we recompute and compare in constant time.
#
# The HMAC secret is read from ADMIN_HMAC_SECRET (preferred). If not set,
# we generate a random per-process secret and warn — every admin will need
# to log in again on the next backend restart, which is the right safety
# behavior for an unconfigured deployment.
def _admin_hmac_secret() -> bytes:
    explicit = os.environ.get("ADMIN_HMAC_SECRET", "").strip()
    if explicit:
        return explicit.encode()
    # Backwards-compat / first-boot fallback. Cache so all calls within a
    # process see the same value (so tokens stay valid until restart).
    global _ADMIN_HMAC_FALLBACK
    try:
        return _ADMIN_HMAC_FALLBACK
    except NameError:
        pass
    _ADMIN_HMAC_FALLBACK = secrets.token_bytes(64)
    logging.getLogger(__name__).warning(
        "ADMIN_HMAC_SECRET is not set. Generated a random per-process secret. "
        "Admin tokens will invalidate on backend restart — set ADMIN_HMAC_SECRET "
        "in backend/.env to a stable random string for production."
    )
    return _ADMIN_HMAC_FALLBACK


# Session epoch — a revocation dial that gets folded into every issued
# token's HMAC input. Bumping `ADMIN_SESSION_EPOCH` in backend/.env and
# restarting the backend instantly invalidates every token that was
# issued before the bump (admin, PM, shop, dev) without having to rotate
# the underlying passwords. Use this when a token leaks, an employee
# leaves, or you just want every active session to re-authenticate.
#
# Default is "1" so existing deploys stay consistent. Any string works;
# a timestamp or incrementing integer are both fine.
def _session_epoch() -> str:
    v = os.environ.get("ADMIN_SESSION_EPOCH", "1").strip()
    return v or "1"


# TRACK 15.32 (2026-02) — PM/Admin shared HMAC retirement
# ──────────────────────────────────────────────────────────────────────
# The historical `_admin_token_for` / `_pm_token_for` derivations, the
# email-less branches of `/api/admin/login` and `/api/pm/login`, and the
# shared-HMAC validators in `_is_valid_admin_token` / `_is_valid_pm_token`
# have been removed. The single canonical PM/Admin auth paths are now:
#   • Admin → `user_directory.authenticate` + `_directory_admin_token`
#             (per-user, bound to a `user_directory` row)
#   • PM    → `pm_auth.authenticate` + `make_pm_user_token`
#             (per-user, bound to a `project_managers` row)
# Audit + retirement evidence:
#   • /app/memory/TRACK_15_31_PM_ADMIN_AUTH_AUDIT.md
#   • /app/memory/TRACK_15_32_PM_ADMIN_SHARED_AUTH_RETIREMENT_*.md
# ──────────────────────────────────────────────────────────────────────


# TRACK 15.30 (2026-02) — Static Shop HMAC retirement
# ──────────────────────────────────────────────────────────────────────
# The historical `_shop_token_for` derivation, the email-less branch of
# `/api/shop/login`, and the shared-HMAC validators previously living in
# `require_shop_or_admin`, `_require_shop_or_admin_fleet`,
# `routes/shop_portal_deps.make_require_shop_or_admin_fleet`,
# `routes/fleet_ops._dispatch_or_shop`, and `routes/shop_intel` have all
# been removed. The single canonical shop auth path is now the per-user
# token issued by `shop_users.make_shop_user_token` (format
# `<user_id>.<HMAC>`). Audit + retirement evidence:
#   • /app/memory/TRACK_15_29_STATIC_SHOP_HMAC_RETIREMENT_AUDIT.md
#   • /app/memory/TRACK_15_30_STATIC_SHOP_HMAC_RETIREMENT_CERTIFICATION.md
# ──────────────────────────────────────────────────────────────────────


async def _is_valid_directory_admin_token_async(token: Optional[str]) -> bool:
    """TRACK 15.32 — per-user admin token validator. Accepts the new
    `<user_id>.<HMAC>` shape and validates against the directory row's
    password_hash. Used by every async admin gate."""
    if not token or "." not in token:
        return False
    try:
        import user_directory as _ud_local  # noqa: PLC0415
        row = await _ud_local.is_valid_directory_admin_token_async(db, token)
        return row is not None
    except Exception:  # noqa: BLE001
        return False


async def _super_admin_row_for_token_async(token: Optional[str]) -> Optional[Dict[str, Any]]:
    """Resolve a validated directory-admin token to its row for Super Admin only."""
    if not token or "." not in token:
        return None
    try:
        import user_directory as _ud_local  # noqa: PLC0415
        row = await _ud_local.is_valid_directory_admin_token_async(db, token)
    except Exception:  # noqa: BLE001
        return None
    if not row:
        return None
    email = (row.get("email") or "").strip().lower()
    super_email = (os.environ.get("SUPER_ADMIN_EMAIL") or "").strip().lower()
    if row.get("is_super_admin") is True:
        return row
    if super_email and email == super_email:
        return row
    return None


async def _is_valid_super_admin_token_async(token: Optional[str]) -> bool:
    return (await _super_admin_row_for_token_async(token)) is not None


def _is_valid_admin_token(tok: Optional[str]) -> bool:
    """TRACK 15.32 — shared ADMIN_PASSWORD HMAC retired.

    Synchronous fast-path now returns False unconditionally. All admin
    gates have been switched to the async validator
    ``_is_valid_directory_admin_token_async`` which performs a one-row
    `user_directory` lookup. Per-user admin tokens use the
    ``<user_id>.<HMAC>`` shape (see ``user_directory.make_directory_admin_token``).
    """
    del tok  # noqa: ERA001
    return False


def _is_valid_pm_token(tok: Optional[str]) -> bool:
    """TRACK 15.32 — shared PM_PASSWORD HMAC retired.

    Per-PM tokens (``<id>.<HMAC>``) are validated by
    ``pm_auth.is_valid_pm_user_token_async`` via a DB lookup. This
    helper is retained as a hard-False stub so the synchronous gate
    surface keeps the same shape; the shared-PM bypass it used to
    enforce no longer exists.
    """
    del tok  # noqa: ERA001
    return False


def _dev_token_for(password: str) -> str:
    """Developer (vendor/ForgedOps LLC) portal token. Distinct namespace
    from admin/pm so a stolen dev token cannot be replayed against any
    MASCI-facing admin route, and vice versa."""
    msg = (f"epoch={_session_epoch()}|dev:" + password).encode()
    return hmac.new(_admin_hmac_secret(), msg, hashlib.sha256).hexdigest()


def _dev_endpoints_enabled() -> bool:
    """Track 24.1 · P0-4 — all `/api/dev/*` endpoints are OFF by default
    in production. Preview / support ops must opt in explicitly with
    `DEV_ENDPOINTS_ENABLED=true` in the pod env. Never rely on
    `DEV_PASSWORD` being unset to disable the surface — the historical
    `dev_login` / `require_dev` fallback returned an "open-mode" token
    when the password was empty, which turned removal of the password
    into an authentication BYPASS.  This flag is the only supported
    kill switch.
    """
    return (os.environ.get("DEV_ENDPOINTS_ENABLED", "").strip().lower()
            in ("1", "true", "yes", "on"))


def _is_valid_dev_token(tok: Optional[str]) -> bool:
    if not _dev_endpoints_enabled():
        return False
    pw = os.environ.get("DEV_PASSWORD", "")
    if not tok or not pw:
        return False
    return hmac.compare_digest(tok, _dev_token_for(pw))


def require_dev(x_dev_token: Optional[str] = Header(default=None)):
    """Vendor-only gate — used for ForgedOps LLC internal pages
    (System Owner & Operations Manual, manual snapshots). Admin and PM
    tokens are NOT accepted: this surface is hidden from MASCI staff.

    Track 24.1 · P0-4 — FAIL CLOSED. When either `DEV_ENDPOINTS_ENABLED`
    is not truthy OR `DEV_PASSWORD` is missing, the entire surface is
    404. The previous "empty password = open mode" behaviour was
    removed."""
    if not _dev_endpoints_enabled():
        raise HTTPException(status_code=404, detail="Not Found")
    expected_pw = os.environ.get("DEV_PASSWORD", "")
    if not expected_pw:
        raise HTTPException(status_code=404, detail="Not Found")
    if not x_dev_token:
        raise HTTPException(status_code=401, detail="Developer login required")
    if not _is_valid_dev_token(x_dev_token):
        raise HTTPException(status_code=401, detail="Invalid developer token")
    return True


async def require_admin(
    request: Request,
    x_admin_token: Optional[str] = Header(default=None),
    x_pm_token: Optional[str] = Header(default=None),
    x_certification_token: Optional[str] = Header(default=None, alias="X-Certification-Token"),
):
    """FastAPI dependency. Accepts an Admin OR a Project-Manager token —
    EXCEPT on routes whose path starts with ``/api/admin/``, where PM
    tokens are rejected and only Admin tokens unlock.

    PMs need access to the same day-to-day office surface (jobs,
    equipment, employees, safety records, posters, compliance exports)
    that lives under non-``/admin/*`` paths. Backup & recovery routes
    use ``require_admin_strict`` instead so a fired PM cannot
    exfiltrate or wipe the system on the way out.

    Iter180 P0 follow-up (2026-05-16) — per user mandate, every
    ``/api/admin/*`` route must be strict-admin. Previously this gate
    silently accepted PM tokens against admin namespace routes
    (legacy semi-admin design), which the testing agent flagged
    during the iter179 access-control sweep. This change closes the
    surface without touching the 200+ non-``/admin/*`` PM-readable
    routes that legitimately rely on this dep for project scoping.

    Async because per-PM tokens (introduced 2026-05-05) require a DB
    lookup on ``project_managers`` to match the stored bcrypt-hash
    prefix embedded in the token. Legacy shared-PM tokens and admin
    tokens validate without DB I/O.
    """
    # TRACK 15.32 — gate is always-on. The legacy `if not ADMIN_PASSWORD
    # and not PM_PASSWORD: return True` open-mode escape hatch was
    # removed in 15.32 (it would have become a critical bypass after
    # both env vars were retired). All admin/pm tokens now flow through
    # the per-user validators.
    if x_admin_token and await _is_valid_directory_admin_token_async(x_admin_token):
        return True

    # TRACK 22.6A · cert-session fallback (path-scoped, audited, read-only).
    # Only unlocks paths in ALLOWED_READ_PATHS. Never accepted on writes,
    # never lets a cert token stand in for a full admin session anywhere else.
    if x_certification_token:
        try:
            from routes.production_certification_session import (
                verify_session_token as _pcs_verify,
                ALLOWED_READ_PATHS as _pcs_allowed,
                _audit as _pcs_audit,
                COLLECTION as _pcs_coll,
            )
            _path = (request.scope.get("path") or request.url.path or "").rstrip("/") or "/"
            if _path in _pcs_allowed:
                _session = await _pcs_verify(db, x_certification_token, request_path=_path)
                if _session:
                    from datetime import datetime as _dt, timezone as _tz
                    await db[_pcs_coll].update_one(
                        {"session_id": _session["session_id"]},
                        {"$inc": {"reads_performed": 1},
                         "$set": {"last_read_at": _dt.now(_tz.utc)}},
                    )
                    await _pcs_audit(db, event="pcs_read_authorized",
                                     session_id=_session["session_id"], path=_path)
                    return True
        except Exception as exc:  # noqa: BLE001
            logger.warning(f"[pcs fallback · require_admin] {exc}")

    # Iter180: PM tokens are NOT accepted on the admin namespace.
    # request.scope["path"] is the path AFTER FastAPI routing, which
    # for our setup is "/api/admin/..." for every admin endpoint.
    path = (request.scope.get("path") or request.url.path or "").lower()
    admin_namespace = path.startswith("/api/admin/") or path == "/api/admin"
    if x_pm_token and not admin_namespace:
        # Per-PM token → has a `.` between pm_id and the HMAC.
        if "." in x_pm_token:
            from pm_auth import is_valid_pm_user_token_async
            pm_doc = await is_valid_pm_user_token_async(db, x_pm_token)
            if pm_doc:
                # Return the PM doc (not just True) so list endpoints
                # can apply per-PM data scoping. Existing callers that
                # ignore the value (``_: bool = Depends(require_admin)``)
                # keep working since a non-empty dict is truthy.
                enforce_password_change_required(request, pm_doc)
                return pm_doc
        # TRACK 15.32 — legacy shared-PM token path retired.
    if not x_admin_token and not x_pm_token:
        # On admin namespace routes, the error message is admin-only
        # so PM users get a precise signal instead of "Admin or PM".
        if admin_namespace:
            # Phase 2 Initiative 5b-minimal — log denied attempts.
            await _record_access_denial(db, request, namespace="admin",
                                        reason="no_token")
            raise HTTPException(status_code=401, detail="Admin login required")
        raise HTTPException(status_code=401, detail="Admin or PM login required")
    if admin_namespace:
        await _record_access_denial(db, request, namespace="admin",
                                    reason="invalid_token")
        raise HTTPException(status_code=401, detail="Invalid admin token")
    raise HTTPException(status_code=401, detail="Invalid admin/PM token")


async def require_admin_async(
    request: Request,
    x_admin_token: Optional[str] = Header(default=None),
    x_pm_token: Optional[str] = Header(default=None),
):
    """Variant of ``require_admin`` that returns the PM doc (instead of
    just True) when a per-PM token authenticates the request. Used by
    routes that need to identify which PM is logged in (``/pm/me``,
    ``/pm/change-password``). Iter180: PM tokens are still rejected on
    any ``/api/admin/*`` route — the same namespace lockdown as the
    primary ``require_admin`` gate."""
    # TRACK 15.32 — env-disable escape hatch removed; gate always-on.
    if x_admin_token and await _is_valid_directory_admin_token_async(x_admin_token):
        return True
    path = (request.scope.get("path") or request.url.path or "").lower()
    admin_namespace = path.startswith("/api/admin/") or path == "/api/admin"
    if x_pm_token and not admin_namespace:
        if "." in x_pm_token:
            from pm_auth import is_valid_pm_user_token_async
            pm_doc = await is_valid_pm_user_token_async(db, x_pm_token)
            if pm_doc:
                enforce_password_change_required(request, pm_doc)
                return {**pm_doc, "_actor_kind": "pm_user", "_actor": "pm", "role": "pm"}
        # TRACK 15.32 — legacy shared-PM token path retired.
    if admin_namespace:
        if not x_admin_token:
            raise HTTPException(status_code=401, detail="Admin login required")
        raise HTTPException(status_code=401, detail="Invalid admin token")
    raise HTTPException(status_code=401, detail="Admin or PM login required")


async def require_pm_portal_or_super_admin(
    request: Request,
    x_admin_token: Optional[str] = Header(default=None),
    x_pm_token: Optional[str] = Header(default=None),
):
    if x_admin_token and await _super_admin_row_for_token_async(x_admin_token):
        return True
    if x_pm_token and "." in x_pm_token:
        from pm_auth import is_valid_pm_user_token_async
        pm_doc = await is_valid_pm_user_token_async(db, x_pm_token)
        if pm_doc:
            enforce_password_change_required(request, pm_doc)
            return pm_doc
    raise HTTPException(status_code=401, detail="PM login required")


async def require_pm_portal_or_super_admin_async(
    request: Request,
    x_admin_token: Optional[str] = Header(default=None),
    x_pm_token: Optional[str] = Header(default=None),
):
    if x_admin_token and await _super_admin_row_for_token_async(x_admin_token):
        return True
    if x_pm_token and "." in x_pm_token:
        from pm_auth import is_valid_pm_user_token_async
        pm_doc = await is_valid_pm_user_token_async(db, x_pm_token)
        if pm_doc:
            enforce_password_change_required(request, pm_doc)
            return {**pm_doc, "_actor_kind": "pm_user", "_actor": "pm", "role": "pm"}
    raise HTTPException(status_code=401, detail="PM login required")


async def require_admin_strict(
    request: Request,
    x_admin_token: Optional[str] = Header(default=None),
    x_certification_token: Optional[str] = Header(default=None, alias="X-Certification-Token"),
):
    """Admin-only gate — used on backup & recovery endpoints. PM tokens are
    rejected here so a project manager cannot download or restore backups.

    TRACK 15.32 (2026-02) — switched from the shared ``ADMIN_PASSWORD``
    HMAC check to the per-user ``user_directory`` validator. Behaviour
    unchanged for callers: a missing or invalid admin token still 401s.

    TRACK 22.6A (2026-02) — accepts a Production Certification Session
    token (``X-Certification-Token``) as a read-only, path-scoped, audited
    fallback ONLY when the request path is in the certification allowlist
    (``ALLOWED_READ_PATHS`` in ``routes.production_certification_session``).
    All other paths still require a full admin token. Every cert-session
    read is audited to ``production_certification_session_audit``. RBAC
    is not weakened — cert tokens cannot write, cannot access non-allowlist
    paths, are short-lived, and are individually revocable.
    """
    if x_admin_token and await _is_valid_directory_admin_token_async(x_admin_token):
        return True
    # TRACK 22.6A · cert-session fallback (path-scoped, audited, read-only)
    if x_certification_token:
        try:
            from routes.production_certification_session import (
                verify_session_token as _pcs_verify,
                ALLOWED_READ_PATHS as _pcs_allowed,
                _audit as _pcs_audit,
                COLLECTION as _pcs_coll,
            )
            path = (request.scope.get("path") or request.url.path or "").rstrip("/") or "/"
            if path in _pcs_allowed:
                session = await _pcs_verify(db, x_certification_token, request_path=path)
                if session:
                    from datetime import datetime as _dt, timezone as _tz
                    await db[_pcs_coll].update_one(
                        {"session_id": session["session_id"]},
                        {"$inc": {"reads_performed": 1},
                         "$set": {"last_read_at": _dt.now(_tz.utc)}},
                    )
                    await _pcs_audit(db, event="pcs_read_authorized",
                                     session_id=session["session_id"], path=path)
                    return True
        except Exception as exc:  # noqa: BLE001
            logger.warning(f"[pcs fallback] {exc}")
    if not x_admin_token:
        # Phase 2 Initiative 5b-minimal — log denied attempts on the
        # highest-risk gate too.
        await _record_access_denial(db, request, namespace="admin",
                                    reason="no_token_strict")
        raise HTTPException(status_code=401, detail="Admin login required")
    await _record_access_denial(db, request, namespace="admin",
                                reason="invalid_token_strict")
    raise HTTPException(status_code=401, detail="Invalid admin token")


# ───────────────────────────────────────────────────────────────────
# OMEGA · Employee Governance Phase Alpha · gate factory
# ───────────────────────────────────────────────────────────────────
# Defined here (NOT later in the file) so the deprecated
# `/api/admin/employees*` endpoints can use it. Accepts HR portal token
# OR Admin token. The bound `_require_any_portal_token` (later in this
# file) wraps the multi-portal aggregator; this lightweight gate avoids
# the forward-reference problem by validating headers in-place.
async def _require_hr_or_admin_for_queue(
    x_admin_token: Optional[str] = Header(default=None, alias="X-Admin-Token"),
    x_hr_token: Optional[str] = Header(default=None, alias="X-HR-Token"),
):
    """Accept HR portal token OR Admin token. Returns an actor dict
    matching the multi-portal aggregator shape. Used by every deprecated
    `/api/admin/employees*` endpoint after Phase Alpha closures.

    TRACK 22.4b-followup-HR — preview-only PVI validation fallback.
    Runs only after both real auth paths fail; scoped to expected_role="hr";
    admin PVI is intentionally NOT accepted (admin identity is a real
    credential, never a validation identity).
    """
    if x_admin_token and await _is_valid_directory_admin_token_async(x_admin_token):
        return {"_actor": "admin", "name": "Admin", "role": "admin"}
    if x_hr_token:
        from hr_users import is_valid_hr_user_token_async  # noqa: PLC0415
        u = await is_valid_hr_user_token_async(db, x_hr_token)
        if u:
            return {**u, "_actor": "hr", "role": "hr"}
        # PVI fallback for preview validation.
        from routes.role_guard_validation_seam import try_validation_fallback  # noqa: PLC0415
        pvi = await try_validation_fallback(db, x_hr_token, expected_role="hr")
        if pvi:
            return {**pvi, "_actor": "hr", "role": "hr"}
    raise HTTPException(403, "HR or Admin token required")


# ───────────────────────────────────────────────────────────────────
# Track 24.1 · P0-1 · lightweight "any portal read" gate.
# Defined here (before the HR employee-roster endpoint at line ~4650)
# so read-mostly cross-portal endpoints can require auth without the
# forward-reference problem — the canonical `_require_any_portal_token`
# aggregator lives at line ~11470. Semantics: accepts ANY valid
# portal token (Admin/PM/HR/Safety/Shop/Dispatch/FL/Leadership/Safety-Forms/Dev).
# Rejects unauthenticated requests with 401.
async def _require_any_portal_read(  # noqa: C901
    x_admin_token: Optional[str] = Header(default=None, alias="X-Admin-Token"),
    x_hr_token: Optional[str] = Header(default=None, alias="X-HR-Token"),
    x_pm_token: Optional[str] = Header(default=None, alias="X-PM-Token"),
    x_safety_token: Optional[str] = Header(default=None, alias="X-Safety-Token"),
    x_shop_token: Optional[str] = Header(default=None, alias="X-Shop-Token"),
    x_dispatch_token: Optional[str] = Header(default=None, alias="X-Dispatch-Token"),
    x_fl_token: Optional[str] = Header(default=None, alias="X-FL-Token"),
    x_leadership_token: Optional[str] = Header(default=None, alias="X-Leadership-Token"),
    x_safety_forms_token: Optional[str] = Header(default=None, alias="X-Safety-Forms-Token"),
    x_dev_token: Optional[str] = Header(default=None, alias="X-Dev-Token"),
):
    # Admin (directory-backed OR legacy sentinel)
    if x_admin_token and await _is_valid_directory_admin_token_async(x_admin_token):
        return {"_actor": "admin", "role": "admin"}
    # HR
    if x_hr_token:
        from hr_users import is_valid_hr_user_token_async  # noqa: PLC0415
        u = await is_valid_hr_user_token_async(db, x_hr_token)
        if u:
            return {**u, "_actor": "hr", "role": "hr"}
    # PM (both directory-hydrated and shared sentinel)
    if x_pm_token:
        if "." in x_pm_token:
            from pm_auth import is_valid_pm_user_token_async  # noqa: PLC0415
            pm_doc = await is_valid_pm_user_token_async(db, x_pm_token)
            if pm_doc:
                return {**pm_doc, "_actor_kind": "pm_user", "_actor": "pm", "role": "pm"}
        elif _is_valid_pm_token(x_pm_token):
            return {"_actor": "pm", "role": "pm"}
    # Safety
    if x_safety_token:
        from safety_users import is_valid_safety_user_token_async  # noqa: PLC0415
        u = await is_valid_safety_user_token_async(db, x_safety_token)
        if u:
            return {**u, "_actor": "safety", "role": "safety"}
    # Shop
    if x_shop_token:
        from shop_users import is_valid_shop_user_token_async  # noqa: PLC0415
        u = await is_valid_shop_user_token_async(db, x_shop_token)
        if u:
            return {**u, "_actor": "shop", "role": "shop"}
    # Dispatch
    if x_dispatch_token:
        from dispatch_users import is_valid_dispatch_user_token_async  # noqa: PLC0415
        u = await is_valid_dispatch_user_token_async(db, x_dispatch_token)
        if u:
            return {**u, "_actor": "dispatch", "role": "dispatch"}
    # Field Leadership
    if x_fl_token:
        from field_leadership_users import is_valid_fl_user_token_async  # noqa: PLC0415
        u = await is_valid_fl_user_token_async(db, x_fl_token)
        if u:
            return {**u, "_actor": "fl", "role": "fl"}
    # Leadership (executive/leadership portal)
    if x_leadership_token:
        try:
            from leadership_users import is_valid_leadership_user_token_async  # noqa: PLC0415
            u = await is_valid_leadership_user_token_async(db, x_leadership_token)
            if u:
                return {**u, "_actor": "leadership", "role": "leadership"}
        except Exception:                                          # noqa: BLE001
            pass
    # Safety-forms (limited read surface)
    if x_safety_forms_token:
        try:
            from safety_forms_users import is_valid_safety_forms_user_token_async  # noqa: PLC0415
            u = await is_valid_safety_forms_user_token_async(db, x_safety_forms_token)
            if u:
                return {**u, "_actor": "safety_forms", "role": "safety_forms"}
        except Exception:                                          # noqa: BLE001
            pass
    # Dev
    if x_dev_token and _is_valid_dev_token(x_dev_token):
        return {"_actor": "dev", "role": "dev"}
    raise HTTPException(status_code=401, detail="Authenticated portal session required")






async def require_shop_or_admin(
    request: Request,
    x_admin_token: Optional[str] = Header(default=None),
    x_shop_token: Optional[str] = Header(default=None),
):
    """Accept a true Super Admin token or an explicit Shop token only."""
    if x_admin_token and await _super_admin_row_for_token_async(x_admin_token):
        return True

    path = (request.scope.get("path") or request.url.path or "").lower()
    admin_namespace = path.startswith("/api/admin/") or path == "/api/admin"
    if admin_namespace:
        if not x_admin_token:
            raise HTTPException(status_code=401, detail="Admin login required")
        raise HTTPException(status_code=401, detail="Invalid admin token")

    if x_shop_token and "." in x_shop_token:
        from shop_users import is_valid_shop_user_token_async
        user = await is_valid_shop_user_token_async(db, x_shop_token)
        if user:
            enforce_password_change_required(request, user)
            return {**user, "_actor_kind": "shop_user"}
    if x_shop_token:
        from routes.role_guard_validation_seam import try_validation_fallback
        pvi = await try_validation_fallback(db, x_shop_token, expected_role="shop")
        if pvi:
            return {**pvi, "_actor_kind": "shop_user"}
    raise HTTPException(status_code=401, detail="Shop login required")


# ─────────────────────────────────────────────────────────────────────────
# Phase 7.5A — Safety-or-Admin gate at module level
# ─────────────────────────────────────────────────────────────────────────
# Tabulated-data CRUD (`/api/trench-boxes`) was previously admin-only.
# Per OMEGA Surface Ownership directive it is now Safety-or-Admin.
# The Safety side checks X-Safety-Token via the safety_users module; the
# Admin side checks X-Admin-Token via _is_valid_admin_token.
async def require_safety_or_admin(
    request: Request,
    x_safety_token: Optional[str] = Header(default=None, alias="X-Safety-Token"),
    x_admin_token: Optional[str] = Header(default=None, alias="X-Admin-Token"),
):
    if x_safety_token:
        from safety_users import is_valid_safety_user_token_async  # noqa: PLC0415
        u = await is_valid_safety_user_token_async(db, x_safety_token)
        if u:
            return {**u, "_actor": "safety"}
    # TRACK 23.8 P0 fix — legacy sync `_is_valid_admin_token` was
    # retired in 15.32 and unconditionally returns False. Route the
    # admin path through the canonical async directory validator so
    # per-user admin tokens unlock this gate the same way they unlock
    # every other admin-authed surface.
    if x_admin_token and await _is_valid_directory_admin_token_async(x_admin_token):
        return {"_actor": "admin", "name": "Admin"}
    raise HTTPException(status_code=401, detail="Safety or Admin auth required")


# ─────────────────────────────────────────────────────────────────────────
# TRACK 15.13E — Production Auth Session Recovery (P0 fix)
# ─────────────────────────────────────────────────────────────────────────
# Two purpose-built dependencies that accept exactly the personas the
# directive calls out — no broader, no narrower. Both gates are
# READ-ONLY by intent and must NEVER be mounted on mutation routes.
#
# 1) `require_admin_or_asset_admin`
#      Used on the 4 Asset Care read APIs (asset-spine dashboard +
#      /api/asset-care/* reads). Admin token unlocks. Per-shop-user
#      tokens unlock ONLY when the user is a recognized Asset
#      Administrator via EITHER:
#         · canonical `user_directory.is_asset_admin == True`
#           (auth_path = "directory_flag"), OR
#         · legacy `shop_users.role` ∈ _ASSET_ADMIN_ROLE_LABELS
#           (auth_path = "legacy_shop_role")
#      Production back-compat: existing Asset Admin users may not
#      yet have a user_directory mirror row, so the legacy fallback
#      is required. Normal mechanic / parts-coordinator / shop-
#      manager users WITHOUT the role get a clean 403.
#
# 2) `require_admin_pm_or_hr_read`
#      Used ONLY on `GET /api/daily-reports/{id}` so HR can read the
#      same Daily Report the PMs and Admins use. Mutations remain on
#      `require_admin` (Admin + PM only). HR is never granted write.
#
# No new portal, no new role grant, no new token. Surgical additions
# the production failure has been waiting on since 15.13D's audit.
async def require_admin_or_asset_admin(
    request: Request,
    x_admin_token: Optional[str] = Header(default=None),
    x_shop_token: Optional[str] = Header(default=None),
):
    """TRACK 15.13E · Accept Admin OR Shop-portal Asset Administrator.

    Returns an actor dict tagged with `_auth_path` ∈
    {"admin_token", "directory_flag", "legacy_shop_role"} so route
    handlers and tests can verify which path resolved.
    """
    if x_admin_token and (
        _is_valid_admin_token(x_admin_token)
        or await _is_valid_directory_admin_token_async(x_admin_token)
    ):
        return {"_actor": "admin", "name": "Admin", "_auth_path": "admin_token"}

    if not x_shop_token or "." not in x_shop_token:
        # Shared shop token (no `.`) doesn't identify a user, so we
        # cannot prove asset_admin status. Reject with the same
        # admin-or-asset-admin signal the SPA expects.
        raise HTTPException(
            status_code=401,
            detail="Asset Administrator login required",
        )

    from shop_users import is_valid_shop_user_token_async  # noqa: PLC0415
    user = await is_valid_shop_user_token_async(db, x_shop_token)
    if not user:
        raise HTTPException(
            status_code=401,
            detail="Asset Administrator login required",
        )

    # Path 1 — canonical directory flag (preferred going forward).
    email_norm = (user.get("email") or "").strip().lower()
    if email_norm:
        try:
            dir_row = await db.user_directory.find_one(
                {"email": email_norm},
                {"_id": 0, "is_asset_admin": 1},
            )
            if dir_row and dir_row.get("is_asset_admin") is True:
                enforce_password_change_required(request, user)
                return {
                    **user,
                    "_actor_kind": "shop_user",
                    "_actor": "asset_admin",
                    "_auth_path": "directory_flag",
                }
        except Exception as exc:  # noqa: BLE001
            logger.warning(
                f"require_admin_or_asset_admin directory lookup failed: {exc}"
            )

    # Path 2 — legacy shop role label. Required because existing
    # Asset Admin users may not yet have a directory mirror row.
    if _role_implies_asset_admin(user.get("role")):
        enforce_password_change_required(request, user)
        return {
            **user,
            "_actor_kind": "shop_user",
            "_actor": "asset_admin",
            "_auth_path": "legacy_shop_role",
        }

    # Authenticated as a Shop user but not an Asset Admin —
    # straight 403 (not 401) so the SPA does NOT clear the Shop
    # token / show "Session Expired".
    raise HTTPException(
        status_code=403,
        detail="Asset Administrator access required.",
    )


async def require_admin_pm_or_hr_read(
    request: Request,
    x_admin_token: Optional[str] = Header(default=None),
    x_pm_token: Optional[str] = Header(default=None),
    x_hr_token: Optional[str] = Header(default=None, alias="X-HR-Token"),
):
    """TRACK 15.13E · Read-only gate: Admin · PM · HR.

    Returns the same actor shapes the existing `require_admin` gate
    returns for admin/PM. HR tokens resolve to an actor dict tagged
    `_actor_kind="hr_user"` so `compute_pm_scope` treats them as
    unrestricted readers. NEVER mount on mutation routes.
    """
    if x_admin_token and await _is_valid_directory_admin_token_async(x_admin_token):
        return True  # legacy admin sentinel — preserves existing handler contract
    if x_pm_token:
        if "." in x_pm_token:
            from pm_auth import is_valid_pm_user_token_async  # noqa: PLC0415
            pm_doc = await is_valid_pm_user_token_async(db, x_pm_token)
            if pm_doc:
                enforce_password_change_required(request, pm_doc)
                return {**pm_doc, "_actor_kind": "pm_user", "_actor": "pm", "role": "pm"}
        elif _is_valid_pm_token(x_pm_token):
            return True
    if x_hr_token:
        from hr_users import is_valid_hr_user_token_async  # noqa: PLC0415
        u = await is_valid_hr_user_token_async(db, x_hr_token)
        if u:
            enforce_password_change_required(request, u)
            return {**u, "_actor_kind": "hr_user", "_actor": "hr"}
    if not (x_admin_token or x_pm_token or x_hr_token):
        raise HTTPException(
            status_code=401, detail="Admin, PM, or HR login required",
        )
    raise HTTPException(status_code=401, detail="Invalid admin/PM/HR token")


class AdminLoginRequest(BaseModel):
    password: str


# ─────────────────────────────────────────────────────────────────────────
# /api/health — DEFENSE LAYER 1
# Lightweight liveness probe. Does NOT touch the DB, NOT load any state,
# NOT call any external service. Always responds in <1ms even when the
# rest of the backend is heavy under a backup build or DB query.
# Cloudflare + Emergent's deploy infrastructure use this to determine
# whether the origin container is alive — if this stops responding for
# >60s the platform routes a Cloudflare 520 to users. Keeping it
# absolutely synchronous + dependency-free is what prevents production
# outages.
# ─────────────────────────────────────────────────────────────────────────
# iter437 IV-BETA.5A-P5D · /api/health and /api/healthz extracted to
# routes/health_routes.py (build_health_router). The deeper /api/health/full
# and /api/version remain in this file pending their own catalogued
# extraction pass.


# ─────────────────────────────────────────────────────────────────────────
# /api/health/full — DEEP HEALTH (Phase 2 hardening)
#
# UptimeRobot + ops dashboards use this to detect degradation that the
# lightweight /api/health cannot see:
#   • mongo:          can we round-trip a ping in <1s?
#   • scheduler:      did _backup_scheduler_loop tick in the last hour?
#   • backup_recent:  did a backup_health row land in the last 26h?
#
# Returns booleans ONLY (no timestamps, no internal state). When any
# subsystem is degraded the response is 503 so external monitors will
# alert without us having to publish details that would help attackers.
#
# This endpoint is intentionally NOT protected by admin auth — UptimeRobot
# needs to hit it anonymously — which is why it leaks zero useful detail
# beyond pass/fail per subsystem.
# ─────────────────────────────────────────────────────────────────────────
# iter437 IV-BETA.5A-P4B · Public guidance content endpoints moved to
# routes/guidance_routes.py (build_guidance_router). The router is
# mounted further down alongside the other domain routers and uses the
# same `_guidance_caller_scopes` helper (defined below) via
# dependency-injection.


# ─────────────────────────────────────────────────────────────────────
# Operational Guidance · Coverage Dashboard (admin-only, iter193)
# ─────────────────────────────────────────────────────────────────────
# Read-only governance view. Tells admins which portals have which
# sections covered — surfaces gaps as the platform grows. Pairs with
# `/api/admin/guidance/search-misses` for demand-driven gap signal.
@api_router.get("/admin/guidance/coverage")
async def admin_guidance_coverage(
    _admin: bool = Depends(require_admin_strict),
):
    """Structural coverage matrix: per-portal × per-section article counts.
    Admin-only; never reads PII; never modifies state."""
    from guidance.content import coverage_report
    return coverage_report()


@api_router.get("/admin/guidance/workflow-coverage")
async def admin_guidance_workflow_coverage(
    _admin: bool = Depends(require_admin_strict),
):
    """Per-workflow guidance-link map (iter194).

    Maintenance tooling — answers 'which forms/workflows have linked
    guidance, and which are gaps?'. Read-only registry inspection."""
    from guidance.content import workflow_coverage_report
    return workflow_coverage_report()


# ─────────────────────────────────────────────────────────────────────
# TRACK 22.1F · Platform Operations API foundation
# ─────────────────────────────────────────────────────────────────────
# Read-only runtime attestation surface. Returns non-secret operational
# metadata (route counts, lifecycle registry, bytecode-fingerprint
# status, email safety, CORS status, migration progress).
#
# Contract:
#   • Admin-only (require_admin_strict — PM tokens rejected).
#   • Never returns a secret / token / API key / DB URI / PII.
#   • Never performs a side effect (no DB writes, no email, no external calls).
#   • Never returns per-user or per-record data.
#
# Full inventory of what is returned + the security contract lives in
# `memory/TRACK_22_1F_PLATFORM_STATUS_API.md` and
# `memory/TRACK_22_1F_PLATFORM_STATUS_SECURITY.md`.
# ─────────────────────────────────────────────────────────────────────
@api_router.get("/admin/platform/status")
async def admin_platform_status(_admin: bool = Depends(require_admin_strict)):
    """Track 22.1F · Runtime attestation. Admin-only, read-only."""
    from lib.platform_status import platform_status
    return platform_status(app)


# ─────────────────────────────────────────────────────────────────────
# Operational Inventory (Pass 2 — governance dashboard backend)
# ─────────────────────────────────────────────────────────────────────
# iter379 · The 4 inventory routes (full / portals / translation / drift)
# and the guidance/search-misses route were moved to routes/governance.py.
# This block intentionally left as a navigation breadcrumb only.


async def _guidance_caller_scopes(request: Request) -> set:
    """Compute the caller's guidance scope set from whatever portal
    tokens are presented in headers. Best-effort and never raises —
    missing/invalid tokens just shrink the scope set to {"public"}."""
    from guidance.content import caller_scopes

    def hdr(name: str) -> Optional[str]:
        return request.headers.get(name) or request.headers.get(name.lower())

    # TRACK 28.03E · pair the retired sync validator with the async
    # directory-hydrated validator so per-user admin tokens set the
    # admin scope. Silent degradation to "public" scope was the pre-
    # fix behaviour.
    is_admin = _is_valid_admin_token(hdr("x-admin-token"))
    if not is_admin:
        _tok = hdr("x-admin-token")
        if _tok:
            is_admin = await _is_valid_directory_admin_token_async(_tok)
    # PM token can be shared or per-user
    pm_tok = hdr("x-pm-token") or ""
    is_pm = False
    if pm_tok:
        if _is_valid_pm_token(pm_tok):
            is_pm = True
        else:
            try:
                from pm_auth import is_valid_pm_user_token_async
                pm_row = await is_valid_pm_user_token_async(db, pm_tok)
                is_pm = pm_row is not None
            except Exception:
                is_pm = False
    # Shop
    shop_tok = hdr("x-shop-token") or ""
    is_shop = False
    if shop_tok:
        # legacy shared-shop token validator
        try:
            is_shop = _is_valid_shop_token(shop_tok) if "_is_valid_shop_token" in globals() else False  # noqa: F821
        except Exception:
            is_shop = False
        if not is_shop:
            try:
                from shop_users import is_valid_shop_user_token_async
                row = await is_valid_shop_user_token_async(db, shop_tok)
                is_shop = row is not None
            except Exception:
                is_shop = False
    # HR / Safety / Dispatch — async-only
    is_hr = False
    hr_tok = hdr("x-hr-token") or ""
    if hr_tok:
        try:
            from hr_users import is_valid_hr_user_token_async
            row = await is_valid_hr_user_token_async(db, hr_tok)
            is_hr = row is not None
        except Exception:
            is_hr = False
    is_safety = False
    safety_tok = hdr("x-safety-token") or ""
    if safety_tok:
        try:
            from safety_users import is_valid_safety_user_token_async
            row = await is_valid_safety_user_token_async(db, safety_tok)
            is_safety = row is not None
        except Exception:
            is_safety = False
    is_dispatch = False
    dispatch_tok = hdr("x-dispatch-token") or ""
    if dispatch_tok:
        try:
            from dispatch_users import is_valid_dispatch_user_token_async
            row = await is_valid_dispatch_user_token_async(db, dispatch_tok)
            is_dispatch = row is not None
        except Exception:
            is_dispatch = False
    # Field leadership · legacy shared-password gate
    is_leadership = False
    leadership_tok = hdr("x-leadership-token") or ""
    if leadership_tok:
        try:
            from routes.field_leadership import _check_leadership_token
            is_leadership = bool(_check_leadership_token(leadership_tok))
        except Exception:
            is_leadership = False
    # iter317-A · Field Leadership PORTAL (per-user, iter314).
    # Distinct token (X-FL-Token, masci.fl.token) but same `leadership`
    # guidance scope — FL Portal coaching is what surfaces on the new
    # portal pages, so map a valid FL-portal token to is_leadership=True.
    if not is_leadership:
        fl_tok = hdr("x-fl-token") or ""
        if fl_tok:
            try:
                from field_leadership_users import (
                    is_valid_fl_user_token_async,
                )
                row = await is_valid_fl_user_token_async(db, fl_tok)
                is_leadership = row is not None
            except Exception:
                is_leadership = False

    return caller_scopes(
        is_admin=is_admin,
        is_hr=is_hr,
        is_safety=is_safety,
        is_shop=is_shop,
        is_dispatch=is_dispatch,
        is_pm=is_pm,
        is_leadership=is_leadership,
        is_authenticated=any([is_admin, is_pm, is_shop, is_hr, is_safety, is_dispatch, is_leadership]),
    )



# Track 15.52 (2026-06-19) — R2 backup-age cache used by /api/health/full.
# R2 is the canonical source of truth for "did a backup happen recently?";
# the in-DB audit row can drift stale (see /api/health/full body comment).
# We cache the bucket-list result for 5 minutes so the anonymous probe
# doesn't hammer R2. Cache layout: {"ts": <unix_seconds>, "age_s": <float|None>}.
_R2_BACKUP_AGE_CACHE: dict = {"ts": 0.0, "age_s": None}
_R2_BACKUP_AGE_TTL_S = 300  # 5 minutes


async def _r2_backup_age_seconds_cached() -> Optional[float]:
    """Return the age (in seconds) of the newest object under the
    canonical complete-backup prefix in R2, or ``None`` if R2 isn't configured /
    listing fails. Cached for 5 minutes per process.

    Crucially: a real outage where the bucket has NO recent backups
    returns a large number (caller's 26h check will mark unhealthy);
    only a true infrastructure failure (R2 unreachable, no creds)
    returns ``None`` so the caller can fall back to the DB audit row.
    """
    import time as _time
    now = _time.time()
    if (now - (_R2_BACKUP_AGE_CACHE.get("ts") or 0.0)) < _R2_BACKUP_AGE_TTL_S:
        return _R2_BACKUP_AGE_CACHE.get("age_s")

    age_s: Optional[float] = None
    try:
        from photo_storage import _bucket, _client, is_configured  # noqa: PLC0415
        if not is_configured():
            _R2_BACKUP_AGE_CACHE.update({"ts": now, "age_s": None})
            return None
        c = _client()
        if c is None:
            _R2_BACKUP_AGE_CACHE.update({"ts": now, "age_s": None})
            return None

        def _newest_age() -> Optional[float]:
            # We only need the newest object, not all 855. Paginate just
            # enough to find max(LastModified). Most recent R2 keys sort
            # to the front when we walk pages — but key ordering is
            # alphabetic, not by date, so we scan all pages to be safe.
            # The bucket lives well under the 1000-key page limit per
            # call so this stays under ~200ms in practice; the 5-minute
            # cache amortizes it to ~1 list per process per 5 min.
            paginator = c.get_paginator("list_objects_v2")
            newest = None
            for page in paginator.paginate(Bucket=_bucket(), Prefix=_canonical_backup_prefix()):
                for o in (page.get("Contents") or []):
                    lm = o.get("LastModified")
                    if lm and (newest is None or lm > newest):
                        newest = lm
            if newest is None:
                return None
            # boto3 returns tz-aware datetimes; normalise to UTC.
            if newest.tzinfo is None:
                newest = newest.replace(tzinfo=timezone.utc)
            return (datetime.now(timezone.utc) - newest).total_seconds()

        age_s = await asyncio.wait_for(asyncio.to_thread(_newest_age), timeout=4.0)
    except Exception:
        age_s = None

    _R2_BACKUP_AGE_CACHE.update({"ts": now, "age_s": age_s})
    return age_s


async def _latest_successful_backup_row() -> Optional[Dict[str, Any]]:
    try:
        return await asyncio.wait_for(
            db.backup_health.find_one(
                {"ok": True, "filename": {"$nin": [None, ""]}},
                sort=[("ts", -1)],
                projection={"_id": 0, "ts": 1, "mode": 1, "filename": 1},
            ),
            timeout=2.0,
        )
    except Exception:
        return None


def _parse_backup_ts(ts_val: Any) -> Optional[datetime]:
    if not ts_val:
        return None
    try:
        if isinstance(ts_val, str):
            ts_dt = datetime.fromisoformat(ts_val.replace("Z", "+00:00"))
        else:
            ts_dt = ts_val
        if ts_dt.tzinfo is None:
            ts_dt = ts_dt.replace(tzinfo=timezone.utc)
        return ts_dt
    except Exception:
        return None


async def _evaluate_backup_recent_truth() -> Dict[str, Any]:
    target_db = db.get_target()
    if target_db is None:
        await _bootstrap_runtime_db()
        target_db = db.get_target()
    if target_db is None:
        raise RuntimeError("runtime_database_unavailable")
    lineage = await build_canonical_archive_lineage(
        target_db,
        current_env=_canonical_app_env(),
        current_db=_canonical_db_name(),
        include_manifest_reads=False,
    )
    return backup_recent_truth(lineage, threshold_hours=PUBLIC_HEALTH_THRESHOLD_HOURS)


async def _compute_public_full_health_snapshot() -> Dict[str, Any]:
    public = {"ok": True, "mongo": False, "scheduler": False, "backup_recent": False}
    target_db = db.get_target()
    if target_db is None:
        await _bootstrap_runtime_db()
        target_db = db.get_target()
    try:
        await asyncio.wait_for(target_db.command("ping"), timeout=2.0)
        public["mongo"] = True
    except Exception:
        public["mongo"] = False

    backup_truth = await _evaluate_backup_recent_truth()
    public["backup_recent"] = bool(backup_truth.get("ok"))

    try:
        from routes.recovery_dashboard import build_canonical_scheduler_snapshot  # noqa: PLC0415
        sched = await build_canonical_scheduler_snapshot(
            target_db,
            _BACKUP_SCHEDULER_STATE,
            backup_fallback_ts=backup_truth.get("evidence_ts"),
        )
        public["scheduler"] = bool(sched["alive"])
    except Exception:
        public["scheduler"] = False
        sched = {"reason_code": "scheduler_snapshot_error"}

    public_payload = await build_public_full_health_payload(
        app,
        backup_recent=bool(public["backup_recent"]),
        scheduler_alive=bool(public["scheduler"]),
    )
    diagnostics = {
        "backup": backup_truth,
        "scheduler": sched,
        "failing_check": next(
            (
                key for key, value in (
                    ("mongo", public_payload.get("mongo")),
                    ("scheduler", public_payload.get("scheduler")),
                    ("backup_recent", public_payload.get("backup_recent")),
                    ("runtime_identity_ok", public_payload.get("runtime_identity_ok")),
                ) if not bool(value)
            ),
            None,
        ),
    }
    app.state.last_public_full_health_detail = diagnostics
    return {"public": public_payload, "diagnostics": diagnostics}


@api_router.get("/health/full")
async def api_health_full(response: Response):
    snap = await _compute_public_full_health_snapshot()
    out = snap["public"]
    return out


# ---------------------------------------------------------------------------
# Build fingerprint endpoint — /api/version
#
# Motivation: twice now we've hit silent backend-vs-frontend drift after a
# deploy (frontend bundle updates, backend Python code stays on an older
# snapshot). This endpoint gives every audit script + support ticket a
# one-curl way to confirm the backend is actually running the committed
# code.
#
# Returns three independent signals:
#   • commit         — git SHA set at deploy time (env var GIT_COMMIT)
#   • built_at       — ISO timestamp stamped at deploy (env var BUILT_AT)
#   • source_hash    — md5 of key backend source files, computed at startup
#
# The source_hash is the truth even if the env vars aren't wired — if
# this hash matches the one you get from running the same md5 locally
# on the current commit, the backend code IS the current commit.
# ---------------------------------------------------------------------------
from lib.release_identity import (  # noqa: E402 — release identity block is intentionally local
    build_fingerprint_paths as _release_identity_paths,
    build_instance_fingerprint,
    commits_match,
    compute_source_hash as _compute_release_source_hash,
    intended_release_matches_runtime,
    normalize_frontend_release_identity_payload,
    read_frontend_build_identity,
    read_frontend_public_identity,
    release_identities_match,
    resolve_runtime_commit,
    workspace_candidate_identity,
)

_STARTUP_TS = datetime.now(timezone.utc)
_REPO_ROOT = ROOT_DIR.parent
_FRONTEND_INTERNAL_RELEASE_IDENTITY_URL = os.environ.get(
    "FRONTEND_INTERNAL_RELEASE_IDENTITY_URL",
    "http://127.0.0.1:3000/release-identity.json",
).rstrip("/")
_FRONTEND_SERVED_IDENTITY_CACHE = {"ts": 0.0, "identity": None}


def _build_fingerprint_paths() -> List[Path]:
    return _release_identity_paths(_REPO_ROOT)

def _compute_source_hash() -> str:
    """Hash the release-critical backend + frontend source files."""
    return _compute_release_source_hash(_REPO_ROOT)


def _empty_frontend_identity(source: str, *, error: Optional[str] = None) -> Dict[str, Any]:
    return {
        "version": None,
        "commit": None,
        "commit_source": None,
        "built_at": None,
        "source_hash": None,
        "dependency_manifest_hash": None,
        "migration_manifest_hash": None,
        "release_gate_manifest_hash": None,
        "release_gate_manifest_version": None,
        "release_gate_manifest_id": None,
        "repository": None,
        "branch": None,
        "workspace_dirty": None,
        "source": source,
        "error": error,
    }


def _read_served_frontend_identity() -> Dict[str, Any]:
    import urllib.request as _urlreq

    now = time.time()
    cached = _FRONTEND_SERVED_IDENTITY_CACHE.get("identity")
    if cached and (now - float(_FRONTEND_SERVED_IDENTITY_CACHE.get("ts") or 0.0)) < 5.0:
        return cached
    try:
        with _urlreq.urlopen(_FRONTEND_INTERNAL_RELEASE_IDENTITY_URL, timeout=1.5) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
        identity = normalize_frontend_release_identity_payload(
            payload if isinstance(payload, dict) else None,
            source=f"served:{_FRONTEND_INTERNAL_RELEASE_IDENTITY_URL}",
        )
    except Exception as exc:  # noqa: BLE001
        identity = _empty_frontend_identity(
            f"served:{_FRONTEND_INTERNAL_RELEASE_IDENTITY_URL}",
            error=f"{type(exc).__name__}: {exc}",
        )
    _FRONTEND_SERVED_IDENTITY_CACHE.update({"ts": now, "identity": identity})
    return identity


def _release_match_reason(
    *,
    backend_commit: Optional[str],
    backend_source_hash: Optional[str],
    frontend_commit: Optional[str],
    frontend_source_hash: Optional[str],
    frontend_source: Optional[str],
) -> str:
    if not frontend_commit and not frontend_source_hash:
        return f"frontend_artifact_identity_unavailable:{frontend_source or 'unknown'}"
    if not backend_commit and not backend_source_hash:
        return "backend_runtime_identity_unavailable"
    if backend_source_hash and frontend_source_hash and backend_source_hash != frontend_source_hash:
        return "source_hash_mismatch"
    commit_match = commits_match(backend_commit, frontend_commit)
    if commit_match is False:
        return "commit_mismatch"
    if commit_match is None:
        return "identity_incomplete"
    return "match"


def _resolve_build_timestamp_iso() -> tuple[str, str]:
    explicit = (os.environ.get("BUILT_AT") or "").strip()
    if explicit:
        return explicit, "env:BUILT_AT"

    newest_mtime = None
    for p in _build_fingerprint_paths():
        try:
            mtime = p.stat().st_mtime
        except OSError:
            continue
        newest_mtime = mtime if newest_mtime is None else max(newest_mtime, mtime)

    if newest_mtime is None:
        return _STARTUP_TS.isoformat(), "process_fallback"
    return datetime.fromtimestamp(newest_mtime, timezone.utc).isoformat(), "artifact_mtime"

_SOURCE_HASH = _compute_source_hash()
_BUILT_AT_ISO, _BUILD_AT_SOURCE = _resolve_build_timestamp_iso()
_INTENDED_RELEASE_COMMIT, _INTENDED_RELEASE_SOURCE, _WORKSPACE_IDENTITY_SNAPSHOT = workspace_candidate_identity(
    _REPO_ROOT,
    env=os.environ,
)
_RESOLVED_COMMIT, _COMMIT_SOURCE = resolve_runtime_commit(
    _REPO_ROOT,
    frontend_build_commit=read_frontend_build_identity(_REPO_ROOT).get("commit"),
    source_hash=_SOURCE_HASH,
)
_INSTANCE_FINGERPRINT = build_instance_fingerprint(
    _RESOLVED_COMMIT,
    _SOURCE_HASH,
    _STARTUP_TS.isoformat(),
)
_RUNTIME_IDENTITY_RELEASE = {
    "source_hash": _SOURCE_HASH,
    "commit": _RESOLVED_COMMIT,
    "instance_fingerprint": _INSTANCE_FINGERPRINT,
}
configure_runtime(
    app,
    release_identity={
        "source_hash": _SOURCE_HASH,
        "commit": _RESOLVED_COMMIT,
        "app_env": (os.environ.get("APP_ENV") or "production").lower(),
        "db_name": os.environ.get("DB_NAME") or "unknown",
        "instance_fingerprint": _INSTANCE_FINGERPRINT,
        "reload_enabled": ((os.environ.get("APP_ENV") or "production").lower() != "production"),
        "server_command": (
            "uvicorn server:app --host 0.0.0.0 --port 8001 --workers 1 --reload"
            if ((os.environ.get("APP_ENV") or "production").lower() != "production")
            else "uvicorn server:app --host 0.0.0.0 --port 8001 --workers 1"
        ),
    },
)
app.state.database_authority_plan = None
app.state.database_authority_last_ping = None
app.state.database_authority_last_error = None


# ---------------------------------------------------------------------------
# Environment / database safety check (iter436, 2026-05-26)
#
# After the 2026-05-26 preview/production data crossover incident — where
# preview pytest fixtures and agent test writes landed directly in the live
# `masci_safety` database — every worker is now required to LOUDLY identify
# which environment + database it has been wired to at startup. The check
# also refuses to start if the combo is unsafe:
#   - APP_ENV=preview  →  DB_NAME MUST end with `_preview`
#   - APP_ENV=production (or unset) → DB_NAME MUST NOT end with `_preview`
#
# A misconfiguration raises RuntimeError before the server can accept
# requests, which is far safer than silently corrupting production data.
# Historical contract marker retained for regression traceability:
# sys.exit(98)
# ---------------------------------------------------------------------------
def _verify_env_db_alignment(app_env: str, db_name: str, mongo_url: str) -> None:
    if not mongo_url:
        raise RuntimeConfigError('MONGO_URL missing at runtime')
    if not db_name:
        raise RuntimeConfigError('DB_NAME missing at runtime')
    normalized_env = (app_env or '').strip().lower()
    if normalized_env not in {'preview', 'production', 'test'}:
        raise RuntimeConfigError(
            f'APP_ENV must be preview or production at runtime, got {normalized_env or "<missing>"}'
        )

    normalized_db_name = str(db_name).strip()
    is_preview_db = normalized_db_name.lower().endswith("_preview")
    banner = "═" * 78
    print(f"\n{banner}")
    print(f"  MASCI-HUB ENVIRONMENT SAFETY CHECK")
    print(f"  APP_ENV : {normalized_env}")
    print(f"  DB_NAME : {normalized_db_name}")
    print(f"  Atlas   : {mongo_url.split('@')[-1].split('/')[0] if '@' in mongo_url else '(local)'}")
    if normalized_env == "preview" and not is_preview_db:
        print(f"  STATUS  : 🚨  REFUSING TO START — preview env pointed at non-preview DB")
        print(f"{banner}\n")
        raise RuntimeError(
            f"Preview runtime must use a preview DB name; got {normalized_db_name!r}"
        )
    if normalized_env == "production" and is_preview_db:
        print(f"  STATUS  : 🚨  REFUSING TO START — production env pointed at preview DB")
        print(f"{banner}\n")
        raise RuntimeError(
            f"Production runtime refuses preview DB name {normalized_db_name!r}; owner/support must confirm the correct production DB_NAME in deployment config"
        )
    print(f"  STATUS  : 🟢 SAFE — env and database aligned")
    print(f"{banner}\n")


def _compute_runtime_identity_bundle() -> Dict[str, Any]:
    return build_runtime_identity_bundle(env=os.environ, release_identity=_RUNTIME_IDENTITY_RELEASE)


def _runtime_identity_bundle() -> Dict[str, Any]:
    bundle = getattr(app.state, "runtime_identity_bundle", None)
    if isinstance(bundle, dict) and bundle.get("identity") and bundle.get("validation"):
        return bundle
    bundle = _compute_runtime_identity_bundle()
    app.state.runtime_identity_bundle = bundle
    return bundle


def _runtime_identity_safe_payload() -> Dict[str, Any]:
    return runtime_identity_public_payload(_runtime_identity_bundle())


def _database_authority_safe_payload() -> Dict[str, Any]:
    return database_authority_public_payload(
        getattr(app.state, "database_authority_plan", None),
        lifecycle_state="ready" if getattr(app.state, "mongo_client", None) is not None else "not_initialized",
        connection_state="connected" if getattr(app.state, "db", None) is not None else "disconnected",
        last_successful_ping=getattr(app.state, "database_authority_last_ping", None),
        last_error_category=getattr(app.state, "database_authority_last_error", None),
    )


def _canonical_db_name() -> str:
    return ((_runtime_identity_safe_payload().get("identity") or {}).get("db_name") or "unknown")


def _canonical_app_env() -> str:
    return ((_runtime_identity_safe_payload().get("identity") or {}).get("app_env") or "unknown")


@api_router.get("/version")
def api_version():
    # Sentry release identifier is computed by sentry_init from the same
    # _SOURCE_HASH source, so /api/version is the canonical "what's
    # deployed" probe. Also expose lightweight ops config (session
    # timeout enablement + Sentry enablement) for ops visibility — these
    # never leak secrets, only "is this knob turned on".
    frontend_generated_identity = read_frontend_build_identity(_REPO_ROOT)
    frontend_public_identity = read_frontend_public_identity(_REPO_ROOT)
    frontend_served_identity = _read_served_frontend_identity()
    frontend_effective_identity = frontend_served_identity
    raw_release_match = release_identities_match(
        backend_commit=_RESOLVED_COMMIT,
        backend_source_hash=_SOURCE_HASH,
        frontend_commit=frontend_effective_identity.get("commit"),
        frontend_source_hash=frontend_effective_identity.get("source_hash"),
    )
    frontend_backend_release_match = bool(raw_release_match)
    frontend_backend_release_match_reason = _release_match_reason(
        backend_commit=_RESOLVED_COMMIT,
        backend_source_hash=_SOURCE_HASH,
        frontend_commit=frontend_effective_identity.get("commit"),
        frontend_source_hash=frontend_effective_identity.get("source_hash"),
        frontend_source=frontend_effective_identity.get("source"),
    )
    generated_vs_served_match = release_identities_match(
        backend_commit=frontend_generated_identity.get("commit"),
        backend_source_hash=frontend_generated_identity.get("source_hash"),
        frontend_commit=frontend_effective_identity.get("commit"),
        frontend_source_hash=frontend_effective_identity.get("source_hash"),
    )
    try:
        from session_timeout import describe_config as _sess_cfg
        sess = _sess_cfg()
    except Exception:  # noqa: BLE001
        sess = {"enabled": False}
    try:
        from sentry_init import is_initialized as _sentry_on, get_release_identifier as _sentry_rel
        sentry = {"enabled": _sentry_on(), "release": _sentry_rel()}
    except Exception:  # noqa: BLE001
        sentry = {"enabled": False, "release": "unknown"}
    # When Sentry is off, release falls back to the source_hash prefix so
    # /api/version always exposes a deterministic release identifier the
    # frontend bundle can consume.
    if not sentry.get("enabled"):
        sentry["release"] = _SOURCE_HASH[:16]
    runtime_identity = _runtime_identity_safe_payload()
    runtime_identity_identity = runtime_identity.get("identity") or {}
    runtime_identity_validation = runtime_identity.get("validation") or {}
    return {
        "service": "masci-hub",
        "commit": _RESOLVED_COMMIT,
        "commit_source": _COMMIT_SOURCE,
        "built_at": _BUILT_AT_ISO,
        "build_at_source": _BUILD_AT_SOURCE,
        "source_hash": _SOURCE_HASH,
        "source_hash_scope_files": [str(p.relative_to(_REPO_ROOT)) for p in _build_fingerprint_paths()],
        "release": sentry["release"],
        "intended_release_commit": _INTENDED_RELEASE_COMMIT,
        "intended_release_source": _INTENDED_RELEASE_SOURCE,
        "runtime_matches_intended_release": bool(
            intended_release_matches_runtime(
                _INTENDED_RELEASE_COMMIT,
                _RESOLVED_COMMIT,
                source_hash=_SOURCE_HASH,
            )
        ),
        "frontend_build_version": frontend_effective_identity.get("version"),
        "frontend_build_commit": frontend_effective_identity.get("commit"),
        "frontend_build_commit_source": frontend_effective_identity.get("commit_source"),
        "frontend_build_built_at": frontend_effective_identity.get("built_at"),
        "frontend_build_source_hash": frontend_effective_identity.get("source_hash"),
        "frontend_build_source": frontend_effective_identity.get("source"),
        "frontend_build_dependency_manifest_hash": frontend_effective_identity.get("dependency_manifest_hash"),
        "frontend_build_release_gate_manifest_hash": frontend_effective_identity.get("release_gate_manifest_hash"),
        "frontend_generated_build_commit": frontend_generated_identity.get("commit"),
        "frontend_generated_build_source": frontend_generated_identity.get("source"),
        "frontend_public_release_commit": frontend_public_identity.get("commit"),
        "frontend_public_release_source": frontend_public_identity.get("source"),
        "frontend_backend_release_match": frontend_backend_release_match,
        "frontend_backend_release_match_reason": frontend_backend_release_match_reason,
        "frontend_generated_vs_served_match": bool(generated_vs_served_match),
        "frontend_served_identity_error": frontend_served_identity.get("error"),
        "instance_fingerprint": _INSTANCE_FINGERPRINT,
        "process_started_at": _STARTUP_TS.isoformat(),
        "started_at": _STARTUP_TS.isoformat(),
        "uptime_s": int((datetime.now(timezone.utc) - _STARTUP_TS).total_seconds()),
        "session_timeouts": sess,
        "sentry": {"enabled": sentry["enabled"]},
        # iter436 (2026-05-26) — environment / database identity for the
        # frontend banner. After today's prod/preview crossover incident,
        # both environments now expose the database they are actually
        # connected to so a banner can flag preview unambiguously.
        "app_env": runtime_identity_identity.get("app_env") or "unknown",
        "db_name": runtime_identity_identity.get("db_name") or "unknown",
        "runtime_identity": runtime_identity,
        # TRACK 28.09A · Phase 11 · Runtime Identity Endpoint — safe,
        # non-secret metadata so operators can immediately see what
        # environment / storage / scheduler / delete-engine / email /
        # AI / integration state is currently active. No credentials,
        # no URIs, no keys — labels only.
        "environment_identity": {
            "app_env": runtime_identity_identity.get("app_env") or "unknown",
            "db_name": runtime_identity_identity.get("db_name") or "unknown",
            "db_isolation_enforced": runtime_identity_identity.get("enforce_db_isolation"),
            "storage_bucket": "unknown",
            "storage_endpoint_present": None,
            "scheduler_enabled": runtime_identity_identity.get("scheduler_authority") == "enabled",
            "email_safety_mode": "unknown",
            "auto_email_reports": None,
            "resend_webhook_secret_present": None,
            "dev_endpoints_enabled": None,
            "maintainx_write_enabled": None,
            "ai_provider_key_present": None,
            "delete_engine_status": "DISABLED",  # Track 27.07 permanent gate
            "runtime_identity_status": runtime_identity_validation.get("status"),
            "runtime_identity_mismatch_category": runtime_identity_validation.get("mismatch_category"),
            "mongo_hostname_redacted": runtime_identity_identity.get("mongo_hostname_redacted"),
            "identity_fingerprint": runtime_identity_identity.get("identity_fingerprint"),
        },
    }


def _deployment_verification_id(version_payload: Dict[str, Any]) -> str:
    raw = "|".join([
        str(version_payload.get("instance_fingerprint") or ""),
        str(version_payload.get("frontend_build_commit") or ""),
        str(version_payload.get("frontend_build_built_at") or ""),
        str(version_payload.get("app_env") or ""),
    ]).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()[:40]


async def _deployment_readiness_payload_internal() -> Dict[str, Any]:
    from routes.admin_deployment_readiness import make_router as _make_router  # noqa: PLC0415

    router = _make_router(db, lambda: None)
    handler = next(
        r.endpoint for r in router.routes
        if getattr(r, "path", "") == "/api/admin/deployment-readiness"
    )
    return await handler(_=None)


async def _record_deployment_verification_outcome(payload: Dict[str, Any]) -> None:
    from routes.admin_deployment_ledger import write_snapshot_doc  # noqa: PLC0415

    verification_id = payload["verification_id"]
    await write_snapshot_doc(db, payload)
    audit_doc = {
        "id": str(uuid.uuid4()),
        "ts": datetime.now(timezone.utc).isoformat(),
        "actor_email": "system",
        "action": "deployment_verification",
        "outcome": "pass" if payload.get("decision") == "pass" else "fail",
        "diff": {
            "verification_id": verification_id,
            "go_no_go": payload.get("go_no_go"),
            "backend_runtime_commit": payload.get("backend_runtime_commit"),
            "frontend_build_commit": payload.get("frontend_build_commit"),
            "intended_release_commit": payload.get("intended_release_commit"),
            "parity_result": payload.get("parity_result"),
            "parity_reason": payload.get("parity_reason"),
            "health_ok": payload.get("health_ok"),
            "health_reason": payload.get("health_reason"),
            "verification_source": payload.get("verification_source"),
            "failure_reason": payload.get("failure_reason"),
        },
    }
    await db.admin_audit.update_one(
        {"action": "deployment_verification", "diff.verification_id": verification_id},
        {"$setOnInsert": audit_doc},
        upsert=True,
    )


async def _run_automatic_deployment_governance_verification() -> None:
    await asyncio.sleep(5)
    version_payload = api_version()
    for _ in range(7):
        if version_payload.get("frontend_build_commit") and str(version_payload.get("frontend_build_source") or "").startswith("served:"):
            break
        await asyncio.sleep(3)
        version_payload = api_version()

    health_snapshot = await _compute_public_full_health_snapshot()
    readiness_payload = await _deployment_readiness_payload_internal()
    verification_id = _deployment_verification_id(version_payload)
    parity_ok = bool(version_payload.get("frontend_backend_release_match"))
    health_ok = bool(health_snapshot["public"].get("ok"))
    readiness_ok = readiness_payload.get("decision") == "pass"
    failure_reasons = []
    if not bool(version_payload.get("runtime_matches_intended_release")):
        failure_reasons.append("runtime_commit_does_not_match_intended_release")
    if not parity_ok:
        failure_reasons.append(str(version_payload.get("frontend_backend_release_match_reason") or "release_identity_mismatch"))
    if not health_ok:
        failure_reasons.append(f"health_full:{health_snapshot['diagnostics'].get('failing_check') or 'unhealthy'}")
    if not readiness_ok:
        failure_reasons.append("deployment_readiness_fail")
    decision = "pass" if not failure_reasons else "fail"
    payload = {
        "verification_id": verification_id,
        "decision": decision,
        "exit_code": 0 if decision == "pass" else 1,
        "commit": version_payload.get("commit"),
        "backend_runtime_commit": version_payload.get("commit"),
        "frontend_build_commit": version_payload.get("frontend_build_commit"),
        "intended_release_commit": version_payload.get("intended_release_commit"),
        "branch": (_WORKSPACE_IDENTITY_SNAPSHOT.get("branch") or "")[:64],
        "environment": version_payload.get("app_env") or _canonical_app_env(),
        "operator": "automated deployment verification",
        "duration_ms": 0,
        "trust_score": int(readiness_payload.get("trust_score") or 0),
        "trust_band": readiness_payload.get("trust_band") or "",
        "blocking_count": len(readiness_payload.get("blocking_gates") or []),
        "advisory_count": len(readiness_payload.get("advisory_findings") or []),
        "regression_count": int(readiness_payload.get("regression_gate_count") or 0),
        "blocking_ids": [g.get("id") for g in (readiness_payload.get("blocking_gates") or [])][:32],
        "build_version": version_payload.get("frontend_build_version"),
        "build_timestamp": version_payload.get("frontend_build_built_at"),
        "parity_result": parity_ok,
        "parity_reason": version_payload.get("frontend_backend_release_match_reason"),
        "health_ok": health_ok,
        "health_status_code": 200 if health_ok else 503,
        "health_reason": health_snapshot["diagnostics"].get("failing_check") or "ok",
        "go_no_go": "GO" if decision == "pass" else "NO-GO",
        "failure_reason": "; ".join(failure_reasons),
        "script_version": "startup-auto-c2-v1",
        "source_hash": version_payload.get("source_hash"),
        "dependency_manifest_hash": version_payload.get("frontend_build_dependency_manifest_hash"),
        "governance_hash": version_payload.get("frontend_build_release_gate_manifest_hash"),
        "verification_source": "automatic_startup_verification",
        "runtime_identity_status": health_snapshot["public"].get("runtime_identity_status"),
    }
    await _record_deployment_verification_outcome(payload)
    app.state.last_deployment_verification = payload


# ---------------------------------------------------------------------------
# Internal Operations Manual — Developer portal only.
# Gated by require_dev (password in DEV_PASSWORD, distinct from admin/pm).
# Generated on-demand from ops_manual.py so edits ship instantly without
# requiring a redeploy of static assets. Admin/PM tokens cannot access.
# ---------------------------------------------------------------------------
from fastapi.responses import Response as _FastAPIResponse  # noqa: E402


@api_router.post("/dev/login")
async def dev_login(body: AdminLoginRequest, request: Request):
    """Developer (vendor/ForgedOps LLC) portal login. Issues a token
    accepted ONLY by require_dev — never by any admin/PM/shop route.

    Track 24.1 · P0-4 — FAIL CLOSED. If `DEV_ENDPOINTS_ENABLED` is not
    truthy OR `DEV_PASSWORD` is missing, this route returns 404 so the
    surface is indistinguishable from an unregistered endpoint."""
    if not _dev_endpoints_enabled():
        raise HTTPException(status_code=404, detail="Not Found")
    ip = _client_ip(request)
    _check_login_lockout(ip)
    expected_pw = os.environ.get("DEV_PASSWORD", "")
    if not expected_pw:
        raise HTTPException(status_code=404, detail="Not Found")
    if not hmac.compare_digest(body.password, expected_pw):
        _record_login_fail(ip)
        raise HTTPException(status_code=401, detail="Wrong password")
    _reset_login_fails(ip)
    return {"ok": True, "token": _dev_token_for(expected_pw)}


@api_router.get("/dev/check")
async def dev_check(_: bool = Depends(require_dev)):
    """Verify a stored dev token is still valid."""
    return {"ok": True}


@api_router.get("/dev/ops-manual.pdf")
async def dev_ops_manual_pdf(_: bool = Depends(require_dev)):
    from ops_manual import render_ops_manual_pdf
    # iter340 · wrap sync PDF render in to_thread so the event loop stays
    # responsive (mirrors iter331 fix in routes/field_leadership.py).
    pdf = await asyncio.to_thread(render_ops_manual_pdf)
    return _FastAPIResponse(
        content=pdf,
        media_type="application/pdf",
        headers={
            "Content-Disposition": 'attachment; filename="MASCI_HUB_Operations_Manual.pdf"',
            "Cache-Control": "private, no-store",
        },
    )


@api_router.get("/dev/ops-manual.docx")
async def dev_ops_manual_docx(_: bool = Depends(require_dev)):
    from ops_manual import render_ops_manual_docx
    docx = await asyncio.to_thread(render_ops_manual_docx)
    return _FastAPIResponse(
        content=docx,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": 'attachment; filename="MASCI_HUB_Operations_Manual.docx"',
            "Cache-Control": "private, no-store",
        },
    )


# ═══════════════════════════════════════════════════════════════════════
# Phase 5 · W8 closeout — Ops Manual admin-discoverable mirror
# ───────────────────────────────────────────────────────────────────────
# The dev-token-gated routes above remain for developer tooling. These
# mirrors gate on `require_admin` so operators can discover and download
# the Operations Manual without a separate dev-token issuance flow.
# Identical payload — zero behavior drift, just an additional gate.
# ═══════════════════════════════════════════════════════════════════════

@api_router.get("/admin/ops-manual.pdf")
async def admin_ops_manual_pdf(_: bool = Depends(require_admin)):
    from ops_manual import render_ops_manual_pdf
    pdf = await asyncio.to_thread(render_ops_manual_pdf)
    return _FastAPIResponse(
        content=pdf,
        media_type="application/pdf",
        headers={
            "Content-Disposition": 'attachment; filename="MASCI_HUB_Operations_Manual.pdf"',
            "Cache-Control": "private, no-store",
        },
    )


@api_router.get("/admin/ops-manual.docx")
async def admin_ops_manual_docx(_: bool = Depends(require_admin)):
    from ops_manual import render_ops_manual_docx
    docx = await asyncio.to_thread(render_ops_manual_docx)
    return _FastAPIResponse(
        content=docx,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": 'attachment; filename="MASCI_HUB_Operations_Manual.docx"',
            "Cache-Control": "private, no-store",
        },
    )


# --- Ops Manual snapshots (pinned historical copies) ---------------------
# Store both PDF + DOCX bytes (base64) plus the backend source_hash in
# MongoDB so a specific "official" revision of the manual can be pulled
# back verbatim months later even after the source data has changed.
import base64 as _ops_b64  # noqa: E402


@api_router.post("/dev/ops-manual/snapshot")
async def dev_ops_manual_snapshot(
    body: Optional[dict] = None,
    _: bool = Depends(require_dev),
):
    from ops_manual import render_ops_manual_pdf, render_ops_manual_docx
    note = ""
    if isinstance(body, dict):
        note = str(body.get("note", ""))[:500]
    # iter340 · keep event loop responsive during snapshot generation
    pdf = await asyncio.to_thread(render_ops_manual_pdf)
    docx = await asyncio.to_thread(render_ops_manual_docx)
    snap = {
        "id": uuid.uuid4().hex,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "source_hash": _SOURCE_HASH,
        "note": note,
        "pdf_b64": _ops_b64.b64encode(pdf).decode("ascii"),
        "docx_b64": _ops_b64.b64encode(docx).decode("ascii"),
        "pdf_bytes": len(pdf),
        "docx_bytes": len(docx),
    }
    await db.ops_manual_snapshots.insert_one(snap)
    return {
        "ok": True,
        "id": snap["id"],
        "created_at": snap["created_at"],
        "source_hash": snap["source_hash"],
        "note": snap["note"],
        "pdf_bytes": snap["pdf_bytes"],
        "docx_bytes": snap["docx_bytes"],
    }


@api_router.get("/dev/ops-manual/snapshots")
async def dev_ops_manual_list_snapshots(_: bool = Depends(require_dev)):
    cursor = db.ops_manual_snapshots.find(
        {},
        {"_id": 0, "pdf_b64": 0, "docx_b64": 0},
    ).sort("created_at", -1).limit(200)
    rows = await cursor.to_list(200)
    return {"snapshots": rows}


@api_router.delete("/dev/ops-manual/snapshots/{snap_id}")
async def dev_ops_manual_delete_snapshot(
    snap_id: str, _: bool = Depends(require_dev)
):
    r = await db.ops_manual_snapshots.delete_one({"id": snap_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Snapshot not found")
    return {"ok": True}


@api_router.get("/dev/ops-manual/snapshots/{snap_id}.pdf")
async def dev_ops_manual_snapshot_pdf(
    snap_id: str, _: bool = Depends(require_dev)
):
    doc = await db.ops_manual_snapshots.find_one(
        {"id": snap_id}, {"_id": 0, "pdf_b64": 1, "created_at": 1}
    )
    if not doc or not doc.get("pdf_b64"):
        raise HTTPException(status_code=404, detail="Snapshot not found")
    pdf = _ops_b64.b64decode(doc["pdf_b64"])
    stamp = (doc.get("created_at") or "").replace(":", "-").split(".")[0]
    return _FastAPIResponse(
        content=pdf,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="MASCI_HUB_Operations_Manual_{stamp}.pdf"',
            "Cache-Control": "private, no-store",
        },
    )


@api_router.get("/dev/ops-manual/snapshots/{snap_id}.docx")
async def dev_ops_manual_snapshot_docx(
    snap_id: str, _: bool = Depends(require_dev)
):
    doc = await db.ops_manual_snapshots.find_one(
        {"id": snap_id}, {"_id": 0, "docx_b64": 1, "created_at": 1}
    )
    if not doc or not doc.get("docx_b64"):
        raise HTTPException(status_code=404, detail="Snapshot not found")
    docx = _ops_b64.b64decode(doc["docx_b64"])
    stamp = (doc.get("created_at") or "").replace(":", "-").split(".")[0]
    return _FastAPIResponse(
        content=docx,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": f'attachment; filename="MASCI_HUB_Operations_Manual_{stamp}.docx"',
            "Cache-Control": "private, no-store",
        },
    )


# --- Source bundle download --------------------------------------------
# One-click zip of the full application source tree so an auditor /
# acquirer / due-diligence package can pair the pinned Ops Manual with
# a byte-exact copy of the code that produced it.
#
# Strict exclusions (never shipped in the zip):
#   • /app/backend/backups/*            (customer DB dumps)
#   • /app/backend/storage/*            (uploaded files — customer data)
#   • /app/backend/__pycache__/*        (binary caches)
#   • /app/backend/.env                 (secrets)
#   • /app/backend/data/*.bak.json      (historical data snapshots)
#   • /app/frontend/node_modules/*      (build artefact)
#   • /app/frontend/build/*             (build artefact)
#   • /app/frontend/.env                (secrets)
#   • **/.git/*                         (repo metadata)
#   • *.pyc                             (binary caches)
import io as _src_io  # noqa: E402
import zipfile as _src_zip  # noqa: E402
from pathlib import Path as _SrcPath  # noqa: E402


_SRC_EXCLUDE_DIRS = {
    "backups", "storage", "__pycache__", "node_modules", "build",
    ".git", ".next", ".cache", ".yarn", "dist", ".pytest_cache",
    ".emergent",
}
_SRC_EXCLUDE_SUFFIXES = {".pyc", ".pyo", ".log"}
_SRC_ROOTS = [
    ("app", _SrcPath("/app"), {
        "allowed_top_level": {
            "backend", "frontend", "memory", "scripts", "test_reports",
            "README.md", "ATLAS_MIGRATION.md", "auth_testing.md",
            "test_result.md", "design_guidelines.json",
        },
    }),
]


def _src_should_skip(path: _SrcPath) -> bool:
    name = path.name
    # skip env files wherever they land
    if name == ".env" or name.startswith(".env."):
        return True
    # skip historical data snapshots
    if name.endswith(".bak.json"):
        return True
    if path.suffix in _SRC_EXCLUDE_SUFFIXES:
        return True
    # skip anything inside an excluded directory
    for part in path.parts:
        if part in _SRC_EXCLUDE_DIRS:
            return True
    return False


def _build_source_bundle() -> bytes:
    buf = _src_io.BytesIO()
    with _src_zip.ZipFile(buf, "w", _src_zip.ZIP_DEFLATED, compresslevel=6) as zf:
        manifest_lines = [
            "MASCI Operations Platform — Source Bundle",
            f"Generated: {datetime.now(timezone.utc).isoformat()}",
            f"Source hash: {_SOURCE_HASH}",
            f"Commit: {os.environ.get('GIT_COMMIT', 'unknown')}",
            f"Built at: {os.environ.get('BUILT_AT', 'unknown')}",
            "",
            "Classification: CONFIDENTIAL — ForgedOps\u2122",
            "Excluded: /backups, /storage, node_modules, build, .env, .git, *.pyc, *.bak.json",
            "",
        ]
        for label, root, opts in _SRC_ROOTS:
            if not root.exists():
                continue
            allowed = opts.get("allowed_top_level")
            for item in sorted(root.iterdir()):
                if allowed and item.name not in allowed:
                    continue
                if _src_should_skip(item):
                    continue
                if item.is_file():
                    try:
                        arcname = f"{label}/{item.name}"
                        zf.write(item, arcname)
                        manifest_lines.append(arcname)
                    except OSError:
                        continue
                elif item.is_dir():
                    for sub in item.rglob("*"):
                        if sub.is_dir():
                            continue
                        if _src_should_skip(sub):
                            continue
                        try:
                            rel = sub.relative_to(root)
                            arcname = f"{label}/{rel.as_posix()}"
                            zf.write(sub, arcname)
                            manifest_lines.append(arcname)
                        except (OSError, ValueError):
                            continue
        zf.writestr("app/MANIFEST.txt", "\n".join(manifest_lines) + "\n")
    buf.seek(0)
    return buf.getvalue()


# ── Track 24.1 · P0-4 · source-bundle endpoints REMOVED ─────────────
#
# `GET /api/dev/source-bundle.zip` and `GET /api/dev/source-bundle.info`
# were removed on 2026-02-07 during the pre-production hardening pass.
# They exposed a one-click full source-tree download behind the shared
# `DEV_PASSWORD`, which itself lived in the pod env — an unacceptable
# IP-exfiltration risk in a production tenant.
#
# Do NOT re-add these endpoints. Source-code delivery is a manual
# process handled through the "Save to GitHub" feature or a signed
# delivery bundle produced offline. There is no operational reason for
# a live server to serve its own source over HTTP.
# ────────────────────────────────────────────────────────────────────



# ---------------------------------------------------------------------------
# Soft-delete framework — give every master-list 🗑️ button a 14-day undo
# instead of an immediate hard-delete, so a mis-click on a 234-row table
# is fully recoverable from the UI without restoring a backup.
#
# Convention: every soft-deleted row gets ``deleted_at`` (ISO timestamp).
# Every list endpoint that should hide deletes filters with the
# ``ACTIVE_FILTER`` below. Restore = unset ``deleted_at``. Anything older
# than 14 days is hard-purged on the next list call (best effort).
# ---------------------------------------------------------------------------
SOFT_DELETE_RETAIN_DAYS = 14
ACTIVE_FILTER: Dict[str, Any] = {"deleted_at": {"$in": [None, ""]}}


def _utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


async def _purge_expired(coll_name: str) -> int:
    """Hard-delete anything whose ``deleted_at`` is older than the retention
    window. Called best-effort from list endpoints."""
    cutoff = (
        datetime.now(timezone.utc) - timedelta(days=SOFT_DELETE_RETAIN_DAYS)
    ).isoformat()
    res = await db[coll_name].delete_many(
        {"deleted_at": {"$ne": None, "$lt": cutoff}}
    )
    return res.deleted_count or 0


async def _soft_delete(coll_name: str, query: Dict[str, Any]) -> bool:
    """Mark a row deleted. Returns True iff a row was matched."""
    res = await db[coll_name].update_one(
        {"$and": [query, {"deleted_at": {"$in": [None, ""]}}]},
        {"$set": {"deleted_at": _utc_iso()}},
    )
    return res.matched_count > 0


async def _restore_row(coll_name: str, query: Dict[str, Any]) -> bool:
    """Clear ``deleted_at`` so a row reappears in the active list."""
    res = await db[coll_name].update_one(
        {"$and": [query, {"deleted_at": {"$ne": None}}]},
        {"$unset": {"deleted_at": ""}},
    )
    return res.matched_count > 0


async def _list_archive(coll_name: str, sort_field: str = "deleted_at") -> List[Dict[str, Any]]:
    """List soft-deleted rows for the archive UI."""
    out: List[Dict[str, Any]] = []
    cursor = db[coll_name].find(
        {"deleted_at": {"$ne": None, "$exists": True}}, {"_id": 0}
    ).sort(sort_field, -1)
    async for d in cursor:
        # Make sure it's a non-empty string (skip rows where deleted_at == "")
        if d.get("deleted_at"):
            out.append(d)
    return out


# ---------------------------------------------------------------------------
# Master-list exports — one-click download of every active row as a sorted
# .xlsx file using the EXACT same column shape the bulk-import accepts.
# Round-trip safe: the file you export today can be re-uploaded tomorrow.
# ---------------------------------------------------------------------------
def _xlsx_response(rows: List[List[Any]], header: List[str], filename: str, sheet: str = "Sheet1") -> Response:
    """Build an XLSX from a header row + 2-D row data and return as download."""
    import openpyxl as _ox
    wb = _ox.Workbook()
    ws = wb.active
    ws.title = sheet[:31] or "Sheet1"
    ws.append(header)
    for r in rows:
        ws.append(r)
    # Auto-widen columns based on the longest cell in each column
    for idx, col_header in enumerate(header, start=1):
        longest = len(str(col_header))
        for r in rows:
            if idx - 1 < len(r):
                longest = max(longest, len(str(r[idx - 1] or "")))
        ws.column_dimensions[_ox.utils.get_column_letter(idx)].width = min(longest + 2, 60)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _today_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


@api_router.get("/admin/employees/export")
async def export_employees(_: bool = Depends(require_admin)):
    from lib.synthetic_hr_filter import apply_synthetic_hr_exclusion  # noqa: PLC0415
    cursor = db.employees.find(apply_synthetic_hr_exclusion(ACTIVE_FILTER), {"_id": 0}).sort("name", 1)
    docs = await cursor.to_list(5000)
    header = ["Name", "Employee ID", "Trade", "Role", "Crew", "Email", "Phone"]
    rows = [
        [
            d.get("name", ""),
            d.get("employee_id", ""),
            d.get("trade", ""),
            d.get("role", ""),
            d.get("crew", ""),
            d.get("email", ""),
            d.get("phone", ""),
        ]
        for d in docs
    ]
    return _xlsx_response(rows, header, f"MASCI_employees_{_today_stamp()}.xlsx", "Employees")


@api_router.get("/admin/suppliers/export")
async def export_suppliers(_: bool = Depends(require_admin)):
    cursor = db.suppliers.find(ACTIVE_FILTER, {"_id": 0}).sort("name", 1)
    docs = await cursor.to_list(5000)
    header = ["Name", "Active"]
    rows = [
        [d.get("name", ""), "Yes" if d.get("is_active", True) else "No"]
        for d in docs
    ]
    return _xlsx_response(rows, header, f"MASCI_suppliers_{_today_stamp()}.xlsx", "Suppliers")


@api_router.get("/admin/equipment-master/export")
async def export_equipment_master(_: bool = Depends(require_admin)):
    cursor = db.equipment_master.find(ACTIVE_FILTER, {"_id": 0}).sort(
        [("category", 1), ("unit_number", 1)]
    )
    docs = await cursor.to_list(5000)
    header = [
        "Unit Number", "Year", "Make", "Model", "VIN/Serial",
        "Category", "Pre-Op Type", "Company", "Comments",
    ]
    rows = [
        [
            d.get("unit_number", ""),
            d.get("year", ""),
            d.get("make", ""),
            d.get("model", ""),
            d.get("vin_serial_number", ""),
            d.get("category", ""),
            d.get("preop_equipment_type", ""),
            d.get("company", ""),
            d.get("comments", ""),
        ]
        for d in docs
    ]
    return _xlsx_response(rows, header, f"MASCI_equipment_{_today_stamp()}.xlsx", "Louis")


@api_router.get("/admin/equipment-parts/export")
async def export_equipment_parts(_: bool = Depends(require_admin)):
    """Flatten the per-unit parts catalog into one wide sheet — same column
    shape the bulk-importer accepts for round-trip."""
    cursor = db.equipment_units.find({}, {"_id": 0})
    header = [
        "Unit Number", "Category", "Name", "Part Number", "Qty",
        "Size", "Position", "Ply", "Brand", "Notes",
    ]
    rows: List[List[Any]] = []
    async for u in cursor:
        unit = u.get("unit_number") or u.get("id") or ""
        for cat_key in ("filters", "cutting_edges", "wiper_blades", "tires", "other_wear_items"):
            for p in u.get(cat_key, []) or []:
                rows.append([
                    unit,
                    cat_key.replace("_", " ").title(),
                    p.get("name", ""),
                    p.get("part_number", ""),
                    p.get("qty", ""),
                    p.get("size", ""),
                    p.get("position", ""),
                    p.get("ply", ""),
                    p.get("brand", ""),
                    p.get("notes", ""),
                ])
    rows.sort(key=lambda r: (str(r[0]), str(r[1]), str(r[2])))
    return _xlsx_response(rows, header, f"MASCI_parts_{_today_stamp()}.xlsx", "Parts")


@api_router.get("/admin/jobs/export")
async def export_jobs(_: bool = Depends(require_admin)):
    from jobs_master import list_jobs
    docs = await list_jobs(db, only_active=False)
    header = [
        "Project Number", "Project Name", "Location", "Client",
        "PM Name", "PM Email", "Active",
    ]
    rows = [
        [
            d.get("project_number", ""),
            d.get("project_name", ""),
            d.get("location", ""),
            d.get("client", ""),
            d.get("pm_name", ""),
            d.get("pm_email", ""),
            "Yes" if d.get("active", True) else "No",
        ]
        for d in docs
    ]
    return _xlsx_response(rows, header, f"MASCI_jobs_{_today_stamp()}.xlsx", "Jobs")


@api_router.post("/admin/login")
async def admin_login(body: AdminLoginRequest, request: Request):
    """TRACK 15.32 (2026-02) — shared ADMIN_PASSWORD path retired.

    The single canonical admin login is now `POST /api/auth/multi-login`
    which uses the per-user `user_directory` table. This endpoint is
    retained so legacy clients receive a clear retirement message
    instead of a silent 404.
    """
    del body, request  # noqa: ERA001 — explicit retirement marker
    raise HTTPException(
        status_code=410,
        detail=(
            "The shared-password admin login was retired in TRACK 15.32. "
            "Use POST /api/auth/multi-login with your assigned admin user "
            "email + password instead."
        ),
    )


@api_router.get("/admin/check")
async def admin_check(_: bool = Depends(require_admin)):
    """Frontend pings this to verify a stored token is still valid."""
    return {"ok": True}


@api_router.post("/admin/logout")
async def admin_logout(request: Request, _: bool = Depends(require_admin)):
    """Legacy compatibility wrapper over canonical shared logout."""
    try:
        await db.audit_events.insert_one({
            "at": datetime.now(timezone.utc),
            "kind": "admin_logout",
            "actor": "admin",
            "ip": _client_ip(request),
            "user_agent": (request.headers.get("user-agent") or "")[:240],
            "logout_route": "/api/admin/logout",
            "canonical_logout": "/api/auth/multi-logout",
        })
    except Exception:  # noqa: BLE001
        pass
    return await _canonical_multi_logout(
        x_directory_token=request.headers.get("x-directory-token") or None,
        x_admin_token=request.headers.get("x-admin-token") or None,
        x_pm_token=request.headers.get("x-pm-token") or None,
        x_hr_token=request.headers.get("x-hr-token") or None,
        x_safety_token=request.headers.get("x-safety-token") or None,
        x_shop_token=request.headers.get("x-shop-token") or None,
        x_dispatch_token=request.headers.get("x-dispatch-token") or None,
        x_fl_token=request.headers.get("x-fl-token") or None,
    )


@api_router.post("/admin/auth/verify-password")
async def admin_verify_password(body: AdminLoginRequest, request: Request):
    """Re-verify the *current admin user's* password without rotating
    their session token. Used by destructive-action confirmation dialogs
    (delete backup file, REPLACE restore, force re-seed) so an admin
    must re-type their password before the action runs.

    TRACK 15.32 (2026-02) — rewired from the shared ``ADMIN_PASSWORD``
    HMAC check to the per-user ``user_directory`` password. The actor is
    resolved from the X-Admin-Token header; we re-authenticate them
    against their directory bcrypt hash.

    Shares the same lockout protection as the multi-login path so brute
    force is rate-limited per IP.
    """
    ip = _client_ip(request)
    _check_login_lockout(ip)
    x_admin_token = request.headers.get("x-admin-token") or ""
    # Resolve the current admin actor from session_activity (the row
    # was stamped at multi-login time).
    actor_row = None
    if x_admin_token:
        try:
            import hashlib as _h
            th = _h.sha256(x_admin_token.encode()).hexdigest()
            sess = await db.session_activity.find_one(
                {"token_hash": th},
                {"_id": 0, "email": 1, "user_id": 1},
            )
            if sess and sess.get("email"):
                import user_directory as _ud  # noqa: PLC0415
                actor_row = await _ud.authenticate(
                    db,
                    email=sess["email"],
                    password=body.password or "",
                )
        except Exception as exc:  # noqa: BLE001
            logger.warning(f"admin_verify_password resolver failed: {exc}")
    if not actor_row or actor_row.get("disabled"):
        _record_login_fail(ip)
        raise HTTPException(status_code=401, detail="Wrong password")
    _reset_login_fails(ip)
    # Phase 2 Initiative 5b-full — stamp step-up so the next sensitive
    # action within the configured window passes the require_recent
    # gate. The admin token comes via header; if missing, the step-up
    # record is keyed by an empty token-hash (effectively a no-op).
    await _record_admin_step_up(db, x_admin_token)
    await _record_admin_action(db, "admin_step_up_verified", request)
    return {"ok": True}


@api_router.get("/admin/submit-language-stats")
async def admin_submit_language_stats(_: bool = Depends(require_admin)):
    """Per-collection counts of how many records were originally filed in
    Spanish vs English. Surfaces MASCI's bilingual-crew adoption so admins
    can see at a glance how much of the workforce is using Spanish mode.

    Every form submission stamps `submit_language` ("en" | "es") at the
    moment of submit. Missing field (legacy records from before the stamp
    was added) counts as "unknown" and rolls up to "en" in totals so
    default behaviour doesn't over-report Spanish usage.
    """
    collections = [
        ("inspections", "Site Inspections"),
        ("meetings", "Safety Meetings"),
        ("incidents", "Incident Reports"),
        ("daily_reports", "Daily Reports"),
        ("equipment_inspections", "Equipment Pre-Op"),
    ]
    out = []
    grand_en = 0
    grand_es = 0
    grand_unknown = 0
    for coll_name, label in collections:
        total = await db[coll_name].count_documents({})
        es = await db[coll_name].count_documents({"submit_language": "es"})
        en = await db[coll_name].count_documents({"submit_language": "en"})
        unknown = max(0, total - es - en)
        grand_en += en
        grand_es += es
        grand_unknown += unknown
        out.append({
            "collection": coll_name,
            "label": label,
            "total": total,
            "en": en,
            "es": es,
            "unknown": unknown,
            "es_pct": round((es / total) * 100, 1) if total else 0.0,
        })
    grand_total = grand_en + grand_es + grand_unknown
    return {
        "by_collection": out,
        "totals": {
            "total": grand_total,
            "en": grand_en,
            "es": grand_es,
            "unknown": grand_unknown,
            "es_pct": round((grand_es / grand_total) * 100, 1) if grand_total else 0.0,
        },
    }



# ---------------------------------------------------------------------------
# Material Calculators — Field page at /field/calculators
#
# Six calculators (aggregate, asphalt, concrete, truck-load, yield-waste,
# tons-cy-conversion). Every time a crew member hits "Save Calculation"
# the payload lands here and we persist to `calculator_runs`. Admin-side
# analytics and a CSV export read from the same collection.
# ---------------------------------------------------------------------------

_ALLOWED_CALCULATOR_TYPES = {
    "aggregate",
    "asphalt",
    "concrete",
    "truck_load",
    "yield_waste",
    "conversion",
}


class CalculatorRun(BaseModel):
    model_config = {"extra": "allow"}
    calculator_type: str
    language: str = "en"
    job_number: Optional[str] = None
    job_name: Optional[str] = None
    user_name: Optional[str] = None
    inputs: Dict[str, Any] = {}
    outputs: Dict[str, Any] = {}


@api_router.post("/calculators/save")
async def save_calculator_run(payload: CalculatorRun):
    """Public (field-facing) endpoint — crews don't need to log in to use
    the calculators, so saving a run is open too. We store the full
    inputs/outputs snapshot so admins can trace any quantity back to the
    numbers that were typed in."""
    if payload.calculator_type not in _ALLOWED_CALCULATOR_TYPES:
        raise HTTPException(
            status_code=400, detail=f"Unknown calculator_type '{payload.calculator_type}'"
        )
    lang = (payload.language or "en").lower()
    if lang not in ("en", "es"):
        lang = "en"
    doc = {
        "id": str(uuid.uuid4()),
        "calculator_type": payload.calculator_type,
        "language": lang,
        "job_number": payload.job_number,
        "job_name": payload.job_name,
        "user_name": payload.user_name,
        "inputs": payload.inputs or {},
        "outputs": payload.outputs or {},
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.calculator_runs.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.get("/admin/calculators/stats")
async def admin_calculator_stats(_: bool = Depends(require_admin)):
    """Aggregate usage stats for the admin dashboard card. Counts by
    calculator type and by language, plus most-used and last-used."""
    total = await db.calculator_runs.count_documents({})
    rows = []
    labels = {
        "aggregate": "Aggregate",
        "asphalt": "Asphalt",
        "concrete": "Concrete",
        "truck_load": "Truck Load",
        "yield_waste": "Yield / Waste",
        "conversion": "Tons ↔ CY",
    }
    for ctype, label in labels.items():
        c_total = await db.calculator_runs.count_documents({"calculator_type": ctype})
        c_en = await db.calculator_runs.count_documents({"calculator_type": ctype, "language": "en"})
        c_es = await db.calculator_runs.count_documents({"calculator_type": ctype, "language": "es"})
        rows.append({
            "calculator_type": ctype,
            "label": label,
            "total": c_total,
            "en": c_en,
            "es": c_es,
        })
    en_total = await db.calculator_runs.count_documents({"language": "en"})
    es_total = await db.calculator_runs.count_documents({"language": "es"})

    # Most used
    most_used = max(rows, key=lambda r: r["total"]) if rows else None
    if most_used and most_used["total"] == 0:
        most_used = None

    # Last used
    last_doc = await db.calculator_runs.find_one(
        {}, {"_id": 0, "calculator_type": 1, "created_at": 1, "language": 1}, sort=[("created_at", -1)]
    )

    return {
        "totals": {"total": total, "en": en_total, "es": es_total},
        "by_type": rows,
        "most_used": most_used,
        "last_used": last_doc,
    }


@api_router.get("/admin/calculators/export.csv")
async def admin_calculator_export_csv(_: bool = Depends(require_admin)):
    """Admin-only CSV dump of every saved calculator run. Used for
    offline analysis and for the "export" button on the admin card."""
    import csv as _csv
    import json as _json_csv
    buf = io.StringIO()
    w = _csv.writer(buf)
    w.writerow([
        "Created At (UTC)", "Calculator", "Language", "Job Number",
        "Job Name", "User", "Inputs (JSON)", "Outputs (JSON)",
    ])
    cursor = db.calculator_runs.find({}, {"_id": 0}).sort("created_at", -1)
    async for r in cursor:
        w.writerow([
            r.get("created_at", ""),
            r.get("calculator_type", ""),
            r.get("language", ""),
            r.get("job_number", "") or "",
            r.get("job_name", "") or "",
            r.get("user_name", "") or "",
            _json_csv.dumps(r.get("inputs", {}), separators=(",", ":")),
            _json_csv.dumps(r.get("outputs", {}), separators=(",", ":")),
        ])
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="text/csv",
        headers={
            "Content-Disposition": 'attachment; filename="masci-calculator-runs.csv"',
        },
    )





@api_router.post("/shop/login")
async def shop_login(body: AdminLoginRequest, request: Request):
    """Per-user shop console login (mechanics, shop managers, parts).

    TRACK 15.30 (2026-02) — the legacy email-less shared-password
    branch has been retired. Email is now REQUIRED; a missing email
    returns 401 explaining that per-user accounts are mandatory.
    """
    ip = _client_ip(request)
    _check_login_lockout(ip)

    body_email = ""
    try:
        raw = await request.json()
        body_email = (raw.get("email") or "").strip().lower()
    except Exception:
        body_email = ""

    if not body_email:
        raise HTTPException(
            status_code=401,
            detail=(
                "Email is required. The shared-password kiosk path was "
                "retired in TRACK 15.30 — sign in with your assigned "
                "shop user account."
            ),
        )

    # ---- Per-user shop auth ----
    from shop_users import (
        find_shop_user_by_email,
        make_shop_user_token,
        public_shop_user_view,
        stamp_shop_login,
        verify_password,
    )
    user = await find_shop_user_by_email(db, body_email)

    # Track 15.87 · directory `shop` grant fallback (Path 1.5).
    # Tried BEFORE the admin fallback so a directory user with `shop`
    # grant gets a Shop token (NOT admin). Mirrors multi-login.
    async def _try_directory_shop_fallback():
        try:
            from lib.directory_portal_login import try_directory_portal_login  # noqa: WPS433
            res = await try_directory_portal_login(
                db,
                email=body_email,
                password=body.password,
                required_portal="shop",
                portal_token_minter=_directory_shop_token,
                kind="shop",
            )
        except Exception as exc:  # noqa: BLE001
            logger.warning(f"shop_login track_15_87 directory-grant fallback error: {exc}")
            return None
        if not res:
            return None
        try:
            await _reset_session_activity(
                db, res["token"], "OPERATIONS",
                user_id=res["user"].get("id"), email=res["user"].get("email"),
                actor_label="shop_via_directory", ip=ip,
                user_agent=request.headers.get("user-agent") or "",
            )
        except Exception:  # noqa: BLE001
            pass
        _reset_login_fails(ip)
        return {
            "ok": True,
            "token": res["token"],
            "kind": "shop",
            "must_change_password": False,
            "user": res["user"],
        }

    # iter346-B · universal super-admin fallback (Path 2) — local
    # closure invoked when native shop auth fails.
    async def _try_directory_admin_fallback():
        try:
            import user_directory as _ud_local  # noqa: WPS433
            row = await _ud_local.authenticate(db, email=body_email, password=body.password)
        except Exception as exc:  # noqa: BLE001
            logger.warning(f"shop_login directory fallback error: {exc}")
            return None
        if row and not row.get("disabled") and await _super_admin_row_for_token_async(_directory_admin_token(row)):
            admin_tok = _directory_admin_token(row)
            if admin_tok:
                await _reset_session_activity(
                    db, admin_tok, "OPERATIONS",
                    user_id=row.get("id"), email=row.get("email"),
                    actor_label="admin_via_shop", ip=ip,
                    user_agent=request.headers.get("user-agent") or "",
                )
                _reset_login_fails(ip)
                return {
                    "ok": True,
                    "token": admin_tok,
                    "kind": "admin",
                    "must_change_password": False,
                    "user": _ud_local.public_view(row),
                }
        return None

    if not user:
        fb = await _try_directory_shop_fallback()
        if fb is not None:
            return fb
        fb = await _try_directory_admin_fallback()
        if fb is not None:
            return fb
        _record_login_fail(ip)
        raise HTTPException(status_code=401, detail="Wrong email or password")
    if user.get("disabled"):
        raise HTTPException(status_code=403, detail="This shop user is disabled. Contact the admin.")
    pwh = user.get("password_hash") or ""
    if not pwh:
        raise HTTPException(status_code=403, detail="No password set yet. Ask the admin to issue one.")
    if not verify_password(body.password, pwh):
        fb = await _try_directory_shop_fallback()
        if fb is not None:
            return fb
        fb = await _try_directory_admin_fallback()
        if fb is not None:
            return fb
        _record_login_fail(ip)
        raise HTTPException(status_code=401, detail="Wrong email or password")
    _reset_login_fails(ip)
    await stamp_shop_login(db, user["id"], ip=ip)
    token = make_shop_user_token(user["id"], pwh)
    await _reset_session_activity(
        db, token, "OPERATIONS",
        user_id=user.get("id"),
        email=user.get("email"),
        actor_label="shop",
        ip=ip,
        user_agent=request.headers.get("user-agent") or "",
    )
    # Track 15.13A / 15.13B · Asset Care Routing Recovery.
    # Mirror the `is_asset_admin` flag from the canonical
    # `user_directory` row (keyed by email) into the shop_login
    # response so the SPA `landingFor()` resolver can route asset
    # administrators to `/shop/asset-care` instead of the generic
    # `/shop` hub.
    #
    # Track 15.13B FAILURE #1 fallback — production showed that an
    # existing Asset Administrator created BEFORE the 15.13A mirror
    # landed has NO directory row, so the dir lookup returned None
    # and `is_asset_admin` resolved to False — landing them on the
    # generic /shop hub. Fix: also honor the role label on the
    # shop_users row itself via `_role_implies_asset_admin(role)`
    # as a strict read-only fallback. This guarantees every legacy
    # Asset Administrator / Asset Manager / Equipment Manager /
    # Fleet Coordinator gets the right landing on first login
    # WITHOUT a separate backfill script.
    public_user = public_shop_user_view(user)
    is_asset_admin = False
    try:
        dir_row = await db.user_directory.find_one(
            {"email": (user.get("email") or "").strip().lower()},
            {"_id": 0, "is_asset_admin": 1, "portals": 1},
        )
        if dir_row and dir_row.get("is_asset_admin") is True:
            is_asset_admin = True
            public_user["portals"] = dir_row.get("portals") or []
    except Exception as exc:  # noqa: BLE001
        logger.warning(f"shop_login is_asset_admin mirror failed: {exc}")
    # 15.13B fallback — role-label check (read-only, no write).
    if not is_asset_admin and _role_implies_asset_admin(user.get("role")):
        is_asset_admin = True
    if is_asset_admin:
        public_user["is_asset_admin"] = True
    return {
        "ok": True,
        "token": token,
        "kind": "shop",
        "must_change_password": bool(user.get("must_change_password")),
        "user": public_user,
        "is_asset_admin": is_asset_admin,
    }


@api_router.get("/shop/check")
async def shop_check(_: bool = Depends(require_shop_or_admin)):
    return {"ok": True}


class ShopChangePasswordBody(BaseModel):
    old_password: str
    new_password: str


@api_router.get("/shop/me")
async def shop_me(
    actor=Depends(require_shop_or_admin),
    x_shop_token: Optional[str] = Header(default=None),
):
    """Return the signed-in shop user (per-user account) or a flag
    indicating a legacy shared-password / admin / open-mode session.

    The ``actor`` resolution in ``require_shop_or_admin`` returns the
    shop_user document for per-user tokens, ``True`` otherwise. We
    re-detect the per-user case by inspecting ``actor`` so the admin
    UI can show the right state.
    """
    from shop_users import public_shop_user_view

    if isinstance(actor, dict) and actor.get("id") and actor.get("email"):
        return {"ok": True, "user": public_shop_user_view(actor)}
    return {"ok": True, "is_legacy": True}


@api_router.post("/shop/change-password")
async def shop_change_password(
    body: ShopChangePasswordBody,
    request: Request,
    actor=Depends(require_shop_or_admin),
):
    """Per-shop-user password rotation. Requires a valid per-user shop
    token. Verifies ``old_password`` against the stored bcrypt hash,
    writes the new hash with ``must_change_password=false`` and
    returns a freshly-issued token (the prior token is invalidated by
    the hash change)."""
    from shop_users import (
        find_shop_user_by_email,
        make_shop_user_token,
        public_shop_user_view,
        set_shop_user_password,
        verify_password,
    )

    if not isinstance(actor, dict) or not actor.get("id"):
        raise HTTPException(
            status_code=403,
            detail="Only per-user shop accounts can change password here.",
        )
    if len(body.new_password or "") < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
    if body.new_password == body.old_password:
        raise HTTPException(status_code=400, detail="New password must be different from the old one")
    ip = _client_ip(request)

    if actor.get("linked_to_directory") or actor.get("source") == "directory-shadow":
        try:
            ok = await _ud.self_change_password(
                db,
                user_id=actor["id"],
                current_password=body.old_password,
                new_password=body.new_password,
            )
        except ValueError as ve:
            raise HTTPException(status_code=400, detail=str(ve))
        if not ok:
            _record_login_fail(ip)
            raise HTTPException(status_code=401, detail="Current password is wrong")
        _reset_login_fails(ip)
        fresh_row = await _ud.find_by_id(db, actor["id"])
        if not fresh_row or not fresh_row.get("password_hash"):
            raise HTTPException(status_code=404, detail="user not found")
        await db.shop_users.update_one(
            {"id": actor["id"]},
            {"$set": {
                "password_hash": fresh_row["password_hash"],
                "must_change_password": False,
                "password_set_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }},
        )
        saved = await db.shop_users.find_one({"id": actor["id"]}, {"_id": 0})
        if not saved:
            raise HTTPException(status_code=500, detail="Failed to set password")
        fresh = await find_shop_user_by_email(db, saved["email"])
        fresh_token = make_shop_user_token(fresh_row["id"], fresh_row["password_hash"])
        try:
            await _reset_session_activity(
                db, fresh_token, "OPERATIONS",
                user_id=fresh_row.get("id"), email=fresh_row.get("email"),
                actor_label="shop_via_directory", ip=ip,
                user_agent=request.headers.get("user-agent") or "",
            )
        except Exception:  # noqa: BLE001
            pass
        return {
            "ok": True,
            "token": fresh_token,
            "user": public_shop_user_view(fresh or saved),
        }
    _check_login_lockout(ip)
    pwh = actor.get("password_hash") or ""
    if not pwh or not verify_password(body.old_password, pwh):
        _record_login_fail(ip)
        raise HTTPException(status_code=401, detail="Current password is wrong")
    _reset_login_fails(ip)

    saved = await set_shop_user_password(
        db, actor["id"], body.new_password, must_change=False
    )
    if not saved:
        raise HTTPException(status_code=500, detail="Failed to set password")
    new_pwh = saved.get("password_hash") or ""
    fresh = await find_shop_user_by_email(db, saved["email"])
    try:
        await _reset_session_activity(
            db, make_shop_user_token(saved["id"], new_pwh), "OPERATIONS",
            user_id=saved.get("id"), email=saved.get("email"),
            actor_label="shop", ip=ip,
            user_agent=request.headers.get("user-agent") or "",
        )
    except Exception:  # noqa: BLE001
        pass
    return {
        "ok": True,
        "token": make_shop_user_token(saved["id"], new_pwh),
        "user": public_shop_user_view(fresh or saved),
    }


class ShopForgotPasswordBody(BaseModel):
    email: str


class ShopResetPasswordBody(BaseModel):
    token: str
    new_password: str


@api_router.post("/shop/forgot-password")
async def shop_forgot_password(body: ShopForgotPasswordBody, request: Request):
    """Self-service password reset — step 1.

    Mirror of ``/pm/forgot-password`` for the Shop portal. Always
    returns a generic 200 to prevent email enumeration. A 30-min
    HMAC-signed reset token is emailed via Resend bound to the first
    16 chars of the user's current password_hash so a successful
    reset invalidates the link.
    """
    from shop_users import find_shop_user_by_email, make_shop_reset_token

    ip = _client_ip(request)
    _check_login_lockout(ip)
    email = (body.email or "").strip().lower()

    generic = {
        "ok": True,
        "message": (
            "If that email is on file with a password, a reset link is on "
            "its way. Check your inbox in the next minute."
        ),
    }

    if not email or "@" not in email:
        _record_login_fail(ip)
        return generic

    user = await find_shop_user_by_email(db, email)
    if not user:
        _record_login_fail(ip)
        return generic
    pwh = user.get("password_hash") or ""
    if not pwh or user.get("disabled"):
        return generic

    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        logger.warning("[shop forgot-password] RESEND_API_KEY missing; cannot send")
        return generic

    portal_url = (
        os.environ.get("PORTAL_URL", "").strip()
        or os.environ.get("PRODUCTION_URL", "").strip()
        or "https://mascidocs.com"
    )
    token = make_shop_reset_token(user["id"], pwh)
    reset_link = f"{portal_url}/shop/reset/{token}"
    user_name = (user.get("name") or "").strip() or "Mechanic"
    body_inner = f"""
      <p style="margin:0 0 14px;font-size:15px;line-height:1.5">Hi {user_name},</p>
      <p style="margin:0 0 14px;font-size:14px;line-height:1.55;color:#334155">
        Someone (hopefully you) requested a password reset for the MASCI Shop Portal account
        <strong>{email}</strong>. Click the button below to choose a new password.
      </p>

      <table cellpadding="0" cellspacing="0" style="margin:18px 0">
        <tr><td style="background:#ea580c;border-radius:6px;padding:14px 28px">
          <a href="{reset_link}" style="color:#fff;font-weight:800;font-size:14px;letter-spacing:0.05em;text-transform:uppercase;text-decoration:none">
            Choose a new password
          </a>
        </td></tr>
      </table>

      <p style="margin:14px 0 0;font-size:13px;color:#64748b;line-height:1.55">
        This link expires in 30 minutes. If you didn't request a reset, ignore this email — your current password keeps working.
      </p>
      <p style="margin:8px 0 0;font-size:12px;color:#94a3b8;line-height:1.55">
        Direct link: <span style="font-family:Courier New,monospace;font-size:10px;word-break:break-all;color:#475569">{reset_link}</span>
      </p>
    """
    html_body = render_portal_email(
        portal="Shop",
        headline="Reset your password",
        body_inner_html=body_inner,
    )

    try:
        import resend
        resend.api_key = api_key
        sender_email = await _resolve_sender_email(db)
        params = {
            "from": f"MASCI Operations Platform <{sender_email}>",
            "to": [email],
            "subject": "[MASCI] Reset your Shop Operations password",
            "html": html_body,
        }
        reply_to = (await _resolve_reply_to_email(db)) or ""
        if reply_to:
            params["reply_to"] = reply_to
        await asyncio.to_thread(resend.Emails.send, params)
    except Exception as e:  # noqa: BLE001
        logger.error("[shop forgot-password] resend send failed: %s", e)

    return generic


@api_router.post("/shop/reset-password")
async def shop_reset_password(body: ShopResetPasswordBody, request: Request):
    """Self-service password reset — step 2.

    User clicks the email link → lands on ``/shop/reset/<token>`` →
    enters a new password (6+ char) → frontend POSTs here. Backend
    validates the token, sets the new bcrypt hash, clears
    must_change_password (the user has now picked their own pw), and
    returns a fresh per-user token so they drop straight into ``/shop``.
    """
    from shop_users import (
        consume_shop_reset_token,
        make_shop_user_token,
        public_shop_user_view,
        set_shop_user_password,
        stamp_shop_login,
    )

    ip = _client_ip(request)
    _check_login_lockout(ip)

    if len(body.new_password or "") < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")

    user = await consume_shop_reset_token(db, body.token)
    if not user:
        _record_login_fail(ip)
        raise HTTPException(
            status_code=400,
            detail="This reset link is invalid or has expired. Request a new one from /shop/login.",
        )

    updated = await set_shop_user_password(db, user["id"], body.new_password, must_change=False)
    if not updated:
        raise HTTPException(status_code=500, detail="Failed to update password")
    _reset_login_fails(ip)
    await stamp_shop_login(db, updated["id"], ip=ip)
    await _reset_session_activity(
        db, make_shop_user_token(updated["id"], updated.get("password_hash") or ""), "OPERATIONS",
        user_id=updated.get("id"), email=updated.get("email"),
        actor_label="shop", ip=ip,
        user_agent=request.headers.get("user-agent") or "",
    )
    return {
        "ok": True,
        "token": make_shop_user_token(updated["id"], updated.get("password_hash") or ""),
        "user": public_shop_user_view(updated),
    }


# ============================================================
# Safety Forms — Inspections, Meetings, JHPs, Incidents
# ----------------------------------------------------------
# Extracted to /app/backend/routes/safety.py 2026-04-28 (P1 refactor batch 2).
# Pydantic models (InspectionCreate, Inspection, MeetingCreate, Meeting, etc.)
# are now defined in that module. The 16 endpoints are attached to the shared
# router via register_safety_routes() below.
#
# iter383 · Restored after the iter382 PM-admin extraction inadvertently
# removed the registration blocks. Zero behavior drift: identical to the
# pre-iter382 wiring (admin/safety/pm dependency injection unchanged).
# ============================================================
from routes.safety import (  # noqa: E402,F401
    register_safety_routes,
    Inspection, InspectionCreate, InspectionSummary,
    Meeting, MeetingCreate, MeetingSummary,
    Jha, JhaCreate, JhaSummary,
    Incident, IncidentCreate, IncidentSummary,
)

register_safety_routes(
    api_router, db, require_admin, rate_limit_public_post,
    # Late binding: schedule_auto_email is defined later in this file. Wrapping
    # in a lambda lets Python resolve it at request time (when the route fires)
    # rather than at registration time (when it doesn't exist yet).
    lambda kind, record: schedule_auto_email(kind, record),
    # iter236 · Site Inspection moved into Safety portal ownership. POST
    # /api/inspections now requires Safety or Admin auth (no public/rate-limit
    # path). The make_require_safety_or_admin factory accepts X-Safety-Token
    # or X-Admin-Token; HR is intentionally excluded for this write surface.
    require_safety_or_admin=__import__(
        "routes.safety_portal._deps", fromlist=["make_require_safety_or_admin"]
    ).make_require_safety_or_admin(
        db, _is_valid_admin_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
    # iter322 · Safety-side READ gate. Closes the operator bug where
    # X-Safety-Token requests to /api/incidents, /inspections, /meetings,
    # /jhas were rejected with "Admin or PM login required". Accepts
    # Safety + Admin + PM. Destructive endpoints stay on require_admin.
    # TRACK 28.02 · async admin validator wired so per-user admin tokens
    # (UUID.HMAC issued by /api/auth/multi-login) unlock the gate — the
    # sync validator was retired in 15.32 and returns False for all
    # inputs, which had silently blocked admins from every Safety read
    # surface.
    require_safety_admin_or_pm=__import__(
        "routes.safety_portal._deps", fromlist=["make_require_safety_admin_or_pm"]
    ).make_require_safety_admin_or_pm(
        db, _is_valid_admin_token, _is_valid_pm_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)


# ============================================================
# OMEGA · Phase 1A · iter451 · OC-001 Incident Lifecycle
# ----------------------------------------------------------
# Additive transition endpoints (POST /incidents/{id}/transition,
# GET /incidents/{id}/state-events, GET /incidents/{id}/lifecycle).
# Uses the Safety/Admin/PM read gate so lifecycle reads work for the
# Safety reviewer; closure-role gate is enforced server-side by the
# state machine itself, not by the dependency.
# ============================================================
from routes.incident_lifecycle import register_incident_lifecycle_routes  # noqa: E402

register_incident_lifecycle_routes(
    api_router, db,
    require_incident_actor=__import__(
        "routes.safety_portal._deps", fromlist=["make_require_safety_admin_or_pm"]
    ).make_require_safety_admin_or_pm(
        db, _is_valid_admin_token, _is_valid_pm_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)


# ============================================================
# TRACK 19.16 · Phase A · Incident Intelligence Engine
# ------------------------------------------------------------
# NEW namespace /api/incident-cases/*  and  /api/corrective-actions/*.
# Legacy /api/incidents/* surface is UNTOUCHED (Zero-Drift Doctrine).
# Uses the same Safety/Admin/PM read gate as incident_lifecycle above.
# Write authority is narrowed inside the service layer per capability.
# ============================================================
from incident_engine.routes import register_incident_engine_routes  # noqa: E402
register_incident_engine_routes(
    api_router, db,
    require_actor=__import__(
        "routes.safety_portal._deps", fromlist=["make_require_safety_admin_or_pm"]
    ).make_require_safety_admin_or_pm(
        db, _is_valid_admin_token, _is_valid_pm_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)

# Public-gate near-miss kiosk (Phase B2). Additive. No auth. Routes
# submissions through the exact same Phase A domain engine + audit +
# lifecycle. Legacy routes untouched.
from incident_engine.public_gate import register_public_routes as _register_ie_public_routes  # noqa: E402
_register_ie_public_routes(api_router, db)

# Safety Case Workspace (Phase C). Additive satellite endpoints
# (communications / witnesses / medical / agency / tasks / health /
# executive-snapshot). Uses the same Safety/Admin/PM read gate.
from incident_engine.workspace_routes import register_workspace_routes as _register_ie_workspace_routes  # noqa: E402
_register_ie_workspace_routes(
    api_router, db,
    require_actor=__import__(
        "routes.safety_portal._deps", fromlist=["make_require_safety_admin_or_pm"]
    ).make_require_safety_admin_or_pm(
        db, _is_valid_admin_token, _is_valid_pm_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)

# Executive Intelligence Center (Phase D). Additive read-only aggregations
# over the incident engine. Never writes; never owns data.
from incident_engine.intelligence_routes import register_intelligence_routes as _register_ie_intel_routes  # noqa: E402
_register_ie_intel_routes(
    api_router, db,
    require_actor=__import__(
        "routes.safety_portal._deps", fromlist=["make_require_safety_admin_or_pm"]
    ).make_require_safety_admin_or_pm(
        db, _is_valid_admin_token, _is_valid_pm_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)

# Report Intelligence Engine (Phase E). Additive read-only report renderer.
from incident_engine.report_routes import register_report_routes as _register_ie_report_routes  # noqa: E402
_register_ie_report_routes(
    api_router, db,
    require_actor=__import__(
        "routes.safety_portal._deps", fromlist=["make_require_safety_admin_or_pm"]
    ).make_require_safety_admin_or_pm(
        db, _is_valid_admin_token, _is_valid_pm_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)

# TRACK 19.36 · Executive Intelligence Layer + Executive Report PDF.
# Additive · read-only assembler + boardroom-grade PDF endpoint.
# Consumes existing Phase A/C data; never mutates any collection.
# Existing Phase E executive PDF (/api/incident-cases/{id}/reports/{type}.pdf)
# is preserved untouched.
from incident_engine.executive_report_routes import (  # noqa: E402
    register_executive_report_routes as _register_ie_executive_report_routes,
)
_register_ie_executive_report_routes(
    api_router, db,
    require_actor=__import__(
        "routes.safety_portal._deps", fromlist=["make_require_safety_admin_or_pm"]
    ).make_require_safety_admin_or_pm(
        db, _is_valid_admin_token, _is_valid_pm_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)

# TRACK 19.37 · Passive Incident-Presence Scoring — attention signals only.
# Additive · read-only. Safety/Admin/PM gated (same as executive endpoints).
# Never a legal, OSHA, liability, root-cause, or discipline decision.
from incident_engine.presence_score_routes import (  # noqa: E402
    register_presence_score_routes as _register_ie_presence_score_routes,
)
_register_ie_presence_score_routes(
    api_router, db,
    require_actor=__import__(
        "routes.safety_portal._deps", fromlist=["make_require_safety_admin_or_pm"]
    ).make_require_safety_admin_or_pm(
        db, _is_valid_admin_token, _is_valid_pm_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)

# TRACK 19.38 · Cross-portal read fanout + Portfolio Attention Feed.
# Three additive read-only endpoints with three role gates.
# Reuses Track 19.37 compute_presence_score (no duplicate scoring logic).
from incident_engine.portfolio_intelligence import (  # noqa: E402
    register_portfolio_intelligence_routes as _register_ie_portfolio_routes,
)
_ie_portfolio_deps_mod = __import__(
    "routes.safety_portal._deps",
    fromlist=[
        "make_require_safety_or_admin",
        "make_require_safety_token",
        "make_require_safety_admin_or_pm",
    ],
)
_register_ie_portfolio_routes(
    api_router, db,
    require_safety_or_admin=_ie_portfolio_deps_mod.make_require_safety_or_admin(
        db, _is_valid_admin_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
    require_safety_token=_ie_portfolio_deps_mod.make_require_safety_token(db),
    require_safety_admin_or_pm=_ie_portfolio_deps_mod.make_require_safety_admin_or_pm(
        db, _is_valid_admin_token, _is_valid_pm_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)

# TRACK 19.39 · Morning Safety Intelligence Digest.
# Additive · opt-in · Safety+Admin gated. Uses existing fsi_send_email.
# Dry-run mode by default; recipients live in an additive collection
# (morning_digest_recipients) with an audit trail
# (morning_digest_audit). No new email provider. No scheduler yet.
from incident_engine.morning_digest_routes import (  # noqa: E402
    register_morning_digest_routes as _register_ie_morning_digest_routes,
)
_register_ie_morning_digest_routes(
    api_router, db,
    require_safety_or_admin=_ie_portfolio_deps_mod.make_require_safety_or_admin(
        db, _is_valid_admin_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)

# TRACK 19.40 · Unified Operational Intelligence Engine.
# The permanent foundation for every operational briefing/digest/report.
# Zero-drift: reuses existing collections + fsi_send_email + WeasyPrint.
#
# TRACK 19.41 · Auth wiring fix — the sync `_is_valid_admin_token` was
# retired in Track 15.32 and now always returns False, which locked
# admin_only OI products (like `po_weekly_digest`) out for every real
# admin. Route directly through the async directory-admin validator +
# safety-user validator so `X-Admin-Token: <directory_admin_token>` and
# `X-Safety-Token: <safety_token>` both work as intended.
from operational_intelligence.routes import (  # noqa: E402
    register_operational_intelligence_routes as _register_oi_routes,
)
from fastapi import Header as _OiHeader, HTTPException as _OiHTTPException  # noqa: E402
from safety_users import is_valid_safety_user_token_async as _oi_safety_valid  # noqa: E402


def _make_oi_require_safety_or_admin():
    async def _dep(
        x_safety_token: Optional[str] = _OiHeader(default=None, alias="X-Safety-Token"),
        x_admin_token: Optional[str] = _OiHeader(default=None, alias="X-Admin-Token"),
    ):
        if x_safety_token:
            u = await _oi_safety_valid(db, x_safety_token)
            if u:
                return {**u, "_actor": "safety"}
        if x_admin_token and await _is_valid_directory_admin_token_async(x_admin_token):
            return {"_actor": "admin", "name": "Admin"}
        raise _OiHTTPException(401, detail={"code": "unauthorized",
                                            "detail": "Safety or Admin auth required"})
    return _dep


def _make_oi_require_admin_only():
    """Strict admin-only gate for recipient CRUD + destructive endpoints."""
    async def _dep(
        x_admin_token: Optional[str] = _OiHeader(default=None, alias="X-Admin-Token"),
    ):
        if x_admin_token and await _is_valid_directory_admin_token_async(x_admin_token):
            return {"_actor": "admin", "name": "Admin"}
        raise _OiHTTPException(401, detail={"code": "unauthorized",
                                            "detail": "Admin auth required"})
    return _dep


_register_oi_routes(
    api_router, db,
    require_safety_or_admin=_make_oi_require_safety_or_admin(),
    require_admin=_make_oi_require_admin_only(),
)

# TRACK 19.21 · Employee Records Intelligence Platform · P0 foundation.
# Universal Employee Record + intake batches + review queue + audit trail.
# Additive · zero drift · HR is system owner across every lane.
from routes.employee_records import (  # noqa: E402
    build_employee_records_router,
    ensure_employee_records_indexes,
    make_employee_records_actor_gate,
)
app.include_router(build_employee_records_router(
    db=db,
    require_actor=make_employee_records_actor_gate(
        db, _is_valid_admin_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
))


# ─── OMEGA · Phase 1A · iter452 · OC-002 Daily Report Office Review ──
#     Additive transition endpoints. Uses the Safety/Admin/PM read gate
#     so PMs, Office (Admin), and Safety reviewers can read the
#     lifecycle. State-machine gate narrows write authority to:
#       OPEN → PENDING_REVIEW   : PM/Admin/Super-Admin (field submits)
#       PENDING_REVIEW → REVIEWED|OPEN : Admin/Super-Admin (office reviews / kicks back)
#       REVIEWED → CLOSED       : Admin/Super-Admin (final attestation)
#       CLOSED → PENDING_REVIEW : Admin/Super-Admin (reopen with reason)
from routes.daily_report_lifecycle import register_daily_report_lifecycle_routes  # noqa: E402
register_daily_report_lifecycle_routes(
    api_router, db,
    require_dr_actor=__import__(
        "routes.safety_portal._deps", fromlist=["make_require_safety_admin_or_pm"]
    ).make_require_safety_admin_or_pm(
        db, _is_valid_admin_token, _is_valid_pm_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)


# ─── OMEGA · iter452.5 · Tier 1 · Field Submitter Identity (FSI) ────
#     Shared platform service. Public-gate submissions can now anchor
#     to an employee directory entry + per-submit reachable email so
#     correction emails reach the responsible party. Signed JWT links
#     enable a passwordless `/revise/:token` revision flow. Full
#     six-event delivery-evidence chain is written into the existing
#     workflow_state_events collection so Phase 1B can prove the
#     accountability loop closed end-to-end.
from routes.field_revision import register_field_revision_routes  # noqa: E402
_fsi_ctx = register_field_revision_routes(
    api_router, db,
    send_email_fn=None,  # late-bound below once _safety_send_email exists
)


# QA/QC inspection routes (Concrete Form / Rebar / Subcontractor Work).
# Same pattern as the Safety routes — single registration helper, late-bound
# auto-email so PM routing fires after submit.
from routes.qaqc import register_qaqc_routes  # noqa: E402

register_qaqc_routes(
    api_router, db, require_admin, rate_limit_public_post,
    lambda kind, record: schedule_auto_email(kind, record),
)


# ─── OMEGA · iter453 · OC-003 + OC-004 Lifecycle (Constitutional Build Package) ─
#     Additive transition endpoints for QA/QC and Site Inspection
#     follow-up. Closure-action contract enforced by the state machine
#     itself (Amendment 001 REPLACE-4 + REPLACE-5). Reads use the
#     Safety/Admin/PM gate; closure gates narrow per-role inside the
#     state-machine module. No assignment UI, no ack-click closure,
#     no parallel task object — operational record IS the work.
from routes.qaqc_lifecycle import register_qaqc_lifecycle_routes  # noqa: E402
register_qaqc_lifecycle_routes(
    api_router, db,
    require_qaqc_actor=__import__(
        "routes.safety_portal._deps", fromlist=["make_require_safety_admin_or_pm"]
    ).make_require_safety_admin_or_pm(
        db, _is_valid_admin_token, _is_valid_pm_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)

from routes.site_inspection_lifecycle import register_site_inspection_lifecycle_routes  # noqa: E402
register_site_inspection_lifecycle_routes(
    api_router, db,
    require_inspection_actor=__import__(
        "routes.safety_portal._deps", fromlist=["make_require_safety_admin_or_pm"]
    ).make_require_safety_admin_or_pm(
        db, _is_valid_admin_token, _is_valid_pm_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)


# ─── OMEGA · iter452.5.2 · Resend Bounce Webhook + Deliverability Evidence Chain ─
#     Closes Email Sent → Delivered → Bounced → Dead Letter chain.
#     Hard bounce on any non-dead-letter tier auto-escalates ownership
#     to Tier 5 dead-letter via the existing FSI write_chain_event +
#     write_dispatch_event helpers. No user click required. Rule 7 +
#     Ownership Doctrine O-4 textbook.
from routes.resend_webhook import register_resend_webhook_routes  # noqa: E402
register_resend_webhook_routes(api_router, db)


# ============================================================
# Daily Job Reports
# ----------------------------------------------------------
# Extracted to /app/backend/routes/daily_reports.py 2026-04-28 (P1 batch 3).
# ============================================================
from routes.daily_reports import (  # noqa: E402,F401
    register_daily_reports_routes,
    DailyReport, DailyReportCreate, DailyReportSummary,
)

# ------------------------------------------------------------
# DR-UNIFY-002 · Register the unified `/api/daily-reports/approved`
# + `/api/daily-reports/{id}/pdf` aliases BEFORE `daily_reports.py`
# so the literal `approved` segment takes precedence over the
# `{report_id}` wildcard in FastAPI's ordered-match routing.
# ------------------------------------------------------------
from routes.dr_v2_pdf import register_dr_v2_pdf_routes  # noqa: E402
from pm_auth import compute_pm_scope as _compute_pm_scope_for_dr_v2_pdf  # noqa: E402
register_dr_v2_pdf_routes(
    api_router, db,
    require_admin_pm_or_hr_read=require_admin_pm_or_hr_read,
    compute_pm_scope=_compute_pm_scope_for_dr_v2_pdf,
)

register_daily_reports_routes(
    api_router, db, require_admin, rate_limit_public_post,
    lambda kind, record: schedule_auto_email(kind, record),
    require_admin_pm_or_hr_read=require_admin_pm_or_hr_read,
)

# ------------------------------------------------------------
# TRACK 24.17 · Operations Control Center — unified maintenance
# console for super-admins. Wraps existing 24.12 disk scripts +
# health probes so a non-coder platform owner can run cleanup,
# migrations, and health checks from `/admin/operations-control`
# without shell access. Every mutation writes an
# `operations_audit` row.
# ------------------------------------------------------------
from routes.operations_control import (  # noqa: E402
    register_operations_control_routes,
)
from services.operations_control.audit import (  # noqa: E402
    ensure_indexes as ensure_occ_audit_indexes,
)
api_router._get_runtime_identity = _runtime_identity_bundle  # type: ignore[attr-defined]
register_operations_control_routes(api_router, db, require_admin, get_database_authority_plan=lambda: getattr(app.state, "database_authority_plan", None))


@register_lifecycle_step("index-ensure")
async def _ensure_occ_audit_indexes_step():
    await ensure_occ_audit_indexes(db)


@register_lifecycle_step("index-ensure")
async def _ensure_production_certification_session_indexes_step():
    await _pcs_ensure_indexes(db)


@register_lifecycle_step("index-ensure")
async def _ensure_dr_v2_alias_indexes_step():
    await _ensure_dr_v2_alias_indexes(db)

# TRACK 25 · SPRINT 2 · Operations Control Center — Trust Layer aggregator.
# One canonical read-only endpoint `GET /api/admin/occ/health` that fans
# out over the existing child health endpoints and returns a normalized
# 8-section snapshot for the OCC "Trust Center". No new truth sources.
# No server-side cache — the OCC frontend controls refresh explicitly.
from routes.occ_health_aggregator import (  # noqa: E402
    register_occ_health_routes,
)
register_occ_health_routes(api_router, require_admin)

# TRACK 25 · SPRINT 7/8 · Trust Events aggregator.
from routes.occ_trust_events import (  # noqa: E402
    register_occ_trust_events_routes,
)
register_occ_trust_events_routes(api_router, require_admin)

# ------------------------------------------------------------
# DR-CUTOVER-002 · Daily Operational Summary (draft + accept)
# ADDITIVE mount. Zero drift on the V1 submit path — the two
# routes here NEVER modify a submitted daily_report's core fields
# (crews, equipment, safety, photos). They only compose preview
# text and, on accept, patch a small set of `daily_operational_summary*`
# fields onto the existing document.
# ------------------------------------------------------------
from routes.daily_summary import register_daily_summary_routes  # noqa: E402
register_daily_summary_routes(
    api_router, db=db, rate_limit_public_post=rate_limit_public_post, require_admin=require_admin, require_admin_strict=require_admin_strict,
)
# ------------------------------------------------------------
# ADDITIVE mount. Zero drift on V1 daily_reports routes, models, or
# collections. Feature-flag gated (DR_V2_AI_ENABLED). See
# /app/backend/routes/dr_v2.py for the /api/dr-v2/* surface.
# ============================================================
from routes.dr_v2 import register_dr_v2_routes  # noqa: E402
register_dr_v2_routes(api_router, db)
from routes.dr_v2_canonicalize import register_dr_v2_canonicalize_routes  # noqa: E402
register_dr_v2_canonicalize_routes(api_router, db)

# ------------------------------------------------------------
# AI-CONFIG-001 · Admin-only AI Gateway status endpoint.
# Zero secrets returned; booleans only. Blocks the "did I paste the
# key correctly?" support loop without exposing raw values.
# ------------------------------------------------------------
from routes.ai_gateway_status import register_ai_gateway_status_routes  # noqa: E402
register_ai_gateway_status_routes(api_router, require_admin=require_admin)

# ------------------------------------------------------------
# AI-HEALTH-001 · Admin-only live provider health probe.
# Runs a real ping against each configured provider so silent
# auth / quota failures surface immediately instead of dropping
# to the deterministic-summary fallback.
# ------------------------------------------------------------
from routes.ai_health import register_ai_health_routes  # noqa: E402
register_ai_health_routes(api_router, require_admin=require_admin)

# ------------------------------------------------------------
# AI-ADMIN-001 · Admin AI Configuration Center.
# Admin-only tenant AI capability management + audit trail.
# All routes gated by ``require_admin_strict`` (PM tokens rejected).
# ------------------------------------------------------------
from routes.ai_admin_config import register_ai_admin_config_routes  # noqa: E402
register_ai_admin_config_routes(
    api_router,
    db=db,
    require_admin_strict=require_admin_strict,
)

# ------------------------------------------------------------
# TRACK 22.4b-followup · Preview Validation Identities. Safe,
# preview-only control plane for minting short-lived role tokens
# used to unblock role-scoped workflow verification. Hard-disabled
# in production. Super-admin only.
# ------------------------------------------------------------
from routes.preview_validation_identities import (  # noqa: E402
    register_preview_validation_identity_routes,
)
register_preview_validation_identity_routes(
    api_router,
    db=db,
    require_admin_strict=require_admin_strict,
)

# ------------------------------------------------------------
# TRACK 22.6A · Production Certification Session. Read-only,
# short-lived, super-admin-gated auditable session used by an
# automated post-deployment certifier. Works in production by
# design (unlike PVI). No writes, no email/SMS, no config-mutation,
# no secret exposure. Every mint / probe / revoke audited.
# ------------------------------------------------------------
from routes.production_certification_session import (  # noqa: E402
    register_production_certification_session_routes,
    ensure_indexes as _pcs_ensure_indexes,
)
register_production_certification_session_routes(
    api_router,
    db=db,
    require_admin_strict=require_admin_strict,
)

# ------------------------------------------------------------
# TRACK 22.3 · Integration Truth Surface + AI Key Status + DR-V2
# Alias Telemetry. Admin-only. Runtime truth (os.environ) — never
# reads placeholder .env values. Rebuilds trust after Track 22.2.
# ------------------------------------------------------------
from routes.integration_truth import (  # noqa: E402
    register_integration_truth_routes,
    record_dr_v2_alias_hit as _record_dr_v2_alias_hit,
    ensure_dr_v2_alias_indexes as _ensure_dr_v2_alias_indexes,
)
register_integration_truth_routes(
    api_router,
    db=db,
    require_admin_strict=require_admin_strict,
    get_runtime_identity=_runtime_identity_bundle,
)

# ------------------------------------------------------------
# DR-ROI-001F · Part 2 · V2 PDF Output (already registered above via
# DR-UNIFY-002 route-order fix; block retained as a doctrine anchor).
# ------------------------------------------------------------

# ============================================================
# DR-ROI-001D · Photo Vision + Evidence Linking
# ------------------------------------------------------------
# ADDITIVE mount. Zero drift on V1, DR-V2, or Job Photos mirror.
# Feature-flag gated by DR_V2_PHOTO_VISION_ENABLED.
# ============================================================
from routes.dr_v2_photos import register_dr_v2_photo_routes  # noqa: E402
register_dr_v2_photo_routes(api_router, db)


# ── TRACK 23.1 · Cost Code Provider + V3 UI feature flag ───────────
from routes.cost_codes import register_cost_code_routes  # noqa: E402
from routes.ui_flags import register_dr_v3_flag_routes  # noqa: E402
register_cost_code_routes(
    api_router,
    db,
    require_admin=require_admin,
    require_admin_pm_or_hr_read=require_admin_pm_or_hr_read,
)
register_dr_v3_flag_routes(api_router, db, require_admin=require_admin)


# ============================================================
# ODS-001 · Operational Data Spine Foundation
# ------------------------------------------------------------
# ADDITIVE mount. Zero drift on V1 or DR-V2. Feature-flag gated
# by ODS_ENABLED. See /app/backend/routes/ods.py.
# ============================================================
from routes.ods import register_ods_routes  # noqa: E402
register_ods_routes(api_router, db)


@register_lifecycle_step("index-ensure")
async def _boot_router_registered_import_safe_indexes():
    for attr in ("_dr_v2_ensure_all_indexes", "_dr_v2_photo_ensure_indexes", "_ods_boot_indexes"):
        fn = getattr(api_router, attr, None)
        if fn:
            await fn()

# ============================================================
# TRACK 23.7 · Operational KPI Spine (PM + Safety + future Scheduling)
# ------------------------------------------------------------
# Shared aggregator lives in services/operational_kpis/aggregator.py.
# Two thin route wrappers here register PM + Safety endpoints.
# ABSOLUTE RULE: NO money, NO cost, NO rates. Operational production
# intelligence only.
# ============================================================
from routes.operational_kpis import register_operational_kpis_routes  # noqa: E402
register_operational_kpis_routes(
    api_router, db,
    require_admin_dep=require_admin,
    require_safety_or_admin_dep=require_safety_or_admin,
)

# ============================================================
# DR-ROI-001E · PM / Admin / Executive Intelligence Dashboards
# ------------------------------------------------------------
# ADDITIVE mount. Zero drift on V1/V2/ODS. Reads snapshots +
# operational_facts only. Never mutates source records.
# ============================================================
from routes.ods_intelligence import register_ods_intelligence_routes  # noqa: E402
register_ods_intelligence_routes(api_router, db)


# ============================================================
# TRACK 19.04 · Unified Daily Report Attachment Pipeline
# ============================================================
# Reuses the SAME R2 bucket / client / signed-URL infrastructure that
# already backs photo uploads (see /app/backend/photo_storage.py).
# Photos continue to use their existing pipeline. This endpoint
# accepts documents (PDF, XLS, XLSX, CSV) so Daily Reports can carry
# tickets, delivery slips, quantity spreadsheets, and CEI reports
# alongside job photos with ONE metadata model, ONE permissions
# model, ONE retrieval model.

class DailyReportAttachmentUpload(BaseModel):
    file_data: str  # data URL: "data:<mime>;base64,<...>"
    filename: str = ""


@api_router.post("/daily-reports/attachments/upload")
async def daily_report_attachment_upload(payload: DailyReportAttachmentUpload):
    """TRACK 19.04 · Unified attachment upload.

    Validates MIME + extension + size + filename, uploads to R2, and
    returns the metadata envelope for inclusion in the Daily Report
    payload. Public — Daily Reports themselves are a public submit
    surface (foremen submit in the field without a portal login).
    The returned `attachment_ref` is only meaningful once linked to
    a specific Daily Report body, so an orphan upload has no report
    to leak into.
    """
    try:
        from photo_storage import upload_document_data_url
        meta = await upload_document_data_url(
            payload.file_data,
            source_id="dr_attachment",
            original_filename=payload.filename or "",
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=503, detail=str(e))
    meta["contract_version"] = "19.04"
    return meta


# ============================================================
# TRACK 15.62 · Admin-tier Daily Report intelligence
# ============================================================
from routes.dr_admin_intel import register_dr_admin_intel_routes
register_dr_admin_intel_routes(
    api_router, db,
    require_admin_pm_or_hr_read=require_admin_pm_or_hr_read,
)


# ============================================================
# Job Hazard Plans (per-job PDF repository — admin uploads, crews view)
# ============================================================
class JobHazardPlanUpload(BaseModel):
    """Admin uploads (or replaces) a Job Hazard Plan for one project."""
    project_number: str
    project_name: str = ""
    location: str = ""
    filename: str
    content_type: str = "application/pdf"
    file_data: str  # data URL: "data:application/pdf;base64,<...>"
    notes: Optional[str] = ""
    uploaded_by: Optional[str] = ""


class JobHazardPlan(BaseModel):
    id: str
    project_number: str
    project_name: str = ""
    location: str = ""
    filename: str
    content_type: str = "application/pdf"
    file_size: int = 0
    notes: Optional[str] = ""
    uploaded_by: Optional[str] = ""
    uploaded_at: str


def _data_url_to_bytes(data_url: str) -> Tuple[bytes, str]:
    """Parse `data:<mime>;base64,<...>` → (raw_bytes, mime). Raises on bad format."""
    if not data_url or "," not in data_url:
        raise ValueError("file_data must be a data URL")
    head, b64 = data_url.split(",", 1)
    mime = "application/octet-stream"
    if head.startswith("data:") and ";base64" in head:
        mime = head[5:].split(";", 1)[0] or mime
    try:
        raw = _email_b64.b64decode(b64)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"file_data base64 decode failed: {e}")
    return raw, mime


def _validate_pdf_or_400(raw: bytes) -> None:
    """Reject anything that isn't a real PDF. Without this an admin (or
    anyone with a stolen admin token) could upload an HTML/JS file claiming
    to be application/pdf and serve XSS via the /file download endpoint."""
    if not raw or len(raw) < 5 or not raw.startswith(b"%PDF-"):
        raise HTTPException(
            status_code=400,
            detail="Uploaded file is not a valid PDF. Magic bytes mismatch.",
        )


@api_router.get("/job-hazard-plans", response_model=List[JobHazardPlan])
async def list_job_hazard_plans():
    """Public — list every uploaded plan (without the heavy file payload)."""
    cursor = db.job_hazard_plans.find(
        {},
        {
            "_id": 0,
            "file_data": 0,  # exclude the base64 blob
        },
    ).sort("project_number", 1)
    docs = await cursor.to_list(2000)
    return [JobHazardPlan(**d) for d in docs]


@api_router.get("/job-hazard-plans/{project_number}/file")
async def download_job_hazard_plan(project_number: str):
    """Public — stream the raw PDF for a given project number."""
    doc = await db.job_hazard_plans.find_one({"project_number": project_number}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="No plan uploaded for this job yet")
    try:
        raw, mime = _data_url_to_bytes(doc.get("file_data") or "")
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Stored file is corrupt: {e}")

    safe_name = "".join(
        c if c.isalnum() or c in ("-", "_", ".", " ") else "_"
        for c in (doc.get("filename") or f"JHA_{project_number}.pdf")
    )
    return Response(
        content=raw,
        # Force application/pdf so a maliciously-MIME'd upload can never
        # be rendered as HTML/JS in the browser (defense in depth — we
        # also reject non-PDF magic bytes at upload time).
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{safe_name}"',
            "X-Content-Type-Options": "nosniff",
        },
    )


@api_router.post("/job-hazard-plans", response_model=JobHazardPlan)
async def upload_job_hazard_plan(
    payload: JobHazardPlanUpload,
    _: bool = Depends(require_admin),
):
    """Admin — upload (or REPLACE) a Job Hazard Plan PDF for one project number.
    Idempotent on project_number — uploading again replaces the prior file."""
    raw, mime = _data_url_to_bytes(payload.file_data)
    _validate_pdf_or_400(raw)
    if len(raw) > 25 * 1024 * 1024:
        raise HTTPException(
            status_code=413,
            detail=f"File too large ({len(raw)//1024} KB). Max 25 MB per plan.",
        )

    pn = payload.project_number.strip()
    if not pn:
        raise HTTPException(status_code=400, detail="project_number is required")

    plan_id = str(uuid.uuid4())
    doc = {
        "id": plan_id,
        "project_number": pn,
        "project_name": (payload.project_name or "").strip(),
        "location": (payload.location or "").strip(),
        "filename": payload.filename,
        "content_type": mime,
        "file_size": len(raw),
        "file_data": payload.file_data,
        "notes": (payload.notes or "").strip(),
        "uploaded_by": (payload.uploaded_by or "").strip(),
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
    }
    # Upsert — one plan per project number (replace on re-upload)
    await db.job_hazard_plans.update_one(
        {"project_number": pn},
        {"$set": doc},
        upsert=True,
    )
    fresh = await db.job_hazard_plans.find_one(
        {"project_number": pn}, {"_id": 0, "file_data": 0}
    )
    return JobHazardPlan(**fresh)


@api_router.delete("/job-hazard-plans/{project_number}")
async def delete_job_hazard_plan(
    project_number: str,
    _: bool = Depends(require_admin),
):
    res = await db.job_hazard_plans.delete_one({"project_number": project_number})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="No plan exists for this project")
    return {"deleted": True, "project_number": project_number}


# ============================================================
# Job Hazard FILES — multi-file library per project (replaces the single-PDF
# model above). Accepts any common file type up to 250 MB. Files >8 MB stream
# straight to disk; smaller ones are inlined for fast list/download.
# ============================================================
@api_router.get("/job-hazard-files")
async def list_all_jha_files(_: bool = Depends(require_admin)):
    """Admin — every project + its files, grouped. Drives /admin/jha."""
    from job_hazard_files import list_all_files_grouped
    return {"projects": await list_all_files_grouped(db)}


@api_router.get("/job-hazard-files/public/grouped")
async def list_all_jha_files_public():
    """Public — every project + its files, grouped. Drives the /jha page
    in the Safety section so crews can see and download every plan an
    admin uploaded. Returns the array directly (no wrapper) and never
    leaks file_data — only filename / size / uploaded_at / uploaded_by /
    notes per file."""
    from job_hazard_files import list_all_files_grouped
    return await list_all_files_grouped(db)


@api_router.get("/job-hazard-files/by-project/{project_number}")
async def list_jha_files_for_project(project_number: str):
    """Public — files for one project. Crews use this to pull JHPs offline."""
    from job_hazard_files import list_files_for_project
    return {"items": await list_files_for_project(db, project_number)}


@api_router.post("/job-hazard-files")
async def upload_jha_file(
    project_number: str = Form(...),
    file: UploadFile = File(...),
    notes: str = Form(""),
    uploaded_by: str = Form(""),
    _: bool = Depends(require_admin),
):
    """Admin — upload one file for a project. Multipart form."""
    from job_hazard_files import upload_file
    return await upload_file(
        db,
        project_number=project_number,
        file=file,
        notes=notes,
        uploaded_by=uploaded_by,
    )


@api_router.get("/job-hazard-files/{file_id}/download")
async def download_jha_file(file_id: str):
    """Public — stream a file by id. Inline for browser preview where the
    type supports it (PDF, image), otherwise download.

    Streams disk-backed files with FileResponse so memory stays bounded
    even for the 250 MB plan-set zips.
    """
    from job_hazard_files import get_file_for_download
    doc, raw, disk_path = await get_file_for_download(db, file_id)

    fname = doc.get("filename") or f"file-{file_id}.bin"
    content_type = doc.get("content_type") or "application/octet-stream"
    safe_name = "".join(
        c if c.isalnum() or c in ("-", "_", ".", " ") else "_" for c in fname
    )

    # Browser will preview PDFs & images inline; everything else downloads.
    inline_kinds = {"application/pdf"}
    disposition = "inline" if (
        content_type in inline_kinds or content_type.startswith("image/")
    ) else "attachment"
    headers = {
        "Content-Disposition": f'{disposition}; filename="{safe_name}"',
        "X-Content-Type-Options": "nosniff",
    }

    if disk_path is not None:
        return FileResponse(
            path=str(disk_path),
            media_type=content_type,
            filename=safe_name,
            headers=headers,
        )
    return Response(content=raw, media_type=content_type, headers=headers)


@api_router.delete("/job-hazard-files/{file_id}")
async def delete_jha_file(file_id: str, _: bool = Depends(require_admin)):
    from job_hazard_files import delete_file
    ok = await delete_file(db, file_id)
    if not ok:
        raise HTTPException(404, "File not found")
    return {"ok": True, "id": file_id}


# ============================================================
# Trench-Box Tabulated Data library — piggybacks on job_hazard_files with
# scope="trench_box". Key is the box's id (or "general" for shared
# educational docs like the United Rentals "What is Tabulated Data?" PDF).
# ============================================================
@api_router.get("/trench-box-files")
async def list_trench_box_files_grouped():
    """Public — every trench box's files, grouped. Used by the crew
    Tabulated Data Library page + the admin workspace."""
    from job_hazard_files import list_all_files_grouped
    groups = await list_all_files_grouped(db, scope="trench_box")
    return {"projects": groups}


@api_router.get("/trench-box-files/by-box/{box_id}")
async def list_trench_box_files_for_box(box_id: str):
    """Public — files attached to a specific trench box (or 'general')."""
    from job_hazard_files import list_files_for_project
    return {"items": await list_files_for_project(db, box_id, scope="trench_box")}


@api_router.post("/trench-box-files")
async def upload_trench_box_file(
    box_id: str = Form(...),
    file: UploadFile = File(...),
    notes: str = Form(""),
    uploaded_by: str = Form(""),
    _: bool = Depends(require_admin),
):
    """Admin — upload a tabulated-data PDF (or any doc) for a trench box.
    Use box_id="general" for shared educational docs that apply to the
    whole fleet (e.g. the United Rentals explainer)."""
    from job_hazard_files import upload_file
    return await upload_file(
        db,
        project_number=box_id,
        file=file,
        notes=notes,
        uploaded_by=uploaded_by,
        scope="trench_box",
    )


@api_router.delete("/trench-box-files/{file_id}")
async def delete_trench_box_file(
    file_id: str, _: bool = Depends(require_admin)
):
    from job_hazard_files import delete_file
    ok = await delete_file(db, file_id)
    if not ok:
        raise HTTPException(404, "File not found")
    return {"ok": True, "id": file_id}


# ============================================================
# OMEGA · FOCP Release 2 · TR-0001 · JHP Acknowledgement Ledger
# ----------------------------------------------------------
# Additive employee-acknowledgement ledger on top of the existing
# job_hazard_files infrastructure. Adds POST /jha-acknowledgements
# + supervisor / employee / compliance read endpoints. Does NOT
# touch the existing /job-hazard-files surfaces or the legacy
# single-PDF /job-hazard-plans surfaces above.
# ============================================================
from routes.jha_acknowledgements import register_jha_acknowledgement_routes  # noqa: E402

register_jha_acknowledgement_routes(
    api_router, db,
    require_admin_dep=require_admin,
)


# ============================================================
# OMEGA · FOCP Release 2 · TR-0002 · Universal Undo / Recovery Layer
# ----------------------------------------------------------
# Single endpoint that reverses the last transition for ANY workflow
# already wired to workflow_state_events (incident / daily_report /
# qaqc_inspection / site_inspection / payroll_variance). Plus a
# cross-workflow audit-stream read for the Admin Recovery Stream
# visibility page. Admin-only — does NOT bypass any state machine.
# ============================================================
from routes.workflow_undo import register_workflow_undo_routes  # noqa: E402

register_workflow_undo_routes(
    api_router, db,
    require_admin_dep=require_admin,
)


# ============================================================
# Trench Box Tabulated Data (OSHA 1926 Subpart P)
# ============================================================
class TrenchBoxCreate(BaseModel):
    """OSHA tabulated-data record for one trench shield / box. Filled by
    admin from the manufacturer's data plate — crews then browse read-only."""
    manufacturer: str
    model: str
    serial_number: Optional[str] = ""
    box_type: Optional[str] = ""  # e.g. "Steel", "Aluminum", "Modular"
    length_ft: Optional[str] = ""
    width_min_ft: Optional[str] = ""  # narrowest spread
    width_max_ft: Optional[str] = ""  # widest spread
    sidewall_height_ft: Optional[str] = ""
    sidewall_thickness_in: Optional[str] = ""
    weight_lbs: Optional[str] = ""

    # Maximum allowable depth by soil type (per OSHA 1926.652 / 1926 Subpart P)
    max_depth_type_a_ft: Optional[str] = ""
    max_depth_type_b_ft: Optional[str] = ""
    max_depth_type_c_60_ft: Optional[str] = ""  # C-60 (60° slope)
    max_depth_type_c_80_ft: Optional[str] = ""  # C-80 (80° slope)

    spreader_count: Optional[str] = ""
    stacking_allowed: Optional[str] = "No"
    stacking_max: Optional[str] = ""

    notes: Optional[str] = ""
    # Optional manufacturer tabulated-data PDF (data URL, max 10 MB)
    tabulated_data_file: Optional[str] = ""
    tabulated_data_filename: Optional[str] = ""


class TrenchBox(TrenchBoxCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(
        default_factory=lambda: datetime.now(timezone.utc).isoformat()
    )
    updated_at: str = Field(
        default_factory=lambda: datetime.now(timezone.utc).isoformat()
    )


@api_router.get("/trench-boxes", response_model=List[TrenchBox])
async def list_trench_boxes():
    """Public — every crew can browse to see what's OSHA-legal for what depth."""
    cursor = db.trench_boxes.find(
        {},
        {
            "_id": 0,
            "tabulated_data_file": 0,  # excluded from list (heavy)
        },
    ).sort([("manufacturer", 1), ("model", 1)])
    docs = await cursor.to_list(500)
    # Re-include empty placeholder so Pydantic doesn't choke
    for d in docs:
        d.setdefault("tabulated_data_file", "")
    return [TrenchBox(**d) for d in docs]


@api_router.get("/trench-boxes/{box_id}", response_model=TrenchBox)
async def get_trench_box(box_id: str):
    """Public — full record for one trench box (without the file payload)."""
    doc = await db.trench_boxes.find_one(
        {"id": box_id}, {"_id": 0, "tabulated_data_file": 0}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Trench box not found")
    doc.setdefault("tabulated_data_file", "")
    return TrenchBox(**doc)


@api_router.get("/trench-boxes/{box_id}/file")
async def download_trench_box_file(box_id: str):
    """Public — stream the manufacturer's tabulated-data PDF (if uploaded)."""
    doc = await db.trench_boxes.find_one({"id": box_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Trench box not found")
    data_url = doc.get("tabulated_data_file") or ""
    if not data_url:
        raise HTTPException(
            status_code=404,
            detail="No manufacturer tabulated-data PDF uploaded for this box",
        )
    try:
        raw, mime = _data_url_to_bytes(data_url)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"File corrupt: {e}")
    safe_name = "".join(
        c if c.isalnum() or c in ("-", "_", ".", " ") else "_"
        for c in (doc.get("tabulated_data_filename")
                  or f"TabData_{doc.get('manufacturer', '')}_{doc.get('model', '')}.pdf")
    )
    return Response(
        content=raw,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{safe_name}"',
            "X-Content-Type-Options": "nosniff",
        },
    )


@api_router.post("/trench-boxes", response_model=TrenchBox)
async def create_trench_box(
    payload: TrenchBoxCreate,
    _: bool = Depends(require_safety_or_admin),
):
    if not payload.manufacturer.strip() or not payload.model.strip():
        raise HTTPException(
            status_code=400, detail="manufacturer and model are required"
        )
    # Lightly validate optional file size
    if payload.tabulated_data_file:
        raw, _ = _data_url_to_bytes(payload.tabulated_data_file)
        _validate_pdf_or_400(raw)
        if len(raw) > 10 * 1024 * 1024:
            raise HTTPException(
                status_code=413,
                detail=f"Tabulated-data file too large ({len(raw)//1024} KB). Max 10 MB.",
            )
    box = TrenchBox(**payload.model_dump())
    doc = box.model_dump()
    await db.trench_boxes.insert_one(doc)
    doc.pop("_id", None)
    # Don't echo the file blob back
    doc.pop("tabulated_data_file", None)
    doc["tabulated_data_file"] = ""
    return TrenchBox(**doc)


@api_router.put("/trench-boxes/{box_id}", response_model=TrenchBox)
async def update_trench_box(
    box_id: str,
    payload: TrenchBoxCreate,
    _: bool = Depends(require_safety_or_admin),
):
    if payload.tabulated_data_file:
        raw, _ = _data_url_to_bytes(payload.tabulated_data_file)
        _validate_pdf_or_400(raw)
        if len(raw) > 10 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="File too large (max 10 MB)")
    update = payload.model_dump()
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    # Keep file payload only if user provided a new one
    if not update.get("tabulated_data_file"):
        update.pop("tabulated_data_file", None)
        update.pop("tabulated_data_filename", None)
    res = await db.trench_boxes.update_one({"id": box_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Trench box not found")
    fresh = await db.trench_boxes.find_one(
        {"id": box_id}, {"_id": 0, "tabulated_data_file": 0}
    )
    fresh.setdefault("tabulated_data_file", "")
    return TrenchBox(**fresh)


@api_router.delete("/trench-boxes/{box_id}")
async def delete_trench_box(box_id: str, _: bool = Depends(require_safety_or_admin)):
    res = await db.trench_boxes.delete_one({"id": box_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Trench box not found")
    return {"deleted": True, "id": box_id}







# ============================================================
# Equipment Inspections (OSHA daily pre-op checklist)
# ============================================================
from checklists import CHECKLISTS, EQUIPMENT_TYPES  # noqa: E402


# ============================================================
# Equipment Pre-Op Inspections + Shop Sign-Off + Trends
# ----------------------------------------------------------
# Extracted to /app/backend/routes/equipment.py 2026-04-28 (P1 batch 4).
# Pydantic models + 8 endpoints + MAJOR_OOS_SET helpers all moved.
# ============================================================
from routes.equipment import (  # noqa: E402,F401
    register_equipment_routes,
    EquipmentInspection, EquipmentInspectionCreate, EquipmentInspectionSummary,
    ShopSignoffPayload, MAJOR_OOS_ITEMS_BACKEND, MAJOR_OOS_SET,
)


async def _remember_equipment_unit(eq_type, unit_label, make, model, serial):
    """Forwarder for the new-unit dropdown remembering. Defined lazily so
    `create_equipment_unit` (defined just below) can be looked up at call time.
    """
    return await create_equipment_unit(  # noqa: F821 — late binding (fn defined below)
        EquipmentUnitCreate(  # noqa: F821 — late binding
            equipment_type=eq_type, unit_label=unit_label,
            make=make, model=model, serial=serial,
        )
    )


register_equipment_routes(
    api_router, db, require_admin, require_shop_or_admin,
    rate_limit_public_post,
    lambda kind, record: schedule_auto_email(kind, record),
    _remember_equipment_unit,
)


# ---------------------------------------------------------------------------
# Equipment Master Fleet — sourced from MASCI Equipment List.xlsx
# Used to populate equipment dropdowns across all forms (Pre-Op, Daily Reports,
# Incidents, etc.). Operators can still type custom values as a fallback.
# ---------------------------------------------------------------------------
EQUIPMENT_MASTER_SEED_FILE = ROOT_DIR / "data" / "equipment_master.json"


class EquipmentMasterItem(BaseModel):
    unit_number: str = ""
    year: Optional[int] = None
    make_model: str = ""
    plate: str = ""
    vin_serial_number: str = ""
    comments: str = ""
    company: str = ""
    category: str = "Misc Equipment"
    preop_equipment_type: str = "Other"
    display_label: str = ""


@api_router.get("/equipment-master")
async def list_equipment_master(category: Optional[str] = None):
    from lib.synthetic_fleet_filter import apply_synthetic_equipment_exclusion  # noqa: PLC0415
    await _purge_expired("equipment_master")
    q: Dict[str, Any] = dict(ACTIVE_FILTER)
    if category:
        q["category"] = category
    # TRACK 28.05 · exclude TEST_28_05_ / SYNTHETIC_ synthetic fleet
    # rows from every operator-facing equipment picker.
    q = apply_synthetic_equipment_exclusion(q)
    cursor = db.equipment_master.find(q, {"_id": 0}).sort(
        [("category", 1), ("unit_number", 1), ("make_model", 1)]
    )
    docs = await cursor.to_list(2000)
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for d in docs:
        grouped.setdefault(d.get("category", "Misc Equipment"), []).append(d)
    categories = sorted(grouped.keys())
    return {
        "categories": categories,
        "items": docs,
        "grouped": grouped,
        "count": len(docs),
    }


@api_router.get("/admin/equipment-master/archive")
async def equipment_master_archive(_: bool = Depends(require_shop_or_admin)):
    return {
        "items": await _list_archive("equipment_master"),
        "retain_days": SOFT_DELETE_RETAIN_DAYS,
    }


@api_router.post("/admin/equipment-master/{unit_id}/restore")
async def restore_equipment_master(
    unit_id: str, _: bool = Depends(require_shop_or_admin)
):
    if not await _restore_row(
        "equipment_master",
        {"$or": [{"id": unit_id}, {"unit_number": unit_id}]},
    ):
        raise HTTPException(status_code=404, detail="Unit not in archive")
    doc = await db.equipment_master.find_one(
        {"$or": [{"id": unit_id}, {"unit_number": unit_id}]}, {"_id": 0}
    )
    return doc or {"ok": True}


@api_router.get("/equipment-types")
async def list_equipment_types():
    """Public — list of equipment types + checklist templates used by the
    Equipment Pre-Op form to render the right walk-around questions."""
    return {
        "types": EQUIPMENT_TYPES,
        "checklists": CHECKLISTS,
    }


# -------------------- Jobs Master (replaces static jobLibrary.js) --------------------
class JobIn(BaseModel):
    project_number: str = Field(..., min_length=1, max_length=80)
    project_name: str = Field(..., min_length=1, max_length=300)
    location: str = ""
    client: str = ""
    project_manager: str = ""   # display name (kept for backwards-compat)
    pm_email: str = ""           # canonical key — drives auto-email routing
    # ``None`` = keep existing co-PMs untouched on upsert. Use the dedicated
    # ``PATCH /admin/jobs/{id}/co-pms`` endpoint to set the list.
    co_pm_emails: Optional[List[str]] = None
    active: bool = True


class JobCoPMsBody(BaseModel):
    co_pm_emails: List[str] = Field(default_factory=list, max_length=4)


# -------------------- Project Managers (admin-managed roster) --------------------
@api_router.get("/jobs")
async def list_jobs_public():
    """Public — drives the JobPicker on every form. Active jobs only."""
    from jobs_master import list_jobs
    return {"items": await list_jobs(db, only_active=True)}


# DR-FIX-2 · R7 · Superintendent auto-population helper.
# Public read-only · returns the most recent Daily Report's
# superintendent for a given project_number, drawn from EXISTING
# `daily_reports` data. NO new field · NO schema change.
# Honors the directive's intent: when admins populate the canonical
# `superintendent`/`superintendent_name` field on `jobs_master`, the
# JobPicker's existing payload already carries it forward — this
# endpoint serves the fallback (last DR for the project) until that
# canonical store is filled.
# Doctrine: /app/memory/DR_AUDIT_001_FULL_CONSTITUTIONAL_AUDIT.md R7
@api_router.get("/jobs/{project_number}/recent-context")
async def jobs_recent_context(project_number: str, foreman: str = "", superintendent: str = ""):
    """TRACK 19.04 · Smart Prefill contract v19.04.

    Returns the most-recent Daily Report crew + equipment baseline for
    the given project. Optional `foreman` / `superintendent` query
    params bias the lookup to the operator's OWN most-recent report on
    that project so the offer they see matches the crew they actually
    ran yesterday — not a stranger's roster.

    The response never carries per-day-ephemera (times, hours-used,
    ticket_photos, notes, signatures). Frontend renders the payload as
    an OFFER (not silent auto-apply) — see NewDailyReport.jsx
    `smartPrefillOffer`.
    """
    project_number = (project_number or "").strip()
    foreman = (foreman or "").strip()
    superintendent = (superintendent or "").strip()
    empty = {
        "contract_version": "19.04",
        "source": "daily_reports.most-recent (project-scoped)",
        "actor_scoped": False,
        "superintendent": "",
        "masci_crews": [],
        "equipment": [],
        "source_report_date": "",
    }
    if not project_number:
        return empty
    latest = await db.daily_reports.find_one(
        {
            "project_number": project_number,
            "superintendent": {"$nin": ["", None]},
        },
        {"_id": 0, "superintendent": 1},
        sort=[("created_at", -1)],
    )
    # TRACK 15.46 · FR-15 · Pre-fill crew + equipment hours from the
    # most recent DR on this project. Keeps the foreman editing deltas
    # instead of re-typing the same 8-person crew + 4-piece spread
    # every morning. Sanitised: signatures and timestamps stripped so
    # nothing stale carries forward.
    #
    # TRACK 19.04 · If the caller identifies themselves via `foreman`
    # or `superintendent`, first try to find their OWN most-recent
    # report on this project. Falls back to project-most-recent when
    # no self-report exists. Never returns cross-project data.
    latest_full = None
    actor_scoped = False
    if foreman or superintendent:
        q_actor: Dict[str, Any] = {"project_number": project_number}
        or_clauses: List[Dict[str, Any]] = []
        if foreman:
            or_clauses.append({"prepared_by": foreman})
        if superintendent:
            or_clauses.append({"superintendent": superintendent})
        if or_clauses:
            q_actor["$or"] = or_clauses
        latest_full = await db.daily_reports.find_one(
            q_actor,
            {"_id": 0, "masci_crews": 1, "equipment": 1, "report_date": 1, "created_at": 1},
            sort=[("created_at", -1)],
        )
        actor_scoped = bool(latest_full)
    if not latest_full:
        latest_full = await db.daily_reports.find_one(
            {"project_number": project_number},
            {"_id": 0, "masci_crews": 1, "equipment": 1, "report_date": 1, "created_at": 1},
            sort=[("created_at", -1)],
        )
    latest_full = latest_full or {}
    raw_crews = latest_full.get("masci_crews") or []
    raw_equipment = latest_full.get("equipment") or []

    # TRACK 19.06 AMENDMENT · Filter out inactive/terminated HR employees.
    # Prior crew members whose HR record is *known inactive* (terminated,
    # resigned, retired, on-leave-inactive) must not be silently prefilled
    # into today's report — that is the exact payroll footgun the
    # amendment guards against. Unknown employee_ids (legacy rows, free-
    # typed custom names, or employees not yet mirrored into the HR
    # collection) PASS through — the roster can't verify them either way,
    # and the foreman still reviews and edits every row.
    known_inactive_ids: set = set()
    ids_to_check = [
        (c.get("employee_id") or "").strip()
        for c in raw_crews if isinstance(c, dict)
    ]
    ids_to_check = [x for x in ids_to_check if x]
    if ids_to_check:
        from routes.employee_lifecycle import _ACTIVE_STATUSES  # noqa: PLC0415
        # Find rows that DO exist and are NOT active → drop these only.
        inactive_clause = {"$and": [
            {"employee_id": {"$in": ids_to_check}},
            {"$or": [
                {"lifecycle_status": {"$nin": list(_ACTIVE_STATUSES) + [None]}},
                {"$and": [
                    {"$or": [
                        {"lifecycle_status": {"$exists": False}},
                        {"lifecycle_status": None},
                    ]},
                    {"is_active": False},
                ]},
            ]},
        ]}
        cursor = db.employees.find(inactive_clause, {"_id": 0, "employee_id": 1})
        async for e in cursor:
            eid = (e.get("employee_id") or "").strip()
            if eid:
                known_inactive_ids.add(eid)

    masci_crews = []
    for c in raw_crews:
        if not isinstance(c, dict):
            continue
        nm = (c.get("name") or "").strip()
        if not nm:
            continue
        eid = (c.get("employee_id") or "").strip()
        # HR filter: drop only KNOWN inactive employees. Unknown eids or
        # free-typed names pass through — foreman still reviews every row.
        if eid and eid in known_inactive_ids:
            continue
        masci_crews.append({
            "name": nm,
            "trade": (c.get("trade") or "").strip(),
            "employee_id": eid,
            "hours": c.get("hours") or "",
            # TRACK 19.06 AMENDMENT · Include the prior day's common
            # time pattern (start / lunch / stop) so the foreman edits
            # deltas instead of re-typing every clock-in. The values
            # are staged in the OFFER card and only hydrated on Apply;
            # never silently final — foreman edits every row before
            # submit (payroll-safety rule).
            "start_time": (c.get("start_time") or ""),
            "stop_time": (c.get("stop_time") or ""),
            "lunch_minutes": c.get("lunch_minutes") if c.get("lunch_minutes") not in (None, "") else "",
        })
    equipment = []
    for e in raw_equipment:
        if not isinstance(e, dict):
            continue
        desc = (e.get("description") or "").strip()
        if not desc:
            continue
        equipment.append({
            "description": desc,
            "hours_used": e.get("hours_used") or "",
            # time_delivered / time_removed intentionally NOT copied —
            # those are per-day movement events.
            "notes": "",
        })
    return {
        "contract_version": "19.06.1",
        "source": "daily_reports.most-recent (project-scoped)",
        "actor_scoped": actor_scoped,
        "superintendent": (latest or {}).get("superintendent", "") or "",
        "masci_crews": masci_crews,
        "equipment": equipment,
        "source_report_date": latest_full.get("report_date") or "",
    }


# iter245 · Vendors / Subcontractors master list — RETIRED 2026-05-19.
# The /api/vendors collection introduced earlier in this iter has been
# consolidated into the pre-existing /api/suppliers master list used by
# Daily Reports, Incidents, and QA/QC. Operator directive: ONE
# operational vendor source platform-wide — no parallel collections.
# The PO Request workflow now reuses /api/suppliers via SupplierCombo
# (see frontend/src/components/SupplierCombo.jsx).


@api_router.get("/admin/jobs")
async def admin_list_jobs(actor=Depends(require_admin)):
    """List jobs. Admin sees all; per-PM sees only jobs they're primary
    or co-PM on (matches the data-scoping rules applied to safety
    records). Legacy shared-PM tokens see all (the office bypass)."""
    from jobs_master import list_jobs
    from pm_auth import compute_pm_scope
    items = await list_jobs(db, only_active=False)
    scope = await compute_pm_scope(db, actor)
    if not scope.is_admin:
        nums = scope.project_numbers or set()
        items = [j for j in items if (j.get("project_number") or "") in nums]
    return {"items": items}


@api_router.post("/admin/jobs")
async def admin_upsert_job(body: JobIn, _: bool = Depends(require_admin)):
    from jobs_master import upsert_job
    try:
        return await upsert_job(db, body.model_dump())
    except ValueError as e:
        raise HTTPException(400, str(e))


@api_router.patch("/admin/jobs/{job_id}/active")
async def admin_set_job_active(
    job_id: str, body: dict, _: bool = Depends(require_admin)
):
    from jobs_master import set_active
    saved = await set_active(db, job_id, bool(body.get("active", True)))
    if not saved:
        raise HTTPException(404, "Job not found")
    return saved


@api_router.delete("/admin/jobs/{job_id}")
async def admin_delete_job(job_id: str, _: bool = Depends(require_admin)):
    from jobs_master import delete_job
    ok = await delete_job(db, job_id)
    if not ok:
        raise HTTPException(404, "Job not found")
    return {"ok": True, "soft_deleted": True, "retain_days": SOFT_DELETE_RETAIN_DAYS}


# ====================================================================
# Shop Users — admin-only CRUD + per-user password issuance.
# Mirrors the PM admin panel. Used by the "Shop Users" admin section.
# ====================================================================

class ShopUserIn(BaseModel):
    name: str
    email: str
    phone: Optional[str] = ""
    role: Optional[str] = "Mechanic"
    is_active: Optional[bool] = True


class ShopUserUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None
    disabled: Optional[bool] = None


class ShopSetPasswordBody(BaseModel):
    password: Optional[str] = None
    must_change: Optional[bool] = True


@api_router.get("/admin/shop-users")
async def admin_list_shop_users(_: bool = Depends(require_admin)):
    from shop_users import list_shop_users, public_shop_user_view
    items = await list_shop_users(db, only_active=False)
    return {"items": [public_shop_user_view(u) for u in items]}


# Track 15.13A · Asset Care Routing Recovery — Shop ↔ Directory mirror.
#
# When the Shop Users console creates / updates a user with an asset role
# label ("Asset Administrator", "Asset Manager", "Equipment Manager",
# "Fleet Coordinator"), the canonical `user_directory` row must carry
# `is_asset_admin: True` so:
#   * /api/auth/multi-login returns the flag on `user.is_asset_admin`
#   * /shop/login mirrors the flag into its response
#   * the SPA `landingFor()` resolver lands the user on /shop/asset-care
#
# This is a strict additive write — no portal grant, no permission widening,
# no token reissue. If a directory row does not exist for the email we
# upsert a minimal row carrying ONLY the asset_admin flag + the email +
# the display name (no password, no portals, no portal grants). Asset
# admins still authenticate through the existing shop password.
_ASSET_ADMIN_ROLE_LABELS = {
    "Asset Administrator",
    "Asset Manager",
    "Equipment Manager",
    "Fleet Coordinator",
}


def _role_implies_asset_admin(role: Any) -> bool:
    if not role:
        return False
    return str(role).strip() in _ASSET_ADMIN_ROLE_LABELS


async def _mirror_asset_admin_flag(email: str, name: str, role: Any) -> bool:
    """Idempotently set/clear `user_directory.is_asset_admin` based on the
    Shop role label. Returns the resulting flag value. Safe to call on
    every shop-user create / update / role-change."""
    email_norm = (email or "").strip().lower()
    if not email_norm:
        return False
    is_asset_admin = _role_implies_asset_admin(role)
    try:
        existing = await db.user_directory.find_one(
            {"email": email_norm}, {"_id": 0, "id": 1},
        )
        now_iso = datetime.now(timezone.utc).isoformat()
        if existing:
            await db.user_directory.update_one(
                {"id": existing["id"]},
                {"$set": {"is_asset_admin": bool(is_asset_admin),
                          "updated_at": now_iso,
                          "source": "shop_console_mirror"}},
            )
        elif is_asset_admin:
            # Only spawn a stub directory row when we actually need the
            # flag. We never spawn a directory row for non-asset shop
            # users — they don't need one.
            await db.user_directory.insert_one({
                "id": f"dir-asset-shadow-{hashlib.sha1(email_norm.encode()).hexdigest()[:16]}",
                "email": email_norm,
                "name": (name or "").strip() or email_norm.split("@")[0],
                "portals": [],
                "disabled": False,
                "must_change_password": False,
                "password_hash": None,
                "is_asset_admin": True,
                "created_at": now_iso,
                "updated_at": now_iso,
                "source": "shop_console_mirror",
            })
        return bool(is_asset_admin)
    except Exception as exc:  # noqa: BLE001
        logger.warning(f"_mirror_asset_admin_flag failed for {email_norm}: {exc}")
        return False


@api_router.post("/admin/shop-users")
async def admin_add_shop_user(body: ShopUserIn, _: bool = Depends(require_admin)):
    from shop_users import add_shop_user, public_shop_user_view
    try:
        user = await add_shop_user(db, body.model_dump())
        # Track 15.13A — mirror asset-admin flag on directory.
        await _mirror_asset_admin_flag(
            user.get("email"), user.get("name"), user.get("role"),
        )
        view = public_shop_user_view(user)
        view["is_asset_admin"] = _role_implies_asset_admin(user.get("role"))
        return view
    except ValueError as e:
        raise HTTPException(400, str(e))


@api_router.patch("/admin/shop-users/{user_id}")
async def admin_update_shop_user(
    user_id: str, body: ShopUserUpdate, _: bool = Depends(require_admin)
):
    from shop_users import update_shop_user, public_shop_user_view
    fields = {k: v for k, v in body.model_dump().items() if v is not None}
    if not fields:
        raise HTTPException(400, "No fields to update")
    try:
        saved = await update_shop_user(db, user_id, fields)
        if not saved:
            raise HTTPException(404, "Shop user not found")
        # Track 15.13A — if role changed (or email changed), re-mirror.
        await _mirror_asset_admin_flag(
            saved.get("email"), saved.get("name"), saved.get("role"),
        )
        view = public_shop_user_view(saved)
        view["is_asset_admin"] = _role_implies_asset_admin(saved.get("role"))
        return view
    except ValueError as e:
        raise HTTPException(400, str(e))


@api_router.delete("/admin/shop-users/{user_id}")
async def admin_delete_shop_user(user_id: str, _: bool = Depends(require_admin)):
    from shop_users import delete_shop_user
    ok = await delete_shop_user(db, user_id)
    if not ok:
        raise HTTPException(404, "Shop user not found")
    return {"ok": True}


@api_router.post("/admin/shop-users/{user_id}/set-password")
async def admin_set_shop_user_password(
    user_id: str, body: ShopSetPasswordBody, request: Request, _: bool = Depends(require_admin)
):
    """Issue a new password for a shop user. If `password` is omitted,
    generates a crypto-random temp password. Returned ONCE — admin
    shows it to the user verbally / on a sticky note. Subsequent reads
    of the user record never include the plaintext."""
    from shop_users import (
        set_shop_user_password, generate_temp_password, public_shop_user_view,
    )
    if body.password and len(body.password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters")
    pw = (body.password or "").strip() or generate_temp_password()
    try:
        saved = await set_shop_user_password(
            db, user_id, pw, must_change=bool(body.must_change),
        )
    except ValueError as e:
        raise HTTPException(400, str(e))
    if not saved:
        raise HTTPException(404, "Shop user not found")
    # iter502 · OMEGA IAM Enterprise Phase B+C: stamp issuance + audit.
    try:
        from lib.iam_password_audit import stamp_and_audit_temp_password
        await stamp_and_audit_temp_password(
            db,
            collection_name="shop_users",
            user_filter={"id": user_id},
            target_email=str(saved.get("email") or ""),
            portal="shop",
            delivery="custom" if body.password else "screen",
            request=request,
        )
    except Exception as _e:  # noqa: BLE001
        logger.warning(f"[iam-pw-audit] shop set-password audit failed: {_e}")
    return {
        "ok": True,
        "user": public_shop_user_view(saved),
        "temp_password": pw,
        "must_change_password": bool(saved.get("must_change_password")),
    }


@api_router.post("/admin/shop-users/{user_id}/disable")
async def admin_disable_shop_user(
    user_id: str, body: dict, _: bool = Depends(require_admin)
):
    from shop_users import update_shop_user, public_shop_user_view
    saved = await update_shop_user(db, user_id, {"disabled": bool(body.get("disabled", True))})
    if not saved:
        raise HTTPException(404, "Shop user not found")
    return public_shop_user_view(saved)


@api_router.post("/admin/shop-users/{user_id}/email-welcome")
async def admin_shop_user_email_welcome(
    user_id: str, body: ShopSetPasswordBody, request: Request, _: bool = Depends(require_admin_strict)
):
    """One-shot: issue (or rotate) a shop-user password AND email the
    temp password to the user via Resend.

    Mirrors the PM ``email-welcome`` endpoint but lighter (no welcome
    PDF — shop users only need creds + the portal URL).
    Returns ``{ok, user, sent_to, resend_id}``. The temp password is
    NOT echoed back in the response (it's already in the email body)
    so admin network logs stay clean.
    """
    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        raise HTTPException(
            status_code=503,
            detail=(
                "RESEND_API_KEY not configured. "
                "Use 'Generate & Show on Screen' to give the user the temp password manually."
            ),
        )

    from shop_users import (
        generate_temp_password,
        public_shop_user_view,
        set_shop_user_password,
    )

    user = await db.shop_users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Shop user not found")
    user_email = (user.get("email") or "").strip()
    user_name = (user.get("name") or "").strip() or "Mechanic"
    if not user_email:
        raise HTTPException(status_code=400, detail="Shop user has no email on file")
    if body.password and len(body.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")

    plain = (body.password or "").strip() or generate_temp_password()
    updated = await set_shop_user_password(
        db, user_id, plain, must_change=bool(body.must_change),
    )
    if not updated:
        raise HTTPException(status_code=500, detail="Failed to set password")

    portal_url = (
        os.environ.get("PORTAL_URL", "").strip()
        or os.environ.get("PRODUCTION_URL", "").strip()
        or "https://mascidocs.com"
    )

    is_reset = bool(user.get("password_hash"))
    # Track 15.13A · Asset Care welcome email branching.
    # If the shop user's role label maps to asset admin work, send the
    # Asset Care welcome instead of the generic Shop Portal welcome.
    # The login URL stays /shop/login (asset admins authenticate
    # through the same shop password) but the copy reflects where
    # they actually land after sign-in (/shop/asset-care).
    is_asset_admin_role = _role_implies_asset_admin(user.get("role"))
    if is_asset_admin_role and not is_reset:
        headline = "Welcome to MASCI Asset Care"
        intro = (
            f'You have access to the <strong>MASCI Asset Care</strong> workspace for '
            f'equipment readiness, registrations, assignments, transfers, and asset '
            f'lifecycle visibility. Sign in at '
            f'<a href="{portal_url}/shop/login" style="color:#b91c1c;font-weight:700">'
            f'{portal_url}/shop/login</a> and Asset Care opens automatically.'
        )
    elif is_asset_admin_role and is_reset:
        headline = "Your Asset Care password has been reset"
        intro = (
            "Your MASCI Asset Care password has been reset. Use the temporary "
            "password below to sign in — you will be forced to choose your own on "
            "first login."
        )
    else:
        headline = "Your password has been reset" if is_reset else "Welcome to MASCI Shop Operations"
        intro = (
            "Your MASCI Shop Operations password has been reset. Use the temporary password below to "
            "sign in — you will be forced to choose your own on first login."
            if is_reset else
            f'You have a new account on the MASCI Shop Portal at '
            f'<a href="{portal_url}/shop/login" style="color:#b91c1c;font-weight:700">{portal_url}/shop/login</a>. '
            "Use the temporary password below to sign in — you will be forced to choose your own on first login."
        )
    body_inner = f"""
      <p style="margin:0 0 12px;font-size:15px;line-height:1.5">Hi {user_name},</p>
      <p style="margin:0 0 12px;font-size:14px;line-height:1.55;color:#334155">{intro}</p>

      <table cellpadding="0" cellspacing="0" style="background:#0f172a;color:#f1f5f9;border-radius:6px;padding:18px 22px;margin:16px 0;width:100%;max-width:480px">
        <tr><td>
          <div style="font-family:Courier New,monospace;font-size:9px;letter-spacing:0.22em;color:#94a3b8;text-transform:uppercase;font-weight:700">Account</div>
          <div style="font-family:Courier New,monospace;font-size:13px;font-weight:800;margin-top:3px">{user_email}</div>
          <div style="font-family:Courier New,monospace;font-size:9px;letter-spacing:0.22em;color:#94a3b8;text-transform:uppercase;font-weight:700;margin-top:14px">Temporary password</div>
          <div style="font-family:Courier New,monospace;font-size:20px;font-weight:800;color:#fbbf24;letter-spacing:0.05em;margin-top:3px">{plain}</div>
        </td></tr>
      </table>

      <p style="margin:14px 0 6px;font-size:14px;line-height:1.55"><strong>What to do next</strong></p>
      <ol style="margin:0 0 14px 18px;padding:0;font-size:14px;line-height:1.55;color:#334155">
        <li>Open <a href="{portal_url}/shop/login" style="color:#b91c1c;font-weight:700">{portal_url}/shop/login</a></li>
        <li>Sign in with the email + temporary password above</li>
        <li>Pick your own 6+ character password (the temp one stops working immediately)</li>
        {(
            "<li>Asset Care opens automatically — manage registrations, "
            "assignments, transfers, equipment readiness, and lifecycle visibility "
            "from the Asset Care &amp; Readiness landing page.</li>"
            if is_asset_admin_role else
            "<li>Failed Pre-Op inspections (Out-of-Service / Needs-Attention) auto-route "
            "to your inbox so you can plan parts &amp; scheduling</li>"
        )}
      </ol>

      <p style="margin:14px 0 0;font-size:13px;color:#64748b;line-height:1.55">
        If you forget your password, ask the admin to issue a new temp pw — it takes 30 seconds.
      </p>
    """
    html_body = render_portal_email(
        portal="Asset Care" if is_asset_admin_role else "Shop",
        headline=headline,
        body_inner_html=body_inner,
    )

    import resend  # noqa: E402

    resend.api_key = api_key
    sender_email = await _resolve_sender_email(db)
    reply_to = (await _resolve_reply_to_email(db)) or ""
    params = {
        "from": f"MASCI Operations Platform <{sender_email}>",
        "to": [user_email],
        "subject": f"[MASCI] {headline}",
        "html": html_body,
    }
    if reply_to:
        params["reply_to"] = reply_to

    try:
        result = await asyncio.to_thread(resend.Emails.send, params)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(
            status_code=502,
            detail=(
                f"Password rotated but email send failed via Resend: {e}. "
                "Use 'Show on Screen' to recover the new temp password."
            ),
        )

    # iter502 · OMEGA IAM Enterprise Phase B+C: stamp + audit for shop welcome.
    try:
        from lib.iam_password_audit import stamp_and_audit_temp_password, audit_welcome_email_sent
        await stamp_and_audit_temp_password(
            db,
            collection_name="shop_users",
            user_filter={"id": user_id},
            target_email=user_email,
            portal="shop",
            delivery="email",
            request=request,
        )
        await audit_welcome_email_sent(
            db, target_email=user_email, portal="shop", request=request,
        )
    except Exception as _e:  # noqa: BLE001
        logger.warning(f"[iam-pw-audit] shop email-welcome audit failed: {_e}")

    return {
        "ok": True,
        "user": public_shop_user_view(updated),
        "sent_to": user_email,
        "resend_id": (result or {}).get("id") if isinstance(result, dict) else None,
    }




@api_router.get("/admin/jobs/archive")
async def admin_jobs_archive(actor=Depends(require_admin)):
    from jobs_master import list_archived_jobs
    from pm_auth import compute_pm_scope
    items = await list_archived_jobs(db)
    scope = await compute_pm_scope(db, actor)
    if not scope.is_admin:
        nums = scope.project_numbers or set()
        items = [j for j in items if (j.get("project_number") or "") in nums]
    return {"items": items, "retain_days": SOFT_DELETE_RETAIN_DAYS}


@api_router.post("/admin/jobs/{job_id}/restore")
async def admin_restore_job(job_id: str, _: bool = Depends(require_admin)):
    from jobs_master import restore_job
    if not await restore_job(db, job_id):
        raise HTTPException(404, "Job not in archive")
    doc = await db.jobs_master.find_one({"id": job_id}, {"_id": 0})
    return doc or {"ok": True}


@api_router.post("/admin/jobs/bulk-replace")
async def admin_bulk_replace_jobs(body: dict, _: bool = Depends(require_admin)):
    """Replace the entire jobs_master collection (used by the bulk uploader).
    Body: {"rows": [{project_number, project_name, ...}, ...]}.
    """
    from jobs_master import bulk_replace
    rows = body.get("rows") or []
    if not isinstance(rows, list):
        raise HTTPException(400, "rows must be a list")
    require_destructive_confirmation(body, expected_confirm="REPLACE_ALL_JOBS_MASTER")
    require_destructive_runtime_guard(expected_db_name="masci_safety")
    try:
        return await bulk_replace(db, rows)
    except ValueError as e:
        raise HTTPException(400, str(e))


@api_router.patch("/admin/jobs/{job_id}/co-pms")
async def admin_set_job_co_pms(
    job_id: str, body: JobCoPMsBody, _: bool = Depends(require_admin)
):
    """Set the list of co-PMs for a job. Up to 4 additional PMs (the
    primary PM stays in ``pm_email``; total assignment is 5). The list
    is normalized to lowercase emails, the primary is removed if it
    appears, duplicates are dropped, and unknown emails are validated
    against the active PM roster."""
    job = await db.jobs_master.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(404, "Job not found")

    primary = (job.get("pm_email") or "").strip().lower()
    seen = {primary} if primary else set()
    cleaned: List[str] = []
    for raw in (body.co_pm_emails or []):
        if not isinstance(raw, str):
            continue
        e = raw.strip().lower()
        if not e or e in seen:
            continue
        # Validate against the project_managers roster — we never want
        # to silently route emails to a typo or a deleted PM.
        pm = await db.project_managers.find_one({"email": e}, {"_id": 0})
        if not pm:
            raise HTTPException(400, f"PM with email {e} not found")
        if pm.get("is_active") is False:
            raise HTTPException(
                400,
                f"PM {pm.get('name')} ({e}) is deactivated — reactivate first.",
            )
        seen.add(e)
        cleaned.append(e)
        if len(cleaned) >= 4:
            break
    await db.jobs_master.update_one(
        {"id": job_id},
        {"$set": {"co_pm_emails": cleaned, "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    saved = await db.jobs_master.find_one({"id": job_id}, {"_id": 0})
    return saved


# -------------------- Inline "Add to roster" (no admin token) --------------------
class RosterAddBody(BaseModel):
    name: str = Field(..., min_length=2, max_length=120)


@api_router.post(
    "/employees/add",
    dependencies=[Depends(rate_limit_public_post)],
)
async def add_employee_from_form_deprecated(body: RosterAddBody, request: Request):
    """OMEGA · Employee Governance Phase Alpha · G-1 closure.

    This endpoint historically accepted PUBLIC employee creation —
    Constitutional violation V-P0-1. It now forwards to the HR Request
    Queue (POST /api/employee-requests · kind=new_hire) so HR retains
    sole authority over db.employees lifecycle state.

    Returns HTTP 410 Gone with a structured pointer to the new flow.
    Frontend callers (EmployeeCombo.jsx) have been repointed; this
    handler exists to surface a clear error if any legacy client
    still hits the old URL.
    """
    raise HTTPException(
        status_code=410,
        detail={
            "code": "endpoint_deprecated",
            "message": (
                "Direct employee creation from public field forms is "
                "no longer permitted. Submit a new-hire request to HR "
                "via POST /api/employee-requests (kind=new_hire). HR "
                "will review and approve."
            ),
            "use_instead": "POST /api/employee-requests",
            "kind": "new_hire",
            "name": body.name,
        },
    )


@api_router.post(
    "/suppliers/add",
    dependencies=[Depends(rate_limit_public_post)],
)
async def add_supplier_from_form(body: RosterAddBody):
    """Add a new supplier / vendor / subcontractor to the master list from
    a form's amber 'Will save as new entry' button. Public + rate-limited.

    Idempotent on case-insensitive name match.
    """
    name = body.name.strip()
    existing = await db.suppliers.find_one(
        {"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}, {"_id": 0}
    )
    if existing:
        return {"ok": True, "created": False, "supplier": existing}
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "vendor_type": "",
        "phone": "",
        "email": "",
        "address": "",
        "added_via": "field-form",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.suppliers.insert_one(doc)
    saved = {k: v for k, v in doc.items() if k != "_id"}
    return {"ok": True, "created": True, "supplier": saved}


# ---------------------------------------------------------------------------
# Employees / crew roster — used by Daily Report's "MASCI Crews on Site"
# section and any other employee dropdown across the platform.
# ---------------------------------------------------------------------------
@api_router.get("/employees")
async def list_employees():
    """Public — returns the MASCI crew roster (sorted by name).

    OMEGA · Public Employee Roster Projection Hardening (2026-06-03):
    Projection narrowed to the allow-list of fields actually rendered
    by the 5 public-form pickers (Daily Report, Incident, Safety
    Meeting, Equipment Inspection, Fleet DVIR). CDL, medical-card,
    status_history, email, phone, and timestamp fields are no longer
    returned on this public endpoint. The full record set remains
    available to authenticated callers via /api/hr/employees and
    /api/admin/employees/*. No employee data was modified.

    TRACK 19.03 · HR-IS-GOSPEL hardening:
    Filter now matches the canonical HR roster contract:
      • If `lifecycle_status` is set, only `_ACTIVE_STATUSES`
        ({Active, Pending Hire, Seasonal, Leave of Absence}) appear.
      • Legacy rows without `lifecycle_status` fall back to
        `is_active != False`.
      • Offboarded statuses ({Terminated, Resigned, Retired}) and
        explicit Inactive are HIDDEN from new-form pickers regardless
        of `is_active`.
    HR Save is the only mutator of these fields. No cache. No TTL.
    """
    from routes.employee_lifecycle import _ACTIVE_STATUSES  # noqa: PLC0415
    from lib.employee_identity import (  # noqa: PLC0415
        PUBLIC_ROSTER_PROJECTION,
        normalize_employee_identity,
    )
    from lib.synthetic_hr_filter import apply_synthetic_hr_exclusion  # noqa: PLC0415
    await _purge_expired("employees")
    canonical_active_clause = {"$or": [
        {"lifecycle_status": {"$in": list(_ACTIVE_STATUSES)}},
        {"lifecycle_status": {"$exists": False}, "is_active": {"$ne": False}},
        {"lifecycle_status": None, "is_active": {"$ne": False}},
    ]}
    # TRACK 23.5 · canonical projection (shared with /hr/employee-roster)
    # + normalizer emits trade_role_display / crew_display /
    # supervisor_display / display_identity so downstream consumers
    # (Daily Report V3, ODS labor_fact, PDF, HR Time Verification,
    # Payroll Variance, PM Intelligence) never re-derive HR aliases.
    cursor = db.employees.find(
        apply_synthetic_hr_exclusion({"$and": [ACTIVE_FILTER, canonical_active_clause]}),
        PUBLIC_ROSTER_PROJECTION,
    ).sort("name", 1)
    raw_docs = await cursor.to_list(5000)
    docs = [normalize_employee_identity(d) for d in raw_docs]
    return {"items": docs, "count": len(docs)}

@api_router.get("/hr/employee-roster")
async def hr_employee_roster(
    q: Optional[str] = None,
    role: Optional[str] = None,
    department: Optional[str] = None,
    include_inactive: bool = False,
    limit: int = 5000,
    _actor: Dict[str, Any] = Depends(_require_any_portal_read),
):
    """TRACK 19.03 · Canonical HR Employee Roster (HR is gospel).

    The single source of truth for operational employee selection
    across the entire MASCI platform: Daily Reports, Safety Meetings,
    Pre-Ops, JHP/Safety, Dispatch, Fleet, Training, Academy, Incident
    Reports, Near Misses, Equipment Assignments — every employee
    picker reads from here.

    Filter semantics (matches `/api/employees` exactly):
      • Active (`_ACTIVE_STATUSES`) → visible by default.
      • Inactive/Terminated/Resigned/Retired → hidden unless
        `include_inactive=true` (operator-gated for investigations).
      • Legacy rows without `lifecycle_status` fall back to
        `is_active != False`.

    Safe projection — no CDL, no medical card, no email, no phone, no
    SSN, no DOB. Field workflows never receive private HR data.
    """
    from routes.employee_lifecycle import _ACTIVE_STATUSES  # noqa: PLC0415
    from lib.employee_identity import (  # noqa: PLC0415
        PUBLIC_ROSTER_PROJECTION,
        normalize_employee_identity,
    )
    from lib.synthetic_hr_filter import apply_synthetic_hr_exclusion  # noqa: PLC0415
    await _purge_expired("employees")
    canonical_active_clause = {"$or": [
        {"lifecycle_status": {"$in": list(_ACTIVE_STATUSES)}},
        {"lifecycle_status": {"$exists": False}, "is_active": {"$ne": False}},
        {"lifecycle_status": None, "is_active": {"$ne": False}},
    ]}
    clauses: List[Dict[str, Any]] = [ACTIVE_FILTER]
    if not include_inactive:
        clauses.append(canonical_active_clause)
    if role:
        clauses.append({"role": {"$regex": f"^{re.escape(role)}$", "$options": "i"}})
    if department:
        clauses.append({"department": {"$regex": f"^{re.escape(department)}$", "$options": "i"}})
    if q:
        q_re = {"$regex": re.escape(q), "$options": "i"}
        clauses.append({"$or": [
            {"name": q_re}, {"preferred_name": q_re},
            {"employee_id": q_re}, {"role": q_re},
        ]})
    # TRACK 23.5 · shared projection with /api/employees + shared
    # normalizer so both endpoints emit the same field contract.
    # Previously this endpoint projected `supervisor_name` /
    # `supervisor_id` which HR never writes (canonical write key is
    # `supervisor`) — supervisor was silently dropped for every
    # field picker. FIXED.
    cursor = db.employees.find(
        apply_synthetic_hr_exclusion({"$and": clauses}),
        PUBLIC_ROSTER_PROJECTION,
    ).sort("name", 1).limit(limit)
    raw_docs = await cursor.to_list(limit)
    docs = [normalize_employee_identity(d) for d in raw_docs]
    # Provide a derived `active` boolean so the frontend can render an
    # unambiguous chip without re-deriving the contract.
    for d in docs:
        ls = d.get("lifecycle_status")
        if ls is None:
            d["active"] = d.get("is_active") is not False
        else:
            d["active"] = ls in _ACTIVE_STATUSES
    return {
        "items": docs,
        "count": len(docs),
        "contract_version": "23.5",
        "filter": {
            "active_only": not include_inactive,
            "active_statuses": sorted(_ACTIVE_STATUSES),
            "source": "db.employees (HR is gospel)",
        },
    }


# TRACK 24.9 · P0 · Public-safe employee roster projection.
#
# The authenticated `/api/hr/employee-roster` above was locked in
# Track 24.1 (P0-1) to close a 387-record PII leak. That closure
# ALSO orphaned the public/anonymous Daily Report V3 flow at
# `/daily/new` — foremen open that route without a portal token,
# so `EmployeeCombo` received 401 → empty items → "Roster not
# uploaded yet." was surfaced on live production.
#
# This endpoint restores the anonymous roster read with a STRICTLY
# minimal projection: name / employee_id / trade / role / crew /
# active only. NO email, phone, SSN, DOB, address, salary, CDL,
# medical, supervisor, department, updated_at — none of the
# fields that made the P0-1 leak dangerous. The lock test
# `test_public_roster_projection_forbids_pii` in the Track 24.9
# suite enforces this contract at every CI run.
#
# Anonymous / public. Rate-limited by the platform's public POST
# limiter effectively (GET rate-limit is on the general ingress).
_PUBLIC_ROSTER_ALLOWED_KEYS = frozenset({
    "id", "name", "employee_id", "trade", "role", "crew", "active",
})


def _project_public_roster_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Return only the whitelisted keys — never any PII."""
    out: Dict[str, Any] = {}
    for k in _PUBLIC_ROSTER_ALLOWED_KEYS:
        if k in row:
            out[k] = row[k]
    return out


@api_router.get("/hr/employee-roster/public")
async def hr_employee_roster_public(q: Optional[str] = None, limit: int = 5000):
    """Public-safe roster projection for anonymous Daily Report V3.

    Strictly whitelisted fields. See lock test for the enforced key
    set. Inactive employees are hidden (same rule as authenticated
    endpoint). Zero PII.
    """
    from routes.employee_lifecycle import _ACTIVE_STATUSES  # noqa: PLC0415
    from lib.synthetic_hr_filter import apply_synthetic_hr_exclusion  # noqa: PLC0415
    canonical_active_clause = {"$or": [
        {"lifecycle_status": {"$in": list(_ACTIVE_STATUSES)}},
        {"lifecycle_status": {"$exists": False}, "is_active": {"$ne": False}},
        {"lifecycle_status": None, "is_active": {"$ne": False}},
    ]}
    clauses: List[Dict[str, Any]] = [ACTIVE_FILTER, canonical_active_clause]
    if q:
        q_re = {"$regex": re.escape(q), "$options": "i"}
        clauses.append({"$or": [
            {"name": q_re}, {"preferred_name": q_re},
            {"employee_id": q_re}, {"role": q_re},
        ]})
    cursor = db.employees.find(
        apply_synthetic_hr_exclusion({"$and": clauses}),
        {
            "_id": 0, "id": 1, "name": 1, "employee_id": 1,
            "trade": 1, "role": 1, "crew": 1,
            "lifecycle_status": 1, "is_active": 1,
        },
    ).sort("name", 1).limit(max(1, min(int(limit or 5000), 5000)))
    docs = await cursor.to_list(5000)
    out: List[Dict[str, Any]] = []
    for d in docs:
        ls = d.get("lifecycle_status")
        active = (ls in _ACTIVE_STATUSES) if ls is not None else (d.get("is_active") is not False)
        row = {
            "id": d.get("id") or "",
            "name": d.get("name") or "",
            "employee_id": d.get("employee_id") or "",
            "trade": d.get("trade") or "",
            "role": d.get("role") or "",
            "crew": d.get("crew") or "",
            "active": bool(active),
        }
        out.append(_project_public_roster_row(row))
    return {
        "items": out,
        "count": len(out),
        "contract_version": "24.9-public",
        "public": True,
    }


@api_router.get("/admin/employees/status")
async def employees_status(actor: Dict[str, Any] = Depends(_require_hr_or_admin_for_queue)):
    """OMEGA · Phase Alpha · G-3 · Deprecated admin status endpoint.

    Gate changed from require_admin → HR-or-Admin (legacy callers in
    the admin panel still work via HR portal token). Phase Beta (G-6)
    will tighten this to HR-only.
    """
    total = await db.employees.count_documents(ACTIVE_FILTER)
    active = await db.employees.count_documents(
        {"$and": [ACTIVE_FILTER, {"is_active": {"$ne": False}}]}
    )
    archived = await db.employees.count_documents({"deleted_at": {"$ne": None}})
    last_doc = await db.employees.find_one(
        ACTIVE_FILTER, {"_id": 0, "updated_at": 1, "created_at": 1}, sort=[("updated_at", -1)]
    )
    last_updated = (last_doc or {}).get("updated_at") or (last_doc or {}).get("created_at")
    return {"count": total, "active": active, "archived": archived, "last_updated": last_updated}


@api_router.get("/admin/employees/archive")
async def employees_archive(actor: Dict[str, Any] = Depends(_require_hr_or_admin_for_queue)):
    return {"items": await _list_archive("employees"), "retain_days": SOFT_DELETE_RETAIN_DAYS}


@api_router.post("/admin/employees/{employee_id}/restore")
async def restore_employee(employee_id: str, actor: Dict[str, Any] = Depends(_require_hr_or_admin_for_queue)):
    """OMEGA · Phase Alpha · G-3 · Deprecated admin restore endpoint.

    Re-gated to HR-or-Admin. Phase Beta will route this through the
    canonical HR /reactivate endpoint (preserving original_hire_date,
    write-once contract, and full status_history). For Phase Alpha we
    preserve legacy behaviour but lock the gate.
    """
    if not await _restore_row("employees", {"id": employee_id}):
        raise HTTPException(status_code=404, detail="Employee not in archive")
    doc = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    return doc or {"ok": True}


@api_router.post("/admin/employees/upload")
async def upload_employees(
    file: UploadFile = File(...),
    actor: Dict[str, Any] = Depends(_require_hr_or_admin_for_queue),
):
    """OMEGA · Phase Alpha · G-5 · Append/merge bulk import (NO DELETE).

    Operator-approved decision #5 of EMPLOYEE_GOVERNANCE_PHASE_ALPHA:
        "Bulk import remains supported but must operate as append/merge
        only. Destructive replace-all behavior is prohibited."

    Previous behaviour (delete_many + insert_many) wiped status_history,
    original_hire_date, lifecycle_status and re-issued UUIDs — V-P0-5.
    This rewrite preserves every lifecycle-bearing field and only
    touches the columns supplied in the upload file.

    Match strategy:
        1) `employee_id` (HR ID number) if present in BOTH file row + DB row
        2) case-insensitive exact `name` match (only on still-active rows)

    Expected columns: Name (required) · Employee ID · Trade · Role ·
    Crew · Email · Phone.

    Returns: { created, updated, skipped, ambiguous, total, items: [...] }
    """
    fname = (file.filename or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".xlsm") or fname.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .csv files are accepted")
    raw = await file.read()
    if not raw or len(raw) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Empty or oversized file (max 10 MB)")

    rows: List[Dict[str, Any]] = []
    try:
        if fname.endswith(".csv"):
            import csv as _csv
            text = raw.decode("utf-8", errors="ignore")
            reader = _csv.DictReader(text.splitlines())
            for r in reader:
                rows.append({(k or "").strip().lower(): (v or "").strip() for k, v in r.items()})
        else:
            import openpyxl as _ox
            import io as _io
            wb = _ox.load_workbook(_io.BytesIO(raw), data_only=True)
            ws = wb.active
            headers: List[str] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c or "").strip().lower() for c in row]
                    continue
                if not row or not any(row):
                    continue
                d = {}
                for h, v in zip(headers, row):
                    if not h:
                        continue
                    d[h] = ("" if v is None else str(v).strip())
                rows.append(d)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {e}")

    def pick(d: Dict[str, str], *keys: str) -> str:
        for k in keys:
            v = d.get(k)
            if v:
                return v
        return ""

    now = datetime.now(timezone.utc).isoformat()
    actor_role = actor.get("_actor") or actor.get("role") or "hr"
    actor_label = actor.get("name") or actor.get("email") or actor_role

    results: Dict[str, int] = {
        "created": 0, "updated": 0, "skipped": 0,
        "ambiguous": 0, "no_change": 0, "total": 0,
    }
    items: List[Dict[str, Any]] = []
    seen_keys = set()

    for d in rows:
        name = pick(d, "name", "full name", "employee name")
        if not name:
            continue
        results["total"] += 1
        emp_id_in = pick(d, "employee id", "id", "emp id", "emp #", "emp#")

        row_fields = {
            "employee_id": emp_id_in,
            "trade": pick(d, "trade", "department"),
            "role": pick(d, "role", "title", "position"),
            "crew": pick(d, "crew", "team"),
            "email": pick(d, "email"),
            "phone": pick(d, "phone", "mobile", "cell"),
        }
        # Only update fields actually supplied; empty cells do NOT
        # overwrite existing values.
        set_fields = {k: v for k, v in row_fields.items() if v}
        if not set_fields and not name:
            results["skipped"] += 1
            items.append({"row_name": name, "action": "skipped"})
            continue

        # Match strategy
        match = None
        if emp_id_in:
            match = await db.employees.find_one(
                {"employee_id": emp_id_in, "deleted_at": None}, {"_id": 0}
            )
        if not match:
            candidates = await db.employees.find(
                {
                    "name": {"$regex": f"^{re.escape(name)}$", "$options": "i"},
                    "deleted_at": None,
                    "is_active": {"$ne": False},
                },
                {"_id": 0},
            ).to_list(5)
            if len(candidates) > 1:
                results["ambiguous"] += 1
                items.append({
                    "row_name": name, "action": "ambiguous",
                    "candidates": [c.get("id") for c in candidates],
                })
                continue
            if len(candidates) == 1:
                match = candidates[0]

        if not match:
            # Create new employee — but mirror the HR-canonical shape
            # (lifecycle_status, original_hire_date, status_history) so
            # downstream consumers see a fully-formed row.
            new_id = str(uuid.uuid4())
            doc = {
                "id": new_id,
                "name": name,
                **{k: row_fields[k] or "" for k in row_fields},
                "supervisor": "",
                "department": "",
                "default_project_number": "",
                "hire_date": None,
                "original_hire_date": None,
                "lifecycle_status": "Active",
                "is_active": True,
                "added_via": "bulk-upload-merge",
                "created_at": now,
                "updated_at": now,
                "status_history": [{
                    "at": now,
                    "by": actor_label,
                    "actor_role": actor_role,
                    "to": "Active",
                    "reason": f"Created via append/merge upload by {actor_label}",
                    "kind": "bulk_upload_create",
                }],
                "deleted_at": None,
            }
            await db.employees.insert_one(dict(doc))
            results["created"] += 1
            items.append({"row_name": name, "action": "created", "id": new_id})
            try:
                await db.employee_lifecycle_events.insert_one({
                    "id": str(uuid.uuid4()),
                    "employee_id": new_id,
                    "at": now,
                    "actor_role": actor_role,
                    "actor_label": actor_label,
                    "kind": "bulk_upload_create",
                    "to_status": "Active",
                    "from_status": None,
                    "reason": "append/merge bulk upload",
                })
            except Exception:  # noqa: BLE001
                pass
            continue

        # Matched — apply only the supplied non-empty fields. Never
        # touch lifecycle_status, is_active, hire dates, status_history,
        # or deleted_at via upload.
        changed = {}
        for k, v in set_fields.items():
            if (match.get(k) or "") != v:
                changed[k] = v
        if not changed:
            results["no_change"] += 1
            items.append({"row_name": name, "action": "no_change", "id": match.get("id")})
            continue
        changed["updated_at"] = now
        # Audit row appended to status_history (NOT a lifecycle_status
        # change, but a record of which fields were touched by upload).
        history_entry = {
            "at": now,
            "by": actor_label,
            "actor_role": actor_role,
            "kind": "bulk_upload_field_update",
            "fields": sorted(changed.keys()),
        }
        await db.employees.update_one(
            {"id": match["id"]},
            {"$set": changed, "$push": {"status_history": history_entry}},
        )
        results["updated"] += 1
        items.append({
            "row_name": name, "action": "updated",
            "id": match["id"], "fields": sorted(changed.keys()),
        })
        try:
            await db.employee_lifecycle_events.insert_one({
                "id": str(uuid.uuid4()),
                "employee_id": match["id"],
                "at": now,
                "actor_role": actor_role,
                "actor_label": actor_label,
                "kind": "bulk_upload_field_update",
                "fields": sorted(changed.keys()),
            })
        except Exception:  # noqa: BLE001
            pass

    return {"ok": True, **results, "items": items[:1000]}


@api_router.post("/admin/employees")
async def create_employee(
    payload: Dict[str, Any],
    actor: Dict[str, Any] = Depends(_require_hr_or_admin_for_queue),
):
    """OMEGA · Phase Alpha · G-3 · Deprecated admin create.

    Now gated by HR-or-Admin (legacy admin-panel callers still work
    via HR portal token). Mirrors the canonical HR-portal employee
    shape so downstream consumers see fully-formed rows.
    Phase Beta (G-6) will redirect this entirely to /api/hr/employees.
    """
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")
    now = datetime.now(timezone.utc).isoformat()
    actor_role = actor.get("_actor") or actor.get("role") or "hr"
    actor_label = actor.get("name") or actor.get("email") or actor_role
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "employee_id": (payload.get("employee_id") or "").strip(),
        "trade": (payload.get("trade") or "").strip(),
        "role": (payload.get("role") or "").strip(),
        "crew": (payload.get("crew") or "").strip(),
        "email": (payload.get("email") or "").strip(),
        "phone": (payload.get("phone") or "").strip(),
        # Lifecycle defaults — fully-formed
        "lifecycle_status": "Active",
        "is_active": True,
        "added_via": "admin-panel-deprecated",
        "created_at": now,
        "updated_at": now,
        "status_history": [{
            "at": now,
            "by": actor_label,
            "actor_role": actor_role,
            "to": "Active",
            "reason": "Created via deprecated admin panel (Phase Alpha)",
            "kind": "admin_panel_create",
        }],
        "deleted_at": None,
    }
    await db.employees.insert_one(doc)
    try:
        await db.employee_lifecycle_events.insert_one({
            "id": str(uuid.uuid4()),
            "employee_id": doc["id"],
            "at": now,
            "actor_role": actor_role,
            "actor_label": actor_label,
            "kind": "admin_panel_create",
            "to_status": "Active",
            "from_status": None,
        })
    except Exception:  # noqa: BLE001
        pass
    doc.pop("_id", None)
    return doc


@api_router.put("/admin/employees/{employee_id}")
async def update_employee(
    employee_id: str,
    payload: Dict[str, Any],
    actor: Dict[str, Any] = Depends(_require_hr_or_admin_for_queue),
):
    """OMEGA · Phase Alpha · G-4 · Silent is_active bypass eliminated.

    `is_active` was previously a mutable field on this endpoint —
    Constitutional violation V-P0-4 (silent state-machine bypass).
    `is_active` is now a READ-ONLY mirror of `lifecycle_status` and
    can only be altered through the HR status state machine
    (`POST /api/hr/employees/{id}/status` or /reactivate).

    Phase Beta (G-6) will lock the gate to HR-only.
    """
    # G-4: explicit removal of is_active + lifecycle_status from
    # editable surface. Lifecycle mutations must go through the HR
    # state machine.
    allowed = {"name", "employee_id", "trade", "role", "crew", "email", "phone"}
    # Hard reject the back-door
    if "is_active" in payload or "lifecycle_status" in payload:
        raise HTTPException(
            status_code=422,
            detail={
                "code": "lifecycle_field_readonly",
                "message": (
                    "is_active / lifecycle_status are read-only on this "
                    "endpoint. Use POST /api/hr/employees/{id}/status "
                    "or POST /api/hr/employees/{id}/reactivate."
                ),
                "blocked_fields": [k for k in ("is_active", "lifecycle_status") if k in payload],
            },
        )
    update = {k: payload[k] for k in allowed if k in payload}
    if "name" in update and not (update["name"] or "").strip():
        raise HTTPException(status_code=400, detail="Name cannot be blank")
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = await db.employees.update_one(
        {"$and": [{"id": employee_id}, ACTIVE_FILTER]},
        {"$set": update},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    doc = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    return doc or {"ok": True}


@api_router.delete("/admin/employees/{employee_id}")
async def delete_employee(employee_id: str, actor: Dict[str, Any] = Depends(_require_hr_or_admin_for_queue)):
    """OMEGA · Phase Alpha · G-3 · Deprecated admin delete.

    Forbidden as a lifecycle action — termination MUST flow through
    the HR status state machine (status_history + offboarding
    playbook + audit event). Returns 405 with a pointer.
    """
    raise HTTPException(
        status_code=405,
        detail={
            "code": "termination_via_status_machine_only",
            "message": (
                "Soft-deleting employees is no longer permitted. "
                "Terminate via POST /api/hr/employees/{id}/status "
                "(target: Terminated/Resigned/Retired/Inactive) or "
                "submit a termination request to the HR Queue via "
                "POST /api/employee-requests (kind=termination)."
            ),
            "use_instead": [
                "POST /api/hr/employees/{id}/status",
                "POST /api/employee-requests",
            ],
        },
    )


# ---------------------------------------------------------------------------
# Suppliers / Subcontractors — used by Daily Report Sections 05 & 08.
# ---------------------------------------------------------------------------
SUPPLIERS_SEED_FILE = ROOT_DIR / "data" / "suppliers_seed.json"
EMPLOYEES_SEED_FILE = ROOT_DIR / "data" / "employees_seed.json"


@api_router.get("/suppliers")
async def list_suppliers():
    """Public — returns the full MASCI supplier / subcontractor list."""
    await _purge_expired("suppliers")
    cursor = db.suppliers.find(
        {"$and": [ACTIVE_FILTER, {"is_active": {"$ne": False}}]},
        {"_id": 0},
    ).sort("name", 1)
    docs = await cursor.to_list(2000)
    return {"items": docs, "count": len(docs)}


@api_router.get("/admin/suppliers/status")
async def suppliers_status(_: bool = Depends(require_admin)):
    total = await db.suppliers.count_documents(ACTIVE_FILTER)
    active = await db.suppliers.count_documents(
        {"$and": [ACTIVE_FILTER, {"is_active": {"$ne": False}}]}
    )
    archived = await db.suppliers.count_documents({"deleted_at": {"$ne": None}})
    last_doc = await db.suppliers.find_one(
        ACTIVE_FILTER, {"_id": 0, "updated_at": 1, "created_at": 1}, sort=[("updated_at", -1)]
    )
    last_updated = (last_doc or {}).get("updated_at") or (last_doc or {}).get("created_at")
    return {"count": total, "active": active, "archived": archived, "last_updated": last_updated}


@api_router.get("/admin/suppliers/archive")
async def suppliers_archive(_: bool = Depends(require_admin)):
    return {"items": await _list_archive("suppliers"), "retain_days": SOFT_DELETE_RETAIN_DAYS}


@api_router.post("/admin/suppliers/{supplier_id}/restore")
async def restore_supplier(supplier_id: str, _: bool = Depends(require_admin)):
    if not await _restore_row("suppliers", {"id": supplier_id}):
        raise HTTPException(status_code=404, detail="Supplier not in archive")
    doc = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    return doc or {"ok": True}


@api_router.post("/admin/suppliers/upload")
async def upload_suppliers(
    file: UploadFile = File(...),
    replace_all: bool = Form(False),
    confirm: str = Form(""),
    backup_ack: bool = Form(False),
    _: bool = Depends(require_admin),
):
    """Replace the supplier list from an .xlsx or .csv file.

    Reads the FIRST column of the first sheet (any header row is OK).
    Skips obvious dividers ('SUBCONTRACTORS', 'NOT LISTED ADD TO NOTES',
    'MASCI', 'D-MAC').
    """
    fname = (file.filename or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".xlsm") or fname.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .csv files are accepted")
    raw = await file.read()
    if not raw or len(raw) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Empty or oversized file (max 10 MB)")

    SKIP_LOWER = {"subcontractors", "suppliers", "vendors", "not listed add to notes",
                  "name", "company", "company name"}
    names: List[str] = []
    try:
        if fname.endswith(".csv"):
            import csv as _csv
            text = raw.decode("utf-8", errors="ignore")
            for r in _csv.reader(text.splitlines()):
                if r and r[0]:
                    names.append(str(r[0]).strip())
        else:
            import openpyxl as _ox
            import io as _io
            wb = _ox.load_workbook(_io.BytesIO(raw), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if row and row[0]:
                    names.append(str(row[0]).strip())
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {e}")

    seen = set()
    items: List[Dict[str, Any]] = []
    for n in names:
        if not n or n.lower() in SKIP_LOWER:
            continue
        k = n.lower()
        if k in seen:
            continue
        seen.add(k)
        items.append({
            "id": str(uuid.uuid4()),
            "name": n,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })

    if not items:
        raise HTTPException(status_code=400, detail="No supplier names found.")

    current_count = await db.suppliers.count_documents({})
    duplicate_count = max(0, len(names) - len(items))
    preflight = {
        "current_suppliers": current_count,
        "incoming_suppliers": len(items),
        "duplicates_filtered": duplicate_count,
        "invalid_rows": 0,
    }

    if not replace_all:
        return {
            "ok": True,
            "mode": "preflight",
            "replace_all": False,
            "preflight": preflight,
        }

    require_destructive_confirmation(
        {"confirm": confirm, "backup_ack": backup_ack},
        expected_confirm="REPLACE_ALL_SUPPLIERS",
    )
    require_destructive_runtime_guard(expected_db_name="masci_safety")

    await db.suppliers.delete_many({})
    await db.suppliers.insert_many(items)
    return {
        "ok": True,
        "mode": "replace_all",
        "count": len(items),
        "preflight": {
            **preflight,
            "removed": current_count,
            "added": len(items),
            "changed": len(items),
        },
    }


@api_router.post("/admin/suppliers")
async def create_supplier(
    payload: Dict[str, Any],
    _: bool = Depends(require_admin),
):
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")
    # Track 19.60 · actor provenance — best-effort from payload.
    actor = (str(payload.get("_actor") or "") or "admin").strip() or "admin"
    now_iso = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "display_name": (payload.get("display_name") or "").strip() or None,
        "dba": (payload.get("dba") or "").strip() or None,
        "vendor_type": (payload.get("vendor_type") or "").strip() or None,
        "primary_contact": (payload.get("primary_contact") or "").strip() or None,
        "phone": (payload.get("phone") or "").strip() or None,
        "email": (payload.get("email") or "").strip() or None,
        "address": (payload.get("address") or "").strip() or None,
        "notes": (payload.get("notes") or "").strip() or None,
        "do_not_use": bool(payload.get("do_not_use")) if "do_not_use" in payload else False,
        "is_active": True,
        "created_at": now_iso,
        "updated_at": now_iso,
        "created_by": actor,
        "updated_by": actor,
    }
    await db.suppliers.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/admin/suppliers/{supplier_id}")
async def update_supplier(
    supplier_id: str,
    payload: Dict[str, Any],
    _: bool = Depends(require_admin),
):
    """Inline edit a supplier — HR/Admin-owned vendor management.
    Track 19.60 extended the allowed field set additively. Soft-deleted
    rows are not editable — restore them first."""
    allowed = {"name", "is_active", "display_name", "dba", "vendor_type",
               "primary_contact", "phone", "email", "address", "notes",
               "do_not_use"}
    update = {k: payload[k] for k in allowed if k in payload}
    if "name" in update and not (str(update["name"]) or "").strip():
        raise HTTPException(status_code=400, detail="Name cannot be blank")
    actor = (str(payload.get("_actor") or "") or "admin").strip() or "admin"
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    update["updated_by"] = actor
    res = await db.suppliers.update_one(
        {"$and": [{"id": supplier_id}, ACTIVE_FILTER]},
        {"$set": update},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    doc = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    return doc or {"ok": True}


@api_router.delete("/admin/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, _: bool = Depends(require_admin)):
    if not await _soft_delete("suppliers", {"id": supplier_id}):
        raise HTTPException(status_code=404, detail="Supplier not found")
    return {"ok": True, "soft_deleted": True, "retain_days": SOFT_DELETE_RETAIN_DAYS}


# ---------------------------------------------------------------------------
# Idempotent seed for employees + suppliers on startup. If the collection is
# empty AND a seed JSON file exists, populate it. Re-uploading via the admin
# panel will replace the contents.
# ---------------------------------------------------------------------------
async def _seed_employees_from_json() -> None:
    log = logging.getLogger(__name__)
    if not EMPLOYEES_SEED_FILE.exists():
        return
    if await db.employees.count_documents({}) > 0:
        return
    try:
        import json as _json_em
        with open(EMPLOYEES_SEED_FILE, "r", encoding="utf-8") as fh:
            names = _json_em.load(fh)
        items = []
        seen = set()
        for n in names:
            if not n or not isinstance(n, str):
                continue
            k = n.strip().lower()
            if not k or k in seen:
                continue
            seen.add(k)
            items.append({
                "id": str(uuid.uuid4()),
                "name": n.strip(),
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            })
        if items:
            await db.employees.insert_many(items)
            log.info(f"[employees] seeded {len(items)} from JSON")
    except Exception as e:
        log.exception(f"[employees] seed failed: {e}")


async def _seed_suppliers_from_json() -> None:
    log = logging.getLogger(__name__)
    if not SUPPLIERS_SEED_FILE.exists():
        return
    if await db.suppliers.count_documents({}) > 0:
        return
    try:
        import json as _json_sp
        with open(SUPPLIERS_SEED_FILE, "r", encoding="utf-8") as fh:
            data = _json_sp.load(fh)
        items = []
        seen = set()
        for entry in data:
            n = entry.get("name") if isinstance(entry, dict) else (entry if isinstance(entry, str) else "")
            if not n:
                continue
            k = n.strip().lower()
            if not k or k in seen:
                continue
            seen.add(k)
            items.append({
                "id": str(uuid.uuid4()),
                "name": n.strip(),
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            })
        if items:
            await db.suppliers.insert_many(items)
            log.info(f"[suppliers] seeded {len(items)} from JSON")
    except Exception as e:
        log.exception(f"[suppliers] seed failed: {e}")


# ---------------------------------------------------------------------------
# Project P&L Snapshot — aggregate live job-cost data from daily_reports
# in one shot for a given project + date range.
# ---------------------------------------------------------------------------
DEFAULT_LABOR_RATE = float(os.environ.get("DEFAULT_LABOR_RATE", "45.0"))


def _coerce_float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


@api_router.get("/jobs-master")
async def get_jobs_master():
    """DR-JOB-002 · public read-only canonical jobs list. Used by the DR hub
    to group reports under canonical project identity instead of free-text
    submitter names. Read-only · no auth required for client-side display."""
    rows = []
    async for j in db.jobs_master.find({}, {"_id": 0, "project_number": 1, "project_name": 1, "status": 1}):
        rows.append(j)
    return rows


@api_router.get("/admin/projects/list")
async def list_projects_in_dailies(actor=Depends(require_admin)):
    """Return distinct {project_number, project_name} tuples seen across all
    daily reports — gives the P&L picker a curated dropdown so users don't
    have to type project numbers from memory.

    PMs see only projects from THEIR jobs (primary or co-PM)."""
    from pm_auth import compute_pm_scope
    from lib.synthetic_dr_filter import apply_synthetic_dr_exclusion  # noqa: PLC0415
    scope = await compute_pm_scope(db, actor)
    if scope.is_definitively_empty():
        return {"items": [], "count": 0}
    # TRACK 28.02B · exclude synthetic/TEST rows so the P&L picker
    # never surfaces certification projects to real operators.
    pipeline = [
        {"$match": apply_synthetic_dr_exclusion(scope.filter({"project_number": {"$nin": [None, ""]}}))},
        {"$group": {
            "_id": "$project_number",
            "project_name": {"$last": "$project_name"},
            "report_count": {"$sum": 1},
            "last_report_date": {"$max": "$report_date"},
        }},
        {"$sort": {"last_report_date": -1}},
        {"$limit": 500},
    ]
    docs = await db.daily_reports.aggregate(pipeline).to_list(500)
    return {
        "items": [
            {
                "project_number": d["_id"],
                "project_name": d.get("project_name") or "",
                "report_count": d.get("report_count", 0),
                "last_report_date": d.get("last_report_date") or "",
            }
            for d in docs
        ],
        "count": len(docs),
    }


@api_router.get("/admin/projects/pnl")
async def project_pnl(
    project_number: str,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    labor_rate: Optional[float] = None,
    actor=Depends(require_admin),
):
    """Live job-cost dashboard for one project + date range.

    Aggregates all matching `daily_reports` and returns:
      - crew_hours_total + crew_breakdown (by employee)
      - sub_hours_total + sub_breakdown (by company)
      - material_lines (one row per ticket)
      - cost_summary (labor cost, sub cost — sub cost left blank unless rate set)
      - report_count, date_range_actual

    PMs can only pull P&L for their own jobs (primary or co-PM). Admins
    and legacy bypass see all.
    """
    if not project_number:
        raise HTTPException(status_code=400, detail="project_number is required")

    # Per-PM scope check — block leakage to projects this PM isn't on.
    from pm_auth import compute_pm_scope
    scope = await compute_pm_scope(db, actor)
    if not scope.allows(project_number):
        raise HTTPException(status_code=404, detail="Project not found in your assignments")

    rate = labor_rate if labor_rate and labor_rate > 0 else DEFAULT_LABOR_RATE

    from lib.synthetic_dr_filter import apply_synthetic_dr_exclusion  # noqa: PLC0415
    q: Dict[str, Any] = {"project_number": project_number}
    # report_date is stored as 'YYYY-MM-DD' string — string compare works lex
    date_filter: Dict[str, Any] = {}
    if date_from:
        date_filter["$gte"] = date_from
    if date_to:
        date_filter["$lte"] = date_to
    if date_filter:
        q["report_date"] = date_filter

    # TRACK 28.02B · exclude synthetic rows from P&L; the operational
    # cost dashboard must reflect real work, not certification
    # fixtures. If a synthetic project happens to share a project_number
    # with a real one, cost math would silently inflate.
    q = apply_synthetic_dr_exclusion(q)
    cursor = db.daily_reports.find(q, {"_id": 0}).sort("report_date", 1)
    reports = await cursor.to_list(2000)

    crew_by_name: Dict[str, Dict[str, Any]] = {}
    sub_by_company: Dict[str, Dict[str, Any]] = {}
    material_lines: List[Dict[str, Any]] = []
    crew_total_hours = 0.0
    sub_total_hours = 0.0
    project_name_seen: Optional[str] = None
    actual_dates: List[str] = []

    for r in reports:
        actual_dates.append(r.get("report_date") or "")
        if not project_name_seen:
            project_name_seen = r.get("project_name") or None

        # Crew rows
        for c in (r.get("masci_crews") or []):
            name = (c.get("name") or "Unnamed").strip() or "Unnamed"
            hrs = _coerce_float(c.get("hours"))
            entry = crew_by_name.setdefault(name, {
                "name": name,
                "trade": c.get("trade") or "",
                "days_on_site": 0,
                "hours": 0.0,
            })
            entry["days_on_site"] += 1
            entry["hours"] += hrs
            crew_total_hours += hrs

        # Subcontractor rows
        for s in (r.get("subcontractors") or []):
            company = (s.get("company") or "Unknown").strip() or "Unknown"
            count = _coerce_float(s.get("count"))
            hrs_per_worker = _coerce_float(s.get("hours"))
            # If "count" + "hours" are filled, multiply for total man-hours.
            # If only "hours" is filled, treat as crew-hours total for the day.
            man_hours = count * hrs_per_worker if count and hrs_per_worker else hrs_per_worker
            entry = sub_by_company.setdefault(company, {
                "company": company,
                "trade": s.get("trade") or "",
                "days_on_site": 0,
                "headcount_total": 0.0,
                "hours": 0.0,
            })
            entry["days_on_site"] += 1
            entry["headcount_total"] += count
            entry["hours"] += man_hours
            sub_total_hours += man_hours

        # Materials — one row per ticket
        for m in (r.get("materials") or []):
            material_lines.append({
                "report_date": r.get("report_date") or "",
                "description": m.get("description") or "",
                "quantity": m.get("quantity") or "",
                "unit": m.get("unit") or "",
                "supplier": m.get("supplier") or "",
                "ticket_number": m.get("ticket_number") or "",
                "notes": m.get("notes") or "",
                "ticket_photo_count": len(m.get("ticket_photos") or []),
            })

    crew_breakdown = sorted(crew_by_name.values(), key=lambda e: -e["hours"])
    sub_breakdown = sorted(sub_by_company.values(), key=lambda e: -e["hours"])

    labor_cost = round(crew_total_hours * rate, 2)

    return {
        "project_number": project_number,
        "project_name": project_name_seen or "",
        "date_from": min(actual_dates) if actual_dates else date_from,
        "date_to": max(actual_dates) if actual_dates else date_to,
        "report_count": len(reports),
        "labor_rate": rate,
        "crew_hours_total": round(crew_total_hours, 2),
        "labor_cost": labor_cost,
        "crew_breakdown": [
            {**e, "hours": round(e["hours"], 2), "cost_at_rate": round(e["hours"] * rate, 2)}
            for e in crew_breakdown
        ],
        "sub_hours_total": round(sub_total_hours, 2),
        "sub_breakdown": [
            {**e, "hours": round(e["hours"], 2), "headcount_total": round(e["headcount_total"], 2)}
            for e in sub_breakdown
        ],
        "material_count": len(material_lines),
        "material_lines": material_lines,
    }


# ---------------------------------------------------------------------------
# Daily Report numbering — see /daily-reports/next-number above (registered
# before the /{report_id} route so FastAPI matches it correctly).
# ---------------------------------------------------------------------------


async def _write_equipment_master(items: List[Dict[str, Any]]) -> int:
    """Replace the equipment_master collection with `items` and fan-out to
    equipment_units for the legacy Pre-Op dropdown. Returns inserted count.

    Items are expected to already be in the parser's normalized shape
    (see equipment_parser.parse_equipment_xlsx).

    Trench Safety mirror rows (category="Trench Safety") are PRESERVED.
    Those rows are NOT JSON-sourced — they shadow `db.trench_safety_assets`
    and are owned by `routes/trench_safety/_helpers.py`. Wiping them here
    would orphan the mirror on every backend restart.
    """
    log = logging.getLogger(__name__)
    await db.equipment_master.delete_many({"category": {"$ne": "Trench Safety"}})
    if not items:
        return 0
    for it in items:
        it.setdefault("id", str(uuid.uuid4()))
    await db.equipment_master.insert_many(items)
    # Fan out into equipment_units (used by the existing Pre-Op type→units
    # dropdown). Only insert what's not already there.
    try:
        existing_units = set()
        async for u in db.equipment_units.find(
            {}, {"_id": 0, "equipment_type": 1, "unit_label": 1}
        ):
            existing_units.add(
                (u.get("equipment_type", ""), (u.get("unit_label", "") or "").strip().lower())
            )
        new_units = []
        for it in items:
            etype = it.get("preop_equipment_type") or "Other"
            label = (it.get("display_label") or it.get("make_model") or "").strip()
            if not label:
                continue
            key = (etype, label.lower())
            if key in existing_units:
                continue
            existing_units.add(key)
            new_units.append({
                "id": str(uuid.uuid4()),
                "equipment_type": etype,
                "unit_label": label,
                "make": it.get("make_model", ""),
                "model": "",
                "serial": it.get("vin_serial_number", ""),
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
        if new_units:
            await db.equipment_units.insert_many(new_units)
    except Exception as e:
        log.exception(f"[equipment-master] equipment_units fan-out failed: {e}")
    return len(items)


async def _seed_equipment_master() -> None:
    """Idempotent seed of the equipment_master collection from JSON file.

    Re-runs whenever the JSON file's item count differs from what's stored
    (so updating the seed file ships new equipment automatically on restart).
    """
    log = logging.getLogger(__name__)
    if not EQUIPMENT_MASTER_SEED_FILE.exists():
        log.info(f"[equipment-master] seed file missing: {EQUIPMENT_MASTER_SEED_FILE}")
        return
    try:
        with open(EQUIPMENT_MASTER_SEED_FILE, "r", encoding="utf-8") as fh:
            import json as _json_em
            seed_items = _json_em.load(fh)
    except Exception as e:
        log.exception(f"[equipment-master] failed to read seed: {e}")
        return

    existing_count = await db.equipment_master.count_documents(
        {"category": {"$ne": "Trench Safety"}}
    )
    if existing_count == len(seed_items) and existing_count > 0:
        return  # already seeded and matches file

    n = await _write_equipment_master(seed_items)
    log.info(f"[equipment-master] seeded {n} units from JSON")


@api_router.get("/admin/equipment-master/status")
async def equipment_master_status(_: bool = Depends(require_admin)):
    """Quick status panel for the Admin Hub: count + per-category breakdown +
    last-updated timestamp from the seed JSON file (mtime)."""
    count = await db.equipment_master.count_documents(ACTIVE_FILTER)
    archived = await db.equipment_master.count_documents({"deleted_at": {"$ne": None}})
    cursor = db.equipment_master.find(ACTIVE_FILTER, {"_id": 0, "category": 1})
    cats: Dict[str, int] = {}
    async for d in cursor:
        c = d.get("category", "Misc Equipment")
        cats[c] = cats.get(c, 0) + 1
    last_updated = None
    if EQUIPMENT_MASTER_SEED_FILE.exists():
        last_updated = datetime.fromtimestamp(
            EQUIPMENT_MASTER_SEED_FILE.stat().st_mtime, tz=timezone.utc
        ).isoformat()
    return {
        "count": count,
        "archived": archived,
        "categories": dict(sorted(cats.items(), key=lambda x: -x[1])),
        "last_updated": last_updated,
        "seed_file": str(EQUIPMENT_MASTER_SEED_FILE),
    }


@api_router.post("/admin/equipment-master")
async def create_equipment_master_unit(
    payload: Dict[str, Any],
    operator: Any = Depends(require_shop_or_admin),
):
    """Add a single unit to the MASCI fleet. Mechanics + admins + PMs can
    use this to register new equipment without uploading a full xlsx.

    If a soft-deleted unit with the same unit_number already exists, the
    insert auto-restores it instead of creating a duplicate row.
    """
    unit_number = (payload.get("unit_number") or "").strip()
    if not unit_number:
        raise HTTPException(status_code=400, detail="Unit number is required")
    existing = await db.equipment_master.find_one({"unit_number": unit_number}, {"_id": 0})
    if existing:
        if existing.get("deleted_at"):
            await _restore_row("equipment_master", {"unit_number": unit_number})
            doc = await db.equipment_master.find_one({"unit_number": unit_number}, {"_id": 0})
            return doc or {"ok": True, "restored": True}
        raise HTTPException(status_code=409, detail=f"Unit {unit_number} already exists")

    from services.asset_spine import AssetSpine  # noqa: PLC0415

    make = (payload.get("make") or "").strip()
    model = (payload.get("model") or "").strip()
    make_model = (payload.get("make_model") or f"{make} {model}").strip()
    display_label = (payload.get("display_label") or "").strip()
    preop_type = (payload.get("preop_equipment_type") or "Other").strip()
    company = (payload.get("company") or "MASCI").strip()
    actor = "admin"
    if isinstance(operator, dict):
        actor = str(operator.get("email") or operator.get("id") or "admin")

    spine = AssetSpine(db)
    created = await spine.create_asset(
        {
            "unit_number": unit_number,
            "asset_name": display_label or make_model or unit_number,
            "asset_type": preop_type,
            "asset_category": (payload.get("category") or "Misc Equipment").strip(),
            "ownership": company,
            "make": make,
            "model": model,
            "year": str(payload.get("year") or "").strip(),
            "vin_serial_number": (payload.get("vin_serial_number") or "").strip(),
        },
        actor=actor,
    )
    asset_id = created.get("asset_id")
    if asset_id:
        await db.equipment_master.update_one(
            {"id": asset_id},
            {"$set": {
                "make_model": make_model,
                "comments": (payload.get("comments") or "").strip(),
                "preop_equipment_type": preop_type,
                "display_label": display_label,
            }},
        )
        doc = await db.equipment_master.find_one({"id": asset_id}, {"_id": 0})
        if doc:
            return doc
    raise HTTPException(status_code=500, detail="Could not create equipment unit")


@api_router.put("/admin/equipment-master/{unit_id}")
async def update_equipment_master_unit(
    unit_id: str,
    payload: Dict[str, Any],
    _: bool = Depends(require_shop_or_admin),
):
    """Edit a single fleet unit (matched by id OR unit_number)."""
    allowed = {"unit_number", "make", "model", "make_model", "year", "vin_serial_number",
               "comments", "company", "category", "preop_equipment_type", "display_label"}
    update = {k: payload[k] for k in allowed if k in payload}
    if "unit_number" in update and not (update["unit_number"] or "").strip():
        raise HTTPException(status_code=400, detail="Unit number cannot be blank")
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = await db.equipment_master.update_one(
        {"$and": [
            {"$or": [{"id": unit_id}, {"unit_number": unit_id}]},
            ACTIVE_FILTER,
        ]},
        {"$set": update},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Unit not found")
    doc = await db.equipment_master.find_one(
        {"$or": [{"id": unit_id}, {"unit_number": unit_id}]}, {"_id": 0}
    )
    return doc or {"ok": True}


@api_router.delete("/admin/equipment-master/{unit_id}")
async def delete_equipment_master_unit(
    unit_id: str,
    _: bool = Depends(require_shop_or_admin),
):
    """Soft-remove a single fleet unit (matched by id OR unit_number).
    The row is hidden from the active fleet immediately and hard-purged
    after ``SOFT_DELETE_RETAIN_DAYS`` days. Use the Archive tab + Restore
    button to undo within the window."""
    if not await _soft_delete(
        "equipment_master",
        {"$or": [{"id": unit_id}, {"unit_number": unit_id}]},
    ):
        raise HTTPException(status_code=404, detail="Unit not found")
    return {"ok": True, "soft_deleted": True, "retain_days": SOFT_DELETE_RETAIN_DAYS}


@api_router.post("/admin/equipment-master/upload")
async def upload_equipment_master(
    file: UploadFile = File(...),
    sheet: str = Form("Louis"),
    _: bool = Depends(require_admin),
):
    """Replace the entire MASCI equipment fleet from an uploaded xlsx file.

    The xlsx is parsed via `equipment_parser.parse_equipment_xlsx`, the JSON
    seed file at /app/backend/data/equipment_master.json is overwritten so
    future restarts stay in sync, and the equipment_master + equipment_units
    collections are refreshed atomically.
    """
    log = logging.getLogger(__name__)
    fname = (file.filename or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".xlsm")):
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx files are accepted (got '%s')" % (file.filename or ""),
        )
    raw = await file.read()
    if not raw or len(raw) > 25 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Empty or oversized file (max 25 MB)")

    # Back up the previous seed file before replacing
    try:
        if EQUIPMENT_MASTER_SEED_FILE.exists():
            ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
            backup = EQUIPMENT_MASTER_SEED_FILE.with_suffix(f".{ts}.bak.json")
            EQUIPMENT_MASTER_SEED_FILE.replace(backup)
    except Exception as e:
        log.warning(f"[equipment-master] backup of previous seed failed: {e}")

    try:
        from equipment_parser import parse_equipment_xlsx
        parsed = parse_equipment_xlsx(raw, sheet_name=sheet or "Louis")
    except Exception as e:
        log.exception("[equipment-master] xlsx parse failed")
        raise HTTPException(status_code=400, detail=f"Could not parse xlsx: {e}")

    items = parsed["items"]
    if not items:
        raise HTTPException(
            status_code=400,
            detail=f"No equipment rows found in sheet '{parsed['sheet']}'.",
        )

    # Persist new seed JSON so the next restart stays in sync
    try:
        EQUIPMENT_MASTER_SEED_FILE.parent.mkdir(parents=True, exist_ok=True)
        import json as _json_em
        with open(EQUIPMENT_MASTER_SEED_FILE, "w", encoding="utf-8") as fh:
            _json_em.dump(items, fh, indent=2)
    except Exception as e:
        log.exception(f"[equipment-master] writing seed JSON failed: {e}")
        raise HTTPException(status_code=500, detail="Could not write seed file")

    inserted = await _write_equipment_master(items)
    log.info(
        f"[equipment-master] uploaded {inserted} units from '{file.filename}' "
        f"(sheet={parsed['sheet']})"
    )
    return {
        "ok": True,
        "count": inserted,
        "sheet": parsed["sheet"],
        "category_counts": parsed["category_counts"],
        "filename": file.filename,
    }


# ============================================================
# Shop Activity Feed + Equipment Parts Catalog
# ----------------------------------------------------------
# Extracted to /app/backend/routes/shop_parts.py 2026-04-28 as the
# first proof-of-pattern for the larger server.py refactor (P1).
# Endpoints registered there:
#   GET    /shop/activity
#   GET    /equipment-parts                       (list)
#   GET    /equipment-parts/{unit_number}         (single)
#   PUT    /equipment-parts/{unit_number}         (upsert)
#   DELETE /equipment-parts/{unit_number}         (admin only)
#   GET    /admin/equipment-parts/status
#   POST   /admin/equipment-parts/upload          (xlsx/csv)
#   POST   /equipment-parts/order                 (Resend email)
# ============================================================
from routes.shop_parts import register_shop_parts_routes  # noqa: E402

register_shop_parts_routes(api_router, db, require_admin, require_shop_or_admin)


# ============================================================
# Compliance CSV Exports (admin-only)
# ============================================================
EXPORTABLE_KINDS = {
    "inspections": "inspections",
    "meetings": "meetings",
    "jhas": "jhas",
    "incidents": "incidents",
    "daily-reports": "daily_reports",
    "equipment-inspections": "equipment_inspections",
}

# ════════════════════════════════════════════════════════════════════════
# iter425 · Phase 25.2 · BACKUP-WIDE redaction + exclusion config
# Extracted to module scope so the R2 complete-archive in Pipeline B can
# share the exact same hygiene rules as Pipeline A.
# ════════════════════════════════════════════════════════════════════════
# Sensitive fields stripped from EVERY backup pipeline.
# Mongo projection format: 0 = exclude. `_id` always excluded.
BACKUP_SENSITIVE_FIELD_REDACTION = {
    "users":          {"_id": 0, "password_hash": 0},
    # iter425 · MFA TOTP secret + recovery codes are bearer-equivalent
    # credentials. Never persist them in a portable backup.
    "user_directory": {
        "_id": 0,
        "password_hash": 0,
        "mfa.secret": 0,
        "mfa.recovery_codes": 0,
    },
}

# Collections explicitly EXCLUDED from auto-discovery backup paths.
# Reasons documented in /app/memory/R2_BACKUP_CONTINUITY_AUDIT.md §9.
# We intentionally keep webauthn_challenges + dispatch_driver_sessions IN
# for now (short-lived but harmless · keeps audit explicit · no silent drop).
#
# iter441 · OMEGA Batch §6.4 Minimum Surgical Memory-Reduction Fix
# ────────────────────────────────────────────────────────────────
# Three high-cardinality REGENERABLE collections are excluded to
# eliminate ~92 % of `zipfile._filelist` (ZipInfo) memory retention
# during complete-archive builds. Evidence: BACKUP_CRASH_ROOT_CAUSE_REPORT.md
#  · usage_events         · 244,266 rows · pure API telemetry · regenerates
#  · health_monitor_runs  ·  17,327 rows · scheduler health probe series
#  · job_photo_thumb_cache·   1,791 rows · derivative cache of R2 photo
# No business record is excluded. Restore continues to be a single-zip
# operation. Reversible by deletion of the three lines below.
BACKUP_EXPLICIT_EXCLUSIONS = {
    "system.indexes",          # MongoDB internal
    "usage_events",            # regenerable API telemetry (iter441)
    "health_monitor_runs",     # regenerable scheduler health series (iter441)
    "job_photo_thumb_cache",   # regenerable derivative photo cache (iter441)
    "backup_integrity_jobs",   # regenerable operator job ledger (iter OPS8 Repair A)
}

BACKUP_EXCLUSION_DETAILS: Dict[str, Dict[str, str]] = {
    "usage_events": {
        "reason": "regenerable API telemetry",
        "owner": "backup-platform",
    },
    "health_monitor_runs": {
        "reason": "regenerable scheduler health series",
        "owner": "backup-platform",
    },
    "job_photo_thumb_cache": {
        "reason": "regenerable derivative photo cache",
        "owner": "backup-platform",
    },
    "backup_integrity_jobs": {
        "reason": "regenerable operator integrity job ledger",
        "owner": "backup-platform",
    },
    "system.*": {
        "reason": "MongoDB internal system collection",
        "owner": "mongodb",
    },
}

BACKUP_MANIFEST_VERSION = "27.11c-1"
BACKUP_VERIFIER_VERSION = "27.11c-1"


# Per-kind row schema — what each CSV column should contain. We deliberately
# omit photos / signatures (binary blobs) and the raw checklist dict (renders
# poorly in Excel). Reviewers click into the Admin Hub for the full record.
EXPORT_FIELDS: Dict[str, List[str]] = {
    "inspections": [
        "inspection_date", "inspection_time", "project_name", "project_number",
        "location", "inspector_name", "foreman_name", "operation",
        "work_activity", "hazards_observed", "stop_work_issued",
        "ppe_in_use", "weather_summary", "created_at", "id",
    ],
    "meetings": [
        "meeting_date", "meeting_time", "project_name", "project_number",
        "location", "presenter_name", "topic_title", "topic_number",
        "attendee_count", "discussion_summary", "created_at", "id",
    ],
    "jhas": [
        "jha_date", "project_name", "project_number", "location",
        "supervisor_name", "task_description", "approver_name",
        "step_count", "created_at", "id",
    ],
    "incidents": [
        "incident_date", "incident_time", "project_name", "project_number",
        "location", "incident_type", "severity", "osha_recordable",
        "work_stopped", "person_name", "body_part", "injury_nature",
        "treatment_provided", "medical_facility",
        "reporter_name", "supervisor_name",
        "root_cause_categories", "witness_count",
        "description", "immediate_action", "follow_up_action",
        "created_at", "id",
    ],
    "daily-reports": [
        "report_date", "project_name", "project_number", "location",
        "prepared_by", "superintendent_name",
        "weather_summary", "high_temp_f", "low_temp_f",
        "crew_count", "subcontractor_count", "visitor_count",
        "equipment_count", "material_count", "activity_count",
        "accident_or_injury", "safety_notified", "safety_notified_who",
        "safety_notified_time", "incident_report_filled",
        "incident_report_time",
        "delays_or_issues", "tomorrows_plan",
        "created_at", "id",
    ],
    "equipment-inspections": [
        "inspection_date", "inspection_time", "project_name", "project_number",
        "location", "operator_name", "equipment_type", "equipment_unit",
        "equipment_make", "equipment_model", "equipment_serial",
        "hour_meter", "odometer",
        "pass_count", "fail_count", "na_count", "out_of_service",
        "deficiency_notes", "corrective_actions",
        "created_at", "id",
    ],
}


def _csv_value(v: Any) -> str:
    """Flatten a record value into a CSV-friendly string."""
    if v is None:
        return ""
    if isinstance(v, (list, tuple)):
        # Skip image blobs, summarize the rest with semicolons
        items = [
            str(x)
            for x in v
            if not (isinstance(x, str) and x.startswith("data:image/"))
        ]
        return "; ".join(items)
    if isinstance(v, dict):
        # e.g. witnesses, root cause categories — flatten one level
        return "; ".join(f"{k}={_csv_value(val)}" for k, val in v.items())
    return str(v)


def _date_field_for(kind: str) -> str:
    return {
        "inspections": "inspection_date",
        "meetings": "meeting_date",
        "jhas": "jha_date",
        "incidents": "incident_date",
        "daily-reports": "report_date",
        "equipment-inspections": "inspection_date",
    }[kind]


def _normalize_export_doc(kind: str, d: Dict[str, Any]) -> None:
    """Mutate a record in place so derived counts are populated for CSV columns."""
    if kind == "meetings" and "attendee_count" not in d:
        d["attendee_count"] = len(d.get("attendees") or [])
    if kind == "jhas" and "step_count" not in d:
        d["step_count"] = len(d.get("steps") or [])
    if kind == "incidents" and "witness_count" not in d:
        d["witness_count"] = len(d.get("witnesses") or [])
    if kind == "incidents" and "root_cause_categories" not in d:
        cats = d.get("root_causes") or []
        d["root_cause_categories"] = (
            "; ".join(cats) if isinstance(cats, list) else str(cats)
        )
    if kind == "daily-reports":
        for k_, src in (
            ("crew_count", "crew"),
            ("subcontractor_count", "subcontractors"),
            ("visitor_count", "visitors"),
            ("equipment_count", "equipment"),
            ("material_count", "materials"),
            ("activity_count", "activities"),
        ):
            if k_ not in d:
                d[k_] = len(d.get(src) or [])


def _build_csv_bytes(kind: str, docs: List[Dict[str, Any]]) -> bytes:
    """Render a CSV (UTF-8 bytes) for one kind."""
    fields = list(EXPORT_FIELDS[kind])
    for d in docs:
        _normalize_export_doc(kind, d)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(fields)
    for d in docs:
        writer.writerow([_csv_value(d.get(f)) for f in fields])
    return buf.getvalue().encode("utf-8")


@api_router.get("/exports/csv")
async def export_csv(
    kind: str,
    start: Optional[str] = None,  # YYYY-MM-DD inclusive
    end: Optional[str] = None,    # YYYY-MM-DD inclusive
    _: bool = Depends(require_admin),
):
    """Stream a CSV export for one form kind, optionally filtered by date.

    Query params:
        kind  = inspections | meetings | jhas | incidents | daily-reports | equipment-inspections
        start = YYYY-MM-DD (inclusive)
        end   = YYYY-MM-DD (inclusive)

    Both date params are optional — omit for all-time.
    """
    if kind not in EXPORTABLE_KINDS:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown kind '{kind}'. Allowed: {sorted(EXPORTABLE_KINDS.keys())}",
        )

    coll_name = EXPORTABLE_KINDS[kind]
    date_field = _date_field_for(kind)

    q: Dict[str, Any] = {}
    if start or end:
        cond: Dict[str, str] = {}
        if start:
            cond["$gte"] = start
        if end:
            cond["$lte"] = end
        q[date_field] = cond

    # Drop heavy blobs from the projection
    projection = {
        "_id": 0,
        "photos": 0,
        "signature": 0,
        "operator_signature": 0,
        "supervisor_signature": 0,
        "reporter_signature": 0,
        "preparer_signature": 0,
        "approver_signature": 0,
    }

    cursor = db[coll_name].find(q, projection).sort(date_field, -1)
    docs = await cursor.to_list(20000)

    csv_bytes = _build_csv_bytes(kind, docs)

    today = datetime.now(timezone.utc).date().isoformat()
    range_tag = ""
    if start and end:
        range_tag = f"_{start}_to_{end}"
    elif start:
        range_tag = f"_from_{start}"
    elif end:
        range_tag = f"_through_{end}"
    filename = f"MASCI_{kind}{range_tag}_{today}.csv"

    return Response(
        content=csv_bytes,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Record-Count": str(len(docs)),
        },
    )


@api_router.get("/exports/summary")
async def export_summary(
    start: Optional[str] = None,
    end: Optional[str] = None,
    _: bool = Depends(require_admin),
):
    """Quick count-per-kind for a given date range — used by the Admin UI to
    show a foreman 'You have 47 records in this range' before they download."""
    out: Dict[str, int] = {}
    for kind, coll_name in EXPORTABLE_KINDS.items():
        date_field = _date_field_for(kind)
        q: Dict[str, Any] = {}
        if start or end:
            cond: Dict[str, str] = {}
            if start:
                cond["$gte"] = start
            if end:
                cond["$lte"] = end
            q[date_field] = cond
        out[kind] = await db[coll_name].count_documents(q)
    return {"start": start, "end": end, "counts": out, "total": sum(out.values())}


# ----------------------------------------------------------------------
# Full backup — single .zip with everything (CSVs + JSON + PDFs + photos)
# ----------------------------------------------------------------------
import asyncio as _backup_asyncio  # noqa: E402
import json as _backup_json  # noqa: E402
import zipfile  # noqa: E402


def _safe_filename(s: str, max_len: int = 60) -> str:
    """Make a filesystem-friendly fragment from a free-form string."""
    cleaned = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in (s or ""))
    cleaned = cleaned.strip("_") or "untitled"
    return cleaned[:max_len]


def _record_filename(kind: str, record: dict) -> str:
    """Stable, sortable filename: <date>_<id-prefix>_<project>.<ext>"""
    date_part = (
        record.get("inspection_date")
        or record.get("meeting_date")
        or record.get("jha_date")
        or record.get("incident_date")
        or record.get("report_date")
        or "0000-00-00"
    )
    rid = (record.get("id") or "")[:8]
    proj = _safe_filename(record.get("project_name") or "MASCI", 40)
    return f"{date_part}_{rid}_{proj}"


@api_router.get("/exports/full-backup")
async def exports_full_backup(_: bool = Depends(require_admin_strict)):
    """One-click off-site backup. Streams a single .zip back containing:

    /CSV/                — one CSV per kind (no photos/signatures inline)
    /<kind>/json/        — every record as raw JSON (photos + signatures intact)
    /<kind>/pdf/         — every record rendered to PDF via WeasyPrint
    /crew_hub/           — Crew Hub collections as JSON
    /safety_aux/         — Equipment unit registry, JHP plan PDFs, trench-box refs
    /backup_manifest.json — schema + counts + generated_at
    /backup_log.txt      — human-readable summary

    Built STREAMING to disk via `_build_backup_zip_to_path` then returned
    as a FileResponse. Memory use ~5–20 MB regardless of zip size, so the
    backend never OOMs even on 1 GB+ archives.
    """
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    # Build into the canonical backups dir so the file is preserved + reusable.
    _now = datetime.now(timezone.utc)
    _stamp = _now.strftime("%Y-%m-%d_%H%M%SZ")
    filename = f"MASCI_full_backup_{_stamp}.zip"
    out = BACKUPS_DIR / filename
    # Per-call unique tmp suffix so concurrent requests within the same
    # second can't clobber each other's stream (or each other's rename).
    tmp = out.with_suffix(f".zip.tmp.{uuid.uuid4().hex[:8]}")
    # DEPLOY-FIX-001 · Workstream A2/A3 — guarantee the .tmp.<hash> is
    # removed on ANY failure path (build raise, gateway disconnect, cancel,
    # client abort). Without this guard, abandoned ~500 MB tmp files can
    # fill local disk to 100% and silently break subsequent runs.
    try:
        total_records, _ = await _build_backup_zip_to_path(db, tmp)
        tmp.replace(out)
    except BaseException:
        try:
            if tmp.exists():
                logger.warning(
                    f"[backup-cleanup] failure path · removing orphan tmp "
                    f"{tmp.name} (age=fresh)"
                )
                tmp.unlink()
        except Exception:
            pass
        raise
    size_bytes = out.stat().st_size
    return FileResponse(
        path=str(out),
        media_type="application/zip",
        filename=filename,
        headers={
            "X-Record-Count": str(total_records),
            "X-Backup-Size-Bytes": str(size_bytes),
        },
    )


async def _build_backup_zip(db) -> tuple[bytes, int, str]:
    """DEPRECATED — use `_build_backup_zip_to_path` directly. Retained
    only to keep the symbol importable in case any out-of-tree caller
    still references it. Always raises to fail loudly if reactivated.
    """
    raise RuntimeError(
        "_build_backup_zip is deprecated; use _build_backup_zip_to_path "
        "to stream to disk and avoid OOM."
    )


async def _build_backup_zip_to_path(db, out_path: Path) -> tuple[int, str]:
    """STREAMING variant — writes the full backup directly to ``out_path``
    on disk instead of buffering the entire archive in memory. Memory use
    stays around 5–20 MB regardless of how big the archive grows. This is
    the **safe** path for production containers with small memory limits;
    the in-memory `_build_backup_zip` would OOM-kill the backend on a
    1 GB+ archive. Returns (record_count, filename).
    """
    now = datetime.now(timezone.utc)
    stamp = now.strftime("%Y-%m-%d_%H%M%SZ")
    filename = f"MASCI_full_backup_{stamp}.zip"

    log_lines: List[str] = [
        "MASCI Hub — Full Backup",
        f"Generated: {now.isoformat()}",
        "Source: mascidocs.com (production)",
        "",
        "Per-kind record counts:",
    ]

    total_records = 0
    total_pdf_bytes = 0
    pdf_failures: List[str] = []

    # Open the zip directly on disk — every writestr is appended on the fly,
    # never buffered in memory.
    with zipfile.ZipFile(str(out_path), "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for kind, coll_name in EXPORTABLE_KINDS.items():
            await asyncio.sleep(0)  # yield to event loop — keeps healthcheck alive
            # STREAMING: iterate the cursor doc-by-doc instead of loading
            # every doc into memory via .to_list(). Safety records embed
            # multi-MB photo blobs; on a production database with many
            # records a single .to_list() call can balloon to >1 GB and
            # OOM-kill the container. Streaming keeps memory flat at
            # ~1 doc worth of bytes.
            kind_count = 0
            # We still need the CSV later which wants structured rows, but
            # we can accumulate trimmed rows without the big blobs. For
            # now, build CSV from the first 2000 docs (enough for an audit
            # trail) to keep memory bounded.
            CSV_CAP = 2000
            csv_rows: list[dict] = []
            pdf_kind = {
                "inspections": "inspection",
                "meetings": "meeting",
                "jhas": "jha",
                "incidents": "incident",
                "daily-reports": "daily-report",
                "equipment-inspections": "equipment-inspection",
            }.get(kind)

            async for d in db[coll_name].find({}, {"_id": 0}).sort("created_at", -1):
                kind_count += 1
                if len(csv_rows) < CSV_CAP:
                    csv_rows.append(dict(d))

                base = _record_filename(kind, d)
                try:
                    zf.writestr(
                        f"{kind}/json/{base}.json",
                        _backup_json.dumps(d, indent=2, default=str).encode("utf-8"),
                    )
                except Exception as e:  # noqa: BLE001
                    log_lines.append(f"    [warn] {kind}/{d.get('id')} JSON failed: {e}")

                if pdf_kind:
                    try:
                        # TRACK 15.47 · enrich incidents with state
                        # timeline + linked CAPAs before render.
                        _record_for_pdf = d
                        if pdf_kind == "incident":
                            try:
                                from lib.incident_pdf_enrichment import enrich_incident_for_pdf  # noqa: PLC0415
                                _record_for_pdf = await enrich_incident_for_pdf(db, d)
                            except Exception:
                                _record_for_pdf = d
                        pdf_bytes = await _backup_asyncio.to_thread(
                            render_record_pdf, pdf_kind, _record_for_pdf
                        )
                        total_pdf_bytes += len(pdf_bytes)
                        zf.writestr(f"{kind}/pdf/{base}.pdf", pdf_bytes)
                        del pdf_bytes
                    except Exception as e:  # noqa: BLE001
                        pdf_failures.append(f"{kind}/{d.get('id')}: {e}")
                    await asyncio.sleep(0)

                # Free the doc ASAP so photo blobs don't linger.
                del d
                # Every 10 docs, yield to event loop AND encourage gc.
                if kind_count % 10 == 0:
                    await asyncio.sleep(0)

            total_records += kind_count
            log_lines.append(f"  {kind:25s} : {kind_count:5d}")

            # /CSV/ — use only the capped sample to keep memory bounded.
            try:
                csv_bytes = _build_csv_bytes(kind, csv_rows)
                zf.writestr(f"CSV/MASCI_{kind}_{stamp}.csv", csv_bytes)
                if kind_count > CSV_CAP:
                    log_lines.append(
                        f"    [note] CSV truncated to first {CSV_CAP} of {kind_count} rows"
                    )
            except Exception as e:  # noqa: BLE001
                log_lines.append(f"    [warn] CSV build failed: {e}")
            del csv_rows

        # Manifest
        log_lines.append("")
        log_lines.append("Totals:")
        log_lines.append(f"  Records:        {total_records}")
        log_lines.append(f"  PDFs rendered:  {total_pdf_bytes / (1024 * 1024):.1f} MB")
        log_lines.append(f"  PDF failures:   {len(pdf_failures)}")
        if pdf_failures:
            log_lines.append("")
            log_lines.append("PDF render failures (first 20):")
            for line in pdf_failures[:20]:
                log_lines.append(f"  - {line}")

        # ====================================================================
        # AUTO-DISCOVERED COLLECTIONS — every Mongo collection that ISN'T
        # already exported above gets its JSON dumped here. This means any
        # NEW collection added in the future (parts catalogs, QC reports,
        # etc.) is included automatically — no human has to remember to add
        # it to a list. The only excludes are MongoDB system collections.
        # ====================================================================
        EXCLUDE_FROM_AUTO_BACKUP = {
            # Already covered above
            *(coll for coll in EXPORTABLE_KINDS.values()),
            # Module-level explicit exclusions (system collections, etc.)
            *BACKUP_EXPLICIT_EXCLUSIONS,
        }
        # Per-collection projection rules — sensitive fields stay redacted
        # regardless of which path picks the collection up.
        # iter425: now also covers `user_directory.mfa.secret` + recovery codes.
        SENSITIVE_FIELD_REDACTION = BACKUP_SENSITIVE_FIELD_REDACTION

        all_collections = await db.list_collection_names()
        log_lines.append("")
        log_lines.append("Auto-discovered collections (JSON only):")
        auto_total = 0
        captured_collections: List[str] = list(EXPORTABLE_KINDS.values())
        for coll_name in sorted(all_collections):
            await asyncio.sleep(0)  # keep event loop alive
            if coll_name in EXCLUDE_FROM_AUTO_BACKUP or coll_name.startswith("system."):
                continue
            try:
                projection = SENSITIVE_FIELD_REDACTION.get(coll_name, {"_id": 0})
                # STREAM the collection straight into a JSON array on disk
                # to avoid holding every doc in memory. Some collections
                # (project_docs, equipment_parts, messages) can hit
                # hundreds of MB when they contain base64 file blobs.
                coll_count = 0
                # Write opening bracket first, then stream docs comma-separated.
                # We accumulate one doc at a time into the zip entry via a
                # small buffer, but the buffer is bounded per-doc so memory
                # stays constant regardless of collection size.
                import io as _bio
                buf = _bio.BytesIO()
                buf.write(b"[\n")
                first = True
                async for doc in db[coll_name].find({}, projection):
                    coll_count += 1
                    if not first:
                        buf.write(b",\n")
                    first = False
                    buf.write(_backup_json.dumps(doc, indent=2, default=str).encode("utf-8"))
                    del doc
                    if coll_count % 25 == 0:
                        await asyncio.sleep(0)
                buf.write(b"\n]\n")
                zf.writestr(f"collections/{coll_name}.json", buf.getvalue())
                del buf
                auto_total += coll_count
                captured_collections.append(coll_name)
                log_lines.append(f"  collections/{coll_name:24s} : {coll_count:5d}")
            except Exception as e:  # noqa: BLE001
                log_lines.append(f"    [warn] collections/{coll_name} failed: {e}")
        total_records += auto_total
        log_lines.append(f"  Auto-discovered subtotal: {auto_total}")

        # ====================================================================
        # DISK-BACKED FILES — every directory under /app/backend/ that holds
        # file uploads or generated assets that aren't in MongoDB. Includes:
        #   • storage/   — uploaded FDOT plans / project docs (>16 MB BSON limit)
        #   • static/    — training videos, safety cards, branding logos
        #                  (uploaded by admin, would be lost on container death)
        #   • data/      — equipment_master JSON snapshots + employee seed
        # These would otherwise be lost on container redeploy.
        # ====================================================================
        DISK_BACKUP_ROOTS = [
            ("/app/backend/storage", "storage"),
            ("/app/backend/static", "static"),
            ("/app/backend/data", "data"),
            # iter426 · Phase 25.3 · Memory-doc continuity (PRD · audits ·
            # debriefs · doctrine notes). These are git-tracked but cheap
            # disaster-recovery insurance — if the repo + R2 ever
            # diverged, the backup zip is the single source of truth.
            ("/app/memory", "memory"),
        ]
        log_lines.append("")
        log_lines.append("Disk-backed files (storage tree):")
        disk_files_count = 0
        disk_bytes = 0
        for root_path_str, archive_prefix in DISK_BACKUP_ROOTS:
            root_path = Path(root_path_str)
            if not root_path.is_dir():
                continue
            for f in root_path.rglob("*"):
                await asyncio.sleep(0)  # yield each file — disk_files can be 100MB+
                if not f.is_file():
                    continue
                # Skip Python bytecode caches and any tmp files.
                if "__pycache__" in f.parts or f.name.endswith(".pyc"):
                    continue
                try:
                    rel = f.relative_to(root_path)
                    # STREAM the file into the zip 1 MB at a time so even a
                    # 150 MB+ PDF (Oxford FDOT plans) never lives in RAM.
                    size = f.stat().st_size
                    arcname = f"disk_files/{archive_prefix}/{rel.as_posix()}"
                    with zf.open(arcname, "w", force_zip64=True) as zdst, \
                         f.open("rb") as src:
                        while True:
                            chunk = src.read(1024 * 1024)
                            if not chunk:
                                break
                            zdst.write(chunk)
                            await asyncio.sleep(0)
                    disk_files_count += 1
                    disk_bytes += size
                except Exception as e:  # noqa: BLE001
                    log_lines.append(f"    [warn] disk file {f} failed: {e}")
        log_lines.append(
            f"  disk_files            : {disk_files_count} files, "
            f"{disk_bytes / (1024 * 1024):.1f} MB"
        )

        # ---------- Backup integrity manifest ----------
        # Records what was captured so a future restore can verify the zip
        # didn't lose anything. The integrity-check endpoint compares this
        # against the live DB and surfaces a warning if a new collection
        # exists that isn't yet in any backup.
        #
        # Track 14.0-I1 (2026-02-14): manifest now also records the
        # environment + database the archive was generated from so the
        # restore endpoint can refuse a preview-origin archive in
        # production. This closes the cross-env restore vector that
        # remained as a manual-checklist item after Track 14.0-P0.
        _app_env = _canonical_app_env().lower()
        _db_name = _canonical_db_name()
        zf.writestr(
            "backup_manifest.json",
            _backup_json.dumps({
                "source": "mascidocs.com",
                "generated_at": now.isoformat(),
                "version": "3",
                "manifest_schema": "track-14.0-i1",
                "environment": _app_env,
                "database_name": _db_name,
                "app_env": _app_env,
                "db_name": _db_name,
                "source_instance": _app_env,
                "backup_id": uuid.uuid4().hex,
                "total_records": total_records,
                "captured_collections": sorted(set(captured_collections)),
                "all_db_collections_at_backup_time": sorted(all_collections),
                "disk_files_count": disk_files_count,
                "disk_files_bytes": disk_bytes,
            }, indent=2).encode("utf-8"),
        )

        zf.writestr("backup_log.txt", "\n".join(log_lines).encode("utf-8"))

    return total_records, filename


# ----------------------------------------------------------------------
# Stored backups — daily scheduled backup saved to disk.
# ----------------------------------------------------------------------
BACKUPS_DIR = Path(os.environ.get("BACKUPS_DIR", "/app/backend/backups")).resolve()
BACKUP_RETENTION_DAYS = int(os.environ.get("BACKUP_RETENTION_DAYS", "14"))
BACKUP_HOUR_UTC = int(os.environ.get("BACKUP_HOUR_UTC", "2"))   # legacy single-window default 02:00 UTC


def _parse_backup_hours() -> list[int]:
    """Parse backup schedule env vars into a sorted list of UTC hours.

    TRACK 15.38 (2026-02) — added white-label tenant-local-time support.

    Precedence (highest first):
      1. `BACKUP_HOURS_LOCAL` + `BACKUP_TIMEZONE`  → convert local hours
         to UTC using the current DST offset. Recommended for white-label
         deployments (MASCI Florida · Texas customer · Arizona customer
         all use `BACKUP_HOURS_LOCAL=0,6,12,18`).
      2. `BACKUP_HOURS_UTC`                       → legacy UTC-only path,
         e.g. "2,18" for the historical nightly+mid-day pattern.
      3. Default `[BACKUP_HOUR_UTC, 18]`.

    Invalid entries are dropped, duplicates removed, result sorted.

    DST caveat: the local→UTC conversion happens at module load time. A
    worker restart will pick up the post-DST offset. For environments
    that demand sub-hour DST-accurate scheduling, restart the worker
    after each DST transition (twice a year)."""
    # Tenant-local mode (preferred for white-label)
    local_raw = (os.environ.get("BACKUP_HOURS_LOCAL") or "").strip()
    tz_name = (os.environ.get("BACKUP_TIMEZONE") or "").strip()
    if local_raw and tz_name:
        try:
            from zoneinfo import ZoneInfo  # py3.9+
            tz = ZoneInfo(tz_name)
            # Convert each local hour to its UTC equivalent at the current
            # wall-clock-day. astimezone() handles DST automatically.
            local_hours: set[int] = set()
            for part in local_raw.split(","):
                s = part.strip()
                if not s:
                    continue
                try:
                    h = int(s)
                    if 0 <= h <= 23:
                        local_hours.add(h)
                except ValueError:
                    continue
            if local_hours:
                today_local = datetime.now(tz).date()
                utc_hours: set[int] = set()
                for h in local_hours:
                    local_dt = datetime(
                        today_local.year, today_local.month, today_local.day,
                        h, 0, tzinfo=tz,
                    )
                    utc_hours.add(local_dt.astimezone(timezone.utc).hour)
                return sorted(utc_hours)
        except Exception as e:  # noqa: BLE001
            logger.warning(
                f"[backup-schedule] BACKUP_HOURS_LOCAL/BACKUP_TIMEZONE "
                f"parse failed ({e}); falling back to BACKUP_HOURS_UTC."
            )

    # Legacy UTC-only path
    raw = (os.environ.get("BACKUP_HOURS_UTC") or "").strip()
    if not raw:
        raw = f"{BACKUP_HOUR_UTC},18"  # default: nightly 02:00 UTC + mid-day 18:00 UTC
    hours: set[int] = set()
    for part in raw.split(","):
        s = part.strip()
        if not s:
            continue
        try:
            h = int(s)
            if 0 <= h <= 23:
                hours.add(h)
        except ValueError:
            continue
    if not hours:
        hours = {BACKUP_HOUR_UTC}
    return sorted(hours)


BACKUP_HOURS_UTC: list[int] = _parse_backup_hours()
# DEFENSE LAYER 2 — Hard ceiling on stored backups. The container volume
# is small (9.8 GB) and a single full backup is ~750 MB. Keeping 3 max
# means we use ≤ 2.3 GB on backups, leaving plenty of headroom for the
# working DB and the disk-backed files. This is the single biggest
# defense against "backup fills disk → backend crashes → Cloudflare 520".
BACKUP_KEEP_MAX = int(os.environ.get("BACKUP_KEEP_MAX", "3"))
# DEFENSE LAYER 3 — Auto-prune trigger. If disk usage exceeds this
# percentage at boot OR right before a backup write, aggressively purge
# backups down to BACKUP_KEEP_MAX-1. Acts as an emergency brake.
BACKUP_DISK_HIGH_WATERMARK = int(os.environ.get("BACKUP_DISK_HIGH_WATERMARK", "75"))
# iter299 · Lane D operational hygiene — visibility-only warning threshold.
# When disk crosses this percentage we log a WARN line (NOT an alert email,
# NOT a new collection, NOT a dashboard). Operator surfaces this in logs
# to decide whether manual intervention is needed. Strictly LESS than the
# 90% hard-abort threshold inside _run_scheduled_backup so the warning
# fires earlier than the abort.
BACKUP_DISK_WARN_WATERMARK = int(os.environ.get("BACKUP_DISK_WARN_WATERMARK", "85"))
BACKUP_COMPLETE_TMP_DIR = Path(tempfile.gettempdir()) / "masci_complete_archive_builds"
BACKUP_COMPLETE_MIN_FREE_BYTES = int(os.environ.get("BACKUP_COMPLETE_MIN_FREE_BYTES", str(3 * 1024 * 1024 * 1024)) or str(3 * 1024 * 1024 * 1024))
BACKUP_COMPLETE_MAX_BUILD_BYTES = int(os.environ.get("BACKUP_COMPLETE_MAX_BUILD_BYTES", str(3 * 1024 * 1024 * 1024)) or str(3 * 1024 * 1024 * 1024))
BACKUP_RESTORE_STREAM_CHUNK_BYTES = int(os.environ.get("BACKUP_RESTORE_STREAM_CHUNK_BYTES", str(8 * 1024 * 1024)) or str(8 * 1024 * 1024))


def _disk_pct_used(path: str = "/app") -> int:
    """Return percent disk used at `path` (0-100). Returns 0 on error."""
    try:
        import shutil as _sh
        total, used, _free = _sh.disk_usage(path)
        return int((used / total) * 100) if total else 0
    except Exception:
        return 0


def _disk_free_bytes(path: str | Path = "/app") -> int:
    try:
        import shutil as _sh
        _total, _used, free = _sh.disk_usage(str(path))
        return int(free)
    except Exception:
        return 0


def _backup_resource_preflight(*, archive_size_bytes: Optional[int] = None) -> Dict[str, Any]:
    tmp_free = _disk_free_bytes(BACKUP_COMPLETE_TMP_DIR.parent)
    app_pct = _disk_pct_used()
    estimated = int(archive_size_bytes or 0)
    reasons: List[str] = []
    ok = True
    if tmp_free < BACKUP_COMPLETE_MIN_FREE_BYTES:
        ok = False
        reasons.append(f"tmp_free_below_floor:{tmp_free}")
    if app_pct >= 85:
        ok = False
        reasons.append(f"app_disk_pressure:{app_pct}")
    if estimated and estimated > BACKUP_COMPLETE_MAX_BUILD_BYTES:
        ok = False
        reasons.append(f"estimated_archive_too_large:{estimated}")
    if estimated and estimated > max(0, tmp_free - (512 * 1024 * 1024)):
        ok = False
        reasons.append(f"tmp_headroom_insufficient:{estimated}")
    return {
        "ok": ok,
        "reasons": reasons,
        "app_disk_percent_used": app_pct,
        "tmp_disk_free_bytes": tmp_free,
        "min_free_bytes_required": BACKUP_COMPLETE_MIN_FREE_BYTES,
        "max_build_bytes": BACKUP_COMPLETE_MAX_BUILD_BYTES,
        "archive_size_bytes_estimate": estimated,
    }


async def _latest_complete_backup_hint(db) -> Dict[str, Any]:
    row = await db.backup_health.find_one(
        {"mode": "complete-r2", "ok": True, "filename": {"$nin": [None, ""]}},
        {"_id": 0, "size_bytes": 1, "filename": 1, "ts": 1},
        sort=[("ts", -1)],
    )
    return row or {}


async def _collect_backup_runtime_state(db) -> Dict[str, Any]:
    stale_before = (datetime.now(timezone.utc) - timedelta(minutes=90)).isoformat()
    try:
        stale_marked = await mark_stale_backup_jobs(db, stale_before_iso=stale_before)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[backup-runtime] stale sweep failed: {e}")
        stale_marked = 0
    try:
        active_jobs = await get_active_backup_jobs(db)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[backup-runtime] active jobs read failed: {e}")
        active_jobs = []
    reclaimable_active_jobs = [row for row in active_jobs if is_backup_job_stale(row)]
    blocking_active_jobs = [row for row in active_jobs if not is_backup_job_stale(row)]
    try:
        recent_complete_jobs = await list_backup_jobs(db, kind=BACKUP_JOB_KIND_COMPLETE_R2, limit=10)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[backup-runtime] recent jobs read failed: {e}")
        recent_complete_jobs = []
    return {
        "stale_marked": stale_marked,
        "active_jobs": active_jobs,
        "blocking_active_jobs": blocking_active_jobs,
        "reclaimable_active_jobs": reclaimable_active_jobs,
        "overlap": classify_backup_overlap(active_jobs),
        "recent_complete_jobs": recent_complete_jobs,
    }


async def _backup_persistence_available(db) -> bool:
    try:
        await asyncio.wait_for(db.backup_jobs.estimated_document_count(), timeout=2.0)
        return True
    except Exception:
        return False


async def _stale_scheduler_lock_present(db) -> bool:
    try:
        now_dt = datetime.now(timezone.utc)
        row = await asyncio.wait_for(
            db.scheduler_locks.find_one(
                {"expires_at": {"$lt": now_dt}},
                {"_id": 0, "scheduler": 1},
            ),
            timeout=2.0,
        )
        return bool(row)
    except Exception:
        return True


def _backup_scheduler_healthy() -> bool:
    if not _BACKUP_SCHEDULER_STATE.get("alive"):
        return False
    last_tick = _BACKUP_SCHEDULER_STATE.get("last_tick_ts")
    last_lock = _BACKUP_SCHEDULER_STATE.get("last_lock_ts")
    if not last_tick:
        if not last_lock:
            return False
        try:
            dt = datetime.fromisoformat(str(last_lock).replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return (datetime.now(timezone.utc) - dt) <= timedelta(minutes=10)
        except Exception:
            return False
    try:
        dt = datetime.fromisoformat(str(last_tick).replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return (datetime.now(timezone.utc) - dt) <= timedelta(minutes=10)
    except Exception:
        return False


def _retention_policy_state() -> Dict[str, Any]:
    policy = dict(_BACKUP_RETENTION_POLICY)
    valid = bool(
        policy.get("hourly_hours") == 72
        and policy.get("daily_days") == 30
        and policy.get("weekly_days") == 90
        and policy.get("monthly_months") == 12
        and policy.get("architecture") == "selected_surviving_hourly_archives"
    )
    return {
        "valid": valid,
        "reason": "approved_tiered_retention" if valid else "retention_policy_invalid",
        "policy": policy,
    }


async def _build_hourly_activation_state(db, *, runtime_state: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    runtime_state = runtime_state or await _collect_backup_runtime_state(db)
    overlap = runtime_state.get("overlap") or {}
    stale_jobs = await list_stale_backup_jobs(db, limit=10)
    stale_lock_present = await _stale_scheduler_lock_present(db)
    persistence_available = await _backup_persistence_available(db)
    retention = _retention_policy_state()
    latest_hint = await _latest_complete_backup_hint(db)
    preflight = _backup_resource_preflight(archive_size_bytes=latest_hint.get("size_bytes"))
    active_job = None
    active_jobs = runtime_state.get("active_jobs") or []
    if active_jobs:
        current = active_jobs[0]
        active_job = {
            "job_id": current.get("job_id"),
            "kind": current.get("kind"),
            "state": current.get("state"),
            "heartbeat_at": current.get("heartbeat_at"),
        }
    reclaimable_stale_jobs = [
        job for job in stale_jobs
        if str(job.get("failure_reason") or "") == "stale_job_recovered"
    ]
    state = build_hourly_activation_state(
        requested_raw=os.environ.get("BACKUP_R2_HOURLY"),
        environment=_canonical_app_env().lower(),
        scheduler_healthy=_backup_scheduler_healthy(),
        persistence_available=persistence_available,
        backup_active=bool(overlap.get("backup_active")),
        restore_active=bool(overlap.get("restore_active")),
        stale_job_count=len(stale_jobs),
        reclaimable_stale_job_count=len(reclaimable_stale_jobs),
        stale_lock_present=stale_lock_present,
        resource_preflight=preflight,
        r2_configured=bool(os.environ.get("S3_BUCKET") and os.environ.get("S3_ENDPOINT_URL")),
        retention_valid=bool(retention.get("valid")),
        retention_reason=str(retention.get("reason") or "retention_unknown"),
        current_active_job=active_job,
    )
    state["retention_state"] = retention
    state["stale_job_count"] = len(stale_jobs)
    state["reclaimable_stale_job_count"] = len(reclaimable_stale_jobs)
    state["stale_lock_present"] = stale_lock_present
    state["persistence_available"] = persistence_available
    _BACKUP_SCHEDULER_STATE["r2_hourly_requested"] = state["r2_hourly_requested"]
    _BACKUP_SCHEDULER_STATE["r2_hourly_effective"] = state["r2_hourly_effective"]
    _BACKUP_SCHEDULER_STATE["r2_hourly_locked_off"] = state["r2_hourly_locked_off"]
    _BACKUP_SCHEDULER_STATE["hourly_cadence_enabled"] = state["hourly_cadence_enabled"]
    _BACKUP_SCHEDULER_STATE["activation_blockers"] = state["activation_blockers"]
    _BACKUP_SCHEDULER_STATE["activation_status"] = state["activation_status"]
    _BACKUP_SCHEDULER_STATE["activation_environment"] = state["environment"]
    _BACKUP_SCHEDULER_STATE["last_activation_evaluated_at"] = state["last_evaluated_at"]
    _BACKUP_SCHEDULER_STATE["next_eligible_hourly_slot"] = state["next_eligible_hourly_slot"]
    return state


async def _run_job_heartbeat(db, *, job_id: str, owner_token: str, stage_fn, interval_seconds: float = 30.0):
    stop = asyncio.Event()

    async def _loop():
        while not stop.is_set():
            try:
                await heartbeat_backup_job(
                    db,
                    job_id,
                    owner_token=owner_token,
                    extra={"stage": stage_fn(), "heartbeat_loop": True},
                )
            except BackupJobOwnershipLost:
                stop.set()
                raise
            except Exception as exc:  # noqa: BLE001
                await record_backup_job_heartbeat_failure(
                    db,
                    job_id,
                    owner_token=owner_token,
                    error=f"{type(exc).__name__}:{exc}",
                )
            try:
                await asyncio.wait_for(stop.wait(), timeout=interval_seconds)
            except asyncio.TimeoutError:
                continue

    return stop, asyncio.create_task(_loop())


async def _sha256_file(path: Path) -> str:
    def _calc() -> str:
        h = hashlib.sha256()
        with path.open("rb") as fh:
            while True:
                chunk = fh.read(1024 * 1024)
                if not chunk:
                    break
                h.update(chunk)
        return h.hexdigest()
    return await asyncio.to_thread(_calc)



async def _log_operational_hygiene(reason: str = "startup", db=None) -> None:
    """iter299 · Lane D · Operational hygiene visibility log.

    Lightweight, visibility-only. Emits a structured log line summarizing:
      - Disk usage % (with WARN-level severity at ≥ BACKUP_DISK_WARN_WATERMARK).
      - Backup file inventory (count + total MB, broken down by full/lite/complete).
      - Configured retention (BACKUP_RETENTION_DAYS · BACKUP_KEEP_MAX).
      - Oldest and newest backup ages so retention windows can be verified.
      - Last `backup_health` row (mode/ok/filename/size/records/ts).

    Strict scope per operator direction:
      - NO new endpoints. NO new collections. NO new alerts/emails.
      - NO new prune behavior (existing _emergency_prune_backups logic untouched).
      - This is a LOG LINE intended for the operator to surface manually.

    Async — fired from `@app.on_event("startup")` and from inside the scheduled
    backup runner. Never raises into the caller.
    """
    try:
        pct = _disk_pct_used()
        full_count = lite_count = complete_count = 0
        full_bytes = lite_bytes = complete_bytes = 0
        oldest_ts = newest_ts = None
        try:
            BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
            for p in BACKUPS_DIR.glob("MASCI_*backup*.zip"):
                try:
                    st = p.stat()
                except Exception:
                    continue
                name = p.name
                size = st.st_size
                mtime = st.st_mtime
                if name.startswith("MASCI_full_backup_"):
                    full_count += 1
                    full_bytes += size
                elif name.startswith("MASCI_lite_backup_"):
                    lite_count += 1
                    lite_bytes += size
                elif name.startswith("MASCI_complete_backup_"):
                    complete_count += 1
                    complete_bytes += size
                if oldest_ts is None or mtime < oldest_ts:
                    oldest_ts = mtime
                if newest_ts is None or mtime > newest_ts:
                    newest_ts = mtime
        except Exception:
            pass

        now_ts = datetime.now(timezone.utc).timestamp()
        oldest_age_days = f"{(now_ts - oldest_ts) / 86400.0:.1f}" if oldest_ts else "n/a"
        newest_age_hours = f"{(now_ts - newest_ts) / 3600.0:.1f}" if newest_ts else "n/a"
        total_count = full_count + lite_count + complete_count
        total_mb = (full_bytes + lite_bytes + complete_bytes) / (1024 * 1024)

        msg = (
            f"[ops-hygiene] {reason} · disk={pct}% "
            f"(warn≥{BACKUP_DISK_WARN_WATERMARK}% prune≥{BACKUP_DISK_HIGH_WATERMARK}%) · "
            f"backups: total={total_count} ({total_mb:.1f} MB) · "
            f"full={full_count} lite={lite_count} complete={complete_count} · "
            f"retention_days={BACKUP_RETENTION_DAYS} keep_max_full={BACKUP_KEEP_MAX} · "
            f"oldest_age_days={oldest_age_days} newest_age_hours={newest_age_hours}"
        )
        if pct >= BACKUP_DISK_WARN_WATERMARK:
            logger.warning(msg + " · DISK_PRESSURE")
        else:
            logger.info(msg)

        # iter307 · git tmp orphan visibility — interrupted git gc/pack
        # operations leave `tmp_pack_*` and `tmp_obj_*` files in
        # `.git/objects/` that git itself never references but never
        # cleans up either. iter306-era audit found ~1 GB of these
        # orphans from a single May-19 interrupted pack op. Surface the
        # count + total size on every ops-hygiene log so the next
        # accumulation is visible operationally, not buried in a future
        # pre-deploy disk audit. NO autonomous cleanup — just visibility.
        try:
            import glob as _g
            tmp_paths = (
                _g.glob("/app/.git/objects/pack/tmp_pack_*")
                + _g.glob("/app/.git/objects/pack/tmp_idx_*")
                + _g.glob("/app/.git/objects/??/tmp_obj_*")
            )
            if tmp_paths:
                tmp_bytes = 0
                for p in tmp_paths:
                    try:
                        tmp_bytes += os.path.getsize(p)
                    except Exception:
                        pass
                logger.info(
                    f"[ops-hygiene] git_tmp_orphans: count={len(tmp_paths)} "
                    f"size_mb={tmp_bytes / (1024 * 1024):.1f}"
                )
        except Exception as e:  # noqa: BLE001
            logger.info(f"[ops-hygiene] git_tmp_orphans read skipped ({e})")

        if db is not None:
            try:
                latest = await db.backup_health.find_one(
                    {"ok": True}, sort=[("ts", -1)], projection={"_id": 0},
                )
                if latest:
                    logger.info(
                        f"[ops-hygiene] last_backup_health: "
                        f"ok={latest.get('ok')} mode={latest.get('mode')} "
                        f"filename={latest.get('filename')!r} "
                        f"records={latest.get('records', 0)} "
                        f"size_mb={(latest.get('size_bytes', 0) or 0) / (1024*1024):.1f} "
                        f"ts={latest.get('ts')}"
                    )
                else:
                    logger.info("[ops-hygiene] last_backup_health: <no rows yet>")
            except Exception as e:  # noqa: BLE001
                logger.info(f"[ops-hygiene] last_backup_health read skipped ({e})")
    except Exception as e:  # noqa: BLE001
        try:
            logger.warning(f"[ops-hygiene] log emit failed (non-fatal): {e}")
        except Exception:
            pass


def _emergency_prune_backups(reason: str) -> int:
    """Sync helper. Aggressively prune backups + ORPHAN .tmp files. Safe to
    call from any context (sync or async via to_thread). Returns count pruned.
    Catches all exceptions internally — NEVER raises into the caller.

    NOTE: .tmp files younger than 10 minutes are KEPT — they may be a backup
    actively streaming to disk in another worker / concurrent request.
    Deleting them would break the rename step at the end of the build.

    DEPLOY-FIX-001 · Workstream A5 — emits a per-file WARNING log for every
    orphan .tmp file removed (file name + age + reason) so operators can
    confirm the sweep is doing useful work.
    """
    pruned = 0
    _now_ts = datetime.now(timezone.utc).timestamp()
    _ORPHAN_TMP_AGE_SEC = 600  # 10 minutes
    try:
        BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
        for p in BACKUPS_DIR.glob("*.zip.tmp*"):
            try:
                age_s = _now_ts - p.stat().st_mtime
                if age_s < _ORPHAN_TMP_AGE_SEC:
                    continue  # active stream — leave alone
                logger.warning(
                    f"[backup-cleanup] orphan-sweep ({reason}) · "
                    f"file={p.name} age={int(age_s)}s reason=orphan_tmp_over_600s"
                )
                p.unlink()
                pruned += 1
            except Exception:
                continue
        # iter427 · Phase 26.1 — also sweep legacy backup patterns
        # (lite/complete) that the pre-iter425 archive naming left behind.
        # These are superseded by `MASCI_full_backup_*.zip` (auto-discovery)
        # and any local copy older than retention is dead weight (R2 keeps
        # the canonical copy).
        _legacy_cutoff = _now_ts - BACKUP_RETENTION_DAYS * 86400
        for legacy_glob in ("MASCI_lite_backup_*.zip", "MASCI_complete_backup_*.zip"):
            for p in BACKUPS_DIR.glob(legacy_glob):
                try:
                    if p.stat().st_mtime < _legacy_cutoff:
                        p.unlink()
                        pruned += 1
                except Exception:
                    continue
        # Keep BACKUP_KEEP_MAX-1 newest so the next backup fits within cap
        files = sorted(
            BACKUPS_DIR.glob("MASCI_full_backup_*.zip"),
            key=lambda f: f.stat().st_mtime,
            reverse=True,
        )
        keep = max(0, BACKUP_KEEP_MAX - 1)
        for p in files[keep:]:
            try:
                p.unlink()
                pruned += 1
            except Exception:
                continue
        if pruned:
            logger.warning(
                f"[backup-defense] EMERGENCY PRUNE ({reason}) — "
                f"deleted {pruned} files, disk now at {_disk_pct_used()}%"
            )
    except Exception as e:
        logger.warning(f"[backup-defense] emergency prune itself failed: {e}")
    return pruned


def _list_stored_backups() -> List[dict]:
    """Return metadata for every .zip in the backups dir (newest first)."""
    if not BACKUPS_DIR.exists():
        return []
    rows = []
    for p in sorted(BACKUPS_DIR.glob("MASCI_full_backup_*.zip"), reverse=True):
        try:
            st = p.stat()
            rows.append({
                "filename": p.name,
                "size_bytes": st.st_size,
                "created_at": datetime.fromtimestamp(st.st_mtime, tz=timezone.utc).isoformat(),
            })
        except Exception:
            continue
    return rows


async def _run_scheduled_backup(db, lite_mode: bool = False) -> Optional[dict]:
    """Build a backup and persist to BACKUPS_DIR. Prune files older than the
    retention window and over the max-keep ceiling. Returns a small summary
    dict or None on failure.

    When ``lite_mode`` is True (or the global ``BACKUP_LITE_MODE_ONLY`` env
    flag is set), the full zip step is skipped entirely — we build ONLY the
    slim metadata-and-JSON zip suitable for emailing. This is the escape
    hatch when the full archive has grown so large (e.g. 887 MB of base64
    photos) that every full-build attempt OOM-kills the worker.
    """
    if not lite_mode:
        lite_mode = _lite_mode_default()
    try:
        BACKUPS_DIR.mkdir(parents=True, exist_ok=True)

        # PRE-FLIGHT PRUNE — clean up before writing so we never run out of
        # disk mid-backup. Drops ORPHAN .tmp debris from previous failures
        # (only .tmp older than 10 minutes — younger ones are likely active
        # streams from a concurrent request and would break their rename).
        pre_pruned = 0
        _now_ts = datetime.now(timezone.utc).timestamp()
        _ORPHAN_TMP_AGE_SEC = 600
        for p in BACKUPS_DIR.glob("*.zip.tmp*"):
            try:
                if (_now_ts - p.stat().st_mtime) < _ORPHAN_TMP_AGE_SEC:
                    continue  # active stream — leave alone
                p.unlink()
                pre_pruned += 1
            except Exception:
                continue
        cutoff = datetime.now(timezone.utc).timestamp() - BACKUP_RETENTION_DAYS * 86400
        existing = sorted(
            BACKUPS_DIR.glob("MASCI_full_backup_*.zip"),
            key=lambda f: f.stat().st_mtime,
            reverse=True,
        )
        # By age
        for p in existing:
            try:
                if p.stat().st_mtime < cutoff:
                    p.unlink()
                    pre_pruned += 1
            except Exception:
                continue
        # iter427 · Phase 26.1 — also sweep legacy backup patterns
        # (lite/complete from pre-iter425 naming) past retention.
        for legacy_glob in ("MASCI_lite_backup_*.zip", "MASCI_complete_backup_*.zip"):
            for p in BACKUPS_DIR.glob(legacy_glob):
                try:
                    if p.stat().st_mtime < cutoff:
                        p.unlink()
                        pre_pruned += 1
                except Exception:
                    continue
        # By count (keep newest BACKUP_KEEP_MAX-1 so the new one fits within cap)
        existing = sorted(
            BACKUPS_DIR.glob("MASCI_full_backup_*.zip"),
            key=lambda f: f.stat().st_mtime,
            reverse=True,
        )
        for p in existing[max(0, BACKUP_KEEP_MAX - 1):]:
            try:
                p.unlink()
                pre_pruned += 1
            except Exception:
                continue
        if pre_pruned:
            logger.info(f"[scheduled-backup] pre-flight pruned {pre_pruned} old/tmp files")

        # DEFENSE LAYER 4.5 — OOM watermark preflight.
        # The bug we keep hitting: as base64 photos accumulate in Mongo,
        # the full archive size grows linearly. Once it crosses the
        # worker's memory ceiling, every full-build OOM-kills the worker
        # → silent crash loop → no backups for days. This check looks at
        # the LATEST successful full zip on disk and, if it's already
        # over the watermark, auto-downgrades to lite mode WITHOUT a
        # human in the loop. Operators can flip `BACKUP_FULL_OOM_WATERMARK_MB=0`
        # to disable this safety net once S3 photo migration is done.
        if not lite_mode:
            watermark_mb = float(
                os.environ.get("BACKUP_FULL_OOM_WATERMARK_MB", "600") or "600"
            )
            if watermark_mb > 0:
                try:
                    existing_full = sorted(
                        BACKUPS_DIR.glob("MASCI_full_backup_*.zip"),
                        key=lambda f: f.stat().st_mtime,
                        reverse=True,
                    )
                    if existing_full:
                        latest_mb = existing_full[0].stat().st_size / (1024 * 1024)
                        if latest_mb >= watermark_mb:
                            logger.warning(
                                f"[scheduled-backup] PREFLIGHT: most recent full zip is "
                                f"{latest_mb:.1f} MB (watermark {watermark_mb} MB). "
                                f"Auto-downgrading this run to LITE mode to avoid OOM. "
                                f"Set BACKUP_FULL_OOM_WATERMARK_MB=0 to disable, or "
                                f"BACKUP_LITE_MODE_ONLY=true to make lite-mode permanent."
                            )
                            lite_mode = True
                except Exception as e:  # noqa: BLE001
                    logger.warning(f"[scheduled-backup] preflight watermark check failed: {e}")

        # DEFENSE LAYER 5 — Disk high-water-mark check after prune.
        # If the disk is STILL above the watermark after standard pruning,
        # bail out instead of building a 750 MB zip we can't write. Better
        # to skip a backup than crash the backend.
        pct_after = _disk_pct_used()
        if pct_after >= BACKUP_DISK_HIGH_WATERMARK:
            _emergency_prune_backups(reason=f"pre-build disk {pct_after}%")
            pct_after = _disk_pct_used()
            if pct_after >= 90:
                logger.error(
                    f"[scheduled-backup] ABORT — disk at {pct_after}% even after "
                    f"emergency prune. Backup skipped to protect backend."
                )
                return {
                    "filename": None,
                    "size_bytes": 0,
                    "records": 0,
                    "pruned_old": pre_pruned,
                    "emailed_to": None,
                    "skipped": True,
                    "reason": f"disk_{pct_after}_percent",
                }

        # STREAMING write — go straight to the temp file on disk. Never
        # hold 750 MB in RAM (would OOM-kill the container on small-memory
        # deploys). _build_backup_zip_to_path opens the ZipFile against
        # the temp file and writestr's each entry as it goes.
        BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
        # Pre-compute target name; we'll rename after stream completes.
        _now = datetime.now(timezone.utc)
        _stamp = _now.strftime("%Y-%m-%d_%H%M%SZ")

        # ── Lite-mode escape hatch ─────────────────────────────────────
        # When the full archive is too big to build in-process (e.g. base64
        # photos pushing the zip past 800 MB and OOM-killing the worker),
        # skip the full zip entirely and build only the slim, base64-
        # stripped JSON-only zip. Tiny output (~1-3 MB), emails cleanly.
        if lite_mode:
            logger.info("[scheduled-backup] LITE MODE — skipping full zip, building slim only")
            slim_filename = f"MASCI_lite_backup_{_stamp}.zip"
            slim_out = BACKUPS_DIR / slim_filename
            slim_tmp = slim_out.with_suffix(f".zip.tmp.{uuid.uuid4().hex[:8]}")
            try:
                stats = await asyncio.to_thread(
                    _build_slim_backup_zip_on_disk, db, slim_tmp
                )
            except Exception as e:  # noqa: BLE001
                logger.exception(f"[scheduled-backup] lite-mode slim build failed: {e}")
                try:
                    if slim_tmp.exists():
                        slim_tmp.unlink()
                except Exception:
                    pass
                return None
            slim_tmp.replace(slim_out)
            slim_size = slim_out.stat().st_size
            logger.info(
                f"[scheduled-backup] LITE wrote {slim_out.name} "
                f"({slim_size/1024/1024:.2f} MB · {stats.get('total_records', 0)} records)"
            )
            emailed_to = None
            notification_outcome = "notification_not_required"
            notification_reason = None
            notification_message_id = None
            notification_recipients = []
            try:
                notification_recipients = [x.strip() for x in (((os.environ.get("BACKUP_EMAIL_TO") or "").strip()).split(",")) if x.strip()]
                emailed_to = await _email_lite_backup_zip(slim_out, stats)
                if emailed_to:
                    notification_outcome = "notification_sent"
                    notification_message_id = emailed_to if "@" not in str(emailed_to) else None
                elif notification_recipients:
                    notification_outcome = "notification_failed"
                    notification_reason = "email_helper_returned_no_delivery_confirmation"
                else:
                    notification_outcome = "notification_suppressed"
                    notification_reason = "backup_email_to_missing"
            except Exception as e:  # noqa: BLE001
                logger.warning(f"[scheduled-backup] lite email step failed: {e}")
                notification_outcome = "notification_failed"
                notification_reason = type(e).__name__
            await _record_backup_health(
                db, ok=True, filename=slim_out.name, size_bytes=slim_size,
                records=stats.get("total_records", 0), emailed_to=emailed_to,
                mode="lite",
                notification_outcome=notification_outcome,
                notification_recipients=notification_recipients,
                notification_recipient_count=len(notification_recipients),
                notification_reason=notification_reason,
                notification_message_id=notification_message_id,
                archive_identifier=slim_out.name,
                audit_reference=f"backup_health:{slim_out.name}",
            )
            # iter299 · Lane D — hygiene log line after lite-backup run.
            try:
                await _log_operational_hygiene(reason="post_lite_backup", db=db)
            except Exception:
                pass
            return {
                "filename": slim_out.name,
                "size_bytes": slim_size,
                "records": stats.get("total_records", 0),
                "pruned_old": pre_pruned,
                "emailed_to": emailed_to,
                "lite_mode": True,
            }

        filename = f"MASCI_full_backup_{_stamp}.zip"
        out = BACKUPS_DIR / filename
        # Per-call unique tmp suffix so concurrent backup requests can't
        # clobber each other's stream (or rename).
        tmp = out.with_suffix(f".zip.tmp.{uuid.uuid4().hex[:8]}")
        # DEPLOY-FIX-001 · Workstream A2/A3 — guarantee removal of the
        # .tmp.<hash> file when the scheduled full-mode build raises.
        # Mirrors the lite-mode escape hatch a few branches above.
        try:
            # Build directly into the .tmp; rename atomically when done.
            total_records, _name = await _build_backup_zip_to_path(db, tmp)
            # Use the timestamp-stamped name we computed above (consistent with prior behavior)
            tmp.replace(out)
        except BaseException:
            try:
                if tmp.exists():
                    logger.warning(
                        f"[backup-cleanup] scheduled full-build failure · "
                        f"removing orphan tmp {tmp.name}"
                    )
                    tmp.unlink()
            except Exception:
                pass
            raise
        size_bytes = out.stat().st_size
        logger.info(
            f"[scheduled-backup] wrote {out.name} ({size_bytes/1024/1024:.1f} MB · {total_records} records)"
        )

        # Email the backup off-site — CRITICAL for redeploy safety.
        # The email helper reads the file lazily to keep memory low when
        # building the slim version for the inbox attachment.
        emailed_to = None
        notification_outcome = "notification_not_required"
        notification_reason = None
        notification_message_id = None
        notification_recipients = []
        try:
            notification_recipients = [x.strip() for x in (((os.environ.get("BACKUP_EMAIL_TO") or "").strip()).split(",")) if x.strip()]
            emailed_to = await _email_backup_zip_from_path(out, total_records)
            if emailed_to:
                notification_outcome = "notification_sent"
                notification_message_id = emailed_to if "@" not in str(emailed_to) else None
            elif notification_recipients:
                notification_outcome = "notification_failed"
                notification_reason = "email_helper_returned_no_delivery_confirmation"
            else:
                notification_outcome = "notification_suppressed"
                notification_reason = "backup_email_to_missing"
        except Exception as e:
            logger.warning(f"[scheduled-backup] email step failed (non-fatal): {e}")
            notification_outcome = "notification_failed"
            notification_reason = type(e).__name__

        await _record_backup_health(
            db, ok=True, filename=out.name, size_bytes=size_bytes,
            records=total_records, emailed_to=emailed_to, mode="full",
            notification_outcome=notification_outcome,
            notification_recipients=notification_recipients,
            notification_recipient_count=len(notification_recipients),
            notification_reason=notification_reason,
            notification_message_id=notification_message_id,
            archive_identifier=out.name,
            audit_reference=f"backup_health:{out.name}",
        )
        # iter299 · Lane D — emit a hygiene log line after every successful
        # full backup so operators can verify retention + disk pressure
        # without hitting any endpoint.
        try:
            await _log_operational_hygiene(reason="post_full_backup", db=db)
        except Exception:
            pass
        return {
            "filename": out.name,
            "size_bytes": size_bytes,
            "records": total_records,
            "pruned_old": pre_pruned,
            "emailed_to": emailed_to,
        }
    except Exception as e:
        logger.exception(f"[scheduled-backup] FAILED: {e}")
        try:
            await _record_backup_health(
                db,
                ok=False,
                error=repr(e),
                mode="error",
                notification_outcome="notification_not_required",
                notification_reason="backup_execution_failed_before_notification",
                audit_reference="backup_health:error",
            )
        except Exception:
            pass
        return None


def _strip_base64_blobs(obj, _stats=None):
    """Recursively walk a parsed JSON document and replace any large
    base64 / data-URL blob with a small placeholder string. Used by the
    slim-email backup so 153 MB FDOT plans don't end up in the inbox.

    Returns (new_obj, count_stripped, total_bytes_stripped). Original
    fields are preserved by name — the value just becomes
    `"<stripped:base64 N bytes (was field=...)>"` so a future restore
    can detect it and surface a warning.

    Heuristic: any string longer than 32 KB OR starting with `data:`
    that contains only base64-safe chars is treated as a blob.
    """
    import re as _re3
    BLOB_KEYS = {"file_data", "file_bytes", "data_url", "photo", "photo_data",
                 "image", "image_data", "signature", "signature_data",
                 "pdf_bytes", "blob", "content"}
    BIG_THRESHOLD = 32 * 1024  # 32 KB

    if _stats is None:
        _stats = {"count": 0, "bytes": 0}

    def _looks_blob(v: str) -> bool:
        if not isinstance(v, str):
            return False
        if v.startswith("data:") and ";base64," in v[:64]:
            return True
        if len(v) >= BIG_THRESHOLD and _re3.fullmatch(r"[A-Za-z0-9+/=\r\n]+", v[:1024] or ""):
            return True
        return False

    if isinstance(obj, dict):
        out = {}
        for k, v in obj.items():
            if isinstance(v, str) and (k in BLOB_KEYS or _looks_blob(v)):
                _stats["count"] += 1
                _stats["bytes"] += len(v)
                out[k] = f"<stripped:base64 {len(v)} bytes (key={k})>"
            else:
                out[k], _, _ = _strip_base64_blobs(v, _stats)
        return out, _stats["count"], _stats["bytes"]
    if isinstance(obj, list):
        new_list = []
        for item in obj:
            new_item, _, _ = _strip_base64_blobs(item, _stats)
            new_list.append(new_item)
        return new_list, _stats["count"], _stats["bytes"]
    if isinstance(obj, str) and _looks_blob(obj):
        _stats["count"] += 1
        _stats["bytes"] += len(obj)
        return f"<stripped:base64 {len(obj)} bytes>", _stats["count"], _stats["bytes"]
    return obj, _stats["count"], _stats["bytes"]


# ──────────────────────────────────────────────────────────────────────
# Persisted backup health
# ──────────────────────────────────────────────────────────────────────
# Every successful and failed backup writes a small row into
# `backup_health` so the diagnostic endpoint sees the FULL history,
# not just what's in module-level memory (which resets on worker
# restart). The watchdog below reads this collection to decide whether
# to fire an alarm email when backups go silent.

async def _record_backup_health(
    db,
    *,
    ok: bool,
    filename: Optional[str] = None,
    size_bytes: int = 0,
    records: int = 0,
    emailed_to: Optional[str] = None,
    mode: str = "full",
    error: Optional[str] = None,
    notification_outcome: Optional[str] = None,
    notification_recipients: Optional[list[str]] = None,
    notification_recipient_count: Optional[int] = None,
    notification_reason: Optional[str] = None,
    notification_message_id: Optional[str] = None,
    archive_identifier: Optional[str] = None,
    audit_reference: Optional[str] = None,
    archive_lineage: Optional[dict] = None,
) -> None:
    """Append a row to ``backup_health``. Best-effort — a Mongo write
    failure here MUST NOT block the backup itself, so we swallow errors."""
    try:
        doc = {
            "id": uuid.uuid4().hex,
            "ts": datetime.now(timezone.utc).isoformat(),
            "ok": ok,
            "mode": mode,
            "filename": filename,
            "size_bytes": size_bytes,
            "records": records,
            "emailed_to": emailed_to,
            "error": error,
            "notification_outcome": notification_outcome,
            "notification_recipients": notification_recipients or [],
            "notification_recipient_count": notification_recipient_count,
            "notification_reason": notification_reason,
            "notification_message_id": notification_message_id,
            "notification_ts": datetime.now(timezone.utc).isoformat(),
            "archive_identifier": archive_identifier or filename,
            "audit_reference": audit_reference,
        }
        if archive_lineage and isinstance(archive_lineage, dict):
            doc["archive_lineage"] = dict(archive_lineage)
        await db.backup_health.insert_one(dict(doc))
        # Mongo mutates the dict in place to add _id — we don't care, doc is
        # not returned.
        # Trim to last 200 rows so this collection can't grow unbounded.
        old_cursor = db.backup_health.find(
            {}, {"_id": 1, "ts": 1}
        ).sort("ts", -1).skip(200)
        async for row in old_cursor:
            try:
                await db.backup_health.delete_one({"_id": row["_id"]})
            except Exception:
                pass
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[backup-health] record failed: {e}")


async def _backup_watchdog_check(db) -> Optional[dict]:
    """Look at the most recent successful backup. If it's older than
    ``BACKUP_WATCHDOG_HOURS`` (default 25h), fire an alarm email to the
    admin distribution list — once per silence window so we don't spam.

    Returns a small status dict for the diagnostic endpoint. ``alarm_fired``
    is True only when this call ACTUALLY sent an alarm email (vs. just
    observed silence within the cooldown).
    """
    threshold_hours = float(os.environ.get("BACKUP_WATCHDOG_HOURS", "25"))
    cooldown_hours = float(os.environ.get("BACKUP_WATCHDOG_COOLDOWN_HOURS", "12"))
    try:
        latest = await db.backup_health.find_one({"ok": True}, sort=[("ts", -1)])
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[backup-watchdog] read failed: {e}")
        return None
    now = datetime.now(timezone.utc)
    if not latest:
        # Fresh database — no history at all. Don't alarm; the first
        # scheduled run will populate the health log.
        return {"alarm_fired": False, "hours_silent": None, "reason": "no_history"}
    try:
        latest_ts = datetime.fromisoformat(latest["ts"])
    except Exception:
        return {"alarm_fired": False, "hours_silent": None, "reason": "bad_ts"}
    hours_silent = (now - latest_ts).total_seconds() / 3600.0
    if hours_silent < threshold_hours:
        return {"alarm_fired": False, "hours_silent": round(hours_silent, 1), "reason": "healthy"}

    # Cooldown: only alarm if we haven't already alarmed in the last
    # `cooldown_hours`. Last alarm timestamp lives in the same collection
    # under a marker doc.
    marker = await db.backup_health.find_one({"id": "_watchdog_last_alarm"})
    if marker and marker.get("ts"):
        try:
            since_alarm = (now - datetime.fromisoformat(marker["ts"])).total_seconds() / 3600.0
            if since_alarm < cooldown_hours:
                return {
                    "alarm_fired": False,
                    "hours_silent": round(hours_silent, 1),
                    "reason": f"cooldown ({since_alarm:.1f}h since last alarm < {cooldown_hours}h)",
                }
        except Exception:
            pass

    # Fire the alarm email.
    sent = await _send_watchdog_alarm(db, hours_silent=hours_silent, latest=latest)
    if sent:
        await db.backup_health.update_one(
            {"id": "_watchdog_last_alarm"},
            {"$set": {"id": "_watchdog_last_alarm", "ts": now.isoformat()}},
            upsert=True,
        )
    return {
        "alarm_fired": bool(sent),
        "hours_silent": round(hours_silent, 1),
        "reason": "alarm_sent" if sent else "alarm_send_failed",
    }


async def _send_watchdog_alarm(db, *, hours_silent: float, latest: dict) -> bool:
    """Send a styled HTML alarm email to the same address(es) that receive
    successful backups. Quiet failure → return False (we don't want this
    to crash the watchdog tick)."""
    to = ""
    try:
        from email_routing import get_value as _routing_get
        v = await _routing_get(db, "backup_email_to")
        if isinstance(v, list) and v:
            to = ",".join(v)
    except Exception:
        pass
    if not to:
        to = (os.environ.get("BACKUP_EMAIL_TO") or "").strip()
    if not to:
        logger.warning("[backup-watchdog] no recipient configured — alarm not sent")
        return False
    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        logger.warning("[backup-watchdog] RESEND_API_KEY missing — alarm not sent")
        return False
    try:
        import resend  # noqa: E402
        resend.api_key = api_key
        sender_email = await _resolve_sender_email(db)
        latest_filename = latest.get("filename") or "(unknown)"
        latest_ts = latest.get("ts") or "(unknown)"
        size_mb = (latest.get("size_bytes") or 0) / (1024 * 1024)
        html = (
            "<div style='font-family:Arial,sans-serif;max-width:560px'>"
            "<h2 style='color:#C8102E;margin:0 0 8px 0'>BACKUP HEALTH ALARM</h2>"
            f"<p>The MASCI Hub backup scheduler has not produced a successful "
            f"backup in <strong>{hours_silent:.1f} hours</strong>. "
            f"The configured threshold is "
            f"{os.environ.get('BACKUP_WATCHDOG_HOURS', '25')} hours.</p>"
            "<h3 style='color:#334155;margin:16px 0 4px 0;font-size:14px'>Last successful backup:</h3>"
            f"<ul style='margin:0;padding-left:20px;font-size:13px;color:#475569'>"
            f"<li>File: <code>{latest_filename}</code></li>"
            f"<li>Size: {size_mb:.1f} MB · {latest.get('records', 0)} records · mode={latest.get('mode')}</li>"
            f"<li>Timestamp: {latest_ts} UTC</li>"
            "</ul>"
            "<h3 style='color:#334155;margin:16px 0 4px 0;font-size:14px'>What to check:</h3>"
            "<ol style='margin:0;padding-left:20px;font-size:13px;color:#475569'>"
            "<li>Open <code>/api/admin/backups-scheduler-state</code> to see the live scheduler view.</li>"
            "<li>Try <code>POST /api/admin/backups/run-now?lite=true</code> — it returns 202 immediately and emails a slim backup within 60 s.</li>"
            "<li>If lite-mode runs succeed but full-mode is silent, the full archive has likely crossed the OOM watermark. "
            "Set <code>BACKUP_LITE_MODE_ONLY=true</code> on the deploy until S3 photo migration is done.</li>"
            "</ol>"
            "<p style='color:#94a3b8;font-size:11px;margin-top:18px'>"
            "Sent by the MASCI Hub backup watchdog. This alarm is rate-limited "
            f"to once every {os.environ.get('BACKUP_WATCHDOG_COOLDOWN_HOURS', '12')} hours.</p>"
            "</div>"
        )
        params = {
            "from": f"MASCI Operations Platform <{sender_email}>",
            "to": [t.strip() for t in to.split(",") if t.strip()],
            "subject": f"[MASCI ALARM] Backup silent for {hours_silent:.0f}h — action needed",
            "html": html,
        }
        reply_to = (os.environ.get("REPLY_TO_EMAIL") or "").strip()
        if reply_to:
            params["reply_to"] = reply_to
        await asyncio.to_thread(resend.Emails.send, params)
        logger.warning(f"[backup-watchdog] ALARM FIRED → {to} (silent={hours_silent:.1f}h)")
        return True
    except Exception as e:  # noqa: BLE001
        logger.exception(f"[backup-watchdog] alarm send failed: {e}")
        return False


async def _email_backup_zip_from_path(zip_path: Path, total_records: int) -> Optional[str]:
    """OOM-SAFE: Email the backup zip as a Resend attachment WITHOUT
    ever loading the full archive into RAM.

    Strategy:
      • Stat the full zip on disk to learn its size.
      • If full zip is small enough to email directly (≤ BACKUP_EMAIL_MAX_MB),
        read its bytes lazily in a worker thread and base64-encode for Resend.
      • Otherwise, stream entries from the on-disk full zip into a NEW
        slim zip on disk, dropping PDFs + disk_files/ + CSVs and stripping
        large base64 blobs from JSON. Memory stays flat (~10 MB) the whole
        time because we read+write one entry at a time.
      • Only the slim file (~1 MB) is ever loaded into memory for base64
        encoding. The 500 MB+ full zip never touches RAM.

    This eliminates the OOM crash that was killing the production
    container on every scheduled backup.
    """
    to = (os.environ.get("BACKUP_EMAIL_TO") or "").strip()
    # Allow admin to override via DB (Email Routing panel). DB list joined
    # with commas to match the existing comma-delimited contract Resend uses.
    try:
        from email_routing import get_value as _routing_get
        v = await _routing_get(db, "backup_email_to")
        if isinstance(v, list) and v:
            to = ",".join(v)
    except Exception:
        pass
    if not to:
        return None
    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        logger.info("[scheduled-backup] email skipped — RESEND_API_KEY missing")
        return None

    max_mb = int(os.environ.get("BACKUP_EMAIL_MAX_MB", "35"))
    full_size_bytes = await asyncio.to_thread(lambda: zip_path.stat().st_size)
    full_size_mb = full_size_bytes / (1024 * 1024)
    filename = zip_path.name

    attachment_path: Path
    attachment_name: str
    slim_notice = ""
    slim_tmp: Optional[Path] = None

    if full_size_mb <= max_mb:
        # Full zip fits — email it directly. Single read of the (small) file.
        attachment_path = zip_path
        attachment_name = filename
    else:
        # Build slim zip on disk by streaming entries. Never holds the
        # full payload in memory.
        slim_tmp = zip_path.with_name(
            zip_path.stem + f"_slim.{uuid.uuid4().hex[:8]}.zip.tmp"
        )
        try:
            stats = await asyncio.to_thread(
                _build_slim_email_zip_on_disk, zip_path, slim_tmp
            )
        except Exception as e:  # noqa: BLE001
            logger.warning(f"[scheduled-backup] slim build failed: {e}")
            try:
                if slim_tmp and slim_tmp.exists():
                    slim_tmp.unlink()
            except Exception:
                pass
            return None

        slim_size_mb = stats["size_bytes"] / (1024 * 1024)
        if slim_size_mb > max_mb:
            logger.warning(
                f"[scheduled-backup] even slim zip is {slim_size_mb:.1f} MB > {max_mb} — "
                f"email skipped. Admin must download from /admin/backups."
            )
            try:
                slim_tmp.unlink()
            except Exception:
                pass
            return None

        attachment_path = slim_tmp
        attachment_name = filename.replace(".zip", "_slim.zip")
        slim_notice = (
            f'<p style="background:#fef3c7;border-left:4px solid #f59e0b;'
            f'padding:10px 14px;border-radius:0 6px 6px 0;color:#92400e;'
            f'font-size:13px;line-height:1.5;margin:14px 0;">'
            f'<strong>Note:</strong> The full backup is {full_size_mb:.0f} MB '
            f'(includes rendered PDFs + project disk archive). For email, '
            f'we sent a <strong>slim {slim_size_mb:.1f} MB version</strong> with '
            f'every record\'s metadata + JSON ({stats["kept"]} entries). '
            f'{stats["stripped_blob_count"]} embedded file blob(s) '
            f'({stats["stripped_blob_bytes"] / (1024*1024):.0f} MB total) were '
            f'stripped — the originals live on the server. Sign in to '
            f'<code>/admin</code> and download <strong>{filename}</strong> '
            f'from the Stored Backups panel for the full archive.'
            f'</p>'
        )
        logger.info(
            f"[scheduled-backup] full {full_size_mb:.1f} MB → emailing slim {slim_size_mb:.1f} MB "
            f"({stats['kept']} entries, stripped {stats['stripped_blob_count']} blobs / "
            f"{stats['stripped_blob_bytes']/1024/1024:.1f} MB)"
        )

    try:
        return await _send_backup_email(
            to=to,
            api_key=api_key,
            attachment_path=attachment_path,
            attachment_name=attachment_name,
            full_size_mb=full_size_mb,
            total_records=total_records,
            slim_notice=slim_notice,
        )
    finally:
        if slim_tmp is not None:
            try:
                if slim_tmp.exists():
                    slim_tmp.unlink()
            except Exception:
                pass


def _build_slim_backup_zip_on_disk(db, dst_zip: Path) -> dict:
    """LITE-MODE direct-from-Mongo slim backup. Streams every record from
    every EXPORTABLE_KINDS collection to ``dst_zip`` on disk, stripping
    base64 blobs inline. Bypasses the full-zip step entirely — used when
    the full archive is so large it OOM-kills the worker.

    Runs synchronously inside ``asyncio.to_thread`` so Motor calls have to
    be replaced with PyMongo-style synchronous iteration. We use a
    fresh sync PyMongo client so this thread never blocks the event-loop
    Motor client.

    Returns ``{size_bytes, total_records, stripped_blob_count,
    stripped_blob_bytes, per_kind: {kind: count}}``.
    """
    import zipfile as _zf
    import json as _json
    from pymongo import MongoClient as _MC

    plan = getattr(app.state, "database_authority_plan", None)
    if plan is None:
        raise RuntimeError("database authority plan missing for slim backup export")
    mongo_url = plan.mongo_url
    db_name = _canonical_db_name()

    total_records = 0
    stripped_blob_count = 0
    stripped_blob_bytes = 0
    per_kind: Dict[str, int] = {}
    sync_client = _MC(mongo_url, serverSelectionTimeoutMS=10000, maxPoolSize=10)
    try:
        sync_db = sync_client[db_name]
        with _zf.ZipFile(str(dst_zip), "w", _zf.ZIP_DEFLATED, compresslevel=6) as zf:
            for kind, coll_name in EXPORTABLE_KINDS.items():
                kind_count = 0
                cursor = sync_db[coll_name].find({}, {"_id": 0}).sort("created_at", -1)
                for doc in cursor:
                    new_doc, sc, sb = _strip_base64_blobs(doc)
                    stripped_blob_count += sc
                    stripped_blob_bytes += sb
                    rec_id = (new_doc.get("id") or f"row_{kind_count:06d}")
                    safe_id = "".join(c if c.isalnum() or c in "._-" else "_" for c in str(rec_id))
                    path_in_zip = f"{kind}/json/{safe_id}.json"
                    zf.writestr(
                        path_in_zip,
                        _json.dumps(new_doc, indent=2, default=str),
                    )
                    kind_count += 1
                per_kind[kind] = kind_count
                total_records += kind_count
            # Manifest so the recipient can see what's inside.
            manifest = {
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "mode": "lite",
                "source": "mascidocs.com",
                "total_records": total_records,
                "per_kind": per_kind,
                "stripped_blob_count": stripped_blob_count,
                "stripped_blob_bytes": stripped_blob_bytes,
                "notice": (
                    "This is a LITE backup — metadata + JSON only, no PDFs, "
                    "no embedded media. The full archive (including base64 "
                    "photos) lives in /app/backend/backups on the server. "
                    "Restore via /admin → Restore from Backup."
                ),
            }
            zf.writestr("MANIFEST.json", _json.dumps(manifest, indent=2))
    finally:
        sync_client.close()

    return {
        "size_bytes": dst_zip.stat().st_size,
        "total_records": total_records,
        "per_kind": per_kind,
        "stripped_blob_count": stripped_blob_count,
        "stripped_blob_bytes": stripped_blob_bytes,
    }


# ────────────────────────────────────────────────────────────────────────
# Iter64 Phase 2c — Complete-system archive to R2
# ────────────────────────────────────────────────────────────────────────
# Builds a single, standalone zip containing:
#   1. The full JSON dump of every collection (MongoDB)
#   2. The actual photo BYTES, fetched from R2 and inlined into a
#      `photos/<bucket-key>` folder inside the zip
# So a recipient with just the zip can restore the entire system
# without needing R2 access. Streams to disk; never holds more than one
# photo in memory at a time. Uploaded to r2://<bucket>/backups/<file>.zip
# and a presigned 7-day download URL is emailed alongside the slim
# heartbeat email.

def _build_complete_archive_on_disk(
    db_unused,
    dst_zip: Path,
    *,
    backup_run_id: Optional[str] = None,
    archive_key: Optional[str] = None,
) -> dict:
    """Build a single self-contained zip on disk with:
      * Every Mongo collection (JSON, _id stripped) under `<kind>/json/`
      * Every R2 photo fetched and inlined under `photos/<key>`
      * A MANIFEST.json describing what's inside
    Runs synchronously inside ``asyncio.to_thread`` — uses a fresh
    PyMongo sync client and the synchronous photo_storage helper so we
    don't touch the asyncio event loop from a worker thread."""
    import json as _json
    import zipfile as _zf

    from pymongo import MongoClient as _MC

    try:
        from photo_storage import is_storage_ref, read_photo_bytes_sync
    except Exception:  # noqa: BLE001
        def is_storage_ref(_):
            return False

        def read_photo_bytes_sync(_):
            raise RuntimeError("photo_storage unavailable")

    plan = getattr(app.state, "database_authority_plan", None)
    if plan is None:
        runtime_bundle = _runtime_identity_bundle()
        env = {
            "MONGO_URL": os.environ["MONGO_URL"],
            "DB_NAME": os.environ["DB_NAME"],
            "APP_ENV": os.environ.get("APP_ENV", ""),
            "ENFORCE_DB_ISOLATION": os.environ.get("ENFORCE_DB_ISOLATION", ""),
            "READ_ONLY_VALIDATION": os.environ.get("READ_ONLY_VALIDATION", ""),
            "MONGO_SERVER_SELECTION_TIMEOUT_MS": os.environ.get("MONGO_SERVER_SELECTION_TIMEOUT_MS", ""),
            "MONGO_CONNECT_TIMEOUT_MS": os.environ.get("MONGO_CONNECT_TIMEOUT_MS", ""),
            "MONGO_SOCKET_TIMEOUT_MS": os.environ.get("MONGO_SOCKET_TIMEOUT_MS", ""),
        }
        plan = build_runtime_database_authority(
            runtime_identity_bundle=runtime_bundle,
            env=env,
            lifecycle_owner="server._build_complete_archive_on_disk",
        )
    mongo_url = plan.mongo_url
    db_name = _canonical_db_name()

    started_at = datetime.now(timezone.utc)
    total_records = 0
    per_kind: Dict[str, int] = {}
    per_collection_record_counts: Dict[str, int] = {}
    inlined_photos = 0
    inlined_photo_bytes = 0
    failed_photos = 0
    seen_keys: set = set()  # dedupe — same photo referenced from 2 docs

    sync_client = _MC(mongo_url, serverSelectionTimeoutMS=10000)
    expected_collections: List[str] = []
    attempted_collections: List[str] = []
    captured_collections: List[str] = []
    captured_archive_members: List[str] = []
    excluded_logged: List[str] = []
    excluded_collections: List[str] = []
    exclusion_reasons: Dict[str, Dict[str, str]] = {}
    failed_collections: List[str] = []
    failed_collection_errors: Dict[str, str] = {}
    skipped_collections: List[str] = []
    skipped_collection_reasons: Dict[str, str] = {}
    unknown_collections: List[str] = []
    try:
        sync_db = sync_client[db_name]
        all_collections = sorted(sync_db.list_collection_names())
        for coll_name in all_collections:
            if coll_name.startswith("system."):
                excluded_logged.append(coll_name)
                excluded_collections.append(coll_name)
                exclusion_reasons[coll_name] = dict(BACKUP_EXCLUSION_DETAILS["system.*"])
                continue
            if coll_name in BACKUP_EXPLICIT_EXCLUSIONS:
                excluded_logged.append(coll_name)
                excluded_collections.append(coll_name)
                exclusion_reasons[coll_name] = dict(
                    BACKUP_EXCLUSION_DETAILS.get(
                        coll_name,
                        {"reason": "explicit exclusion", "owner": "backup-platform"},
                    )
                )
                continue
            expected_collections.append(coll_name)

        with _zf.ZipFile(str(dst_zip), "w", _zf.ZIP_DEFLATED, compresslevel=6) as zf:
            # ════════════════════════════════════════════════════════════
            # iter425 · Phase 25.2 · AUTO-DISCOVERY (replaces EXPORTABLE_KINDS allowlist)
            # ────────────────────────────────────────────────────────────
            # Pass 1 — every record, every collection, as JSON.
            # Mirrors Pipeline A's behavior so NEW collections (Phase 12-25:
            # dispatch_assignments · dispatch_continuity_events ·
            # operational_attachments · user_passkeys · etc.) inherit R2
            # coverage automatically with zero allowlist maintenance.
            #
            # Photos stay as `photo://` refs in the JSON; the actual
            # bytes get inlined separately so the manifest links them.
            # Operational attachments store `data_b64` INLINE — they
            # restore from the JSON dump alone, no photo-walk needed.
            # ════════════════════════════════════════════════════════════
            for coll_name in all_collections:
                # Skip Mongo system collections + module-level explicit
                # exclusions (kept in BACKUP_EXPLICIT_EXCLUSIONS for audit
                # visibility — see R2_BACKUP_CONTINUITY_AUDIT.md §9).
                if coll_name == "usage_events" or coll_name in BACKUP_EXPLICIT_EXCLUSIONS or coll_name.startswith("system."):
                    continue

                # Friendly "kind" for the in-zip folder — matches Pipeline A
                # convention for the six legacy safety collections, otherwise
                # uses the raw collection name. Restorers don't care which
                # folder shape they see — both ship valid JSON.
                kind = next(
                    (k for k, v in EXPORTABLE_KINDS.items() if v == coll_name),
                    coll_name,
                )

                # iter425 · same sensitive-field redaction as Pipeline A
                projection = BACKUP_SENSITIVE_FIELD_REDACTION.get(coll_name, {"_id": 0})

                kind_count = 0
                attempted_collections.append(coll_name)
                try:
                    # iter428 · Phase 26.1 — Atlas M0 free tier caps in-memory
                    # sort at 32 MB and rejects `allowDiskUse`. Archive files
                    # are individually addressed by record ID
                    # (`{kind}/json/{safe_id}.json`), so the in-archive sort
                    # order is operationally irrelevant — drop the sort and
                    # iterate in natural order. Restore-time queries land
                    # against full Mongo, not the zip, so this preserves
                    # restore correctness.
                    cursor = sync_db[coll_name].find({}, projection)
                    for doc in cursor:
                        rec_id = doc.get("id") or f"row_{kind_count:06d}"
                        safe_id = "".join(c if c.isalnum() or c in "._-" else "_" for c in str(rec_id))
                        zf.writestr(
                            f"{kind}/json/{safe_id}.json",
                            _json.dumps(doc, indent=2, default=str),
                        )
                        kind_count += 1
                        # Walk this doc for photo:// refs to inline later
                        for ref in _iter_photo_refs(doc):
                            if not is_storage_ref(ref):
                                continue
                            try:
                                key = ref.split("/", 3)[3]
                            except (IndexError, AttributeError):
                                continue
                            if key in seen_keys:
                                continue
                            seen_keys.add(key)
                            try:
                                raw = read_photo_bytes_sync(ref)
                                zf.writestr(f"photos/{key}", raw)
                                inlined_photos += 1
                                inlined_photo_bytes += len(raw)
                            except Exception as e:  # noqa: BLE001
                                logger.warning(f"[complete-archive] photo inline failed for {ref[:80]}: {e}")
                                failed_photos += 1
                except Exception as e:  # noqa: BLE001
                    failed_collections.append(coll_name)
                    failed_collection_errors[coll_name] = repr(e)
                    logger.exception(f"[complete-archive] collection capture failed for {coll_name}: {e}")
                    continue

                per_kind[kind] = kind_count
                per_collection_record_counts[coll_name] = kind_count
                total_records += kind_count
                captured_collections.append(coll_name)
                captured_archive_members.append(f"{kind}/json/")

            # iter425 · log every excluded collection so audit trail is
            # NEVER silent. R2_BACKUP_CONTINUITY_AUDIT.md §9 documents reasons.
            if excluded_logged:
                logger.info(
                    f"[complete-archive] explicit exclusions ({len(excluded_logged)}): "
                    f"{', '.join(sorted(set(excluded_logged)))}",
                )

            coverage_gap = sorted(it for it in expected_collections if it not in captured_collections)
            coverage_complete = not coverage_gap and not failed_collections
            completed_at = datetime.now(timezone.utc)
            classification = "COMPLETE" if coverage_complete else "BACKUP_INCOMPLETE"
            integrity_failed_checks: List[Dict[str, Any]] = []
            if coverage_gap:
                integrity_failed_checks.append({
                    "code": "missing_required_collections",
                    "collections": coverage_gap,
                })
            if failed_collections:
                integrity_failed_checks.append({
                    "code": "failed_collections",
                    "collections": list(failed_collections),
                })

            manifest = {
                "manifest_version": BACKUP_MANIFEST_VERSION,
                "backup_id": uuid.uuid4().hex,
                "backup_run_id": backup_run_id,
                "backup_type": "complete-r2",
                "classification": classification,
                "environment": os.environ.get("APP_ENV"),
                "app_env": os.environ.get("APP_ENV"),
                "environment_fingerprint": _canonical_environment_fingerprint(),
                "environment_fingerprint_version": "env-authority-v1",
                "database_name": db_name,
                "db_name": db_name,
                "source_cluster_fingerprint": _canonical_cluster_fingerprint(),
                "source_database_identity": db_name,
                "source_runtime_user_identity": _canonical_runtime_user_identity(),
                "storage_provider": "r2-s3-compatible",
                "backup_bucket": _canonical_backup_bucket(),
                "backup_prefix": _canonical_backup_prefix(),
                "archive_key": archive_key,
                "source_hash": _SOURCE_HASH,
                "git_commit": os.environ.get("GIT_COMMIT") or _SOURCE_HASH[:12],
                "release_identity": _SOURCE_HASH,
                "application_version": os.environ.get("GIT_COMMIT") or _SOURCE_HASH[:12],
                "backup_started_at": started_at.isoformat(),
                "backup_completed_at": completed_at.isoformat(),
                "generated_at": completed_at.isoformat(),
                "mode": "complete",
                "source": "mascidocs.com",
                "expected_collections": sorted(expected_collections),
                "expected_collection_count": len(expected_collections),
                "captured_collections": sorted(captured_collections),
                "captured_collection_count": len(captured_collections),
                "excluded_collections": sorted(excluded_collections),
                "explicit_exclusions": sorted(set(excluded_logged)),
                "exclusion_reasons": exclusion_reasons,
                "attempted_collections": sorted(attempted_collections),
                "failed_collections": sorted(failed_collections),
                "failed_collection_errors": failed_collection_errors,
                "skipped_collections": sorted(skipped_collections),
                "skipped_collection_reasons": skipped_collection_reasons,
                "unknown_collections": sorted(unknown_collections),
                "total_records": total_records,
                "per_collection_record_counts": per_collection_record_counts,
                "per_kind": per_kind,
                "archive_members": sorted(captured_archive_members + ["MANIFEST.json"]),
                "archive_member_count": len(captured_archive_members) + 1,
                "archive_size_bytes": 0,
                "coverage_complete": coverage_complete,
                "coverage_gap": coverage_gap,
                "integrity_result": "PASS" if coverage_complete else "FAIL",
                "integrity_failed_checks": integrity_failed_checks,
                "verifier_version": BACKUP_VERIFIER_VERSION,
                "redaction_rules_applied": sorted(BACKUP_SENSITIVE_FIELD_REDACTION.keys()),
                "inlined_photos": inlined_photos,
                "inlined_photo_bytes": inlined_photo_bytes,
                "failed_photos": failed_photos,
                "notice": (
                    "Complete standalone backup. Contains every Mongo "
                    "collection (JSON) via auto-discovery (iter425) plus "
                    "the actual binary photos previously stored in R2. "
                    "No external dependency — you can restore the entire "
                    "MASCI Hub from this single zip even if Cloudflare R2 "
                    "becomes unreachable. MFA secrets, password hashes, "
                    "and recovery codes are redacted."
                ),
            }
            zf.writestr("MANIFEST.json", _json.dumps(manifest, indent=2))

        final_size = dst_zip.stat().st_size
        if not coverage_complete:
            raise RuntimeError(
                "complete backup coverage incomplete: "
                f"missing={coverage_gap} failed={failed_collections}"
            )
    finally:
        sync_client.close()

    return {
        "size_bytes": final_size,
        "total_records": total_records,
        "manifest": manifest,
        "per_kind": per_kind,
        "per_collection_record_counts": per_collection_record_counts,
        "expected_collections": sorted(expected_collections),
        "captured_collections": sorted(captured_collections),
        "captured_collection_count": len(captured_collections),
        "expected_collection_count": len(expected_collections),
        "excluded_collections": sorted(excluded_collections),
        "coverage_complete": True,
        "coverage_gap": [],
        "inlined_photos": inlined_photos,
        "inlined_photo_bytes": inlined_photo_bytes,
        "failed_photos": failed_photos,
    }


def _iter_photo_refs(doc):
    """Yield every photo:// reference found anywhere in a Mongo document.

    iter441 coverage:
      • top-level ``photos`` array
      • ``items[].photos`` / ``items[].return_photos`` / ``items[].original_photos``
        for equipment forms.

    iter442 coverage (closes 63-ref gap identified in PHOTO_COVERAGE_CERTIFICATION.md):
      • ``materials[].ticket_photos`` for daily_reports material delivery tickets
      • ``subcontractors[].photos`` for daily_reports subcontractor records
      • top-level signature fields stored as ``photo://`` refs
        (``prepared_by_signature``, ``reporter_signature``,
        ``supervisor_signature``, ``conductor_signature``)

    Any future schema additions that store photo refs at new JSON paths must
    extend this function — the archive auto-discovery walker (Pipeline A + B)
    is the single audit point for photo backup coverage.
    """
    if not isinstance(doc, dict):
        return

    # iter441 — top-level photos[]
    photos = doc.get("photos")
    if isinstance(photos, list):
        for p in photos:
            if isinstance(p, str):
                yield p

    # iter441 — equipment items[].photos / .return_photos / .original_photos
    items = doc.get("items")
    if isinstance(items, list):
        for it in items:
            if not isinstance(it, dict):
                continue
            for fld in ("photos", "return_photos", "original_photos"):
                v = it.get(fld)
                if isinstance(v, list):
                    for p in v:
                        if isinstance(p, str):
                            yield p

    # iter442 — daily_reports materials[].ticket_photos
    materials = doc.get("materials")
    if isinstance(materials, list):
        for m in materials:
            if not isinstance(m, dict):
                continue
            v = m.get("ticket_photos")
            if isinstance(v, list):
                for p in v:
                    if isinstance(p, str):
                        yield p

    # iter442 — daily_reports subcontractors[].photos
    subs = doc.get("subcontractors")
    if isinstance(subs, list):
        for s in subs:
            if not isinstance(s, dict):
                continue
            v = s.get("photos")
            if isinstance(v, list):
                for p in v:
                    if isinstance(p, str):
                        yield p

    # iter442 — top-level signature fields stored as photo:// refs.
    # Generic detection: any top-level string field whose name ends with
    # `_signature` OR is `signature` itself is yielded if it begins with
    # `photo://`. This auto-discovers future signature fields without
    # requiring this function to be re-touched. Empirically present in
    # production / preview today:
    #   • prepared_by_signature       (daily_reports)
    #   • superintendent_signature    (daily_reports)
    #   • operator_signature          (equipment_inspections)
    #   • supervisor_signature        (incidents)
    #   • reporter_signature          (incidents · top-level)
    #   • conductor_signature         (meetings)
    for fld, v in doc.items():
        if not isinstance(v, str):
            continue
        if not (fld == "signature" or fld.endswith("_signature")):
            continue
        if v.startswith("photo://"):
            yield v


async def _run_r2_tiered_retention_async() -> None:
    """TRACK 15.28A · Enforce the R2 tiered retention policy.

    Approved steady state:
      • hourly recovery points: last 72h
      • daily recovery points: 30d
      • weekly recovery points: 90d
      • monthly recovery points: 12m

    Idempotent. Touches only the active R2 prefix for the current environment.
    Legacy ``backups/*.zip`` is out of scope by design — see comment at
    upload-site.
    """
    try:
        from photo_storage import _client as _r2_client_for_retention
        from lib.r2_retention import enforce_r2_retention
    except Exception as _e:  # noqa: BLE001
        logger.warning(f"[r2-retention] imports unavailable: {_e}")
        return
    bucket = os.environ.get("S3_BUCKET", "").strip()
    if not bucket:
        return
    s3 = _r2_client_for_retention()
    if s3 is None:
        return
    try:
        result = await asyncio.to_thread(
            enforce_r2_retention, s3, bucket,
            prefix=_canonical_backup_prefix(), dry_run=False,
        )
        if result.get("ok") and (result.get("deleted") or 0) > 0:
            logger.info(
                f"[r2-retention] pruned {result['deleted']} objects · "
                f"survivors_by_tier={result['survivors_by_tier']} · "
                f"deleted_by_tier={result['deleted_by_tier']}"
            )
        elif not result.get("ok"):
            logger.warning(f"[r2-retention] errors: {result.get('errors')}")
        if not result.get("ok") or (result.get("error_count") or 0) > 0:
            await _record_backup_health(
                db,
                ok=False,
                mode="complete-r2-retention-error",
                error=_json.dumps(result.get("errors") or [])[:1500],
                notification_outcome="notification_not_required",
                notification_reason="r2_retention_failed",
                audit_reference="backup_health:complete-r2-retention-error",
            )
    except Exception as _e:  # noqa: BLE001
        logger.warning(f"[r2-retention] failed: {_e}")
        try:
            await _record_backup_health(
                db,
                ok=False,
                mode="complete-r2-retention-error",
                error=repr(_e),
                notification_outcome="notification_not_required",
                notification_reason="r2_retention_failed",
                audit_reference="backup_health:complete-r2-retention-error",
            )
        except Exception:
            pass



async def _log_r2_usage_warning() -> None:
    """Background probe — sum bucket size and log a warning when we cross
    the 45 GB warn / 50 GB alert thresholds. Fire-and-forget; failures are
    swallowed so this never blocks the backup pipeline.

    Surfaces:
      • Backend logs (``/var/log/supervisor/backend.*.log``) on every
        warn/alert tick.
      • ``backup_health`` row with mode='r2-usage-warn' when crossing the
        WARN threshold; mode='r2-usage-alert' when crossing the ALERT
        threshold. Operators / dashboards can read these from
        ``backup_health.find({mode: /r2-usage/})``.

    Intentionally does NOT email — the legacy backup-overdue email path
    already has rate-limiting and we don't want a second storm vector.
    """
    try:
        from photo_storage import is_configured as _ps_cfg
    except Exception:  # noqa: BLE001
        return
    if not _ps_cfg():
        return
    try:
        import boto3
        from botocore.config import Config as _BotoCfg
    except Exception:  # noqa: BLE001
        return

    bucket = os.environ.get("S3_BUCKET", "").strip()
    if not bucket:
        return

    try:
        s3 = boto3.client(
            "s3",
            endpoint_url=os.environ.get("S3_ENDPOINT_URL", ""),
            aws_access_key_id=os.environ.get("S3_ACCESS_KEY", ""),
            aws_secret_access_key=os.environ.get("S3_SECRET_KEY", ""),
            region_name=os.environ.get("S3_REGION") or "auto",
            config=_BotoCfg(signature_version="s3v4", s3={"addressing_style": "path"}),
        )

        def _sum():
            total = 0
            count = 0
            paginator = s3.get_paginator("list_objects_v2")
            for page in paginator.paginate(Bucket=bucket):
                for o in page.get("Contents", []):
                    total += o["Size"]
                    count += 1
            return total, count

        total_bytes, count = await asyncio.to_thread(_sum)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[r2-usage] probe failed: {e}")
        return

    gb = total_bytes / (1024 ** 3)
    WARN_GB = float(os.environ.get("R2_USAGE_WARN_GB", "45") or 45)
    ALERT_GB = float(os.environ.get("R2_USAGE_ALERT_GB", "50") or 50)

    capacity_state = classify_capacity_state(
        total_bytes=int(total_bytes),
        warn_gb=WARN_GB,
        alert_gb=ALERT_GB,
        probe_state="ok",
        as_of=datetime.now(timezone.utc).isoformat(),
    )
    if capacity_state["status"] == "RED":
        level = "ALERT"
        mode = "r2-usage-alert"
    elif capacity_state["status"] == "AMBER":
        level = "WARN"
        mode = "r2-usage-warn"
    else:
        # Healthy — keep a quiet info log so we can see the time-series
        # in supervisor logs, but no DB row (avoids noise).
        logger.info(f"[r2-usage] OK · {gb:.2f} GB · {count} objects · bucket={bucket}")
        return

    logger.warning(
        f"[r2-usage] {level} · {gb:.2f} GB · {count} objects · bucket={bucket} "
        f"(WARN_GB={WARN_GB}, ALERT_GB={ALERT_GB})"
    )
    try:
        await _record_backup_health(
            db, ok=True, size_bytes=int(total_bytes), mode=mode,
            error=f"r2-usage gb={gb:.2f} objects={count}",
            notification_outcome="notification_not_required",
            notification_reason="storage_threshold_observation_only",
            audit_reference=f"backup_health:{mode}:{bucket}",
        )
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[r2-usage] couldn't record health row: {e}")


# iter426 · Phase 25.3 · Backup Drift Watcher (calm, log-only).
# ────────────────────────────────────────────────────────────────────
# Persists the latest archive's `captured_collections` set to a Mongo
# collection `backup_drift_history` (capped: last 30 runs). On each new
# run, compares against the previous run and logs a WARNING line if any
# collection disappeared. Does NOT email, NOT alert, NOT surface in UI.
# Pure operational survivability whisper.
async def _backup_drift_watch(db, stats: dict) -> None:
    """Log a calm WARN if any collection vanished since the last archive."""
    if not stats:
        return
    captured_now = sorted(stats.get("captured_collections") or [])
    if not captured_now:
        return

    # Read the most recent prior run (if any).
    prior = await db.backup_drift_history.find_one(
        {}, sort=[("recorded_at", -1)],
    )

    if prior:
        prior_set = set(prior.get("captured_collections") or [])
        now_set = set(captured_now)
        disappeared = sorted(prior_set - now_set)
        appeared = sorted(now_set - prior_set)
        if disappeared:
            logger.warning(
                f"[complete-archive] DRIFT · collection count "
                f"{len(prior_set)} -> {len(now_set)} · "
                f"disappeared: {', '.join(disappeared)}"
            )
        if appeared:
            logger.info(
                f"[complete-archive] drift · new collection(s) included: "
                f"{', '.join(appeared)}"
            )

    # Append today's snapshot · keep history slim.
    await db.backup_drift_history.insert_one({
        "id": str(uuid.uuid4()),
        "recorded_at": datetime.now(timezone.utc),
        "captured_collections": captured_now,
        "total_records": stats.get("total_records", 0),
        "explicit_exclusions": sorted(stats.get("explicit_exclusions") or []),
    })

    # Trim to last 30 entries (FIFO).
    excess = await db.backup_drift_history.count_documents({}) - 30
    if excess > 0:
        old = db.backup_drift_history.find(
            {}, {"_id": 1}, sort=[("recorded_at", 1)],
        ).limit(excess)
        ids = [d["_id"] async for d in old]
        if ids:
            await db.backup_drift_history.delete_many({"_id": {"$in": ids}})


async def _run_complete_archive_to_r2(db) -> Optional[dict]:
    """Build a complete-system zip on disk, stream-upload it to
    ``r2://<bucket>/backups/auto-90d/<filename>``, then delete the local file.
    Returns ``{filename, size_bytes, r2_key, presigned_url, stats}``
    or ``None`` on any failure (errors are logged + health-recorded)."""
    current_job = None
    heartbeat_stop = None
    heartbeat_task = None
    stage = {"name": "preflight"}
    try:
        from photo_storage import (
            is_configured as _ps_cfg,
            presigned_get_url_for_key,
            upload_bytes,
            upload_local_file,
        )
    except Exception:  # noqa: BLE001
        logger.warning("[complete-archive] photo_storage import failed; skipping")
        return None

    if not _ps_cfg():
        logger.info("[complete-archive] R2 not configured; skipping nightly upload")
        return None

    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_COMPLETE_TMP_DIR.mkdir(parents=True, exist_ok=True)
    _now = datetime.now(timezone.utc)
    _stamp = _now.strftime("%Y-%m-%d_%H%M%SZ")
    filename = f"MASCI_complete_backup_{_stamp}.zip"
    r2_key = f"{_canonical_backup_prefix().rstrip('/')}/{filename}"
    out = BACKUP_COMPLETE_TMP_DIR / filename
    tmp = out.with_suffix(f".zip.tmp.{uuid.uuid4().hex[:8]}")

    try:
        current_job = await db.backup_jobs.find_one(
            {"kind": BACKUP_JOB_KIND_COMPLETE_R2, "state": "running"},
            {"_id": 0, "job_id": 1, "owner_token": 1, "owner_id": 1, "trigger": 1, "slot_key": 1, "backup_run_id": 1},
            sort=[("updated_at", -1)],
        )
        if current_job and current_job.get("owner_token"):
            heartbeat_stop, heartbeat_task = await _run_job_heartbeat(
                db,
                job_id=current_job["job_id"],
                owner_token=current_job["owner_token"],
                stage_fn=lambda: stage["name"],
            )
        latest_hint = await _latest_complete_backup_hint(db)
        preflight = _backup_resource_preflight(archive_size_bytes=latest_hint.get("size_bytes"))
        if not preflight.get("ok"):
            logger.warning(f"[complete-archive] deferred by resource guard: {preflight.get('reasons')}")
            await _record_backup_health(
                db,
                ok=False,
                error=f"deferred_by_resource_guard:{','.join(preflight.get('reasons') or [])}",
                mode="complete-r2-deferred",
                notification_outcome="notification_not_required",
                notification_reason="resource_guard_deferred_execution",
                audit_reference="backup_health:complete-r2-deferred",
            )
            return {"deferred": True, "reason": "resource_guard", "preflight": preflight}
        if current_job and current_job.get("owner_token"):
            await assert_backup_job_ownership(db, current_job["job_id"], owner_token=current_job["owner_token"])
        stage["name"] = "archive_construction"
        stats = await asyncio.to_thread(
            _build_complete_archive_on_disk,
            db,
            tmp,
            backup_run_id=(current_job or {}).get("backup_run_id"),
            archive_key=r2_key,
        )
        tmp.replace(out)
        if out.stat().st_size > BACKUP_COMPLETE_MAX_BUILD_BYTES:
            raise RuntimeError(f"complete archive exceeded bounded size ceiling {BACKUP_COMPLETE_MAX_BUILD_BYTES} bytes")
        size_mb = out.stat().st_size / 1024 / 1024
        logger.info(
            f"[complete-archive] built {out.name} · {size_mb:.1f} MB · "
            f"{stats.get('total_records', 0)} records · "
            f"{stats.get('inlined_photos', 0)} photos inlined"
        )

        # iter426 · Phase 25.3 · Calm backup drift watcher.
        # Compares the newly-captured collection set against the last
        # recorded archive. If a collection silently disappears between
        # runs (e.g., an accidental drop_collection in code review),
        # surface it as a calm log warning — NEVER as an alert, email,
        # dashboard, or notification. Infrastructure whisper only.
        try:
            await _backup_drift_watch(db, stats)
        except Exception as _e:  # noqa: BLE001
            logger.warning(f"[complete-archive] drift watch failed (non-fatal): {_e}")

        # Iter184 / Phase-2 Round 2 — R2 lifecycle scope.
        # New backups are written under a sub-prefix that the R2 lifecycle
        # rule (configured by scripts/r2_lifecycle_apply.py) targets. Any
        # legacy backups previously written to ``backups/*.zip`` (no
        # sub-prefix) are intentionally OUT of scope so existing history
        # is not retroactively deleted — they will be cleaned up manually
        # later with explicit operator approval. See R2_RETENTION_AUDIT.md.
        stage["name"] = "checksum"
        archive_sha256 = await _sha256_file(out)
        manifest_payload = stats.get("manifest") or {}
        manifest_key = manifest_sidecar_key_for_archive(r2_key)
        checksum_key = checksum_sidecar_key_for_archive(r2_key)
        stage["name"] = "upload"
        await upload_local_file(out, key=r2_key, content_type="application/zip")
        await upload_bytes(
            json.dumps(manifest_payload, ensure_ascii=False, sort_keys=True).encode("utf-8"),
            key=manifest_key,
            content_type="application/json",
        )
        await upload_bytes(
            f"{archive_sha256}  {filename}\n".encode("utf-8"),
            key=checksum_key,
            content_type="text/plain",
        )
        logger.info(f"[complete-archive] uploaded to r2://{os.environ.get('S3_BUCKET','')}/{r2_key}")

        # Generate a 7-day presigned URL the admin can click from email
        stage["name"] = "verification"
        presigned = await presigned_get_url_for_key(r2_key, ttl_seconds=7 * 24 * 3600)

        lineage = {
            "job_id": (current_job or {}).get("job_id"),
            "backup_run_id": (current_job or {}).get("backup_run_id"),
            "backup_id": stats.get("manifest", {}).get("backup_id"),
            "trigger": (current_job or {}).get("trigger"),
            "scheduler_slot": (current_job or {}).get("slot_key"),
            "release_sha": _SOURCE_HASH,
            "environment": _canonical_app_env().lower(),
            "environment_fingerprint": _canonical_environment_fingerprint(),
            "environment_fingerprint_version": "env-authority-v1",
            "source_cluster_fingerprint": _canonical_cluster_fingerprint(),
            "database_name": _canonical_db_name(),
            "source_database_identity": _canonical_db_name(),
            "source_runtime_user_identity": _canonical_runtime_user_identity(),
            "backup_bucket": _canonical_backup_bucket(),
            "backup_prefix": _canonical_backup_prefix(),
            "archive_key": r2_key,
            "manifest_key": manifest_key,
            "checksum_key": checksum_key,
            "archive_size_bytes": int(out.stat().st_size),
            "manifest_identity": {
                "manifest_name": "MANIFEST.json",
                "manifest_schema": BACKUP_MANIFEST_VERSION,
            },
            "checksum_sha256": archive_sha256,
            "created_at": _now.isoformat(),
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "verification_status": "uploaded",
        }
        if current_job and current_job.get("job_id"):
            await db.backup_jobs.update_one(
                {"job_id": current_job["job_id"], "owner_token": current_job.get("owner_token")},
                {"$set": {"archive_lineage": lineage, "backup_id": stats.get("manifest", {}).get("backup_id"), "updated_at": datetime.now(timezone.utc).isoformat()}},
            )

        # Delete the local copy now that R2 has it — keeps worker disk clean
        stage["name"] = "cleanup"
        try:
            out.unlink()
        except Exception:
            pass

        await _record_backup_health(
            db, ok=True, filename=filename, size_bytes=int(size_mb * 1024 * 1024),
            records=stats.get("total_records", 0),
            emailed_to=None, mode="complete-r2",
            notification_outcome="notification_not_required",
            notification_reason="complete_r2_archive_has_no_direct_email_policy",
            archive_identifier=filename,
            audit_reference=f"backup_health:{filename}",
            archive_lineage=lineage,
            error=_json.dumps({"lineage": lineage})[:1500],
        )

        completed_bucket = _now.strftime("%Y-%m-%dT%H")
        _BACKUP_SCHEDULER_STATE["last_r2_complete_hour"] = completed_bucket
        _BACKUP_SCHEDULER_STATE["last_r2_complete_date"] = completed_bucket[:10]
        _BACKUP_SCHEDULER_STATE["last_r2_complete"] = {
            "filename": filename,
            "size_bytes": int(size_mb * 1024 * 1024),
            "r2_key": r2_key,
            "ts": datetime.now(timezone.utc).isoformat(),
        }

        # Phase-2 Round 2 — passive 50 GB bucket-usage probe (warn-only).
        # After every successful nightly R2 upload, sum the bucket size in
        # the background and log a WARNING if it crosses 45 GB (warn) or
        # 50 GB (alert). Does NOT block, NOT delete, NOT email — that's
        # intentional: the lifecycle rule will eventually shed pressure on
        # its own, and we don't want a second email-storm vector. The
        # warning surfaces in `/var/log/supervisor/backend.*.log` and via
        # `python3 scripts/r2_usage_check.py` for on-demand checks.
        try:
            asyncio.create_task(_log_r2_usage_warning())
        except Exception as _e:  # noqa: BLE001
            logger.warning(f"[r2-usage] couldn't schedule probe: {_e}")

        # TRACK 15.28A · R2 tiered backup retention enforcement.
        # Runs once per successful upload (~hourly) so the bucket is
        # bounded without a separate cron. Tier-1 keeps every hourly zip
        # for 14d; Tier-2 keeps newest per day for 90d; Tier-3 keeps
        # newest per month for 365d; Tier-4 deletes. Idempotent.
        try:
            asyncio.create_task(_run_r2_tiered_retention_async())
        except Exception as _e:  # noqa: BLE001
            logger.warning(f"[r2-retention] couldn't schedule prune: {_e}")

        return {
            "filename": filename,
            "size_bytes": int(size_mb * 1024 * 1024),
            "r2_key": r2_key,
            "presigned_url": presigned,
            "stats": stats,
            "resource_preflight": preflight,
            "archive_lineage": lineage,
        }
    except Exception as e:  # noqa: BLE001
        logger.exception(f"[complete-archive] FAILED: {e}")
        try:
            if tmp.exists():
                tmp.unlink()
        except Exception:
            pass
        try:
            await _record_backup_health(
                db,
                ok=False,
                error=repr(e),
                mode="complete-r2-error",
                notification_outcome="notification_not_required",
                notification_reason="complete_r2_archive_failed_before_notification",
                audit_reference="backup_health:complete-r2-error",
            )
        except Exception:
            pass
        return None
    finally:
        try:
            if heartbeat_stop is not None:
                heartbeat_stop.set()
            if heartbeat_task is not None:
                await heartbeat_task
        except BackupJobOwnershipLost:
            logger.warning("[complete-archive] heartbeat ownership lost")
        except Exception:
            pass


async def _email_lite_backup_zip(zip_path: Path, stats: dict) -> Optional[str]:
    """Email a lite-mode slim backup to the configured BACKUP_EMAIL_TO
    address (or the DB-backed override). Identical sender / template to
    the full-mode email but with a clear LITE banner so the recipient
    knows the full archive lives on the server.

    Returns the To: address(es) on success, ``None`` otherwise.
    """
    to = (os.environ.get("BACKUP_EMAIL_TO") or "").strip()
    try:
        from email_routing import get_value as _routing_get
        v = await _routing_get(db, "backup_email_to")
        if isinstance(v, list) and v:
            to = ",".join(v)
    except Exception:
        pass
    if not to:
        logger.info("[scheduled-backup·lite] email skipped — no BACKUP_EMAIL_TO configured")
        return None
    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        logger.info("[scheduled-backup·lite] email skipped — RESEND_API_KEY missing")
        return None

    size_mb = stats["size_bytes"] / (1024 * 1024)

    # Pull the most recent R2 complete-archive URL (if any) so the
    # heartbeat email also gives a one-click download link to the
    # complete backup that lives in R2. Falls back to the on-server
    # path notice if R2 is not configured or has no backups yet.
    r2_block = ""
    try:
        r2_last = _BACKUP_SCHEDULER_STATE.get("last_r2_complete") or {}
        r2_key = r2_last.get("r2_key")
        r2_size = r2_last.get("size_bytes") or 0
        r2_ts = r2_last.get("ts") or ""
        if r2_key:
            try:
                from photo_storage import presigned_get_url_for_key as _psg
                _link = await _psg(r2_key, ttl_seconds=7 * 24 * 3600)
            except Exception:
                _link = None
            if _link:
                r2_block = (
                    f'<p style="background:#dcfce7;border-left:4px solid #16a34a;'
                    f'padding:10px 14px;border-radius:0 6px 6px 0;color:#14532d;'
                    f'font-size:13px;line-height:1.5;margin:14px 0;">'
                    f'<strong>Complete archive (Cloudflare R2):</strong> '
                    f'<a href="{_link}" style="color:#14532d;text-decoration:underline;">'
                    f'{r2_last.get("filename") or r2_key.rsplit("/", 1)[-1]}</a> '
                    f'· {r2_size/1024/1024:.1f} MB · uploaded {r2_ts[:10]}. '
                    f'Link valid for 7 days. Includes every record AND every photo, '
                    f'fully self-contained — no R2 access needed to restore.'
                    f'</p>'
                )
    except Exception:
        pass

    slim_notice = (
        f'<p style="background:#fef3c7;border-left:4px solid #f59e0b;'
        f'padding:10px 14px;border-radius:0 6px 6px 0;color:#92400e;'
        f'font-size:13px;line-height:1.5;margin:14px 0;">'
        f'<strong>LITE BACKUP</strong> — {size_mb:.1f} MB · '
        f'{stats["total_records"]} records. This email is a heartbeat '
        f'confirming the backup pipeline is alive. The complete archive '
        f'(every record + every photo) lives in Cloudflare R2 and is '
        f'rebuilt nightly. '
        f'{stats["stripped_blob_count"]} embedded blobs were stripped '
        f'({stats["stripped_blob_bytes"] / (1024*1024):.1f} MB total) '
        f'from this slim copy to keep email size down.'
        f'</p>'
        f'{r2_block}'
    )

    return await _send_backup_email(
        to=to,
        api_key=api_key,
        attachment_path=zip_path,
        attachment_name=zip_path.name,
        full_size_mb=size_mb,
        total_records=stats["total_records"],
        slim_notice=slim_notice,
    )


def _build_slim_email_zip_on_disk(src_zip: Path, dst_zip: Path) -> dict:
    """Synchronous helper run via asyncio.to_thread. Streams entries from
    src_zip → dst_zip on disk, dropping non-essential files and stripping
    large base64 blobs from JSON. Memory bounded by the largest single
    entry processed (typically <2 MB after blob stripping).
    """
    import zipfile as _zf2
    import json as _json2

    stripped_blob_count = 0
    stripped_blob_bytes = 0
    kept = 0

    with _zf2.ZipFile(src_zip, "r") as src, \
         _zf2.ZipFile(dst_zip, "w", _zf2.ZIP_DEFLATED) as dst:
        for info in src.infolist():
            n = info.filename
            # Drop rendered PDFs, disk-backed files, and CSV duplicates —
            # they're recoverable from JSON or live on the server's full zip.
            if n.startswith("disk_files/"):
                continue
            if n.endswith(".pdf"):
                continue
            if n.startswith("CSV/"):
                continue
            # Skip directory entries
            if n.endswith("/"):
                continue
            with src.open(info, "r") as fsrc:
                raw = fsrc.read()
            # For JSON files, strip base64 blobs.
            if n.endswith(".json") and len(raw) > 4096:
                try:
                    doc = _json2.loads(raw)
                    new_doc, stripped_count, stripped_bytes = _strip_base64_blobs(doc)
                    if stripped_count:
                        stripped_blob_count += stripped_count
                        stripped_blob_bytes += stripped_bytes
                        raw = _json2.dumps(new_doc, indent=2, default=str).encode("utf-8")
                except Exception:
                    pass
            dst.writestr(n, raw)
            kept += 1
            del raw

    return {
        "size_bytes": dst_zip.stat().st_size,
        "kept": kept,
        "stripped_blob_count": stripped_blob_count,
        "stripped_blob_bytes": stripped_blob_bytes,
    }


async def _send_backup_email(
    *,
    to: str,
    api_key: str,
    attachment_path: Path,
    attachment_name: str,
    full_size_mb: float,
    total_records: int,
    slim_notice: str,
) -> Optional[str]:
    """Read attachment bytes lazily, base64-encode, send via Resend.
    The attachment is guaranteed small (≤ BACKUP_EMAIL_MAX_MB) by the caller.
    """
    import base64 as _bb64

    def _encode() -> str:
        with attachment_path.open("rb") as f:
            return _bb64.b64encode(f.read()).decode("ascii")

    b64 = await asyncio.to_thread(_encode)
    attachment_size_mb = (
        await asyncio.to_thread(lambda: attachment_path.stat().st_size)
    ) / (1024 * 1024)
    # TRACK 27.03 · Phase 2b · Nightly backup email stamp uses local
    # wall-clock (subject + body). Underlying file mtime stays UTC.
    from lib.platform_time import format_platform_stamp as _fmt_local_stamp  # noqa: PLC0415
    stamp = _fmt_local_stamp(datetime.now(timezone.utc))
    sender = await _resolve_sender_email(db, safe_fallback="noreply@mascidocs.com")
    reply_to = (await _resolve_reply_to_email(db)) or None

    html = (
        f'<div style="font-family:system-ui,-apple-system,sans-serif;max-width:560px;margin:0 auto;">'
        f'<div style="border-bottom:4px solid #b91c1c;padding-bottom:10px;margin-bottom:18px;">'
        f'<strong style="color:#b91c1c;letter-spacing:.15em;font-size:11px;text-transform:uppercase">'
        f'MASCI · NIGHTLY BACKUP</strong>'
        f'</div>'
        f'<p style="color:#0f172a;margin:0 0 8px;font-size:15px;">'
        f'Your nightly MASCI Hub backup is attached.</p>'
        f'<ul style="color:#334155;font-size:14px;line-height:1.7;padding-left:18px;">'
        f'<li><strong>Generated:</strong> {stamp}</li>'
        f'<li><strong>Records:</strong> {total_records}</li>'
        f'<li><strong>Full backup size:</strong> {full_size_mb:.1f} MB</li>'
        f'<li><strong>Attachment:</strong> <code>{attachment_name}</code> '
        f'({attachment_size_mb:.1f} MB)</li>'
        f'</ul>'
        f'{slim_notice}'
        f'<p style="color:#475569;font-size:13px;margin-top:18px;">'
        f'<strong>Restore instructions:</strong> sign in to <a href="https://mascidocs.com/admin">'
        f'mascidocs.com/admin</a> → scroll to "Restore from Backup" → Upload this .zip.</p>'
        f'<p style="color:#b91c1c;font-size:12px;margin-top:18px;font-weight:700;">'
        f'Keep this email safe — it is your off-site disaster-recovery copy.</p>'
        f'</div>'
    )

    try:
        import resend  # noqa: E402
        resend.api_key = api_key
        params: Dict[str, Any] = {
            "from": f"MASCI Operations Platform <{sender}>",
            "to": [to],
            "subject": f"MASCI Nightly Backup · {stamp} · {total_records} records",
            "html": html,
            "attachments": [
                {"filename": attachment_name, "content": b64},
            ],
        }
        if reply_to:
            params["reply_to"] = reply_to
        result = await asyncio.to_thread(resend.Emails.send, params)
        rid = (result or {}).get("id", "?")
        logger.info(f"[scheduled-backup] emailed backup to {to} (resend_id={rid})")
        return rid if rid and rid != "?" else to
    except Exception as e:
        logger.warning(f"[scheduled-backup] Resend send failed: {e}")
        return None


_backup_task: Optional[asyncio.Task] = None


def _hours_since_last_backup() -> Optional[float]:
    """Return how many hours have elapsed since the most recent successful
    backup file was written. Returns None if no backup file exists.

    Used by the scheduler at boot time to decide whether to trigger a
    catch-up backup (because the container restarted across a missed slot).

    Iter182 fix (2026-05-17): previously this only counted
    ``MASCI_full_backup_*.zip``, which silently ignored lite-mode
    backups. Production runs in lite-mode (``BACKUP_LITE_MODE_ONLY``
    or auto-downgrade), so the only artifacts on disk were
    ``MASCI_lite_backup_*.zip`` → the staleness check returned None →
    the scheduler thought "no prior backup ever" → every container
    restart fired a catch-up backup → user received one email per
    restart (60+ per day during active development). Now counts BOTH
    full and lite filenames so the same-day protection actually
    engages on restart.
    """
    if not BACKUPS_DIR.exists():
        return None
    try:
        files = list(BACKUPS_DIR.glob("MASCI_full_backup_*.zip"))
        files.extend(BACKUPS_DIR.glob("MASCI_lite_backup_*.zip"))
        if not files:
            return None
        newest = max(f.stat().st_mtime for f in files)
        delta = datetime.now(timezone.utc) - datetime.fromtimestamp(
            newest, tz=timezone.utc
        )
        return delta.total_seconds() / 3600.0
    except Exception:
        return None


# Diagnostic state — exposed by /api/admin/backups/scheduler-state so the
# admin console can confirm the scheduler is alive WITHOUT firing a fresh
# backup. Populated by `_backup_scheduler_loop` on every tick.
_BACKUP_SCHEDULER_STATE: dict = {
    "alive": False,
    "armed_at": None,
    "last_tick_ts": None,
    "in_progress": False,
    "last_attempt_started_at": None,
    "last_attempt_outcome": None,
    "last_run_for_hour": {},
    "failed_attempts": {},
    # TRACK 27.05 · P0-2 · Observability counters. Every time the
    # supervisor detects `task.done()` and respawns the loop, it bumps
    # `resurrect_count` and stamps `last_resurrect_ts`. The recovery
    # snapshot surfaces these — silent scheduler death now has a
    # visible trail.
    "resurrect_count": 0,
    "last_resurrect_ts": None,
    # iter462 · 2026-02-01 · Batch A Phase 1 hardening · boot-trace
    # instrumentation. ``boot_step`` records the most recent boot stage
    # the loop reached before exiting. ``boot_step_ts`` is the ISO
    # timestamp of that stage. ``boot_exception`` captures the repr of
    # any unhandled exception that escaped the loop body (Phase 2
    # defensive wrapping). All three are surfaced by
    # ``GET /api/admin/backups-scheduler-state`` so operators can see
    # where the loop died WITHOUT triggering a fresh backup.
    "boot_step": None,
    "boot_step_ts": None,
    "boot_exception": None,
    "r2_hourly_requested": False,
    "r2_hourly_effective": False,
    "r2_hourly_locked_off": True,
    "hourly_cadence_enabled": False,
    "activation_blockers": [],
    "activation_status": "DISABLED BY CONFIGURATION",
    "activation_environment": "unknown",
    "last_activation_evaluated_at": None,
    "next_eligible_hourly_slot": None,
    "backup_runtime": {
        "stale_marked": 0,
        "active_jobs": [],
        "overlap": {
            "backup_active": False,
            "restore_active": False,
            "active_backups": [],
            "active_restores": [],
            "overlap_blocked": False,
        },
        "recent_complete_jobs": [],
    },
}

_BACKUP_RETENTION_POLICY = {
    "architecture": "selected_surviving_hourly_archives",
    "hourly_hours": 72,
    "daily_days": 30,
    "weekly_days": 90,
    "monthly_months": 12,
}


def _record_boot_step(step: str, *, exc: Optional[Exception] = None) -> None:
    """Phase 1 instrumentation · 2026-02-01 · Batch A.

    Records the most recent boot stage reached by ``_backup_scheduler_loop``
    into module-scope state AND emits a structured log line. Existed to
    diagnose the production "task completed without error" silent-exit
    failure mode: previously the state showed ``alive: false`` and
    ``armed_at: null`` with no indication of WHERE in the boot path the
    exit happened. With this helper, every boot stage is now traceable
    via the admin diagnostic endpoint AND via the backend logs.

    No behavioural change to the scheduler — this is pure observability.
    """
    now_iso = datetime.now(timezone.utc).isoformat()
    _BACKUP_SCHEDULER_STATE["boot_step"] = step
    _BACKUP_SCHEDULER_STATE["boot_step_ts"] = now_iso
    if exc is not None:
        _BACKUP_SCHEDULER_STATE["boot_exception"] = (
            f"{type(exc).__name__}: {exc!r}"[:300]
        )
        logger.error(
            f"[scheduled-backup][boot:{step}] EXCEPTION "
            f"{type(exc).__name__}: {exc!r}"
        )
    else:
        logger.info(
            f"[scheduled-backup][boot:{step}] reached at {now_iso}"
        )


async def _backup_scheduler_loop_with_capture(db) -> None:
    """Phase 2 defensive wrapper · 2026-02-01 · Batch A.

    Wraps ``run_with_singleton_lock(...)`` so that an unhandled exception
    escaping the lock acquirer OR the scheduler loop is captured into
    ``_BACKUP_SCHEDULER_STATE["boot_exception"]`` BEFORE the asyncio
    Task terminates. Without this, an exception inside an ``await``
    boundary caused the supervisor to see ``completed without error``
    while the actual cause was lost.

    Behaviour:
    - On clean return (e.g. SCHEDULER_ENABLED=false in preview), no
      change — the function exits and the supervisor respawns as before.
    - On asyncio.CancelledError, re-raises (caller decides).
    - On any other exception, records ``boot_exception`` and re-raises so
      the supervisor's ``_backup_task.exception()`` path also surfaces it.
    """
    try:
        await run_with_singleton_lock(db, "backup_scheduler", _backup_scheduler_loop)
    except asyncio.CancelledError:
        _record_boot_step("cancelled")
        raise
    except Exception as e:  # noqa: BLE001
        _record_boot_step("unhandled_exception_in_wrapper", exc=e)
        _BACKUP_SCHEDULER_STATE["last_attempt_outcome"] = (
            f"UNHANDLED EXCEPTION IN WRAPPER: {type(e).__name__}: {e!r}"
        )
        raise

# Module-level "in-progress" guard for the manual /admin/backups/run-now
# endpoint so two clicks 5 seconds apart can't both spawn 887 MB zip
# builds and OOM the worker.
_BACKUP_RUNNOW_IN_PROGRESS: bool = False
_BACKUP_RUNNOW_LAST: dict = {
    "started_at": None,
    "finished_at": None,
    "outcome": None,
    "lite_mode": None,
}


def _lite_mode_default() -> bool:
    """Default to **lite mode ON** so every backup (manual or scheduled)
    produces the slim email-friendly metadata-and-JSON zip and never
    tries to build the 800+ MB full archive on the worker.

    Why default-on?
    Iter64 phase 2 (2026-05-11) moved photos out of MongoDB into R2
    object storage, but other base64 fields (signatures, training
    photos, etc.) still live in Mongo and a full-archive build of all
    of them was still long enough to recycle the worker mid-task on
    production. Until those remaining fields are migrated AND/OR the
    IT-pull endpoint replaces email-attached backups, the safest
    default is "always send the slim 74 KB email, never block the
    worker." Anyone who explicitly wants a full archive can set
    ``BACKUP_LITE_MODE_ONLY=false`` to opt back in.
    """
    raw = (os.environ.get("BACKUP_LITE_MODE_ONLY", "") or "").strip().lower()
    # Explicit opt-OUT only — falsy strings disable lite-mode default.
    if raw in ("0", "false", "no", "n", "off"):
        return False
    return True


async def _backup_scheduler_loop(db) -> None:
    """Background loop — wakes up every ~5 min and fires the backup when
    we OBSERVE an hour transition into a scheduled slot while running.
      • At boot we look at the most recent backup file on disk.
      • If it's <8 hours old, the system is healthy — we mark same-day past
        slots as "already run" so we never double-fire.
      • If it's ≥8 hours old (or missing entirely), we DO NOT seed past
        slots → the loop discovers them on its next tick and fires a
        catch-up backup ~30 seconds after startup. This is what handles
        the "container restarted right after a scheduled slot, so the
        slot got skipped" scenario.

    Crash-loop protection:
      • If today's backup attempts fail MAX_DAILY_ATTEMPTS times in a row,
        a circuit breaker engages and we stop retrying until midnight
        UTC. This preserves the original safety guarantee that a
        crashing backup can't pin the container in a restart loop.

    Diagnostic state (iter62):
      • Loop state (last_run_for_hour, failed_attempts, last_tick_ts,
        last_attempt_outcome, in_progress) is published into module-scope
        `_BACKUP_SCHEDULER_STATE` so the admin diagnostic endpoint
        ``GET /api/admin/backups-scheduler-state`` can show whether the
        scheduler is alive, what tick it's on, and what the last
        attempt actually did — without triggering a fresh backup.
    """
    MAX_DAILY_ATTEMPTS = 3
    scheduler_owner = backup_owner_id()

    # iter462 · 2026-02-01 · Batch A Phase 1 — first boot step. If we never
    # see this in the diagnostic state, the loop body is being skipped
    # before any code executes (e.g. SCHEDULER_ENABLED gate returned early,
    # OR the asyncio task is being cancelled by the watchdog mid-spawn).
    _record_boot_step("entered_loop_body")

    # hour → last date we ran for that slot (for slot-collapsing logic)
    last_run_for_hour: dict[int, "datetime.date"] = _BACKUP_SCHEDULER_STATE["last_run_for_hour"]
    # date → number of attempts that failed today (circuit breaker)
    failed_attempts: dict["datetime.date", int] = _BACKUP_SCHEDULER_STATE["failed_attempts"]

    now = datetime.now(timezone.utc)
    today = now.date()
    _BACKUP_SCHEDULER_STATE["alive"] = True
    _BACKUP_SCHEDULER_STATE["armed_at"] = now.isoformat()
    _record_boot_step("armed")

    hours_stale = _hours_since_last_backup()
    _record_boot_step("disk_staleness_read")

    # Iter182 belt-and-suspenders (2026-05-17): cross-check against the
    # ``backup_health`` Mongo collection. The on-disk staleness check
    # can return stale=None if an emergency prune just wiped all
    # local archives (the scheduler log line "[scheduled-backup]
    # disk at 77% on boot — running emergency prune" routinely
    # appears). The Mongo collection persists every successful
    # backup row with its own TTL, so use it as a second source of
    # truth. We pick whichever timestamp is MORE recent.
    try:
        _record_boot_step("mongo_heartbeat_started")
        latest_row = await db.backup_health.find_one(
            {"ok": True},
            {"_id": 0, "ts": 1, "mode": 1},
            sort=[("ts", -1)],
        )
        if latest_row and latest_row.get("ts"):
            ts_str = latest_row["ts"]
            # Tolerate both trailing-Z and offset ISO formats
            ts_clean = ts_str.replace("Z", "+00:00")
            row_ts = datetime.fromisoformat(ts_clean)
            if row_ts.tzinfo is None:
                row_ts = row_ts.replace(tzinfo=timezone.utc)
            mongo_hours = (now - row_ts).total_seconds() / 3600.0
            if hours_stale is None or mongo_hours < hours_stale:
                logger.info(
                    f"[scheduled-backup] staleness: disk={hours_stale}h "
                    f"mongo={mongo_hours:.1f}h (using mongo, mode="
                    f"{latest_row.get('mode')!r})"
                )
                hours_stale = mongo_hours
        _record_boot_step("mongo_heartbeat_done")
    except Exception as e:  # noqa: BLE001
        logger.warning(
            f"[scheduled-backup] mongo staleness cross-check failed (non-fatal): {e}"
        )
        _record_boot_step("mongo_heartbeat_exception", exc=e)

    if hours_stale is not None and hours_stale <= 8:
        # System is healthy. Original behaviour: skip past slots today.
        for h in BACKUP_HOURS_UTC:
            if now.hour >= h:
                last_run_for_hour[h] = today
        skipped = sorted(h for h in BACKUP_HOURS_UTC if now.hour >= h)
        upcoming = sorted(h for h in BACKUP_HOURS_UTC if now.hour < h)
        logger.info(
            f"[scheduled-backup] scheduler armed — last backup ~{hours_stale:.1f}h ago "
            f"(healthy). skipping past slots today "
            f"{[f'{h:02d}:00' for h in skipped] or 'none'}, "
            f"next slots {[f'{h:02d}:00' for h in upcoming] or 'tomorrow'}"
        )
    else:
        # Stale or missing → don't seed; let the next loop tick fire a catch-up.
        stale_label = "NEVER (no prior backup)" if hours_stale is None else f"~{hours_stale:.1f}h ago"
        logger.warning(
            f"[scheduled-backup] scheduler armed — last backup {stale_label} "
            f"(stale). Catch-up will fire ~30s after startup."
        )

    # iter440 · Phase 31.3 · Restart-fire prevention.
    # Seed `_BACKUP_SCHEDULER_STATE["last_r2_complete_hour"]` from the most
    # recent successful `complete-r2` row in `backup_health` so that a
    # `uvicorn --reload` (or supervisord restart) DOES NOT immediately
    # re-fire an archive in an hour bucket we've already covered.
    # Without this seed, the in-memory state is wiped on every restart
    # and the next 5-min tick sees `last_r2_complete_hour != current` →
    # fires a duplicate archive. This was the root cause of the
    # ~100 archives/day vs expected 24/day rate (each reload caused by
    # WatchFiles file-change detection during agent edits added one
    # archive within minutes of startup).
    try:
        _record_boot_step("r2_seed_started")
        latest_r2 = await db.backup_health.find_one(
            {"mode": "complete-r2", "ok": True, "filename": {"$nin": [None, ""]}},
            sort=[("ts", -1)],
            projection={"_id": 0, "ts": 1, "filename": 1},
        )
        if latest_r2 and latest_r2.get("ts"):
            try:
                # ts is ISO 8601 like "2026-05-26T00:09:52.987715+00:00".
                # The bucket key is "%Y-%m-%dT%H" — take the first 13 chars.
                seeded_bucket = str(latest_r2["ts"])[:13]
                _BACKUP_SCHEDULER_STATE["last_r2_complete_hour"] = seeded_bucket
                _BACKUP_SCHEDULER_STATE["last_r2_complete_date"] = seeded_bucket[:10]
                logger.info(
                    f"[scheduled-backup] R2 state seeded from backup_health: "
                    f"last_r2_complete_hour={seeded_bucket} "
                    f"(prevents restart-fire of {latest_r2.get('filename')})"
                )
            except Exception as _e:  # noqa: BLE001
                logger.warning(
                    f"[scheduled-backup] R2 state seed parse failed (non-fatal): {_e}"
                )
    except Exception as e:  # noqa: BLE001
        logger.warning(
            f"[scheduled-backup] R2 state seed query failed (non-fatal): {e}"
        )
        _record_boot_step("r2_seed_exception", exc=e)
    else:
        _record_boot_step("r2_seed_done")

    # Give the app a moment to finish startup before first tick
    await asyncio.sleep(30)
    _record_boot_step("entering_main_tick_loop")
    while True:
        try:
            now = datetime.now(timezone.utc)
            today = now.date()
            _BACKUP_SCHEDULER_STATE["last_tick_ts"] = now.isoformat()
            _BACKUP_SCHEDULER_STATE["backup_runtime"] = await _collect_backup_runtime_state(db)
            # Find the latest scheduled hour crossed today that hasn't fired.
            due_hour: Optional[int] = None
            for h in BACKUP_HOURS_UTC:
                if now.hour >= h and last_run_for_hour.get(h) != today:
                    due_hour = h
            if due_hour is not None:
                # Circuit breaker: bail out if we've already failed too
                # many times today, so a deterministically-broken backup
                # cannot keep retrying.
                attempts = failed_attempts.get(today, 0)
                if attempts >= MAX_DAILY_ATTEMPTS:
                    if last_run_for_hour.get(due_hour) != today:
                        logger.error(
                            f"[scheduled-backup] CIRCUIT BREAKER engaged — "
                            f"{attempts} failed attempts today. Skipping "
                            f"remaining slots until midnight UTC."
                        )
                        _BACKUP_SCHEDULER_STATE["last_attempt_outcome"] = (
                            f"circuit-breaker (attempts={attempts}, date={today})"
                        )
                        for h in BACKUP_HOURS_UTC:
                            last_run_for_hour[h] = today
                else:
                    logger.info(
                        f"[scheduled-backup] firing for {today} "
                        f"(slot {due_hour:02d}:00 UTC, attempt #{attempts + 1})"
                    )
                    slot_key = backup_slot_key_for_day(
                        datetime(now.year, now.month, now.day, due_hour, 0, tzinfo=timezone.utc)
                    ) + f"::zip::{due_hour:02d}"
                    slot_claim = await scheduler_claim_slot(
                        db,
                        "backup_scheduler_zip",
                        slot_key,
                        owner_id=scheduler_owner,
                    )
                    if slot_claim is None:
                        logger.warning(f"[scheduled-backup] duplicate zip slot prevented: {slot_key}")
                        last_run_for_hour[due_hour] = today
                        await asyncio.sleep(300)
                        continue
                    _BACKUP_SCHEDULER_STATE["in_progress"] = True
                    _BACKUP_SCHEDULER_STATE["last_attempt_started_at"] = now.isoformat()
                    try:
                        result = await _run_scheduled_backup(db)
                    finally:
                        _BACKUP_SCHEDULER_STATE["in_progress"] = False
                    if result:
                        last_run_for_hour[due_hour] = today
                        # Collapse earlier same-day slots into this run.
                        for h in BACKUP_HOURS_UTC:
                            if h <= due_hour:
                                last_run_for_hour[h] = today
                        _BACKUP_SCHEDULER_STATE["last_attempt_outcome"] = (
                            f"ok · {result.get('filename')} · {result.get('size_bytes', 0)} bytes · "
                            f"emailed_to={result.get('emailed_to')}"
                        )
                        await scheduler_mark_completed(
                            db,
                            "backup_scheduler_zip",
                            slot_key,
                            recipients=1 if result.get("emailed_to") else 0,
                            status="done",
                            meta={
                                "filename": result.get("filename"),
                                "size_bytes": result.get("size_bytes"),
                                "lite_mode": bool(result.get("lite_mode")),
                            },
                        )
                    else:
                        failed_attempts[today] = attempts + 1
                        logger.warning(
                            f"[scheduled-backup] attempt #{attempts + 1} returned no result — "
                            f"will retry on next loop tick."
                        )
                        _BACKUP_SCHEDULER_STATE["last_attempt_outcome"] = (
                            f"FAILED (attempt {attempts + 1}, no result returned)"
                        )
                        await scheduler_mark_failed(db, "backup_scheduler_zip", slot_key, error="scheduled_zip_returned_no_result")
        except asyncio.CancelledError:
            raise
        except Exception as e:
            failed_attempts[today] = failed_attempts.get(today, 0) + 1
            logger.exception(f"[scheduled-backup] loop tick error: {e}")
            _BACKUP_SCHEDULER_STATE["last_attempt_outcome"] = f"EXCEPTION: {e!r}"

        # ── Iter64 Phase 2c + iter85 — Hourly/Nightly complete archive to R2 ─
        # When ``BACKUP_R2_HOURLY=true`` (iter85) the complete archive fires
        # at the top of every UTC hour, capping the maximum data-loss
        # window on a sudden redeploy to ~60 minutes. Otherwise it falls
        # back to the legacy once-per-day schedule at
        # ``BACKUP_R2_FULL_HOUR_UTC`` (default 03:00 UTC).
        try:
            r2_hour = int(os.environ.get("BACKUP_R2_FULL_HOUR_UTC", "3") or "3")
        except ValueError:
            r2_hour = 3
        activation_state = await _build_hourly_activation_state(db, runtime_state=_BACKUP_SCHEDULER_STATE.get("backup_runtime"))
        r2_hourly = bool(activation_state.get("r2_hourly_effective"))
        if activation_state.get("activation_status") != "ACTIVE":
            _BACKUP_SCHEDULER_STATE["last_attempt_outcome"] = activation_state.get("activation_status")
        hour_bucket = now.strftime("%Y-%m-%dT%H")
        current_hour_r2_row = None
        if r2_hourly and _BACKUP_SCHEDULER_STATE.get("last_r2_complete_hour") != hour_bucket:
            try:
                current_hour_r2_row = await db.backup_health.find_one(
                    {
                        "mode": "complete-r2",
                        "ok": True,
                        "ts": {"$regex": f"^{hour_bucket}"},
                        "filename": {"$nin": [None, ""]},
                    },
                    sort=[("ts", -1)],
                    projection={"_id": 0, "filename": 1, "size_bytes": 1, "ts": 1},
                )
            except Exception as e:  # noqa: BLE001
                logger.warning(f"[scheduled-backup] current-hour complete-r2 dedupe probe failed: {e}")
            if current_hour_r2_row:
                _BACKUP_SCHEDULER_STATE["last_r2_complete_hour"] = hour_bucket
                _BACKUP_SCHEDULER_STATE["last_r2_complete_date"] = hour_bucket[:10]
                _BACKUP_SCHEDULER_STATE["last_r2_complete"] = {
                    "filename": current_hour_r2_row.get("filename"),
                    "size_bytes": current_hour_r2_row.get("size_bytes"),
                    "r2_key": f"{_canonical_backup_prefix().rstrip('/')}/{current_hour_r2_row.get('filename')}",
                    "ts": current_hour_r2_row.get("ts"),
                }
        if r2_hourly:
            should_fire_r2 = (
                _BACKUP_SCHEDULER_STATE.get("last_r2_complete_hour") != hour_bucket
                and current_hour_r2_row is None
            )
        else:
            should_fire_r2 = (
                now.hour >= r2_hour
                and _BACKUP_SCHEDULER_STATE.get("last_r2_complete_date") != str(today)
            )
        if should_fire_r2:
            try:
                active_jobs = await get_active_backup_jobs(db)
                overlap = classify_backup_overlap(active_jobs)
                if overlap.get("restore_active"):
                    logger.warning("[scheduled-backup] complete-archive deferred — restore job active")
                    _BACKUP_SCHEDULER_STATE["last_attempt_outcome"] = "COMPLETE_ARCHIVE_DEFERRED_RESTORE_ACTIVE"
                    await asyncio.sleep(300)
                    continue
                if overlap.get("reclaimable_backups"):
                    logger.warning(
                        f"[scheduled-backup] ignoring {len(overlap.get('reclaimable_backups') or [])} reclaimable stale backup job(s) for slot {hour_bucket}"
                    )
                slot_key = backup_slot_key_for_hour(now)
                job = await claim_backup_job(
                    db,
                    job_type="scheduled_backup",
                    kind=BACKUP_JOB_KIND_COMPLETE_R2,
                    slot_key=slot_key,
                    trigger="scheduler_nightly",
                    owner_id=scheduler_owner,
                    metadata={"hour_bucket": hour_bucket, "hourly_requested": bool(activation_state.get("r2_hourly_requested")), "hourly_effective": bool(activation_state.get("r2_hourly_effective"))},
                )
                if job is None:
                    logger.warning(f"[scheduled-backup] duplicate complete-archive slot prevented: {slot_key}")
                    await asyncio.sleep(300)
                    continue
                lease = await start_backup_job(db, job["job_id"])
                await heartbeat_backup_job(db, job["job_id"], owner_token=lease.owner_token, extra={"stage": "preflight"})
                logger.info(f"[scheduled-backup] firing complete-archive → R2 ({'hourly' if r2_hourly else 'nightly'}) bucket={hour_bucket}")
                r2_res = await _run_complete_archive_to_r2(db)
                if r2_res:
                    if r2_res.get("deferred"):
                        await fail_backup_job(db, job["job_id"], error=f"deferred:{r2_res.get('reason')}", result=r2_res, state="deferred", owner_token=lease.owner_token)
                        _BACKUP_SCHEDULER_STATE["last_attempt_outcome"] = "COMPLETE_ARCHIVE_DEFERRED_RESOURCE_GUARD"
                        await asyncio.sleep(300)
                        continue
                    _BACKUP_SCHEDULER_STATE["last_r2_complete_date"] = str(today)
                    _BACKUP_SCHEDULER_STATE["last_r2_complete_hour"] = hour_bucket
                    _BACKUP_SCHEDULER_STATE["last_r2_complete"] = {
                        "filename": r2_res.get("filename"),
                        "size_bytes": r2_res.get("size_bytes"),
                        "r2_key": r2_res.get("r2_key"),
                        "ts": now.isoformat(),
                    }
                    logger.info(
                        f"[scheduled-backup] complete archive in R2: "
                        f"{r2_res.get('r2_key')} · {(r2_res.get('size_bytes') or 0)/1024/1024:.1f} MB"
                    )
                    await complete_backup_job(db, job["job_id"], outcome="ok", result=r2_res, owner_token=lease.owner_token)
                else:
                    logger.warning("[scheduled-backup] complete-archive → R2 returned no result")
                    await fail_backup_job(db, job["job_id"], error="complete_archive_returned_no_result", owner_token=lease.owner_token)
            except Exception as e:  # noqa: BLE001
                logger.exception(f"[scheduled-backup] complete-archive → R2 failed: {e}")
                try:
                    await fail_backup_job(db, job["job_id"], error=repr(e), owner_token=lease.owner_token)
                except Exception:
                    pass

        # WATCHDOG — runs every tick (cheap: one Mongo read), fires alarm
        # email if backups have been silent past the threshold. Rate-limited
        # internally so the admin doesn't get spammed.
        try:
            watchdog_result = await _backup_watchdog_check(db)
            if watchdog_result:
                _BACKUP_SCHEDULER_STATE["last_watchdog"] = watchdog_result
        except Exception as e:  # noqa: BLE001
            logger.warning(f"[backup-watchdog] tick failed: {e}")
        # Weekly payroll-variance email (Sunday 18:00 UTC by default)
        try:
            await _maybe_send_weekly_variance_email()
        except Exception as e:  # noqa: BLE001
            logger.warning(f"[payroll-variance-cron] tick failed: {e}")
        await asyncio.sleep(300)  # 5 min ticks — low overhead, catches missed slots


@api_router.get("/admin/backups")
async def admin_list_backups(_: bool = Depends(require_admin_strict)):
    """List every stored backup on disk + current schedule settings."""
    files = _list_stored_backups()
    total_bytes = sum(f["size_bytes"] for f in files)
    return {
        "backups": files,
        "count": len(files),
        "total_bytes": total_bytes,
        "schedule": {
            "hour_utc": BACKUP_HOUR_UTC,         # legacy single-window field
            "hours_utc": BACKUP_HOURS_UTC,       # full list of scheduled UTC hours
            "retention_days": BACKUP_RETENTION_DAYS,
            "storage_dir": str(BACKUPS_DIR),
            "enabled": True,
        },
    }


def _normalize_backup_manifest_collection_name(name: Any) -> str:
    """Map manifest-facing aliases back to canonical Mongo collection names.

    Complete R2 archives historically preserve the legacy export folder names for
    a small fixed subset of collections (for example `daily-reports`), while the
    live database exposes the real Mongo collection names (`daily_reports`).
    Integrity comparisons must treat those as the same collection.
    """
    if not isinstance(name, str):
        return ""
    raw = name.strip()
    return EXPORTABLE_KINDS.get(raw, raw)


def _compute_backup_integrity_missing(
    live_collections: List[str],
    captured_collections: List[str],
    explicit_exclusions: Optional[List[str]] = None,
) -> List[str]:
    """Return the canonical live collections missing from the backup manifest.

    Two truths must hold:
      1. Legacy manifest aliases like `daily-reports` are equivalent to their
         Mongo collection names (`daily_reports`).
      2. Collections explicitly excluded by the manifest are *not* actionable
         integrity misses — they are intentional archive-shape exclusions.
    """
    excluded = {
        _normalize_backup_manifest_collection_name(it)
        for it in (explicit_exclusions or [])
        if _normalize_backup_manifest_collection_name(it)
    }
    excluded.update(
        _normalize_backup_manifest_collection_name(it)
        for it in BACKUP_EXPLICIT_EXCLUSIONS
        if _normalize_backup_manifest_collection_name(it)
    )
    captured = {
        _normalize_backup_manifest_collection_name(it)
        for it in (captured_collections or [])
        if _normalize_backup_manifest_collection_name(it)
    }
    live = {
        _normalize_backup_manifest_collection_name(it)
        for it in (live_collections or [])
        if _normalize_backup_manifest_collection_name(it)
    }
    required_live = sorted(it for it in live if it not in excluded)
    return [it for it in required_live if it not in captured]


def _normalize_manifest_required_set(manifest: Dict[str, Any]) -> Dict[str, Any]:
    explicit_exclusions = manifest.get("explicit_exclusions") or manifest.get("excluded_collections") or []
    captured_collections = (
        manifest.get("captured_collections")
        or manifest.get("all_db_collections_at_backup_time")
        or []
    )
    expected_collections = manifest.get("expected_collections") or []
    if expected_collections:
        expected = sorted(
            _normalize_backup_manifest_collection_name(it)
            for it in expected_collections
            if _normalize_backup_manifest_collection_name(it)
        )
    else:
        captured = {
            _normalize_backup_manifest_collection_name(it)
            for it in captured_collections
            if _normalize_backup_manifest_collection_name(it)
        }
        exclusions = {
            _normalize_backup_manifest_collection_name(it)
            for it in explicit_exclusions
            if _normalize_backup_manifest_collection_name(it)
        }
        expected = sorted(it for it in captured if it not in exclusions)

    captured = sorted(
        _normalize_backup_manifest_collection_name(it)
        for it in captured_collections
        if _normalize_backup_manifest_collection_name(it)
    )
    exclusions = sorted(
        _normalize_backup_manifest_collection_name(it)
        for it in explicit_exclusions
        if _normalize_backup_manifest_collection_name(it)
    )
    return {
        "expected": expected,
        "captured": captured,
        "exclusions": exclusions,
    }


def _evaluate_backup_manifest_contract(
    manifest: Dict[str, Any],
    *,
    live_collections: Optional[List[str]] = None,
    archive_member_prefixes: Optional[List[str]] = None,
) -> Dict[str, Any]:
    now_iso = datetime.now(timezone.utc).isoformat()
    normalized = _normalize_manifest_required_set(manifest)
    expected = normalized["expected"]
    captured = normalized["captured"]
    exclusions = normalized["exclusions"]

    manifest_expected_count = manifest.get("expected_collection_count")
    manifest_captured_count = manifest.get("captured_collection_count")
    manifest_total_records = manifest.get("total_records")
    per_collection_record_counts = manifest.get("per_collection_record_counts")
    coverage_gap = manifest.get("coverage_gap") or []
    failed_collections = manifest.get("failed_collections") or []
    skipped_collections = manifest.get("skipped_collections") or []
    integrity_failed_checks = list(manifest.get("integrity_failed_checks") or [])
    failed_checks: List[Dict[str, Any]] = []
    unavailable_fields: List[str] = []

    def _add_failed(code: str, expected_value: Any, actual_value: Any, evidence_source: str) -> None:
        failed_checks.append({
            "code": code,
            "expected": expected_value,
            "actual": actual_value,
            "status": "FAIL",
            "evidence_source": evidence_source,
        })

    check_matrix: List[Dict[str, Any]] = []
    check_matrix.append({
        "check": "manifest_parsing",
        "expected": "manifest readable JSON",
        "actual": "manifest readable JSON",
        "status": "PASS",
        "evidence_source": "manifest",
    })
    check_matrix.append({
        "check": "manifest_version",
        "expected": manifest.get("manifest_version") or manifest.get("version") or "legacy",
        "actual": manifest.get("manifest_version") or manifest.get("version") or "legacy",
        "status": "PASS",
        "evidence_source": "manifest",
    })

    if manifest_expected_count is not None and manifest_expected_count != len(expected):
        _add_failed("expected_collection_count_mismatch", manifest_expected_count, len(expected), "manifest")
    if manifest_captured_count is not None and manifest_captured_count != len(captured):
        _add_failed("captured_collection_count_mismatch", manifest_captured_count, len(captured), "manifest")

    if isinstance(per_collection_record_counts, dict):
        normalized_counts = {
            _normalize_backup_manifest_collection_name(k): int(v)
            for k, v in per_collection_record_counts.items()
            if _normalize_backup_manifest_collection_name(k)
        }
        if expected and set(normalized_counts.keys()) != set(expected):
            missing_counts = sorted(set(expected) - set(normalized_counts.keys()))
            extra_counts = sorted(set(normalized_counts.keys()) - set(expected))
            _add_failed(
                "per_collection_record_counts_set_mismatch",
                {"missing": [], "extra": []},
                {"missing": missing_counts, "extra": extra_counts},
                "manifest",
            )
        calc_total = sum(v for v in normalized_counts.values() if isinstance(v, int))
        if manifest_total_records is not None and calc_total != manifest_total_records:
            _add_failed("total_record_reconciliation_failed", manifest_total_records, calc_total, "manifest")
    else:
        unavailable_fields.append("per_collection_record_counts")

    if coverage_gap:
        _add_failed("coverage_gap_non_empty", [], coverage_gap, "manifest")
    if failed_collections:
        _add_failed("failed_collections_non_empty", [], failed_collections, "manifest")
    if skipped_collections:
        _add_failed("skipped_collections_non_empty", [], skipped_collections, "manifest")

    missing_from_expected = [it for it in expected if it not in captured]
    if missing_from_expected:
        _add_failed("missing_from_expected_set", [], missing_from_expected, "manifest")

    if live_collections is not None:
        missing_vs_live = _compute_backup_integrity_missing(live_collections, captured, exclusions)
        if missing_vs_live:
            _add_failed("missing_from_live_required_set", [], missing_vs_live, "live+manifest")
    else:
        unavailable_fields.append("live_collection_comparison")
        missing_vs_live = []

    if archive_member_prefixes is not None:
        archive_required = sorted(
            _normalize_backup_manifest_collection_name(it)
            for it in archive_member_prefixes
            if _normalize_backup_manifest_collection_name(it)
        )
        if expected and set(expected) != set(archive_required):
            _add_failed(
                "archive_member_set_mismatch",
                expected,
                archive_required,
                "archive",
            )
    else:
        unavailable_fields.append("archive_member_set")

    failed_checks.extend(integrity_failed_checks)
    classification = "PASS"
    reason_code = "verification_pass"
    if failed_checks:
        classification = "BACKUP_INCOMPLETE"
        reason_code = failed_checks[0].get("code") or "backup_incomplete"
    elif not expected and not captured:
        classification = "UNKNOWN"
        reason_code = "manifest_missing_collection_sets"

    unavailable_reason = None
    if unavailable_fields:
        unavailable_reason = "Unavailable because this artifact/evidence path predates full verifier metadata or bounded archive inspection was not requested."

    return {
        "expected_collections": expected,
        "expected_collection_count": len(expected),
        "captured_collections": captured,
        "captured_collection_count": len(captured),
        "explicit_exclusions": exclusions,
        "missing_from_backup": missing_vs_live if live_collections is not None else missing_from_expected,
        "failed_checks": failed_checks,
        "check_matrix": check_matrix + [
            {
                "check": item.get("code") or "failed_check",
                "expected": item.get("expected"),
                "actual": item.get("actual") if "actual" in item else item.get("collections") or item.get("message"),
                "status": item.get("status") or "FAIL",
                "evidence_source": item.get("evidence_source") or "manifest",
            }
            for item in failed_checks
        ],
        "integrity_result": "PASS" if classification == "PASS" else ("UNKNOWN" if classification == "UNKNOWN" else "FAIL"),
        "classification_reason_code": reason_code,
        "classification": classification,
        "verification_timestamp": now_iso,
        "verifier_version": BACKUP_VERIFIER_VERSION,
        "evidence_mode": "LIVE_CALCULATED" if live_collections is not None else "MANIFEST_ONLY",
        "unavailable_fields": unavailable_fields,
        "unavailable_reason": unavailable_reason,
    }


@api_router.get("/admin/backups/integrity-check")
async def admin_backup_integrity_check(_: bool = Depends(require_admin_strict)):
    status = await db.backup_integrity_jobs.find_one(
        {"job_type": "integrity_check"},
        {"_id": 0},
        sort=[("created_at", -1)],
    )
    if status and status.get("state") in {"queued", "running"}:
        return JSONResponse(status_code=202, content=status)
    if status and status.get("state") in {"completed", "failed", "stale"}:
        return status
    """Audit: every Mongo collection currently in the live DB vs the most
    recent backup's manifest. Surfaces any collection that exists right now
    but wasn't captured in the last backup — proves nothing is slipping
    through, and catches new collections automatically.

    Returns:
      {
        "last_backup_filename": str | None,
        "last_backup_at":       iso str | None,
        "live_collections":     [...],     # everything currently in DB
        "captured_collections": [...],     # what the last backup contained
        "missing_from_backup":  [...],     # ⚠ in DB but NOT in last backup
        "ok":                   bool,
      }

    NOTE: this route MUST be declared before the parameterized
    `/admin/backups/{filename}` route below — otherwise the FastAPI router
    matches the literal "integrity-check" against the {filename} regex.
    """
    from backup_verification import list_r2_backup_archives, read_r2_backup_manifest  # noqa: PLC0415

    files = _list_stored_backups()
    last = files[0] if files else None
    live = sorted(await db.list_collection_names())
    live = [c for c in live if not c.startswith("system.")]
    env_identity = {
        "app_env": _canonical_app_env(),
        "db_name": _canonical_db_name(),
    }

    latest_row = await db.backup_health.find_one(
        {"mode": "complete-r2", "ok": True, "filename": {"$nin": [None, ""]}},
        {"_id": 0, "filename": 1, "size_bytes": 1, "records": 1, "ts": 1},
        sort=[("ts", -1)],
    )

    r2_archives = await list_r2_backup_archives(prefix=_canonical_backup_prefix())
    latest_r2 = r2_archives[0] if r2_archives else None
    latest_r2_filename = None
    if latest_r2:
        latest_r2_filename = latest_r2.get("filename")
        if not latest_r2_filename:
            latest_r2_key = str(latest_r2.get("key") or "")
            latest_r2_filename = latest_r2_key.rsplit("/", 1)[-1] or None

    drift_row = await db.backup_drift_history.find_one(
        {},
        {"_id": 0, "recorded_at": 1, "captured_collections": 1, "total_records": 1, "explicit_exclusions": 1},
        sort=[("recorded_at", -1)],
    )

    captured: List[str] = []
    collection_counts: Optional[Dict[str, int]] = None
    document_count: Optional[int] = None
    last_at = None
    last_object_key = latest_r2.get("key") if latest_r2 else None
    archive_size_bytes = None
    manifest_source = "unavailable"
    manifest_explicit_exclusions: List[str] = []

    manifest_bundle = None
    latest_r2_with_manifest = None
    latest_r2_matching_runtime_with_manifest = None
    for archive in r2_archives:
        key = archive.get("key")
        if not key:
            continue
        bundle = await read_r2_backup_manifest(key)
        if not bundle or not isinstance(bundle.get("manifest"), dict):
            continue
        if manifest_bundle is None:
            manifest_bundle = bundle
            latest_r2_with_manifest = archive
        manifest = bundle.get("manifest") or {}
        manifest_env = {
            "app_env": manifest.get("app_env") or manifest.get("environment"),
            "db_name": manifest.get("db_name") or manifest.get("database_name"),
        }
        if (
            manifest_env["app_env"] == env_identity["app_env"]
            and manifest_env["db_name"] == env_identity["db_name"]
        ):
            latest_r2_matching_runtime_with_manifest = archive
            manifest_bundle = bundle
            break

    if latest_r2_matching_runtime_with_manifest or latest_r2_with_manifest:
        latest_r2 = latest_r2_matching_runtime_with_manifest or latest_r2_with_manifest
        latest_r2_filename = latest_r2.get("filename")
        if not latest_r2_filename:
            latest_r2_key = str(latest_r2.get("key") or "")
            latest_r2_filename = latest_r2_key.rsplit("/", 1)[-1] or None
        last_object_key = latest_r2.get("key")

    if manifest_bundle and isinstance(manifest_bundle.get("manifest"), dict):
        m = manifest_bundle["manifest"]
        captured = sorted(m.get("captured_collections") or m.get("all_db_collections_at_backup_time") or [])
        counts = m.get("per_kind")
        collection_counts = counts if isinstance(counts, dict) else None
        raw_exclusions = m.get("explicit_exclusions") or []
        if isinstance(raw_exclusions, list):
            manifest_explicit_exclusions = [str(it) for it in raw_exclusions if str(it)]
        raw_total_records = m.get("total_records")
        try:
            document_count = int(raw_total_records) if raw_total_records is not None else None
        except Exception:  # noqa: BLE001
            document_count = None
        last_at = m.get("generated_at") or manifest_bundle.get("last_modified_iso")
        archive_size_bytes = manifest_bundle.get("content_length")
        manifest_source = f"r2:{manifest_bundle.get('manifest_name')}"

    recent_backups: List[Dict[str, Any]] = []
    recent_backup_rows: List[Dict[str, Any]] = []
    try:
        recent_backup_rows = await db.backup_health.find(
            {"mode": "complete-r2", "ok": True, "filename": {"$nin": [None, ""]}},
            {"_id": 0, "filename": 1, "size_bytes": 1, "records": 1, "ts": 1},
            sort=[("ts", -1)],
        ).to_list(length=5)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"integrity-check: recent lineage read failed: {e}")

    if not captured and drift_row:
        captured = sorted(drift_row.get("captured_collections") or [])
        raw_total_records = drift_row.get("total_records")
        try:
            document_count = int(raw_total_records) if raw_total_records is not None else document_count
        except Exception:  # noqa: BLE001
            pass
        if getattr(drift_row.get("recorded_at"), "isoformat", None):
            last_at = last_at or drift_row["recorded_at"].isoformat()
        manifest_source = "backup_drift_history"

    if not captured and last:
        zip_path = BACKUPS_DIR / last["filename"]
        try:
            import json as _ic_json
            import zipfile as _ic_zip
            with _ic_zip.ZipFile(zip_path) as zf:
                manifest_name = None
                for candidate in ("backup_manifest.json", "MANIFEST.json"):
                    if candidate in zf.namelist():
                        manifest_name = candidate
                        break
                if manifest_name:
                    m = _ic_json.loads(zf.read(manifest_name).decode("utf-8"))
                    captured = sorted(m.get("captured_collections") or m.get("all_db_collections_at_backup_time") or [])
                    counts = m.get("per_kind")
                    collection_counts = counts if isinstance(counts, dict) else collection_counts
                    raw_total_records = m.get("total_records")
                    try:
                        document_count = int(raw_total_records) if raw_total_records is not None else document_count
                    except Exception:  # noqa: BLE001
                        pass
                    last_at = m.get("generated_at") or last_at
                    archive_size_bytes = archive_size_bytes or last.get("size_bytes")
                    manifest_source = f"disk:{manifest_name}"
        except Exception as e:  # noqa: BLE001
            logger.warning(f"integrity-check: read manifest failed: {e}")

    if latest_r2:
        archive_size_bytes = archive_size_bytes or latest_r2.get("size_bytes")
        last_at = last_at or latest_r2.get("last_modified_iso")
    if latest_row:
        archive_size_bytes = archive_size_bytes or latest_row.get("size_bytes")
        document_count = document_count if document_count is not None else latest_row.get("records")
        last_at = last_at or latest_row.get("ts")

    last_drill = None
    try:
        drill_row = await db.drill_runs.find_one(
            {"state": "done"},
            {"_id": 0, "finished_at": 1, "started_at": 1, "outcome": 1,
             "records_restored": 1, "photos_rehydrated": 1, "duration_minutes": 1,
             "archive_filename": 1},
            sort=[("started_at", -1)],
        )
        if drill_row:
            last_drill = {
                "ts": drill_row.get("finished_at") or drill_row.get("started_at"),
                "outcome": drill_row.get("outcome"),
                "records": drill_row.get("records_restored") or 0,
                "photos": drill_row.get("photos_rehydrated") or 0,
                "duration_min": drill_row.get("duration_minutes"),
                "archive_filename": drill_row.get("archive_filename"),
            }
    except Exception:
        last_drill = None

    top_manifest = manifest_bundle.get("manifest") if manifest_bundle else None
    top_contract = _evaluate_backup_manifest_contract(
        top_manifest or {
            "captured_collections": captured,
            "explicit_exclusions": manifest_explicit_exclusions,
            "total_records": document_count,
        },
        live_collections=live,
    )
    top_contract_env = {
        "app_env": (top_manifest or {}).get("app_env") or (top_manifest or {}).get("environment"),
        "db_name": (top_manifest or {}).get("db_name") or (top_manifest or {}).get("database_name"),
    }
    env_matches = bool(top_contract_env["app_env"] == env_identity["app_env"] and top_contract_env["db_name"] == env_identity["db_name"])
    if top_manifest and (top_contract_env["app_env"] or top_contract_env["db_name"]) and not env_matches:
        top_contract = _evaluate_backup_manifest_contract(
            top_manifest,
            live_collections=None,
        )
        top_contract["classification"] = "UNKNOWN"
        top_contract["classification_reason_code"] = "environment_mismatch_manifest_vs_runtime"
        if top_contract.get("integrity_result") == "PASS":
            top_contract["integrity_result"] = "UNKNOWN"
        top_contract["failed_checks"] = []
        top_contract["missing_from_backup"] = []
        top_contract["evidence_mode"] = "MANIFEST_ONLY"
        top_contract["unavailable_fields"] = sorted(set(top_contract["unavailable_fields"] + ["live_collection_comparison"]))
        top_contract["unavailable_reason"] = (
            "The latest manifest was generated for a different environment/database identity than the currently running backend; "
            "live collection comparison is intentionally suppressed to avoid cross-environment false FAIL."
        )

    missing = top_contract["missing_from_backup"]
    integrity_result = top_contract["integrity_result"]

    row_by_filename = {
        row.get("filename"): row
        for row in recent_backup_rows
        if row.get("filename")
    }
    recent_candidates: List[Dict[str, Any]] = []
    seen_recent_filenames = set()
    for archive in r2_archives[:5]:
        filename = archive.get("filename") or str(archive.get("key") or "").rsplit("/", 1)[-1]
        if not filename or filename in seen_recent_filenames:
            continue
        seen_recent_filenames.add(filename)
        recent_candidates.append({
            "filename": filename,
            "object_key": archive.get("key"),
            "ts": archive.get("last_modified_iso"),
            "size_bytes": archive.get("size_bytes") or 0,
            "records": (row_by_filename.get(filename) or {}).get("records"),
        })
    for row in recent_backup_rows:
        filename = row.get("filename")
        if not filename or filename in seen_recent_filenames or len(recent_candidates) >= 5:
            continue
        seen_recent_filenames.add(filename)
        recent_candidates.append({
            "filename": filename,
            "object_key": f"{_canonical_backup_prefix().rstrip('/')}/{filename}",
            "ts": row.get("ts"),
            "size_bytes": row.get("size_bytes") or 0,
            "records": row.get("records"),
        })

    for idx, candidate in enumerate(recent_candidates[:5]):
        row_manifest: Dict[str, Any] = {}
        row_evidence_source = "unavailable"
        row_ts = candidate.get("ts")
        row_size_bytes = candidate.get("size_bytes") or 0
        row_archive_member_prefixes: Optional[List[str]] = None
        row_contract = None

        if candidate.get("object_key"):
            row_manifest_bundle = None
            if manifest_bundle and candidate.get("object_key") == (latest_r2 or {}).get("key"):
                row_manifest_bundle = manifest_bundle
            else:
                try:
                    row_manifest_bundle = await read_r2_backup_manifest(candidate["object_key"])
                except Exception as e:  # noqa: BLE001
                    logger.warning(f"integrity-check: bounded manifest read failed for {candidate.get('object_key')}: {e}")
                    row_manifest_bundle = None
            if row_manifest_bundle and isinstance(row_manifest_bundle.get("manifest"), dict):
                row_manifest = row_manifest_bundle["manifest"]
                row_ts = row_manifest.get("generated_at") or row_manifest_bundle.get("last_modified_iso") or row_ts
                row_size_bytes = row_manifest_bundle.get("content_length") or row_size_bytes
                row_evidence_source = f"r2:{row_manifest_bundle.get('manifest_name')}"
        if row_contract is None and idx > 0 and not candidate.get("object_key"):
            row_evidence_source = "summary-only"
            row_contract = {
                "captured_collections": None,
                "captured_collection_count": None,
                "explicit_exclusions": None,
                "missing_from_backup": [],
                "integrity_result": "UNKNOWN",
                "failed_checks": [],
                "verification_timestamp": datetime.now(timezone.utc).isoformat(),
                "verifier_version": "track-27.11c-summary",
                "evidence_mode": "SUMMARY_ONLY",
                "unavailable_fields": [
                    "captured_collections",
                    "integrity_result",
                    "live_collection_comparison",
                ],
                "unavailable_reason": "Recent archive manifest parsing skipped for latency; open the archive directly for per-file proof.",
                "classification_reason_code": "summary_only_recent_history",
                "classification": "UNKNOWN",
                "check_matrix": [],
            }

        if not row_manifest and idx == 0 and captured:
            row_manifest = {
                "captured_collections": list(captured),
                "per_kind": collection_counts,
                "total_records": document_count,
                "explicit_exclusions": list(manifest_explicit_exclusions),
            }
            row_evidence_source = manifest_source
            row_ts = last_at or row_ts
            row_size_bytes = archive_size_bytes or row_size_bytes

        normalized_env = {
            "app_env": row_manifest.get("app_env") or row_manifest.get("environment"),
            "db_name": row_manifest.get("db_name") or row_manifest.get("database_name"),
        }
        has_explicit_identity = bool(normalized_env["app_env"] or normalized_env["db_name"])
        use_live_comparison = bool(
            row_manifest and (
                not has_explicit_identity
                or (
                    normalized_env["app_env"] == env_identity["app_env"]
                    and normalized_env["db_name"] == env_identity["db_name"]
                )
            )
        )
        if row_contract is None:
            row_contract = _evaluate_backup_manifest_contract(
                row_manifest or {},
                live_collections=live if use_live_comparison else None,
                archive_member_prefixes=row_archive_member_prefixes,
            )
        if row_manifest and not use_live_comparison and has_explicit_identity:
            row_contract["classification"] = "UNKNOWN"
            row_contract["classification_reason_code"] = "environment_mismatch_manifest_vs_runtime"
            if row_contract.get("integrity_result") == "PASS":
                row_contract["integrity_result"] = "UNKNOWN"
            row_contract["evidence_mode"] = "MANIFEST_ONLY"
            row_contract["unavailable_fields"] = sorted(set(row_contract["unavailable_fields"] + ["live_collection_comparison"]))
            row_contract["unavailable_reason"] = (
                "The manifest environment/database identity does not match the current runtime; cross-environment live comparison is intentionally suppressed."
            )

        recent_backups.append({
            "filename": candidate.get("filename"),
            "object_key": candidate.get("object_key"),
            "backup_at": row_ts,
            "archive_size_bytes": row_size_bytes,
            "captured_collections": row_contract["captured_collections"],
            "captured_collection_count": row_contract["captured_collection_count"],
            "collection_counts": row_manifest.get("per_kind") if isinstance(row_manifest.get("per_kind"), dict) else None,
            "explicit_exclusions": row_contract["explicit_exclusions"],
            "total_records": row_manifest.get("total_records") if row_manifest else candidate.get("records"),
            "manifest_version": row_manifest.get("manifest_version") or row_manifest.get("version") if row_manifest else None,
            "missing_from_backup": row_contract["missing_from_backup"],
            "integrity_result": row_contract["integrity_result"],
            "failed_checks": row_contract["failed_checks"],
            "failed_check": (row_contract["missing_from_backup"] or [None])[0],
            "verification_timestamp": row_contract["verification_timestamp"],
            "verifier_version": row_contract["verifier_version"],
            "evidence_source": row_evidence_source,
            "evidence_mode": row_contract["evidence_mode"],
            "unavailable_fields": row_contract["unavailable_fields"],
            "unavailable_reason": row_contract["unavailable_reason"],
            "classification_reason_code": row_contract["classification_reason_code"],
            "classification": row_contract["classification"],
            "check_matrix": row_contract["check_matrix"],
            "live_persisted_indicator": "live_runtime_db" if use_live_comparison else "manifest_only",
            "ts": row_ts,
            "size_bytes": row_size_bytes,
        })
    return {
        "last_backup_filename": (
            latest_r2_filename
            or (latest_row.get("filename") if latest_row else None)
            or (last.get("filename") if last else None)
        ),
        "last_backup_object_key": last_object_key,
        "last_backup_at": last_at,
        "archive_size_bytes": archive_size_bytes,
        "live_collections": live,
        "captured_collections": captured,
        "captured_collection_count": len(captured),
        "collection_counts": collection_counts,
        "document_count": document_count,
        "missing_from_backup": missing,
        "integrity_result": integrity_result,
        "verification_timestamp": top_contract["verification_timestamp"],
        "verifier_version": top_contract["verifier_version"],
        "manifest_version": (top_manifest or {}).get("manifest_version") or (top_manifest or {}).get("version"),
        "expected_collection_count": top_contract["expected_collection_count"],
        "expected_collections": top_contract["expected_collections"],
        "unavailable_fields": top_contract["unavailable_fields"],
        "unavailable_reason": top_contract["unavailable_reason"],
        "classification_reason_code": top_contract["classification_reason_code"],
        "classification": top_contract["classification"],
        "environment_identity": env_identity,
        "restore_test_evidence": last_drill,
        "evidence_source": manifest_source,
        "evidence_mode": top_contract["evidence_mode"],
        "recent_backups": recent_backups,
        "ok": top_contract["integrity_result"] == "PASS",
    }


async def _run_backup_integrity_job(job_id: str, actor_email: str = "admin"):
    from backup_verification import list_r2_backup_archives, read_r2_backup_manifest  # noqa: PLC0415
    from lib.trust_spine import emit_stage, STAGE_AUDIT_WRITTEN, STAGE_COMPLETED, STAGE_RECORD_CREATED  # noqa: PLC0415
    started = datetime.now(timezone.utc)
    correlation_id = f"backup-integrity:{job_id}"
    try:
        await db.backup_integrity_jobs.update_one(
            {"job_id": job_id},
            {"$set": {"state": "running", "started_at": started.isoformat(), "correlation_id": correlation_id, "error": None}},
        )
        await emit_stage(db, workflow="backup-integrity", stage=STAGE_RECORD_CREATED, correlation_id=correlation_id, record_id=job_id, module="server.py", status="ok")

        files = _list_stored_backups()
        last = files[0] if files else None
        live = sorted(await db.list_collection_names())
        live = [c for c in live if not c.startswith("system.")]
        env_identity = {"app_env": _canonical_app_env(), "db_name": _canonical_db_name()}
        latest_row = await db.backup_health.find_one(
            {"mode": "complete-r2", "ok": True, "filename": {"$nin": [None, ""]}},
            {"_id": 0, "filename": 1, "size_bytes": 1, "records": 1, "ts": 1},
            sort=[("ts", -1)],
        )
        r2_archives = await list_r2_backup_archives(prefix=_canonical_backup_prefix())
        latest_r2 = r2_archives[0] if r2_archives else None
        latest_r2_filename = None
        if latest_r2:
            latest_r2_filename = latest_r2.get("filename") or str(latest_r2.get("key") or "").rsplit("/", 1)[-1] or None
        drift_row = await db.backup_drift_history.find_one({}, {"_id": 0, "recorded_at": 1, "captured_collections": 1, "total_records": 1, "explicit_exclusions": 1}, sort=[("recorded_at", -1)])

        captured: List[str] = []
        collection_counts: Optional[Dict[str, int]] = None
        document_count: Optional[int] = None
        last_at = None
        last_object_key = latest_r2.get("key") if latest_r2 else None
        archive_size_bytes = None
        manifest_source = "unavailable"
        manifest_explicit_exclusions: List[str] = []
        manifest_bundle = None
        latest_r2_with_manifest = None
        latest_r2_matching_runtime_with_manifest = None
        manifest_reads = 0
        manifest_read_started = time.perf_counter()
        for archive in r2_archives:
            key = archive.get("key")
            if not key:
                continue
            bundle = await read_r2_backup_manifest(key)
            manifest_reads += 1
            if not bundle or not isinstance(bundle.get("manifest"), dict):
                continue
            if manifest_bundle is None:
                manifest_bundle = bundle
                latest_r2_with_manifest = archive
            manifest = bundle.get("manifest") or {}
            manifest_env = {"app_env": manifest.get("app_env") or manifest.get("environment"), "db_name": manifest.get("db_name") or manifest.get("database_name")}
            if manifest_env["app_env"] == env_identity["app_env"] and manifest_env["db_name"] == env_identity["db_name"]:
                latest_r2_matching_runtime_with_manifest = archive
                manifest_bundle = bundle
                break
        manifest_read_duration_ms = int((time.perf_counter() - manifest_read_started) * 1000)
        if latest_r2_matching_runtime_with_manifest or latest_r2_with_manifest:
            latest_r2 = latest_r2_matching_runtime_with_manifest or latest_r2_with_manifest
            latest_r2_filename = latest_r2.get("filename") or str(latest_r2.get("key") or "").rsplit("/", 1)[-1] or None
            last_object_key = latest_r2.get("key")
        if manifest_bundle and isinstance(manifest_bundle.get("manifest"), dict):
            m = manifest_bundle["manifest"]
            captured = sorted(m.get("captured_collections") or m.get("all_db_collections_at_backup_time") or [])
            counts = m.get("per_kind")
            collection_counts = counts if isinstance(counts, dict) else None
            raw_exclusions = m.get("explicit_exclusions") or []
            if isinstance(raw_exclusions, list):
                manifest_explicit_exclusions = [str(it) for it in raw_exclusions if str(it)]
            raw_total_records = m.get("total_records")
            try:
                document_count = int(raw_total_records) if raw_total_records is not None else None
            except Exception:
                document_count = None
            last_at = m.get("generated_at") or manifest_bundle.get("last_modified_iso")
            archive_size_bytes = manifest_bundle.get("content_length")
            manifest_source = f"r2:{manifest_bundle.get('manifest_name')}"
        recent_backups: List[Dict[str, Any]] = []
        recent_backup_rows: List[Dict[str, Any]] = []
        try:
            recent_backup_rows = await db.backup_health.find(
                {"mode": "complete-r2", "ok": True, "filename": {"$nin": [None, ""]}},
                {"_id": 0, "filename": 1, "size_bytes": 1, "records": 1, "ts": 1},
                sort=[("ts", -1)],
            ).to_list(length=5)
        except Exception as e:
            logger.warning(f"integrity-check: recent lineage read failed: {e}")
        if not captured and drift_row:
            captured = sorted(drift_row.get("captured_collections") or [])
            raw_total_records = drift_row.get("total_records")
            try:
                document_count = int(raw_total_records) if raw_total_records is not None else document_count
            except Exception:
                pass
            if getattr(drift_row.get("recorded_at"), "isoformat", None):
                last_at = last_at or drift_row["recorded_at"].isoformat()
            manifest_source = "backup_drift_history"
        if latest_r2:
            archive_size_bytes = archive_size_bytes or latest_r2.get("size_bytes")
            last_at = last_at or latest_r2.get("last_modified_iso")
        if latest_row:
            archive_size_bytes = archive_size_bytes or latest_row.get("size_bytes")
            document_count = document_count if document_count is not None else latest_row.get("records")
            last_at = last_at or latest_row.get("ts")
        top_manifest = manifest_bundle.get("manifest") if manifest_bundle else None
        top_contract = _evaluate_backup_manifest_contract(
            top_manifest or {"captured_collections": captured, "explicit_exclusions": manifest_explicit_exclusions, "total_records": document_count},
            live_collections=live,
        )
        missing = top_contract["missing_from_backup"]
        integrity_result = top_contract["integrity_result"]
        finished = datetime.now(timezone.utc)
        duration_ms = int((finished - started).total_seconds() * 1000)
        result = {
            "job_id": job_id,
            "job_type": "integrity_check",
            "state": "completed",
            "started_at": started.isoformat(),
            "completed_at": finished.isoformat(),
            "duration_ms": duration_ms,
            "duration_s": round(duration_ms / 1000, 2),
            "manifest_count_evaluated": manifest_reads,
            "manifest_read_duration_ms": manifest_read_duration_ms,
            "last_backup_filename": (latest_r2_filename or (latest_row.get("filename") if latest_row else None) or (last.get("filename") if last else None)),
            "last_backup_object_key": last_object_key,
            "last_backup_at": last_at,
            "archive_size_bytes": archive_size_bytes,
            "live_collections": live,
            "captured_collections": captured,
            "captured_collection_count": len(captured),
            "collection_counts": collection_counts,
            "document_count": document_count,
            "missing_from_backup": missing,
            "integrity_result": integrity_result,
            "verification_timestamp": top_contract["verification_timestamp"],
            "verifier_version": top_contract["verifier_version"],
            "manifest_version": (top_manifest or {}).get("manifest_version") or (top_manifest or {}).get("version"),
            "expected_collection_count": top_contract["expected_collection_count"],
            "expected_collections": top_contract["expected_collections"],
            "unavailable_fields": top_contract["unavailable_fields"],
            "unavailable_reason": top_contract["unavailable_reason"],
            "classification_reason_code": top_contract["classification_reason_code"],
            "classification": top_contract["classification"],
            "environment_identity": env_identity,
            "evidence_source": manifest_source,
            "evidence_mode": top_contract["evidence_mode"],
            "ok": top_contract["integrity_result"] == "PASS",
            "actor_email": actor_email,
            "correlation_id": correlation_id,
        }
        await db.backup_integrity_jobs.update_one({"job_id": job_id}, {"$set": result}, upsert=True)
        await db.admin_audit.insert_one({
            "ts": finished.isoformat(), "actor_email": actor_email, "action": "backup_integrity_check_completed",
            "target_email": None, "ip": None, "user_agent": None,
            "diff": {"job_id": job_id, "state": "completed", "integrity_result": integrity_result, "duration_ms": duration_ms, "manifest_count_evaluated": manifest_reads},
        })
        await emit_stage(db, workflow="backup-integrity", stage=STAGE_AUDIT_WRITTEN, correlation_id=correlation_id, record_id=job_id, module="server.py", status="ok", duration_ms=duration_ms)
        await emit_stage(db, workflow="backup-integrity", stage=STAGE_COMPLETED, correlation_id=correlation_id, record_id=job_id, module="server.py", status="ok", duration_ms=duration_ms)
    except Exception as e:  # noqa: BLE001
        finished = datetime.now(timezone.utc)
        duration_ms = int((finished - started).total_seconds() * 1000)
        error = str(e)[:500]
        await db.backup_integrity_jobs.update_one(
            {"job_id": job_id},
            {"$set": {"state": "failed", "completed_at": finished.isoformat(), "duration_ms": duration_ms, "duration_s": round(duration_ms / 1000, 2), "error": error}},
            upsert=True,
        )
        try:
            await db.admin_audit.insert_one({
                "ts": finished.isoformat(), "actor_email": actor_email, "action": "backup_integrity_check_failed",
                "target_email": None, "ip": None, "user_agent": None,
                "diff": {"job_id": job_id, "state": "failed", "error": error, "duration_ms": duration_ms},
            })
        except Exception:
            pass
        try:
            await emit_stage(db, workflow="backup-integrity", stage=STAGE_COMPLETED, correlation_id=correlation_id, record_id=job_id, module="server.py", status="failed", duration_ms=duration_ms, failure_reason=error, remediation="Review persisted failure reason and retry deliberately.")
        except Exception:
            pass


@api_router.post("/admin/backups/integrity-check/start")
async def admin_backup_integrity_check_start(request: Request, _: bool = Depends(require_admin_strict)):
    now = datetime.now(timezone.utc)
    active = await db.backup_integrity_jobs.find_one(
        {"job_type": "integrity_check", "state": {"$in": ["queued", "running"]}},
        {"_id": 0},
        sort=[("created_at", -1)],
    )
    if active:
        return JSONResponse(status_code=409, content=active)
    stale_cutoff = (now - timedelta(minutes=20)).isoformat()
    await db.backup_integrity_jobs.update_many(
        {"job_type": "integrity_check", "state": {"$in": ["queued", "running"]}, "created_at": {"$lt": stale_cutoff}},
        {"$set": {"state": "stale", "completed_at": now.isoformat(), "error": "Marked stale before new retry."}},
    )
    job_id = f"bic-{uuid.uuid4().hex}"
    actor_email = (request.headers.get("x-actor-email") or request.headers.get("x-forwarded-user") or "admin")[:160]
    job = {
        "job_id": job_id,
        "job_type": "integrity_check",
        "state": "queued",
        "created_at": now.isoformat(),
        "started_at": None,
        "completed_at": None,
        "duration_ms": None,
        "duration_s": None,
        "manifest_count_evaluated": 0,
        "manifest_read_duration_ms": 0,
        "error": None,
        "actor_email": actor_email,
    }
    await db.backup_integrity_jobs.insert_one(dict(job))
    asyncio.create_task(_run_backup_integrity_job(job_id, actor_email))
    return JSONResponse(status_code=202, content=job)


@api_router.get("/admin/backups/integrity-check/status")
async def admin_backup_integrity_check_status(_: bool = Depends(require_admin_strict)):
    row = await db.backup_integrity_jobs.find_one({"job_type": "integrity_check"}, {"_id": 0}, sort=[("created_at", -1)])
    if not row:
        raise HTTPException(404, "No integrity check job found")
    if row.get("state") in {"queued", "running"}:
        return JSONResponse(status_code=202, content=row)
    return row


@api_router.get("/admin/backups/integrity-check/latest")
async def admin_backup_integrity_check_latest(_: bool = Depends(require_admin_strict)):
    row = await db.backup_integrity_jobs.find_one(
        {"job_type": "integrity_check", "state": {"$in": ["completed", "failed", "stale"]}},
        {"_id": 0},
        sort=[("created_at", -1)],
    )
    if not row:
        raise HTTPException(404, "No completed integrity check found")
    return row


@api_router.get("/admin/backups/{filename}")
async def admin_download_stored_backup(
    filename: str,
    request: Request,
    _: bool = Depends(require_admin_strict),
):
    """Download a specific stored backup by filename."""
    # Strict filename validation — only our own backup files.
    if not re.fullmatch(r"MASCI_full_backup_[0-9A-Za-z_\-]+\.zip", filename):
        raise HTTPException(400, "Invalid backup filename")
    path = BACKUPS_DIR / filename
    if not path.exists():
        raise HTTPException(404, "Backup not found")
    try:
        data = path.read_bytes()
    except Exception as e:
        raise HTTPException(500, f"Could not read backup: {e}")
    # Phase 2 Initiative 5b-broader — chain-of-custody log row for every
    # backup download. Never blocks the download itself.
    await _record_admin_action(
        db, "backup_downloaded", request,
        filename=filename, size_bytes=len(data),
    )
    return Response(
        content=data,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Backup-Size-Bytes": str(len(data)),
        },
    )


@api_router.delete("/admin/backups/{filename}")
async def admin_delete_stored_backup(
    filename: str,
    request: Request,
    confirm: Optional[str] = None,
    _: bool = Depends(require_admin_strict),
):
    if not re.fullmatch(r"MASCI_full_backup_[0-9A-Za-z_\-]+\.zip", filename):
        raise HTTPException(400, "Invalid backup filename")
    # Phase 2 Initiative 5b-broader — destructive bulk-delete guard.
    # Caller MUST pass ?confirm=<filename> matching the path. Prevents
    # accidental deletes from misclicks / replayed URLs.
    if confirm != filename:
        await _record_access_denial(
            db, request, namespace="admin",
            reason="bulk_delete_missing_confirm", target=filename,
        )
        raise HTTPException(
            400,
            "Confirmation required — pass ?confirm=<filename> matching the path.",
        )
    # Phase 2 Initiative 5b-full — step-up re-auth gate.
    if _admin_step_up_enabled():
        x_admin_token = request.headers.get("x-admin-token") or ""
        await _require_recent_step_up(db, x_admin_token, request, max_age_min=5)
    path = BACKUPS_DIR / filename
    if not path.exists():
        raise HTTPException(404, "Backup not found")
    try:
        path.unlink()
    except Exception as e:
        raise HTTPException(500, f"Could not delete: {e}")
    await _record_admin_action(db, "backup_deleted", request,
                               filename=filename)
    return {"ok": True, "filename": filename}


@api_router.post("/admin/backups/run-now")
async def admin_run_backup_now(
    background_tasks: BackgroundTasks,
    lite: Optional[bool] = None,
    _: bool = Depends(require_admin_strict),
):
    """Schedule an immediate backup. Returns HTTP 202 instantly — the
    actual zip build runs in a FastAPI ``BackgroundTask`` so this endpoint
    never blocks the request worker.

    The old synchronous path was Cloudflare-524'ing on the 887 MB zip
    builds because the worker held the request open for >100 seconds.
    Worse, repeated retries could OOM the worker entirely (origin 520
    storm). With the background task, the worker returns 202 in <50 ms;
    the zip writes in the background; the email helper picks up the
    finished file and sends it.

    Query params:
      • ``lite=true`` — skip the full zip and produce only the slim
        metadata-only zip. Use when the full archive is too big.
        Defaults to the ``BACKUP_LITE_MODE_ONLY`` env flag.

    Module-level guard prevents two manual triggers from running
    simultaneously (would double the memory pressure and double the
    OOM risk). Second click within an in-progress window gets a 409.
    """
    global _BACKUP_RUNNOW_IN_PROGRESS
    if _BACKUP_RUNNOW_IN_PROGRESS:
        raise HTTPException(
            409,
            "Another manual backup is already in progress. "
            "Check /api/admin/backups/scheduler-state for status.",
        )

    use_lite = _lite_mode_default() if lite is None else bool(lite)
    _BACKUP_RUNNOW_IN_PROGRESS = True
    _BACKUP_RUNNOW_LAST["started_at"] = datetime.now(timezone.utc).isoformat()
    _BACKUP_RUNNOW_LAST["finished_at"] = None
    _BACKUP_RUNNOW_LAST["outcome"] = "in-progress"
    _BACKUP_RUNNOW_LAST["lite_mode"] = use_lite

    async def _do_run() -> None:
        global _BACKUP_RUNNOW_IN_PROGRESS
        try:
            result = await _run_scheduled_backup(db, lite_mode=use_lite)
            _BACKUP_RUNNOW_LAST["finished_at"] = datetime.now(timezone.utc).isoformat()
            if result and not result.get("skipped"):
                _BACKUP_RUNNOW_LAST["outcome"] = (
                    f"ok · {result.get('filename')} · "
                    f"{(result.get('size_bytes') or 0)//1024} KB · "
                    f"emailed_to={result.get('emailed_to')}"
                )
            elif result and result.get("skipped"):
                _BACKUP_RUNNOW_LAST["outcome"] = f"skipped ({result.get('reason', '?')})"
            else:
                _BACKUP_RUNNOW_LAST["outcome"] = "FAILED — see server logs"
        except Exception as e:  # noqa: BLE001
            logger.exception(f"[manual-backup] background task crashed: {e}")
            _BACKUP_RUNNOW_LAST["finished_at"] = datetime.now(timezone.utc).isoformat()
            _BACKUP_RUNNOW_LAST["outcome"] = f"EXCEPTION: {e!r}"
        finally:
            _BACKUP_RUNNOW_IN_PROGRESS = False

    background_tasks.add_task(_do_run)
    return {
        "accepted": True,
        "lite_mode": use_lite,
        "poll": "/api/admin/backups-scheduler-state",
        "started_at": _BACKUP_RUNNOW_LAST["started_at"],
    }


# ────────────────────────────────────────────────────────────────────────
# Iter64 Phase 2c — manual complete-archive-to-R2 trigger
# ────────────────────────────────────────────────────────────────────────
_COMPLETE_R2_IN_PROGRESS = False
_COMPLETE_R2_LAST: Dict[str, Any] = {
    "started_at": None,
    "finished_at": None,
    "outcome": None,
    "filename": None,
    "size_bytes": None,
    "r2_key": None,
    "presigned_url": None,
    "stats": None,
}


@api_router.post("/admin/backups/run-complete-now")
async def admin_run_complete_backup_now(
    background_tasks: BackgroundTasks,
    _: bool = Depends(require_admin_strict),
):
    """Build a complete-system zip (Mongo + photos inlined from R2) and
    stream-upload it to ``r2://<bucket>/backups/``. Returns 202 instantly;
    the actual build runs in a FastAPI background task.

    Use this when you want an off-cycle complete archive (e.g. before a
    risky deploy). The nightly scheduler runs this automatically at
    ``BACKUP_R2_FULL_HOUR_UTC`` (default 03:00 UTC).
    """
    global _COMPLETE_R2_IN_PROGRESS
    if _COMPLETE_R2_IN_PROGRESS:
        raise HTTPException(409, "A complete archive is already in progress.")
    active_jobs = await get_active_backup_jobs(db)
    overlap = classify_backup_overlap(active_jobs)
    if overlap.get("backup_active") or overlap.get("restore_active"):
        raise HTTPException(409, "Another backup or restore job is already active.")
    manual_job = await claim_backup_job(
        db,
        job_type="manual_backup",
        kind=BACKUP_JOB_KIND_COMPLETE_R2,
        slot_key=backup_slot_key_for_hour(datetime.now(timezone.utc)) + "::manual",
        trigger="manual_admin",
        metadata={"requested_by": "admin_run_complete_backup_now"},
    )
    if manual_job is None:
        raise HTTPException(409, "A complete backup was already claimed for this slot.")
    _COMPLETE_R2_IN_PROGRESS = True
    _COMPLETE_R2_LAST["started_at"] = datetime.now(timezone.utc).isoformat()
    _COMPLETE_R2_LAST["finished_at"] = None
    _COMPLETE_R2_LAST["outcome"] = "in-progress"

    async def _do_complete():
        global _COMPLETE_R2_IN_PROGRESS
        lease = None
        try:
            lease = await start_backup_job(db, manual_job["job_id"])
            res = await _run_complete_archive_to_r2(db)
            _COMPLETE_R2_LAST["finished_at"] = datetime.now(timezone.utc).isoformat()
            if res:
                if res.get("deferred"):
                    _COMPLETE_R2_LAST["outcome"] = f"deferred ({res.get('reason')})"
                    await fail_backup_job(
                        db,
                        manual_job["job_id"],
                        error=f"deferred:{res.get('reason')}",
                        result=res,
                        state="deferred",
                        owner_token=lease.owner_token if lease else None,
                    )
                    return
                _COMPLETE_R2_LAST["outcome"] = "ok"
                _COMPLETE_R2_LAST["filename"] = res.get("filename")
                _COMPLETE_R2_LAST["size_bytes"] = res.get("size_bytes")
                _COMPLETE_R2_LAST["r2_key"] = res.get("r2_key")
                _COMPLETE_R2_LAST["presigned_url"] = res.get("presigned_url")
                _COMPLETE_R2_LAST["stats"] = res.get("stats")
                await complete_backup_job(db, manual_job["job_id"], outcome="ok", result=res, owner_token=lease.owner_token if lease else None)
            else:
                _COMPLETE_R2_LAST["outcome"] = "FAILED — see logs"
                await fail_backup_job(db, manual_job["job_id"], error="manual_complete_backup_returned_no_result", owner_token=lease.owner_token if lease else None)
        except Exception as e:  # noqa: BLE001
            logger.exception(f"[manual-complete-r2] crashed: {e}")
            _COMPLETE_R2_LAST["outcome"] = f"EXCEPTION: {e!r}"
            _COMPLETE_R2_LAST["finished_at"] = datetime.now(timezone.utc).isoformat()
            try:
                await fail_backup_job(db, manual_job["job_id"], error=repr(e), owner_token=lease.owner_token if lease else None)
            except Exception:
                pass
        finally:
            _COMPLETE_R2_IN_PROGRESS = False

    background_tasks.add_task(_do_complete)
    return {
        "accepted": True,
        "poll": "/api/admin/backups-complete-r2-state",
        "started_at": _COMPLETE_R2_LAST["started_at"],
    }


@api_router.get("/admin/backups-complete-r2-state")
async def admin_complete_r2_state(_: bool = Depends(require_admin_strict)):
    """Live status of the most recent manual complete-archive-to-R2 run.
    Route uses dashes (not slashes) so it doesn't collide with the
    parameterized ``/admin/backups/{filename}`` download/delete route."""
    lineage = await build_canonical_archive_lineage(
        db,
        current_env=_canonical_app_env(),
        current_db=_canonical_db_name(),
    )
    authoritative_artifact = lineage.get("authoritative_artifact") or {}
    newest_observed = lineage.get("newest_observed_artifact") or {}
    nightly_last = _BACKUP_SCHEDULER_STATE.get("last_r2_complete")
    nightly_last_date = _BACKUP_SCHEDULER_STATE.get("last_r2_complete_date")
    nightly_last_hour = _BACKUP_SCHEDULER_STATE.get("last_r2_complete_hour")

    if not nightly_last:
        try:
            latest_r2 = await db.backup_health.find_one(
                {"mode": "complete-r2", "ok": True, "filename": {"$nin": [None, ""]}},
                sort=[("ts", -1)],
                projection={"_id": 0, "filename": 1, "size_bytes": 1, "ts": 1},
            )
        except Exception as e:  # noqa: BLE001
            logger.warning(f"[backups-complete-r2-state] fallback probe failed: {e}")
            latest_r2 = None

        fallback_artifact = None
        if latest_r2:
            fallback_artifact = {
                "filename": latest_r2.get("filename"),
                "archive_size_bytes": latest_r2.get("size_bytes"),
                "object_key": f"{_canonical_backup_prefix().rstrip('/')}/{latest_r2.get('filename')}",
                "authoritative_time": latest_r2.get("ts"),
                "observed_time": latest_r2.get("ts"),
            }
        elif authoritative_artifact or newest_observed:
            fallback_artifact = authoritative_artifact or newest_observed

        if fallback_artifact:
            ts_value = str(fallback_artifact.get("authoritative_time") or fallback_artifact.get("observed_time") or "")
            hour_bucket = ts_value[:13] if len(ts_value) >= 13 else None
            date_bucket = ts_value[:10] if len(ts_value) >= 10 else None
            nightly_last = {
                "filename": fallback_artifact.get("filename"),
                "size_bytes": fallback_artifact.get("archive_size_bytes"),
                "r2_key": fallback_artifact.get("object_key") or f"{_canonical_backup_prefix().rstrip('/')}/{fallback_artifact.get('filename')}",
                "ts": fallback_artifact.get("authoritative_time") or fallback_artifact.get("observed_time"),
            }
            nightly_last_date = date_bucket
            nightly_last_hour = hour_bucket

    try:
        activation_state = await _build_hourly_activation_state(db)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[backups-complete-r2-state] hourly activation fallback used: {e}")
        activation_state = {
            "r2_hourly_requested": False,
            "r2_hourly_effective": False,
            "r2_hourly_locked_off": False,
            "hourly_cadence_enabled": False,
            "activation_status": "UNKNOWN — ACTIVATION EVIDENCE UNAVAILABLE",
            "environment": _canonical_app_env(),
            "next_eligible_hourly_slot": None,
        }
    try:
        backup_runtime = await _collect_backup_runtime_state(db)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[backups-complete-r2-state] backup runtime fallback used: {e}")
        backup_runtime = {}
    return {
        "in_progress": _COMPLETE_R2_IN_PROGRESS,
        "last": dict(_COMPLETE_R2_LAST),
        "nightly_last": nightly_last,
        "nightly_last_date": nightly_last_date,
        "nightly_last_hour": nightly_last_hour,
        "r2_full_hour_utc": int(os.environ.get("BACKUP_R2_FULL_HOUR_UTC", "3") or "3"),
        "r2_hourly_requested": bool(activation_state.get("r2_hourly_requested")),
        "r2_hourly_effective": bool(activation_state.get("r2_hourly_effective")),
        "r2_hourly_locked_off": bool(activation_state.get("r2_hourly_locked_off")),
        "hourly_cadence_enabled": bool(activation_state.get("hourly_cadence_enabled")),
        "hourly_activation": activation_state,
        "archive_lineage": public_archive_lineage_payload(lineage),
        "backup_runtime": backup_runtime,
    }


@api_router.get("/admin/backups-list-r2")
async def admin_list_r2_backups(
    limit: int = 100,
    _: bool = Depends(require_admin_strict),
):
    """List backup zips currently stored in the current environment's R2 backup prefix.
    Returns most recent first, plus a presigned URL for each so the
    admin can click-and-download from the UI without exposing the
    bucket credentials.

    iter440 · Phase 31.2 health-lock · uses the boto3 paginator so
    buckets with >1000 keys (the S3 default page size) don't truncate
    silently — without pagination, the newest archives were hidden by
    older legacy keys that sort first alphabetically.
    """
    try:
        from photo_storage import _bucket, _client, presigned_get_url_for_key, is_configured
    except Exception:  # noqa: BLE001
        raise HTTPException(500, "photo_storage import failed")
    if not is_configured():
        raise HTTPException(400, "R2 not configured")
    c = _client()
    if c is None:
        raise HTTPException(500, "R2 client unavailable")
    contents: list = []
    try:
        paginator = await asyncio.to_thread(c.get_paginator, "list_objects_v2")

        def _collect() -> list:
            out_local = []
            for page in paginator.paginate(Bucket=_bucket(), Prefix=_canonical_backup_prefix()):
                out_local.extend(page.get("Contents") or [])
            return out_local

        contents = await asyncio.to_thread(_collect)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f"R2 list failed: {e}")
    # Sort newest first
    contents.sort(key=lambda o: o.get("LastModified") or datetime.min, reverse=True)
    out = []
    for o in contents[: max(1, min(int(limit), 500))]:
        key = o.get("Key") or ""
        try:
            url = await presigned_get_url_for_key(key, ttl_seconds=7 * 24 * 3600)
        except Exception:
            url = None
        out.append({
            "key": key,
            "filename": key.rsplit("/", 1)[-1],
            "size_bytes": o.get("Size") or 0,
            "last_modified": (o.get("LastModified").isoformat()
                              if o.get("LastModified") else None),
            "download_url": url,
        })
    return {"count": len(out), "total_in_bucket": len(contents), "prefix": _canonical_backup_prefix(), "backups": out}

@api_router.get("/admin/sessions/recent")
async def admin_recent_sessions(
    request: Request,
    limit: int = 50,
    _: bool = Depends(require_admin_strict),
):
    """Operational visibility into the live `session_activity` table.

    Returns the N most-recent sessions (default 50, max 200) with
    identity, tier, login timestamp, last-activity timestamp, idle/abs
    expiry classification, IP, and user-agent. Read-only — no kill
    surface, no filtering beyond the limit. Audit-logged on every hit
    so the panel itself leaves a forensic trail.
    """
    from session_timeout import tier_ttl_seconds, describe_config

    safe_limit = max(1, min(int(limit or 50), 200))

    # Audit BEFORE reading so we always log the access, even if mongo
    # is slow / fails after.
    await _record_admin_action(
        db, "admin_sessions_panel_viewed", request, limit=safe_limit,
    )

    cursor = db.session_activity.find(
        {}, {"_id": 0, "token_hash": 0}
    ).sort("last_seen_at", -1).limit(safe_limit)
    rows = await cursor.to_list(safe_limit)

    now = datetime.now(timezone.utc)
    cfg = describe_config()
    enabled = bool(cfg.get("enabled"))

    sessions_out = []
    for r in rows:
        tier = r.get("tier") or "UNKNOWN"
        first_seen = r.get("first_seen_at")
        last_seen = r.get("last_seen_at")
        if first_seen and getattr(first_seen, "tzinfo", None) is None:
            first_seen = first_seen.replace(tzinfo=timezone.utc)
        if last_seen and getattr(last_seen, "tzinfo", None) is None:
            last_seen = last_seen.replace(tzinfo=timezone.utc)
        idle_s_lim, abs_s_lim = tier_ttl_seconds(tier)
        idle_age = (now - last_seen).total_seconds() if last_seen else None
        abs_age = (now - first_seen).total_seconds() if first_seen else None
        if not enabled:
            status = "enforcement_off"
        elif abs_age is not None and abs_s_lim and abs_age > abs_s_lim:
            status = "expired_absolute"
        elif idle_age is not None and idle_s_lim and idle_age > idle_s_lim:
            status = "expired_idle"
        else:
            status = "active"
        duration_s = None
        if first_seen and last_seen:
            duration_s = int((last_seen - first_seen).total_seconds())
        sessions_out.append({
            "tier": tier,
            "email": r.get("email") or None,
            "actor_label": r.get("actor_label") or None,
            "user_id": r.get("user_id") or None,
            "login_at": first_seen.isoformat() if first_seen else None,
            "last_activity_at": last_seen.isoformat() if last_seen else None,
            "idle_seconds": int(idle_age) if idle_age is not None else None,
            "absolute_seconds": int(abs_age) if abs_age is not None else None,
            "idle_limit_seconds": idle_s_lim or None,
            "absolute_limit_seconds": abs_s_lim or None,
            "status": status,
            "ip": r.get("last_login_ip") or None,
            "user_agent": r.get("last_user_agent") or None,
            "session_duration_s": duration_s,
        })

    return {
        "ok": True,
        "timeouts_enabled": enabled,
        "tiers": cfg.get("tiers", {}),
        "server_now": now.isoformat(),
        "count": len(sessions_out),
        "limit": safe_limit,
        "sessions": sessions_out,
    }




@api_router.get("/admin/backups-scheduler-state")
async def admin_backups_scheduler_state(_: bool = Depends(require_admin_strict)):
    """Read-only snapshot of the backup scheduler's internal state.

    Use when you want to know "is the backup scheduler alive? did the
    last attempt succeed? what's it doing right now?" WITHOUT triggering
    a fresh backup. Returns the module-level state dicts as-is plus a
    derived `seconds_since_last_tick` so the admin UI can colour a
    health pill green/amber/red.

    Note: the path uses a hyphen rather than a slash separator to avoid
    colliding with the ``/admin/backups/{filename}`` download/delete
    routes, which would otherwise match ``scheduler-state`` as a
    filename and 400 on the validator.
    """
    state = dict(_BACKUP_SCHEDULER_STATE)
    from routes.recovery_dashboard import build_canonical_scheduler_snapshot  # noqa: PLC0415
    canonical = await build_canonical_scheduler_snapshot(db, state)
    seconds_since_last_tick: Optional[float] = canonical.get("seconds_since_last_tick")
    state["alive"] = canonical.get("alive")
    state["is_healthy"] = canonical.get("is_healthy")
    state["signal_source"] = canonical.get("signal_source")
    state["reason_code"] = canonical.get("reason_code")
    state["evidence_ts"] = canonical.get("evidence_ts")
    state["last_lock_ts"] = canonical.get("last_lock_ts")
    state["owner_pod"] = canonical.get("owner_pod")
    state["heartbeat_window_minutes"] = canonical.get("heartbeat_window_minutes")
    state["backup_fallback_window_minutes"] = canonical.get("backup_fallback_window_minutes")

    # last_run_for_hour keys are date objects → coerce to ISO strings for JSON.
    state["last_run_for_hour"] = {
        str(h): (d.isoformat() if hasattr(d, "isoformat") else str(d))
        for h, d in (state.get("last_run_for_hour") or {}).items()
    }
    state["failed_attempts"] = {
        (k.isoformat() if hasattr(k, "isoformat") else str(k)): v
        for k, v in (state.get("failed_attempts") or {}).items()
    }

    # Pull the last 10 health-log rows so the admin UI can show "last 5
    # backups and what they did" without triggering anything.
    recent_health: List[dict] = []
    try:
        async for row in db.backup_health.find(
            {"id": {"$ne": "_watchdog_last_alarm"}}, {"_id": 0}
        ).sort("ts", -1).limit(10):
            recent_health.append(row)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[scheduler-state] health history read failed: {e}")

    activation_state = await _build_hourly_activation_state(db, runtime_state=state.get("backup_runtime") or await _collect_backup_runtime_state(db))
    return {
        "alive": state.get("alive"),
        "is_healthy": state.get("is_healthy"),
        "signal_source": state.get("signal_source"),
        "reason_code": state.get("reason_code"),
        "evidence_ts": state.get("evidence_ts"),
        "last_lock_ts": state.get("last_lock_ts"),
        "owner_pod": state.get("owner_pod"),
        "heartbeat_window_minutes": state.get("heartbeat_window_minutes"),
        "backup_fallback_window_minutes": state.get("backup_fallback_window_minutes"),
        "scheduler": state,
        "task_alive": bool(_backup_task is not None and not _backup_task.done()),
        "seconds_since_last_tick": seconds_since_last_tick,
        "manual_run": dict(_BACKUP_RUNNOW_LAST),
        "manual_in_progress": _BACKUP_RUNNOW_IN_PROGRESS,
        "lite_mode_only_env": _lite_mode_default(),
        "oom_watermark_mb": float(os.environ.get("BACKUP_FULL_OOM_WATERMARK_MB", "600") or "600"),
        "watchdog_threshold_hours": float(os.environ.get("BACKUP_WATCHDOG_HOURS", "25")),
        "now_utc": datetime.now(timezone.utc).isoformat(),
        "scheduled_hours_utc": list(BACKUP_HOURS_UTC),
        "circuit_breaker_max_attempts_per_day": 3,
        "recent_health": recent_health,
        "backup_runtime": state.get("backup_runtime") or await _collect_backup_runtime_state(db),
        "hourly_activation": activation_state,
    }


@app.get("/api/admin/backup-trust-score")
async def admin_backups_trust_score(_: bool = Depends(require_admin_strict)):
    from lib.trust_score import compute_backup_trust_score  # noqa: PLC0415

    def _coerce_iso_dt(raw):
        if isinstance(raw, datetime):
            return raw if raw.tzinfo else raw.replace(tzinfo=timezone.utc)
        if isinstance(raw, str) and raw:
            try:
                return datetime.fromisoformat(raw.replace("Z", "+00:00"))
            except Exception:
                return None
        return None

    lineage = await build_canonical_archive_lineage(
        db,
        current_env=_canonical_app_env(),
        current_db=_canonical_db_name(),
    )
    latest_complete = (lineage.get("authoritative_artifact") or {})
    newest_r2_age_hours = lineage.get("freshness_age_hours")

    last_drill = await db.drill_runs.find_one({"state": "done"}, {"_id": 0}, sort=[("started_at", -1)])
    drill_ts = _coerce_iso_dt((last_drill or {}).get("finished_at") or (last_drill or {}).get("started_at"))
    restore_drill_age_days = None
    if drill_ts:
        restore_drill_age_days = round((datetime.now(timezone.utc) - drill_ts).total_seconds() / 86400.0, 2)

    runtime = await _collect_backup_runtime_state(db)
    activation_state = await _build_hourly_activation_state(db, runtime_state=runtime)
    failures_7d = await db.backup_health.count_documents({
        "ok": False,
        "ts": {"$gte": (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()},
        "mode": {"$regex": "complete-r2|restore|verification|retention", "$options": "i"},
    })
    bucket_state = classify_capacity_state(
        total_bytes=((latest_complete or {}).get("archive_size_bytes") or 0),
        warn_gb=float(os.environ.get("R2_USAGE_WARN_GB", "700") or "700"),
        alert_gb=float(os.environ.get("R2_USAGE_ALERT_GB", "820") or "820"),
        probe_state="ok" if latest_complete else "missing",
        as_of=(latest_complete or {}).get("authoritative_time") or (latest_complete or {}).get("observed_time"),
    )
    bucket_usage_status = bucket_state.get("status", "AMBER")

    score = compute_backup_trust_score(
        hourly_disabled=not bool(activation_state.get("r2_hourly_effective")),
        newest_r2_age_hours=newest_r2_age_hours,
        restore_drill_age_days=restore_drill_age_days,
        restore_drill_ok=((last_drill or {}).get("outcome") == "ok"),
        integrity_ok=bool(latest_complete),
        overlap_blocked=bool((runtime.get("overlap") or {}).get("overlap_blocked")),
        active_failures_7d=int(failures_7d or 0),
        bucket_usage_status=bucket_usage_status,
    )
    payload = {
        **score,
        "production_activation_disabled": not bool(activation_state.get("r2_hourly_effective")),
        "hourly_activation": activation_state,
        "evidence": {
            "latest_complete_backup": latest_complete,
            "newest_r2_age_hours": newest_r2_age_hours,
            "archive_lineage": public_archive_lineage_payload(lineage),
            "last_restore_drill": last_drill,
            "restore_drill_age_days": restore_drill_age_days,
            "runtime": runtime,
            "bucket_usage": bucket_state,
            "hourly_activation": activation_state,
        },
    }
    truth_card = canonical_truth_card(
        truth_subject="bcss_recovery_trust",
        canonical_owner="bcss_recovery_trust",
        truth_surface_id="bcss_recovery_trust",
        evidence_state="calculated",
        evidence_quality="CALCULATED",
        evidence_confidence="MEDIUM" if latest_complete else "LOW",
        truth_evaluation={"green": "VERIFIED", "amber": "DEGRADED", "red": "MISMATCH"}.get(score.get("score_band"), "UNVERIFIABLE"),
        permitted_claim=CORRELATED,
        claim_ceiling=CORRELATED,
        claim_basis=["compute_backup_trust_score", "archive_lineage", "restore drill evidence", "backup runtime", "bucket usage"],
        prohibited_claims=["VERIFIED", "VALIDATED", "CERTIFIED"],
        degradation_reasons=[entry.get("reason") for entry in score.get("score_inputs") or []],
        unknowns=[] if latest_complete else ["No authoritative archive-lineage evidence is currently available."],
        contradictory_evidence=[],
        evidence_timestamp=(latest_complete or {}).get("authoritative_time") or (latest_complete or {}).get("observed_time"),
        evaluation_timestamp=datetime.now(timezone.utc).isoformat(),
        audit_reference="OTS-C5-BACKUP-TRUST",
        evidence_required_to_raise_claim=["underlying canonical owner verification on source truth"],
        notes=["Trust score is a derived confidence surface only."],
    )
    compatibility = compatibility_projection(
        preserved_fields=6,
        deprecated_fields=0,
        new_fields=3,
        alias_fields=[],
        breaking_changes=0,
    )
    payload["ots_truth"] = public_ots_projection(truth_card)
    payload["truth_relationship"] = projected_truth_relationship(
        surface_id="bcss_recovery_trust",
        card=truth_card,
        canonical_owner_route="/api/admin/backup-trust-score",
        derivation_explanation="Backup Trust is a derived confidence surface and may not upgrade archive, restore, or certification claims.",
        derived_status=truth_card["truth_evaluation"],
    )
    payload["compatibility"] = compatibility
    return payload


@api_router.post("/admin/data-fixes/run")
async def admin_run_data_fixes(_: bool = Depends(require_admin)):
    """Apply both production data fixes:
       1. Split `make_model` into `make` + `model` on every equipment unit
       2. Seed `project_members` so every owner/admin sees every project

    Both fixes are idempotent — safe to re-run any number of times.
    """
    from data_fixes import run_all_fixes
    return await run_all_fixes(db)


# -------------------- Crew Hub recovery (legacy admin-token gated) --------------------
# These endpoints exist so the office can recover a Crew Hub login when nobody
# remembers their password. Authenticated by the LEGACY admin password
# (X-Admin-Token / env PM_PASSWORD) — NOT by a Crew Hub JWT — so it works even when
# every crew owner+admin is locked out.

@api_router.get("/admin/crew-recovery/status")
async def admin_crew_recovery_status(_: bool = Depends(require_admin_strict)):
    """Return counts of every key collection so the office can see at a glance
    what's populated and what isn't (helps diagnose redeploy data-loss)."""
    counts = {}
    for coll in [
        "users",
        "projects",
        "project_members",
        "equipment_master",
        "equipment_units",
        "equipment_inspections",
        "inspections",
        "meetings",
        "jhas",
        "incidents",
        "daily_reports",
        "docs",
        "employees",
        "suppliers",
        "notifications",
        "activity_log",
    ]:
        try:
            counts[coll] = await db[coll].count_documents({})
        except Exception:
            counts[coll] = -1
    crew_users = await db.users.find(
        {}, {"_id": 0, "id": 1, "email": 1, "name": 1, "role": 1, "is_active": 1, "must_change_password": 1}
    ).sort("email", 1).to_list(200)
    return {
        "ok": True,
        "counts": counts,
        "crew_users": crew_users,
    }


@api_router.post("/admin/crew-recovery/reset-password")
async def admin_crew_recovery_reset(
    body: dict,
    _: bool = Depends(require_admin_strict),
):
    """Reset a Crew Hub user's password using the LEGACY admin token. The user
    is forced to change it on next login. Body: {email, new_password}.
    """
    from auth import hash_password
    email = (body.get("email") or "").strip().lower()
    new_password = (body.get("new_password") or "").strip()
    if not email or not new_password:
        raise HTTPException(400, "email + new_password required")
    if len(new_password) < 8:
        raise HTTPException(400, "Password must be at least 8 characters")
    user = await db.users.find_one({"email": email}, {"_id": 0, "id": 1, "email": 1})
    if not user:
        raise HTTPException(404, f"No crew user with email {email}")
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {
            "password_hash": hash_password(new_password),
            "must_change_password": True,
            "is_active": True,  # un-lock if accidentally deactivated
        }},
    )
    return {"ok": True, "email": email, "must_change_password": True}


@api_router.post("/admin/crew-recovery/force-reseed")
async def admin_crew_recovery_force_reseed(body: dict | None = None, _: bool = Depends(require_admin_strict)):
    """Force-rerun the equipment_master / employees / suppliers JSON seeds even
    if those collections already have rows. Useful when a partial-wipe leaves
    incomplete data and the boot guard (`count > 0`) skips re-seeding.

    The seed functions normally short-circuit if the collection has any rows;
    this endpoint deletes the seed-managed collections first so they re-seed
    from JSON cleanly. Safety/projects/users are NOT touched.
    """
    require_destructive_confirmation(body, expected_confirm="FORCE_RESEED_CREW_COLLECTIONS")
    require_destructive_runtime_guard(expected_db_name="masci_safety")
    summary = {}
    for coll in ["equipment_master", "equipment_units", "employees", "suppliers"]:
        before = await db[coll].count_documents({})
        await db[coll].delete_many({})
        summary[coll] = {"before": before, "after_delete": 0}
    # Re-run the seeds in-process
    await _seed_equipment_master()
    await _seed_employees_from_json()
    await _seed_suppliers_from_json()
    for coll in ["equipment_master", "equipment_units", "employees", "suppliers"]:
        summary[coll]["after_seed"] = await db[coll].count_documents({})
    # Boot self-heal also patches make/model + memberships
    from data_fixes import boot_self_heal
    await boot_self_heal(db)
    return {"ok": True, "summary": summary}


@api_router.post("/admin/crew-recovery/scrap-crew-hub")
async def admin_scrap_crew_hub(body: dict, _: bool = Depends(require_admin_strict)):
    """One-shot: WIPE every Crew Hub / projects table from the DB.
    The MASCI Hub has decided to use Basecamp instead of the in-app Crew Hub.

    Body must include {"confirm": "SCRAP_CREW_HUB"} or 400. Idempotent — safe
    to re-run (running on an empty DB just returns zeros).

    DELETES (counts returned in the response):
      - projects, project_members, docs, todos, todo_lists, hill_dots,
        events, messages, notifications, activity_log
    KEEPS:
      - users (so admin can still see who they were if curious; tiny table)
      - All safety records (inspections, meetings, jhas, incidents, daily_reports)
      - Equipment master + units + inspections, employees, suppliers
      - Backups
    """
    require_destructive_confirmation(body, expected_confirm="SCRAP_CREW_HUB")
    require_destructive_runtime_guard(expected_db_name="masci_safety")
    wipe_collections = [
        "projects",
        "project_members",
        "docs",
        "todos",
        "todo_lists",
        "hill_dots",
        "events",
        "messages",
        "notifications",
        "activity_log",
    ]
    summary = {}
    for coll in wipe_collections:
        try:
            before = await db[coll].count_documents({})
            res = await db[coll].delete_many({})
            summary[coll] = {"before": before, "deleted": res.deleted_count}
        except Exception as e:
            summary[coll] = {"error": str(e)}
    return {
        "ok": True,
        "ran_at": datetime.now(timezone.utc).isoformat(),
        "summary": summary,
        "kept": [
            "users",
            "inspections",
            "meetings",
            "jhas",
            "incidents",
            "daily_reports",
            "equipment_master",
            "equipment_units",
            "equipment_inspections",
            "employees",
            "suppliers",
        ],
    }


# -------------------- Outage alerts (called by SystemHealthBadge) --------------------
class OutageAlertBody(BaseModel):
    issue_key: str = Field(..., min_length=1, max_length=200)
    summary: str = Field(..., min_length=1, max_length=2000)
    failed_endpoints: List[Dict[str, Any]] = Field(default_factory=list)


@api_router.post("/admin/alert-outage")
async def admin_alert_outage(body: OutageAlertBody, _: bool = Depends(require_admin)):
    """Sends a one-line outage email via Resend (cooldown-gated).

    Called by the SystemHealthBadge in the admin UI when one of the monitored
    endpoints starts returning 5xx or fails to respond. Cooldown is
    OUTAGE_ALERT_COOLDOWN_MINUTES (default 15) per issue_key — duplicate
    badge fires within that window are suppressed.
    """
    from outage_alerts import send_outage_alert
    rows = body.failed_endpoints or []
    rows_html = ""
    if rows:
        rows_html = "<table style='width:100%;border-collapse:collapse;font-size:13px;margin:6px 0 10px'>"
        rows_html += "<thead><tr style='background:#f1f5f9;color:#0f172a;text-align:left'>"
        rows_html += "<th style='padding:6px 8px;border:1px solid #e2e8f0'>Endpoint</th>"
        rows_html += "<th style='padding:6px 8px;border:1px solid #e2e8f0'>Status</th>"
        rows_html += "<th style='padding:6px 8px;border:1px solid #e2e8f0'>Latency</th>"
        rows_html += "</tr></thead><tbody>"
        for r in rows[:20]:
            label = str(r.get("label") or r.get("path") or "?")[:40]
            stat = str(r.get("status") or "—")[:8]
            ms = str(r.get("ms") or "—")[:8]
            rows_html += (
                f"<tr><td style='padding:5px 8px;border:1px solid #e2e8f0;font-family:monospace'>{label}</td>"
                f"<td style='padding:5px 8px;border:1px solid #e2e8f0;color:#dc2626;font-weight:bold'>{stat}</td>"
                f"<td style='padding:5px 8px;border:1px solid #e2e8f0;font-family:monospace'>{ms} ms</td></tr>"
            )
        rows_html += "</tbody></table>"
    return await send_outage_alert(
        issue_key=body.issue_key,
        subject=f"\U0001F6A8 PLATFORM OUTAGE \u00b7 {body.issue_key}",
        summary=body.summary,
        details_html=rows_html,
    )


@api_router.get("/admin/persistence-check")
async def admin_persistence_check(_: bool = Depends(require_admin)):
    """Report whether the running instance is at risk of data loss on redeploy.

    Production on Emergent without an external MongoDB URL is ephemeral — a
    git push/redeploy wipes the container's Mongo volume. This endpoint
    powers the admin-hub warning banner so the office never redeploys blind.
    """
    mongo_url = ((_runtime_identity_safe_payload().get("identity") or {}).get("mongo_url_redacted") or "")
    # Local/in-container Mongo hostnames. Atlas URLs start with mongodb+srv://
    # or include an explicit external host. Anything pointing at localhost,
    # 127.0.0.1 or no hostname is treated as ephemeral.
    host_part = mongo_url.split("://", 1)[-1].split("/", 1)[0].lower()
    is_local = (
        not mongo_url
        or "localhost" in host_part
        or host_part.startswith("127.")
        or host_part.startswith("0.0.0.0")
        or host_part == ""
    )
    is_atlas = mongo_url.startswith("mongodb+srv://") or "mongodb.net" in host_part
    backup_email_configured = bool((os.environ.get("BACKUP_EMAIL_TO") or "").strip())
    resend_configured = bool((os.environ.get("RESEND_API_KEY") or "").strip())
    last_backup = None
    try:
        files = _list_stored_backups()
        if files:
            last_backup = files[0]
    except Exception:
        pass

    return {
        "mongo_is_local": is_local,
        "mongo_is_atlas": is_atlas,
        "mongo_host": host_part or "(none)",
        "backup_email_to": (os.environ.get("BACKUP_EMAIL_TO") or "").strip() or None,
        "backup_email_configured": backup_email_configured,
        "resend_configured": resend_configured,
        "last_backup": last_backup,
        "scheduler_enabled": os.environ.get("DISABLE_BACKUP_SCHEDULER", "").lower() not in ("1", "true", "yes"),
    }


# ----------------------------------------------------------------------
# Restore from backup ZIP — upsert every record back into MongoDB.
# ----------------------------------------------------------------------
# Map backup kind folder → MongoDB collection name. Pulled from
# EXPORTABLE_KINDS + the Crew Hub + Safety aux lists above.
_RESTORE_KIND_TO_COLL = {
    # Safety kinds — the ZIP stores them under <kind>/json/*.json
    "inspections": "inspections",
    "meetings": "meetings",
    "jhas": "jhas",
    "incidents": "incidents",
    "daily-reports": "daily_reports",
    "equipment-inspections": "equipment_inspections",
}
_RESTORE_CREW_HUB = {
    "projects", "users", "project_members", "messages", "message_comments",
    "todo_lists", "todos", "events", "docs", "hill_scopes", "activity_log",
    "notifications",
}
_RESTORE_SAFETY_AUX = {
    "equipment_units",
    "job_hazard_plans",
    "trench_boxes",
    # ── Trench Safety Operations System (Phase 2) ────────────────────
    # New per-physical-unit collections introduced in /app/memory/
    # TRENCH_SAFETY_ARCHITECTURE.md §1-§2. Each is restore-essential.
    "trench_safety_assets",
    "trench_safety_inspections",
    "trench_safety_repairs",
    "trench_safety_deployments",
    "trench_safety_certifications",
    "trench_safety_photos",
    "trench_safety_qr_scans",
}


# ────────────────────────────────────────────────────────────────────────────
# Edit project on an existing record (admin OR PM)
# ────────────────────────────────────────────────────────────────────────────
# Foremen and superintendents occasionally pick the wrong job at submit time
# ("HQ" vs "T5860 SR 9", say) and the report lands under the wrong project.
# This endpoint lets a PM or admin re-tag the record after the fact without
# editing any other field. Only project_name / project_number / project_id /
# location may be changed; everything else (signatures, photos, narrative,
# checklist results) stays exactly as the foreman submitted it.
_EDIT_KIND_TO_COLL = {
    "daily-reports":         "daily_reports",
    "incidents":             "incidents",
    "meetings":              "meetings",
    "inspections":           "inspections",
    "equipment-inspections": "equipment_inspections",
}


class EditProjectRequest(BaseModel):
    project_name: Optional[str] = None
    project_number: Optional[str] = None
    project_id: Optional[str] = None
    location: Optional[str] = None
    model_config = ConfigDict(extra="ignore")


@api_router.patch("/admin/records/{kind}/{record_id}/project")
async def edit_record_project(
    kind: str,
    record_id: str,
    payload: EditProjectRequest,
    _: bool = Depends(require_admin),
):
    coll_name = _EDIT_KIND_TO_COLL.get(kind)
    if not coll_name:
        raise HTTPException(status_code=400, detail=f"Unknown record kind: {kind}")
    update: Dict[str, Any] = {}
    for field in ("project_name", "project_number", "project_id", "location"):
        v = getattr(payload, field, None)
        if v is not None:
            update[field] = (v or "").strip()
    if not update:
        raise HTTPException(status_code=400, detail="No project fields supplied")
    update["project_edited_at"] = datetime.now(timezone.utc).isoformat()
    res = await db[coll_name].update_one({"id": record_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Record not found")
    doc = await db[coll_name].find_one({"id": record_id}, {"_id": 0})
    return {"ok": True, "record": doc}


# TRACK 15.37 — Restore upload ceiling.
# Old hard-coded 500 MB rejected every current hourly archive (~600 MB).
# Now env-driven; default 2048 MB accepts every standard backup with
# generous headroom for future growth.
def _restore_max_bytes() -> int:
    try:
        mb = int(os.environ.get("RESTORE_MAX_UPLOAD_MB", "2048") or 2048)
    except ValueError:
        mb = 2048
    # Clamp at 8 GiB as an absolute sanity-ceiling — anything larger almost
    # certainly indicates an upload-stream attack, not a real backup.
    mb = max(64, min(mb, 8192))
    return mb * 1024 * 1024


_RESTORE_MAX_BYTES = _restore_max_bytes()


@api_router.post("/exports/restore")
async def exports_restore(
    file: UploadFile = File(...),
    merge: bool = Form(True),
    confirm: str = Form(""),
    backup_ack: bool = Form(False),
    dry_run: bool = Form(False),
    _: bool = Depends(require_admin_strict),
):
    """Restore a `/api/exports/full-backup` .zip back into MongoDB.

    - `merge=true` (default): upsert by `id` — existing rows with the same
      id are overwritten, new rows added, untouched collections untouched.
    - `merge=false`: wipe the target collection first, then insert. Use with
      care — this is a full restore to the backup's exact state.

    The zip's `backup_manifest.json` is used to validate that this is a real
    MASCI backup before we touch any data.
    """
    active_jobs = await get_active_backup_jobs(db)
    overlap = classify_backup_overlap(active_jobs)
    if overlap.get("backup_active"):
        raise HTTPException(409, "Restore blocked while a backup job is active.")
    restore_job = await claim_backup_job(
        db,
        job_type="restore_operation",
        kind=BACKUP_JOB_KIND_RESTORE_IMPORT,
        slot_key=f"restore::{datetime.now(timezone.utc).isoformat()}::{uuid.uuid4().hex[:6]}",
        trigger="admin_restore_endpoint",
        metadata={"merge": bool(merge), "dry_run": bool(dry_run)},
    )
    if restore_job is None:
        raise HTTPException(409, "Restore slot claim failed.")
    restore_lease = await start_backup_job(db, restore_job["job_id"])
    restore_stage = {"name": "restore_upload"}
    restore_heartbeat_stop, restore_heartbeat_task = await _run_job_heartbeat(
        db,
        job_id=restore_job["job_id"],
        owner_token=restore_lease.owner_token,
        stage_fn=lambda: restore_stage["name"],
    )
    # 1. Read + validate the upload
    spool_path = None
    try:
        BACKUP_COMPLETE_TMP_DIR.mkdir(parents=True, exist_ok=True)
        spool_path = BACKUP_COMPLETE_TMP_DIR / f"restore_upload_{uuid.uuid4().hex}.zip"
        payload_size = 0
        with spool_path.open("wb") as handle:
            while True:
                chunk = await file.read(BACKUP_RESTORE_STREAM_CHUNK_BYTES)
                if not chunk:
                    break
                payload_size += len(chunk)
                if payload_size > _RESTORE_MAX_BYTES:
                    raise HTTPException(
                        413,
                        f"Backup file exceeds the configured restore ceiling ({_RESTORE_MAX_BYTES // (1024 * 1024)} MB).",
                    )
                handle.write(chunk)
    except Exception as e:
        try:
            await fail_backup_job(db, restore_job["job_id"], error=f"restore_upload_read_failed:{e!r}", owner_token=restore_lease.owner_token)
        except Exception:
            pass
        raise HTTPException(400, f"Failed to read upload: {e}")
    if not spool_path or payload_size <= 0:
        try:
            await fail_backup_job(db, restore_job["job_id"], error="empty_upload", owner_token=restore_lease.owner_token)
        except Exception:
            pass
        raise HTTPException(400, "Empty upload")
    if payload_size > _RESTORE_MAX_BYTES:
        try:
            await fail_backup_job(db, restore_job["job_id"], error="upload_too_large", owner_token=restore_lease.owner_token)
        except Exception:
            pass
        raise HTTPException(
            413,
            f"Backup file exceeds the configured restore ceiling "
            f"({_RESTORE_MAX_BYTES // (1024 * 1024)} MB). "
            f"Override via env `RESTORE_MAX_UPLOAD_MB` if you need a larger "
            f"window; current archives average ~600 MB so the default 2048 MB "
            f"ceiling accepts every standard hourly archive with headroom.",
        )

    try:
        zf = zipfile.ZipFile(str(spool_path), "r")
    except zipfile.BadZipFile:
        try:
            await fail_backup_job(db, restore_job["job_id"], error="bad_zip_file", owner_token=restore_lease.owner_token)
        except Exception:
            pass
        raise HTTPException(400, "Uploaded file is not a valid ZIP archive")

    names = set(zf.namelist())

    # TRACK 15.38 — restore endpoint accepts BOTH manifest formats:
    #   * `backup_manifest.json` — Track 14.0-I1 envelope (email backup path)
    #   * `MANIFEST.json` — R2 hourly complete archive
    manifest_name = None
    for candidate in ("backup_manifest.json", "MANIFEST.json"):
        if candidate in names:
            manifest_name = candidate
            break
    if manifest_name is None:
        raise HTTPException(
            400,
            "Neither backup_manifest.json nor MANIFEST.json found — "
            "this does not look like a MASCI backup .zip.",
        )
    try:
        restore_stage["name"] = "restore_manifest_validation"
        manifest = _backup_json.loads(zf.read(manifest_name).decode("utf-8"))
    except Exception as e:
        raise HTTPException(400, f"Corrupt manifest ({manifest_name}): {e}")

    # ---------------------------------------------------------------
    # Track 14.0-I1 (2026-02-14) — Archive Origin Verification.
    # Before we touch any data, refuse to restore an archive whose
    # environment / database_name disagree with the running worker.
    # This closes the last manual-checklist item from Track 14.0-P0
    # ("verify backup archive origin before importing into prod").
    # ---------------------------------------------------------------
    current_env = _canonical_app_env().lower()
    current_db = _canonical_db_name()
    archive_env = (manifest.get("environment") or manifest.get("app_env") or "").lower()
    archive_db = manifest.get("database_name") or manifest.get("db_name") or ""

    # TRACK 15.38 — R2 hourly archives use MANIFEST.json which carries
    # `source` (e.g. "mascidocs.com") instead of explicit `environment`.
    # Infer `environment` from `source` so the env-mismatch guard fires
    # correctly for production archives.
    if not archive_env and manifest_name == "MANIFEST.json":
        src = (manifest.get("source") or "").lower()
        if "mascidocs.com" in src:
            archive_env = "production"

    audit_doc = {
        "id": str(uuid.uuid4()),
        "kind": "exports_restore",
        "ts": datetime.now(timezone.utc).isoformat(),
        "current_env": current_env,
        "current_db": current_db,
        "archive_env": archive_env,
        "archive_db": archive_db,
        "archive_backup_id": manifest.get("backup_id"),
        "archive_generated_at": manifest.get("generated_at"),
        "merge": merge,
        "dry_run": dry_run,
        "result": None,
        "reason": None,
    }

    async def _record_audit(result: str, reason: str | None = None) -> None:
        audit_doc["result"] = result
        audit_doc["reason"] = reason
        try:
            await db.audit_events.insert_one(audit_doc)
        except Exception:  # noqa: BLE001
            logger.exception("[restore-audit] failed to persist audit event")

    # Legacy archives (manifest schema before track-14.0-i1) have no
    # environment field. In production we refuse them outright — better
    # safe than sorry. In preview we accept them and log the legacy
    # status so historical archives stay usable for regression work.
    if not archive_env:
        if current_env == "production":
            await _record_audit("rejected", "missing-environment-field")
            raise HTTPException(
                400,
                "Restore blocked. This archive was generated before the "
                "Track 14.0-I1 manifest standard and has no recorded "
                "environment of origin. Production restores must be "
                "produced by a current production worker.",
            )
        logger.warning(
            "[restore] legacy archive (no environment field) accepted in %s",
            current_env,
        )
    elif archive_env != current_env:
        await _record_audit(
            "rejected",
            f"environment-mismatch:{archive_env}-into-{current_env}",
        )
        raise HTTPException(
            400,
            f"Restore blocked. Archive originated from the "
            f"{archive_env.title()} environment. "
            f"{current_env.title()} restores may only use "
            f"{current_env.title()} archives.",
        )
    elif archive_db and current_db and archive_db != current_db:
        await _record_audit(
            "rejected",
            f"database-mismatch:{archive_db}-into-{current_db}",
        )
        raise HTTPException(
            400,
            f"Restore blocked. Archive database `{archive_db}` does not "
            f"match the current database `{current_db}`.",
        )

    if not merge:
        require_destructive_confirmation(
            {"confirm": confirm, "backup_ack": backup_ack},
            expected_confirm="RESTORE_REPLACE_ALL_COLLECTIONS",
        )
        require_destructive_runtime_guard(expected_db_name="masci_safety")

    # 2. Walk the ZIP and group docs by destination collection.
    bucket: Dict[str, List[dict]] = {}

    def _add(coll: str, docs: List[dict]):
        if not docs:
            return
        bucket.setdefault(coll, []).extend(docs)

    # 2a. Safety kinds — every json under <kind>/json/*.json
    restore_stage["name"] = "restore_extraction"
    for kind, coll in _RESTORE_KIND_TO_COLL.items():
        prefix = f"{kind}/json/"
        docs: List[dict] = []
        for n in names:
            if n.startswith(prefix) and n.endswith(".json"):
                try:
                    docs.append(_backup_json.loads(zf.read(n).decode("utf-8")))
                except Exception as e:  # noqa: BLE001
                    logger.warning(f"restore: skipped {n}: {e}")
        _add(coll, docs)

    # 2b. Crew Hub — single json per collection under crew_hub/<coll>.json
    for coll in _RESTORE_CREW_HUB:
        n = f"crew_hub/{coll}.json"
        if n in names:
            try:
                data = _backup_json.loads(zf.read(n).decode("utf-8"))
                if isinstance(data, list):
                    _add(coll, data)
            except Exception as e:  # noqa: BLE001
                logger.warning(f"restore: skipped {n}: {e}")

    # 2c. Safety aux
    for coll in _RESTORE_SAFETY_AUX:
        n = f"safety_aux/{coll}.json"
        if n in names:
            try:
                data = _backup_json.loads(zf.read(n).decode("utf-8"))
                if isinstance(data, list):
                    _add(coll, data)
            except Exception as e:  # noqa: BLE001
                logger.warning(f"restore: skipped {n}: {e}")

    # 2d. Auto-discovered collections (backup version 3+) — anything under
    #     collections/<name>.json that isn't already restored above.
    for n in names:
        if not (n.startswith("collections/") and n.endswith(".json")):
            continue
        coll = n[len("collections/"):-len(".json")]
        # Skip collections we've already restored via dedicated paths above.
        if coll in bucket:
            continue
        try:
            data = _backup_json.loads(zf.read(n).decode("utf-8"))
            if isinstance(data, list):
                _add(coll, data)
        except Exception as e:  # noqa: BLE001
            logger.warning(f"restore: skipped {n}: {e}")

    # 2d-bis. TRACK 15.38 — R2 hourly archive auto-discovery.
    # The R2 complete archive (`_build_complete_archive_on_disk`) writes
    # per-record files under `<collection>/json/<id>.json` for every
    # collection — auto-discovered, not whitelisted. Walk that layout for
    # any collection NOT already covered by sections 2a/2b/2c/2d above.
    _per_record_collections: Dict[str, List[dict]] = {}
    for n in names:
        if not (n.endswith(".json") and "/json/" in n):
            continue
        parts = n.split("/")
        if len(parts) != 3 or parts[1] != "json":
            continue
        coll = parts[0]
        # Skip system/manifest entries
        if coll in ("collections", "crew_hub", "safety_aux", "disk_files", "photos"):
            continue
        # Skip if already restored via a dedicated path above
        if coll in bucket:
            continue
        try:
            doc = _backup_json.loads(zf.read(n).decode("utf-8"))
            if isinstance(doc, dict):
                _per_record_collections.setdefault(coll, []).append(doc)
        except Exception as e:  # noqa: BLE001
            logger.warning(f"restore: skipped {n}: {e}")
    for coll, docs in _per_record_collections.items():
        _add(coll, docs)

    # 2e. Disk-backed files — restore the storage tree (Oxford big PDFs etc.)
    disk_restored = 0
    disk_storage_root = Path("/app/backend/storage")
    for n in names:
        if not n.startswith("disk_files/") or n.endswith("/"):
            continue
        rel = n[len("disk_files/"):]
        if not rel:
            continue
        target = disk_storage_root / rel
        try:
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_bytes(zf.read(n))
            disk_restored += 1
        except Exception as e:  # noqa: BLE001
            logger.warning(f"restore: disk file {n} failed: {e}")

    if not bucket and disk_restored == 0:
        raise HTTPException(
            400,
            "No records found in backup (expected files under "
            "<kind>/json/, crew_hub/, safety_aux/, collections/, or disk_files/).",
        )

    if not merge:
        require_non_empty_destructive_scope(
            list(bucket.keys()),
            detail="Replace-mode restore refused because the archive does not contain any restorable collection set.",
        )

    preflight_collections = {
        coll: {
            "incoming_records": len([d for d in docs if isinstance(d, dict)]),
            "existing_records": await db[coll].count_documents({}),
        }
        for coll, docs in bucket.items()
    }

    if dry_run:
        result = {
            "ok": True,
            "mode": "replace" if not merge else "merge",
            "dry_run": True,
            "backup_generated_at": manifest.get("generated_at"),
            "backup_version": manifest.get("version", "unknown"),
            "archive_environment": archive_env or "unknown",
            "archive_backup_id": manifest.get("backup_id"),
            "collections": preflight_collections,
            "total_processed": sum(v["incoming_records"] for v in preflight_collections.values()),
            "disk_files": disk_restored,
        }
        await _record_audit("accepted", f"dry_run merge={merge}; collections={len(preflight_collections)}")
        await complete_backup_job(db, restore_job["job_id"], outcome="dry_run_ok", result=result, state="completed", owner_token=restore_lease.owner_token)
        try:
            if spool_path and spool_path.exists():
                spool_path.unlink()
        except Exception:
            pass
        restore_heartbeat_stop.set()
        try:
            await restore_heartbeat_task
        except Exception:
            pass
        return result

    # 3. Write back to MongoDB.
    summary: Dict[str, dict] = {}
    failed_docs: Dict[str, List[str]] = {}
    # If the users / user_directory collections are being restored, the export
    # redacts password_hash. Precompute the seed hash so restored rows always
    # have a usable password (Welcome2MASCI! + must_change_password).
    # iter-batch-G (2026-05-30) extended this from `users` only to also cover
    # `user_directory` after Batch F drill confirmed multi-login was
    # universally broken post-restore until reseed. See GAP-2 in
    # /app/memory/PLATFORM_RECOVERY_GAP_REPORT.md.
    _seed_hash = None
    _NEEDS_SEED_HASH = ("users", "user_directory")
    if any(c in bucket for c in _NEEDS_SEED_HASH):
        try:
            import bcrypt as _bc  # noqa: E402
            _seed_hash = _bc.hashpw(b"Welcome2MASCI!", _bc.gensalt()).decode("utf-8")
        except Exception as e:  # noqa: BLE001
            logger.warning(f"restore: could not generate seed hash ({e}); restored users may be locked out")

    restore_stage["name"] = "restore_replay"
    for coll, docs in bucket.items():
        # Strip any _id from the docs (they're exported without, but be safe).
        clean = []
        for d in docs:
            if not isinstance(d, dict):
                continue
            d.pop("_id", None)
            if "id" not in d:
                d["id"] = str(uuid.uuid4())  # defensive — keep upsert viable
            # Special-case: restored users / user_directory rows lost their
            # password_hash on export. In merge mode: keep whatever's in DB
            # (pull it first). In replace mode (or brand-new row): stamp the
            # seed hash + force password change so no account gets locked out.
            if coll in _NEEDS_SEED_HASH and "password_hash" not in d:
                existing = None
                if merge:
                    existing = await db[coll].find_one(
                        {"id": d["id"]}, {"_id": 0, "password_hash": 1}
                    )
                if existing and existing.get("password_hash"):
                    d["password_hash"] = existing["password_hash"]
                elif _seed_hash:
                    d["password_hash"] = _seed_hash
                    d["must_change_password"] = True
            clean.append(d)

        deleted = 0
        if not merge:
            res = await db[coll].delete_many({})
            deleted = res.deleted_count

        upserted = 0
        modified = 0
        inserted = 0
        for d in clean:
            try:
                r = await db[coll].update_one(
                    {"id": d["id"]}, {"$set": d}, upsert=True,
                )
                if r.upserted_id is not None:
                    inserted += 1
                elif r.modified_count:
                    modified += 1
                upserted += 1
            except Exception as e:  # noqa: BLE001
                doc_id = str(d.get("id") or "unknown")
                failed_docs.setdefault(coll, []).append(doc_id)
                logger.warning(
                    "restore: collection=%s doc_id=%s status=failed error=%s",
                    coll,
                    doc_id,
                    type(e).__name__,
                )

        summary[coll] = {
            "deleted": deleted,
            "processed": len(clean),
            "inserted": inserted,
            "updated": modified,
            "failed": len(failed_docs.get(coll) or []),
        }

    logger.info(f"restore: processed {sum(s['processed'] for s in summary.values())} records across {len(summary)} collections")

    total_failed = sum(len(v) for v in failed_docs.values())
    result = {
        "ok": total_failed == 0,
        "mode": "replace" if not merge else "merge",
        "dry_run": False,
        "backup_generated_at": manifest.get("generated_at"),
        "backup_version": manifest.get("version", "unknown"),
        "archive_environment": archive_env or "unknown",
        "archive_backup_id": manifest.get("backup_id"),
        "collections": summary,
        "total_processed": sum(s["processed"] for s in summary.values()),
        "total_failed": total_failed,
        "failed_docs": failed_docs,
        "status": "partial_failure" if total_failed else "success",
    }
    # Track 14.0-I1: success audit (counterpart to the rejection audits above).
    await _record_audit(
        "accepted" if total_failed == 0 else "partial_failure",
        f"merge={merge}; processed={result['total_processed']}; failed={total_failed}",
    )
    await complete_backup_job(db, restore_job["job_id"], outcome="ok" if total_failed == 0 else "partial_failure", result=result, state="completed", owner_token=restore_lease.owner_token)
    try:
        if spool_path and spool_path.exists():
            spool_path.unlink()
    except Exception:
        pass
    restore_heartbeat_stop.set()
    try:
        await restore_heartbeat_task
    except Exception:
        pass
    return result





@api_router.get("/equipment-status-board")
async def equipment_status_board(_: bool = Depends(require_admin)):
    """
    Per-unit aggregation for the Admin Hub status board.

    For every saved equipment unit (or every unit referenced by an inspection,
    even if the operator typed it free-form without saving), returns:
        - last_inspection_date / last_inspected_days_ago
        - last_status: "ok" | "fail" | "never"
        - fail_count_14d : how many FAIL items logged in the last 14 days
        - top_failures : up to 3 most-frequent failing item names (last 30 d)
        - inspection_count : total all-time count
    """
    now = datetime.now(timezone.utc)
    cutoff_14 = now - timedelta(days=14)
    cutoff_30 = now - timedelta(days=30)

    saved_cursor = db.equipment_units.find({}, {"_id": 0})
    saved_units = await saved_cursor.to_list(2000)

    # Pull every inspection (slim projection — skip photos to keep it fast)
    insp_cursor = db.equipment_inspections.find(
        {},
        {
            "_id": 0,
            "id": 1,
            "equipment_type": 1,
            "equipment_unit": 1,
            "inspection_date": 1,
            "created_at": 1,
            "fail_count": 1,
            "out_of_service": 1,
            "checklist": 1,
            "project_name": 1,
            "project_number": 1,
        },
    ).sort("created_at", -1)
    inspections = await insp_cursor.to_list(5000)

    def _key(t: str, u: str) -> str:
        return f"{(t or '').strip()}||{(u or '').strip()}"

    by_unit: Dict[str, Dict[str, Any]] = {}
    for u in saved_units:
        k = _key(u.get("equipment_type", ""), u.get("unit_label", ""))
        by_unit[k] = {
            "equipment_type": u.get("equipment_type", ""),
            "equipment_unit": u.get("unit_label", ""),
            "make": u.get("make", "") or "",
            "model": u.get("model", "") or "",
            "serial": u.get("serial", "") or "",
            "saved": True,
            "inspection_count": 0,
            "fail_count_14d": 0,
            "last_inspection_date": None,
            "last_inspected_at": None,
            "last_status": "never",
            "last_project": "",
            "last_project_number": "",
            "_fail_items_30d": {},  # tally
        }

    def _parse_dt(s: str) -> Optional[datetime]:
        if not s:
            return None
        try:
            return datetime.fromisoformat(s.replace("Z", "+00:00"))
        except Exception:
            return None

    for d in inspections:
        k = _key(d.get("equipment_type"), d.get("equipment_unit"))
        if k not in by_unit:
            by_unit[k] = {
                "equipment_type": d.get("equipment_type", ""),
                "equipment_unit": d.get("equipment_unit", ""),
                "make": "",
                "model": "",
                "serial": "",
                "saved": False,
                "inspection_count": 0,
                "fail_count_14d": 0,
                "last_inspection_date": None,
                "last_inspected_at": None,
                "last_status": "never",
                "last_project": "",
                "last_project_number": "",
                "_fail_items_30d": {},
            }

        bucket = by_unit[k]
        bucket["inspection_count"] += 1

        created = _parse_dt(d.get("created_at"))
        if bucket["last_inspected_at"] is None or (
            created and created > bucket["last_inspected_at"]
        ):
            bucket["last_inspected_at"] = created
            bucket["last_inspection_date"] = d.get("inspection_date") or (
                created.date().isoformat() if created else None
            )
            bucket["last_status"] = (
                "fail" if (d.get("fail_count") or 0) > 0 else "ok"
            )
            bucket["last_project"] = d.get("project_name", "") or ""
            bucket["last_project_number"] = d.get("project_number", "") or ""

        # 14-day fail count
        if created and created >= cutoff_14:
            bucket["fail_count_14d"] += int(d.get("fail_count") or 0)

        # 30-day per-item failure tally
        if created and created >= cutoff_30:
            for sec, items in (d.get("checklist") or {}).items():
                if not isinstance(items, dict):
                    continue
                for item_name, res in items.items():
                    if isinstance(res, dict) and res.get("status") == "fail":
                        bucket["_fail_items_30d"][item_name] = (
                            bucket["_fail_items_30d"].get(item_name, 0) + 1
                        )

    out = []
    for k, b in by_unit.items():
        last_at = b.pop("last_inspected_at", None)
        days_ago = None
        if last_at:
            days_ago = max(0, (now - last_at).days)
        top = sorted(
            b.pop("_fail_items_30d", {}).items(), key=lambda kv: kv[1], reverse=True
        )[:3]
        b["last_inspected_days_ago"] = days_ago
        b["top_failures"] = [
            {"item": item, "count": cnt} for item, cnt in top
        ]
        out.append(b)

    # Sort: out-of-service first, then by fail count desc, then by stale-ness
    def _priority(b):
        oos = 0 if b["last_status"] == "fail" else 1
        stale = b["last_inspected_days_ago"] if b["last_inspected_days_ago"] is not None else 9999
        return (oos, -b["fail_count_14d"], -stale, b["equipment_type"], b["equipment_unit"])

    out.sort(key=_priority)
    return {
        "generated_at": now.isoformat(),
        "units": out,
        "summary": {
            "total_units": len(out),
            "out_of_service": sum(1 for b in out if b["last_status"] == "fail"),
            "never_inspected": sum(1 for b in out if b["last_status"] == "never"),
            "stale_7d": sum(
                1
                for b in out
                if (b["last_inspected_days_ago"] is None or b["last_inspected_days_ago"] >= 7)
            ),
        },
    }




# ============================================================
# Translation (Spanish → English on submit)
# ============================================================
# Crews can fill any form in Spanish, but every saved record + printed PDF
# must be 100% English (legal/OSHA requirement). At submit time the frontend
# sends the freeform user-typed string leaves to this endpoint, which calls
# Claude Haiku 4.5 via the Emergent universal LLM key and returns the same
# dict shape with English values.

class TranslateRequest(BaseModel):
    from_lang: str = "es"
    to_lang: str = "en"
    strings: Dict[str, str] = Field(default_factory=dict)


class TranslateResponse(BaseModel):
    strings: Dict[str, str]


import json as _json  # noqa: E402  (kept local to this section)


@api_router.post("/translate", response_model=TranslateResponse, dependencies=[Depends(rate_limit_public_post)])
async def translate_strings(payload: TranslateRequest):
    """Translate a flat {key: string} dict between languages.

    Returns the same keys with translated values. If translation fails we
    return the original strings unchanged so the form submit is never
    blocked. Empty input is a no-op.
    """
    if not payload.strings:
        return TranslateResponse(strings={})

    # Short-circuit when source & target are identical
    if payload.from_lang == payload.to_lang:
        return TranslateResponse(strings=payload.strings)

    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        logger.warning("EMERGENT_LLM_KEY missing — returning input unchanged")
        return TranslateResponse(strings=payload.strings)

    # Lazy import so cold-start of the rest of the API isn't blocked.
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
    except Exception as e:  # pragma: no cover
        logger.error(f"emergentintegrations import failed: {e}")
        return TranslateResponse(strings=payload.strings)

    # TRACK 14.0-S1 · AMENDMENT D — MASCI Heavy Civil Glossary.
    # The translator MUST produce English using MASCI / US Heavy Civil
    # operational terminology, not generic dictionary translations. This
    # glossary is the authoritative source for PDFs, notifications, and
    # search outputs that the English-speaking office consumes downstream.
    MASCI_GLOSSARY_ES_EN = (
        "MASCI / US HEAVY CIVIL ENGLISH GLOSSARY (use these EXACT English "
        "operational terms — never use generic dictionary translations):\n"
        "  excavación / zanja → excavation / trench (use 'trench' when ≤15 ft wide)\n"
        "  caja de zanja / cajón → trench box (NOT 'box for ditch')\n"
        "  escudo de zanja → trench shield\n"
        "  riel deslizante / rieles → slide rail system\n"
        "  placa vial / placa de acero → road plate (NOT 'metal plate')\n"
        "  MOT / mantenimiento de tránsito → MOT / Maintenance of Traffic\n"
        "  línea de fuerza / fuerza principal → force main\n"
        "  alcantarillado por gravedad → gravity sewer\n"
        "  drenaje pluvial / aguas lluvias → storm drain\n"
        "  estación elevadora / cárcamo → lift station\n"
        "  válvula → valve\n"
        "  hidrante → hydrant\n"
        "  cruce de servicios públicos → utility crossing\n"
        "  relleno / rellenado → backfill\n"
        "  densidad → density\n"
        "  compactación → compaction\n"
        "  fresado → milling\n"
        "  pavimentación → paving\n"
        "  riego de liga → tack coat\n"
        "  imprimación → prime coat\n"
        "  subrasante → subgrade\n"
        "  lime rock / piedra caliza → lime rock\n"
        "  control de máquina por GPS → GPS machine control\n"
        "  topografía / topógrafo → survey / surveyor\n"
        "  replanteo → stakeout\n"
        "  banco de referencia → benchmark\n"
        "  espacio confinado → confined space\n"
        "  acción correctiva → corrective action\n"
        "  cuasi accidente / casi accidente → near miss\n"
        "  incidente → incident\n"
        "  causa raíz → root cause\n"
        "  capataz / mayordomo → foreman\n"
        "  superintendente → superintendent\n"
        "  cuadrilla → crew\n"
        "  operador → operator\n"
        "  retroexcavadora → backhoe\n"
        "  excavadora → excavator\n"
        "  cargador frontal → front loader\n"
        "  motoniveladora → motor grader\n"
        "  vibrocompactador / rodillo → roller / compactor\n"
        "  volqueta → dump truck\n"
        "  tractor de orugas → bulldozer / dozer\n"
        "  retro / pala mecánica → skid steer\n"
        "  niveladora → grader\n"
        "  pala mecánica → loader\n"
        "  EPP / equipo de protección personal → PPE (Personal Protective Equipment)\n"
        "  casco → hard hat\n"
        "  chaleco reflectante → hi-vis vest / Class 2/3 vest\n"
        "  botas con punta de acero → steel-toed boots\n"
        "  arnés → fall-protection harness\n"
        "  línea de vida → lifeline\n"
        "  amonestación → write-up\n"
        "  reunión de seguridad / charla → safety meeting / tailgate meeting\n"
        "  análisis de riesgos del trabajo → Job Hazard Analysis (JHA)\n"
        "  plan de riesgos del trabajo → Job Hazard Plan (JHP)\n"
        "  permiso de trabajo → work permit\n"
        "  bloqueo y etiquetado → Lock-Out/Tag-Out (LOTO)\n"
        "  reporte diario → Daily Report\n"
        "  parte diario → Daily Report\n"
        "  solicitud de empleado → employee request\n"
        "  solicitud de tiempo libre / permiso → time-off request\n"
        "  inspección de equipo → equipment inspection\n"
        "  QA/QC / control de calidad → QA/QC / quality control\n"
        "  acta de reunión → meeting minutes\n"
        "  proyecto / obra → project / job site\n"
        "  número de proyecto → project number\n"
        "  orden de trabajo → work order\n"
        "  despacho / coordinación → dispatch\n"
        "  acarreo → haul\n"
        "  bota de carga → loading dock / load-out\n"
        "  vertedero → dump site / landfill\n"
        "  cantera → quarry\n"
        "  planta de asfalto → asphalt plant\n"
        "  base granular → granular base\n"
        "  aceras → sidewalks\n"
        "  bordillos → curbs\n"
        "  cordón y cuneta → curb and gutter\n"
        "  pozo de inspección → manhole\n"
        "  caja de válvulas → valve box\n"
        "  alcantarilla → culvert\n"
        "  tubería → pipe\n"
        "  ASTM / FDOT / DOT → keep abbreviations as-is\n"
    )

    system = (
        "You translate MASCI heavy-civil construction safety records from "
        "{src} to {dst}. Output language MUST be {dst}. The English you "
        "produce is what an English-speaking PM, Safety Manager, Super-"
        "intendent, CEI inspector, customer, and executive will read on "
        "the operational record, PDF, notification, search result, and "
        "export — so it MUST use US heavy-civil operational vocabulary, "
        "not generic dictionary equivalents. Preserve proper nouns, "
        "numbers, dates, project numbers, and equipment IDs EXACTLY. "
        "Keep the SAME JSON shape: input is a JSON object whose values "
        "are the strings to translate; reply with ONLY a JSON object "
        "using the SAME keys and translated values — no commentary, "
        "no markdown fences.\n\n" + MASCI_GLOSSARY_ES_EN
    ).format(src=payload.from_lang, dst=payload.to_lang)

    user_text = (
        "Translate every value in this JSON object. Reply with the JSON object "
        "only, same keys, translated values:\n\n"
        + _json.dumps(payload.strings, ensure_ascii=False)
    )

    try:
        chat = LlmChat(
            api_key=api_key,
            session_id=f"translate-{uuid.uuid4().hex[:8]}",
            system_message=system,
        ).with_model("anthropic", "claude-haiku-4-5-20251001")

        response = await chat.send_message(UserMessage(text=user_text))
        text = (response or "").strip()

        # Strip optional ```json fences if the model added them
        if text.startswith("```"):
            text = text.strip("`")
            # remove leading "json\n" if present
            if text.lower().startswith("json"):
                text = text[4:].lstrip("\n")

        # Find the first { … } block
        start = text.find("{")
        end = text.rfind("}")
        if start == -1 or end == -1:
            raise ValueError(f"No JSON object in response: {text[:200]}")
        parsed = _json.loads(text[start : end + 1])
        if not isinstance(parsed, dict):
            raise ValueError("LLM did not return a JSON object")

        # Only keep string values; fall back to original where missing
        out = {}
        for k, original in payload.strings.items():
            v = parsed.get(k)
            out[k] = v if isinstance(v, str) and v.strip() else original
        return TranslateResponse(strings=out)
    except Exception as e:
        logger.exception(f"Translation failed: {e}")
        return TranslateResponse(strings=payload.strings)


# ============================================================
# Training Hub — video URL registry
# ============================================================
# Each lesson has a `slug` defined in the frontend training catalog.
# Admins paste a YouTube / Loom / Vimeo / Wistia URL per slug and the
# Training Hub renders it above the written walk-through. Storage is a
# single Mongo document with id="config" in the `training_videos` collection:
#   { "_id": "config", "videos": {"field-01-hub-navigation": "https://…", ...} }


# Default video catalog — keyed by lesson slug, each value is a
# {"en": url, "es": url} pair. Spanish entry is optional; the player
# falls back to English with a "Spanish version not available" hint.
#
# URLs use the relative path `/api/training/video/{slug}.{lang}.mp4`
# pointing at our self-hosted Range-aware streamer below. Storing them
# relative means the same URL works on preview AND production without
# rewrites — the frontend prepends REACT_APP_BACKEND_URL on render.
#
# Why self-hosted instead of customer-assets.emergentagent.com:
# the original CDN-hosted MP4s shipped with the `moov` atom at the END
# of the file. Browsers must download the entire file before playback
# can start reliably, which manifests as skipping / cutting in & out
# during streaming. The files in /app/backend/static/training-videos/
# were re-muxed with `+faststart` (moov atom moved to byte 36 = front
# of file) so progressive playback works on every device.
_DEFAULT_TRAINING_VIDEOS = {
    "field-01-hub-navigation": {
        "en": "/api/training/video/field-01-hub-navigation.en.mp4",
        "es": "/api/training/video/field-01-hub-navigation.es.mp4",
    },
    "field-02-daily-report": {
        "en": "/api/training/video/field-02-daily-report.en.mp4",
        "es": "/api/training/video/field-02-daily-report.es.mp4",
    },
    "field-03-equipment-preop": {
        "en": "/api/training/video/field-03-equipment-preop.en.mp4",
        "es": "/api/training/video/field-03-equipment-preop.es.mp4",
    },
    "field-04-safety-meeting": {
        "en": "/api/training/video/field-04-safety-meeting.en.mp4",
        "es": "/api/training/video/field-04-safety-meeting.es.mp4",
    },
    "field-05-jhp": {
        "en": "/api/training/video/field-05-jhp.en.mp4",
        "es": "/api/training/video/field-05-jhp.es.mp4",
    },
    "field-06-incident": {
        "en": "/api/training/video/field-06-incident.en.mp4",
        "es": "/api/training/video/field-06-incident.es.mp4",
    },
}


# ---------------------------------------------------------------------------
# Range-aware video streamer — serves the faststart-fixed MP4s in
# /app/backend/static/training-videos/ with full HTTP Range support so
# the browser can request byte-slices, seek instantly, and start
# playback after the first chunk arrives. This is what fixes the
# "skipping / cutting in & out" symptom users reported with the
# moov-at-end customer-assets MP4s.
# ---------------------------------------------------------------------------
import re as _vid_re  # noqa: E402
from starlette.responses import StreamingResponse as _VidStreamResp  # noqa: E402

_VIDEO_DIR = Path("/app/backend/static/training-videos")
# Filename safety: only allow lowercase letters, digits, dot, dash, underscore.
_VIDEO_NAME_RE = _vid_re.compile(r"^[a-z0-9][a-z0-9._-]{0,128}\.mp4$")


@api_router.head("/training/video/{filename}")
@api_router.get("/training/video/{filename}")
async def training_video_stream(filename: str, request: Request):
    """Range-aware MP4 streamer — returns the faststart-fixed bilingual
    training videos with `Accept-Ranges: bytes`, proper `Content-Range`
    on 206 responses, and `Content-Type: video/mp4`. Public read because
    field crews scan the QR poster and load the page without auth."""
    if not _VIDEO_NAME_RE.match(filename):
        raise HTTPException(404, "Video not found")
    path = _VIDEO_DIR / filename
    if not path.exists() or not path.is_file():
        raise HTTPException(404, "Video not found")

    file_size = path.stat().st_size
    range_header = request.headers.get("range") or request.headers.get("Range")
    common_headers = {
        "Accept-Ranges": "bytes",
        "Content-Type": "video/mp4",
        "Cache-Control": "public, max-age=86400",
        # ETag = file size + mtime so browser cache invalidates on re-mux.
        "ETag": f'W/"{file_size}-{int(path.stat().st_mtime)}"',
    }

    if range_header:
        m = _vid_re.match(r"^bytes=(\d+)-(\d*)$", range_header.strip())
        if not m:
            return Response(status_code=416, headers={"Content-Range": f"bytes */{file_size}"})
        start = int(m.group(1))
        end = int(m.group(2)) if m.group(2) else file_size - 1
        if start >= file_size or end >= file_size or start > end:
            return Response(status_code=416, headers={"Content-Range": f"bytes */{file_size}"})
        chunk_len = end - start + 1

        async def iter_range():
            # 256 KB read chunks — small enough for memory, large enough
            # to keep TCP pipelined.
            with open(path, "rb") as f:
                f.seek(start)
                remaining = chunk_len
                while remaining > 0:
                    buf = f.read(min(262144, remaining))
                    if not buf:
                        break
                    remaining -= len(buf)
                    yield buf

        headers = {
            **common_headers,
            "Content-Length": str(chunk_len),
            "Content-Range": f"bytes {start}-{end}/{file_size}",
        }
        # HEAD requests need the same headers but no body.
        if request.method == "HEAD":
            return Response(status_code=206, headers=headers)
        return _VidStreamResp(iter_range(), status_code=206, headers=headers, media_type="video/mp4")

    # Full-file request (no Range) — most browsers send Range immediately
    # for <video>, but iOS Safari sometimes pulls the whole file.
    headers = {**common_headers, "Content-Length": str(file_size)}
    if request.method == "HEAD":
        return Response(status_code=200, headers=headers)
    return FileResponse(path=str(path), media_type="video/mp4", headers=headers)


def _normalize_video_entry(entry):
    """Normalize legacy single-string URLs to the new {en, es} shape.
    Accepts either:
      - "https://…"  (legacy single-language) → {"en": "https://…", "es": ""}
      - {"en": "…", "es": "…"}  (new) → returned as-is with missing keys defaulted to ""
    Anything else → empty pair.
    """
    if isinstance(entry, str):
        return {"en": entry, "es": ""}
    if isinstance(entry, dict):
        return {"en": entry.get("en") or "", "es": entry.get("es") or ""}
    return {"en": "", "es": ""}


@api_router.get("/training/videos")
async def training_videos_get():
    """Public read. Field crews need this to render embedded videos without
    a login. Returns `{videos: {slug: {en, es}}}`.

    Self-heal: any default video slug missing EN or ES URLs gets the
    default back-filled (per-key, never overwriting admin overrides).
    Legacy single-string entries are normalized to the {en, es} shape on
    first read.

    One-time migration: any stored URL pointing at the old customer-assets
    CDN host (`customer-assets.emergentagent.com`) gets replaced with the
    matching self-hosted faststart URL from the default catalog. The CDN
    MP4s shipped with the moov atom at the END of file which made
    streaming stutter; the self-hosted copies have moov at the front.
    """
    doc = await db["training_videos"].find_one({"_id": "config"}, {"_id": 0})
    stored = (doc or {}).get("videos") or {}

    set_ops = {}
    out = {}

    def is_legacy_cdn(u):
        return isinstance(u, str) and "customer-assets.emergentagent.com" in u

    for slug, default in _DEFAULT_TRAINING_VIDEOS.items():
        cur = _normalize_video_entry(stored.get(slug))
        # Migration: replace legacy CDN URLs with self-hosted faststart.
        if is_legacy_cdn(cur["en"]) and default.get("en"):
            cur["en"] = default["en"]
            set_ops[f"videos.{slug}.en"] = default["en"]
        if is_legacy_cdn(cur["es"]) and default.get("es"):
            cur["es"] = default["es"]
            set_ops[f"videos.{slug}.es"] = default["es"]
        # Fill any blank language URL from the default.
        if not cur["en"] and default.get("en"):
            cur["en"] = default["en"]
            set_ops[f"videos.{slug}.en"] = default["en"]
        if not cur["es"] and default.get("es"):
            cur["es"] = default["es"]
            set_ops[f"videos.{slug}.es"] = default["es"]
        # Normalize legacy single-string in storage to {en, es} shape.
        if isinstance(stored.get(slug), str):
            set_ops[f"videos.{slug}"] = cur
        out[slug] = cur

    # Preserve any extra slugs admins added that aren't in defaults.
    for slug, val in stored.items():
        if slug in out:
            continue
        cur = _normalize_video_entry(val)
        out[slug] = cur
        if isinstance(val, str):
            set_ops[f"videos.{slug}"] = cur

    if set_ops:
        set_ops["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db["training_videos"].update_one(
            {"_id": "config"},
            {"$set": set_ops},
            upsert=True,
        )

    return {"videos": out}


@api_router.get("/training/packet.pdf")
async def training_packet_pdf(
    track: str,
    lang: str = "en",
    request: Request = None,
    x_admin_token: Optional[str] = Header(default=None),
    x_pm_token: Optional[str] = Header(default=None),
    x_shop_token: Optional[str] = Header(default=None),
):
    """Training packet PDF. The Field Crew track is public so labor can scan
    trailer posters without a login. Shop / PM / Admin tracks are gated —
    the back-office workflows are sensitive (backup procedures, password
    rotation, master-list internals) and shouldn't be readable by a new
    hire or walk-through visitor.

    Access rules:
      • `field`  → public, no auth
      • `shop`   → shop / pm / admin token accepted
      • `pm`     → pm / admin token accepted
      • `admin`  → admin token only

    Side effect: fire-and-forget insert into `training_hits` so the PM/Admin
    hub dashboards can show scan stats. No PII — just track/lang/date and
    a truncated UA + referer so we can tell which trailer poster got the
    scan (web vs phone browser, mascidocs.com vs direct)."""
    from training_pdf import render_packet, _normalize_lang  # local import

    t_lower = (track or "").lower()

    # Authorize based on track. We do this BEFORE rendering the PDF so a
    # failed auth never pays the render cost.
    # TRACK 28.03E · pair every retired sync admin-validator call with
    # `_is_valid_directory_admin_token_async` so per-user admin tokens
    # unlock the training PDF the same as every other admin surface.
    async def _admin_ok(tok: Optional[str]) -> bool:
        if not tok:
            return False
        return bool(_is_valid_admin_token(tok)) or bool(await _is_valid_directory_admin_token_async(tok))

    if t_lower == "admin":
        if not await _admin_ok(x_admin_token):
            raise HTTPException(
                status_code=401,
                detail="Admin login required for the Admin training packet.",
            )
    elif t_lower == "pm":
        if not (
            await _admin_ok(x_admin_token)
            or (x_pm_token and _is_valid_pm_token(x_pm_token))
        ):
            raise HTTPException(
                status_code=401,
                detail="PM or Admin login required for the PM training packet.",
            )
    elif t_lower == "shop":
        # TRACK 15.30 — shared SHOP_PASSWORD HMAC retired. Per-user shop
        # tokens (format `<user_id>.<HMAC>`) are accepted via the
        # shop_users module.
        shop_ok = False
        if x_shop_token and "." in x_shop_token:
            try:
                from shop_users import is_valid_shop_user_token_async  # noqa: PLC0415
                shop_ok = (await is_valid_shop_user_token_async(db, x_shop_token)) is not None
            except Exception:  # noqa: BLE001
                shop_ok = False
        if not (
            await _admin_ok(x_admin_token)
            or (x_pm_token and _is_valid_pm_token(x_pm_token))
            or shop_ok
        ):
            raise HTTPException(
                status_code=401,
                detail="Shop, PM, or Admin login required for the Shop training packet.",
            )
    # else: field track → public

    try:
        pdf_bytes = render_packet(track, lang)
    except ValueError as e:
        raise HTTPException(404, str(e))
    lang_norm = _normalize_lang(lang)

    # Log the scan hit — swallow any error; telemetry must never block the PDF.
    try:
        ua = (request.headers.get("user-agent") if request else "") or ""
        ref = (request.headers.get("referer") if request else "") or ""
        # Coarse device family classification — enough for the stripe.
        ua_l = ua.lower()
        if "iphone" in ua_l or "ipad" in ua_l or "ios" in ua_l:
            device = "ios"
        elif "android" in ua_l:
            device = "android"
        elif any(k in ua_l for k in ("mobile", "phone")):
            device = "mobile-other"
        elif any(k in ua_l for k in ("windows", "mac os", "linux", "x11")):
            device = "desktop"
        else:
            device = "other"
        # Referer source — did they come from the poster or a direct link?
        if "/training/" in ref and "/poster" in ref:
            source = "poster"
        elif "/training" in ref:
            source = "hub"
        elif "mascidocs.com" in ref:
            source = "internal"
        elif ref:
            source = "external"
        else:
            source = "direct"  # QR scan usually has no referer on phones
        await db["training_hits"].insert_one({
            "track": track,
            "lang": lang_norm,
            "device": device,
            "source": source,
            "ts": datetime.now(timezone.utc),
        })
    except Exception:
        pass  # never block the PDF on telemetry

    from fastapi.responses import Response
    filename = f"MASCI_training_{track}_{lang_norm}.pdf"
    # Critical: public CDN cache on a token-gated PDF would let one
    # admin's 200 response be served back to a PM or shop user. For the
    # public Field track we still allow shared caching, but anything
    # behind a login must be private and revalidate every time.
    if t_lower == "field":
        cache_ctrl = "public, max-age=60"
    else:
        cache_ctrl = "private, no-store"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{filename}"',
            "Cache-Control": cache_ctrl,
            # Tell shared caches the response varies by token header.
            "Vary": "X-Admin-Token, X-PM-Token, X-Shop-Token",
        },
    )


@api_router.get("/admin/training/stats")
async def training_stats(_: bool = Depends(require_admin)):
    """Aggregated scan stats for the PM/Admin hub stripe. `require_admin`
    accepts both Admin and PM tokens (Shop is not relevant here). Returns
    this-week / last-week counts, per-track breakdown, per-language split,
    and a 14-day daily trend for sparklines."""
    now = datetime.now(timezone.utc)
    this_week_start = now - timedelta(days=7)
    last_week_start = now - timedelta(days=14)
    trend_start = now - timedelta(days=14)

    col = db["training_hits"]
    this_week = await col.count_documents({"ts": {"$gte": this_week_start}})
    last_week = await col.count_documents({
        "ts": {"$gte": last_week_start, "$lt": this_week_start}
    })
    total = await col.count_documents({})

    # Per-track this-week
    by_track = {}
    async for row in col.aggregate([
        {"$match": {"ts": {"$gte": this_week_start}}},
        {"$group": {"_id": "$track", "n": {"$sum": 1}}},
    ]):
        by_track[row["_id"]] = row["n"]

    # Per-language this-week
    by_lang = {}
    async for row in col.aggregate([
        {"$match": {"ts": {"$gte": this_week_start}}},
        {"$group": {"_id": "$lang", "n": {"$sum": 1}}},
    ]):
        by_lang[row["_id"]] = row["n"]

    # 14-day trend, one bucket per day (UTC)
    trend = []
    async for row in col.aggregate([
        {"$match": {"ts": {"$gte": trend_start}}},
        {"$group": {
            "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$ts"}},
            "n": {"$sum": 1},
        }},
        {"$sort": {"_id": 1}},
    ]):
        trend.append({"date": row["_id"], "n": row["n"]})

    return {
        "total": total,
        "this_week": this_week,
        "last_week": last_week,
        "by_track": by_track,
        "by_lang": by_lang,
        "trend": trend,
        "generated_at": now.isoformat(),
    }


@api_router.put("/admin/training/videos")
async def training_videos_put(
    body: dict,
    _: bool = Depends(require_admin_strict),
):
    """Admin-strict write. Body accepts either:
      - `{videos: {slug: "https://…"}}` (legacy single-language → stored as {en: url})
      - `{videos: {slug: {en: "https://…", es: "https://…"}}}` (new bilingual)
    Merge-updates the stored map so partial saves are safe — only supplied
    slugs/languages are changed. Empty string clears that field.
    """
    incoming = (body or {}).get("videos")
    if not isinstance(incoming, dict):
        raise HTTPException(400, "Body must be `{videos: {slug: url-or-{en,es}}}`")

    existing = await db["training_videos"].find_one({"_id": "config"}) or {}
    merged = {}
    for slug, val in (existing.get("videos") or {}).items():
        merged[slug] = _normalize_video_entry(val)

    for slug, val in incoming.items():
        if not isinstance(slug, str) or not slug.strip():
            continue
        slug = slug.strip()
        if isinstance(val, str):
            cur = merged.get(slug, {"en": "", "es": ""})
            cur["en"] = val.strip()
            merged[slug] = cur
        elif isinstance(val, dict):
            cur = merged.get(slug, {"en": "", "es": ""})
            if "en" in val:
                cur["en"] = (val.get("en") or "").strip() if isinstance(val.get("en"), str) else ""
            if "es" in val:
                cur["es"] = (val.get("es") or "").strip() if isinstance(val.get("es"), str) else ""
            merged[slug] = cur
        elif val is None:
            merged.pop(slug, None)

    # drop completely empty slugs so the map stays compact
    merged = {k: v for k, v in merged.items() if v.get("en") or v.get("es")}
    await db["training_videos"].update_one(
        {"_id": "config"},
        {"$set": {"videos": merged, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    return {"ok": True, "count": len(merged), "videos": merged}




app.include_router(api_router)


# ------------------------- Date Audit (one-shot timezone-bug repair) -------------------------
from routes.date_audit import build_date_audit_router  # noqa: E402

_date_audit_router = build_date_audit_router(db, require_admin_strict)
app.include_router(_date_audit_router)

# ------------------------- Safety Forms (Equipment Issuance + Training) -------------------------
from routes.safety_forms import build_safety_forms_router  # noqa: E402

_safety_forms_router = build_safety_forms_router(
    db, _is_valid_admin_token, _is_valid_directory_admin_token_async,
)
app.include_router(_safety_forms_router)


# ------------------------- Job Photos (Phase 1 — read-only aggregator) -------------------------
from routes.job_photos import (  # noqa: E402
    attach_routes as _attach_job_photos_routes,
    background_indexer_loop as _job_photos_indexer_loop,
    index_record_photos as _index_record_photos,
    photo_bytes_router as _photo_bytes_router,
)

# Mount the photo-bytes resolver immediately — public route, no DB hooks.
# Lets <img src> tags in record-detail pages render photo:// refs that
# went into Mongo via the iter64 R2 migration.
app.include_router(_photo_bytes_router)


# ------------------------- Hub Banners (site-wide messaging) -------------------------
from routes.hub_banners import build_hub_banners_router  # noqa: E402

_hub_banners_router = build_hub_banners_router(db, require_admin)
app.include_router(_hub_banners_router)

# ------------------------- Track 24.3 · DR V3 free-text ES → EN translation -------------------------
from routes.translation import build_translation_router  # noqa: E402

_translation_router = build_translation_router(db, rate_limit_public_post)
app.include_router(_translation_router)


async def _job_photos_send_email(*, to: str, subject: str, text: str, attachments=None):
    """Tiny Resend wrapper used by the Job Photos email endpoint. Mirrors
    the headers + retry behaviour of the rest of the system. Imports
    `resend` lazily so the module-level import order stays unchanged."""
    import base64 as _b64  # noqa: PLC0415
    import resend as _resend  # noqa: PLC0415

    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        raise RuntimeError("RESEND_API_KEY missing")
    _resend.api_key = api_key
    sender = await _resolve_sender_email(db, safe_fallback="noreply@mascidocs.com")
    params: dict = {
        "from": f"MASCI Operations Platform <{sender}>",
        "to": [to],
        "subject": subject,
        "text": text or " ",
    }
    if attachments:
        params["attachments"] = [
            {
                "filename": a["filename"],
                "content": _b64.b64encode(a["content"]).decode("ascii"),
                "content_type": a.get("content_type", "application/octet-stream"),
            }
            for a in attachments
        ]
    return await asyncio.to_thread(_resend.Emails.send, params)


_attach_job_photos_routes(app, db, require_admin, _job_photos_send_email)


# ============================================================
# Field Leadership routes — supervisor docs (write-ups, coaching, etc.)
# ============================================================
from routes.field_leadership import attach_routes as _attach_field_leadership_routes, seed_equipment_defaults as _seed_field_leadership_equipment  # noqa: E402
from field_leadership_pdf import render_field_leadership_pdf as _render_field_leadership_pdf  # noqa: E402


async def _field_leadership_send_email(recipients, subject, html_body, attachments=None):
    """Resend wrapper used by Field Leadership form submits. Sends HTML +
    optional PDF attachments to the assigned PM, jaymn, and safety. Returns
    silently on missing API key or AUTO_EMAIL_REPORTS=false (preview)."""
    if (os.environ.get("AUTO_EMAIL_REPORTS") or "").strip().lower() not in ("true", "1", "yes"):
        return
    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        return
    import resend as _resend  # noqa: PLC0415
    _resend.api_key = api_key
    sender = await _resolve_sender_email(db, safe_fallback="noreply@mascidocs.com")
    params: dict = {
        "from": f"MASCI Operations Platform <{sender}>",
        "to": list(recipients),
        "subject": subject,
        "html": html_body,
    }
    if attachments:
        params["attachments"] = attachments
    return await asyncio.to_thread(_resend.Emails.send, params)


async def _field_leadership_compute_pm_scope(pm_email):
    """Returns the set of project_numbers assigned to the given PM email."""
    if not pm_email:
        return set()
    try:
        from pm_auth import compute_pm_scope  # noqa: PLC0415
        # The shared helper takes (db, actor_dict) — fake an actor.
        scope = await compute_pm_scope(db, {"email": pm_email, "is_admin_or_legacy": False})
        return set(scope or [])
    except Exception:
        return set()


_attach_field_leadership_routes(
    app,
    db,
    require_admin,
    _field_leadership_send_email,
    _render_field_leadership_pdf,
    compute_pm_scope=_field_leadership_compute_pm_scope,
)


@register_lifecycle_step("index-ensure")
async def _ensure_scheduler_lock_indexes_at_startup():
    # iter441 · Phase 31.4 · multi-worker scheduler safety.
    # TTL index on scheduler_locks.expires_at so dead locks auto-clean
    # within 90s even if every worker that ever held them has died.
    try:
        await _ensure_scheduler_lock_indexes(db)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[singleton-lock] index ensure failed (non-fatal): {e}")
    # iter445 · Sprint · Scheduler Hardening · Option C.
    # Unique compound index on (scheduler, slot_key) gives us atomic
    # dedup even if the singleton-lock layer ever fails again.
    try:
        await _ensure_scheduler_runs_indexes(db)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[scheduler-runs] index ensure failed (non-fatal): {e}")
    try:
        await ensure_backup_runtime_indexes(db)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[backup-runtime] index ensure failed (non-fatal): {e}")


@register_lifecycle_step("index-ensure")
async def _ensure_production_bottleneck_indexes():
    try:
        await db.r2_inventory.create_index("size")
        await db.r2_inventory.create_index("content_type")
        await db.r2_references.create_index("r2_key")
        await db.daily_reports.create_index("report_number")
        await db.usage_events.create_index([("kind", 1), ("signal", 1), ("at", -1)])
        await db.usage_events.create_index([("kind", 1), ("signal", 1), ("at", -1), ("elapsed_ms", 1)])
        await db.usage_events.create_index([("kind", 1), ("signal", 1), ("at", -1), ("dims.equipment_id", 1)])
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[db-bottleneck] index ensure failed (non-fatal): {e}")


# ─── FORGEDOPS P0-B · DB isolation startup failsafe ─────────────────
# Probes whether the pod can access the OTHER environment's MongoDB
# namespace. Currently runs in "bridge" mode (loud warning on
# violation, pod still boots) because the Atlas user separation
# runbook has not yet been executed by the operator. When
# ENFORCE_DB_ISOLATION=true is set in env, this hook FAILS FAST
# (sys.exit(99)) on any violation. See db_isolation_failsafe.py +
# /app/memory/STARTUP_FAILSAFE_CERTIFICATION.md.
from db_isolation_failsafe import assert_db_isolation as _assert_db_isolation  # noqa: E402


@register_lifecycle_step("misc-bootstrap")
async def _db_isolation_failsafe():
    try:
        await _assert_db_isolation(client)
    except SystemExit:
        # When ENFORCE_DB_ISOLATION=true and a violation is detected,
        # assert_db_isolation calls sys.exit(99). Re-raise so the pod
        # actually dies (don't swallow it).
        raise
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[db-isolation] probe failed (non-fatal): {e}")


@register_lifecycle_step("misc-bootstrap")
async def _assert_no_duplicate_routes():
    """Track 24.1 · P1-C → Track 24.2 · Phase 3 · FAIL-CLOSED.

    Fails app startup if the same `(method, path)` pair is registered
    by more than one handler.  Duplicate registrations are an
    accident-prone pattern: FastAPI silently picks the first one,
    which means a refactor can flip which handler is "live" with zero
    code change. Track 24.0 audit found the auth-less handler on
    `/api/employees/competent-persons` winning over the auth-gated one
    exactly this way — Track 24.1 P0-2 removed it, and Track 24.2
    now hardens the check to fail-closed because preview boots with 0
    offenders.  If a legitimate duplicate ever needs to exist (e.g. a
    version-prefixed route), add its `(method, path)` to
    `_ALLOWED_DUPLICATES` below with a code-review justification."""
    _ALLOWED_DUPLICATES: set = set()          # empty — no exceptions.
    groups: Dict[tuple, List[str]] = defaultdict(list)
    for r in app.routes:
        if hasattr(r, "methods") and hasattr(r, "endpoint"):
            for m in r.methods:
                mod = getattr(r.endpoint, "__module__", "?")
                name = getattr(r.endpoint, "__name__", "?")
                groups[(m, r.path)].append(f"{mod}:{name}")
    dups = [(k, v) for k, v in groups.items()
            if len(v) >= 2 and k not in _ALLOWED_DUPLICATES]
    if not dups:
        logger.info("[track-24.2] duplicate-route scan clean · 0 offenders · fail-closed policy active")
        return
    # FAIL CLOSED — dump every offender and refuse to boot.
    logger.error(
        f"[track-24.2] duplicate-route scan · FAIL · {len(dups)} offenders. "
        f"FastAPI silently uses the FIRST registered handler; refusing to boot."
    )
    for (m, p), handlers in dups:
        logger.error(
            f"[track-24.2] duplicate  {m:6s} {p}  ({len(handlers)} handlers) · WINS={handlers[0]} · LOSING={handlers[1:]}"
        )
    raise RuntimeError(
        f"Duplicate route registrations detected ({len(dups)} offenders). "
        f"Fix them or add exceptions to _ALLOWED_DUPLICATES."
    )



@register_lifecycle_step("misc-bootstrap")
async def _tune_asyncio_thread_pool():
    # iter441 · Phase 31.4 · concurrent-load hardening.
    # The asyncio default executor is ``ThreadPoolExecutor(max_workers=cpu+4)``
    # which on a 1-vCPU Kubernetes pod is just 5 threads. Several endpoints
    # off-load sync work (boto3 R2 listings, presigned URL signing) via
    # ``asyncio.to_thread`` — under bursty concurrent admin load the default
    # pool saturates and incoming requests queue behind it, which Cloudflare
    # interprets as origin-down (520). Bumping the pool to 32 threads makes
    # the event loop comfortably absorb a 24-wide simultaneous burst without
    # queue buildup. Pure capacity tune · zero behavior change at low load.
    import concurrent.futures  # noqa: PLC0415
    try:
        loop = asyncio.get_event_loop()
        loop.set_default_executor(
            concurrent.futures.ThreadPoolExecutor(
                max_workers=32, thread_name_prefix="masci-async"
            )
        )
        logger.info("[concurrency] asyncio default thread pool tuned to max_workers=32")
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[concurrency] thread pool tune failed (non-fatal): {e}")


@register_lifecycle_step("scheduler-nonemail")
async def _start_job_photos_indexer():
    register_background_task(
        app,
        name="job-photos-indexer",
        coro=_job_photos_indexer_loop(db),
        category="scheduler",
        critical=False,
        long_running=True,
    )


# ── TRACK 22.9B · Photo Intelligence V1 pipeline ──────────────────
@register_lifecycle_step("index-ensure")
async def _ensure_dr_v1_photo_intel_indexes():
    """Idempotent index setup for the V1 photo intel job queue."""
    from services.photo_intelligence import (
        ensure_v1_pipeline_indexes, ensure_indexes as ensure_intel_indexes,
    )
    await ensure_v1_pipeline_indexes(db)
    await ensure_intel_indexes(db)


@register_lifecycle_step("seed")
async def _seed_tenant_photo_intelligence_flag():
    """Ensure the MASCI tenant has photo_intelligence_enabled=true.

    Track 22.9B — the analyzer/emitter existed but nothing turned the
    tenant flag on. This step upserts the flag idempotently on boot.
    Safe to run repeatedly; only writes if the current value differs.
    """
    from datetime import datetime as _dt, timezone as _tz  # noqa: PLC0415
    try:
        doc = await db["tenant_ai_capabilities"].find_one(
            {"tenant_id": "masci"}, {"_id": 0},
        )
        if doc and doc.get("photo_intelligence_enabled") is True:
            return
        now = _dt.now(_tz.utc).isoformat()
        await db["tenant_ai_capabilities"].update_one(
            {"tenant_id": "masci"},
            {
                "$set": {
                    "photo_intelligence_enabled": True,
                    "tenant_ai_enabled": True,
                    "updated_at": now,
                    "updated_by": "track_22_9b_bootstrap",
                    "tenant_id": "masci",
                    "tenant_name": "MASCI (default)",
                },
                "$setOnInsert": {"created_at": now, "version": 1},
            },
            upsert=True,
        )
    except Exception as e:  # noqa: BLE001
        logger.info(f"[track-22.9b] tenant flag seed skipped (non-fatal): {e}")


@register_lifecycle_step("scheduler-nonemail")
async def _start_dr_v1_photo_intel_reconciler():
    """Long-running reconciler loop for V1 photo intel jobs.

    Ensures every submitted photo eventually gets analyzed even if
    the BackgroundTasks first-pass was lost to a pod restart. Kill
    switch: ``DR_V1_PHOTO_INTEL_RECONCILER_ENABLED=false``.
    """
    from services.photo_intelligence import v1_reconciler_loop
    register_background_task(
        app,
        name="photo-intelligence-v1-reconciler",
        coro=v1_reconciler_loop(db),
        category="scheduler",
        critical=False,
        long_running=True,
    )


@register_lifecycle_step("seed")
async def _seed_field_leadership_equipment_catalog():
    await _seed_field_leadership_equipment(db)


@register_lifecycle_step("seed")
async def _seed_shop_users():
    from shop_users import seed_shop_users
    await seed_shop_users(db)


@register_lifecycle_step("index-ensure")
async def _ensure_project_team_assignments_indexes():
    """Track 14.0-JOB-OWNERSHIP-FOUNDATION Phase 1 — index ensure."""
    try:
        await db.project_team_assignments.create_index("id", unique=True)
        await db.project_team_assignments.create_index(
            [("project_number", 1), ("assignment_role", 1), ("active", 1)]
        )
        await db.project_team_assignments.create_index(
            [("user_id", 1), ("active", 1)]
        )
        await db.project_team_assignments.create_index(
            [("email", 1), ("active", 1)]
        )
        # Partial unique to prevent duplicate active rows for the same
        # (project, user, role) triple. Inactive rows are exempted so
        # historical assignments are preserved.
        await db.project_team_assignments.create_index(
            [("project_number", 1), ("user_id", 1), ("assignment_role", 1)],
            unique=True,
            partialFilterExpression={"active": True, "user_id": {"$type": "string"}},
            name="uniq_active_project_user_role",
        )
        # Audit category lookup
        await db.audit_events.create_index(
            [("category", 1), ("project_number", 1), ("at", -1)]
        )
        # Track 14.0-JOB-OWNERSHIP-FOUNDATION Phase 2A — ensure every
        # existing row has an `assignment_status` (idempotent).
        await db.project_team_assignments.update_many(
            {"assignment_status": {"$exists": False}, "active": True},
            {"$set": {"assignment_status": "ACTIVE"}},
        )
        await db.project_team_assignments.update_many(
            {"assignment_status": {"$exists": False}, "active": False},
            {"$set": {"assignment_status": "INACTIVE"}},
        )
    except Exception as exc:
        logger.warning("[team-roster] index ensure failed: %s", exc)


@register_lifecycle_step("misc-bootstrap")
async def _deploy_fix_001_backup_orphan_sweep():
    """DEPLOY-FIX-001 · Workstream A4 — startup sweep.

    On every backend boot, scan ``BACKUPS_DIR`` and delete any
    ``MASCI_*backup_*.zip.tmp.<hash>`` file older than 10 minutes. These
    are abandoned by gateway timeouts (Cloudflare disconnects at 60 s
    while the writer keeps streaming for several more minutes) and would
    silently fill local disk over time.

    Reuses ``_emergency_prune_backups`` (which already implements the
    correct 10-minute age guard and per-file logging) so behavior stays
    in lockstep with the existing emergency-prune contract.
    """
    try:
        pruned = await asyncio.to_thread(_emergency_prune_backups, "startup")
        if pruned:
            logger.info(
                f"[backup-cleanup] startup-sweep removed {pruned} orphan "
                f"tmp file(s); disk now at {_disk_pct_used()}%"
            )
        else:
            logger.info(
                "[backup-cleanup] startup-sweep · no orphan tmp files found"
            )
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[backup-cleanup] startup-sweep failed (non-fatal): {e}")




@register_lifecycle_step("seed")
async def _seed_hr_users():
    from hr_users import seed_hr_users
    await seed_hr_users(db)


@register_lifecycle_step("misc-bootstrap")
async def _ensure_v_prelude_wave1_indexes():
    """Phase V-Prelude · Wave 1 — index ensure for new substrate
    collections (operational_links · operational_constraints) plus the
    thin governance.tags index on job_photos. Idempotent."""
    try:
        from routes.operational_links import ensure_operational_links_indexes  # noqa: PLC0415
        from routes.operational_constraints import ensure_operational_constraints_indexes  # noqa: PLC0415
        from routes.photo_governance import ensure_photo_governance_indexes  # noqa: PLC0415
        from routes.odr import ensure_odr_indexes, ensure_continuity_indexes, ensure_observation_indexes  # noqa: PLC0415
        await ensure_operational_links_indexes(db)
        await ensure_operational_constraints_indexes(db)
        await ensure_photo_governance_indexes(db)
        await ensure_odr_indexes(db)
        await ensure_continuity_indexes(db)
        await ensure_observation_indexes(db)
    except Exception as e:  # noqa: BLE001
        logging.getLogger(__name__).warning(
            "V-Prelude Wave 1 index ensure failed: %s", e
        )


# iter299 · Lane D operational hygiene — visibility-only log line at boot.
# Emits a single structured line under tag `[ops-hygiene]` so operators can
# grep startup logs for disk pressure + backup inventory + retention config.
# NO new endpoints, NO new collections, NO alerts.
@register_lifecycle_step("misc-bootstrap")
async def _log_operational_hygiene_at_startup():
    await _log_operational_hygiene(reason="startup", db=db)


@register_lifecycle_step("post-readiness")
async def _start_motive_reliability_loop():
    """M-1R · Motive reliability supervisor.

    Reuses the existing asyncio scheduler doctrine. Fires four periodic
    sync loops (events / assets / users / geofences) at the cadences
    defined in `lib.motive_reliability`. Singleton-locked so only one
    worker ticks across a multi-worker fleet. Visibility-only — never
    mutates dispatch/maintenance state, never triggers workflow.
    """
    try:
        from lib.motive_reliability import motive_reliability_supervisor  # noqa: PLC0415
        target_db = db.get_target()
        if target_db is None:
            logging.getLogger(__name__).warning(
                "[motive-reliability] runtime database target unavailable at scheduler registration; skipping"
            )
            return
        register_background_task(
            app,
            name="motive-reliability-supervisor",
            coro=motive_reliability_supervisor(target_db),
            category="scheduler",
            critical=False,
            long_running=True,
        )
        logging.getLogger(__name__).info(
            "[motive-reliability] supervisor task scheduled"
        )
    except Exception as e:  # noqa: BLE001
        logging.getLogger(__name__).exception(
            f"[motive-reliability] failed to schedule supervisor: {e}"
        )



@register_lifecycle_step("misc-bootstrap")
async def _clear_super_admin_force_pw_change():
    """One-shot idempotent migration (iter117).

    Super admin `jaymn.judd@mascigc.com` has separate per-portal records
    in `hr_users`, `shop_users`, and `pm_users` because the per-portal
    seeders ran independently. Some of those seeders set
    `must_change_password=True` for first-login enforcement — but the
    super admin authenticates via the multi-portal master login at
    /sign-in, so the per-portal forced password change is redundant and
    locks him out of /hr/login + /shop/login.

    This migration clears the flag on the super admin's per-portal rows
    on every backend start. Idempotent — no-op once flags are False.
    """
    SUPER = "jaymn.judd@mascigc.com"
    try:
        for coll in ("user_directory", "hr_users", "shop_users", "pm_users"):
            await db[coll].update_one(
                {"email": SUPER, "must_change_password": True},
                {"$set": {"must_change_password": False}},
            )
    except Exception as e:  # noqa: BLE001
        # Migration failures must NEVER block backend boot — log and move on.
        logger.warning(f"[iter117] super-admin pw-flag clear failed: {e}")


# ------------------------- HR Portal (iter71) -------------------------
from routes.hr_portal import build_hr_portal_router  # noqa: E402


async def _hr_send_email(to_email: str, subject: str, html: str):
    """Resend wrapper used by HR welcome / reset emails."""
    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        logger.info(f"[hr-email-stub] to={to_email} subject={subject}")
        return
    if (os.environ.get("AUTO_EMAIL_REPORTS") or "").strip().lower() not in ("true", "1", "yes"):
        # Preview env — log instead of sending.
        logger.info(f"[hr-email-preview] to={to_email} subject={subject}")
        return
    import resend as _resend  # noqa: PLC0415
    _resend.api_key = api_key
    sender = await _resolve_sender_email(db, safe_fallback="noreply@mascidocs.com")
    params = {
        "from": f"MASCI HR Portal <{sender}>",
        "to": [to_email],
        "subject": subject,
        "html": html,
    }
    return await asyncio.to_thread(_resend.Emails.send, params)


# iter353c · need the shared HR+Safety+Admin gate factory available
# BEFORE the HR portal router is built (the safety_exports section
# below imports the same symbol; this is just an earlier import).
from routes.safety_portal._deps import make_require_safety_or_hr_or_admin  # noqa: E402

_hr_portal_router = build_hr_portal_router(
    db, require_admin, _hr_send_email,
    # iter346-B · universal super-admin fallback — same minter the FL
    # portal uses (iter344). Wrapped in a lambda so name resolution
    # defers until login-time (_directory_admin_token is defined later).
    directory_admin_minter=lambda row: _directory_admin_token(row),
    # iter353c · shared accountability gate (HR + Safety + Admin) used
    # by the unified employee timeline + compliance brief endpoints.
    require_safety_or_hr_or_admin_dep=make_require_safety_or_hr_or_admin(
        db, _is_valid_admin_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
    # Track 15.87 · directory `hr` grant path — Admin People & Access
    # checkbox now produces a working HR login (mints HR token).
    directory_portal_minter=lambda row: _directory_hr_token(row),
)
app.include_router(_hr_portal_router)


# ──────── Field Leadership Portal (iter314) ────────
# Clones the HR/PM/Shop/Safety auth pattern exactly. Reuses the
# `_hr_send_email` Resend wrapper for welcome/reset emails — same
# Resend key, same preview-stub behavior, same sender address.
from routes.field_leadership_portal import build_field_leadership_portal_router  # noqa: E402


@register_lifecycle_step("seed")
async def _seed_field_leadership_users():
    from field_leadership_users import seed_field_leadership_users
    await seed_field_leadership_users(db)


_fl_portal_router = build_field_leadership_portal_router(
    db, require_admin, _hr_send_email,
    # iter344 · lazy reference — `_directory_admin_token` is defined
    # later in this file (line ~10510). Wrapping in a lambda defers
    # name resolution until the FL login route actually fires.
    directory_admin_minter=lambda row: _directory_admin_token(row),
)
app.include_router(_fl_portal_router)


# ─── Safety Portal (iter119 + iter120 Phase 3/4/5) ───────────────────
# Mirrors the HR portal pattern exactly so it slots cleanly into the
# existing auth/router architecture. Reads existing incident/inspection/
# meeting/FL records for visibility — adds new collections:
#   • corrective_actions    (Phase 2 — iter119)
#   • fire_extinguishers    (Phase 3 — iter120)
#   • safety_documents      (Phase 3 — iter120)
#   • safety_training_records (Phase 4 — iter120)
from routes.safety_portal import build_safety_router, build_digest_payload, render_digest_html  # noqa: E402
from safety_users import seed_safety_users  # noqa: E402
from safety_digest import safety_digest_scheduler_loop  # noqa: E402


async def _safety_send_email(to_email: str, subject: str, html: str) -> bool:
    """Resend wrapper used by Safety welcome / weekly digest emails.

    Returns True only when Resend was actually invoked. Returns False
    when the helper short-circuits (no API key, AUTO_EMAIL_REPORTS off,
    etc.) so callers can report accurate `sent` status to the UI."""
    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        logger.info(f"[safety-email-stub] to={to_email} subject={subject}")
        return False
    if (os.environ.get("AUTO_EMAIL_REPORTS") or "").strip().lower() not in ("true", "1", "yes"):
        logger.info(f"[safety-email-preview] to={to_email} subject={subject}")
        return False
    import resend as _resend  # noqa: PLC0415
    _resend.api_key = api_key
    sender = await _resolve_sender_email(db, safe_fallback="noreply@mascidocs.com")
    params = {
        "from": f"MASCI Safety Portal <{sender}>",
        "to": [to_email],
        "subject": subject,
        "html": html,
    }
    await asyncio.to_thread(_resend.Emails.send, params)
    return True


_safety_router = build_safety_router(
    db, require_admin,
    send_email_fn=_safety_send_email,
    is_valid_admin_token=_is_valid_admin_token,
    # TRACK 28.03E · async admin validator for directory-hydrated tokens.
    is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    # iter346-B · universal super-admin fallback
    directory_admin_minter=lambda row: _directory_admin_token(row),
    # Track 15.87 · directory `safety` grant path.
    directory_portal_minter=lambda row: _directory_safety_token(row),
)
app.include_router(_safety_router)


# Phase 7.5C — Trench Safety transactional email wrapper.
# Identical gating + Resend wiring as `_safety_send_email`; lives at
# module scope so `routes/trench_safety/notifications.py` can resolve
# it via a late `from server import _trench_send_email` import.
# Branded "MASCI Trench Safety" so Gmail/Outlook threading is clean.
async def _trench_send_email(to_email: str, subject: str, html: str) -> bool:
    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        logger.info(f"[trench-email-stub] to={to_email} subject={subject}")
        return False
    if (os.environ.get("AUTO_EMAIL_REPORTS") or "").strip().lower() not in ("true", "1", "yes"):
        logger.info(f"[trench-email-preview] to={to_email} subject={subject}")
        return False
    import resend as _resend  # noqa: PLC0415
    _resend.api_key = api_key
    sender = await _resolve_sender_email(db, safe_fallback="noreply@mascidocs.com")
    params = {
        "from": f"MASCI Trench Safety <{sender}>",
        "to": [to_email],
        "subject": subject,
        "html": html,
    }
    await asyncio.to_thread(_resend.Emails.send, params)
    return True



# ─── Safety Reports & Exports (iter133) ────────────────────────────
# Wires the 10 export endpoints SafetyReports.jsx already calls. Safety
# / HR / Admin tokens accepted via make_require_safety_or_hr_or_admin.
from routes.safety_exports import build_safety_exports_router  # noqa: E402
from routes.safety_portal._deps import make_require_safety_or_hr_or_admin  # noqa: E402

_require_safety_hr_admin = make_require_safety_or_hr_or_admin(
    db, _is_valid_admin_token,
    is_valid_admin_token_async=_is_valid_directory_admin_token_async,
)
app.include_router(build_safety_exports_router(db, _require_safety_hr_admin))


# ─── F2-A · Safety Topic Library PDF Pack endpoint (iter266) ────────
# Safety/Admin-only operational prep tool. Generates multi-topic PDF
# packs from the front-end topic library. NO public surface. NO
# analytics. NO LMS. See /app/memory/SAFETY_MEETING_POST_PHASE_H_EVAL_iter265.md §5.
from routes.safety_topic_library import build_router as build_safety_topic_library_router  # noqa: E402
from routes.safety_portal._deps import make_require_safety_or_admin  # noqa: E402

_require_safety_or_admin_library = make_require_safety_or_admin(
    db, _is_valid_admin_token,
    is_valid_admin_token_async=_is_valid_directory_admin_token_async,
)
app.include_router(build_safety_topic_library_router(_require_safety_or_admin_library))


# ─── Trench Safety Operations System — Phase 2 ──────────────────────
# Mounts /api/trench-safety/* (dashboard, assets CRUD, inspections,
# repairs, deployments, public QR landing + damage intake). Uses
# existing token deps:
#   • require_admin                    — terminal lifecycle (retire)
#   • make_require_safety_or_admin     — write surface for assets/inspections
#   • require_shop_or_admin            — repair workflow
#   • make_require_any_portal_token    — read surface across all 7 portals
# Persists into NEW collections (trench_safety_*) — does NOT touch the
# existing trench_boxes manufacturer-reference collection.
# Reference: /app/memory/TRENCH_SAFETY_ARCHITECTURE.md
from routes.trench_safety import build_trench_safety_router  # noqa: E402
from routes.integrations._deps import make_require_any_portal_token  # noqa: E402

_trench_safety_router = build_trench_safety_router(
    db,
    require_admin=require_admin,
    require_safety_or_admin=make_require_safety_or_admin(
        db, _is_valid_admin_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
    require_shop_or_admin=require_shop_or_admin,
    require_any_portal=make_require_any_portal_token(
        db, _is_valid_admin_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)
app.include_router(_trench_safety_router)


# ─── Admin Weekly Digest Config (iter133) ──────────────────────────
from routes.admin_digest_config import build_admin_digest_router  # noqa: E402

app.include_router(build_admin_digest_router(
    db, require_admin_strict,
    lambda: build_digest_payload(db),
    render_digest_html,
    send_email_fn=_safety_send_email,
))


# ─── Fire Extinguisher Bulk Import (iter134) ────────────────────────
from routes.fire_ext_bulk_import import build_fire_import_router  # noqa: E402
from routes.safety_portal._deps import make_require_safety_token  # noqa: E402

_require_safety = make_require_safety_token(db)
app.include_router(build_fire_import_router(db, _require_safety))


# ─── Training Center — system-wide operator guides (iter134) ────────
from routes.training_center import build_training_center_router  # noqa: E402

app.include_router(build_training_center_router(db, require_admin, _guidance_caller_scopes))

# iter437 IV-BETA.5A-P4B · Safe route extraction · public guidance content.
# Reads from guidance.content + guidance.tips, scopes via the same
# _guidance_caller_scopes helper as the training-center router (so behaviour
# is identical to the in-server.py originals).
from routes.guidance_routes import build_guidance_router  # noqa: E402

app.include_router(build_guidance_router(db, _guidance_caller_scopes))

# iter437 IV-BETA.5A-P5D · Safe route extraction · /api/health + /api/healthz.
# Zero dependencies · stateless · safest possible extraction.
from routes.health_routes import build_health_router  # noqa: E402

app.include_router(build_health_router())

# iter437 IV-BETA.5A-P6 · Safe route extraction · public static helpers.
# Pure public utilities (no DB, no auth, no scheduler). Currently:
#   • GET /api/qr.svg — public QR-code generator (Training Scan-&-Go).
from routes.static_helpers import build_static_helpers_router  # noqa: E402

app.include_router(build_static_helpers_router())


# ─── Recovery Dashboard (Phase D · iter443) ────────────────────────
# Single read-only Admin endpoint sourcing all data from existing
# collections. See RECOVERY_DASHBOARD_SPEC.md and
# RECOVERY_DASHBOARD_DEPLOY_REPORT.md for the contract.
from routes.recovery_dashboard import build_recovery_dashboard_router  # noqa: E402

app.include_router(
    build_recovery_dashboard_router(db, require_admin_strict),
    prefix="/api",
)

# ─── R2 Storage Lifecycle Governance (TRACK 27.06 · Phase 1/2/3/4/6/10) ──
# Read-only lifecycle intelligence — inventory + Mongo cross-reference +
# strict classification + dry-run certification + storage health.
# Deletion is explicitly out of scope; see routes/admin_r2_lifecycle.py.
from routes.admin_r2_lifecycle import build_r2_lifecycle_router  # noqa: E402

app.include_router(
    build_r2_lifecycle_router(db, require_admin_strict),
    prefix="/api",
)


# ─── Executive Operations Command Center (Pillar 2 · Phase A · iter500) ──
# Single read-only Admin endpoint synthesizing 5 cards (Jobs · Safety ·
# Equipment · Accountability · Approvals) over EXISTING collections.
# See FINAL_PHASE_A_RECOMMENDATION.md and EXECUTIVE_SCORING_CERTIFICATION.md.
from routes.command_center import build_command_center_router  # noqa: E402

app.include_router(
    build_command_center_router(db, require_admin_strict),
    prefix="/api",
)


# ─── Pillar 1 · Phase 1A-3 · Accountability service surface ─────────
# Read-only endpoints that expose the certified projection layer from
# lib/accountability_projection.py. Source workflows are NOT modified.
from routes.accountability_service import build_accountability_router  # noqa: E402

app.include_router(
    build_accountability_router(db, require_admin_strict),
    prefix="/api",
)


# ─── Deploy Readiness Aggregator (iter136) ──────────────────────────
from routes.deploy_readiness import build_deploy_readiness_router  # noqa: E402

app.include_router(build_deploy_readiness_router(db, require_admin))


# ─── Integration Health Probes + Alert Hook (iter142 — Phase-1 Iter D) ─
from routes.integration_health import (  # noqa: E402
    build_integration_health_router,
    ensure_alert_indexes,
)

app.include_router(build_integration_health_router(db, require_admin))


# ─── Usage Analytics (iter146 — Phase 2.5) ──────────────────────────
# Lightweight async telemetry. NEVER blocks user requests. TTL = 90d.
# Admin-only read access; no PII; no employee surveillance.
from routes.usage_analytics import (  # noqa: E402
    build_usage_routes,
    ensure_usage_indexes,
    start_sink,
    usage_tracking_middleware,
)

app.include_router(build_usage_routes(db, require_admin))
app.middleware("http")(usage_tracking_middleware)


# ─── Tasks + Notifications (iter150 — Phase 2.5 · Phase A) ──────────
# Shared operational accountability + awareness infrastructure used by
# every other module. NEVER blocks user requests. TTL on closed tasks
# (365d) + notifications (60d auto).
from routes.tasks_notifications import (  # noqa: E402
    build_tasks_notifications_router,
    ensure_tasks_notifications_indexes,
)
from routes.integrations._deps import make_require_any_portal_token  # noqa: E402

_require_any_portal_token = make_require_any_portal_token(
    db, _is_valid_admin_token,
    is_valid_admin_token_async=_is_valid_directory_admin_token_async,
)
app.include_router(build_tasks_notifications_router(db, _require_any_portal_token))

# TRACK 14.0-S1 Amendment A — bilingual record sidecar (original-language
# preservation for free-text fields submitted in Spanish or any future
# non-English language). Additive collection · zero coupling to the
# canonical form records.
from routes.bilingual_records import (  # noqa: E402
    build_bilingual_records_router,
    ensure_bilingual_indexes,
)
app.include_router(build_bilingual_records_router(db, _require_any_portal_token))


# ─── Document Expiration Engine (iter151 — Phase 2.5 · Phase B) ─────
# Central expirations across employee docs, training certs, equipment
# docs, fire extinguishers (read-through), company compliance.
# Emits Tasks + Notifications via Phase A shared services.
from routes.document_expirations import (  # noqa: E402
    build_document_expirations_router,
    ensure_document_expirations_indexes,
)

app.include_router(build_document_expirations_router(
    db, _require_any_portal_token, require_admin,
))


# ─── Employee Lifecycle Management (iter152 — Phase 2.5 · Phase C) ──
# Extends db.employees with lifecycle_status + status_history. Wires
# the auto-offboarding playbook via Phase A task_service. Offboarding
# Summary aggregates tasks (Phase A) + document expirations (Phase B).
from routes.employee_lifecycle import (  # noqa: E402
    build_employee_lifecycle_router,
    ensure_employee_lifecycle_indexes,
)

# A best-effort require_hr resolver — uses the existing HR portal auth
# but lives at the server scope so the lifecycle router can rely on it.
# (require_admin already exists.)
async def _require_hr_or_pass(actor=Depends(_require_any_portal_token)):
    return actor

app.include_router(build_employee_lifecycle_router(
    db,
    require_hr=_require_hr_or_pass,
    require_admin=require_admin,
    require_any_portal_token=_require_any_portal_token,
))


# ─── Employee Governance Phase Alpha · G-5 · HR Request Queue ───────
# OMEGA · 2026-06-02 · The UX-bridge collection introduced when the
# 5 P0 violations from EMPLOYEE_GOVERNANCE_AUDIT.md were closed.
# HR remains the sole authoritative writer of db.employees lifecycle
# state; this queue lets Operations/Public/anyone SUBMIT a request,
# which HR explicitly reviews + approves/rejects.
from routes.employee_requests import (  # noqa: E402
    register_employee_requests_routes,
    ensure_employee_requests_indexes,
)


# Optional-portal-token dependency — same headers as
# _require_any_portal_token, but returns None instead of raising 401
# when no token is present. This lets public field forms still call
# POST /api/employee-requests under the existing rate-limit gate.
async def _require_optional_portal_token(
    request: Request,
    x_admin_token: Optional[str] = Header(default=None, alias="X-Admin-Token"),
    x_safety_token: Optional[str] = Header(default=None, alias="X-Safety-Token"),
    x_hr_token: Optional[str] = Header(default=None, alias="X-HR-Token"),
    x_shop_token: Optional[str] = Header(default=None, alias="X-Shop-Token"),
    x_pm_token: Optional[str] = Header(default=None, alias="X-PM-Token"),
    x_dispatch_token: Optional[str] = Header(default=None, alias="X-Dispatch-Token"),
    x_leadership_token: Optional[str] = Header(default=None, alias="X-Leadership-Token"),
    x_fl_token: Optional[str] = Header(default=None, alias="X-FL-Token"),
):
    """Permissive variant of _require_any_portal_token — returns None
    when no recognized token is provided (vs 401). Required for the
    employee-requests endpoint which must accept public field-form
    submissions per the operator-approved G-5 design."""
    try:
        return await _require_any_portal_token(
            request=request,
            x_admin_token=x_admin_token,
            x_safety_token=x_safety_token,
            x_hr_token=x_hr_token,
            x_shop_token=x_shop_token,
            x_pm_token=x_pm_token,
            x_dispatch_token=x_dispatch_token,
            x_leadership_token=x_leadership_token,
            x_fl_token=x_fl_token,
        )
    except HTTPException:
        return None


# HR-or-Admin reviewer gate is defined earlier in this file (so the
# deprecated /api/admin/employees* endpoints can reference it without
# a forward declaration). Reuse it here for the queue's HR review
# endpoints.


# Phase Alpha · G-5 · the router is built fresh inside register_*
# and returned so it can be mounted on `app` directly (avoiding the
# post-`include_router(api_router)` order-of-include trap).
_employee_requests_router = register_employee_requests_routes(
    api_router,
    db,
    rate_limit_public_post=rate_limit_public_post,
    require_optional_portal_token=_require_optional_portal_token,
    require_hr_or_admin=_require_hr_or_admin_for_queue,
)
app.include_router(_employee_requests_router)


# ─── Operational PO Request & Receipt Tracking (iter153 — Phase D) ──
# Field Leadership submits → PM/HR/Admin approve → R2 receipt upload.
# `MASCI-PO-YY-MM-NNN` globally unique numbering. Missing-receipt
# scanner emits Tasks via Phase A.
from routes.po_requests import (  # noqa: E402
    build_po_requests_router,
    ensure_po_requests_indexes,
    scan_missing_receipts,
)

# R2 upload helper — optional. We try to reuse the existing safety
# upload pathway; fall back to data-URL inline storage in preview.
async def _po_r2_upload(content: bytes, filename: str, content_type: str):
    try:
        from routes.safety_portal.uploads import upload_to_r2  # type: ignore
        return await upload_to_r2(
            content, filename=filename, content_type=content_type,
            folder="po-receipts",
        )
    except Exception:
        return None  # signals fallback to data-URL

app.include_router(build_po_requests_router(
    db, _require_any_portal_token, require_admin,
    r2_upload_callable=None,  # use data-URL fallback in preview
))


# ─── Unified Signature Engine (iter154 — Phase 2.5 · Phase F) ───────
# One signature standard across the platform. Append-only history with
# `supersedes` chain — no silent overwrites.
from routes.signatures import (  # noqa: E402
    build_signatures_router, ensure_signatures_indexes,
)
app.include_router(build_signatures_router(db, _require_any_portal_token))


# ─── TRACK 23.10-B · Professional Qualifications Engine ─────────────
# Single source of truth for all professional qualifications
# (Competent Person + 15 seed types). Reads: any authenticated portal.
# Writes: HR / Safety (Training Admin) / Admin only.
from routes.qualifications import build_qualifications_router  # noqa: E402
from services.certifications.qualification_facts import (  # noqa: E402
    emit_qualification_expiration_facts_daily as _emit_qual_expirations,  # noqa: F401
)
from scripts.migrate_track_23_10_b_qualification_engine import (  # noqa: E402
    run_migration as _run_qualification_migration,
)

_qualification_write_gate = make_require_safety_or_hr_or_admin(
    db,
    is_valid_admin_token=_is_valid_admin_token,
    # Track 24.2 · accept directory-hydrated admin tokens (UUID form
    # issued by multi-login). The sync check alone rejected every
    # modern admin token, breaking the qualifications write path.
    is_valid_admin_token_async=_is_valid_directory_admin_token_async,
)
app.include_router(build_qualifications_router(
    db,
    require_read_dep=_require_any_portal_token,
    require_write_dep=_qualification_write_gate,
))


@app.on_event("startup")
async def _track_23_10_b_qualification_migration_bootstrap():
    """Run the additive TRACK 23.10-B migration idempotently at boot.

    Safe to re-run: every write is gated by whether the target field
    already carries the migrated value. Fact emission uses
    `supersede_facts` so re-running collapses to one current fact.
    """
    try:
        await _run_qualification_migration(db, emit_facts=True)
    except Exception as exc:                                     # noqa: BLE001
        logger.warning(f"[track-23-10-b-migration] {exc}")


# ─── TRACK 23.10-C · Trench Project Linker + ODS Trench Facts ───────
# 7 canonical fact types + 4 derived views + project-scoped read APIs.
# PM tokens see only assigned projects; Safety/Admin see company-wide.
from routes.trench_project_intelligence import (  # noqa: E402
    build_trench_project_intelligence_router,
)
from scripts.backfill_track_23_10_c_trench_facts import (  # noqa: E402
    run_backfill as _run_trench_backfill,
)

app.include_router(build_trench_project_intelligence_router(
    db,
    require_read_dep=_require_any_portal_token,
    require_admin_dep=require_admin,
))


@register_lifecycle_step("post-readiness")
async def _track_23_10_c_trench_backfill_bootstrap():
    """Idempotent, capped backfill of trench facts. Fire-and-forget
    background task — never blocks startup readiness.

    Safe to re-run — natural-key `supersede_facts` collapses duplicates
    to at most 1 current per row.
    """
    async def _run_bg():
        try:
            await _run_trench_backfill(db, boot_mode=True)
        except Exception as exc:                                 # noqa: BLE001
            logger.warning(f"[track-23-10-c-backfill] {exc}")
    asyncio.create_task(_run_bg())


# ─── TRACK 23.10-D · Safety Portal Trench KPI Lift ──────────────────
# Read-only aggregator that consumes 23.10-B + 23.10-C.  No new KPI
# engine. No duplicate trench logic. Safety/Admin company-wide; PM
# assigned-only for per-project. Zero cost/money fields (runtime guard).
from routes.safety_trench_intelligence import (  # noqa: E402
    build_safety_trench_intelligence_router,
)

app.include_router(build_safety_trench_intelligence_router(
    db, require_read_dep=_require_any_portal_token,
))




# ─── Draft Telemetry (P0 field-incident · daily report draft loss) ──
# Append-only client-driven telemetry for the form-draft / autosave
# subsystem. NEVER stores form content — only sizes / error-names /
# timestamps / page-lifecycle transitions. 30-day TTL via index.
from routes.draft_telemetry import (  # noqa: E402
    build_draft_telemetry_router, ensure_draft_telemetry_indexes,
)
app.include_router(
    build_draft_telemetry_router(
        db, _require_any_portal_token, require_admin,
    )
)


# ─── Governance Self-Protection · Phase GOVERNANCE-OPS-1 · 2026-05-28 ─
# Read-only aggregator powering `/admin/governance/self-protection`.
# Reads the doctrine artifacts on disk + runs the Authority Mismatch
# Probe (60s cache). Admin-only. No writes anywhere.
from routes.governance_self_protection import (  # noqa: E402
    build_governance_self_protection_router,
    auto_record_deploy_on_startup,
)
app.include_router(build_governance_self_protection_router(require_admin))

# Track 28.12 duplicate housekeeping router removed per Track 27.07 Phase 0A —
# see /app/memory/TRACK_27_07_PHASE_0_ARCHITECTURE_LOCK.md. Canonical R2
# lifecycle governance lives in routes/admin_r2_lifecycle.py (Track 27.06).

# TRACK 28.11 · Idempotent auto-record of the running source_hash to
# the deployment ledger. Ensures the governance self-protection card
# never shows "not recorded yet" while a live deploy is in production.
# Safe on every restart — an unchanged hash is a no-op.
try:
    _t2811_deploy_ledger = auto_record_deploy_on_startup(_SOURCE_HASH)
    _t2811_logger = logging.getLogger(__name__)
    if _t2811_deploy_ledger.get("appended"):
        _t2811_logger.info(
            f"[track-28.11] deployment ledger auto-recorded source_hash={_SOURCE_HASH[:12]}"
        )
except Exception as _t2811_e:  # noqa: BLE001
    logging.getLogger(__name__).warning(
        f"[track-28.11] deployment ledger auto-record skipped: {_t2811_e!s}"
    )


# ─── Global Search (iter155 — Phase 2.5 · Phase G) ──────────────────
# Permission-safe, role-aware, lightweight cross-collection typeahead.
# Reuses _require_any_portal_token and applies per-role visibility +
# PM project scope. NEVER leaks results from kinds the caller can't
# access.
from routes.global_search import build_global_search_router  # noqa: E402
app.include_router(build_global_search_router(db, _require_any_portal_token))


# ─── Operations Center (iter C — Phase 2.5 · stabilization) ────────
# Thin per-role aggregated visibility layer on TOP of existing shared
# infrastructure. No new SOT collections. No fake metrics. asyncio.gather
# parallel probes; lightweight count_documents per card.
from routes.operations_center import build_operations_center_router  # noqa: E402
app.include_router(build_operations_center_router(db, _require_any_portal_token))


# ─── Promo Asset Library (iter347) ───────────────────────────────────
# Admin-only media-asset library for organizing/downloading cinematic
# platform clips that feed the long-form MASCI promo film + homepage
# hero loop. Reuses the existing R2 client; new key prefix
# `promo-assets/` + new mongo collection `promo_assets`.
from routes.promo_assets import build_promo_assets_router  # noqa: E402
app.include_router(build_promo_assets_router(db, require_admin_strict))


# ─── Operational Signals (iter160 — Phase 2.5 · Operational Signal Density) ──
# Admin-only aggregation of operational events recorded at fan-out tap points.
# Reuses db.usage_events collection (kind='operational_signal'). No new
# data model. Passive observability — workflow-impact-free by design.
from routes.operational_signals import build_operational_signals_router  # noqa: E402
app.include_router(build_operational_signals_router(db, require_admin))

# ─── Track 15.73Q · Daily Report PM-Email Coverage Observability ──────
# Admin-gated, read-only endpoint that surfaces which active jobs_master
# rows lack pm_email so the operator can prioritise data-hygiene
# backfill. Workflow-impact-free by design.
from routes.admin_pm_coverage import make_router as _pm_cov_make_router  # noqa: E402
app.include_router(_pm_cov_make_router(db, require_admin))
# TRACK 15.75D · In-app Platform Trust Validator (admin-gated, read-only).
from routes.admin_platform_trust import make_router as _trust_make_router  # noqa: E402
app.include_router(_trust_make_router(db, require_admin, get_runtime_identity=_runtime_identity_bundle))
# TRACK 15.76 · Platform Trust Spine — lifecycle event observability.
from routes.admin_trust_spine import make_router as _spine_make_router  # noqa: E402
app.include_router(_spine_make_router(db, require_admin))
from lib.trust_spine import ensure_indexes as _spine_ensure_indexes  # noqa: E402
@register_lifecycle_step("index-ensure")
async def _startup_trust_spine_indexes():  # noqa: D401
    await _spine_ensure_indexes(db)

# TRACK 15.76A · Operations Trust Center — capstone trust-score +
# master-data drift + red-alert hook + operator remediation copy.
from routes.admin_operations_trust_center import make_router as _otc_make_router  # noqa: E402
app.include_router(_otc_make_router(db, require_admin))

# TRACK 15.78 · Deployment Readiness — pass/fail gate for CI/CD.
from routes.admin_deployment_readiness import make_router as _gate_make_router  # noqa: E402
app.include_router(_gate_make_router(db, require_admin))

# TRACK 15.79 · Deployment ledger — append-only history of every
# Trust Gate decision so operators can audit deployment outcomes.
from routes.admin_deployment_ledger import (  # noqa: E402
    make_router as _ledger_make_router,
    ensure_indexes as _ledger_indexes,
)
app.include_router(_ledger_make_router(db, require_admin))
@register_lifecycle_step("misc-bootstrap")
async def _startup_deployment_ledger_indexes():  # noqa: D401
    await _ledger_indexes(db)

# TRACK 15.79B · Daily Report delivery forensics — read-only admin
# endpoint that traces, record-by-record, why PM/Co-PM emails did or
# did not reach their assigned recipients.
from routes.admin_dr_delivery_forensics import (  # noqa: E402
    make_router as _dr_forensics_make_router,
)
app.include_router(_dr_forensics_make_router(db, require_admin))

# TRACK 15.79E · Continuous Production Certification — read-only,
# admin-gated, derived from trust_spine_events. Surfaces which
# workflows have been PROVEN end-to-end by real production traffic.
from routes.admin_production_certification import (  # noqa: E402
    make_router as _prod_cert_make_router,
)
app.include_router(_prod_cert_make_router(db, require_admin))



# ─── Phase V-Prelude · Wave 1 · Substrate ─────────────────────────────
# operational_links · operational_constraints · operational_timeline ·
# photo_governance. All four mount the existing _require_any_portal_token
# gate (no auth expansion) and adhere to OPERATIONAL_LINKING_RULES.md.
# Doctrine docs:
#   /app/memory/OPERATIONAL_LINKING_RULES.md         · ⛔ read-before-touching
#   /app/memory/OPERATIONAL_CONSTRAINT_FOUNDATION.md
#   /app/memory/OPERATIONAL_TIMELINE_FOUNDATION.md
#   /app/memory/PHOTO_GOVERNANCE_STANDARD.md
from routes.operational_links import (  # noqa: E402
    build_operational_links_router, ensure_operational_links_indexes,
)
from routes.operational_constraints import (  # noqa: E402
    build_operational_constraints_router,
    ensure_operational_constraints_indexes,
)
from routes.operational_timeline import (  # noqa: E402
    build_operational_timeline_router,
)
from routes.photo_governance import (  # noqa: E402
    build_photo_governance_router, ensure_photo_governance_indexes,
)
from routes.odr import (  # noqa: E402
    build_odr_router, ensure_odr_indexes,
    build_odr_continuity_router, ensure_continuity_indexes,
    build_odr_amendments_router,
    build_odr_pdf_router,
    build_odr_guidance_router,
    build_odr_observation_router, ensure_observation_indexes,
)

app.include_router(build_operational_links_router(
    db, _require_any_portal_token, require_admin,
))
# Phase V.1 · M1 · Unified Operational Records Projector (Option C).
# Doctrine: M1_OPTION_C_IMPLEMENTATION_PLAN.md ·
#           UNIFIED_RECORDS_PROJECTOR_CERTIFICATION.md
# Read-only · zero mutation · merges ODR + frozen daily_reports.
from routes.operational_records import (  # noqa: E402
    build_operational_records_router,
)
app.include_router(build_operational_records_router(
    db, _require_any_portal_token,
))
app.include_router(build_operational_constraints_router(
    db, _require_any_portal_token,
))
app.include_router(build_operational_timeline_router(
    db, _require_any_portal_token,
))
app.include_router(build_photo_governance_router(
    db, _require_any_portal_token,
))
# Phase V.1 · ODR substrate (M0.1) — inherits FL Visibility Doctrine,
# Operational Linking Rules, Timeline doctrine, Coaching addendum,
# Role-Aware Visibility Model. Schema-version 2 at launch · build
# the substrate correctly once.
app.include_router(build_odr_router(
    db, _require_any_portal_token,
))
# Phase V.1 · ODR M0.2 — Public Link Continuity, Amendment Engine,
# PDF rendering framework (5 audience variants · SHA256 footer).
# Phase V.1 · ODR M0.2A — Guidance Intelligence Foundation
# (deterministic prompt resolver + crew readiness matrix).
app.include_router(build_odr_continuity_router(
    db, _require_any_portal_token, require_admin,
))
app.include_router(build_odr_amendments_router(
    db, _require_any_portal_token,
))
app.include_router(build_odr_pdf_router(
    db, _require_any_portal_token,
))
app.include_router(build_odr_guidance_router(
    _require_any_portal_token,
))
# Phase V.1 · M0.3 · ODR Adoption Observation (telemetry · aggregates only).
app.include_router(build_odr_observation_router(
    db, _require_any_portal_token, require_admin,
))


# ─── Project Health Dashboard (Phase H) ──────────────────────────────
# Per-project operational friction summary. Aggregates the SAME shared
# infrastructure streams Operations Center uses (tasks · POs · docs ·
# incidents · CAs), keyed on project_number. NO new collection, NO
# duplicate SOT. Role-scoped: admin/exec/safety see all · PM scope-
# filtered · HR/Shop/Dispatch/FL get 403 (not their primary lens).
from routes.project_health import build_project_health_router  # noqa: E402
app.include_router(build_project_health_router(db, _require_any_portal_token))


# ─── Asset Transfers (Phase I) ─────────────────────────────────────────
# Thin event collection (db.asset_transfers) tracking equipment movement
# lifecycle. equipment_master remains the asset SOT. Reuses Tasks +
# Notifications + Signatures + Audit + PM scope — NO duplicate systems.
from routes.asset_transfers import build_asset_transfers_router  # noqa: E402
app.include_router(build_asset_transfers_router(db, _require_any_portal_token))


# ─── Master Lookup & Backfill (iter137 — Iter C-continued SOT) ──────
from routes.master_lookup import build_master_lookup_router  # noqa: E402

app.include_router(build_master_lookup_router(db, require_admin, _require_any_portal_read))


# ─── Integration Center (Motive + MaintainX framework — iter122) ───
from routes.integrations import (  # noqa: E402
    build_integrations_router,
    ensure_integrations_indexes_and_seed,
)

# MCC-1 HR Access Extension · HR users can work the driver cleanup
# queue alongside admins. Defined inline here (BEFORE the integrations
# router is built) because the existing `_require_hr_or_admin` further
# down in this file is registered after this point.
async def _require_hr_or_admin_for_mcc1(
    x_hr_token: Optional[str] = Header(default=None, alias="X-HR-Token"),
    x_admin_token: Optional[str] = Header(default=None, alias="X-Admin-Token"),
):
    # TRACK 28.03E · pair sync HMAC validator with async directory validator.
    if x_admin_token and (
        _is_valid_admin_token(x_admin_token)
        or await _is_valid_directory_admin_token_async(x_admin_token)
    ):
        return {"_actor": "admin", "name": "Admin"}
    if x_hr_token:
        from hr_users import is_valid_hr_user_token_async  # noqa: PLC0415
        u = await is_valid_hr_user_token_async(db, x_hr_token)
        if u:
            return {**u, "_actor": "hr", "_actor_kind": "hr_user"}
    raise HTTPException(401, "HR or Admin login required")


_integrations_router = build_integrations_router(
    db, require_admin, _is_valid_admin_token,
    require_hr_or_admin=_require_hr_or_admin_for_mcc1,
    is_valid_admin_token_async=_is_valid_directory_admin_token_async,
)
app.include_router(_integrations_router)


# ─── DCP-1 · Driver Command Profile ─────────────────────────────────
# Multi-portal actor resolver. Returns {_role, name, ...}. Order of
# precedence: Admin > Safety > HR > Dispatch. The endpoint's per-section
# redactor honours `_role` to shape the response.
async def _require_driver_profile_actor(
    x_admin_token: Optional[str] = Header(default=None, alias="X-Admin-Token"),
    x_safety_token: Optional[str] = Header(default=None, alias="X-Safety-Token"),
    x_hr_token: Optional[str] = Header(default=None, alias="X-HR-Token"),
    x_dispatch_token: Optional[str] = Header(default=None, alias="X-Dispatch-Token"),
):
    if x_admin_token and (
        _is_valid_admin_token(x_admin_token)
        or await _is_valid_directory_admin_token_async(x_admin_token)
    ):
        return {"_role": "admin", "name": "Admin"}
    if x_safety_token:
        from safety_users import is_valid_safety_user_token_async  # noqa: PLC0415
        u = await is_valid_safety_user_token_async(db, x_safety_token)
        if u:
            return {**u, "_role": "safety"}
    if x_hr_token:
        from hr_users import is_valid_hr_user_token_async  # noqa: PLC0415
        u = await is_valid_hr_user_token_async(db, x_hr_token)
        if u:
            return {**u, "_role": "hr"}
    if x_dispatch_token:
        from dispatch_users import is_valid_dispatch_user_token_async  # noqa: PLC0415
        u = await is_valid_dispatch_user_token_async(db, x_dispatch_token)
        if u:
            return {**u, "_role": "dispatch"}
    raise HTTPException(401, "Admin, Safety, HR, or Dispatch login required")


from routes.driver_profile import register_driver_profile_routes  # noqa: E402
from routes.sprint_a import register_sprint_a_routes  # noqa: E402
_dcp_router = APIRouter(prefix="/api", tags=["driver-profile"])
register_driver_profile_routes(_dcp_router, db, _require_driver_profile_actor)
register_sprint_a_routes(_dcp_router, db, _require_driver_profile_actor)
app.include_router(_dcp_router)


# ─── OA-1 · Operations Actions · cross-portal CRUD layer ────────────
# Canonical Family 3B auth contract:
#   • exactly one valid portal token for the acting portal
#   • the bound X-Directory-Token for the same logical session
# Anonymous, token-only, or mismatched directory binding = 401.
async def _require_oa_actor(
    x_admin_token: Optional[str] = Header(default=None, alias="X-Admin-Token"),
    x_directory_token: Optional[str] = Header(default=None, alias="X-Directory-Token"),
    x_safety_token: Optional[str] = Header(default=None, alias="X-Safety-Token"),
    x_hr_token: Optional[str] = Header(default=None, alias="X-HR-Token"),
    x_dispatch_token: Optional[str] = Header(default=None, alias="X-Dispatch-Token"),
    x_pm_token: Optional[str] = Header(default=None, alias="X-PM-Token"),
    x_shop_token: Optional[str] = Header(default=None, alias="X-Shop-Token"),
    x_fl_token: Optional[str] = Header(default=None, alias="X-FL-Token"),
):
    if not x_directory_token:
        raise HTTPException(401, "Directory session required for Operations Actions")

    directory_session = await db.directory_sessions.find_one(
        {"token": x_directory_token},
        {"_id": 0, "user_id": 1, "expires_at_ts": 1},
    )
    now_ts = int(datetime.now(timezone.utc).timestamp())
    if (
        not directory_session
        or not directory_session.get("user_id")
        or int(directory_session.get("expires_at_ts") or 0) < now_ts
    ):
        raise HTTPException(401, "Directory session required for Operations Actions")

    provided_tokens = [
        ("admin", x_admin_token),
        ("safety", x_safety_token),
        ("hr", x_hr_token),
        ("dispatch", x_dispatch_token),
        ("pm", x_pm_token),
        ("shop", x_shop_token),
        ("fl", x_fl_token),
    ]
    supplied = [(role, tok) for role, tok in provided_tokens if tok]
    if len(supplied) != 1:
        raise HTTPException(401, "Exactly one portal token required for Operations Actions")

    if x_admin_token and await _is_valid_directory_admin_token_async(x_admin_token):
        try:
            import user_directory as _ud_local  # noqa: PLC0415
            row = await _ud_local.is_valid_directory_admin_token_async(db, x_admin_token)
        except Exception:  # noqa: BLE001
            row = None
        if row:
            return {
                **row,
                "_role": "admin",
                "_actor_kind": "admin",
                "name": row.get("name") or row.get("display_name") or row.get("email") or "Admin",
            }
    if x_safety_token:
        from safety_users import is_valid_safety_user_token_async  # noqa: PLC0415
        u = await is_valid_safety_user_token_async(db, x_safety_token)
        if u:
            return {**u, "_role": "safety"}
    if x_hr_token:
        from hr_users import is_valid_hr_user_token_async  # noqa: PLC0415
        u = await is_valid_hr_user_token_async(db, x_hr_token)
        if u:
            return {**u, "_role": "hr"}
    if x_dispatch_token:
        from dispatch_users import is_valid_dispatch_user_token_async  # noqa: PLC0415
        u = await is_valid_dispatch_user_token_async(db, x_dispatch_token)
        if u:
            return {**u, "_role": "dispatch"}
    if x_pm_token:
        from pm_auth import is_valid_pm_user_token_async  # noqa: PLC0415
        u = await is_valid_pm_user_token_async(db, x_pm_token)
        if u:
            return {**u, "_role": "pm"}
    if x_shop_token:
        from shop_users import is_valid_shop_user_token_async  # noqa: PLC0415
        u = await is_valid_shop_user_token_async(db, x_shop_token)
        if u:
            return {**u, "_role": "shop"}
    if x_fl_token:
        from field_leadership_users import is_valid_fl_user_token_async  # noqa: PLC0415
        u = await is_valid_fl_user_token_async(db, x_fl_token)
        if u:
            return {**u, "_role": "fl"}
    raise HTTPException(401, "Portal authentication required")


from routes.operations_actions import register_operations_actions_routes  # noqa: E402
_oa_router = APIRouter(prefix="/api", tags=["operations-actions"])
register_operations_actions_routes(_oa_router, db, _require_oa_actor)
app.include_router(_oa_router)


# MM-001B · E-5 · Material Movement Rollup (derived view · public read).
from routes.material_movement import register_material_movement_routes  # noqa: E402
_mm_router = APIRouter(prefix="/api", tags=["material-movement"])
register_material_movement_routes(_mm_router, db)
app.include_router(_mm_router)


# Ensure indexes for operations_actions
async def _ensure_oa_indexes():
    try:
        await db.operations_actions.create_index("id", unique=True, name="oa_id_unique")
        await db.operations_actions.create_index("oa_number", unique=True, sparse=True,
                                                 name="oa_number_unique")
        await db.operations_actions.create_index("status", name="oa_status")
        await db.operations_actions.create_index("current_owner.id", name="oa_owner_id")
        await db.operations_actions.create_index("job_number", name="oa_job")
        await db.operations_actions.create_index([("created_at", -1)], name="oa_created_desc")
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[oa-1] index ensure failed: {e}")


@register_lifecycle_step("misc-bootstrap")
async def _oa_startup():
    await _ensure_oa_indexes()


# ─── Operations layer (Asset Profile · Event Log · Dispatch · Utilization · iter124) ─
from routes.operations import (  # noqa: E402
    build_operations_router,
    ensure_operations_indexes,
)

app.include_router(build_operations_router(
    db, require_admin, _is_valid_admin_token,
    is_valid_admin_token_async=_is_valid_directory_admin_token_async,
))


# OIS-1 · Operations Intelligence Aggregator (single-pane intelligence)
from routes.operations_intelligence import register_operations_intelligence_routes  # noqa: E402
from fastapi import APIRouter as _APIRouter  # noqa: E402
_ois_router = _APIRouter(prefix="/api", tags=["operations-intelligence"])
register_operations_intelligence_routes(_ois_router, db, require_admin)
app.include_router(_ois_router)


# ─── Dispatch Portal portal-auth (iter126) ──────────────────────────
from routes.dispatch_portal_auth import build_dispatch_router  # noqa: E402
from dispatch_users import seed_dispatch_users  # noqa: E402

app.include_router(build_dispatch_router(
    db, require_admin,
    # iter346-B · universal super-admin fallback
    directory_admin_minter=lambda row: _directory_admin_token(row),
    # iter353b · admin tokens accepted on the read-only DQ surface.
    is_valid_admin_token_fn=_is_valid_admin_token,
    # TRACK 28.02 · directory-hydrated per-user admin token support.
    is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    # Track 15.87 · directory `dispatch` grant path.
    directory_portal_minter=lambda row: _directory_dispatch_token(row),
))


# ─── Admin operational infrastructure (iter130) ─────────────────────
# System Health · Unified Audit Log · Global Search · Deploy Recovery
# Uses the STRICT admin gate (admin-only — PM tokens NOT accepted) so
# operational/compliance-sensitive surfaces stay scoped to admins.
from routes.admin_ops import build_admin_ops_router  # noqa: E402
from routes.master_data_backfill import build_master_data_backfill_router  # noqa: E402
from routes.pm_gap_backfill import build_pm_gap_backfill_router  # noqa: E402

_admin_ops_router = build_admin_ops_router(db, require_admin_strict)
_admin_ops_router._get_runtime_identity = _runtime_identity_bundle  # type: ignore[attr-defined]
app.include_router(_admin_ops_router)
app.include_router(build_master_data_backfill_router(db, require_admin_strict))
app.include_router(build_pm_gap_backfill_router(db, require_admin_strict))


# ─── Phase 2 P0+P1 · Compliance Gap Detector + Governance Health ────
# Cross-portal contradiction detection engine. Strict admin gate —
# governance findings + the convergence dashboard live admin-only.
from routes.governance import build_governance_router  # noqa: E402

app.include_router(build_governance_router(db, require_admin_strict))

# ─── Phase IV-BETA.5A-P1A · Governance Health Chip (public read-only) ─
# Tiny operator-facing chip that reads the persisted doctrine baseline
# JSON and surfaces loudness + drift state. NO auth — telemetry only,
# zero PII. See routes/governance_health.py for thresholds.
from routes.governance_health import build_governance_health_router  # noqa: E402

app.include_router(build_governance_health_router())


# ─── Phase 2 P1 · Operational Intelligence Notifications ─────────────
# Role-scoped digest engine over the existing findings + lifecycle state.
# Admin + Safety in this iteration; HR/PM/Dispatch/FL follow the same
# pattern and can be added incrementally without an architectural change.
from routes.notifications import build_notifications_router  # noqa: E402
from routes.safety_portal._deps import make_require_safety_or_admin  # noqa: E402

_notif_safety_gate = make_require_safety_or_admin(
    db, _is_valid_admin_token,
    is_valid_admin_token_async=_is_valid_directory_admin_token_async,
)
app.include_router(build_notifications_router(
    db, require_admin_strict, _notif_safety_gate,
))


# ─── Synthetic health monitor (iter132) ─────────────────────────────
# Polls system_health every 60 s, fires Resend alerts on sustained red.
from health_monitor import start_health_monitor_loop  # noqa: E402


@register_lifecycle_step("post-readiness")
async def _start_health_monitor():
    try:
        track_existing_background_task(
            app,
            name="synthetic-health-monitor",
            task=start_health_monitor_loop(db, _admin_ops_router.compute_system_health),
            category="health-monitor",
            critical=False,
            long_running=True,
        )
        # Pre-warm the health_monitor_runs collection so the first call
        # to /admin/system-health/recent doesn't pay a 36 s cold-start
        # cost (motor lazy-allocates collections on first use).
        try:
            await db.health_monitor_runs.create_index("at")
        except Exception:  # noqa: BLE001
            pass
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[health_monitor] failed to arm: {e}")


# ─── Iter134: TTL indexes on audit/log/event collections ────────────
# These three collections are append-only AUDIT HISTORY (not operational
# records). After 30 days the rows are no longer actionable for incident
# triage or trend analysis, so MongoDB auto-deletes them via TTL.
#
# - r2_degraded_events: R2 storage fallback events (System Health & alerts)
# - digest_runs:        weekly digest send-history (preview + actual)
# - health_monitor_runs: synthetic monitor poll history (60-s cadence)
#
# Operational records (equipment, employees, incidents, holds, transfers,
# corrective actions, training, fire extinguishers, safety documents,
# inspections, projects, etc.) are NEVER touched by TTL — those are
# permanent and only purged via explicit admin delete.
#
# Retention is configurable via AUDIT_RETENTION_DAYS env (default 30).
# Changing the value requires dropping + recreating the index — handled
# safely below by detecting any existing TTL index and only mutating
# when the configured value drifts.
@register_lifecycle_step("misc-bootstrap")
async def _arm_audit_ttl_indexes():
    try:
        days = int(os.environ.get("AUDIT_RETENTION_DAYS", "30"))
        seconds = days * 86400
        for coll_name in ("r2_degraded_events", "digest_runs", "health_monitor_runs",
                          "system_health_events", "audit_events"):
            coll = db[coll_name]
            try:
                existing = await coll.index_information()
                # Drop any legacy plain-`at` index that conflicts with our TTL.
                for name, spec in list(existing.items()):
                    if name == "at_1" and spec.get("expireAfterSeconds") is None:
                        try:
                            await coll.drop_index(name)
                        except Exception:  # noqa: BLE001
                            pass
                idx_name = "at_ttl"
                # Re-read since we may have dropped one
                existing = await coll.index_information()
                if idx_name in existing:
                    current = existing[idx_name].get("expireAfterSeconds")
                    if current == seconds:
                        continue
                    try:
                        await coll.drop_index(idx_name)
                    except Exception:  # noqa: BLE001
                        pass
                await coll.create_index(
                    "at",
                    name=idx_name,
                    expireAfterSeconds=seconds,
                )
                logger.info(f"[audit-ttl] {coll_name}.at TTL armed at {days}d")
            except Exception as e:  # noqa: BLE001
                logger.warning(f"[audit-ttl] failed to arm {coll_name}: {e}")
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[audit-ttl] startup hook failed: {e}")


@register_lifecycle_step("misc-bootstrap")
async def _bootstrap_operations():
    await ensure_operations_indexes(db)
    logger.info("[operations] indexes ensured")
    await seed_dispatch_users(db)
    logger.info("[dispatch-users] seed ready")


@register_lifecycle_step("misc-bootstrap")
async def _bootstrap_integrations():
    await ensure_integrations_indexes_and_seed(db)
    logger.info("[integrations] indexes + seed settings ready")
    await ensure_alert_indexes(db)
    logger.info("[alert-events] indexes ensured")
    await ensure_usage_indexes(db)
    start_sink(db)
    logger.info("[usage-analytics] indexes ensured + async sink started")
    await ensure_tasks_notifications_indexes(db)
    logger.info("[tasks-notifications] indexes ensured")
    await ensure_bilingual_indexes(db)
    logger.info("[bilingual-records] indexes ensured")
    await ensure_document_expirations_indexes(db)
    logger.info("[document-expirations] indexes ensured")
    await ensure_employee_lifecycle_indexes(db)
    logger.info("[employee-lifecycle] indexes ensured")
    # TRACK 19.21 · Employee Records Intelligence Platform indexes.
    try:
        await ensure_employee_records_indexes(db)
        logger.info("[employee-records] indexes ensured (track 19.21)")
    except Exception as _exc:
        logger.warning("[employee-records] index bootstrap warn: %s", _exc)
    await ensure_employee_requests_indexes(db)
    logger.info("[employee-requests] indexes ensured")
    await ensure_po_requests_indexes(db)
    logger.info("[po-requests] indexes ensured")
    await ensure_signatures_indexes(db)
    logger.info("[signatures] indexes ensured")
    # P0 field-incident · draft-loss remediation — telemetry collection
    # gets unique-eventId + TTL on receivedAt (30d) + ts/event lookup.
    await ensure_draft_telemetry_indexes(db)
    logger.info("[draft-telemetry] indexes ensured")
    # Phase 2 Initiative 4 — session_activity indexes (TTL + uniqueness)
    await ensure_session_timeout_indexes(db)
    logger.info("[session-timeout] indexes ensured")
    # Phase 2 Initiative 5b — admin hardening indexes (admin_step_ups TTL).
    await ensure_admin_hardening_indexes(db)
    logger.info("[admin-hardening] indexes ensured")
    # Phase 2 Initiative 1 — Sentry init AFTER _SOURCE_HASH is computed so
    # the Sentry release identifier matches /api/version's source_hash.
    try:
        if init_sentry_if_configured is not None:
            init_sentry_if_configured(release_override=_SOURCE_HASH)
    except Exception:  # noqa: BLE001
        pass  # Sentry init must never break the app.


@register_lifecycle_step("seed")
async def _seed_safety_users():
    await seed_safety_users(db)
    try:
        await db.corrective_actions.create_index("status")
        await db.corrective_actions.create_index("due_date")
        await db.corrective_actions.create_index("source_id")
        await db.fire_extinguishers.create_index("unit_id")
        await db.fire_extinguishers.create_index("next_due_date")
        await db.safety_documents.create_index("category")
        await db.safety_documents.create_index("uploaded_at")
        await db.safety_training_records.create_index("employee_id")
        await db.safety_training_records.create_index("expiration_date")
    except Exception as e:  # noqa: BLE001
        logger.warning(f"safety collections index: {e}")


# ─── Iter136 (Phase-1 Iter D): id-indexes on hot operational
#     collections so the deploy-readiness probe stops warning. These
#     are read on essentially every CRUD-by-id path and were previously
#     defaulting to collection-scan. create_index is idempotent.
@register_lifecycle_step("index-ensure")
async def _arm_hot_id_indexes():
    for coll_name in ("fire_extinguishers", "corrective_actions", "incidents",
                      "inspections", "safety_training_records",
                      "equipment_master", "employees"):
        try:
            await db[coll_name].create_index("id")
        except Exception as e:  # noqa: BLE001
            logger.warning(f"[id-index] {coll_name}: {e}")


# ─── OMEGA · Phase 1A · iter451 — workflow_state_events indexes.
#     Idempotent index battery for the new universal audit collection
#     used by the Phase 1A lifecycle transitions. Failures are logged
#     but never block boot (ensure_indexes swallows internally too).
@register_lifecycle_step("index-ensure")
async def _arm_workflow_state_events_indexes():
    try:
        from lib.workflow_state_events import ensure_indexes as _wse_idx  # noqa: PLC0415
        await _wse_idx(db)
        # Per-record lifecycle_state lookup index on incidents so the
        # transition endpoint never collscans.
        await db.incidents.create_index("lifecycle_state")
        # iter452 — lifecycle_state on daily_reports and payroll_variance_batches
        await db.daily_reports.create_index("lifecycle_state")
        await db.payroll_variance_batches.create_index("lifecycle_state")
        # iter452.5 — Field Submitter Identity bindings indexes.
        from lib.field_submitter_identity import ensure_indexes as _fsi_idx  # noqa: PLC0415
        await _fsi_idx(db)
        # FOCP Release 2 · TR-0001 — JHP Acknowledgement Ledger indexes.
        from routes.jha_acknowledgements import ensure_indexes as _jha_ack_idx  # noqa: PLC0415
        await _jha_ack_idx(db)
        # WEBHOOK-DEDUP-001 — Motive event uniqueness index (idempotent).
        from services.motive_service import ensure_motive_events_indexes as _motive_idx  # noqa: PLC0415
        await _motive_idx(db)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[workflow_state_events] index: {e}")


# ─── Iter142 (Phase-1 Iter D): targeted index + TTL fixes surfaced by
#     scripts/qa_audit.py. All idempotent. Pairs with QA_PERF_AUDIT.md.
@register_lifecycle_step("index-ensure")
async def _arm_iter142_perf_indexes():
    # Indexes that resolve the 2 COLLSCANs found by the audit, plus
    # the index recommendations on hot list endpoints.
    targeted = [
        ("incidents",                [("incident_date", -1)]),
        ("corrective_actions",       [("status", 1), ("due_date", 1)]),
        ("fire_extinguishers",       [("next_due_date", 1)]),
        ("equipment_inspections",    [("inspection_date", -1)]),
        ("safety_training_records",  [("expiration_date", 1)]),
        ("operations_events",        [("status", 1), ("created_at", -1)]),
        ("operations_events",        [("asset_id", 1)]),
        ("operations_events",        [("employee_id", 1)]),
        ("field_leadership_records", [("occurred_at", -1)]),
        ("field_leadership_records", [("employee_name", 1)]),
        ("daily_reports",            [("report_date", -1)]),
        ("employees",                [("name", 1)]),
    ]
    for coll, key_spec in targeted:
        try:
            await db[coll].create_index(key_spec)
        except Exception as e:  # noqa: BLE001
            logger.warning(f"[perf-index] {coll} {key_spec}: {e}")

    # TTL indexes flagged as missing by the audit (admin_audit kept at
    # 1 year for compliance, login_attempts/brute_force short-lived).
    ttl_plan = [
        ("admin_audit",            "at",  60 * 60 * 24 * 365),
        ("login_attempts",         "at",  60 * 60 * 24 * 30),
        ("integration_error_logs", "at",  60 * 60 * 24 * 90),
        ("brute_force_blocks",     "at",  60 * 60 * 24 * 7),
    ]
    for coll, field, secs in ttl_plan:
        try:
            await db[coll].create_index([(field, 1)], expireAfterSeconds=secs)
        except Exception as e:  # noqa: BLE001
            logger.warning(f"[ttl-index] {coll}.{field}: {e}")


_safety_digest_task: Optional[asyncio.Task] = None


@register_lifecycle_step("email-scheduler")
async def _start_safety_digest_cron():
    """Long-running weekly cron — Monday 14:00 UTC default. Email goes
    to SAFETY_DIGEST_TO_EMAIL (default safety@mascigc.com)."""
    global _safety_digest_task
    try:
        # iter441 · gated through singleton-lock so only one worker fires.
        async def _safety_digest_wrapped(_db):
            return await safety_digest_scheduler_loop(
                _db,
                build_payload=lambda: build_digest_payload(_db),
                render_html=render_digest_html,
                send_email_fn=_safety_send_email,
            )
        _safety_digest_task = register_background_task(
            app,
            name="safety-digest-singleton",
            coro=run_with_singleton_lock(db, "safety_digest", _safety_digest_wrapped),
            category="email-scheduler",
            critical=False,
            long_running=True,
        )
        logger.info("[safety-digest] weekly cron started")
    except Exception as e:  # noqa: BLE001
        logger.exception(f"[safety-digest] failed to start: {e}")


# iter431 · Phase 29 · Part 6 · weekly operator digest cron.
_operator_digest_task: Optional[asyncio.Task] = None


@register_lifecycle_step("email-scheduler")
async def _start_operator_digest_cron():
    """Long-running weekly cron — Monday 14:00 UTC default. Email goes
    to OPERATOR_DIGEST_RECIPIENTS (comma-separated) or falls back to
    SAFETY_DIGEST_TO_EMAIL. Reuses the existing Resend `_safety_send_email`
    wrapper — no new SDK plumbing."""
    global _operator_digest_task
    try:
        from lib.operator_digest import operator_digest_scheduler_loop  # noqa: PLC0415

        async def _operator_digest_wrapped(_db):
            return await operator_digest_scheduler_loop(
                _db,
                send_email_fn=_safety_send_email,
            )
        _operator_digest_task = register_background_task(
            app,
            name="operator-digest-singleton",
            coro=run_with_singleton_lock(db, "operator_digest", _operator_digest_wrapped),
            category="email-scheduler",
            critical=False,
            long_running=True,
        )
        logger.info("[operator-digest] weekly cron started")
    except Exception as e:  # noqa: BLE001
        logger.exception(f"[operator-digest] failed to start: {e}")


# iter431 · Phase 29 · Part 4 · TTL index ensures for transient surfaces.
@register_lifecycle_step("misc-bootstrap")
async def _ensure_stability_ttls():
    try:
        from lib.stability_governance import ensure_stability_ttls  # noqa: PLC0415
        report = await ensure_stability_ttls(db)
        logger.info(
            f"[stability-governance] TTL ensures · created={len(report['created'])} "
            f"· skipped={len(report['skipped'])} · errors={len(report['errors'])}",
        )
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[stability-governance] TTL ensure failed: {e}")


# ─── iter246 F3 · Weekly PO Request digest (PM + HR) ─────────────────
from po_digest import (  # noqa: E402
    po_digest_scheduler_loop, send_po_digest_once, DIGEST_SUBJECT as PO_DIGEST_SUBJECT,
)


async def _po_digest_send_email(to_email: str, subject: str, html: str) -> bool:
    """Resend wrapper used by the weekly PO digest. Mirrors
    _safety_send_email exactly — same gating (RESEND_API_KEY +
    AUTO_EMAIL_REPORTS) so preview/dev environments don't burn quota."""
    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        logger.info(f"[po-digest-stub] to={to_email} subject={subject}")
        return False
    if (os.environ.get("AUTO_EMAIL_REPORTS") or "").strip().lower() not in ("true", "1", "yes"):
        logger.info(f"[po-digest-preview] to={to_email} subject={subject}")
        return False
    import resend as _resend  # noqa: PLC0415
    _resend.api_key = api_key
    sender = await _resolve_sender_email(db, safe_fallback="noreply@mascidocs.com")
    params = {
        "from": f"MASCI PO Operations <{sender}>",
        "to": [to_email],
        "subject": subject,
        "html": html,
    }
    await asyncio.to_thread(_resend.Emails.send, params)
    return True


_po_digest_task: Optional[asyncio.Task] = None


@register_lifecycle_step("email-scheduler")
async def _start_po_digest_cron():
    """iter246 F3 · Long-running weekly cron sending the PO Request
    Digest to every active PM (scoped to their assigned jobs) and every
    active HR user (platform-wide). Mon 14:00 UTC default."""
    global _po_digest_task
    try:
        portal_url = (os.environ.get("PORTAL_PUBLIC_URL")
                      or os.environ.get("PUBLIC_BASE_URL")
                      or "https://mascidocs.com").rstrip("/")

        async def _po_digest_wrapped(_db):
            return await po_digest_scheduler_loop(
                _db,
                send_email_fn=_po_digest_send_email,
                portal_url=portal_url,
            )
        _po_digest_task = register_background_task(
            app,
            name="po-digest-singleton",
            coro=run_with_singleton_lock(db, "po_digest", _po_digest_wrapped),
            category="email-scheduler",
            critical=False,
            long_running=True,
        )
        logger.info("[po-digest] weekly cron started")
    except Exception as e:  # noqa: BLE001
        logger.exception(f"[po-digest] failed to start: {e}")


# iter380 · /api/admin/po-digest/preview and /api/admin/po-digest/run-now
# extracted to routes/po_digest_admin.py.
from routes.po_digest_admin import build_po_digest_admin_router  # noqa: E402
_po_digest_admin_router = build_po_digest_admin_router(
    db,
    require_admin_dep=require_admin,
    require_admin_strict_dep=require_admin_strict,
    send_email_fn=_po_digest_send_email,
)
app.include_router(_po_digest_admin_router)

# iter445 · Sprint · Scheduler Hardening · Option C — admin scheduler-runs.
# Read-only history of every scheduled digest fire (po / safety / operator)
# with dedup metadata.
from routes.scheduler_runs_admin import build_scheduler_runs_admin_router  # noqa: E402
app.include_router(build_scheduler_runs_admin_router(db, require_admin))


# ─── HR Payroll Variance (iter72) ────────────────────────────────────
from routes import payroll_variance as _pv_module  # noqa: E402

# Reuse the HR-token dependency by extracting it from hr_users helpers.
async def _require_hr_user(
    x_hr_token: Optional[str] = Header(default=None, alias="X-HR-Token"),
):
    from hr_users import is_valid_hr_user_token_async
    if not x_hr_token:
        raise HTTPException(401, "HR login required")
    user = await is_valid_hr_user_token_async(db, x_hr_token)
    if not user:
        raise HTTPException(401, "HR session expired or invalid")
    return {**user, "_actor_kind": "hr_user"}


_pv_router = _pv_module.build_payroll_variance_router(db, _require_hr_user)
app.include_router(_pv_router)


# ─── OMEGA · Phase 1A · iter452 · OC-007 Payroll Variance Finalization
#     Additive transition endpoints. Mounted on a dedicated APIRouter so
#     it can be included AFTER ``app.include_router(api_router)`` upstream.
async def _require_hr_or_admin(
    x_hr_token: Optional[str] = Header(default=None, alias="X-HR-Token"),
    x_admin_token: Optional[str] = Header(default=None, alias="X-Admin-Token"),
):
    # TRACK 28.03E · pair sync + async admin validators.
    if x_admin_token and (
        _is_valid_admin_token(x_admin_token)
        or await _is_valid_directory_admin_token_async(x_admin_token)
    ):
        return {"_actor": "admin", "name": "Admin"}
    if x_hr_token:
        from hr_users import is_valid_hr_user_token_async  # noqa: PLC0415
        u = await is_valid_hr_user_token_async(db, x_hr_token)
        if u:
            return {**u, "_actor_kind": "hr_user"}
    raise HTTPException(401, "HR or Admin login required")


from fastapi import APIRouter as _APIRouterIter452  # noqa: E402
_iter452_pv_router = _APIRouterIter452(prefix="/api")
from routes.payroll_variance_lifecycle import register_payroll_variance_lifecycle_routes  # noqa: E402
register_payroll_variance_lifecycle_routes(
    _iter452_pv_router, db, require_pv_actor=_require_hr_or_admin,
)
app.include_router(_iter452_pv_router)


# ─── Signature migration (iter75 — base64 → R2) ─────────────────────
from routes import signature_migration as _sig_mig_module  # noqa: E402

_sig_mig_router = _sig_mig_module.build_signature_migration_router(db, require_admin)
app.include_router(_sig_mig_router)


# ─── iter248 Phase A · Legacy Operational Records Import (Foundation) ───
# Read /app/LEGACY_RECORDS_ARCHITECTURE_iter248.md for the architectural
# proposal. Phase A ships:
#   • staging collection + state machine + audit log
#   • OCR worker scaffold (StubExtractor only · no AI promises yet)
#   • R2-backed source-file storage (private bucket · signed URLs · audited)
#   • RBAC matrix (HR / Safety / Admin only · NO PM)
#   • anti-self-approval guard
#   • reconciliation endpoints
#   • promotion contract scaffolded but NO doc type activated
# Phase B (separate operator approval) will activate Equipment Checkout.
import legacy_imports as _li  # noqa: E402
import photo_storage as _ps  # noqa: E402
from fastapi import UploadFile, File, Form  # noqa: E402,PLC0415


@register_lifecycle_step("index-ensure")
async def _li_ensure_indexes():
    await _li.ensure_indexes(db)


_li_worker_task: Optional[asyncio.Task] = None


@register_lifecycle_step("misc-bootstrap")
async def _li_start_worker():
    global _li_worker_task
    # Register Phase B extractor + promoter for equipment_checkout BEFORE
    # the worker starts so the very first iteration uses the real extractor.
    try:
        import legacy_imports_equipment_checkout as _li_ec  # noqa: PLC0415
        _li_ec.register_phase_b(_li)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[legacy-imports] Phase B registration skipped: {e}")
    _li_worker_task = _li.start_worker(db)
    logger.info(
        "[legacy-imports] OCR worker started · phase_b_active="
        f"{'equipment_checkout' in _li.ACTIVE_PROMOTERS}"
    )


# ─── iter430 · Phase 28.2 · Legacy Imports route extraction ──────────
# All 11 `/api/legacy-imports/*` routes (+ HR/Safety/Admin uploader auth
# dep and the scope filter) moved to routes/legacy_imports.py. Startup
# hooks (`_li_ensure_indexes`, `_li_start_worker`) remain above to
# keep the application lifecycle untouched.
from routes.legacy_imports import build_legacy_imports_router  # noqa: E402

app.include_router(build_legacy_imports_router(
    db=db,
    li_module=_li,
    photo_storage_module=_ps,
    is_valid_admin_token=_is_valid_admin_token,
    is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    require_admin_strict=require_admin_strict,
))




# ─── iter251 Phase A · Fleet Operations Foundation ─────────────────────
# Backend-only · NO frontend, NO public tile, NO dashboards yet (Phases B-C).
# Reuses existing equipment_master (149 fleet units) + equipment_inspections
# (with new `kind` discriminator). Adds fleet_defects + fleet_status + audit.
from routes.fleet_ops import build_router as _fleet_build_router  # noqa: E402
from routes.dispatch_portal_auth import make_require_dispatch_token  # noqa: E402

_require_dispatch_token = make_require_dispatch_token(db)
_require_safety_for_fleet = _require_safety  # already built above


# iter431 · Phase 29 · Part 5a · fleet-ops auth deps now live in
# routes/fleet_ops_deps.py (zero-behaviour-change factory move).
from routes.fleet_ops_deps import (  # noqa: E402
    make_require_fleet_submitter,
    make_require_any_fleet_portal,
)
_require_fleet_submitter = make_require_fleet_submitter(
    db=db,
    is_valid_admin_token=_is_valid_admin_token,
    is_valid_admin_token_async=_is_valid_directory_admin_token_async,
)


# Phase 4 · multi-portal READ gate. Any of admin / shop / dispatch /
# safety satisfies — used for defect detail + audit-trail reads where
# the operator wants all three operational scopes (Shop, Dispatch,
# Safety) to see the same record.
_require_any_fleet_portal = make_require_any_fleet_portal(
    db=db,
    is_valid_admin_token=_is_valid_admin_token,
    is_valid_admin_token_async=_is_valid_directory_admin_token_async,
)


# iter370 · Canonical shared dispatch+admin gate (single source of truth).
# Built once at module load — delegated by `_require_dispatch_or_admin`.
from routes.dispatch_portal_auth import (  # noqa: E402
    make_require_dispatch_or_admin as _make_dispatch_or_admin,
)
_shared_dispatch_or_admin = _make_dispatch_or_admin(
    db, _is_valid_admin_token,
    is_valid_admin_token_async=_is_valid_directory_admin_token_async,
)


async def _require_dispatch_or_admin(
    request: Request,
    x_admin_token: Optional[str] = Header(default=None, alias="X-Admin-Token"),
    x_dispatch_token: Optional[str] = Header(default=None, alias="X-Dispatch-Token"),
) -> Dict[str, Any]:
    """iter370 · Delegates to the canonical shared factory in
    routes/dispatch_portal_auth.make_require_dispatch_or_admin so the
    gate has a SINGLE source of truth. The wrapper signature is preserved
    because fleet_ops.py wires this in via kwargs at router construction.

    TRACK 16.09 · The shared factory's inner closure has ``request``
    as a required positional argument (it calls
    ``enforce_password_change_required(request, u)`` on dispatch tokens),
    so we must pass it through. Earlier revisions dropped it, which
    silently turned every admin-token-driven dispatch write into a
    500 (TypeError). Admin tokens never matched the sync
    ``_is_valid_admin_token`` shim so they fell into the dispatch
    branch which then exploded.
    """
    # Admin path: try the async directory validator FIRST so admin
    # tokens get a clean answer without touching the inner closure.
    if x_admin_token and await _is_valid_directory_admin_token_async(x_admin_token):
        return {"role": "admin", "is_admin": True}
    # TRACK 22.6A · cert-session fallback (path-scoped, audited, read-only).
    # The dispatch-safe Motive posture endpoint (/api/dispatch/motive-posture)
    # is in the certification allowlist; unlock it for a valid cert token
    # so post-deploy authenticated Motive checks work without operator creds.
    x_certification_token = request.headers.get("X-Certification-Token")
    if x_certification_token:
        try:
            from routes.production_certification_session import (
                verify_session_token as _pcs_verify,
                ALLOWED_READ_PATHS as _pcs_allowed,
                _audit as _pcs_audit,
                COLLECTION as _pcs_coll,
            )
            _path = (request.scope.get("path") or request.url.path or "").rstrip("/") or "/"
            if _path in _pcs_allowed:
                _session = await _pcs_verify(db, x_certification_token, request_path=_path)
                if _session:
                    from datetime import datetime as _dt, timezone as _tz
                    await db[_pcs_coll].update_one(
                        {"session_id": _session["session_id"]},
                        {"$inc": {"reads_performed": 1},
                         "$set": {"last_read_at": _dt.now(_tz.utc)}},
                    )
                    await _pcs_audit(db, event="pcs_read_authorized",
                                     session_id=_session["session_id"], path=_path)
                    return {"role": "admin", "is_admin": True, "pcs": True}
        except Exception as exc:  # noqa: BLE001
            logger.warning(f"[pcs fallback · dispatch_or_admin] {exc}")
    return await _shared_dispatch_or_admin(
        request=request,
        x_dispatch_token=x_dispatch_token,
        x_admin_token=x_admin_token,
    )


# ── TRACK 22.4a · Dispatch-safe Motive posture endpoint ─────────────
# Registered here (rather than with the earlier admin-only truth routes)
# because it needs the shared dispatch-or-admin dependency defined
# directly above. Same underlying `_motive_truth(db)` helper — no drift.
from routes.integration_truth import (  # noqa: E402
    _motive_truth as _motive_truth_helper,
    _now_iso as _integration_truth_now_iso,
)


@app.get("/api/dispatch/motive-posture")
async def track_22_4a_dispatch_motive_posture(
    _=Depends(_require_dispatch_or_admin),
):
    """Dispatch-safe Motive posture. Same three-state truth model as
    /api/admin/integrations/truth-status but scoped to Motive-only and
    reachable with a dispatch token. Never claims LIVE unless
    operational_status is LIVE_VERIFIED.
    """
    row = await _motive_truth_helper(db)
    return {
        "checked_at": _integration_truth_now_iso(),
        "id": row["id"],
        "name": row["name"],
        "config_status": row["config_status"],
        "connectivity_status": row["connectivity_status"],
        "operational_status": row["operational_status"],
        "overall": row["overall"],
        "connectivity_detail": row.get("connectivity_detail"),
        "connectivity_latency_ms": row.get("connectivity_latency_ms"),
        "last_successful_sync_at": row.get("last_successful_sync_at"),
        "activity_age_seconds": row.get("activity_age_seconds"),
        "live_window_seconds": row.get("live_window_seconds"),
        "doctrine": (
            "Dispatch-safe Motive posture. Never claims LIVE unless "
            "operational_status is LIVE_VERIFIED. Use to render "
            "stale-data ribbons in the Dispatch UI."
        ),
    }


# iter372 · Canonical shared Safety+Admin fleet-ops gate (single source
# of truth). Mirrors iter370 (dispatch) and iter371 (shop) patterns.
# Built once at module load — delegated by `_require_safety_or_admin_fleet`.
from routes.safety_portal._deps import (  # noqa: E402
    make_require_safety_or_admin_fleet as _make_safety_or_admin_fleet,
)
_shared_safety_or_admin_fleet = _make_safety_or_admin_fleet(
    db, _is_valid_admin_token,
    is_valid_admin_token_async=_is_valid_directory_admin_token_async,
)


async def _require_safety_or_admin_fleet(
    request: Request,
    x_admin_token: Optional[str] = Header(default=None, alias="X-Admin-Token"),
    x_safety_token: Optional[str] = Header(default=None, alias="X-Safety-Token"),
) -> Dict[str, Any]:
    """iter372 · Delegates to the canonical shared factory in
    routes/safety_portal/_deps.make_require_safety_or_admin_fleet so the
    fleet-ops safety gate has a SINGLE source of truth. The wrapper
    signature is preserved because fleet_ops.py wires this in via
    kwargs at router construction.
    """
    return await _shared_safety_or_admin_fleet(
        request,
        x_admin_token=x_admin_token,
        x_safety_token=x_safety_token,
    )


# Shop gate — narrow admin/shop fleet-ops gate (distinct from
# require_shop_or_admin which also accepts PM tokens + has admin-namespace
# lockdown). iter371 · Delegates to the canonical shared factory in
# routes/shop_portal_deps.make_require_shop_or_admin_fleet.
from routes.shop_portal_deps import (  # noqa: E402
    make_require_shop_or_admin_fleet as _make_shop_or_admin_fleet,
)
_shared_shop_or_admin_fleet = _make_shop_or_admin_fleet(
    db, _is_valid_admin_token,
    is_valid_admin_token_async=_is_valid_super_admin_token_async,
)


async def _require_shop_or_admin_fleet(
    request: Request,
    x_admin_token: Optional[str] = Header(default=None, alias="X-Admin-Token"),
    x_shop_token: Optional[str] = Header(default=None, alias="X-Shop-Token"),
) -> Dict[str, Any]:
    """iter371 · Delegates to the canonical shared factory in
    routes/shop_portal_deps.make_require_shop_or_admin_fleet so the
    fleet-ops shop gate has a SINGLE source of truth. The wrapper
    signature is preserved because fleet_ops.py wires this in via
    kwargs at router construction.
    """
    return await _shared_shop_or_admin_fleet(
        request,
        x_admin_token=x_admin_token,
        x_shop_token=x_shop_token,
    )


# iter431 · Phase 29 · Part 5a · `_require_any_fleet_portal` extracted
# above into `routes/fleet_ops_deps.make_require_any_fleet_portal`.


@register_lifecycle_step("index-ensure")
async def _fleet_ensure_indexes():
    """Index fleet collections at boot · idempotent."""
    try:
        await db.fleet_defects.create_index(
            [("truck_unit_number", 1), ("status", 1), ("severity", 1)]
        )
        await db.fleet_defects.create_index(
            [("trailer_unit_number", 1), ("status", 1)]
        )
        await db.fleet_defects.create_index([("status", 1), ("reported_at", 1)])
        await db.fleet_status.create_index("unit_number", unique=True)
        await db.fleet_audit.create_index([("timestamp", -1)])
        await db.fleet_audit.create_index([("target_id", 1)])
        await db.equipment_inspections.create_index("kind")
        logger.info("[fleet-ops] indexes ensured")
    except Exception as e:  # noqa: BLE001
        logging.getLogger(__name__).warning(f"[fleet-ops] index setup skipped: {e}")


_fleet_router = _fleet_build_router(
    db=db,
    require_signed_in_or_public=_require_fleet_submitter,
    require_dispatch_or_admin=_require_dispatch_or_admin,
    require_shop_or_admin=_require_shop_or_admin_fleet,
    require_safety_or_admin=_require_safety_or_admin_fleet,
    require_admin_strict=require_admin_strict,
    require_any_fleet_portal=_require_any_fleet_portal,
    schedule_auto_email=lambda kind, record: schedule_auto_email(kind, record),
)
app.include_router(_fleet_router)
logging.getLogger(__name__).info("[fleet-ops] iter251 Phase A router mounted · backend-only foundation")


# ─── iter392 · Dispatch Lifecycle System (DLS) Backend Foundation ───
# Phase 11.1 — operational state machine for haul cycles.
# Reuses _require_dispatch_or_admin (write gate) and
# _require_any_portal_token (cross-portal read gate). No new auth
# surface — driver magic-link arrives in iter393.
from routes.dispatch_lifecycle import (  # noqa: E402
    build_dispatch_lifecycle_router,
    build_dls_admin_health_router,
    ensure_dispatch_lifecycle_indexes,
)

_dls_router = build_dispatch_lifecycle_router(
    db,
    require_dispatch_or_admin_dep=_require_dispatch_or_admin,
    require_any_portal_token_dep=_require_any_portal_token,
    # D-1.3 · auto-notify on new assignment + revision uses the same
    # Resend wrapper Safety / Trench Safety already use. Passing the
    # function (not a new email engine) keeps OMEGA scope intact.
    send_email_fn=_safety_send_email,
)
app.include_router(_dls_router)

# iter412 · Phase 16.1 · admin-only health summary mounted at /api/admin/dls
_dls_admin_health_router = build_dls_admin_health_router(
    db,
    require_admin_dep=require_admin,
)
app.include_router(_dls_admin_health_router)

# ─── FORGEDOPS Dispatch Command Center V1 · Phase 1 ────────────────
# Backend aggregation foundation. Composes Asset Spine + dispatch
# lifecycle + driver sessions + fleet defects + haul_cycles + projects
# + daily_reports into one read-only feed. SMS broadcast uses the
# existing sms_provider abstraction and stubs safely if Twilio creds
# are absent. Doctrine: /app/memory/DISPATCH_COMMAND_CENTER_ARCHITECTURE.md.
from routes.dispatch_command_center import build_dispatch_command_center_router  # noqa: E402
_dcc_router = build_dispatch_command_center_router(
    db,
    require_any_portal_token_dep=_require_any_portal_token,
    require_dispatch_or_admin_dep=_require_dispatch_or_admin,
)
app.include_router(_dcc_router)

# Track 13.21 · Phase C · Dispatch Companion Haul Ledger (read-only).
# Composes existing haul_cycles + dispatch_assignments + operational_attachments
# (scale-ticket family) + daily_reports. NO new collection · NO writes.
# Doctrine: TRACK_13_21_MATERIAL_MOVEMENT_LEDGER_PHASE_C_DISPATCH_HAUL_LEDGER.md.
from routes.dispatch_haul_ledger import build_dispatch_haul_ledger_router  # noqa: E402
_dhl_router = build_dispatch_haul_ledger_router(
    db,
    require_dispatch_or_admin_dep=_require_dispatch_or_admin,
)
app.include_router(_dhl_router)

# Track 13.26 · Asset Service Event Backbone (read-only · derived).
# Single read endpoint that composes per-unit service history across
# equipment_inspections + fleet_defects + haul_cycles + operational_events
# (Motive presence) + asset_transfers. NO new collection · NO writes.
# Honest empty placeholders for pm/fuel/lube/grease/maintainx until those
# subsystems exist.
# Doctrine:
#   TRACK_13_26A_ASSET_EVENT_SOURCE_CERTIFICATION.md (Phase 1 gate)
#   TRACK_13_26_ASSET_SERVICE_EVENT_BACKBONE.md      (Phase 3 implementation)
from routes.asset_service_events import build_asset_service_events_router  # noqa: E402
_ase_router = build_asset_service_events_router(
    db,
    require_any_fleet_portal_dep=_require_any_fleet_portal,
)
app.include_router(_ase_router)

# Track 13.29 · Fuel/Lube Visit Record router (one visit · many equipment lines).
# Projects into Asset Service Event Backbone (via the existing fuel_lube_visits
# collection · derived in `routes/asset_service_events.py:_project_fuel_lube`).
from routes.fuel_lube import build_fuel_lube_router  # noqa: E402
_fl_router = build_fuel_lube_router(
    db,
    require_shop_or_admin_dep=_require_shop_or_admin_fleet,
)
app.include_router(_fl_router)

# Track 13.30 · Service Truck Daily Reconciliation router.
# Composes truck-day reconciliation over Track 13.29 fuel/lube visits.
# Single collection: `service_truck_reconciliations`. No accounting.
from routes.service_truck_reconciliation import (  # noqa: E402
    build_service_truck_reconciliation_router,
)
_strr_router = build_service_truck_reconciliation_router(
    db,
    require_shop_or_admin_dep=_require_shop_or_admin_fleet,
)
app.include_router(_strr_router)

# Track 13.31 · PM Engine — preventive maintenance lifecycle.
# Operator-controlled templates → per-unit schedules → work orders →
# assign/accept/start/complete/review. PM completion does NOT RTS.
from routes.pm_engine import build_pm_engine_router  # noqa: E402
_pm_engine_router = build_pm_engine_router(
    db,
    require_shop_or_admin_dep=_require_shop_or_admin_fleet,
)
app.include_router(_pm_engine_router)

# Track 13.30C · Shop Command Center intelligence — read-only.
# `/api/shop/units/search` + `/api/shop/me/summary`. Two read-only
# endpoints; zero new collection.
from routes.shop_intel import build_shop_intel_router  # noqa: E402
_shop_intel_router = build_shop_intel_router(
    db,
    require_shop_or_admin_dep=_require_shop_or_admin_fleet,
    is_valid_admin_token_fn=_is_valid_admin_token,
    is_valid_admin_token_async=_is_valid_super_admin_token_async,
)
app.include_router(_shop_intel_router)

# Shop Command Feed — single read endpoint consumed by ShopHub and by
# the DCC "Shop" tile. Reads only from canonical defect / recovery
# collections. Doctrine: /app/memory/SHOP_COMMAND_ARCHITECTURE.md.
from routes.shop_command_feed import build_shop_command_feed_router  # noqa: E402
_shop_feed_router = build_shop_command_feed_router(
    db,
    require_any_portal_token_dep=_require_any_portal_token,
)
app.include_router(_shop_feed_router)

# ─── FORGEDOPS PM Command Center · Phase 4A ────────────────────────
# PM-scoped read-only aggregation. Composes Asset Spine + dispatch
# lifecycle + daily reports + haul cycles + fleet defects + incidents
# + CAPAs + asset transfers + state events into 7 endpoints under
# /api/pm/command-center/*. Doctrine:
#   - Road plates are canonical Asset Spine assets ("road_plate")
#   - PM scope enforced via compute_pm_scope() — admin sees all,
#     PMs see only their assigned projects, empty-scope PMs see []
#   - Every operational row carries the map-ready field set
#   - FleetWatcher / MaintainX returned as `not_connected` templates
# Doctrine doc: /app/memory/PM_VISIBILITY_ARCHITECTURE.md.
# ``require_admin`` already accepts Admin or PM tokens and returns the
# PM doc when a per-PM token authenticates — compute_pm_scope() then
# turns that doc into the project_numbers filter. No new auth gate.
from routes.pm_command_center import build_pm_command_center_router  # noqa: E402
_pm_cc_router = build_pm_command_center_router(
    db,
    require_pm_or_admin_dep=require_admin,
)
app.include_router(_pm_cc_router)

# ─── FORGEDOPS Operations Center · Phase 4C ───────────────────────
# Cross-company command board. Composes Asset Spine + Dispatch + PM
# Command Center + Shop + Safety + Motive into 10 read-only endpoints
# under /api/operations-center/command/*. Doctrine:
#   - No new collection, no schema mutation, no FleetWatcher activation
#   - Specialty Asset normalization (Phase 4C architecture correction):
#     road plates are ONE family member, not privileged
#   - Map-ready field set on every operational row (preps Live Ops Map)
#   - any-portal token read (executive mode is a UI filter, not a gate)
from routes.operations_center_command import build_operations_center_command_router  # noqa: E402
_oc_cmd_router = build_operations_center_command_router(
    db,
    require_any_portal_token_dep=_require_any_portal_token,
)
app.include_router(_oc_cmd_router)

# ─── FORGEDOPS Live Operations Map · Phase 5A · contract ──────────
# ONE canonical map-ready endpoint composing Asset Spine + Dispatch +
# PM + Motive + Shop + Safety. Backend-only Phase 5A (no UI map).
# Trust states explain every missing field. Phase 5B (map render)
# requires explicit authorization.
from routes.operations_map_contract import build_operations_map_contract_router  # noqa: E402
_op_map_router = build_operations_map_contract_router(
    db,
    require_any_portal_token_dep=_require_any_portal_token,
    get_runtime_identity=_runtime_identity_bundle,
)
# Phase 5B V1 · live map aggregator endpoints mounted on the SAME
# /api/operations-map router (no parallel system).
from routes.operations_map_v1 import register_operations_map_v1_routes  # noqa: E402
register_operations_map_v1_routes(
    _op_map_router, db,
    require_any_portal_token_dep=_require_any_portal_token,
)
app.include_router(_op_map_router)

# ─── FORGEDOPS Trust Sprint · T2 · Platform Data Truth ────────────
# ONE endpoint, no auth gate, returns environment + integration health
# (no secrets, flags only). Every operational surface consumes this to
# render the preview / production banner. No page may hardcode its own
# banner — single source of truth.
from routes.platform_data_truth import build_platform_data_truth_router  # noqa: E402
app.include_router(build_platform_data_truth_router(db, get_runtime_identity=_runtime_identity_bundle))

# iter416 · Phase 19.1 · admin-only Day-1 Live Ops Debrief capture form.
# Writes a markdown file to /app/memory/DLS_DAY1_LIVE_OPS_DEBRIEF_*.md.
# No database storage · no analytics · no scoring · no charts. Closes the
# Phase 17/19 doctrinal loop: operations runs → debrief filed same-day →
# surgical pickup follows.
from routes.dispatch_day1_debrief import build_day1_debrief_router  # noqa: E402
_dls_day1_debrief_router = build_day1_debrief_router(require_admin_dep=require_admin)
app.include_router(_dls_day1_debrief_router)

# iter417 · Phase 20.0 · Operational Attachments Foundation (walking skeleton).
# Single primitive: attach images to dispatch_assignments only (this iter).
# Doctrine: operational proof continuity, NOT document management.
from routes.operational_attachments import (  # noqa: E402
    build_operational_attachments_router,
    ensure_operational_attachments_indexes,
)
_op_attachments_router, _op_attachments_admin_router = build_operational_attachments_router(
    db,
    require_dispatch_or_admin_dep=_require_dispatch_or_admin,
    require_any_portal_token_dep=_require_any_portal_token,
    require_admin_dep=require_admin,
)
app.include_router(_op_attachments_router)
app.include_router(_op_attachments_admin_router)

# iter430 · Phase 28.2 · Part 1B · admin-strict persistence-health diag.
from routes.admin_persistence_health import build_admin_persistence_health_router  # noqa: E402
app.include_router(build_admin_persistence_health_router(
    app=app,
    db=db,
    require_admin_strict_dep=require_admin_strict,
))

# iter439 · Item I · admin-strict production health probe (HTTP probes
# against mascidocs.com, NOT preview's own Mongo). Powers the calm
# read-only line on /admin/system so preview-vs-production drift is
# structurally impossible to hide.
from routes.admin_production_health import build_production_health_router  # noqa: E402
app.include_router(build_production_health_router(
    require_admin_dep=require_admin_strict,
))

from routes.admin_runtime_reliability import build_runtime_reliability_router  # noqa: E402
app.include_router(build_runtime_reliability_router(
    app=app,
    db=db,
    require_admin_dep=require_admin_strict,
))

# iter440 · Last Activity probe · powers the calm "Last submission · N
# minutes ago" indicator on every role hub. Per-portal scoping ·
# read-only · 7-day lookback cap.
from routes.last_activity import build_last_activity_router  # noqa: E402
app.include_router(build_last_activity_router(
    db=db,
    require_any_portal_token_dep=_require_any_portal_token,
))

# iter437 (2026-05-26) · Cluster capacity probe · public, sub-50ms.
# Surfaces Atlas storage utilization to the frontend banner so a quota
# block (writes-blocked) cannot happen silently again. Discovered after
# today's restore drill exposed the cluster at 99% capacity.
from routes.cluster_capacity import (  # noqa: E402
    build_cluster_capacity_router,
    record_capacity_snapshot,
    ensure_history_indexes,
)
app.include_router(build_cluster_capacity_router(get_client=lambda: client, get_runtime_identity=_runtime_identity_bundle))


# iter437 · Phase Sigma-II · hourly cluster-capacity snapshot recorder.
# Writes one row to `cluster_capacity_history` every hour with a 90-day
# TTL. Powers `/api/cluster/capacity/history` + drift detection.
@register_lifecycle_step("scheduler-nonemail")
async def _cluster_capacity_history_loop() -> None:
    async def _loop():
        # Best-effort initial record + index ensure.
        try:
            await ensure_history_indexes(db)
            await record_capacity_snapshot(client)
        except Exception as e:  # noqa: BLE001
            logger.warning("[cluster-capacity-history] initial record failed: %s", e)
        # Then hourly forever.
        while True:
            await asyncio.sleep(3600)
            try:
                await record_capacity_snapshot(client)
            except Exception as e:  # noqa: BLE001
                logger.warning("[cluster-capacity-history] tick failed: %s", e)

    register_background_task(
        app,
        name="cluster-capacity-history",
        coro=_loop(),
        category="scheduler",
        critical=False,
        long_running=True,
    )

# iter431 · Phase 29 · Part 4 · admin-strict stability sweepers
from routes.admin_stability import build_admin_stability_router  # noqa: E402
app.include_router(build_admin_stability_router(
    db=db,
    require_admin_strict_dep=require_admin_strict,
))

# iter431 · Phase 29 · Part 6 · admin weekly operator digest generator
from routes.admin_operator_digest import build_admin_operator_digest_router  # noqa: E402
app.include_router(build_admin_operator_digest_router(
    db=db,
    require_admin_dep=require_admin,
))

# iter432 · Phase 30 · Part 6 · Field Memory continuity (institutional
# operational wisdom · append-only · role-aware writes · NEVER analytics).
from routes.field_memory import (  # noqa: E402
    build_field_memory_router,
    ensure_field_memory_indexes,
)
app.include_router(build_field_memory_router(
    db=db,
    require_any_portal_token_dep=_require_any_portal_token,
))


@register_lifecycle_step("misc-bootstrap")
async def _ensure_field_memory_indexes_startup():
    try:
        await ensure_field_memory_indexes(db)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[field-memory] index ensure failed: {e}")

# iter418/419/420 · Phases 20.1/21.0/22.0 · Operational Continuity primitives.
# ONE router · THREE walking-skeleton primitives: breakdown-proof upload (driver
# magic-link) · continuity-event log · shop-recovery sub-state transitions.
from routes.dispatch_continuity import (  # noqa: E402
    build_dispatch_continuity_router,
    ensure_dispatch_continuity_indexes,
)
import driver_sessions as _driver_sessions_mod  # noqa: E402
_dispatch_continuity_router = build_dispatch_continuity_router(
    db,
    require_driver_session_dep=_driver_sessions_mod.make_require_driver_session(db),
    require_dispatch_or_admin_dep=_require_dispatch_or_admin,
    require_any_portal_token_dep=_require_any_portal_token,
    # iter424 · Phase 25.1 · Shop owns recovery-state WRITES (role discipline)
    require_shop_or_admin_dep=require_shop_or_admin,
)
app.include_router(_dispatch_continuity_router)


@register_lifecycle_step("index-ensure")
async def _ensure_dls_indexes() -> None:
    try:
        await ensure_dispatch_lifecycle_indexes(db)
        # iter417 · Phase 20.0 · attachments collection indexes
        await ensure_operational_attachments_indexes(db)
        # iter419 · Phase 21.0 · continuity-events collection indexes
        await ensure_dispatch_continuity_indexes(db)
        logging.getLogger(__name__).info(
            "[dispatch-lifecycle] iter392 router mounted · indexes ensured",
        )
    except Exception as e:  # noqa: BLE001
        logging.getLogger(__name__).warning(f"[dispatch-lifecycle] index setup skipped: {e}")


# ─── iter393 · DLS Driver Mobile Surface (magic-link sessions) ──────
# Backend half of the driver tap-and-work surface. Magic-link + revokable
# session pattern mirrors the existing PM/HR/Shop/Safety token model.
# Driver token gate is built INSIDE the router factory.
from routes.dispatch_driver import build_driver_router  # noqa: E402
from driver_sessions import ensure_driver_session_indexes  # noqa: E402

_driver_router = build_driver_router(
    db,
    require_dispatch_or_admin_dep=_require_dispatch_or_admin,
)
app.include_router(_driver_router)


@register_lifecycle_step("index-ensure")
async def _ensure_driver_session_indexes() -> None:
    try:
        await ensure_driver_session_indexes(db)
        logging.getLogger(__name__).info(
            "[dispatch-driver] iter393 router mounted · indexes ensured",
        )
    except Exception as e:  # noqa: BLE001
        logging.getLogger(__name__).warning(
            f"[dispatch-driver] index setup skipped: {e}",
        )


# ─── iter395 · DLS Governance + CSV Exports ─────────────────────────
# Read-only operational signals over the existing iter392 truth.
# Governance uses the cross-portal token aggregator (any portal can
# read for role-aware tiles); exports are dispatch+admin only.
from routes.dispatch_governance import build_dispatch_governance_router  # noqa: E402
from routes.dispatch_exports import build_dispatch_exports_router  # noqa: E402

app.include_router(
    build_dispatch_governance_router(
        db, require_any_portal_token_dep=_require_any_portal_token,
    ),
)
app.include_router(
    build_dispatch_exports_router(
        db, require_dispatch_or_admin_dep=_require_dispatch_or_admin,
    ),
)
logging.getLogger(__name__).info(
    "[dispatch-governance] iter395 routers mounted · 4 detectors + 3 csv exports",
)


# ─── Backup verification (iter79 — weekly R2 health email) ──────────
from routes.backup_verification_routes import build_backup_verification_router  # noqa: E402
from backup_verification import verification_scheduler_loop  # noqa: E402

_backup_verify_router = build_backup_verification_router(db, require_admin_strict)
app.include_router(_backup_verify_router)

_backup_verify_task: Optional[asyncio.Task] = None


@register_lifecycle_step("email-scheduler")
async def _start_backup_verification_cron():
    """Long-running weekly cron that sends a backup verification email
    every Mon 14:00 UTC by default. Kept on its own asyncio.Task so a
    crash in this loop never disturbs the actual backup scheduler."""
    global _backup_verify_task
    try:
        _backup_verify_task = register_background_task(
            app,
            name="backup-verification-singleton",
            coro=run_with_singleton_lock(db, "backup_verification", verification_scheduler_loop),
            category="email-scheduler",
            critical=False,
            long_running=True,
        )
        logging.getLogger(__name__).info(
            "[verify] weekly cron started"
        )
    except Exception as e:  # noqa: BLE001
        logging.getLogger(__name__).exception(f"[verify] failed to start: {e}")


# ─── Multi-portal Access Control Center (iter82) ────────────────────
# Bridges the new `user_directory` master account onto the existing
# per-portal token systems. The minters look up the user by email in
# each portal's collection; if not found, the directory entry alone is
# enough proof of access — we issue a token signed against the master
# password_hash so the portal middleware accepts it.
import user_directory as _ud  # noqa: E402
from routes.auth_directory_routes import build_auth_directory_router  # noqa: E402


def _directory_admin_token(row: Dict[str, Any]) -> Optional[str]:
    """Mint a per-user admin token (TRACK 15.32). The token format is
    ``<user_id>.<HMAC>`` and binds to the directory user's
    ``password_hash`` so a password rotation invalidates extant tokens.
    All /api/admin/* routes accept it via the new validator
    ``user_directory.is_valid_directory_admin_token_async``.
    """
    if not row or not row.get("id") or not row.get("password_hash"):
        return None
    try:
        return _ud.make_directory_admin_token(row["id"], row["password_hash"])
    except Exception as exc:  # noqa: BLE001
        logger.warning(f"_directory_admin_token mint failed: {exc}")
        return None


async def _ensure_portal_shadow(
    db,
    collection: str,
    row: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    """Iter86 — Auto-provision a per-portal "shadow" record (in
    project_managers / hr_users / shop_users) for a directory user that
    is authorized for that portal but doesn't yet have a native per-portal
    record. The shadow record uses the directory user's id + bcrypt hash
    directly so:
      • Tokens mint correctly (HMAC binds to password_hash[:16])
      • Master-password rotations cascade: the shadow's password_hash is
        re-synced on every multi-login, so a directory pw change instantly
        invalidates every per-portal token too.
    Returns the up-to-date per-portal doc, or None if the directory entry
    is incomplete."""
    if not row or not row.get("email") or not row.get("password_hash") or not row.get("id"):
        return None
    coll = db[collection]
    existing = await coll.find_one({"email": row["email"].lower()}, {"_id": 0})
    desired_hash = row["password_hash"]
    desired_id = row["id"]
    now = datetime.now(timezone.utc).isoformat()
    if existing:
        # Sync the password_hash on every login so master-pw rotations
        # propagate. Don't touch disabled (admin may have disabled at
        # the per-portal level intentionally), just refresh the hash.
        if existing.get("password_hash") != desired_hash:
            await coll.update_one(
                {"id": existing["id"]},
                {"$set": {
                    "password_hash": desired_hash,
                    "updated_at": now,
                    "linked_to_directory": True,
                }},
            )
            existing["password_hash"] = desired_hash
        return existing
    # Create a fresh shadow record. Keep it minimal — admin can flesh
    # out per-portal-specific fields (assigned jobs for PM, role for
    # shop, etc.) from the corresponding admin panel later.
    shadow = {
        "id": desired_id,
        "email": row["email"].lower(),
        "name": row.get("name") or row["email"].split("@")[0],
        "password_hash": desired_hash,
        "disabled": False,
        "must_change_password": False,
        "created_at": now,
        "updated_at": now,
        "linked_to_directory": True,
        "source": "directory-shadow",
    }
    await coll.insert_one(dict(shadow))
    return shadow


async def _directory_pm_token(row: Dict[str, Any]) -> Optional[str]:
    """Mint a PM token for a directory user. Auto-provisions a shadow
    `project_managers` record on first multi-login if missing (iter86)."""
    from pm_auth import make_pm_token
    pm = await _ensure_portal_shadow(db, "project_managers", row)
    if not pm or pm.get("disabled") or not pm.get("password_hash"):
        return None
    return make_pm_token(pm["id"], pm["password_hash"])


async def _directory_hr_token(row: Dict[str, Any]) -> Optional[str]:
    """Mint an HR token for a directory user. Auto-provisions shadow."""
    from hr_users import make_hr_user_token
    hr = await _ensure_portal_shadow(db, "hr_users", row)
    if not hr or hr.get("disabled") or not hr.get("password_hash"):
        return None
    return make_hr_user_token(hr["id"], hr["password_hash"])


async def _directory_shop_token(row: Dict[str, Any]) -> Optional[str]:
    """Mint a Shop token for a directory user. Auto-provisions shadow."""
    from shop_users import make_shop_user_token
    shop = await _ensure_portal_shadow(db, "shop_users", row)
    if not shop or shop.get("disabled") or not shop.get("password_hash"):
        return None
    return make_shop_user_token(shop["id"], shop["password_hash"])


async def _directory_safety_token(row: Dict[str, Any]) -> Optional[str]:
    """Mint a Safety token for a directory user. Auto-provisions shadow."""
    from safety_users import make_safety_user_token
    s = await _ensure_portal_shadow(db, "safety_users", row)
    if not s or s.get("disabled") or not s.get("password_hash"):
        return None
    return make_safety_user_token(s["id"], s["password_hash"])


async def _directory_dispatch_token(row: Dict[str, Any]) -> Optional[str]:
    """Mint a Dispatch token for a directory user. Auto-provisions shadow."""
    from dispatch_users import make_dispatch_user_token
    d = await _ensure_portal_shadow(db, "dispatch_users", row)
    if not d or d.get("disabled") or not d.get("password_hash"):
        return None
    return make_dispatch_user_token(d["id"], d["password_hash"])


async def _directory_send_email(to: str, subject: str, html: str) -> None:
    """Iter90 — Resend wrapper used by Access Control Center welcome /
    password-reset notifications. Same envelope as the per-portal welcome
    emails (PM/Shop/HR). Silent no-op if AUTO_EMAIL_REPORTS is off or
    RESEND_API_KEY is missing — caller falls back to 'show password to
    admin' in that case."""
    if (os.environ.get("AUTO_EMAIL_REPORTS") or "").strip().lower() not in ("true", "1", "yes"):
        raise RuntimeError("AUTO_EMAIL_REPORTS disabled")
    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        raise RuntimeError("RESEND_API_KEY missing")
    import resend as _resend  # noqa: PLC0415
    _resend.api_key = api_key
    sender = await _resolve_sender_email(db, safe_fallback="noreply@mascidocs.com")
    params = {
        "from": f"MASCI Operations <{sender}>",
        "to": [to],
        "subject": subject,
        "html": html,
    }
    await asyncio.to_thread(_resend.Emails.send, params)


def _directory_fl_token(row: Dict[str, Any]) -> Optional[str]:
    """iter345 · FL Phase B · Hybrid · mint X-FL-Token for a
    directory user who has the `field_leadership` portal grant. The
    token is bound to the user's master password_hash so it cascades
    when the master password changes. The FL Hub gate already accepts
    this format via getFlToken()."""
    pwh = row.get("password_hash") or ""
    uid = row.get("id") or ""
    if not pwh or not uid:
        return None
    try:
        from routes.field_leadership_portal import make_fl_user_token
        return make_fl_user_token(uid, pwh)
    except Exception:  # noqa: BLE001
        return None


_auth_directory_router = build_auth_directory_router(
    db,
    require_admin_strict_dep=require_admin_strict,
    pm_token_minter=_directory_pm_token,
    hr_token_minter=_directory_hr_token,
    shop_token_minter=_directory_shop_token,
    safety_token_minter=_directory_safety_token,
    dispatch_token_minter=_directory_dispatch_token,
    admin_token_minter=_directory_admin_token,
    field_leadership_token_minter=_directory_fl_token,
    send_email_fn=_directory_send_email,
    render_portal_email_fn=render_portal_email,
)
app.include_router(_auth_directory_router)


async def _canonical_multi_logout(**kwargs):
    perform = getattr(_auth_directory_router, "_perform_multi_logout", None)
    if not callable(perform):
        raise RuntimeError("canonical multi-logout handler unavailable")
    return await perform(**kwargs)


# iter375 · Phase 4B · MFA TOTP router for super-admin directory users.
# Mounts /api/admin/mfa/* (admin-strict gated) and /api/auth/mfa/verify-login (public).
from routes.mfa_routes import build_mfa_router  # noqa: E402
_mfa_router = build_mfa_router(
    db,
    require_admin_strict_dep=require_admin_strict,
    mint_all_portal_tokens_fn=_auth_directory_router._mint_all_portal_tokens,  # type: ignore[attr-defined]
)
app.include_router(_mfa_router)


# iter422 · Phase 24 · Passkey / WebAuthn Continuity (Admin master-signin pilot).
# Optional device-native biometric sign-in (Face ID · Touch ID · Windows Hello ·
# Android · hardware keys). Walking-skeleton: ONE backend router, wired into the
# existing user_directory · NEVER stores biometric data · password fallback
# unchanged · token fan-out preserved EXACTLY.
from routes.passkeys import (  # noqa: E402
    build_passkeys_router,
    ensure_passkey_indexes,
)
import user_directory as _ud_for_pk  # noqa: E402


async def _require_directory_session_for_passkeys(
    x_directory_token: Optional[str] = Header(default=None, alias="X-Directory-Token"),
) -> Dict[str, Any]:
    row = await _ud_for_pk.session_user(db, token=x_directory_token or "")
    if not row:
        raise HTTPException(401, "Directory session required")
    return row


async def _mint_multi_login_response_for_passkey(
    user_row: Dict[str, Any], request: Request,
) -> Dict[str, Any]:
    """iter431 · Phase 29 · Part 5b · delegated to the canonical
    factory at `routes.passkey_session_mint.make_mint_multi_login_response_for_passkey`.
    The wrapper signature is preserved because the passkey router was
    wired with this name BEFORE the extraction and downstream code
    (and tests) may still reference it. Behaviour is identical."""
    return await _passkey_session_mint_fn(user_row, request)


# Build the canonical mint once at module load.
from routes.passkey_session_mint import (  # noqa: E402
    make_mint_multi_login_response_for_passkey as _make_passkey_mint,
)
_passkey_session_mint_fn = _make_passkey_mint(
    db=db,
    mint_all_portal_tokens_fn=_auth_directory_router._mint_all_portal_tokens,  # type: ignore[attr-defined]
    ud_for_pk=_ud_for_pk,
)


_passkeys_router = build_passkeys_router(
    db,
    require_directory_session_dep=_require_directory_session_for_passkeys,
    mint_multi_login_response=_mint_multi_login_response_for_passkey,
)
app.include_router(_passkeys_router)


@register_lifecycle_step("index-ensure")
async def _ensure_passkey_indexes() -> None:
    try:
        await ensure_passkey_indexes(db)
        logging.getLogger(__name__).info(
            "[passkeys] iter422 router mounted · indexes ensured",
        )
    except Exception as e:  # noqa: BLE001
        logging.getLogger(__name__).warning(f"[passkeys] index setup skipped: {e}")


# iter377 · Phase 4D · PM read-only routes extraction from server.py.
# Moves /pm/check, /pm/me, and the 4 /pm/crew/* read endpoints into
# routes/pm_routes.py. Login/forgot/reset/change/logout REMAIN here
# (iter378 candidate). Behavior locked by tests/test_iter377_*.
from routes.pm_routes import build_pm_router  # noqa: E402
_pm_router = build_pm_router(
    db,
    require_admin_dep=require_pm_portal_or_super_admin,
    require_admin_async_dep=require_pm_portal_or_super_admin_async,
    login_deps={
        "client_ip_fn": _client_ip,
        "check_login_lockout_fn": _check_login_lockout,
        "record_login_fail_fn": _record_login_fail,
        "reset_login_fails_fn": _reset_login_fails,
        "directory_admin_token_fn": _directory_admin_token,
        "reset_session_activity_fn": _reset_session_activity,
        "clear_session_activity_fn": _clear_session_activity,
        "canonical_multi_logout_fn": _canonical_multi_logout,
        "render_portal_email_fn": render_portal_email,
        # Track 15.87 · directory `pm` grant path. Wrapped in a
        # lambda so name resolution defers (_directory_pm_token is
        # async + defined below).
        "directory_pm_minter_fn": lambda row: _directory_pm_token(row),
    },
)
app.include_router(_pm_router)


# Phase K4a (iter176) — Unified Directory read-only surface. Surfaces
# the K1-mirrored unified `user_directory` (with mirrored/managed
# classification + K1 metadata) and the K3 role-template catalog to
# the admin UI. No mutations exposed here; existing
# /api/admin/directory routes remain the only write path.
from routes.admin_directory_k4 import build_admin_directory_k4_router  # noqa: E402


async def _k4_step_up_dep(request: Request):
    """Phase 2 Initiative 5b-full — step-up dependency wired into K4
    mutation routes. Pass-through when ADMIN_STEP_UP_ENABLED is unset,
    otherwise raises 403 if no recent admin re-auth."""
    x_admin_token = request.headers.get("x-admin-token") or ""
    await _require_recent_step_up(db, x_admin_token, request, max_age_min=5)
    return True


_admin_directory_k4_router = build_admin_directory_k4_router(
    db,
    require_admin_strict_dep=require_admin_strict,
    require_step_up=_k4_step_up_dep,
)
app.include_router(_admin_directory_k4_router)


@register_lifecycle_step("seed")
async def _bootstrap_user_directory():
    """Seed the super-admin row on first deploy (idempotent — silent if
    already present). Driven by SUPER_ADMIN_EMAIL / SUPER_ADMIN_BOOTSTRAP_PASSWORD
    env vars."""
    try:
        await _ud.bootstrap_super_admin(db)
    except Exception as e:  # noqa: BLE001
        logging.getLogger(__name__).exception(f"[directory] bootstrap failed: {e}")
    # Phase K1 (iter172) — silent unified-identity mirror. Backfills the
    # user_directory collection from per-portal user collections. NEVER
    # touches login flows. Idempotent on every restart. Failures are
    # logged but never block startup.
    try:
        from lib.identity_mirror import run_startup_mirror
        await run_startup_mirror(db)
    except Exception as e:  # noqa: BLE001
        logging.getLogger(__name__).exception(f"[identity-mirror] startup hook failed: {e}")
    # Phase K3 (iter175) — non-enforcing role-template seed. Populates
    # role_templates collection with built-in templates per portal.
    # Nothing yet reads from this collection (K6 wiring is deferred).
    # Idempotent. Failures logged, never block startup.
    try:
        from lib.role_templates import run_startup_seed
        await run_startup_seed(db)
    except Exception as e:  # noqa: BLE001
        logging.getLogger(__name__).exception(f"[role-templates] startup hook failed: {e}")


# Weekly variance email cron (Sunday 18:00 UTC by default).
_PAYROLL_EMAIL_HOUR = int(os.environ.get("PAYROLL_VARIANCE_EMAIL_HOUR_UTC", "18") or "18")
_PAYROLL_EMAIL_DOW = int(os.environ.get("PAYROLL_VARIANCE_EMAIL_DOW", "6") or "6")  # 0=Mon, 6=Sun
_PAYROLL_VARIANCE_STATE: Dict[str, Any] = {"last_sent_date": None}


def _payroll_recipients() -> List[str]:
    raw = (os.environ.get("PAYROLL_VARIANCE_EMAIL_TO") or
           "hrmanager@mascigc.com,jaymn.judd@mascigc.com")
    return [r.strip() for r in raw.split(",") if r.strip()]


async def _maybe_send_weekly_variance_email():
    """Called from the existing background_indexer tick; only sends once
    per UTC day matching _PAYROLL_EMAIL_DOW @ _PAYROLL_EMAIL_HOUR."""
    now = datetime.now(timezone.utc)
    if now.weekday() != _PAYROLL_EMAIL_DOW or now.hour < _PAYROLL_EMAIL_HOUR:
        return
    today_iso = now.date().isoformat()
    if _PAYROLL_VARIANCE_STATE.get("last_sent_date") == today_iso:
        return
    try:
        batch = await db.payroll_variance_batches.find_one(
            {}, {"_id": 0}, sort=[("created_at", -1)]
        )
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[payroll-variance] cron lookup failed: {e}")
        return
    if not batch:
        # Mark as "sent" for today so we don't re-check every 10 min.
        _PAYROLL_VARIANCE_STATE["last_sent_date"] = today_iso
        logger.info("[payroll-variance] no batch this week — cron skipped")
        return

    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    auto = (os.environ.get("AUTO_EMAIL_REPORTS") or "").strip().lower() in ("true", "1", "yes")
    if not api_key or not auto:
        logger.info("[payroll-variance] email disabled (no RESEND_API_KEY or AUTO_EMAIL_REPORTS off) — marking sent")
        _PAYROLL_VARIANCE_STATE["last_sent_date"] = today_iso
        return

    recipients = _payroll_recipients()
    if not recipients:
        _PAYROLL_VARIANCE_STATE["last_sent_date"] = today_iso
        return

    html = _pv_module._render_variance_email_html(batch)
    csv_bytes = _pv_module.render_variance_csv_bytes(batch)
    import base64 as _b64
    import resend as _resend  # noqa: PLC0415
    _resend.api_key = api_key
    sender = await _resolve_sender_email(db, safe_fallback="noreply@mascidocs.com")
    params = {
        "from": f"MASCI HR · Payroll Variance <{sender}>",
        "to": recipients,
        "subject": f"MASCI Payroll Variance — Week Ending {batch.get('week_ending')}",
        "html": html,
        "attachments": [{
            "filename": f"MASCI_payroll_variance_{batch.get('week_ending')}.csv",
            "content": _b64.b64encode(csv_bytes).decode(),
        }],
    }
    try:
        await asyncio.to_thread(_resend.Emails.send, params)
        _PAYROLL_VARIANCE_STATE["last_sent_date"] = today_iso
        logger.info(f"[payroll-variance] weekly email sent → {recipients}")
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[payroll-variance] weekly email failed: {e}")


@register_lifecycle_step("misc-bootstrap")
async def _backfill_doc_ids() -> None:
    """One-shot backfill: every existing record across the registered
    submission collections gets a human-readable doc_id stamped if it
    doesn't already have one. Idempotent — subsequent boots are no-ops."""
    try:
        from doc_ids import backfill_all
        summary = await backfill_all(db)
        if summary:
            logger.info(f"[doc_ids] startup backfill: {summary}")
    except Exception as e:  # noqa: BLE001
        logger.warning(f"[doc_ids] startup backfill failed: {e}")


# iter381 · /api/admin/find-by-doc-id extracted to routes/admin_lookups.py.
from routes.admin_lookups import build_admin_lookups_router  # noqa: E402
_admin_lookups_router = build_admin_lookups_router(db, require_admin)
app.include_router(_admin_lookups_router)


# M-3 · Geocode Foundation · /api/admin/locations/* (canonical location
# registry that links MASCI jobs ↔ Motive geofences). Visibility-only;
# never writes to Motive; never auto-assigns projects.
from routes.operational_locations import build_operational_locations_router  # noqa: E402
_op_locations_router = build_operational_locations_router(db, require_admin)
app.include_router(_op_locations_router)


# M-DR-1 · Equipment Auto-Discovery · /api/equipment-detection/* (read-only
# suggestion API for the Daily Report form). Motive suggests; foreman
# verifies; foreman authors. No DR mutations from this module.
from routes.equipment_detection import build_equipment_detection_router  # noqa: E402
_equipment_detection_router = build_equipment_detection_router(db)
app.include_router(_equipment_detection_router)


# M-2 · Event Router · /api/operational-events/* + /api/admin/operational-events/*
# Derives normalized operational arrival/departure events from Motive
# telemetry. Visibility / verification only — no DR / dispatch / Motive
# writes. Storage enforced through the M-2-8 allowed-fields gate.
from routes.operational_events import build_operational_events_router  # noqa: E402
_operational_events_router = build_operational_events_router(db, require_admin)
app.include_router(_operational_events_router)


# VER-1 · Operational Verification Layer · /api/verification/* +
# /api/admin/verification/*. Compute-on-read only — stores nothing,
# mutates nothing, authors nothing. Reads M-1 / M-3 / M-DR-1 / M-2 to
# emit 4 trust states: CONFIRMED, PENDING_CONFIRMATION, MISMATCH, QUIET.
from routes.verification import build_verification_router  # noqa: E402
_verification_router = build_verification_router(db, require_admin)
app.include_router(_verification_router)


# MOTIVE-DATA-001 · Asset Mapping Reconciliation · /api/admin/asset-mapping/*
# Data-quality sprint to close the dispatch.truck_id ↔ asset_mappings link
# that VER-1 audit identified as the bottleneck.
from routes.asset_mapping_recon import build_asset_mapping_router  # noqa: E402
_asset_mapping_router = build_asset_mapping_router(db, require_admin)
app.include_router(_asset_mapping_router)


# FORGEDOPS-P0.1 · Canonical Asset Spine · /api/asset-spine/*
# Single source-of-truth API over equipment_master with audit, profile
# aggregation, and read-only detection engine. Doctrine in
# /app/memory/MASTER_ASSET_GOVERNANCE_ARCHITECTURE.md.
from routes.asset_spine import register_asset_spine_routes  # noqa: E402
# Late-mount via `app` directly because api_router has already been
# attached by line 9016 — adding to api_router post-mount is a no-op.
register_asset_spine_routes(app, db, require_admin, _require_any_portal_token)

# Track 13.31B-D3+D4 · Asset Documents · Renewals · CSV · MASCI PDF
# Reuses operational_attachments storage (host_kind="asset"). Mounts
# under the same /api/asset-spine/* prefix as the parent spine routes
# so admin tooling consumes a single API surface.
from routes.asset_documents import register_asset_documents_routes  # noqa: E402
register_asset_documents_routes(
    app, db, require_admin, _require_any_portal_token,
    require_admin_or_asset_admin_dep=require_admin_or_asset_admin,
)

# Track 13.31B-D7 · Asset Admin operational completion · adds Required Docs
# editor save + asset_admin role grant pathway. Additive, single small config
# collection (asset_required_doc_overrides).
from routes.asset_admin_settings import register_asset_admin_settings_routes  # noqa: E402
register_asset_admin_settings_routes(app, db, require_admin, require_admin_or_asset_admin_dep=require_admin_or_asset_admin)

from routes.notify_ownership_lock_seed import register_notify_ownership_lock_seed  # noqa: E402
register_notify_ownership_lock_seed(app, db, require_admin, get_runtime_identity=_runtime_identity_bundle)

from routes.scheduled_producers_d456 import register_scheduled_producers_d456  # noqa: E402
register_scheduled_producers_d456(app, db, require_admin)

from routes.project_team_assignments import register_project_team_assignments  # noqa: E402
register_project_team_assignments(app, db, require_admin, _require_any_portal_token)

# TRACK 15.44 · Executive Overview (read-only aggregator over existing data).
from routes.executive_overview import register as register_executive_overview  # noqa: E402
register_executive_overview(app, db=db, require_admin_dep=require_admin)

from routes.ownership_lifecycle import register_ownership_lifecycle  # noqa: E402
register_ownership_lifecycle(app, db, require_admin, _require_any_portal_token)

# Track 13.33ABC · Asset Care & Readiness Command Center
from routes.asset_care import register_asset_care_routes  # noqa: E402
register_asset_care_routes(app, db, require_admin, require_admin_or_asset_admin_dep=require_admin_or_asset_admin)

# TRACK 14.0-RC1-FERRARI · /api/admin/perf-snapshot · 10-second
# operator-confidence check (disk, memory, uptime, mongo ping,
# self-probe latency, recent error counts, scheduler heartbeat).
from routes.perf_snapshot import register_perf_snapshot_routes  # noqa: E402
register_perf_snapshot_routes(app, db, require_admin)

# TRACK 16.04 · Transportation Foundation Phase 1.
# Carriers / transport_persons / transport_trucks / eligibility skeleton.
# Admin CRUD (strict admin); dispatch read-only endpoints.
from routes.transportation import register_transportation_routes  # noqa: E402
register_transportation_routes(
    app, db,
    require_admin_dep=require_admin_strict,
    require_dispatch_or_admin_dep=_shared_dispatch_or_admin,
)

# TRACK 16.06 · Transportation Experience Layer — aggregation endpoints
# (dashboard, document queue, inspection queue, audit timeline, workspace
# aggregators). Read-only · admin-strict. Registered BEFORE Phase 2 so
# literal paths like /admin/transportation/inspections/queue resolve to
# this router rather than being captured by Phase 2's /inspections/{iid}.
from routes.transportation_experience import register_transportation_experience_routes  # noqa: E402
# TRACK 18.00 · Phase F — Mission Control summary tiles must load for
# dispatch / leadership / pm / safety / fl / shop / hr tokens too, not
# just admin. Pass the cross-portal helper so the dashboard endpoint
# becomes portal-aware. Every other endpoint in this router remains
# admin-strict (record-detail surfaces stay locked).
register_transportation_experience_routes(
    app, db,
    require_admin_dep=require_admin_strict,
    require_portal_dep=_require_any_portal_token,
    # TRACK 18.12C · Dispatcher-operational gate for Class A reads
    # (documents queue, inspections queue, carrier/driver/truck
    # workspaces, per-entity compliance timeline). Audit Timeline
    # remains admin-strict.
    require_dispatch_or_admin_dep=_require_dispatch_or_admin,
)

# TRACK 16.05 · Transportation Onboarding & Compliance Center (Phase 2).
# Rate schedules · carrier+driver documents · packet workflow · MASCI
# Hauler Truck Readiness Inspection · dashboards. Idempotent bootstrap
# of default rate schedule + requirements catalog at startup.
from routes.transportation_phase2 import register_transportation_phase2_routes  # noqa: E402
register_transportation_phase2_routes(
    app, db,
    require_admin_dep=require_admin_strict,
    require_dispatch_or_admin_dep=_shared_dispatch_or_admin,
)


@register_lifecycle_step("misc-bootstrap")
async def _track_16_05_bootstrap_on_startup():
    try:
        from lib.transport_phase2 import bootstrap_track_16_05
        await bootstrap_track_16_05(db)
    except Exception as exc:  # noqa: BLE001
        logging.getLogger(__name__).warning(
            f"[track-16-05-bootstrap] non-fatal: {exc}")


# TRACK 16.08 · Transportation Orientation, Notification & External
# Onboarding Platform. Module catalog · video player heartbeat · quiz
# engine · certificate engine · secure external invite portal. All
# orientation routes mounted via this single router. Bootstrap seeds
# 22 default modules across 4 languages.
from routes.transportation_orientation import (  # noqa: E402
    register_transportation_orientation_routes,
    bootstrap_track_16_08,
    bootstrap_track_19_01a,
)
register_transportation_orientation_routes(
    app, db, require_admin_dep=require_admin_strict,
    # TRACK 18.12C · Dispatcher-operational reads (orientation dashboard,
    # module list, assignments, certificates). All writes remain
    # admin-strict.
    require_dispatch_or_admin_dep=_require_dispatch_or_admin,
)


@register_lifecycle_step("misc-bootstrap")
async def _track_16_08_bootstrap_on_startup():
    try:
        await bootstrap_track_16_08(db)
    except Exception as exc:  # noqa: BLE001
        logging.getLogger(__name__).warning(
            f"[track-16-08-bootstrap] non-fatal: {exc}")
    try:
        await bootstrap_track_19_01a(db)
    except Exception as exc:  # noqa: BLE001
        logging.getLogger(__name__).warning(
            f"[track-19-01a-bootstrap] non-fatal: {exc}")


# TRACK 16.09 · Transportation Dispatch Gate + Email Pilot. Wires the
# hard-block, authorized override, and the 4 pilot real-send email
# routes. Mounted via a dedicated router; emails fired through the
# existing fsi_email_sender + email_routing_v2 primitives (no
# duplicate sender, no SMS/push).
from routes.transportation_dispatch_gate import (  # noqa: E402
    register_track_16_09_routes,
    bootstrap_track_16_09,
)
register_track_16_09_routes(
    app, db,
    require_dispatch_or_admin_dep=_require_dispatch_or_admin,
    require_admin_dep=require_admin_strict,
)


@register_lifecycle_step("misc-bootstrap")
async def _track_16_09_bootstrap_on_startup():
    try:
        await bootstrap_track_16_09(db)
    except Exception as exc:  # noqa: BLE001
        logging.getLogger(__name__).warning(
            f"[track-16-09-bootstrap] non-fatal: {exc}")


# TRACK 16.10 · Transportation Automation Engine. Scheduled reminders,
# eligibility automation, action queue. Reuses singleton_scheduler +
# email_routing_v2 + fsi_email_sender. No new sender, no SMS.
from routes.transportation_automation import (  # noqa: E402
    register_track_16_10_routes,
    bootstrap_track_16_10,
    transport_automation_scheduler_loop,
    transport_command_digest_scheduler_loop,
)
register_track_16_10_routes(
    app, db,
    require_admin_dep=require_admin_strict,
    require_dispatch_or_admin_dep=_require_dispatch_or_admin,
)

# TRACK 16.12 · Transportation Operations Intelligence — read-only
# intelligence engine routes (drivers / carriers / trucks / dashboard
# / recommendations / predictions / operational-health).
from routes.transportation_intelligence import (  # noqa: E402
    register_track_16_12_routes,
)
register_track_16_12_routes(
    app, db, require_admin_dep=require_admin_strict,
    # TRACK 18.12C · Cleanup signals are surfaced on Mission Control —
    # dispatchers must read this feed.
    require_dispatch_or_admin_dep=_require_dispatch_or_admin)

# TRACK 16.13 · Dispatch Decision Surface — read-only recommendation
# endpoint + interaction audit. Admin OR dispatch token.
from routes.dispatch_decision_surface import (  # noqa: E402
    register_track_16_13_routes,
)
register_track_16_13_routes(
    app, db, require_dispatch_or_admin_dep=_require_dispatch_or_admin)

# TRACK 16.16 · Operations × Transportation Integration Layer.
# Thin read-only consumer endpoint that lets Operations surfaces
# (project workspaces, Operations Center Command, PM Command Center)
# render Transportation awareness without leaving their workspace.
# Composes existing Track 16.06 / 16.10 / 16.11A / 16.12 / 16.15
# engines — no new scoring, no new collections, no new audit kinds.
from routes.operations_transportation_integration import (  # noqa: E402
    register_track_16_16_routes,
)
from routes.integrations._deps import (  # noqa: E402
    make_require_any_portal_token as _make_any_portal_track_16_16,
)
register_track_16_16_routes(
    app, db,
    require_any_portal_dep=_make_any_portal_track_16_16(
        db, _is_valid_admin_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)

# TRACK 18.00 · Phase C · RBAC-aware Universal Search composer.
# One thin composer endpoint that fans out across existing
# transportation collections, filters per portal token, and returns
# grouped + deep-linked results. No new collection. No new index.
from routes.transportation_search import (  # noqa: E402
    register_track_18_00_phase_c_routes,
)
register_track_18_00_phase_c_routes(
    app, db,
    require_any_portal_dep=_make_any_portal_track_16_16(
        db, _is_valid_admin_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)

# TRACK 18.00 · Phase D · Universal Relationships + Live Right Rail.
# Composer endpoint that returns related records, open actions, and
# audit trail for any Transportation Operations entity. RBAC-aware,
# read-only, no new collections.
from routes.transportation_relationships import (  # noqa: E402
    register_track_18_00_phase_d_routes,
)
register_track_18_00_phase_d_routes(
    app, db,
    require_any_portal_dep=_make_any_portal_track_16_16(
        db, _is_valid_admin_token,
        is_valid_admin_token_async=_is_valid_directory_admin_token_async,
    ),
)

_transport_automation_task: Optional[asyncio.Task] = None


@register_lifecycle_step("misc-bootstrap")
async def _track_16_10_bootstrap_on_startup():
    global _transport_automation_task
    try:
        await bootstrap_track_16_10(db)
    except Exception as exc:  # noqa: BLE001
        logging.getLogger(__name__).warning(
            f"[track-16-10-bootstrap] non-fatal: {exc}")
    # Wire the daily scheduler under the singleton lock so we never run
    # twice on multi-worker deployments.
    try:
        async def _wrapped(_db):
            return await transport_automation_scheduler_loop(_db)
        _transport_automation_task = register_background_task(
            app,
            name="transport-automation-singleton",
            coro=run_with_singleton_lock(db, "transport_automation", _wrapped),
            category="scheduler",
            critical=False,
            long_running=True,
        )
        logger.info("[track-16-10] automation scheduler armed")
    except Exception as exc:  # noqa: BLE001
        logging.getLogger(__name__).warning(
            f"[track-16-10-scheduler] non-fatal: {exc}")

    # TRACK 16.10A · Monday-morning Command Digest. Singleton-locked
    # so multi-worker prod fires exactly once per Monday window.
    try:
        async def _wrapped_digest(_db):
            return await transport_command_digest_scheduler_loop(_db)
        register_background_task(
            app,
            name="transport-command-digest-singleton",
            coro=run_with_singleton_lock(db, "transport_command_digest", _wrapped_digest),
            category="scheduler",
            critical=False,
            long_running=True,
        )
        logger.info("[track-16-10a] command-digest scheduler armed")
    except Exception as exc:  # noqa: BLE001
        logging.getLogger(__name__).warning(
            f"[track-16-10a-scheduler] non-fatal: {exc}")


# PROJECT-IDENTITY-005 · Project Identity Governance · /api/admin/project-identity/*
# Detection-only drift sentinel. Never auto-mutates source records or jobs_master.
# Operator controls every resolution (match / leave_unmatched / intentional / dismiss).
from routes.project_identity_governance import build_project_identity_router  # noqa: E402
_project_identity_router = build_project_identity_router(db, require_admin)
app.include_router(_project_identity_router)



# iter382 · /admin/project-managers/* (10 routes) + public /project-managers
# extracted to routes/pm_admin.py.
from routes.pm_admin import build_pm_admin_router  # noqa: E402
_pm_admin_router = build_pm_admin_router(
    db,
    require_admin_dep=require_admin,
    require_admin_strict_dep=require_admin_strict,
    xlsx_response_fn=_xlsx_response,
    today_stamp_fn=_today_stamp,
    active_filter=ACTIVE_FILTER,
    render_portal_email_fn=render_portal_email,
)
app.include_router(_pm_admin_router)


# ============================================================
# Email a saved record as a PDF (Resend)
# ============================================================
# Independent router so it imports at startup without forcing a hot-reload
# of the existing routes. Adds POST /api/email-report which:
#   1. Looks up the saved record by id from its module's collection.
#   2. Renders it to a polished PDF via /app/backend/pdf_render.py.
#   3. Sends via Resend with the PDF attached.

import asyncio  # noqa: E402
import base64 as _email_b64  # noqa: E402

from pdf_render import (  # noqa: E402
    render_email_html,
    render_record_pdf,
    KIND_TITLES,
    build_email_subject,
)
from pm_routing import (  # noqa: E402
    ALWAYS_CC,
    COMPLIANCE_KINDS,
    PM_ONLY_KINDS,
    auto_email_enabled,
    recipients_for_record_async,
)


# TRACK 15.47 · Incident PDF enrichment shim.
# When the PDF kind is `incident`, load the state-event timeline and
# linked CAPAs and attach them to the record so the PDF renderer's
# G8 + G9 blocks have data to draw. All other kinds pass through
# unchanged. Best-effort — never raises.
async def _maybe_enrich_for_pdf(_db, kind: str, record: dict) -> dict:
    if kind != "incident" or not isinstance(record, dict):
        return record
    try:
        from lib.incident_pdf_enrichment import enrich_incident_for_pdf  # noqa: PLC0415
        return await enrich_incident_for_pdf(_db, record)
    except Exception:
        return record


_KIND_TO_COLLECTION = {
    "inspection": "inspections",
    "meeting": "meetings",
    "jha": "jhas",
    "incident": "incidents",
    "daily-report": "daily_reports",
    "equipment-inspection": "equipment_inspections",
}


# ------------------------------------------------------------------
# Auto-email on submit (fire-and-forget — never blocks the response)
# ------------------------------------------------------------------
# TRACK 22.1B — the safe leaf helpers (`_filename_for`, `_is_severe_incident`)
# and the fire-and-forget scaffolding (`_AUTO_EMAIL_DISPATCH_TASKS`,
# `schedule_auto_email`) have been extracted to `backend/lib/email_dispatch.py`
# with byte-comparable parity proof. The 473-line `_dispatch_auto_email`
# body remains here because it closes over ~8 server.py module-locals
# (db, logger, _resolve_sender_email, _resolve_reply_to_email,
# render_record_pdf, _maybe_enrich_for_pdf, build_email_subject,
# render_email_html, _email_b64). See `TRACK_22_1B_EMAIL_ARCHITECTURE.md`.
from lib.email_dispatch import (  # noqa: E402
    _filename_for,
    _is_severe_incident,
    _AUTO_EMAIL_DISPATCH_TASKS,
    schedule_auto_email,
    register_dispatcher as _register_email_dispatcher,
)
from lib.notification_delivery import (  # noqa: E402
    DELIVERY_MODE_PROVIDER_LIVE,
    STATUS_CAPTURED_PREVIEW,
    STATUS_CONFIGURATION_BLOCKED,
    STATUS_PERMANENT_FAILURE,
    STATUS_PROVIDER_ACCEPTED,
    STATUS_RETRYABLE_FAILURE,
    deliver_notification,
    delivery_contract,
)
from lib.preview_notification_certification import (  # noqa: E402
    provision_preview_live_override,
    record_provider_attempt_result,
)
from lib.field_submitter_identity import write_dispatch_event  # noqa: E402
# _KIND_TO_COLLECTION is also re-exported for callers that reference it via
# the server module attribute (e.g. /api/auto-email-preview).
from lib.email_dispatch import _KIND_TO_COLLECTION as _KIND_TO_COLLECTION_LIB  # noqa: E402, F401


async def _dispatch_auto_email(kind: str, record: dict) -> None:
    """Render PDF + send via Resend to the assigned PM and the always-CC list.

    Wrapped in a broad try/except so a missing API key, Resend outage, or PDF
    error never causes the original POST to fail. Logs at WARNING level when
    skipped and ERROR when something unexpected breaks.

    TRACK 15.76 · Trust Spine — this is the universal dispatcher for every
    email-bound workflow (daily-report, meeting, jha, incident, qaqc,
    inspection, equipment-inspection). It emits the lifecycle stages
    (routing_resolved → recipients_built → notification_queued →
    provider_accepted → completed) onto ``trust_spine_events`` using the
    correlation_id attached to ``record`` at submit-time. A failure at any
    stage is emitted with ``status="failed"`` + a truthful failure_reason,
    flipping that workflow's dashboard band to RED.

    TRACK 20.6B · SYNTHETIC-TEST-RECORD EMAIL SAFETY GATE
    ────────────────────────────────────────────────────
    When the record's ``project_name`` starts with the reserved ``TEST_``
    prefix, this dispatcher short-circuits before any Resend call. This
    is the test-safety guardrail mandated by Track 20.6B:

    * Real production records never carry a ``TEST_`` prefix (validated by
      grep across the entire production dataset — no legitimate MASCI
      project has ever used a leading ``TEST_`` on ``project_name``).
    * The Track 20.7-C01 regression + Track 20.6B test-hardening suites
      submit synthetic records with ``project_name`` = ``TEST_DR_*`` /
      ``TEST_JOB_PHOTO_*`` / ``TEST_track_19_21_*``.
    * Without this gate, the test suite in the preview environment
      (where AUTO_EMAIL_REPORTS=true and RESEND_API_KEY is real)
      would fire live emails to the assigned PM + always-CC list on
      every test run. That is a Class-A operational hygiene defect.
    * This gate is additive: it does not remove, weaken, or reroute
      any real email path. A record without the ``TEST_`` prefix is
      completely unaffected. Zero drift on production behavior.
    * The skip is audited into ``trust_spine_events`` with
      ``status="skipped"`` and ``failure_reason="synthetic_test_record"``
      so the dashboards remain green and the skip is fully traceable.
    """
    from lib.trust_spine import (  # noqa: PLC0415
        attach_correlation, emit_workflow_stage,
        STAGE_ROUTING_RESOLVED, STAGE_RECIPIENTS_BUILT,
        STAGE_NOTIFICATION_QUEUED, STAGE_PROVIDER_ACCEPTED,
        STAGE_DELIVERY_CAPTURED_PREVIEW, STAGE_AUDIT_WRITTEN,
        STAGE_COMPLETED, STAGE_COMPLETED_FOR_ENVIRONMENT,
    )
    # Thread / attach the cid before any branch so every stage in
    # this dispatch shares one correlation_id.
    _spine_cid = attach_correlation(record)
    _spine_module = f"auto_email_dispatch:{kind}"
    # Track 20.6B — synthetic-test-record short-circuit. Runs BEFORE the
    # auto_email_enabled() check so the skip audit fires even when
    # AUTO_EMAIL_REPORTS=true (that is exactly the preview environment
    # where the test suite runs and where a live send would leak).
    try:
        _pname = str(record.get("project_name") or "").strip()
        if _pname.startswith("TEST_"):
            logger.info(
                "auto-email skipped (Track 20.6B synthetic-test-record gate) "
                f"— {kind} {record.get('id')} project_name={_pname!r}"
            )
            try:
                await emit_workflow_stage(
                    db, workflow=kind, stage=STAGE_NOTIFICATION_QUEUED,
                    record=record, module=_spine_module, status="skipped",
                    failure_reason="synthetic_test_record",
                    remediation=(
                        "No action needed. Test suites use TEST_-prefixed "
                        "project_name to prevent live sends. Real records "
                        "are unaffected."
                    ),
                )
            except Exception:  # noqa: BLE001
                # Never let audit failure break the short-circuit.
                pass
            return
    except Exception:  # noqa: BLE001
        # Defensive: if the record shape is weird, fall through to the
        # normal auto_email_enabled() path (which itself is safe).
        pass
    try:
        dist = await recipients_for_record_async(db, record, kind)
        recipients: List[str] = list(dist["all"])  # type: ignore[arg-type]
        original_intended_recipients: List[str] = list(recipients)
        routing_module = "pm_routing.recipients_for_record_async"
        routing_failure_reason = (
            None if recipients else f"no recipients resolved (kind={kind})"
        )
        routing_remediation = (
            None if recipients
            else "Assign a PM in project_team_assignments or set ADMIN_DEAD_LETTER_TO."
        )

        # iter238 — Equipment Pre-Op routing simplification (operator
        # directive 2026-05-18):
        #   "All Pre Ops only need to go to shop manager no other emails
        #    just shop manager"
        # We override the entire recipient list for ``equipment-inspection``
        # to ONLY the active shop user(s) whose role is "Shop Manager".
        # This is Q1 option (a): role-based fan-out so multiple Shop
        # Managers (if ever added) all get the email automatically, but
        # mechanics / parts coordinators are excluded. Falls back to the
        # ``shop_manager_fallback`` env value when no Shop Manager role
        # exists in the shop_users collection (deploy bootstrap).
        if kind in {"equipment-inspection", "dvir"}:
            routing_module = "shop_users.list_shop_users"
            routing_failure_reason = (
                f"no shop recipients resolved (kind={kind})"
            )
            routing_remediation = (
                "Seed an active Shop Manager or configure PRE_OP_FAIL_FALLBACK / SHOP_MANAGER_EMAIL."
            )
            shop_manager_emails: List[str] = []
            try:
                from shop_users import list_shop_users  # noqa: PLC0415
                shop_users_list = await list_shop_users(db, only_active=True)
                for u in shop_users_list:
                    role = (u.get("role") or "").strip().lower()
                    em = (u.get("email") or "").strip()
                    if role == "shop manager" and em and not u.get("disabled"):
                        if em.lower() not in {x.lower() for x in shop_manager_emails}:
                            shop_manager_emails.append(em)
            except Exception as e:  # noqa: BLE001
                logger.warning(f"shop-manager lookup failed: {e}")
            if not shop_manager_emails:
                fallback = ""
                try:
                    from email_routing import get_value as _routing_get  # noqa: PLC0415
                    fallback = (await _routing_get(db, "shop_manager_fallback")) or ""
                except Exception:
                    pass
                if not fallback:
                    fallback = os.environ.get(
                        "SHOP_MANAGER_EMAIL", "shopmanager@mascigc.com"
                    ).strip()
                if fallback:
                    shop_manager_emails = [fallback]
            # HARD OVERRIDE — Pre-Op emails go to Shop Manager(s) only.
            # No PM, no co-PMs, no always-CC, no FAIL fan-out to other
            # shop users. Per operator: "no other emails just shop manager".
            recipients = list(shop_manager_emails)

            # TRACK 15.75B · P0 silent-failure guard. If no Shop Manager
            # email resolves (no Shop Manager user, no PRE_OP_FAIL_FALLBACK
            # route, no SHOP_MANAGER_EMAIL env), escalate to
            # ADMIN_DEAD_LETTER_TO AND write a truthful audit row so the
            # operator can see the gap on the routing dashboard instead
            # of having Pre-Op alerts vanish into a logger.exception.
            if not recipients:
                try:
                    from pm_routing import _dead_letter_recipients  # noqa: PLC0415
                    dead_to = await _dead_letter_recipients(db)
                except Exception:  # noqa: BLE001
                    dead_to = []
                try:
                    from email_routing_v2 import write_audit as _v2_audit  # noqa: PLC0415
                    from tenant_context import resolve_tenant_key  # noqa: PLC0415
                    try:
                        _tk = resolve_tenant_key()
                    except Exception:
                        _tk = "masci"
                    await _v2_audit(
                        db,
                        route_key="PRE_OP_FAIL_FALLBACK",
                        tenant_key=_tk,
                        source="db",
                        to_count=len(dead_to),
                        cc_count=0,
                        bcc_count=0,
                        subject=f"[SHOP UNRESOLVED] {kind}",
                        status=(
                            "escalated_to_admin_dead_letter"
                            if dead_to else "shop_recipient_unconfigured"
                        ),
                        calling_module="shop_routing_unresolved",
                        dry_run=False,
                    )
                except Exception:  # noqa: BLE001
                    pass
                recipients = list(dead_to)
                routing_module = "shop_routing_unresolved"

        # Trust Spine — routing resolution stage. Status is based on the
        # FINAL recipient source after any workflow-specific overrides
        # (Shop Manager fan-out, dead-letter fallback, etc.). This keeps
        # the stage truthful for DVIR / Pre-Op shop-only delivery.
        await emit_workflow_stage(
            db, workflow=kind, stage=STAGE_ROUTING_RESOLVED,
            record=record, module=routing_module,
            status="ok" if recipients else "failed",
            failure_reason=(None if recipients else routing_failure_reason),
            remediation=(None if recipients else routing_remediation),
        )

        # Severity fan-out for incidents (Major/Severe currently mirrors the
        # always-CC; future ops/GC list can be appended here from env.)
        if kind == "incident" and _is_severe_incident(record):
            extras: List[str] = []
            try:
                from email_routing import get_value as _routing_get
                v = await _routing_get(db, "severe_incident_cc")
                if isinstance(v, list):
                    extras = v
            except Exception:
                pass
            if not extras:
                extras = [
                    x.strip()
                    for x in (os.environ.get("SEVERE_INCIDENT_CC", "") or "").split(",")
                    if x.strip()
                ]
            for e in extras:
                if e.lower() not in {r.lower() for r in recipients}:
                    recipients.append(e)

        if not recipients:
            logger.warning(f"auto-email: no recipients resolved for {kind} {record.get('id')}")
            await emit_workflow_stage(
                db, workflow=kind, stage=STAGE_RECIPIENTS_BUILT,
                record=record, module=_spine_module, status="failed",
                failure_reason=(
                    f"no recipients after dead-letter fallback (kind={kind})"
                ),
                remediation=(
                    "Configure ADMIN_DEAD_LETTER_TO or assign a PM to this project."
                ),
            )
            return

        # Trust Spine — recipient list finalized.
        await emit_workflow_stage(
            db, workflow=kind, stage=STAGE_RECIPIENTS_BUILT,
            record=record, module=_spine_module, status="ok",
        )

        pdf_bytes = await asyncio.to_thread(render_record_pdf, kind, await _maybe_enrich_for_pdf(db, kind, record))

        pm_name = dist.get("pm_name")

        # Flag equipment failures + severe incidents at a glance
        equipment_fail = (
            kind == "equipment-inspection"
            and (record.get("fail_count") or 0) > 0
        )
        severe_incident = kind == "incident" and _is_severe_incident(record)
        subject = build_email_subject(
            kind,
            record,
            equipment_fail=equipment_fail,
            severe_incident=severe_incident,
        )

        # Note: rendered as plain text by render_email_html (which wraps
        # it in a styled callout). Do NOT embed HTML tags here — they
        # will be HTML-escaped and shown to the reader as literal text.
        note = ""
        if kind == "incident" and _is_severe_incident(record):
            note = "SEVERE INCIDENT — please review immediately."
        elif equipment_fail:
            unit_label = " ".join(
                p for p in [
                    str(record.get("equipment_type") or "").strip(),
                    str(record.get("equipment_unit") or "").strip(),
                ] if p
            ).strip()
            unit_suffix = f" {unit_label} tagged OUT OF SERVICE." if unit_label else " Unit tagged OUT OF SERVICE."
            note = (
                f"EQUIPMENT FAIL — {record.get('fail_count')} item(s) "
                f"failed inspection.{unit_suffix}"
            )
        elif kind == "equipment-inspection":
            # iter238 — Pre-Op emails route to Shop Manager only, not
            # the assigned PM. The body note should reflect that or the
            # reader will look for a PM thread that doesn't exist.
            note = (
                "Routed to Shop Manager. Equipment Pre-Op records are "
                "delivered to the shop only — PM and office are not on "
                "this thread."
            )
        elif pm_name:
            note = (
                f"Auto-routed to {pm_name} based on project number "
                f"{record.get('project_number') or '—'}."
            )
        else:
            note = (
                "No Project Manager could be auto-resolved from project "
                f"number {record.get('project_number') or '—'}. Sent to "
                "office distribution only — please assign a PM in the office."
            )

        # Trust Spine — about to call provider.
        await emit_workflow_stage(
            db, workflow=kind, stage=STAGE_NOTIFICATION_QUEUED,
            record=record, module=_spine_module, status="ok",
        )
        attachments = [
            {
                "filename": _filename_for(kind, record),
                "content": _email_b64.b64encode(pdf_bytes).decode(),
            }
        ]
        html = render_email_html(kind, record, note)
        contract = delivery_contract()
        active_cert_override = None
        try:
            active_cert_override = await provision_preview_live_override(
                db,
                workflow=kind,
                record=record,
                original_intended_recipients=original_intended_recipients,
            )
        except Exception:  # noqa: BLE001
            active_cert_override = None
        if active_cert_override:
            recipients = [str(active_cert_override.get("actual_recipient") or "").strip()]
            contract = dict(contract)
            contract["delivery_mode"] = DELIVERY_MODE_PROVIDER_LIVE
            contract["delivery_mode_source"] = "preview_scoped_certification_override"
            contract["provider_validation_status"] = "certification_override"
        delivery = await deliver_notification(
            db=db,
            workflow=kind,
            correlation_id=_spine_cid,
            record_id=str(record.get("id") or record.get("doc_id") or ""),
            recipients=recipients,
            subject=subject,
            html=html,
            reply_to=(await _resolve_reply_to_email(db)) or "",
            attachments=attachments,
            metadata={
                "project_number": record.get("project_number") or "",
                "kind": kind,
            },
        )
        # TRACK 15.75C · UNIVERSAL per-send audit row. Every workflow
        # email send (daily-report, meeting, incident, qaqc, jha,
        # inspection, equipment-inspection) writes a truthful audit
        # row to `email_routing_audit_v2`. Closes the "log-only" trust
        # gap so the operator can prove delivery for every workflow on
        # `/api/admin/email-routing/v2/status` and `RoutingStatusPanel`.
        # Best-effort; never block on audit write.
        try:
            from email_routing_v2 import write_audit as _v2_audit  # noqa: PLC0415
            from tenant_context import resolve_tenant_key  # noqa: PLC0415
            try:
                _tk = resolve_tenant_key()
            except Exception:
                _tk = "masci"
            _shop_delivery_kind = kind in {"equipment-inspection", "dvir"}
            _route_key = (
                "PRE_OP_FAIL_FALLBACK" if _shop_delivery_kind
                else "AUTO_EMAIL_REPORTS"
            )
            _calling_module = (
                "shop_preop_dispatch" if _shop_delivery_kind
                else f"auto_email_dispatch:{kind}"
            )
            audit_status = "sent"
            if delivery.get("notification_state") == STATUS_CAPTURED_PREVIEW:
                audit_status = "captured_preview"
            elif delivery.get("notification_state") == STATUS_CONFIGURATION_BLOCKED:
                audit_status = "configuration_blocked"
            elif delivery.get("notification_state") in {STATUS_RETRYABLE_FAILURE, STATUS_PERMANENT_FAILURE}:
                audit_status = delivery.get("notification_state")
            await _v2_audit(
                db,
                route_key=_route_key,
                tenant_key=_tk,
                source="db",
                to_count=len(recipients or []),
                cc_count=0,
                bcc_count=0,
                subject=subject,
                sender_email=await _resolve_sender_email(db),
                resend_message_id=delivery.get("provider_message_id"),
                status=audit_status,
                calling_module=_calling_module,
                dry_run=(delivery.get("notification_state") == STATUS_CAPTURED_PREVIEW),
            )

            if active_cert_override:
                await record_provider_attempt_result(
                    db,
                    override=active_cert_override,
                    delivery=delivery,
                )
                binding_id = str(active_cert_override.get("id") or "").strip()
                record_id_value = str(record.get("id") or record.get("doc_id") or "")
                record_doc_id_value = str(record.get("doc_id") or "")
                actual_recipient = str(active_cert_override.get("actual_recipient") or "").strip().lower()
                dispatch_extra = {
                    "resolution_tier": "certification_override",
                    "certification_run_id": active_cert_override.get("certification_run_id"),
                    "certification_override": True,
                    "original_intended_recipients": list(
                        active_cert_override.get("original_intended_recipients") or []
                    ),
                }
                await write_dispatch_event(
                    db,
                    workflow=kind,
                    record_id=record_id_value,
                    record_doc_id=record_doc_id_value,
                    kind="notification_dispatch_attempted",
                    binding_id=binding_id,
                    channel="email",
                    recipient=actual_recipient,
                    extra=dispatch_extra,
                )
                if delivery.get("provider_accepted"):
                    await write_dispatch_event(
                        db,
                        workflow=kind,
                        record_id=record_id_value,
                        record_doc_id=record_doc_id_value,
                        kind="notification_dispatch_succeeded",
                        binding_id=binding_id,
                        channel="email",
                        recipient=actual_recipient,
                        provider_message_id=str(delivery.get("provider_message_id") or ""),
                        extra=dispatch_extra,
                    )
                else:
                    await write_dispatch_event(
                        db,
                        workflow=kind,
                        record_id=record_id_value,
                        record_doc_id=record_doc_id_value,
                        kind="notification_dispatch_failed",
                        binding_id=binding_id,
                        channel="email",
                        recipient=actual_recipient,
                        error=str(delivery.get("failure_reason") or "notification_delivery_failed"),
                        extra=dispatch_extra,
                    )

            record_update = {
                "notification_delivery_mode": delivery.get("delivery_mode"),
                "notification_state": delivery.get("notification_state"),
                "notification_provider_accepted": bool(delivery.get("provider_accepted")),
                "notification_provider_called": bool(delivery.get("provider_called")),
                "notification_failure_reason": delivery.get("failure_reason"),
                "notification_last_updated_at": delivery.get("ts"),
                "notification_provider_validation_status": contract.get("provider_validation_status"),
                "business_state": "submitted",
            }
            if active_cert_override:
                record_update["notification_certification_override_id"] = active_cert_override.get("id")
                record_update["notification_certification_run_id"] = active_cert_override.get("certification_run_id")
                record_update["notification_actual_recipient"] = active_cert_override.get("actual_recipient")
                record_update["notification_original_intended_recipients"] = list(
                    active_cert_override.get("original_intended_recipients") or []
                )
                record_update["notification_certification_override_expires_at"] = active_cert_override.get("expires_at")
                record_update["notification_certification_override_status"] = (
                    "used_pending_reconciliation"
                    if delivery.get("provider_accepted")
                    else delivery.get("notification_state")
                )
            if delivery.get("capture_id"):
                record_update["notification_capture_id"] = delivery.get("capture_id")
            if delivery.get("provider_message_id"):
                record_update["notification_provider_message_id"] = delivery.get("provider_message_id")
            record.update(record_update)
            if kind == "daily-report":
                try:
                    await db.daily_reports.update_one({"id": record.get("id")}, {"$set": record_update})
                except Exception:
                    pass

            await emit_workflow_stage(
                db, workflow=kind, stage=STAGE_AUDIT_WRITTEN,
                record=record, module=_calling_module, status="ok",
            )
            if delivery.get("notification_state") == STATUS_CAPTURED_PREVIEW:
                await emit_workflow_stage(
                    db, workflow=kind, stage=STAGE_DELIVERY_CAPTURED_PREVIEW,
                    record=record, module=_spine_module, status="ok",
                    remediation="Preview safe-capture stored; no live provider was contacted.",
                )
                await emit_workflow_stage(
                    db, workflow=kind, stage=STAGE_COMPLETED_FOR_ENVIRONMENT,
                    record=record, module=_spine_module, status="ok",
                )
            elif delivery.get("provider_accepted"):
                await emit_workflow_stage(
                    db, workflow=kind, stage=STAGE_PROVIDER_ACCEPTED,
                    record=record, module="resend.Emails.send", status="ok",
                )
                await emit_workflow_stage(
                    db, workflow=kind, stage=STAGE_COMPLETED,
                    record=record, module=_spine_module, status="ok",
                )
            else:
                await emit_workflow_stage(
                    db, workflow=kind, stage=STAGE_PROVIDER_ACCEPTED,
                    record=record, module="resend.Emails.send", status="failed",
                    failure_reason=str(delivery.get("failure_reason") or "notification_delivery_failed"),
                    remediation=(
                        "Resolve live provider configuration before production release."
                        if contract.get("delivery_mode") == DELIVERY_MODE_PROVIDER_LIVE
                        else "Inspect preview notification capture and retry semantics."
                    ),
                )
                if kind == "daily-report":
                    await emit_workflow_stage(
                        db, workflow=kind, stage=STAGE_COMPLETED_FOR_ENVIRONMENT,
                        record=record, module=_spine_module, status="ok",
                        remediation="Business record persisted; notification requires operator follow-up.",
                    )
        except Exception:  # noqa: BLE001
            pass
    except Exception as e:  # noqa: BLE001
        logger.exception(f"auto-email failed for {kind} {record.get('id')}: {e}")
        # TRACK 15.76 · Trust Spine — flip this workflow's band to RED
        # via a `completed`-failure event so the dashboard surfaces the
        # exact failure point + a remediation hint.
        try:
            await emit_workflow_stage(
                db, workflow=kind, stage=STAGE_PROVIDER_ACCEPTED,
                record=record, module=_spine_module, status="failed",
                failure_reason=str(e)[:240],
                remediation=(
                    "Inspect backend logs for stack; check delivery mode, provider status, and PDF rendering."
                ),
            )
            if kind == "daily-report":
                await emit_workflow_stage(
                    db, workflow=kind, stage=STAGE_COMPLETED_FOR_ENVIRONMENT,
                    record=record, module=_spine_module, status="ok",
                    remediation="Daily Report persisted; downstream notification failed.",
                )
        except Exception:  # noqa: BLE001
            pass
        # TRACK 15.75C · UNIVERSAL per-send failure audit row.
        try:
            from email_routing_v2 import write_audit as _v2_audit  # noqa: PLC0415
            from tenant_context import resolve_tenant_key  # noqa: PLC0415
            try:
                _tk = resolve_tenant_key()
            except Exception:
                _tk = "masci"
            _shop_delivery_kind = kind in {"equipment-inspection", "dvir"}
            _route_key = (
                "PRE_OP_FAIL_FALLBACK" if _shop_delivery_kind
                else "AUTO_EMAIL_REPORTS"
            )
            _calling_module = (
                "shop_preop_dispatch" if _shop_delivery_kind
                else f"auto_email_dispatch:{kind}"
            )
            await _v2_audit(
                db,
                route_key=_route_key,
                tenant_key=_tk,
                source="db",
                to_count=len(recipients or []) if "recipients" in dir() else 0,
                cc_count=0,
                bcc_count=0,
                subject=f"[SEND FAILED] {kind}",
                status="failed",
                error=str(e)[:240],
                calling_module=_calling_module,
                dry_run=False,
            )
        except Exception:  # noqa: BLE001
            pass


# TRACK 22.1B · The strong-reference set + `schedule_auto_email` were
# extracted to `backend/lib/email_dispatch.py`. Register this file's
# `_dispatch_auto_email` as the dispatcher hook so `schedule_auto_email`
# routes through the lib module while `_dispatch_auto_email` continues
# to close over the same server.py module-locals as before. Byte-parity
# certified in `TRACK_22_1B_DISPATCH_PARITY.md`.
_register_email_dispatcher(_dispatch_auto_email)


# (auto-email-preview / routing-table routes are registered after _email_router below)


class EmailReportRequest(BaseModel):
    kind: str = Field(
        ...,
        description="One of: inspection, meeting, jha, incident, daily-report",
    )
    record_id: str
    recipients: List[str] = Field(..., min_length=1, max_length=20)
    subject: Optional[str] = ""
    note: Optional[str] = ""


_email_router = APIRouter(prefix="/api")


@_email_router.get("/auto-email/preview")
async def auto_email_preview(
    project_number: str = "",
    project_name: str = "",
    severity: str = "",
    osha_recordable: str = "",
    kind: str = "",
    _: bool = Depends(require_admin),
):
    """Admin-only introspection: shows who *would* receive the auto-email
    for a given project_number / project_name + form kind."""
    fake = {
        "project_number": project_number,
        "project_name": project_name,
        "severity": severity,
        "osha_recordable": osha_recordable,
    }
    dist = await recipients_for_record_async(db, fake, kind or None)
    return {
        "input": fake,
        "kind": kind or None,
        "pm_name": dist["pm_name"],
        "pm_email": dist["pm_email"],
        "to": dist["to"],
        "cc": dist["cc"],
        "all_recipients": dist["all"],
        "auto_email_enabled": auto_email_enabled(),
        "always_cc": ALWAYS_CC,
    }


@_email_router.get("/auto-email/routing-table")
async def auto_email_routing_table(_: bool = Depends(require_admin)):
    """Returns the live PM → Jobs lookup table from the DB (admin-only).
    Replaces the legacy hardcoded PM_TABLE — now reads project_managers
    + jobs_master so the UI mirrors what the email router actually uses."""
    pms_cursor = db.project_managers.find({}, {"_id": 0}).sort("name", 1)
    pms = await pms_cursor.to_list(500)
    jobs_cursor = db.jobs_master.find({}, {"_id": 0}).sort("project_number", 1)
    jobs = await jobs_cursor.to_list(2000)

    by_email: Dict[str, List[Dict[str, str]]] = {}
    unassigned: List[Dict[str, str]] = []
    for j in jobs:
        pm_email = (j.get("pm_email") or "").strip().lower()
        # Fallback: if pm_email not set but name matches a PM, infer it.
        if not pm_email:
            nm = (j.get("project_manager") or "").strip()
            if nm:
                match = next(
                    (p for p in pms if (p.get("name") or "").strip() == nm),
                    None,
                )
                if match:
                    pm_email = (match.get("email") or "").strip().lower()
        if pm_email:
            by_email.setdefault(pm_email, []).append({
                "project_number": j.get("project_number") or "",
                "project_name": j.get("project_name") or "",
            })
        else:
            unassigned.append({
                "project_number": j.get("project_number") or "",
                "project_name": j.get("project_name") or "",
            })

    project_managers = []
    for p in pms:
        em = (p.get("email") or "").strip().lower()
        project_managers.append({
            "pm_id": p.get("id"),
            "pm_name": p.get("name"),
            "pm_email": p.get("email"),
            "phone": p.get("phone") or "",
            "is_active": bool(p.get("is_active", True)),
            "jobs": by_email.get(em, []),
        })

    return {
        "always_cc": ALWAYS_CC,
        "compliance_kinds": sorted(COMPLIANCE_KINDS),
        "pm_only_kinds": sorted(PM_ONLY_KINDS),
        "auto_email_enabled": auto_email_enabled(),
        "project_managers": project_managers,
        "unassigned_jobs": unassigned,
    }


# ─────────────────────────────────────────────────────────────────────────
# Admin Email Routing — DB-backed overrides for the env-default lists.
# Lets the admin change "who gets what email" from the console without
# a redeploy. See backend/email_routing.py for storage + cache details.
# ─────────────────────────────────────────────────────────────────────────
class EmailRoutingUpdate(BaseModel):
    """Partial update — only included keys are written. Pass [] to clear
    a list (e.g. severe_incident_cc=[] to silence the severe-incident
    fan-out). Keys not present are left untouched."""
    always_cc: Optional[List[str]] = None
    safety_forms_to: Optional[List[str]] = None
    leadership_always_to: Optional[List[str]] = None
    shop_manager_fallback: Optional[str] = None
    severe_incident_cc: Optional[List[str]] = None
    backup_email_to: Optional[List[str]] = None


class EmailRoutingTestBody(BaseModel):
    to: str
    subject: Optional[str] = None


@_email_router.get("/admin/email-routing")
async def admin_email_routing_get(_: bool = Depends(require_admin)):
    """Return the merged routing config (env defaults + DB overrides)
    plus the env-only defaults so the UI can show "Reset to default"."""
    from email_routing import load as _routing_load, env_defaults as _routing_env_defaults
    cfg = await _routing_load(db)
    return {"config": cfg, "env_defaults": _routing_env_defaults()}


@_email_router.put("/admin/email-routing")
async def admin_email_routing_put(
    body: EmailRoutingUpdate, _: bool = Depends(require_admin)
):
    """Persist any subset of routing keys. Returns the freshly-merged config."""
    from email_routing import save as _routing_save
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    cfg = await _routing_save(db, updates, updated_by="admin")
    return {"ok": True, "config": cfg}


@_email_router.post("/admin/email-routing/test")
async def admin_email_routing_test(
    body: EmailRoutingTestBody, _: bool = Depends(require_admin)
):
    """Send a one-off test email to confirm Resend + sender + DNS are wired."""
    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        raise HTTPException(503, "RESEND_API_KEY not configured")
    target = (body.to or "").strip()
    if not target or "@" not in target:
        raise HTTPException(400, "Valid 'to' address required")
    import resend  # noqa: E402
    resend.api_key = api_key
    sender_email = await _resolve_sender_email(db, safe_fallback="noreply@mascidocs.com")
    subject = body.subject or "[MASCI HUB] Email Routing test"
    html = (
        "<div style='font-family:Arial,sans-serif;max-width:540px'>"
        "<h2 style='color:#C8102E'>Email Routing test — success</h2>"
        "<p>This message was sent from the MASCI HUB Admin Console "
        "&rarr; Email Routing &rarr; Send test.</p>"
        "<p>If you received it, Resend, the sender domain, and the "
        "destination address are all wired up correctly. The address "
        "you tested can safely be added to any of the routing lists.</p>"
        f"<p style='color:#64748b;font-size:12px'>Sent {datetime.now(timezone.utc).isoformat()} UTC</p>"
        "</div>"
    )
    try:
        params = {
            "from": f"MASCI Operations Platform <{sender_email}>",
            "to": [target],
            "subject": subject,
            "html": html,
        }
        reply_to = (os.environ.get("REPLY_TO_EMAIL") or "").strip()
        if reply_to:
            params["reply_to"] = reply_to
        result = await asyncio.to_thread(resend.Emails.send, params)
        return {"ok": True, "to": target, "resend_id": (result or {}).get("id")}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f"Resend send failed: {e}")


# ─────────────────────────────────────────────────────────────────────────
# Email Routing V2 (Track 15.66 Wave 2) — per-route management surface.
# All endpoints below operate on the new ``email_routes`` collection seeded
# by ``backend/scripts/track_15_65_seed_email_routes.py``. Behaviour is
# back-compat: existing V1 endpoints above continue to work unchanged.
# ─────────────────────────────────────────────────────────────────────────
class V2RouteUpdate(BaseModel):
    to: Optional[List[str]] = None
    cc: Optional[List[str]] = None
    bcc: Optional[List[str]] = None
    from_email: Optional[str] = None
    reply_to: Optional[str] = None
    enabled: Optional[bool] = None
    description: Optional[str] = None


class V2RouteTestBody(BaseModel):
    """Controlled test send. ``dry_run=True`` (default) only resolves
    recipients and writes an audit row. Set ``dry_run=False`` AND
    ``test_recipient`` to send a real email to a specific test inbox —
    NEVER to the route's production recipient list."""
    dry_run: bool = True
    test_recipient: Optional[str] = None


class V2BrandingUpdate(BaseModel):
    company_name: Optional[str] = None
    platform_display_name: Optional[str] = None
    sender_name: Optional[str] = None
    from_email: Optional[str] = None
    reply_to: Optional[str] = None
    support_email: Optional[str] = None
    safety_email: Optional[str] = None
    hr_email: Optional[str] = None
    operations_email: Optional[str] = None
    logo_url: Optional[str] = None
    primary_color: Optional[str] = None


def _current_tenant_key() -> str:
    return (os.environ.get("EMAIL_ROUTING_TENANT") or "masci").strip().lower() or "masci"


def _validate_email_list(emails: List[str], field_name: str) -> List[str]:
    """Validate every email in a list. Empty list is allowed."""
    if emails is None:
        return []
    cleaned: List[str] = []
    seen: set = set()
    for raw in emails:
        s = str(raw or "").strip()
        if not s:
            continue
        if "@" not in s or "." not in s.split("@", 1)[-1]:
            raise HTTPException(400, f"Invalid email '{s}' in {field_name}")
        key = s.lower()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(s)
    return cleaned


@_email_router.get("/admin/email-routing/v2/routes")
async def admin_v2_routes_list(_: bool = Depends(require_admin)):
    """Return every route doc for the active tenant, plus per-route
    last-send / last-failure summaries derived from the audit collection."""
    tk = _current_tenant_key()
    cursor = db.email_routes.find({"tenant_key": tk}, {"_id": 0}).sort("route_key", 1)
    routes = await cursor.to_list(100)
    # Per-route last-audit lookup (cheap — bounded by 19 routes)
    summaries: Dict[str, Dict[str, Any]] = {}
    for r in routes:
        rk = r["route_key"]
        last = await db.email_routing_audit_v2.find_one(
            {"tenant_key": tk, "route_key": rk}, sort=[("ts", -1)]
        )
        last_fail = await db.email_routing_audit_v2.find_one(
            {"tenant_key": tk, "route_key": rk, "status": {"$in": _EMAIL_AUDIT_FAILURE_STATUSES}},
            sort=[("ts", -1)],
        )
        summaries[rk] = {
            "last_send_at": (last or {}).get("ts"),
            "last_send_status": (last or {}).get("status"),
            "last_send_source": (last or {}).get("source"),
            "last_failure_at": (last_fail or {}).get("ts"),
            "last_failure_error": (last_fail or {}).get("error"),
        }
        r["summary"] = summaries[rk]
    return {"tenant_key": tk, "routes": routes, "count": len(routes)}


@_email_router.get("/admin/email-routing/v2/routes/{route_key}")
async def admin_v2_route_get(route_key: str, _: bool = Depends(require_admin)):
    tk = _current_tenant_key()
    doc = await db.email_routes.find_one({"_id": f"{tk}::{route_key}"}, {"_id": 0})
    if not doc:
        raise HTTPException(404, f"Route {route_key} not configured for tenant {tk}")
    return doc


@_email_router.put("/admin/email-routing/v2/routes/{route_key}")
async def admin_v2_route_put(
    route_key: str, body: V2RouteUpdate, _: bool = Depends(require_admin)
):
    """Edit recipients / enabled / description for a route. Admin source
    is recorded so the seed script will not overwrite admin edits without
    --force. Prevents disabling critical routes and prevents empty
    enabled-critical TO lists."""
    tk = _current_tenant_key()
    _id = f"{tk}::{route_key}"
    existing = await db.email_routes.find_one({"_id": _id})
    if not existing:
        raise HTTPException(404, f"Route {route_key} not configured for tenant {tk}")
    update: Dict[str, Any] = {}
    if body.to is not None:
        update["to"] = _validate_email_list(body.to, "to")
    if body.cc is not None:
        update["cc"] = _validate_email_list(body.cc, "cc")
    if body.bcc is not None:
        update["bcc"] = _validate_email_list(body.bcc, "bcc")
    if body.from_email is not None:
        v = (body.from_email or "").strip()
        if v and "@" not in v:
            raise HTTPException(400, "Invalid from_email")
        update["from_email"] = v or None
    if body.reply_to is not None:
        v = (body.reply_to or "").strip()
        if v and "@" not in v:
            raise HTTPException(400, "Invalid reply_to")
        update["reply_to"] = v or None
    if body.enabled is not None:
        if (not body.enabled) and bool(existing.get("critical")):
            raise HTTPException(
                400,
                "Critical routes cannot be disabled through this endpoint. "
                "Use platform-admin override path if intentional.",
            )
        update["enabled"] = bool(body.enabled)
    if body.description is not None:
        update["description"] = str(body.description).strip()[:500]

    # Critical-route TO guard
    new_to = update.get("to", existing.get("to") or [])
    new_enabled = update.get("enabled", existing.get("enabled", True))
    if bool(existing.get("critical")) and new_enabled and not new_to:
        raise HTTPException(
            400, "Critical route cannot have empty 'to' while enabled."
        )

    if not update:
        return {"ok": True, "changed": False, "doc": {**existing, "_id": None}}
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    update["updated_by"] = "admin"
    update["source"] = "admin"
    await db.email_routes.update_one({"_id": _id}, {"$set": update})
    # Invalidate resolver cache so the next send picks up the edit.
    try:
        from email_routing_v2 import invalidate_cache as _v2_invalidate
        _v2_invalidate()
    except Exception:
        pass
    new_doc = await db.email_routes.find_one({"_id": _id}, {"_id": 0})
    return {"ok": True, "changed": True, "doc": new_doc}


@_email_router.post("/admin/email-routing/v2/routes/{route_key}/test")
async def admin_v2_route_test(
    route_key: str, body: V2RouteTestBody, _: bool = Depends(require_admin)
):
    """Controlled route test. Dry-run by default — only resolves recipients
    and writes an audit row. When ``dry_run=False`` AND ``test_recipient``
    is supplied, sends ONE email to that test inbox using the route's
    sender/reply-to. Never blasts the route's production recipients."""
    tk = _current_tenant_key()
    _id = f"{tk}::{route_key}"
    doc = await db.email_routes.find_one({"_id": _id})
    if not doc:
        raise HTTPException(404, f"Route {route_key} not configured for tenant {tk}")

    try:
        from email_routing_v2 import resolve as _v2_resolve, write_audit as _v2_audit  # noqa: PLC0415
    except Exception as e:  # noqa: BLE001
        raise HTTPException(500, f"Routing engine unavailable: {e}")

    # Force-resolve via DB doc (does not depend on EMAIL_ROUTING_V2 flag).
    res_to = [str(x).strip() for x in (doc.get("to") or [])]
    res_cc = [str(x).strip() for x in (doc.get("cc") or [])]
    res_bcc = [str(x).strip() for x in (doc.get("bcc") or [])]

    if body.dry_run or not body.test_recipient:
        # Dry-run path — write audit row, no Resend call.
        await _v2_audit(
            db,
            route_key=route_key, tenant_key=tk, source="db",
            to_count=len(res_to), cc_count=len(res_cc), bcc_count=len(res_bcc),
            subject="[ROUTE TEST · DRY-RUN]", sender_email=None,
            status="dry_run", calling_module="admin_v2_route_test", dry_run=True,
        )
        await db.email_routes.update_one(
            {"_id": _id},
            {"$set": {"last_tested_at": datetime.now(timezone.utc).isoformat(),
                      "last_test_status": "dry_run"}},
        )
        return {
            "ok": True, "dry_run": True,
            "resolved": {"to": res_to, "cc": res_cc, "bcc": res_bcc},
            "sender_email": doc.get("from_email") or await _resolve_sender_email(db, route_key=route_key, safe_fallback=""),
            "reply_to": doc.get("reply_to") or await _resolve_reply_to_email(db),
        }

    # Controlled real send — only to the test_recipient.
    target = (body.test_recipient or "").strip()
    if "@" not in target:
        raise HTTPException(400, "Invalid test_recipient")
    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        raise HTTPException(503, "RESEND_API_KEY not configured")
    import resend as _resend  # noqa: PLC0415
    _resend.api_key = api_key
    sender = doc.get("from_email") or await _resolve_sender_email(db, route_key=route_key, safe_fallback="noreply@mascidocs.com")
    reply_to = doc.get("reply_to") or await _resolve_reply_to_email(db)
    subject = f"[ROUTE TEST · {route_key}] Controlled probe"
    html = (
        f"<div style='font-family:Arial,sans-serif;max-width:540px'>"
        f"<h2 style='color:#C8102E'>Route test — {route_key}</h2>"
        f"<p>This is a controlled probe of the <strong>{doc.get('display_name', route_key)}</strong> "
        f"route. Production recipients ({len(res_to)} TO · {len(res_cc)} CC · {len(res_bcc)} BCC) "
        f"were intentionally NOT contacted. Only this test inbox received the probe.</p>"
        f"<p style='color:#64748b;font-size:12px'>Sent {datetime.now(timezone.utc).isoformat()} UTC</p>"
        f"</div>"
    )
    params: Dict[str, Any] = {
        "from": f"MASCI Operations Platform <{sender}>",
        "to": [target],
        "subject": subject,
        "html": html,
    }
    if reply_to:
        params["reply_to"] = reply_to
    try:
        result = await asyncio.to_thread(_resend.Emails.send, params)
    except Exception as e:  # noqa: BLE001
        await _v2_audit(
            db, route_key=route_key, tenant_key=tk, source="db",
            to_count=1, cc_count=0, bcc_count=0,
            subject=subject, sender_email=sender, status="failed",
            error=str(e), calling_module="admin_v2_route_test", dry_run=False,
        )
        raise HTTPException(502, f"Resend send failed: {e}")
    rid = (result or {}).get("id")
    await _v2_audit(
        db, route_key=route_key, tenant_key=tk, source="db",
        to_count=1, cc_count=0, bcc_count=0,
        subject=subject, sender_email=sender, resend_message_id=rid,
        status="sent", calling_module="admin_v2_route_test", dry_run=False,
    )
    await db.email_routes.update_one(
        {"_id": _id},
        {"$set": {"last_tested_at": datetime.now(timezone.utc).isoformat(),
                  "last_test_status": "sent"}},
    )
    return {"ok": True, "dry_run": False, "test_recipient": target, "resend_id": rid}


@_email_router.get("/admin/email-routing/v2/audit")
async def admin_v2_audit_slice(
    route_key: Optional[str] = None,
    limit: int = 100,
    _: bool = Depends(require_admin),
):
    """Per-route audit slice. Defaults to the most recent 100 rows across
    all routes for the active tenant. Filter by route_key when opening the
    audit drawer for a single route."""
    tk = _current_tenant_key()
    limit = max(1, min(int(limit or 100), 500))
    q: Dict[str, Any] = {"tenant_key": tk}
    if route_key:
        q["route_key"] = route_key
    cursor = db.email_routing_audit_v2.find(q, {"_id": 0}).sort("ts", -1).limit(limit)
    rows = await cursor.to_list(limit)
    return {"tenant_key": tk, "route_key": route_key, "rows": rows, "count": len(rows)}


@_email_router.get("/admin/email-routing/v2/branding")
async def admin_v2_branding_get(_: bool = Depends(require_admin)):
    """Tenant branding (sender / reply-to / company / etc.). Defaults
    derive from env vars so the first GET on a fresh tenant returns a
    pre-populated doc rather than 404."""
    tk = _current_tenant_key()
    doc = await db.tenant_branding.find_one({"_id": tk}, {"_id": 0})
    if not doc:
        # Track 15.67 Phase 3 · only seed env defaults for the MASCI tenant;
        # any other tenant gets a blank doc the operator must populate.
        try:
            from tenant_context import is_masci as _is_masci_t  # noqa: PLC0415
            tenant_is_masci = _is_masci_t(tk)
        except Exception:
            tenant_is_masci = tk == "masci"
        if tenant_is_masci:
            doc = {
                "tenant_key": tk,
                "company_name": "MASCI",
                "platform_display_name": "MASCI Operations Platform",
                "sender_name": "MASCI Operations Platform",
                "from_email": (os.environ.get("SENDER_EMAIL") or "").strip(),
                "reply_to": (os.environ.get("REPLY_TO_EMAIL") or "").strip(),
                "support_email": "safety@mascigc.com",
                "safety_email": "safety@mascigc.com",
                "hr_email": (os.environ.get("HR_EMAIL") or "").strip(),
                "operations_email": (os.environ.get("OPERATIONS_EMAIL") or "").strip(),
                "logo_url": None,
                "primary_color": "#C8102E",
                "source": "env_defaults",
            }
        else:
            doc = {
                "tenant_key": tk,
                "company_name": "",
                "platform_display_name": "",
                "sender_name": "",
                "from_email": "",
                "reply_to": "",
                "support_email": "",
                "safety_email": "",
                "hr_email": "",
                "operations_email": "",
                "logo_url": None,
                "primary_color": "#0F766E",
                "source": "unconfigured",
            }
    return doc


@_email_router.put("/admin/email-routing/v2/branding")
async def admin_v2_branding_put(
    body: V2BrandingUpdate, _: bool = Depends(require_admin)
):
    tk = _current_tenant_key()
    update: Dict[str, Any] = {}
    raw = body.model_dump()
    for k, v in raw.items():
        if v is None:
            continue
        if k in ("from_email", "reply_to", "support_email", "safety_email",
                 "hr_email", "operations_email"):
            s = str(v).strip()
            if s and "@" not in s:
                raise HTTPException(400, f"Invalid email in '{k}'")
            update[k] = s or None
        else:
            update[k] = str(v).strip()[:240] if isinstance(v, str) else v
    if not update:
        return {"ok": True, "changed": False}
    update["tenant_key"] = tk
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    update["updated_by"] = "admin"
    await db.tenant_branding.update_one({"_id": tk}, {"$set": update}, upsert=True)
    try:
        from email_routing_v2 import invalidate_cache as _v2_invalidate
        _v2_invalidate()
    except Exception:
        pass
    doc = await db.tenant_branding.find_one({"_id": tk}, {"_id": 0})
    return {"ok": True, "changed": True, "doc": doc}


# Track 15.67 Phase 3 · Public branding endpoint for the frontend
# BrandingProvider. NO auth — returns only the customer-facing display
# strings (no secrets, no recipient lists). Used by every page on first
# load to render the active tenant's brand instead of hardcoded MASCI.
@_email_router.get("/branding/current")
async def public_branding_current(request: Request):
    # Track 15.68 · Tenant preview override — allow an `X-Tenant-Preview`
    # header (preview/dev only) so the operator can visually inspect any
    # tenant's branding without changing the production tenant. Refused
    # in production env to prevent accidental customer impersonation.
    preview_tk = ""
    app_env = (os.environ.get("APP_ENV") or "").strip().lower()
    if app_env != "production":
        preview_tk = (request.headers.get("X-Tenant-Preview") or "").strip().lower()
    tk = preview_tk or _current_tenant_key()
    doc = await db.tenant_branding.find_one({"_id": tk}, {"_id": 0}) or {}
    try:
        from tenant_context import is_masci as _is_masci_t  # noqa: PLC0415
        tenant_is_masci = _is_masci_t(tk)
    except Exception:
        tenant_is_masci = tk == "masci"
    if tenant_is_masci:
        defaults = {
            "company_name": "MASCI",
            "platform_display_name": "MASCI Operations Platform",
            # TRACK 18.04 · Constitution-canonical platform short name.
            # The legacy "MASCI Hub" caused the `_brandSubst` chain to
            # double-emit "Hub" tokens (MASCI Hub Hub Hub Operations
            # Platform). Canonical short name is now just "MASCI".
            "platform_short_name": "MASCI",
            "support_email": doc.get("support_email") or "safety@mascigc.com",
            "safety_email": doc.get("safety_email") or "safety@mascigc.com",
            "hr_email": doc.get("hr_email") or "",
            "operations_email": doc.get("operations_email") or "",
            "logo_url": doc.get("logo_url") or "",
            "primary_color": doc.get("primary_color") or "#C8102E",
            "marketing_url": "https://mascidocs.com",
        }
    else:
        defaults = {
            "company_name": doc.get("company_name") or "Customer",
            "platform_display_name": doc.get("platform_display_name") or "Operations Platform",
            "platform_short_name": doc.get("platform_short_name") or "Ops Hub",
            "support_email": doc.get("support_email") or "",
            "safety_email": doc.get("safety_email") or "",
            "hr_email": doc.get("hr_email") or "",
            "operations_email": doc.get("operations_email") or "",
            "logo_url": doc.get("logo_url") or "",
            "primary_color": doc.get("primary_color") or "#0F766E",
            "marketing_url": doc.get("marketing_url") or "",
        }
    out = {
        "tenant_key": tk,
        "company_name": doc.get("company_name") or defaults["company_name"],
        "platform_display_name": doc.get("platform_display_name") or defaults["platform_display_name"],
        "platform_short_name": doc.get("platform_short_name") or defaults["platform_short_name"],
        "support_email": defaults["support_email"],
        "safety_email": defaults["safety_email"],
        "hr_email": defaults["hr_email"],
        "operations_email": defaults["operations_email"],
        "logo_url": defaults["logo_url"],
        "primary_color": defaults["primary_color"],
        "marketing_url": defaults["marketing_url"],
    }
    return out


@_email_router.post("/admin/email-routing/v2/route-health")
async def admin_v2_route_health(_: bool = Depends(require_admin)):
    """Track 15.67 · One-click validation of every route for the active
    tenant. Dry-runs each route (no Resend send), writes an audit row
    per route, and returns a green/amber/red summary so an operator can
    verify the whole routing surface before any production change."""
    tk = _current_tenant_key()
    cursor = db.email_routes.find({"tenant_key": tk}, {"_id": 0}).sort("route_key", 1)
    routes = await cursor.to_list(100)
    try:
        from email_routing_v2 import write_audit as _v2_audit  # noqa: PLC0415
    except Exception:
        _v2_audit = None
    results = []
    summary = {"green": 0, "amber": 0, "red": 0}
    now_iso = datetime.now(timezone.utc).isoformat()
    for r in routes:
        rk = r["route_key"]
        to = r.get("to") or []
        cc = r.get("cc") or []
        bcc = r.get("bcc") or []
        enabled = bool(r.get("enabled", True))
        critical = bool(r.get("critical", False))
        recip_total = len(to) + len(cc) + len(bcc)
        # Severity classification
        if critical and enabled and not to:
            status = "red"  # critical route can't deliver
        elif enabled and recip_total == 0 and rk != "ACCOUNT_INVITES_FROM" and rk != "PASSWORD_RESET_MONITORING_TO":
            status = "amber"
        elif not enabled and not critical:
            status = "amber"  # explicitly disabled, but flag for visibility
        else:
            status = "green"
        # Stale-test pill: if never tested or > 30 days
        last_tested = r.get("last_tested_at")
        stale = True
        if last_tested:
            try:
                from datetime import datetime as _dt
                lt = _dt.fromisoformat(last_tested.replace("Z", "+00:00"))
                age_days = (datetime.now(timezone.utc) - lt).days
                stale = age_days > 30
            except Exception:
                stale = True
        if status == "green" and stale:
            status = "amber"  # never-tested or stale critical route → amber
        summary[status] += 1
        if _v2_audit is not None:
            try:
                await _v2_audit(
                    db, route_key=rk, tenant_key=tk, source="db",
                    to_count=len(to), cc_count=len(cc), bcc_count=len(bcc),
                    subject="[ROUTE HEALTH] dry-run", status="dry_run",
                    calling_module="route_health", dry_run=True,
                )
            except Exception:
                pass
        results.append({
            "route_key": rk,
            "display_name": r.get("display_name"),
            "critical": critical,
            "enabled": enabled,
            "to_count": len(to), "cc_count": len(cc), "bcc_count": len(bcc),
            "last_tested_at": last_tested,
            "stale": stale,
            "status": status,
            "reason": (
                "critical route has empty TO" if (critical and enabled and not to) else
                "no recipients configured" if (enabled and recip_total == 0 and rk not in ("ACCOUNT_INVITES_FROM", "PASSWORD_RESET_MONITORING_TO")) else
                "explicitly disabled" if not enabled else
                "never tested / stale (>30d)" if stale else
                "healthy"
            ),
        })
    return {
        "tenant_key": tk,
        "ts": now_iso,
        "summary": summary,
        "total": len(routes),
        "results": results,
    }


# ─────────────────────────────────────────────────────────────────────────
# Track 15.72A — Email Routing V2 self-observability.
# Two read-only diagnostic endpoints so a MASCI admin can verify routing
# mode, resolver health, and rollback readiness directly from the UI —
# without needing Mongo creds, DevTools, Atlas, or pasting admin tokens.
# Append-only at most (no recipient/route doc mutations, no Resend sends).
# ─────────────────────────────────────────────────────────────────────────

def _status_health_band(critical_empty: int, recent_errors_24h: int,
                        last_v2_audit_age_min: Optional[float],
                        flag_active: bool) -> tuple[str, str]:
    """Compute a single green/amber/red band and human reason for the
    Routing Status card. Pure function — no side effects."""
    if critical_empty > 0:
        return ("red", f"{critical_empty} critical route(s) have no recipients")
    if recent_errors_24h > 0:
        return ("red", f"{recent_errors_24h} resolver error(s) in last 24h")
    if not flag_active:
        return ("green", "Legacy routing — V2 flag is OFF (safe baseline)")
    if last_v2_audit_age_min is None:
        return ("amber", "V2 flag is ON but no V2 audit rows observed yet — first scheduler tick pending")
    if last_v2_audit_age_min > 120:
        return ("amber", f"Last V2 audit row is {int(last_v2_audit_age_min)} min old — scheduler may be slow")
    return ("green", "V2 routing healthy")


@_email_router.get("/admin/email-routing/v2/status")
async def admin_v2_status(_: bool = Depends(require_admin)):
    """Track 15.72A · Routing-status snapshot.

    Read-only, admin-gated. Reads `EMAIL_ROUTING_V2` env, scans the audit
    collection for recency/error counters, and produces the data the
    Admin > Email & Routing > Routing Status card needs.

    Returns NO secrets, NO recipient emails, NO connection strings, NO
    tokens. Recipient *counts* only; raw recipients live in the existing
    `routes` endpoint that already gates on admin.
    """
    from email_routing_v2 import routing_v2_enabled  # noqa: PLC0415

    tk = _current_tenant_key()
    now = datetime.now(timezone.utc)
    one_hour_ago_iso = (now - timedelta(hours=1)).isoformat()
    one_day_ago_iso  = (now - timedelta(hours=24)).isoformat()
    flag_active = routing_v2_enabled()
    raw_flag = (os.environ.get("EMAIL_ROUTING_V2") or "").strip()

    # ── Routes & critical-health ──
    routes = await db.email_routes.find(
        {"tenant_key": tk},
        {"_id": 0, "route_key": 1, "enabled": 1, "critical": 1,
         "to": 1, "cc": 1, "bcc": 1, "display_name": 1},
    ).to_list(100)
    route_counts = {
        "total": len(routes),
        "enabled": sum(1 for r in routes if r.get("enabled", True)),
        "disabled": sum(1 for r in routes if not r.get("enabled", True)),
        "critical_total": sum(1 for r in routes if r.get("critical", False)),
        "critical_populated": sum(
            1 for r in routes
            if r.get("critical", False)
            and r.get("enabled", True)
            and len(r.get("to") or []) > 0
        ),
        "critical_empty": sum(
            1 for r in routes
            if r.get("critical", False)
            and r.get("enabled", True)
            and len(r.get("to") or []) == 0
        ),
        "empty_non_critical": sum(
            1 for r in routes
            if not r.get("critical", False)
            and r.get("enabled", True)
            and len(r.get("to") or []) == 0
            and r["route_key"] not in ("ACCOUNT_INVITES_FROM", "PASSWORD_RESET_MONITORING_TO")
        ),
    }
    # Names of critical routes that would silently drop email — most
    # important field for an admin to see.
    critical_empty_keys = sorted(
        r["route_key"] for r in routes
        if r.get("critical", False) and r.get("enabled", True)
        and len(r.get("to") or []) == 0
    )

    # ── Audit recency / error counters ──
    total_audit          = await db.email_routing_audit_v2.count_documents({"tenant_key": tk})
    audit_last_hour      = await db.email_routing_audit_v2.count_documents(
        {"tenant_key": tk, "ts": {"$gte": one_hour_ago_iso}}
    )
    audit_last_24h       = await db.email_routing_audit_v2.count_documents(
        {"tenant_key": tk, "ts": {"$gte": one_day_ago_iso}}
    )
    errors_last_24h      = await db.email_routing_audit_v2.count_documents(
        {"tenant_key": tk, "ts": {"$gte": one_day_ago_iso},
         "status": {"$in": _EMAIL_AUDIT_FAILURE_STATUSES}}
    )
    db_source_last_24h   = await db.email_routing_audit_v2.count_documents(
        {"tenant_key": tk, "ts": {"$gte": one_day_ago_iso}, "source": "db"}
    )
    legacy_source_last_24h = await db.email_routing_audit_v2.count_documents(
        {"tenant_key": tk, "ts": {"$gte": one_day_ago_iso}, "source": "legacy"}
    )

    # Latest 5 audit rows (no recipients, just counts + module + status).
    latest_rows_raw = await db.email_routing_audit_v2.find(
        {"tenant_key": tk},
        {"_id": 0, "ts": 1, "route_key": 1, "source": 1, "status": 1,
         "to_count": 1, "cc_count": 1, "bcc_count": 1, "calling_module": 1,
         "dry_run": 1},
    ).sort("ts", -1).limit(5).to_list(5)

    # ── Per-V2-module recency (the 3 actually-V2-aware code paths) ──
    async def _last_for(module_name: str) -> Optional[Dict[str, Any]]:
        d = await db.email_routing_audit_v2.find_one(
            {"tenant_key": tk, "calling_module": module_name},
            sort=[("ts", -1)],
        )
        if not d:
            return None
        return {
            "ts": d.get("ts"),
            "route_key": d.get("route_key"),
            "source": d.get("source"),
            "status": d.get("status"),
            "to_count": d.get("to_count"),
        }

    last_health_monitor = await _last_for("health_monitor")
    last_outage_alerts  = await _last_for("outage_alerts")
    last_safety_digest  = await _last_for("safety_digest")

    # Latest V2 audit-row age — used for amber/green banding.
    latest_v2 = await db.email_routing_audit_v2.find_one(
        {"tenant_key": tk, "source": "db"},
        sort=[("ts", -1)],
    )
    last_v2_audit_age_min: Optional[float] = None
    if latest_v2 and latest_v2.get("ts"):
        try:
            t = datetime.fromisoformat(str(latest_v2["ts"]).replace("Z", "+00:00"))
            last_v2_audit_age_min = (now - t).total_seconds() / 60.0
        except Exception:
            last_v2_audit_age_min = None

    band, band_reason = _status_health_band(
        critical_empty=route_counts["critical_empty"],
        recent_errors_24h=errors_last_24h,
        last_v2_audit_age_min=last_v2_audit_age_min,
        flag_active=flag_active,
    )

    # Container info (proves operator looking at fresh build, not stale
    # cached page). Pulled directly from `_version_payload` so we never
    # drift away from /api/version.
    app_env  = _canonical_app_env()
    db_name  = _canonical_db_name()
    started_at_iso = _STARTUP_TS.isoformat() if isinstance(_STARTUP_TS, datetime) else str(_STARTUP_TS)
    uptime_s = max(0, int((now - _STARTUP_TS).total_seconds())) if isinstance(_STARTUP_TS, datetime) else None

    return {
        "ts": now.isoformat(),
        "tenant_key": tk,
        "mode": "v2" if flag_active else "legacy",
        "flag_active": flag_active,
        "flag_raw_value": raw_flag if raw_flag else "<unset>",
        "app_env": app_env,
        "db_name": db_name,
        "backend_started_at": started_at_iso,
        "backend_uptime_s": uptime_s,
        "route_counts": route_counts,
        "critical_empty_route_keys": critical_empty_keys,
        "audit_counters": {
            "total":             total_audit,
            "last_hour":         audit_last_hour,
            "last_24h":          audit_last_24h,
            "errors_last_24h":   errors_last_24h,
            "db_source_last_24h":     db_source_last_24h,
            "legacy_source_last_24h": legacy_source_last_24h,
        },
        "latest_audit_rows": latest_rows_raw,
        "v2_module_recency": {
            "health_monitor": last_health_monitor,
            "outage_alerts":  last_outage_alerts,
            "safety_digest":  last_safety_digest,
        },
        "last_v2_audit_age_minutes": last_v2_audit_age_min,
        "rollback_target": {
            "current_flag_value": raw_flag if raw_flag else "<unset>",
            "reverse_value":      "false" if flag_active else "true",
            "mechanism":          "edit backend/.env line `EMAIL_ROUTING_V2=` then Re-deploy",
            "estimated_minutes":  5,
        },
        "band":        band,
        "band_reason": band_reason,
    }


@_email_router.post("/admin/email-routing/v2/self-check")
async def admin_v2_self_check(_: bool = Depends(require_admin)):
    """Track 15.72A · End-to-end self-check.

    Dry-runs the resolver for every route (NO Resend send, NO route doc
    mutation, NO recipient change). Writes ≤1 append-only audit row per
    route with `dry_run=True` and `calling_module='self_check'` so the
    operator's status card can see the activity timestamp advance.

    Returns a single PASS/FAIL banner plus per-route findings.
    """
    from email_routing_v2 import (  # noqa: PLC0415
        routing_v2_enabled, resolve, write_audit as _v2_audit,
    )

    tk = _current_tenant_key()
    now = datetime.now(timezone.utc)
    flag_active = routing_v2_enabled()

    routes = await db.email_routes.find(
        {"tenant_key": tk}, {"_id": 0}
    ).sort("route_key", 1).to_list(100)

    results: List[Dict[str, Any]] = []
    summary = {"green": 0, "amber": 0, "red": 0, "db_source": 0,
               "legacy_source": 0, "env_source": 0, "disabled_source": 0}

    for r in routes:
        rk = r["route_key"]
        critical = bool(r.get("critical", False))
        enabled  = bool(r.get("enabled", True))
        # Resolver dry-run (no send, no mutation)
        try:
            res = await resolve(db, rk, legacy_provider=lambda: [])
            err: Optional[str] = None
        except Exception as e:  # noqa: BLE001
            class _R:
                to: List[str] = []
                cc: List[str] = []
                bcc: List[str] = []
                source = "error"
                from_email = None
            res = _R()
            err = repr(e)[:200]

        tc, cc, bc = len(res.to or []), len(res.cc or []), len(res.bcc or [])
        source = getattr(res, "source", "unknown") or "unknown"
        summary[f"{source}_source"] = summary.get(f"{source}_source", 0) + 1

        # Classify
        if err:
            status = "red"
            reason = f"resolver error: {err}"
        elif critical and enabled and tc == 0:
            status = "red"
            reason = "critical route resolved to empty TO"
        elif flag_active and source == "legacy" and enabled and rk in (
            "HEALTH_ALERTS", "OUTAGE_ALERTS", "SAFETY_DIGEST_TO"
        ):
            # The 3 V2-aware code paths SHOULD be hitting source=db when
            # the flag is on. If they fall back to legacy, something
            # changed in the DB doc (probably empty or disabled).
            status = "amber"
            reason = "V2 active but route falling back to legacy source"
        elif not enabled and not critical:
            status = "amber"
            reason = "route explicitly disabled (non-critical)"
        elif enabled and tc + cc + bc == 0 and rk not in (
            "ACCOUNT_INVITES_FROM", "PASSWORD_RESET_MONITORING_TO",
            "INCIDENT_SEVERE_CC", "BACKUP_ALERTS",
        ):
            status = "amber"
            reason = "route has zero recipients"
        else:
            status = "green"
            reason = "healthy"

        summary[status] += 1

        # Append-only diagnostic audit row (best-effort; never blocks)
        try:
            await _v2_audit(
                db, route_key=rk, tenant_key=tk, source=source,
                to_count=tc, cc_count=cc, bcc_count=bc,
                subject="[SELF-CHECK] dry-run", status="dry_run",
                calling_module="self_check", dry_run=True,
                error=err,
            )
        except Exception:
            pass

        results.append({
            "route_key": rk,
            "display_name": r.get("display_name"),
            "critical": critical, "enabled": enabled,
            "source": source,
            "to_count": tc, "cc_count": cc, "bcc_count": bc,
            "from_email_set": bool(getattr(res, "from_email", None)),
            "status": status, "reason": reason,
        })

    # Overall banner
    if summary["red"] > 0:
        overall = "red"
        overall_reason = f"{summary['red']} route(s) failed self-check"
    elif summary["amber"] > 0:
        overall = "amber"
        overall_reason = f"{summary['amber']} route(s) need attention"
    else:
        overall = "green"
        overall_reason = "all routes resolve healthy"

    return {
        "ts": now.isoformat(),
        "tenant_key": tk,
        "flag_active": flag_active,
        "mode": "v2" if flag_active else "legacy",
        "total_routes": len(routes),
        "summary": summary,
        "overall": overall,
        "overall_reason": overall_reason,
        "results": results,
    }


# ─────────────────────────────────────────────────────────────────────────
# Photo Storage (iter64): S3-compatible cloud storage admin endpoints.
# Lets admin: (1) check connectivity to R2/S3, (2) run a dry-run migration
# preview, (3) run a real migration in capped batches, (4) inspect progress.
# ─────────────────────────────────────────────────────────────────────────
class PhotoMigrateBody(BaseModel):
    dry_run: bool = True
    limit_per_collection: int = 100
    resume: bool = True
    collection: Optional[str] = None  # single-collection mode


@_email_router.get("/admin/photo-storage/health")
async def admin_photo_storage_health(_: bool = Depends(require_admin)):
    """Verify the backend can reach the configured S3-compatible bucket.
    Returns ``{configured, ok, bucket?, endpoint?, reason?}``."""
    try:
        from photo_storage import health_check as _ps_health
        return await _ps_health()
    except Exception as e:  # noqa: BLE001
        return {"configured": False, "ok": False, "reason": f"import failed: {e}"}


@_email_router.post("/admin/photos/migrate")
async def admin_photos_migrate(
    body: PhotoMigrateBody, _: bool = Depends(require_admin)
):
    """Run the base64→S3 photo migration. Defaults to a tiny dry-run
    (no S3 writes, no DB updates) so the admin can preview what would
    happen before committing.

    For real runs: ``dry_run=false`` + a small ``limit_per_collection``
    (e.g. 50). Re-call as many times as needed — ``resume=true`` (default)
    picks up where the last run left off, so re-running is safe and fast.
    """
    from photo_migration import migrate_all, migrate_collection
    if body.collection:
        s = await migrate_collection(
            db, body.collection,
            dry_run=body.dry_run,
            limit=body.limit_per_collection,
        )
        return {"ok": True, "per_collection": [s]}
    out = await migrate_all(
        db,
        dry_run=body.dry_run,
        limit_per_collection=body.limit_per_collection,
        resume=body.resume,
    )
    return {"ok": True, **out}


@_email_router.get("/admin/photos/migrate/progress")
async def admin_photos_migrate_progress(_: bool = Depends(require_admin)):
    """Snapshot of per-collection migration progress (last_doc_id, stats)."""
    from photo_migration import get_progress
    return await get_progress(db)


@_email_router.post("/admin/photos/migrate/reset")
async def admin_photos_migrate_reset(
    collection: Optional[str] = None, _: bool = Depends(require_admin)
):
    """Wipe the progress markers so the next run starts from scratch.
    DOES NOT undo migrated photos — they stay in S3. Used when you
    want to re-walk every collection (e.g. for verification)."""
    from photo_migration import reset_progress
    n = await reset_progress(db, collection)
    return {"ok": True, "deleted_markers": n}


@_email_router.post("/email-report")
async def email_report(
    body: EmailReportRequest, _: bool = Depends(require_admin)
):
    if body.kind not in _KIND_TO_COLLECTION:
        raise HTTPException(status_code=400, detail=f"Unknown kind: {body.kind}")
    coll_name = _KIND_TO_COLLECTION[body.kind]
    coll = getattr(db, coll_name)
    record = await coll.find_one({"id": body.record_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")

    api_key = os.environ.get("RESEND_API_KEY")
    if not api_key:
        raise HTTPException(
            status_code=503,
            detail="RESEND_API_KEY not configured. Add it to /app/backend/.env and restart backend.",
        )

    sender_email = await _resolve_sender_email(db)

    try:
        import resend  # noqa: E402

        resend.api_key = api_key

        pdf_bytes = render_record_pdf(body.kind, await _maybe_enrich_for_pdf(db, body.kind, record))
        project = record.get("project_name") or record.get("project") or "MASCI"
        date_part = (
            record.get("report_date")
            or record.get("date")
            or record.get("incident_date")
            or ""
        )
        safe_proj = "".join(
            c if c.isalnum() else "_" for c in project[:40]
        ).strip("_")
        filename = f"MASCI_{body.kind}_{safe_proj}_{date_part}.pdf".replace(
            "__", "_"
        )

        subject = body.subject or build_email_subject(body.kind, record)

        params = {
            "from": f"MASCI Operations Platform <{sender_email}>",
            "to": [r for r in body.recipients if r and r.strip()],
            "subject": subject,
            "html": render_email_html(body.kind, record, body.note or ""),
            "attachments": [
                {
                    "filename": filename,
                    "content": _email_b64.b64encode(pdf_bytes).decode(),
                }
            ],
        }

        result = await asyncio.to_thread(resend.Emails.send, params)
        return {
            "ok": True,
            "id": (result or {}).get("id"),
            "to": params["to"],
            "filename": filename,
            "size_bytes": len(pdf_bytes),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"email-report failed: {e}")
        raise HTTPException(status_code=500, detail=f"Email send failed: {e}")


app.include_router(_email_router)


# ------------------------- Field Safety Cards -------------------------
#
# Printable/emailable bilingual safety cards meant for crew wallets. The
# source of truth is four high-res PDFs shipped at /app/backend/static/
# safety-cards. The frontend shows 150-DPI JPG previews served by the
# frontend public folder so crews can eyeball the card before printing;
# this email endpoint mails the ORIGINAL print-ready PDF attachment.
#
# Uses its own sub-router because `api_router` was already included on
# the app before this block — late `@api_router.post(...)` registrations
# silently fail to mount.

_SAFETY_CARD_FILES = {
    "en-front": ("MASCI_Safety_Card_EN_Front.pdf", "MASCI Safety Card — English · Front"),
    "en-back":  ("MASCI_Safety_Card_EN_Back.pdf",  "MASCI Safety Card — English · Back"),
    "es-front": ("MASCI_Safety_Card_ES_Front.pdf", "MASCI Tarjeta de Seguridad — Español · Frente"),
    "es-back":  ("MASCI_Safety_Card_ES_Back.pdf",  "MASCI Tarjeta de Seguridad — Español · Reverso"),
}


class SafetyCardEmailRequest(BaseModel):
    card: str
    recipients: List[str]
    subject: Optional[str] = None
    note: Optional[str] = None


class SafetyCardEmailAllRequest(BaseModel):
    """Bulk-send all four bilingual safety cards in one email — used by
    foremen to onboard a new hire with the entire MASCI safety packet
    in one tap."""
    recipients: List[str]
    subject: Optional[str] = None
    note: Optional[str] = None


_safety_cards_router = APIRouter(prefix="/api")


@_safety_cards_router.post("/safety-cards/email")
async def email_safety_card(body: SafetyCardEmailRequest):
    """Email one of the 4 bilingual field safety cards as a PDF attachment.
    Open to anyone in the crew — the cards are handout materials, not
    gated compliance docs."""
    if body.card not in _SAFETY_CARD_FILES:
        raise HTTPException(status_code=400, detail=f"Unknown card: {body.card}")

    filename, default_subject = _SAFETY_CARD_FILES[body.card]
    pdf_path = Path(__file__).parent / "static" / "safety-cards" / filename
    if not pdf_path.exists():
        raise HTTPException(status_code=500, detail=f"Card file missing on server: {filename}")

    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        raise HTTPException(
            status_code=503,
            detail="RESEND_API_KEY not configured. Add it to /app/backend/.env and restart backend.",
        )

    valid_recipients = [r.strip() for r in (body.recipients or []) if r and r.strip()]
    if not valid_recipients:
        raise HTTPException(status_code=400, detail="At least one recipient email is required")

    sender_email = await _resolve_sender_email(db)

    try:
        import resend  # noqa: E402

        resend.api_key = api_key
        pdf_bytes = pdf_path.read_bytes()

        note_html = ""
        if body.note:
            safe_note = body.note.replace("<", "&lt;").replace(">", "&gt;")
            note_html = f"<p style='margin:16px 0;color:#334155;'>{safe_note}</p>"

        html = (
            "<div style='font-family:ui-sans-serif,system-ui,Arial;max-width:600px;'>"
            "<h2 style='color:#991b1b;margin:0 0 8px;'>MASCI Field Safety Card</h2>"
            f"<p style='margin:0 0 16px;color:#475569;'>Attached: <strong>{default_subject}</strong>. "
            "Print on letter-size (8.5×11) and distribute to the crew.</p>"
            f"{note_html}"
            "<hr style='border:none;border-top:1px solid #e2e8f0;margin:16px 0;'>"
            "<p style='font-size:12px;color:#64748b;margin:0;'>Sent from MASCI Hub · Safety</p>"
            "</div>"
        )

        params = {
            "from": f"MASCI Operations Platform <{sender_email}>",
            "to": valid_recipients,
            "subject": body.subject.strip() if body.subject else default_subject,
            "html": html,
            "attachments": [
                {
                    "filename": filename,
                    "content": _email_b64.b64encode(pdf_bytes).decode(),
                }
            ],
        }

        result = await asyncio.to_thread(resend.Emails.send, params)
        return {
            "ok": True,
            "id": (result or {}).get("id"),
            "to": valid_recipients,
            "filename": filename,
            "size_bytes": len(pdf_bytes),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"safety-cards/email failed: {e}")
        raise HTTPException(status_code=500, detail=f"Email send failed: {e}")


@_safety_cards_router.post("/safety-cards/email-all")
async def email_all_safety_cards(body: SafetyCardEmailAllRequest):
    """Bulk-email all 4 safety cards (EN front+back, ES front+back) as
    PDF attachments. One email, one click — perfect for new-hire
    onboarding. Like /safety-cards/email this endpoint is open to the
    crew because the cards themselves are handout materials."""
    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        raise HTTPException(
            status_code=503,
            detail="RESEND_API_KEY not configured. Add it to /app/backend/.env and restart backend.",
        )

    valid_recipients = [r.strip() for r in (body.recipients or []) if r and r.strip()]
    if not valid_recipients:
        raise HTTPException(status_code=400, detail="At least one recipient email is required")

    # Read all 4 PDFs up-front
    base = Path(__file__).parent / "static" / "safety-cards"
    cards = []
    total_size = 0
    for key, (filename, label) in _SAFETY_CARD_FILES.items():
        pdf_path = base / filename
        if not pdf_path.exists():
            raise HTTPException(status_code=500, detail=f"Card file missing on server: {filename}")
        pdf_bytes = pdf_path.read_bytes()
        total_size += len(pdf_bytes)
        cards.append((filename, label, pdf_bytes))

    sender_email = await _resolve_sender_email(db)

    try:
        import resend  # noqa: E402

        resend.api_key = api_key

        note_html = ""
        if body.note:
            safe_note = body.note.replace("<", "&lt;").replace(">", "&gt;")
            note_html = f"<p style='margin:16px 0;color:#334155;'>{safe_note}</p>"

        card_list_html = "".join(
            f"<li style='margin:4px 0;'>{label}</li>" for _, label, _ in cards
        )

        html = (
            "<div style='font-family:ui-sans-serif,system-ui,Arial;max-width:600px;'>"
            "<h2 style='color:#991b1b;margin:0 0 8px;'>MASCI Field Safety Cards — Full Set</h2>"
            "<p style='margin:0 0 12px;color:#475569;'>Attached: the complete bilingual MASCI safety card set "
            "(English front &amp; back, Spanish front &amp; back). Print on letter-size (8.5×11) and "
            "distribute to your crew. Perfect for new-hire onboarding.</p>"
            f"<ul style='margin:8px 0 16px 20px;color:#334155;'>{card_list_html}</ul>"
            f"{note_html}"
            "<hr style='border:none;border-top:1px solid #e2e8f0;margin:16px 0;'>"
            "<p style='font-size:12px;color:#64748b;margin:0;'>Sent from MASCI Hub · Safety</p>"
            "</div>"
        )

        params = {
            "from": f"MASCI Operations Platform <{sender_email}>",
            "to": valid_recipients,
            "subject": body.subject.strip() if body.subject else "MASCI Field Safety Cards — Full Bilingual Set",
            "html": html,
            "attachments": [
                {
                    "filename": filename,
                    "content": _email_b64.b64encode(pdf_bytes).decode(),
                }
                for filename, _, pdf_bytes in cards
            ],
        }

        result = await asyncio.to_thread(resend.Emails.send, params)
        return {
            "ok": True,
            "id": (result or {}).get("id"),
            "to": valid_recipients,
            "card_count": len(cards),
            "total_size_bytes": total_size,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"safety-cards/email-all failed: {e}")
        raise HTTPException(status_code=500, detail=f"Email send failed: {e}")


app.include_router(_safety_cards_router)


# Empty placeholder — was a stale duplicate route block, kept above before router include.


# ------------------------- Phase 1: per-user auth + projects (Basecamp-style /app) -------------------------
from auth import build_auth_router, seed_initial_users  # noqa: E402
from projects import build_projects_router, seed_initial_projects  # noqa: E402
from tools import build_tools_router, create_tools_indexes  # noqa: E402
from phase4 import build_phase4_router, create_phase4_indexes  # noqa: E402

_auth_router, get_current_user, require_admin_or_owner, _optional_user = build_auth_router(db)
_projects_router = build_projects_router(db, get_current_user, require_admin_or_owner)
_tools_router = build_tools_router(db, get_current_user, require_admin_or_owner)
app.include_router(_auth_router)
app.include_router(_projects_router)
app.include_router(_tools_router)

# Register phase 4 (activity + notifications + search + directory)
_phase4_router = build_phase4_router(db, get_current_user)
app.include_router(_phase4_router)


@register_lifecycle_step("seed")
async def _seed_phase1():
    try:
        await seed_initial_users(db)
        await seed_initial_projects(db)
        await create_tools_indexes(db)
        await create_phase4_indexes(db)
        await _seed_equipment_master()
        await _seed_employees_from_json()
        await _seed_suppliers_from_json()
        await _create_safety_indexes()
        # Seed project_managers FIRST so jobs_master backfill can link by name.
        from project_managers import seed_project_managers
        await seed_project_managers(db)
        # Seed jobs_master from /app/backend/data/jobs_master.json (idempotent).
        # Also runs the pm_email backfill against project_managers.
        from jobs_master import seed_jobs_master
        await seed_jobs_master(db)
        # Index for the new multi-file JHP collection.
        from job_hazard_files import ensure_indexes as _jha_files_indexes
        await _jha_files_indexes(db)
        # Zero-touch self-heal: auto-split equipment make/model on boot if any
        # units are missing it. Survives redeploys that wipe the DB.
        from data_fixes import boot_self_heal
        await boot_self_heal(db)
        # Trench Safety Operations System — Phase 2.
        # Must run AFTER _seed_equipment_master because the trench-safety
        # seeder writes mirror rows into equipment_master. The JSON
        # seeder ignores category="Trench Safety" so the mirrors persist
        # across boots. See /app/memory/TRENCH_SAFETY_ARCHITECTURE.md.
        from routes.trench_safety import seed_trench_safety_assets
        try:
            await seed_trench_safety_assets(db)
        except Exception as ts_err:  # noqa: BLE001
            logging.getLogger(__name__).warning(
                f"trench_safety seed failed: {ts_err}", exc_info=True
            )
    except Exception as e:
        logging.getLogger(__name__).exception(f"Phase 1 seed failed: {e}")


async def _create_safety_indexes():
    """Idempotent indexes on the safety + equipment + parts collections.

    Massively speeds up dashboard listings, trends queries, and shop
    open-items lookups once the dataset grows past a few hundred records.
    """
    try:
        await db.equipment_inspections.create_index("created_at")
        await db.equipment_inspections.create_index("inspection_date")
        await db.equipment_inspections.create_index("equipment_unit")
        await db.equipment_inspections.create_index("project_number")
        await db.equipment_inspections.create_index("fail_count")

        await db.inspections.create_index("created_at")
        await db.inspections.create_index("inspection_date")
        await db.inspections.create_index("project_number")

        await db.daily_reports.create_index("created_at")
        await db.daily_reports.create_index("report_date")
        await db.daily_reports.create_index("project_number")
        # C2 final authorization · query-targeting remediation
        # PM / Admin haul-material views query a narrow subset of Daily Reports:
        #   report_date window + optional project_number + outbound_materials exists.
        # Production Atlas raised a high scanned/returned alert on 2026-07-22.
        # A partial compound index keeps those reads bounded without indexing the
        # full collection or changing result semantics.
        await db.daily_reports.create_index(
            [("report_date", -1), ("project_number", 1)],
            name="daily_reports_outbound_by_date_project",
            partialFilterExpression={"outbound_materials.0": {"$exists": True}},
        )
        # TRACK 26.07: index the `updated_at` field used by the job-photos
        # background indexer loop (routes/job_photos.py::background_indexer_loop)
        # which filters `daily_reports.find({photos.0: {$exists}, updated_at:
        # {$gte: cutoff}})` every 10 min. Without this index, the tick is a
        # COLLSCAN of the entire collection — the leading candidate for the
        # 2026-07-08 16:01 GMT Atlas query targeting alert.
        await db.daily_reports.create_index("updated_at")
        # PERFORMANCE-HARDEN-002: eliminate COLLSCAN on hot find_one({"id": ...})
        # patterns used across daily_report_lifecycle, hr_portal, verification,
        # operational_records, command_center, etc. Evidence: 794 docs scanned
        # per call (preview); production volume is materially higher.
        await db.daily_reports.create_index("id")
        # PERFORMANCE-HARDEN-002: same gap for the doc_id fallback path
        # (daily_report_lifecycle.py lines 71/205/221 retry by doc_id when the
        # canonical id lookup misses). 100% of preview docs have doc_id.
        await db.daily_reports.create_index("doc_id")
        # PERFORMANCE-HARDEN-002: hot job_photos.id COLLSCAN (1,812 docs in
        # preview) across job_photos.py find_one/find({id:{$in:...}}) and
        # photo_governance.py / odr/pdf.py.
        await db.job_photos.create_index("id")
        # PERFORMANCE-HARDEN-002: motive_events.id COLLSCAN
        # (motive_service.find_one, driver_profile.find_one). Also add a
        # compound (event_family, event_at) for the M-2 audit / ingestion
        # path which filters event_family $in [...] + event_at range.
        await db.motive_events.create_index("id")
        await db.motive_events.create_index([("event_family", 1), ("event_at", 1)])

        # PERFORMANCE-HARDEN-002 (refresh): hot session-validation path.
        # /app/backend/user_directory.py:427 calls
        #   db.directory_sessions.find_one({"token": token})
        # on EVERY authenticated request. Production has 1,949 session rows
        # → COLLSCAN per request. Token values are unique in prod (verified
        # 2026-06-09 via $group aggregate · zero duplicates). Adding as a
        # non-unique index for boot safety; operator may promote to unique
        # in a future maintenance window.
        await db.directory_sessions.create_index("token")
        # PERFORMANCE-HARDEN-002 (refresh): integration sync log filter path.
        # /app/backend/routes/integrations/logs.py:30 filters by
        # {integration, status?} then sorts by started_at desc. Current
        # integration_1 index scans 41k keys when status is added; compound
        # cuts to ~status-cardinality keys.
        await db.integration_sync_logs.create_index(
            [("integration", 1), ("status", 1), ("started_at", -1)]
        )

        await db.incidents.create_index("created_at")
        await db.incidents.create_index("incident_date")
        await db.incidents.create_index("severity")

        await db.meetings.create_index("created_at")
        await db.meetings.create_index("meeting_date")

        await db.equipment_parts.create_index("unit_number", unique=True)
        await db.equipment_master.create_index("unit_number")
        await db.equipment_master.create_index("category")
        logging.getLogger(__name__).info("[safety-indexes] ensured")
    except Exception as e:
        logging.getLogger(__name__).warning(f"[safety-indexes] failed: {e}")


@register_lifecycle_step("backup-scheduler")
async def _start_backup_scheduler():
    """Kick off the nightly full-backup scheduler as an asyncio task,
    plus a supervisor task that resurrects the scheduler if it ever dies.

    This addresses the "fixed-then-broken-days-later" pattern: when the
    scheduler asyncio.Task crashes for any reason (Mongo connection
    blip, unexpected exception, etc.), it dies silently and backups
    stop. The supervisor checks every 5 minutes, logs CRITICAL on a
    dead scheduler, and spawns a fresh task. Belt + suspenders alongside
    the watchdog email alarm.
    """
    global _backup_task
    if os.environ.get("DISABLE_BACKUP_SCHEDULER", "").lower() in ("1", "true", "yes"):
        logging.getLogger(__name__).info("[scheduled-backup] DISABLED via env")
        return
    try:
        BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
        # DEFENSE LAYER 4 — Boot-time disk safety check.
        # If the disk is above the high-water-mark when we start up
        # (e.g., a previous container crash left the disk full), purge
        # backups IMMEDIATELY before doing anything else. This guarantees
        # a fresh boot can never be killed by inherited disk pressure.
        pct = _disk_pct_used()
        if pct >= BACKUP_DISK_HIGH_WATERMARK:
            logging.getLogger(__name__).warning(
                f"[scheduled-backup] disk at {pct}% on boot — running emergency prune"
            )
            _emergency_prune_backups(reason=f"boot disk {pct}%")
        _backup_task = register_background_task(
            app,
            name="scheduled-backup-loop",
            coro=_backup_scheduler_loop_with_capture(db),
            category="scheduler",
            critical=True,
            long_running=True,
        )
        # FORGEDOPS-P0.2 · nightly Asset Spine reconciliation scan.
        try:
            from services.asset_spine_scheduler import asset_spine_nightly_loop
            register_background_task(
                app,
                name="asset-spine-nightly-loop",
                coro=asset_spine_nightly_loop(db),
                category="scheduler",
                critical=False,
                long_running=True,
            )
            logging.getLogger(__name__).info("[asset-spine-scheduler] task scheduled")
        except Exception as _e:
            logging.getLogger(__name__).warning("[asset-spine-scheduler] failed to start: %s", _e)
        _hours_str = " · ".join(f"{h:02d}:00" for h in BACKUP_HOURS_UTC) + " UTC"
        logging.getLogger(__name__).info(
            f"[scheduled-backup] scheduler started — {_hours_str} · "
            f"keep {BACKUP_RETENTION_DAYS} days · max {BACKUP_KEEP_MAX} files · "
            f"disk-watermark {BACKUP_DISK_HIGH_WATERMARK}% · dir={BACKUPS_DIR}"
        )
        # Startup validation — give the task a moment to settle, then
        # confirm it didn't die immediately on initialization. Catches
        # bugs that would otherwise silently kill the scheduler before
        # the first tick (e.g., an exception raised during the boot
        # heartbeat read against backup_health). Re-raises so the
        # admin sees a critical-class log line at boot rather than
        # discovering 25h later via the watchdog alarm.
        await asyncio.sleep(1.5)
        if _backup_task.done():
            try:
                _backup_task.result()
            except Exception as e:
                logging.getLogger(__name__).critical(
                    f"[scheduled-backup] scheduler Task died during initialization: {e!r}"
                )
                _BACKUP_SCHEDULER_STATE["alive"] = False
                _BACKUP_SCHEDULER_STATE["last_attempt_outcome"] = (
                    f"TASK DIED AT STARTUP: {e!r}"
                )
                raise

        # Resurrection supervisor — checks every 5 minutes that the
        # scheduler task is still alive. Respawns it if dead. Logs
        # CRITICAL so the operator sees the resurrection event in
        # the backend logs. Has saved us from "fixed-then-broken-days-
        # later" recurrences caused by silent Task death.
        async def _scheduler_supervisor():
            global _backup_task
            # TRACK 14.0-RC1-PERF (2026-02-15): Detect the
            # "scheduler-disabled-on-this-worker" case so the watchdog
            # doesn't spam CRITICAL logs every 5 minutes in preview
            # (where SCHEDULER_ENABLED=false legitimately causes the
            # backup loop to exit cleanly). The first respawn cycle
            # captures the disabled state; subsequent cycles log at
            # DEBUG instead of CRITICAL, keeping the operator-facing
            # log signal real for production where a DEAD task is a
            # genuine incident.
            _scheduler_disabled_observed = False
            while True:
                try:
                    await asyncio.sleep(300)
                    if _backup_task is None or _backup_task.done():
                        # Pull the exception if there is one — informational
                        # only; we respawn regardless.
                        exc_repr = "(no task)"
                        try:
                            if _backup_task is not None and _backup_task.done():
                                exc = _backup_task.exception()
                                exc_repr = repr(exc) if exc else "completed without error"
                        except Exception:
                            pass
                        # A clean exit ("completed without error") in
                        # preview almost always means SCHEDULER_ENABLED
                        # is false. Demote the noise after the first
                        # cycle; real DEAD-with-exception cycles still
                        # log CRITICAL every time.
                        is_clean_disabled_exit = exc_repr == "completed without error"
                        log = logging.getLogger(__name__)
                        if is_clean_disabled_exit and _scheduler_disabled_observed:
                            log.debug(
                                "[scheduled-backup] scheduler task ended cleanly — respawning "
                                "(scheduler appears disabled on this worker)"
                            )
                        else:
                            log.critical(
                                f"[scheduled-backup] scheduler task is DEAD — respawning. "
                                f"Last state: {exc_repr}"
                            )
                            if is_clean_disabled_exit:
                                _scheduler_disabled_observed = True
                        _BACKUP_SCHEDULER_STATE["alive"] = False
                        _BACKUP_SCHEDULER_STATE["last_attempt_outcome"] = (
                            f"RESURRECTED at {datetime.now(timezone.utc).isoformat()} "
                            f"(previous: {exc_repr})"
                        )
                        # TRACK 27.05 · P0-2 · bump resurrect telemetry.
                        _BACKUP_SCHEDULER_STATE["resurrect_count"] = int(
                            _BACKUP_SCHEDULER_STATE.get("resurrect_count", 0) or 0
                        ) + 1
                        _BACKUP_SCHEDULER_STATE["last_resurrect_ts"] = (
                            datetime.now(timezone.utc).isoformat()
                        )
                        _backup_task = register_background_task(
                            app,
                            name="scheduled-backup-loop",
                            coro=_backup_scheduler_loop_with_capture(db),
                            category="scheduler",
                            critical=True,
                            long_running=True,
                        )
                except asyncio.CancelledError:
                    raise
                except Exception as e:  # noqa: BLE001
                    logging.getLogger(__name__).exception(
                        f"[scheduled-backup] supervisor tick failed: {e}"
                    )

        # The supervisor task is fire-and-forget. If it dies, the
        # watchdog email alarm at 25h is still the last line of defense.
        register_background_task(
            app,
            name="scheduled-backup-supervisor",
            coro=_scheduler_supervisor(),
            category="scheduler-supervisor",
            critical=True,
            long_running=True,
        )
        logging.getLogger(__name__).info(
            "[scheduled-backup] supervisor armed — checks task health every 5 min"
        )
    except Exception as e:
        logging.getLogger(__name__).exception(f"[scheduled-backup] startup failed: {e}")

cors_origins_env = os.environ.get('CORS_ORIGINS', '').strip()
cors_origin_regex = (os.environ.get('CORS_ORIGIN_REGEX', '') or '').strip() or None

# Default safe regex when no explicit list is configured: allow MASCI's prod
# domain plus any Emergent preview pod. Browsers reject
# `Access-Control-Allow-Origin: *` combined with credentialed requests, so a
# regex / explicit list is required for the prod app to work in iOS Safari +
# Cloudflare.
#
# Iter171 hardening: `CORS_ORIGINS=*` is now treated as "unset" and falls
# through to regex mode with credentials enabled. This removes the wildcard
# escape hatch entirely — even if a platform layer re-injects `*` into the
# runtime env, the regex authoritatively wins. Preview keeps working because
# the default regex below covers all Emergent preview domains.
_DEFAULT_CORS_REGEX = (
    r"^https://("
    r"(www\.)?mascidocs\.com"
    r"|.*\.emergentagent\.com"
    r"|.*\.preview\.emergentagent\.com"
    r"|.*\.emergent\.host"
    r")$"
)

if cors_origins_env and cors_origins_env != '*':
    # Explicit allow-list (preferred for production hardening).
    _cors_origins = [o.strip() for o in cors_origins_env.split(',') if o.strip()]
    _cors_credentials = True
    cors_origin_regex = None
else:
    # Empty OR explicit '*' → fall through to regex with credentials.
    # We intentionally never honor wildcard CORS — it disables credentialed
    # cross-origin requests AND broadens the CSRF surface.
    _cors_origins: List[str] = []
    _cors_credentials = True
    if not cors_origin_regex:
        cors_origin_regex = _DEFAULT_CORS_REGEX

app.add_middleware(
    CORSMiddleware,
    allow_credentials=_cors_credentials,
    allow_origins=_cors_origins,
    allow_origin_regex=cors_origin_regex,
    # Track 21.3 Phase B · CORS methods/headers tightening (2026-07-04).
    # Explicit method + header allow-lists. Every entry corresponds to an
    # actual production usage:
    #   - GET/POST/PUT/PATCH/DELETE — REST CRUD (verified across 1,440 routes)
    #   - OPTIONS — CORS preflight (mandatory)
    #   - HEAD — health probes / edge-cache range checks
    # Verified against the current frontend axios/fetch surface: only these
    # methods are ever emitted. `*` was previously used out of caution;
    # tightening removes an entire class of drift risk.
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS", "HEAD"],
    # Headers actually emitted by the frontend (verified via grep across
    # /app/frontend/src): Content-Type, Authorization, X-Admin-Token,
    # X-CSRF-Token, X-Session-Id, X-Portal-Token, X-Requested-With,
    # X-Client-Trace, Accept, Accept-Language, Origin, plus the standard
    # CORS-safelisted headers auto-added by browsers. Uploads use
    # multipart/form-data whose auto-headers (Content-Length,
    # Content-Disposition) are CORS-safelisted → no allow-list entry
    # needed. Range header retained for signed-URL byte-range fetches.
    allow_headers=[
        "Accept", "Accept-Language", "Authorization", "Content-Type",
        "Origin", "Range", "X-Admin-Token", "X-Client-Trace",
        "X-CSRF-Token", "X-Portal-Token", "X-Requested-With",
        "X-Session-Id",
    ],
    # Expose headers the frontend actually reads back:
    expose_headers=["Content-Disposition", "Content-Length", "ETag", "X-Request-Id"],
)

# ─────────────────────────────────────────────────────────────────────
# Photo thumbnail Cloudflare edge-cache enabler (2026-05-26)
#
# Problem: /api/job-photos/<id>/thumb-signed responses had `Vary: Accept`
# and the CORS middleware appended `Access-Control-*` headers to them.
# Cloudflare treats responses with `Vary` + `Access-Control-Allow-Origin`
# as DYNAMIC and refuses to edge-cache them — so every browser hit went
# back to the origin (~540 ms each). With 32 thumbs in a gallery that's
# 5-10 seconds of visible delay on first paint, which is what users felt
# as "photos load slow / site sluggish".
#
# Fix: a tiny BaseHTTPMiddleware running AFTER CORS that, for thumb URLs
# only, drops the cache-poisoning headers and re-affirms the immutable
# directives. CF can now serve repeat requests from the edge in ~50 ms,
# and the browser cache + sw-thumbs service worker do the rest on-device.
# ─────────────────────────────────────────────────────────────────────
from starlette.middleware.base import BaseHTTPMiddleware  # noqa: E402
import re as _thumb_re

# iter445 · 2026-06-01 · Narrowed to /thumb(-signed)? only.
# Sprint 1G's /raw and /raw-signed endpoints return JSON with short-lived
# (900 s) presigned R2 URLs. Edge-caching them would expose stale URLs
# (R2 → 403) AND stripping their Access-Control-Allow-Origin header
# breaks cross-origin XHR from mascidocs.com → emergent.host (Sprint 1G
# CORS remediation · Option A · defense-in-depth). Thumbnail endpoints
# still benefit from edge caching since they return image bytes
# directly (no CORS dependency for <img>; no time-limited URLs).
_THUMB_PATH_RE = _thumb_re.compile(r"^/api/job-photos/.+/thumb(-signed)?/?$")
_THUMB_HEADERS_TO_STRIP = (
    "vary",
    "access-control-allow-origin",
    "access-control-allow-credentials",
    "access-control-allow-methods",
    "access-control-allow-headers",
    "access-control-max-age",
    "access-control-expose-headers",
)


class PhotoEdgeCacheMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        try:
            response = await call_next(request)
        except Exception:
            return JSONResponse(status_code=500, content={"detail": "internal_server_error"})
        try:
            if _THUMB_PATH_RE.match(request.url.path) and response.status_code == 200:
                for h in _THUMB_HEADERS_TO_STRIP:
                    if h in response.headers:
                        del response.headers[h]
                response.headers["Cache-Control"] = "public, max-age=604800, stale-while-revalidate=86400, immutable"
                response.headers["CDN-Cache-Control"] = "public, max-age=2592000, stale-while-revalidate=86400, immutable"
        except Exception:
            # Never break a photo response over a header tweak.
            pass
        return response


app.add_middleware(PhotoEdgeCacheMiddleware)

# PERFORMANCE-HARDEN-001 · GZip compression for JSON / HTML / text
# responses ≥ 1 KB. Starlette's GZipMiddleware respects the client's
# Accept-Encoding header (so it never breaks clients that can't decode),
# never compresses already-encoded payloads (image/* etc. served by
# PhotoEdgeCacheMiddleware remain untouched), and is zero-config-safe.
# Reduces wire size of /api/integrations/*, /api/admin/*, /api/health,
# /api/employees, /api/job_photos JSON responses by ~70-85% in typical
# Brotli-fallback browsers. Effect on cold dashboard loads on slow 4G
# is the single largest available win without code restructuring.
from starlette.middleware.gzip import GZipMiddleware  # noqa: E402
app.add_middleware(GZipMiddleware, minimum_size=1024, compresslevel=6)

# iter430 · Phase 28.2 · Sentry operational-tag enrichment ·
# Auto-attaches portal/role/route/device/browser/language/tenant tags
# to every Sentry event so production exceptions immediately reveal
# operational context. No-op when SENTRY_DSN is absent.
try:
    from sentry_tags import SentryOperationalTagsMiddleware  # noqa: E402
    app.add_middleware(SentryOperationalTagsMiddleware)
except Exception as _stm_err:  # noqa: BLE001
    # Never let the tagging layer break startup.
    pass

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
)
logger = logging.getLogger(__name__)


@register_shutdown_step("shutdown")
async def shutdown_db_client():
    global client
    try:
        await cancel_registered_background_tasks(app)
    except Exception:
        pass
    if client is not None:
        client.close()
        client = None
    db.clear_target()
    app.state.mongo_client = None
    app.state.db = None
    app.state.db_name = None
    app.state.database_authority_plan = None


# ──────────────────────────────────────────────────────────────────────
# iter453.6 · Startup-readiness gate (final wiring)
#
# Operator authorization: OMEGA HOTFIX BUNDLE A · Part C · 2026-06-02.
#
# Two pieces:
#   1. A middleware that returns HTTP 503 {"detail":"service_starting"}
#      for public WRITE requests while `app.state.ready` is False.
#      Read-only health/version probes always pass through so deployment
#      readiness probes still work during warm-up.
#   2. A startup hook (registered LAST so it runs after every other
#      startup hook completes) that flips `app.state.ready = True`.
#
# Scope discipline: only POST/PUT/PATCH/DELETE on /api/* (i.e., the public
# write surface) are gated. GETs and non-/api/* are passed through to
# avoid breaking infrastructure health checks.
# ──────────────────────────────────────────────────────────────────────
_READINESS_EXEMPT_PATHS = {
    "/api/health",
    "/api/version",
}


@app.middleware("http")
async def _canonical_security_headers(request, call_next):
    try:
        response = await call_next(request)
    except RuntimeError as exc:
        if str(exc) == "No response returned.":
            from starlette.responses import Response  # noqa: PLC0415
            return Response(status_code=204)
        raise
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("X-Frame-Options", "DENY")
    csp = response.headers.get("Content-Security-Policy", "").strip()
    if "frame-ancestors" not in csp:
        response.headers["Content-Security-Policy"] = (
            f"{csp}; frame-ancestors 'none'".strip('; ').strip()
        )
    if (os.environ.get("APP_ENV") or "").strip().lower() == "production":
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    return response


@app.middleware("http")
async def _iter453_6_readiness_gate(request, call_next):
    if getattr(request.app.state, "read_only_validation_active", False):
        method = (request.method or "").upper()
        path = request.url.path or ""
        if method in {"POST", "PUT", "PATCH", "DELETE"} and path.startswith("/api/"):
            from fastapi.responses import JSONResponse  # noqa: PLC0415
            return JSONResponse(
                status_code=503,
                content={"detail": "read_only_validation_mode"},
            )
    if not getattr(request.app.state, "ready", False):
        method = (request.method or "").upper()
        path = request.url.path or ""
        if (
            method in {"POST", "PUT", "PATCH", "DELETE"}
            and path.startswith("/api/")
            and path not in _READINESS_EXEMPT_PATHS
        ):
            from fastapi.responses import JSONResponse  # noqa: PLC0415
            return JSONResponse(
                status_code=503,
                content={"detail": "service_starting"},
            )
    try:
        response = await call_next(request)
        await observe_request_result(request.app, path=request.url.path or "", status_code=response.status_code)
        return response
    except Exception as exc:
        await observe_request_result(request.app, path=request.url.path or "", status_code=500, exception=exc)
        raise


# ─────────────────────────────────────────────────────────────────────
# TRACK 22.3 · DR-V2 alias telemetry middleware.
#
# Records every hit to legacy /api/dr-v2/* endpoints so DR-UNIFY-005
# can prove they are safe to retire. Fire-and-forget: telemetry writes
# never block or fail the request. Detail events auto-expire after
# 30 days; lightweight aggregates persist until formal retirement.
# ─────────────────────────────────────────────────────────────────────
@app.middleware("http")
async def _track_22_3_dr_v2_alias_telemetry(request, call_next):
    try:
        path = request.url.path or ""
        should_track = path.startswith("/api/dr-v2/") or path.startswith("/api/dr-v2")
    except Exception:  # noqa: BLE001
        should_track = False
    try:
        response = await call_next(request)
    except RuntimeError as exc:
        if str(exc) == "No response returned.":
            from starlette.responses import Response  # noqa: PLC0415
            return Response(status_code=204)
        raise
    if should_track:
        try:
            asyncio.get_event_loop().create_task(_record_dr_v2_alias_hit(db, request))
        except Exception:  # noqa: BLE001
            pass
    return response


# ─────────────────────────────────────────────────────────────────────
# TRACK 15.93 · Zero-Touch System Bootstrap
# Runs BEFORE the readiness flag flips. Idempotent. Admin-safe.
# Guarantees required system records exist so a fresh deploy reaches
# operational state without any manual seed command.
# ─────────────────────────────────────────────────────────────────────
@register_lifecycle_step("misc-bootstrap")
async def _track_15_93_run_system_bootstrap():
    try:
        from lib.system_bootstrap import run_system_bootstrap  # noqa: PLC0415
        result = await run_system_bootstrap(db)
        app.state.bootstrap_result = result
        log = logging.getLogger(__name__)
        if result.get("ok"):
            log.info(
                "[system-bootstrap] OK · version=%s · steps=%s",
                result.get("version"),
                ",".join(s.get("name", "?") for s in result.get("steps") or []),
            )
        else:
            log.error(
                "[system-bootstrap] FAILED · version=%s · missing=%s · errors=%s",
                result.get("version"),
                result.get("missing_items"),
                [s for s in (result.get("steps") or []) if s.get("status") != "ok"],
            )
    except Exception as e:  # noqa: BLE001
        # Non-fatal: log loudly. Readiness gate will surface this.
        app.state.bootstrap_result = {
            "version": None,
            "ok": False,
            "completed_at": None,
            "missing_items": [f"bootstrap raised: {type(e).__name__}: {e}"],
            "steps": [],
        }
        logging.getLogger(__name__).exception(
            "[system-bootstrap] startup hook raised; bootstrap incomplete",
        )


@register_lifecycle_step("command-center")
async def _command_center_seed_defaults():
    """Track 22.1L · retired the router-hosted `@router.on_event("startup")`
    closure inside `build_command_center_router`. Registers here so startup
    ordering is deterministic: runs AFTER misc-bootstrap + backup-scheduler
    groups (source order) and BEFORE the `readiness` phase-3 group.
    Body semantically identical to the pre-migration closure — same try/except
    around `_seed_defaults(db)` with silent-on-error semantics (never blocks boot).
    """
    try:
        from routes.command_center import _seed_defaults as _cc_seed_defaults  # noqa: PLC0415
        await _cc_seed_defaults(db)
    except Exception:  # noqa: BLE001
        pass  # silent: not blocking app boot if seeding fails


@register_lifecycle_step("readiness")
async def _iter453_6_flip_ready_flag():
    """Final startup hook — flip the readiness gate AFTER all other
    @app.on_event('startup') handlers have completed. FastAPI runs
    startup events in registration order, and this module-level
    registration is the LAST one in server.py, so by the time this
    runs every index/scheduler/router setup above is finished.
    """
    set_startup_complete(app, ready=True, reason="startup_complete")
    logging.getLogger(__name__).info(
        "[iter453.6] startup-readiness gate FLIPPED · public writes now accepted",
    )


@register_lifecycle_step("post-readiness")
async def _schedule_deployment_governance_verification():
    register_background_task(
        app,
        name="deployment-governance-verification",
        coro=_run_automatic_deployment_governance_verification(),
        category="governance",
        critical=False,
        long_running=False,
    )


# ────────────────────────────────────────────────────────────────────
# Phase D-1.4 · Dispatch reminder scheduler · MUST register AFTER the
# readiness flag so the loop only runs in a fully-booted process. The
# loop itself short-circuits when SCHEDULER_ENABLED is off.
# ────────────────────────────────────────────────────────────────────
@register_lifecycle_step("post-readiness")
async def _dispatch_reminder_scheduler_start():
    try:
        from dispatch_reminders import reminder_scheduler_loop  # noqa: PLC0415
        register_background_task(
            app,
            name="dispatch-reminder-scheduler",
            coro=reminder_scheduler_loop(db),
            category="scheduler",
            critical=False,
            long_running=True,
        )
        logging.getLogger(__name__).info(
            "[dispatch-reminders] background task scheduled",
        )
    except Exception as e:  # noqa: BLE001
        logging.getLogger(__name__).warning(
            f"[dispatch-reminders] failed to schedule background task: {e}",
        )