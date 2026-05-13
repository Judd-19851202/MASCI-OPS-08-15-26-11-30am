"""Shop Activity Feed + Equipment Parts Catalog routes.

Extracted from server.py 2026-04-28 as the first proof-of-pattern for the
larger server.py refactor (P1 backlog). All endpoints here use either
`require_shop_or_admin` (read + per-unit edit) or `require_admin` (delete +
bulk upload).

Endpoints registered:
    GET    /api/shop/activity                         (shop or admin)
    GET    /api/equipment-parts                       (shop or admin)
    GET    /api/equipment-parts/{unit_number}         (shop or admin)
    PUT    /api/equipment-parts/{unit_number}         (shop or admin)
    DELETE /api/equipment-parts/{unit_number}         (admin only)
    GET    /api/admin/equipment-parts/status          (admin only)
    POST   /api/admin/equipment-parts/upload          (admin only — xlsx/csv)
    POST   /api/equipment-parts/order                 (shop or admin — emails)
"""
from __future__ import annotations

import asyncio
import csv
import io
import logging
import os
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel

logger = logging.getLogger(__name__)

PART_CATEGORIES = ["filters", "cutting_edges", "wiper_blades", "tires", "other_wear_items"]


def _empty_parts_doc(unit_number: str) -> Dict[str, Any]:
    return {
        "unit_number": unit_number,
        "filters": [],
        "cutting_edges": [],
        "wiper_blades": [],
        "tires": [],
        "other_wear_items": [],
        "updated_at": "",
        "updated_by": "",
    }


class EquipmentPartsPayload(BaseModel):
    filters: Optional[List[Dict[str, Any]]] = None
    cutting_edges: Optional[List[Dict[str, Any]]] = None
    wiper_blades: Optional[List[Dict[str, Any]]] = None
    tires: Optional[List[Dict[str, Any]]] = None
    other_wear_items: Optional[List[Dict[str, Any]]] = None
    updated_by: Optional[str] = ""


class PartsOrderItem(BaseModel):
    name: str
    part_number: Optional[str] = ""
    qty: Optional[str] = ""
    category: Optional[str] = ""
    notes: Optional[str] = ""


class PartsOrderRequest(BaseModel):
    unit_number: str
    equipment_label: Optional[str] = ""
    requested_by: str
    send_to: List[str]
    cc: Optional[List[str]] = []
    additional_notes: Optional[str] = ""
    items: List[PartsOrderItem]


def register_shop_parts_routes(api_router: APIRouter, db, require_admin, require_shop_or_admin):
    """Attach Shop Activity + Equipment Parts endpoints to the shared router."""

    # ---------- Shop Activity Feed ----------
    @api_router.get("/shop/activity")
    async def shop_activity(
        limit: int = 20,
        _: bool = Depends(require_shop_or_admin),
    ):
        """Flatten every shop_signoff entry across every equipment inspection,
        newest first. Powers the credibility log on the Shop console and the
        global view on /admin/equipment.
        """
        limit = max(1, min(100, int(limit)))
        cursor = db.equipment_inspections.find(
            {"shop_signoffs.0": {"$exists": True}},
            {"_id": 0, "id": 1, "equipment_type": 1, "equipment_unit": 1,
             "project_name": 1, "project_number": 1, "shop_signoffs": 1},
        )
        out: List[Dict[str, Any]] = []
        async for d in cursor:
            for s in (d.get("shop_signoffs") or []):
                out.append({
                    "inspection_id": d.get("id"),
                    "equipment_type": d.get("equipment_type", ""),
                    "equipment_unit": d.get("equipment_unit", ""),
                    "project_name": d.get("project_name", ""),
                    "project_number": d.get("project_number", ""),
                    "section": s.get("section", ""),
                    "item": s.get("item", ""),
                    "signed_by": s.get("signed_by", ""),
                    "signed_at": s.get("signed_at", ""),
                    "action_taken": s.get("action_taken", ""),
                    "notes": s.get("notes", ""),
                })
        out.sort(key=lambda r: r.get("signed_at", ""), reverse=True)
        return {"items": out[:limit], "count": len(out)}

    # ---------- Equipment Parts Catalog ----------
    @api_router.get("/equipment-parts")
    async def list_equipment_parts(_: bool = Depends(require_shop_or_admin)):
        cursor = db.equipment_parts.find({}, {"_id": 0})
        items = [d async for d in cursor]
        return {"items": items, "count": len(items)}

    @api_router.get("/equipment-parts/{unit_number}")
    async def get_equipment_parts(unit_number: str, _: bool = Depends(require_shop_or_admin)):
        if not unit_number.strip():
            raise HTTPException(status_code=400, detail="unit_number required")
        doc = await db.equipment_parts.find_one({"unit_number": unit_number}, {"_id": 0})
        return doc or _empty_parts_doc(unit_number)

    @api_router.put("/equipment-parts/{unit_number}")
    async def upsert_equipment_parts(
        unit_number: str,
        payload: EquipmentPartsPayload,
        _: bool = Depends(require_shop_or_admin),
    ):
        if not unit_number.strip():
            raise HTTPException(status_code=400, detail="unit_number required")
        doc = _empty_parts_doc(unit_number)
        body = payload.model_dump(exclude_unset=True)
        for cat in PART_CATEGORIES:
            if cat in body and isinstance(body[cat], list):
                doc[cat] = body[cat]
        doc["updated_at"] = datetime.now(timezone.utc).isoformat()
        doc["updated_by"] = (body.get("updated_by") or "").strip()
        await db.equipment_parts.update_one(
            {"unit_number": unit_number},
            {"$set": doc},
            upsert=True,
        )
        return doc

    @api_router.delete("/equipment-parts/{unit_number}")
    async def delete_equipment_parts(unit_number: str, _: bool = Depends(require_admin)):
        res = await db.equipment_parts.delete_one({"unit_number": unit_number})
        return {"deleted": res.deleted_count}

    @api_router.get("/admin/equipment-parts/status")
    async def equipment_parts_status(_: bool = Depends(require_admin)):
        count = await db.equipment_parts.count_documents({})
        last = await db.equipment_parts.find({}, {"_id": 0, "updated_at": 1}).sort("updated_at", -1).to_list(1)
        return {
            "count": count,
            "last_updated": (last[0].get("updated_at") if last else ""),
        }

    @api_router.post("/admin/equipment-parts/upload")
    async def upload_equipment_parts(
        file: UploadFile = File(...),
        _: bool = Depends(require_admin),
    ):
        """Bulk upload the parts catalog. Accepts .xlsx or .csv with columns:

           Unit Number | Category | Name | Part Number | Qty | Size | Position | Ply | Brand | Notes

        Category must be one of: filters, cutting_edges, wiper_blades, tires, other_wear_items.
        Each row appends a part to the corresponding category list for that unit.
        """
        raw = await file.read()
        fname = (file.filename or "").lower()
        rows: List[Dict[str, Any]] = []

        def _norm_key(k: str) -> str:
            return (k or "").strip().lower().replace(" ", "_").replace("-", "_")

        if fname.endswith(".csv"):
            text = raw.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            for r in reader:
                rows.append({_norm_key(k): (v or "").strip() for k, v in r.items()})
        elif fname.endswith(".xlsx") or fname.endswith(".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
                ws = wb.active
                headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    d = {_norm_key(headers[i] if i < len(headers) else f"col{i}"): str(row[i] or "").strip() for i in range(len(row))}
                    rows.append(d)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Could not parse xlsx: {e}")
        else:
            raise HTTPException(status_code=400, detail="Upload .xlsx or .csv only")

        def _norm_cat(c: str) -> Optional[str]:
            c = _norm_key(c)
            if c in PART_CATEGORIES:
                return c
            aliases = {
                "filter": "filters", "cutting_edge": "cutting_edges",
                "wiper_blade": "wiper_blades", "wipers": "wiper_blades",
                "tire": "tires", "other": "other_wear_items",
                "wear_items": "other_wear_items", "wear": "other_wear_items",
            }
            return aliases.get(c)

        by_unit: Dict[str, Dict[str, List[Dict[str, Any]]]] = {}
        skipped = 0
        for r in rows:
            unit = (r.get("unit_number") or r.get("unit") or "").strip()
            cat = _norm_cat(r.get("category") or "")
            if not unit or not cat:
                skipped += 1
                continue
            entry: Dict[str, Any] = {
                "name": r.get("name") or "",
                "part_number": r.get("part_number") or r.get("part") or "",
                "qty": r.get("qty") or r.get("quantity") or "",
                "notes": r.get("notes") or "",
            }
            if cat == "wiper_blades":
                entry["size"] = r.get("size") or ""
            if cat == "tires":
                entry["size"] = r.get("size") or ""
                entry["ply"] = r.get("ply") or ""
                entry["brand"] = r.get("brand") or ""
                entry["position"] = r.get("position") or ""
            if not entry["name"] and not entry["part_number"]:
                skipped += 1
                continue
            by_unit.setdefault(unit, {c: [] for c in PART_CATEGORIES})[cat].append(entry)

        written = 0
        now = datetime.now(timezone.utc).isoformat()
        for unit, cats in by_unit.items():
            existing = await db.equipment_parts.find_one({"unit_number": unit}, {"_id": 0}) or _empty_parts_doc(unit)
            for c in PART_CATEGORIES:
                existing[c] = cats[c]
            existing["updated_at"] = now
            existing["updated_by"] = "admin upload"
            await db.equipment_parts.update_one(
                {"unit_number": unit},
                {"$set": existing},
                upsert=True,
            )
            written += 1

        return {
            "ok": True,
            "units_written": written,
            "rows_total": len(rows),
            "rows_skipped": skipped,
        }

    @api_router.post("/equipment-parts/order")
    async def email_parts_order(
        payload: PartsOrderRequest,
        _: bool = Depends(require_shop_or_admin),
    ):
        """Email a one-click parts order list to the configured address(es)."""
        if not payload.send_to:
            raise HTTPException(status_code=400, detail="send_to email required")
        if not payload.items:
            raise HTTPException(status_code=400, detail="No items to order")
        if not payload.requested_by.strip():
            raise HTTPException(status_code=400, detail="requested_by is required")

        api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
        if not api_key:
            raise HTTPException(status_code=503, detail="Email not configured (RESEND_API_KEY missing)")

        rows_html = "".join(
            f"<tr>"
            f"<td style='padding:6px 10px;border:1px solid #ccc;font-family:monospace'>{(it.part_number or '—')}</td>"
            f"<td style='padding:6px 10px;border:1px solid #ccc'>{it.name}</td>"
            f"<td style='padding:6px 10px;border:1px solid #ccc;text-align:center'>{(it.qty or '—')}</td>"
            f"<td style='padding:6px 10px;border:1px solid #ccc;color:#64748b'>{(it.category or '—')}</td>"
            f"<td style='padding:6px 10px;border:1px solid #ccc;color:#64748b'>{(it.notes or '')}</td>"
            f"</tr>"
            for it in payload.items
        )
        html = f"""
        <div style='font-family:Arial,sans-serif;max-width:680px'>
          <h2 style='color:#C8102E;margin-bottom:4px'>MASCI Shop — Parts Order</h2>
          <div style='font-family:monospace;font-size:11px;color:#64748b;letter-spacing:0.15em;text-transform:uppercase'>Field-mechanic request</div>
          <table style='margin-top:12px;border-collapse:collapse;font-size:14px'>
            <tr><td style='padding:4px 0;color:#64748b'>Unit:</td><td style='padding:4px 0;font-weight:700'>{payload.unit_number}{(' · ' + payload.equipment_label) if payload.equipment_label else ''}</td></tr>
            <tr><td style='padding:4px 0;color:#64748b'>Requested by:</td><td style='padding:4px 0'>{payload.requested_by}</td></tr>
            <tr><td style='padding:4px 0;color:#64748b'>Time:</td><td style='padding:4px 0'>{datetime.now(timezone.utc).isoformat()}</td></tr>
          </table>
          <h3 style='margin-top:18px;border-bottom:2px solid #0f172a;padding-bottom:4px'>Parts requested ({len(payload.items)})</h3>
          <table style='border-collapse:collapse;width:100%;font-size:13px;margin-top:8px'>
            <thead><tr style='background:#0f172a;color:white'>
              <th style='padding:6px 10px;text-align:left'>Part #</th>
              <th style='padding:6px 10px;text-align:left'>Name</th>
              <th style='padding:6px 10px;text-align:center'>Qty</th>
              <th style='padding:6px 10px;text-align:left'>Category</th>
              <th style='padding:6px 10px;text-align:left'>Notes</th>
            </tr></thead>
            <tbody>{rows_html}</tbody>
          </table>
          {f"<p style='margin-top:14px'><b>Notes:</b> {payload.additional_notes}</p>" if payload.additional_notes else ""}
          <p style='margin-top:18px;color:#64748b;font-size:12px'>Sent automatically by the MASCI Shop console.</p>
        </div>
        """

        try:
            import resend  # noqa: E402
            resend.api_key = api_key
            sender_email = os.environ.get("SENDER_EMAIL", "onboarding@resend.dev")
            params = {
                "from": f"MASCI Operations Platform <{sender_email}>",
                "to": payload.send_to,
                "subject": f"[MASCI] Parts Order · {payload.unit_number} · {len(payload.items)} item(s)",
                "html": html,
            }
            if payload.cc:
                params["cc"] = payload.cc
            reply_to = os.environ.get("REPLY_TO_EMAIL", "").strip()
            if reply_to:
                params["reply_to"] = reply_to
            result = await asyncio.to_thread(resend.Emails.send, params)
            return {"ok": True, "resend_id": (result or {}).get("id"), "items": len(payload.items)}
        except Exception as e:
            logger.warning(f"parts-order email failed: {e}")
            raise HTTPException(status_code=502, detail=f"Email failed: {e}")
