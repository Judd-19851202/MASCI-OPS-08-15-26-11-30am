"""
fire_ext_bulk_import.py — Iter134. Bulk-import fire extinguisher
inventory from Excel (.xlsx) or CSV.

Two-step preview/commit pattern matches the Integration Center Mapping
Wizard. The preview NEVER writes; commit applies the previewed plan.

Matching priority (per row):
  1. extinguisher_id   (case-insensitive)
  2. serial_number     (case-insensitive)
  3. truck + location  (composite key fallback)

Match → update. No match → create. Never duplicate.

Endpoints (mounted under /api):
  GET  /safety/fire-extinguishers/import/template     — CSV header template
  POST /safety/fire-extinguishers/import/preview      — multipart .csv/.xlsx → preview plan
  POST /safety/fire-extinguishers/import/commit       — apply previously previewed plan
  GET  /safety/fire-extinguishers/import/history      — last N import runs (read-only)
"""
from __future__ import annotations

import csv
import io
import logging
import re
import uuid
from datetime import datetime, timezone
from typing import Any, Callable, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

# Map of header aliases → canonical field. Case-insensitive whitespace-collapsed.
COLUMN_ALIASES: Dict[str, str] = {
    "extinguisher id": "unit_id",
    "extinguisher_id": "unit_id",
    "unit id": "unit_id",
    "unit_id": "unit_id",
    "id": "unit_id",
    "serial number": "serial_number",
    "serial": "serial_number",
    "serial_number": "serial_number",
    "type": "type",
    "size": "size",
    "location": "location_value",
    "location value": "location_value",
    "assigned job": "project_number",
    "assigned project": "project_number",
    "project": "project_number",
    "project number": "project_number",
    "project_number": "project_number",
    "assigned truck": "truck",
    "assigned equipment": "truck",
    "truck": "truck",
    "equipment": "truck",
    "inspection date": "last_inspection_date",
    "last inspection": "last_inspection_date",
    "next due date": "next_due_date",
    "next due": "next_due_date",
    "status": "last_status",
    "pass fail": "last_status",
    "pass/fail": "last_status",
    "pass_fail": "last_status",
    "deficiencies": "deficiencies",
    "corrective action required": "corrective_action_required",
    "corrective action": "corrective_action_required",
    "notes": "notes",
}

VALID_STATUSES = {"Pass", "Fail", "Needs Service", "Out of Service", ""}


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _norm_status(v: Any) -> str:
    s = str(v or "").strip().lower()
    if s in ("pass", "p", "ok", "good"):
        return "Pass"
    if s in ("fail", "failed", "f", "no", "bad"):
        return "Fail"
    if "service" in s or "repair" in s:
        return "Needs Service" if "needs" in s or "repair" in s else "Out of Service"
    if "out" in s:
        return "Out of Service"
    return ""


def _parse_date(v: Any) -> Optional[str]:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    if not s:
        return None
    # Try the common formats people put in Excel
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%d-%b-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _parse_file(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    """Returns a list of dicts with raw header keys preserved."""
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook  # noqa: PLC0415
        except ImportError as e:
            raise HTTPException(status_code=500, detail=f"openpyxl missing: {e}") from e
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or "").strip() for h in rows[0]]
        out = []
        for r in rows[1:]:
            if all((c is None or str(c).strip() == "") for c in r):
                continue
            out.append({headers[i]: r[i] if i < len(r) else "" for i in range(len(headers))})
        return out
    if name.endswith(".csv") or name.endswith(".txt"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]
    raise HTTPException(status_code=400, detail="Unsupported file type — use .csv or .xlsx")


def _map_row(raw: Dict[str, Any]) -> Dict[str, Any]:
    """Translate user headers to canonical schema fields."""
    out: Dict[str, Any] = {}
    for header, value in raw.items():
        canonical = COLUMN_ALIASES.get(_norm(str(header or "")))
        if canonical:
            out[canonical] = value
    return out


def _validate(row: Dict[str, Any]) -> List[str]:
    errors: List[str] = []
    unit_id = (row.get("unit_id") or "").strip() if isinstance(row.get("unit_id"), str) else row.get("unit_id")
    serial = (row.get("serial_number") or "").strip() if isinstance(row.get("serial_number"), str) else row.get("serial_number")
    if not unit_id and not serial:
        errors.append("Missing Extinguisher ID and Serial Number — need at least one")
    loc = (row.get("location_value") or "").strip() if isinstance(row.get("location_value"), str) else row.get("location_value")
    truck = (row.get("truck") or "").strip() if isinstance(row.get("truck"), str) else row.get("truck")
    if not loc and not truck:
        errors.append("Missing Location and Assigned Truck/Equipment — need at least one")
    # Date sanity
    for key in ("last_inspection_date", "next_due_date"):
        val = row.get(key)
        if val not in (None, "") and _parse_date(val) is None:
            errors.append(f"Bad date format in {key}: {val!r}")
    # Status sanity
    if row.get("last_status") not in (None, ""):
        if _norm_status(row.get("last_status")) == "" and str(row.get("last_status")).strip() != "":
            errors.append(f"Unknown status value: {row.get('last_status')!r}")
    return errors


class CommitPayload(BaseModel):
    preview_id: str = Field(..., min_length=8, max_length=80)


def build_fire_import_router(db, require_safety_token: Callable) -> APIRouter:
    router = APIRouter(
        prefix="/api/safety/fire-extinguishers/import",
        tags=["fire-ext-import"],
    )

    @router.get("/template")
    async def import_template():
        """Downloadable CSV template + 1 example row."""
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow([
            "Extinguisher ID", "Serial Number", "Type", "Size",
            "Location", "Assigned Truck", "Project Number",
            "Inspection Date", "Next Due Date", "Status",
            "Deficiencies", "Corrective Action Required", "Notes",
        ])
        w.writerow([
            "FE-001", "ABCD12345", "ABC", "10 lb",
            "Cab — passenger side", "Truck 12", "P-2026-001",
            "2026-01-15", "2027-01-15", "Pass",
            "", "", "Annual inspection complete",
        ])
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="fire_ext_import_template.csv"'},
        )

    @router.post("/preview", dependencies=[Depends(require_safety_token)])
    async def import_preview(
        request: Request,
        file: UploadFile = File(...),
    ):
        raw = await file.read()
        if not raw:
            raise HTTPException(status_code=400, detail="Empty file")
        if len(raw) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File too large — 10 MB limit")

        try:
            raw_rows = _parse_file(raw, file.filename or "")
        except HTTPException:
            raise
        except Exception as e:  # noqa: BLE001
            raise HTTPException(status_code=400, detail=f"Could not parse file: {e}") from e

        if not raw_rows:
            raise HTTPException(status_code=400, detail="No data rows found")

        # Build a quick lookup of all existing extinguishers for O(1) matching
        existing = await db.fire_extinguishers.find(
            {}, {"_id": 0, "id": 1, "unit_id": 1, "serial_number": 1, "truck": 1, "location_value": 1}
        ).to_list(20000)
        by_unit = {(d.get("unit_id") or "").lower(): d for d in existing if d.get("unit_id")}
        by_serial = {(d.get("serial_number") or "").lower(): d for d in existing if d.get("serial_number")}
        by_truck_loc = {
            f"{(d.get('truck') or '').lower()}|{(d.get('location_value') or '').lower()}": d
            for d in existing if d.get("truck") and d.get("location_value")
        }

        seen_in_file: Dict[str, int] = {}  # for duplicate detection inside the file
        rows: List[Dict[str, Any]] = []
        create = update = skip = 0

        for idx, raw_row in enumerate(raw_rows, start=2):  # row 1 is header
            mapped = _map_row(raw_row)
            errors = _validate(mapped)

            unit_key = (str(mapped.get("unit_id") or "")).strip().lower()
            serial_key = (str(mapped.get("serial_number") or "")).strip().lower()
            dedup_key = unit_key or serial_key
            if dedup_key:
                seen_in_file[dedup_key] = seen_in_file.get(dedup_key, 0) + 1
                if seen_in_file[dedup_key] > 1:
                    errors.append("Duplicate within file — first occurrence will win")

            # Decide create vs update via match priority
            match = None
            match_reason = None
            if unit_key and unit_key in by_unit:
                match = by_unit[unit_key]
                match_reason = "matched on Extinguisher ID"
            elif serial_key and serial_key in by_serial:
                match = by_serial[serial_key]
                match_reason = "matched on Serial Number"
            else:
                truck_key = f"{str(mapped.get('truck') or '').lower()}|{str(mapped.get('location_value') or '').lower()}"
                if truck_key.strip("|") and truck_key in by_truck_loc:
                    match = by_truck_loc[truck_key]
                    match_reason = "matched on Truck + Location"

            action = "skip" if errors else ("update" if match else "create")
            if action == "create":
                create += 1
            elif action == "update":
                update += 1
            else:
                skip += 1

            rows.append({
                "row_number": idx,
                "action": action,
                "match_reason": match_reason,
                "match_id": (match or {}).get("id"),
                "data": {**mapped},
                "errors": errors,
            })

        preview_id = str(uuid.uuid4())
        now_dt = datetime.now(timezone.utc)
        await db.fire_ext_import_runs.insert_one({
            "id": preview_id,
            "at": now_dt,
            "iso_at": now_dt.replace(microsecond=0).isoformat(),
            "status": "preview",
            "file_name": file.filename,
            "uploaded_by": (request.headers.get("X-Safety-Email") or "safety-user"),
            "total_rows": len(rows),
            "to_create": create,
            "to_update": update,
            "to_skip": skip,
            "rows": rows,
        })
        return {
            "preview_id": preview_id,
            "file_name": file.filename,
            "total_rows": len(rows),
            "to_create": create,
            "to_update": update,
            "to_skip": skip,
            "rows": rows,
        }

    @router.post("/commit", dependencies=[Depends(require_safety_token)])
    async def import_commit(payload: CommitPayload):
        run = await db.fire_ext_import_runs.find_one({"id": payload.preview_id}, {"_id": 0})
        if not run:
            raise HTTPException(status_code=404, detail="Preview not found or expired")
        if run.get("status") != "preview":
            raise HTTPException(status_code=409, detail=f"Preview already {run.get('status')}")

        created = updated = skipped = 0
        errors: List[str] = []
        now_iso = datetime.now(timezone.utc).replace(microsecond=0).isoformat()

        for row in run.get("rows", []):
            action = row["action"]
            data = row.get("data") or {}
            if action == "skip":
                skipped += 1
                continue
            # Normalize fields
            doc = {
                "unit_id":             (str(data.get("unit_id") or "")).strip(),
                "serial_number":       (str(data.get("serial_number") or "")).strip(),
                "type":                (str(data.get("type") or "")).strip(),
                "size":                (str(data.get("size") or "")).strip(),
                "location_kind":       "truck" if data.get("truck") else "facility",
                "location_value":      (str(data.get("location_value") or data.get("truck") or "")).strip(),
                "truck":               (str(data.get("truck") or "")).strip(),
                "project_number":      (str(data.get("project_number") or "")).strip(),
                "last_inspection_date": _parse_date(data.get("last_inspection_date")),
                "next_due_date":       _parse_date(data.get("next_due_date")),
                "last_status":         _norm_status(data.get("last_status")),
                "deficiencies":        (str(data.get("deficiencies") or "")).strip(),
                "corrective_action_required": (str(data.get("corrective_action_required") or "")).strip(),
                "notes":               (str(data.get("notes") or "")).strip(),
                "updated_at":          now_iso,
            }
            try:
                if action == "update" and row.get("match_id"):
                    res = await db.fire_extinguishers.update_one(
                        {"id": row["match_id"]}, {"$set": doc},
                    )
                    if res.modified_count or res.matched_count:
                        updated += 1
                    else:
                        skipped += 1
                else:
                    doc["id"] = str(uuid.uuid4())
                    doc["created_at"] = now_iso
                    doc["inspections"] = []
                    await db.fire_extinguishers.insert_one(doc)
                    created += 1
            except Exception as e:  # noqa: BLE001
                skipped += 1
                errors.append(f"Row {row['row_number']}: {e!s}")

        await db.fire_ext_import_runs.update_one(
            {"id": payload.preview_id},
            {"$set": {
                "status": "committed",
                "committed_at": now_iso,
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "commit_errors": errors,
            }},
        )
        return {
            "ok": True,
            "preview_id": payload.preview_id,
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
        }

    @router.get("/history", dependencies=[Depends(require_safety_token)])
    async def import_history(limit: int = 20):
        limit = min(max(int(limit or 20), 1), 100)
        rows: List[Dict[str, Any]] = []
        cursor = db.fire_ext_import_runs.find(
            {}, {"_id": 0, "rows": 0},  # exclude the heavy row blob
        ).sort("iso_at", -1).limit(limit)
        async for r in cursor:
            rows.append(r)
        return {"limit": limit, "rows": rows}

    return router


__all__ = ["build_fire_import_router"]
