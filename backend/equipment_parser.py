"""
MASCI Equipment List parser
---------------------------
Reads an Equipment List.xlsx (Louis sheet — master) and produces normalized
equipment items: { unit_number, year, make_model, plate, vin_serial_number,
comments, company, category, preop_equipment_type, display_label }.

Used by both the startup seed and the admin upload endpoint so the parsing
logic stays in one place.
"""
from __future__ import annotations

import io
import re
from collections import Counter
from typing import Any, Dict, List, Optional

import openpyxl


# Unit-number prefix → category (most reliable source of truth in the sheet)
PREFIX_TO_CAT: Dict[str, str] = {
    "DPT": "Dump Trucks",
    "EXC": "Excavators",
    "DZ":  "Dozers",
    "LDR": "Loaders",
    "BH":  "Backhoes",
    "SKD": "Skid Steers",
    "SWP": "Sweepers",
    "RL":  "Rollers",
    "PX":  "Paving Equipment",
    "TRL": "Trailers",
    "TTR": "Tractor Trailer Trucks",
    "WT":  "Water Trucks",
    "SER": "Service Trucks",
    "PKU": "Pickup Trucks",
    "FBT": "Flatbed Trucks",
    "RG":  "Road Graders",
    "LT":  "Light Towers",
    "MIC": "Misc Equipment",
}

# Make/model keyword → category fallback when no prefix.
KEYWORD_TO_CAT: List[tuple] = [
    ("excavator", "Excavators"),
    ("dozer",     "Dozers"),
    ("backhoe",   "Backhoes"),
    ("loader",    "Loaders"),
    ("skid steer","Skid Steers"),
    ("sweeper",   "Sweepers"),
    ("roller",    "Rollers"),
    ("paver",     "Paving Equipment"),
    ("milling",   "Milling Machines"),
    ("crusher",   "Crushers"),
    ("compactor", "Compactors"),
    ("forklift",  "Forklifts"),
    ("fork lift", "Forklifts"),
    ("telehandler","Telehandlers"),
    ("manlift",   "Man Lifts"),
    ("man lift",  "Man Lifts"),
    ("boom lift", "Boom Lifts"),
    ("scissor",   "Scissor Lifts"),
    ("light tower","Light Towers"),
    ("lighttower","Light Towers"),
    ("generator", "Generators"),
    ("welder",    "Welders"),
    ("compressor","Air Compressors"),
    ("pump",      "Pumps"),
    ("trailer",   "Trailers"),
    ("dump truck","Dump Trucks"),
    ("water truck","Water Trucks"),
    ("service truck","Service Trucks"),
    ("pickup",    "Pickup Trucks"),
    ("tundra",    "Pickup Trucks"),
    ("silverado", "Pickup Trucks"),
    ("ram 1500",  "Pickup Trucks"),
    ("ram 2500",  "Pickup Trucks"),
    ("flatbed",   "Flatbed Trucks"),
    ("grader",    "Road Graders"),
    ("mower",     "Mowers"),
    ("storage",   "Storage / Containers"),
    ("container", "Storage / Containers"),
    ("scotsman",  "Storage / Containers"),
    ("modspace",  "Storage / Containers"),
    ("garmin",    "GPS / Electronics"),
    ("mack ",     "Dump Trucks"),
    ("kenworth",  "Dump Trucks"),
    ("peterbilt", "Tractor Trailer Trucks"),
    ("freightliner","Service Trucks"),
    ("attachment","Attachments"),
    ("bucket",    "Attachments"),
    ("hammer",    "Attachments"),
    ("auger",     "Attachments"),
]

# Category label normalizations
NORMALIZE_CAT: Dict[str, str] = {
    "Misc. Equipment": "Misc Equipment",
    "Misc. Trucks": "Misc Trucks",
    "Light Tower": "Light Towers",
    "Generator": "Generators",
}

# MASCI fleet category → existing Pre-Op `EQUIPMENT_TYPES` value.
PREOP_TYPE_MAP: Dict[str, str] = {
    "Excavators": "Excavator",
    "Dozers": "Dozer",
    "Loaders": "Loader",
    "Backhoes": "Backhoe",
    "Skid Steers": "Skid Steer",
    "Sweepers": "Broom",
    "Rollers": "Steel Drum Asphalt Roller",
    "Paving Equipment": "Paver",
    "Milling Machines": "Asphalt Milling Machine",
    "Trailers": "Other",
    "Tractor Trailer Trucks": "Haul Truck",
    "Water Trucks": "Water Truck",
    "Service Trucks": "Haul Truck",
    "Pickup Trucks": "Other",
    "Flatbed Trucks": "Haul Truck",
    "Dump Trucks": "Haul Truck",
    "Misc Trucks": "Haul Truck",
    "Supervisor / Mgmt Trucks": "Other",
    "Road Graders": "Motor Grader",
    "Light Towers": "Other",
    "Generators": "Other",
    "Welders": "Other",
    "Air Compressors": "Other",
    "Pumps": "Other",
    "Forklifts": "Telehandler / Forklift",
    "Telehandlers": "Telehandler / Forklift",
    "Man Lifts": "Other",
    "Boom Lifts": "Other",
    "Scissor Lifts": "Other",
    "Crushers": "Other",
    "Compactors": "Plate Compactor",
    "Mowers": "Other",
    "Storage / Containers": "Other",
    "GPS / Electronics": "Other",
    "Attachments": "Other",
    "Misc Equipment": "Other",
}

# Header-row keywords (uppercase rows that announce a section in the sheet)
_HEADER_KEYWORDS = {
    "DUMP TRUCK","TRACTOR TRAILER","SUPERVISOR","MGMT","MANAGEMENT","SERVICE TRUCK",
    "WATER TRUCK","PAVING","ROLLER","EXCAVATOR","DOZER","LOADER","BACKHOE","SKID STEER",
    "SWEEPER","TRAILER","PICKUP","VAN","FORKLIFT","FORK LIFT","CRUSHER","MILL","PAVER",
    "COMPACTOR","TACK","DISTRIBUTOR","GENERATOR","LIGHT TOWER","COMPRESSOR","WELDER",
    "MAN LIFT","MANLIFT","BOOM LIFT","SCISSOR","TELEHANDLER","ATTACHMENT","MOWER",
    "ROAD GRADER","GRADER","FLATBED","MISC. TRUCK","MISC TRUCK","MISC EQUIPMENT",
    "STORAGE","CONTAINER","PUMP","ROAD WIDENER","SHUTTLE BUGGY","ARTICULATED",
}


def _is_header(row) -> bool:
    if not row:
        return False
    a = row[0]
    if not isinstance(a, str) or not a.strip():
        return False
    rest_empty = all(
        (c is None or (isinstance(c, str) and not c.strip())) for c in row[1:8]
    )
    if not rest_empty:
        return False
    upper = a.strip().upper()
    if any(k in upper for k in _HEADER_KEYWORDS):
        return True
    if a.strip() == upper and len(a.strip()) > 2 and not any(ch.isdigit() for ch in a.strip()):
        return True
    return False


def _category_from_keywords(make_model: str, comments: str = "") -> Optional[str]:
    s = (make_model + " " + comments).lower()
    for kw, cat in KEYWORD_TO_CAT:
        if kw in s:
            return cat
    return None


def parse_equipment_xlsx(file_bytes: bytes, sheet_name: str = "Louis") -> Dict[str, Any]:
    """
    Parse an Equipment List xlsx and return:
      {
        "items": [ ...normalized equipment dicts... ],
        "count": int,
        "category_counts": { "Excavators": 35, ... },
        "sheet": "Louis",
      }
    Falls back to the first sheet if the requested one is missing.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    use_sheet = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
    ws = wb[use_sheet]

    current_header: Optional[str] = None
    items: List[Dict[str, Any]] = []

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        a = row[0]
        if isinstance(a, str) and a.strip().lower() == "equipment #":
            continue  # column header
        if _is_header(row):
            current_header = a.strip()
            continue

        # Pull first 7 columns
        eq_num, year, make_model, plate, vin, comments, company = (row + (None,) * 7)[:7]
        if not isinstance(make_model, str) or not make_model.strip():
            continue
        mm = make_model.strip()
        if mm.lower() in ("make/model", "make / model"):
            continue

        un = (str(eq_num).strip() if eq_num else "")

        # Determine category — prefix → keyword → header → "Misc Equipment"
        cat: Optional[str] = None
        m = re.match(r"^([A-Z]{2,4})\d", un)
        if m and m.group(1) in PREFIX_TO_CAT:
            cat = PREFIX_TO_CAT[m.group(1)]
        if not cat:
            cat = _category_from_keywords(mm, str(comments or ""))
        if not cat and current_header:
            cat = current_header.title()
        if not cat:
            cat = "Misc Equipment"
        cat = NORMALIZE_CAT.get(cat, cat)

        yr = int(year) if isinstance(year, (int, float)) and year else None
        yr_mm = (f"{yr} " if yr else "") + mm
        display_label = " — ".join(p for p in [un, yr_mm.strip()] if p)

        items.append({
            "unit_number": un,
            "year": yr,
            "make_model": mm,
            "plate": str(plate).strip() if plate else "",
            "vin_serial_number": str(vin).strip() if vin else "",
            "comments": str(comments).strip() if comments else "",
            "company": str(company).strip() if company else "",
            "category": cat,
            "preop_equipment_type": PREOP_TYPE_MAP.get(cat, "Other"),
            "display_label": display_label,
        })

    # Dedupe on (unit_number, make_model, vin)
    seen = set()
    deduped: List[Dict[str, Any]] = []
    for it in items:
        key = (
            it["unit_number"].lower(),
            it["make_model"].lower(),
            it["vin_serial_number"].lower(),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(it)

    deduped.sort(
        key=lambda x: (x["category"], x["unit_number"] or "", x["make_model"])
    )

    counts = dict(Counter(i["category"] for i in deduped))
    return {
        "items": deduped,
        "count": len(deduped),
        "category_counts": counts,
        "sheet": use_sheet,
    }
