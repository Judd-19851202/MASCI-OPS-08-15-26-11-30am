from __future__ import annotations

import asyncio
import csv
import hashlib
import io
import re
from collections import defaultdict
from copy import deepcopy
from datetime import date, datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple
from uuid import uuid4

import pdfplumber
from openpyxl import Workbook, load_workbook

from services.cost_codes.foundation import (
    normalize_job_assignment,
    persist_project_assignments,
)
from services.project_budget_authority import (
    COLL_BUDGET_LINES,
    COLL_BUDGET_VERSIONS,
)
from services.project_controls_authority import (
    _actor_label,
    _clean,
    _load_job,
    _norm,
    _sanitize,
    _status,
    _suggest_work_type,
    _to_float,
    _write_audit,
    ensure_project_controls_foundation,
    get_project_lookahead,
    list_enterprise_work_types,
    save_project_lookahead,
)


COLL_SCHEDULE_VERSIONS = "project_schedule_versions"
COLL_SCHEDULE_ACTIVITIES = "project_schedule_activities"
COLL_SCHEDULE_IMPORTS = "project_schedule_import_sessions"
COLL_SCHEDULE_IMPORT_ROWS = "project_schedule_import_rows"
COLL_SCHEDULE_REVIEW = "project_schedule_review_queue"
COLL_WORK_PACKAGES = "project_work_packages"
COLL_SCHEDULE_DISTRIBUTION = "project_schedule_distribution_audit"
COLL_SCHEDULE_RUNS = "project_schedule_runs"

SOURCE_KINDS = ["csv", "excel", "pdf_review", "primavera_p6", "ms_project"]
VERSION_KINDS = ["master_schedule", "pending_revision"]
PRIORITY_LEVELS = ["low", "normal", "high", "critical"]
EXECUTION_STRATEGIES = ["self_perform", "vendor", "subcontractor", "hybrid"]

CONSTRAINT_CATEGORIES = [
    "weather",
    "utility_conflict",
    "material_delay",
    "rfi",
    "owner_decision",
    "permit",
    "traffic",
    "safety",
    "equipment",
    "unknown",
]

FILE_KINDS = {
    ".csv": "csv",
    ".xlsx": "excel",
    ".xlsm": "excel",
    ".xltx": "excel",
    ".xltm": "excel",
    ".pdf": "pdf_review",
    ".xer": "primavera_p6",
    ".mpp": "ms_project",
    ".xml": "primavera_p6",
}

EVENT_CONTRACTS = [
    {
        "event_key": "schedule.import_staged",
        "producer": "project_schedule_authority",
        "authority_owner": COLL_SCHEDULE_IMPORTS,
        "consumers": ["pm_project_schedule", "admin_project_schedule", COLL_SCHEDULE_REVIEW],
        "idempotency_key": "project_number:import_id",
        "operator_visible_consequence": "Schedule imports stay reviewable until a PM approves activation.",
    },
    {
        "event_key": "schedule.activity_reviewed",
        "producer": "project_schedule_authority",
        "authority_owner": COLL_SCHEDULE_IMPORT_ROWS,
        "consumers": ["pm_project_schedule", "admin_project_schedule"],
        "idempotency_key": "import_id:row_id:reviewed_at",
        "operator_visible_consequence": "Each activity mapping is explicitly reviewed before activation.",
    },
    {
        "event_key": "schedule.version_activated",
        "producer": "project_schedule_authority",
        "authority_owner": COLL_SCHEDULE_VERSIONS,
        "consumers": [COLL_SCHEDULE_ACTIVITIES, COLL_WORK_PACKAGES, "pm_project_schedule"],
        "idempotency_key": "project_number:version_id",
        "operator_visible_consequence": "Baseline history is preserved while the active schedule projection updates.",
    },
    {
        "event_key": "schedule.lookahead_saved",
        "producer": "project_schedule_authority",
        "authority_owner": "project_lookahead",
        "consumers": ["pm_project_schedule", "admin_project_schedule"],
        "idempotency_key": "project_number:lookahead_id:version",
        "operator_visible_consequence": "Lookahead changes remain a governed view and do not overwrite the baseline schedule.",
    },
]


_FOUNDATION_READY_DBS: set[str] = set()
_FOUNDATION_READY_LOCK = asyncio.Lock()

COLUMN_ALIASES = {
    "activity_id": ["activity id", "activity code", "task id", "id", "wbs activity"],
    "activity_name": ["activity name", "task name", "activity", "description", "task"],
    "phase_id": ["phase", "phase id"],
    "work_package_id": ["work package", "work package id", "package", "wbs"],
    "customer_pay_item_number": ["customer pay item", "pay item", "pay item number", "customer cost code"],
    "project_cost_code": ["project cost code", "cost code", "operational cost code"],
    "duration_days": ["duration", "duration days", "days"],
    "planned_start_date": ["start", "start date", "planned start"],
    "planned_finish_date": ["finish", "finish date", "planned finish"],
    "predecessor_activity_ids": ["predecessors", "dependencies", "predecessor ids"],
    "calendar_name": ["calendar", "calendar name"],
    "status": ["status", "activity status"],
    "percent_complete": ["percent complete", "% complete", "progress %"],
    "owner": ["owner", "responsible", "responsible party"],
    "priority": ["priority"],
    "notes": ["notes", "remarks"],
    "planned_crew_ids": ["planned crew", "crew", "crew plan"],
    "planned_employee_ids": ["planned employees", "employees", "labor"],
    "planned_equipment_ids": ["planned equipment", "equipment", "equipment plan"],
    "planned_materials": ["planned materials", "materials", "material plan"],
    "planned_vendor_refs": ["planned vendor", "vendor", "vendor plan"],
    "planned_subcontractor_refs": ["planned subcontractor", "subcontractor", "sub plan"],
    "planned_production_quantity": ["planned production quantity", "production quantity", "planned qty"],
    "planned_hours": ["planned hours", "hours"],
    "planned_constraints": ["planned constraints", "constraints"],
    "execution_strategy": ["execution strategy", "execution"],
}


def _utcnow() -> str:
    return datetime.now(timezone.utc).isoformat()


def _safe_float(value: Any) -> float:
    if isinstance(value, str):
        value = value.replace(",", "").replace("$", "").replace("%", "").strip()
    return round(_to_float(value, 0.0), 4)


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        if value in (None, ""):
            return default
        return int(round(float(value)))
    except (TypeError, ValueError):
        return default


def _canonical_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _clean(value).lower()).strip()


def _clean_filename(filename: str) -> str:
    return "".join(ch if ch.isalnum() or ch in {"-", "_", "."} else "_" for ch in _clean(filename) or "upload")


def _hash_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _extract_extension(filename: str) -> str:
    cleaned = _clean(filename)
    if "." not in cleaned:
        return ""
    return "." + cleaned.rsplit(".", 1)[-1].lower()


def _coerce_list(value: Any) -> List[str]:
    if isinstance(value, list):
        raw = value
    elif isinstance(value, str):
        raw = [piece.strip() for line in value.splitlines() for piece in line.split(",")]
    else:
        raw = []
    seen = set()
    out = []
    for item in raw:
        if isinstance(item, dict):
            text = _clean(
                item.get("label")
                or item.get("title")
                or item.get("description")
                or item.get("vendor_name")
                or item.get("subcontractor_name")
                or item.get("name")
                or item.get("constraint_id")
                or item.get("material_id")
                or item.get("crew_id")
                or item.get("employee_id")
                or item.get("equipment_id")
                or item.get("vendor_id")
                or item.get("id")
            )
        else:
            text = _clean(item)
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(text)
    return out


def _parse_date(value: Any) -> str:
    text = _clean(value)[:10]
    if not text:
        return ""
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError:
        return ""


def _match_alias(headers: Dict[str, Any], logical_key: str) -> Any:
    for alias in COLUMN_ALIASES.get(logical_key) or []:
        key = _canonical_key(alias)
        if key in headers and _clean(headers[key]) != "":
            return headers[key]
    return ""


def _calc_duration_days(start_date: str, finish_date: str, fallback: int = 1) -> int:
    try:
        start = date.fromisoformat(start_date)
        finish = date.fromisoformat(finish_date)
        return max(1, (finish - start).days + 1)
    except Exception:
        return max(1, fallback)


def _normalize_priority(value: Any) -> str:
    return _status(value or "normal", allowed=PRIORITY_LEVELS, default="normal")


def _normalize_execution_strategy(value: Any) -> str:
    return _status(value or "self_perform", allowed=EXECUTION_STRATEGIES, default="self_perform")


def _constraint_tokens(value: Any) -> List[Dict[str, Any]]:
    rows = []
    if isinstance(value, list) and any(isinstance(item, dict) for item in value):
        seen = set()
        for item in value:
            if not isinstance(item, dict):
                continue
            title = _clean(item.get("title") or item.get("label") or item.get("description") or item.get("constraint_id"))
            if not title:
                continue
            dedupe_key = title.lower()
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            rows.append(
                {
                    "constraint_id": _clean(item.get("constraint_id") or item.get("id")),
                    "category": _status(
                        item.get("category") or title.lower().replace("-", "_").replace(" ", "_"),
                        allowed=CONSTRAINT_CATEGORIES,
                        default="unknown",
                    ),
                    "title": title,
                    "status": _clean(item.get("status") or "planned") or "planned",
                    "notes": _clean(item.get("notes")),
                }
            )
        return rows
    for token in _coerce_list(value):
        rows.append(
            {
                "constraint_id": "",
                "category": _status(token.lower().replace("-", "_").replace(" ", "_"), allowed=CONSTRAINT_CATEGORIES, default="unknown"),
                "title": token,
                "status": "planned",
                "notes": "",
            }
        )
    return rows


def _normalize_resource_refs(value: Any, *, kind: str) -> List[Dict[str, Any]]:
    rows = []
    if isinstance(value, list) and any(isinstance(item, dict) for item in value):
        seen = set()
        for item in value:
            if not isinstance(item, dict):
                continue
            row_id = _clean(item.get(f"{kind}_id") or item.get("id"))
            label = _clean(item.get("label") or item.get("name") or item.get("description") or row_id)
            if not label and not row_id:
                continue
            dedupe_key = (row_id or label).lower()
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            rows.append({f"{kind}_id": row_id, "label": label or row_id})
        return rows
    return [{f"{kind}_id": "", "label": token} for token in _coerce_list(value)]


def _normalize_material_refs(value: Any) -> List[Dict[str, Any]]:
    if isinstance(value, list) and any(isinstance(item, dict) for item in value):
        rows = []
        seen = set()
        for item in value:
            if not isinstance(item, dict):
                continue
            material_id = _clean(item.get("material_id") or item.get("id"))
            description = _clean(item.get("description") or item.get("label") or material_id)
            if not material_id and not description:
                continue
            dedupe_key = (material_id or description).lower()
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            rows.append(
                {
                    "material_id": material_id,
                    "description": description,
                    "quantity": _safe_float(item.get("quantity")),
                    "unit": _clean(item.get("unit")),
                }
            )
        return rows
    return [{"material_id": "", "description": token, "quantity": 0.0, "unit": ""} for token in _coerce_list(value)]


def _normalize_vendor_refs(value: Any) -> List[Dict[str, Any]]:
    if isinstance(value, list) and any(isinstance(item, dict) for item in value):
        rows = []
        seen = set()
        for item in value:
            if not isinstance(item, dict):
                continue
            vendor_id = _clean(item.get("vendor_id") or item.get("id"))
            vendor_name = _clean(item.get("vendor_name") or item.get("label") or item.get("name") or vendor_id)
            if not vendor_id and not vendor_name:
                continue
            dedupe_key = (vendor_id or vendor_name).lower()
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            rows.append({"vendor_id": vendor_id, "vendor_name": vendor_name})
        return rows
    return [{"vendor_id": "", "vendor_name": token} for token in _coerce_list(value)]


def _normalize_subcontractor_refs(value: Any) -> List[Dict[str, Any]]:
    if isinstance(value, list) and any(isinstance(item, dict) for item in value):
        rows = []
        seen = set()
        for item in value:
            if not isinstance(item, dict):
                continue
            vendor_id = _clean(item.get("vendor_id") or item.get("id"))
            subcontractor_name = _clean(item.get("subcontractor_name") or item.get("vendor_name") or item.get("label") or item.get("name") or vendor_id)
            if not vendor_id and not subcontractor_name:
                continue
            dedupe_key = (vendor_id or subcontractor_name).lower()
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            rows.append({"vendor_id": vendor_id, "subcontractor_name": subcontractor_name})
        return rows
    return [{"vendor_id": "", "subcontractor_name": token} for token in _coerce_list(value)]


def _normalize_row(source_values: Dict[str, Any]) -> Dict[str, Any]:
    headers = {_canonical_key(key): value for key, value in (source_values or {}).items()}
    activity_id = _clean(_match_alias(headers, "activity_id"))
    activity_name = _clean(_match_alias(headers, "activity_name"))
    start_date = _parse_date(_match_alias(headers, "planned_start_date"))
    finish_date = _parse_date(_match_alias(headers, "planned_finish_date"))
    duration_days = _safe_int(_match_alias(headers, "duration_days"), 0)
    if duration_days <= 0:
        duration_days = _calc_duration_days(start_date, finish_date, 1)
    normalized = {
        "activity_id": activity_id,
        "activity_name": activity_name,
        "phase_id": _clean(_match_alias(headers, "phase_id")),
        "work_package_id": _clean(_match_alias(headers, "work_package_id")),
        "customer_pay_item_number": _clean(_match_alias(headers, "customer_pay_item_number")),
        "project_cost_code": _clean(_match_alias(headers, "project_cost_code")),
        "planned_start_date": start_date,
        "planned_finish_date": finish_date,
        "duration_days": max(1, duration_days),
        "predecessor_activity_ids": _coerce_list(_match_alias(headers, "predecessor_activity_ids")),
        "calendar_name": _clean(_match_alias(headers, "calendar_name") or "Default"),
        "status": _clean(_match_alias(headers, "status") or "not_started"),
        "percent_complete": max(0.0, min(100.0, _safe_float(_match_alias(headers, "percent_complete")))),
        "owner": _clean(_match_alias(headers, "owner")),
        "priority": _normalize_priority(_match_alias(headers, "priority")),
        "notes": _clean(_match_alias(headers, "notes")),
        "planned_crew_ids": _normalize_resource_refs(_match_alias(headers, "planned_crew_ids"), kind="crew"),
        "planned_employee_ids": _normalize_resource_refs(_match_alias(headers, "planned_employee_ids"), kind="employee"),
        "planned_equipment_ids": _normalize_resource_refs(_match_alias(headers, "planned_equipment_ids"), kind="equipment"),
        "planned_materials": _normalize_material_refs(_match_alias(headers, "planned_materials")),
        "planned_vendor_refs": _normalize_vendor_refs(_match_alias(headers, "planned_vendor_refs")),
        "planned_subcontractor_refs": _normalize_subcontractor_refs(_match_alias(headers, "planned_subcontractor_refs")),
        "planned_production_quantity": _safe_float(_match_alias(headers, "planned_production_quantity")),
        "planned_hours": _safe_float(_match_alias(headers, "planned_hours")),
        "planned_constraints": _constraint_tokens(_match_alias(headers, "planned_constraints")),
        "execution_strategy": _normalize_execution_strategy(_match_alias(headers, "execution_strategy")),
    }
    return normalized


def _parse_csv_bytes(data: bytes) -> Tuple[List[Dict[str, Any]], List[str], str]:
    warnings = []
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
        warnings.append("CSV decoded with latin-1 fallback.")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for index, row in enumerate(reader, start=2):
        if not any(_clean(value) for value in (row or {}).values()):
            continue
        rows.append({"row_number": index, "source_values": _sanitize(row or {})})
    if not rows:
        warnings.append("No schedule rows were found in the CSV file.")
    return rows, warnings, "csv_dict_reader"


def _parse_excel_bytes(data: bytes) -> Tuple[List[Dict[str, Any]], List[str], str]:
    warnings = []
    rows = []
    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            continue
        headers = [str(cell or "").strip() for cell in values[0]]
        if not any(headers):
            headers = [f"column_{index + 1}" for index in range(len(values[0]))]
        for row_number, values_row in enumerate(values[1:], start=2):
            raw = {headers[index] if headers[index] else f"column_{index + 1}": values_row[index] for index in range(len(headers))}
            if not any(_clean(value) for value in raw.values()):
                continue
            rows.append({"row_number": row_number, "sheet_name": sheet.title, "source_values": _sanitize(raw)})
    if not rows:
        warnings.append("No schedule rows were found in the Excel workbook.")
    return rows, warnings, "openpyxl"


def _parse_pdf_line(line: str) -> Dict[str, Any]:
    text = _clean(line)
    if not text:
        return {}
    pattern = re.compile(
        r"^(?P<activity>[A-Za-z0-9\-_.]+)\s+(?P<name>.+?)\s+(?P<start>\d{4}-\d{2}-\d{2})\s+(?P<finish>\d{4}-\d{2}-\d{2})\s+(?P<duration>\d+)$"
    )
    match = pattern.match(text)
    if not match:
        return {"activity name": text}
    return {
        "activity id": match.group("activity"),
        "activity name": match.group("name"),
        "start date": match.group("start"),
        "finish date": match.group("finish"),
        "duration": match.group("duration"),
    }


def _parse_pdf_bytes(data: bytes) -> Tuple[List[Dict[str, Any]], List[str], str]:
    warnings = []
    rows = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables() or []
            if tables:
                for table in tables:
                    if not table:
                        continue
                    headers = [str(cell or "").strip() for cell in (table[0] or [])]
                    if not any(headers):
                        headers = [f"column_{i + 1}" for i in range(len(table[0] or []))]
                    for row_number, values_row in enumerate(table[1:], start=2):
                        raw = {headers[i] if i < len(headers) and headers[i] else f"column_{i + 1}": values_row[i] for i in range(len(values_row or []))}
                        if not any(_clean(value) for value in raw.values()):
                            continue
                        rows.append({"row_number": row_number, "page_number": page_number, "source_values": _sanitize(raw)})
            else:
                text = page.extract_text() or ""
                page_rows = []
                for line_number, line in enumerate(text.splitlines(), start=1):
                    parsed = _parse_pdf_line(line)
                    if not parsed:
                        continue
                    page_rows.append({"row_number": line_number, "page_number": page_number, "source_values": _sanitize(parsed)})
                if page_rows:
                    warnings.append(f"Page {page_number} used line-based PDF extraction and still needs PM review.")
                    rows.extend(page_rows)
    if not rows:
        warnings.append("No reviewable activity rows were extracted from the PDF.")
    return rows, warnings, "pdfplumber"


def _parse_extension_point_bytes(filename: str, data: bytes, *, source_kind: str) -> Tuple[List[Dict[str, Any]], List[str], str]:
    warning = f"{source_kind} adapter foundation is wired, but runtime production certification is deferred until representative source files are supplied. File preserved as review-only evidence."
    row = {
        "row_number": 1,
        "source_values": {
            "activity id": _clean(filename),
            "activity name": f"{source_kind} source preserved for governed review",
            "notes": warning,
        },
    }
    return [row], [warning], f"{source_kind}_extension_point"


def _parse_import_file(filename: str, data: bytes, *, source_kind: str) -> Tuple[List[Dict[str, Any]], List[str], str, str]:
    extension = _extract_extension(filename)
    file_kind = FILE_KINDS.get(extension)
    if source_kind in {"primavera_p6", "ms_project"}:
        rows, warnings, parser = _parse_extension_point_bytes(filename, data, source_kind=source_kind)
        return rows, warnings, parser, source_kind
    if not file_kind:
        raise ValueError("unsupported_schedule_import_file")
    if file_kind == "csv":
        rows, warnings, parser = _parse_csv_bytes(data)
    elif file_kind == "excel":
        rows, warnings, parser = _parse_excel_bytes(data)
    elif file_kind == "pdf_review":
        rows, warnings, parser = _parse_pdf_bytes(data)
    else:
        rows, warnings, parser = _parse_extension_point_bytes(filename, data, source_kind=file_kind)
    return rows, warnings, parser, file_kind


async def _ensure_indexes(db) -> None:
    await db[COLL_SCHEDULE_VERSIONS].create_index([("project_number", 1), ("version_id", 1)], unique=True)
    await db[COLL_SCHEDULE_VERSIONS].create_index([("project_number", 1), ("status", 1), ("activated_at", -1)])
    await db[COLL_SCHEDULE_VERSIONS].create_index([("project_number", 1), ("activated_at", -1)])
    await db[COLL_SCHEDULE_ACTIVITIES].create_index([("project_number", 1), ("version_id", 1), ("activity_id", 1)], unique=True)
    await db[COLL_SCHEDULE_ACTIVITIES].create_index([("project_number", 1), ("work_package_id", 1), ("status", 1)])
    await db[COLL_SCHEDULE_ACTIVITIES].create_index([("project_number", 1), ("version_id", 1), ("phase_id", 1), ("work_package_id", 1), ("planned_start_date", 1), ("activity_id", 1)])
    await db[COLL_SCHEDULE_IMPORTS].create_index([("project_number", 1), ("import_id", 1)], unique=True)
    await db[COLL_SCHEDULE_IMPORTS].create_index([("project_number", 1), ("imported_at", -1)])
    await db[COLL_SCHEDULE_IMPORT_ROWS].create_index([("import_id", 1), ("row_id", 1)], unique=True)
    await db[COLL_SCHEDULE_REVIEW].create_index("review_id", unique=True)
    await db[COLL_SCHEDULE_REVIEW].create_index([("project_number", 1), ("priority", -1), ("updated_at", -1)])
    await db[COLL_WORK_PACKAGES].create_index([("project_number", 1), ("version_id", 1), ("work_package_id", 1)], unique=True)
    await db[COLL_SCHEDULE_DISTRIBUTION].create_index([("project_number", 1), ("created_at", -1)])
    await db[COLL_SCHEDULE_RUNS].create_index([("run_type", 1)], unique=True)


async def ensure_project_schedule_foundation(db) -> Dict[str, Any]:
    db_key = str(getattr(db, "name", "")) or COLL_SCHEDULE_VERSIONS
    if db_key not in _FOUNDATION_READY_DBS:
        async with _FOUNDATION_READY_LOCK:
            if db_key not in _FOUNDATION_READY_DBS:
                await _ensure_indexes(db)
                _FOUNDATION_READY_DBS.add(db_key)
    await ensure_project_controls_foundation(db)
    last_run = await db[COLL_SCHEDULE_RUNS].find_one({"run_type": "wp18c4_backfill"}, {"_id": 0})
    return {
        "ok": True,
        "backfill": _sanitize(last_run or {"run_type": "wp18c4_backfill", "status": "pending_manual_run"}),
        "event_contracts": EVENT_CONTRACTS,
    }


async def _list_budget_lines_for_project(db, project_number: str) -> List[Dict[str, Any]]:
    active = await db[COLL_BUDGET_VERSIONS].find_one({"project_number": project_number, "status": "active"}, {"_id": 0}, sort=[("activated_at", -1)])
    if not active:
        return []
    return [_sanitize(row) async for row in db[COLL_BUDGET_LINES].find({"project_number": project_number, "version_id": active["version_id"]}, {"_id": 0})]


def _tokenize(text: str) -> List[str]:
    return [token for token in re.split(r"[^a-z0-9]+", (text or "").lower()) if token]


def _suggest_budget_line(normalized: Dict[str, Any], budget_lines: List[Dict[str, Any]]) -> Dict[str, Any]:
    project_cost_code = _clean(normalized.get("project_cost_code"))
    if project_cost_code:
        for line in budget_lines:
            if _clean(line.get("project_cost_code")).lower() == project_cost_code.lower():
                return {
                    "budget_line_id": line.get("budget_line_id") or "",
                    "customer_pay_item_number": line.get("customer_pay_item_number") or "",
                    "enterprise_work_type_id": line.get("enterprise_work_type_id") or "",
                    "phase_id": line.get("phase_id") or "",
                    "work_package_id": line.get("work_package_id") or "",
                    "project_cost_code": line.get("project_cost_code") or project_cost_code,
                    "confidence": "high",
                    "reason": "Exact project cost-code match found in the active governed budget.",
                }
    customer_pay_item_number = _clean(normalized.get("customer_pay_item_number"))
    if customer_pay_item_number:
        for line in budget_lines:
            if _clean(line.get("customer_pay_item_number")).lower() == customer_pay_item_number.lower():
                return {
                    "budget_line_id": line.get("budget_line_id") or "",
                    "customer_pay_item_number": line.get("customer_pay_item_number") or customer_pay_item_number,
                    "enterprise_work_type_id": line.get("enterprise_work_type_id") or "",
                    "phase_id": line.get("phase_id") or "",
                    "work_package_id": line.get("work_package_id") or "",
                    "project_cost_code": line.get("project_cost_code") or project_cost_code,
                    "confidence": "high",
                    "reason": "Exact customer pay-item match found in the active governed budget.",
                }
    desc_tokens = set(_tokenize(normalized.get("activity_name") or ""))
    best = None
    best_score = 0
    for line in budget_lines:
        line_tokens = set(_tokenize(f"{line.get('description') or ''} {line.get('customer_pay_item_number') or ''} {line.get('project_cost_code') or ''}"))
        score = len(desc_tokens & line_tokens)
        if score > best_score:
            best_score = score
            best = line
    if not best:
        return {"budget_line_id": "", "confidence": "review_required", "reason": "No governed budget line could be matched confidently."}
    return {
        "budget_line_id": best.get("budget_line_id") or "",
        "customer_pay_item_number": best.get("customer_pay_item_number") or customer_pay_item_number,
        "enterprise_work_type_id": best.get("enterprise_work_type_id") or "",
        "phase_id": best.get("phase_id") or "",
        "work_package_id": best.get("work_package_id") or "",
        "project_cost_code": best.get("project_cost_code") or project_cost_code,
        "confidence": "medium" if best_score >= 2 else "review_required",
        "reason": "Description overlap suggested a reusable governed budget line, but PM review remains required.",
    }


def _build_row_suggestion(normalized: Dict[str, Any], budget_lines: List[Dict[str, Any]], work_types: List[Dict[str, Any]]) -> Dict[str, Any]:
    budget_line = _suggest_budget_line(normalized, budget_lines)
    work_type = _suggest_work_type(
        {"customer_pay_item_number": normalized.get("customer_pay_item_number"), "description": normalized.get("activity_name")},
        work_types,
    )
    warnings: List[str] = []
    if not _clean(normalized.get("activity_id")):
        warnings.append("Activity ID could not be resolved from the source file.")
    if not _clean(normalized.get("activity_name")):
        warnings.append("Activity name could not be resolved from the source file.")
    if not _clean(normalized.get("planned_start_date")):
        warnings.append("Planned start date is missing.")
    if int(normalized.get("duration_days") or 0) < 1:
        warnings.append("Duration must be at least one day.")
    if not _clean(budget_line.get("budget_line_id")):
        warnings.append("Budget line still needs governed mapping.")
    if not _clean(normalized.get("work_package_id") or budget_line.get("work_package_id")):
        warnings.append("Work package still needs PM confirmation.")
    confidence_levels = [budget_line.get("confidence") or "review_required", work_type.get("confidence") or "review_required"]
    confidence = "high" if all(level == "high" for level in confidence_levels) else "medium" if "medium" in confidence_levels and "review_required" not in confidence_levels else "review_required"
    return {
        "budget_line_id": budget_line.get("budget_line_id") or "",
        "customer_pay_item_number": budget_line.get("customer_pay_item_number") or normalized.get("customer_pay_item_number") or "",
        "enterprise_work_type_id": budget_line.get("enterprise_work_type_id") or work_type.get("primary_work_type_id") or "",
        "phase_id": budget_line.get("phase_id") or normalized.get("phase_id") or "",
        "work_package_id": budget_line.get("work_package_id") or normalized.get("work_package_id") or "",
        "project_cost_code": budget_line.get("project_cost_code") or normalized.get("project_cost_code") or "",
        "confidence": confidence,
        "matched_terms": work_type.get("matched_terms") or [],
        "reasons": [reason for reason in [budget_line.get("reason"), ("Matched work-type keywords: " + ", ".join(work_type.get("matched_terms") or [])) if work_type.get("matched_terms") else "No strong work-type keyword overlap was found."] if reason],
        "warnings": warnings,
    }


async def _upsert_review_item(db, review: Dict[str, Any]) -> Dict[str, Any]:
    now = _utcnow()
    existing = await db[COLL_SCHEDULE_REVIEW].find_one({"review_id": review["review_id"]}, {"_id": 0})
    doc = {**(existing or {}), **review, "created_at": (existing or {}).get("created_at") or now, "updated_at": now}
    await db[COLL_SCHEDULE_REVIEW].replace_one({"review_id": doc["review_id"]}, doc, upsert=True)
    return _sanitize(doc)


async def _mark_review_resolved(db, review_id: str, *, actor: Optional[Dict[str, Any]] = None, resolution_note: str = "") -> None:
    row = await db[COLL_SCHEDULE_REVIEW].find_one({"review_id": review_id}, {"_id": 0})
    if not row:
        return
    row["status"] = "resolved"
    row["resolution_note"] = _clean(resolution_note) or "Resolved by subsequent governed schedule action."
    row["resolved_at"] = _utcnow()
    row["resolved_by"] = _actor_label(actor)
    row["updated_at"] = _utcnow()
    await db[COLL_SCHEDULE_REVIEW].replace_one({"review_id": review_id}, row, upsert=True)


async def list_schedule_versions(db, project_number: str) -> List[Dict[str, Any]]:
    await ensure_project_schedule_foundation(db)
    return [_sanitize(row) async for row in db[COLL_SCHEDULE_VERSIONS].find({"project_number": project_number}, {"_id": 0}).sort([("activated_at", -1), ("created_at", -1)])]


async def list_schedule_activities(db, project_number: str, *, version_id: str) -> List[Dict[str, Any]]:
    await ensure_project_schedule_foundation(db)
    return [_sanitize(row) async for row in db[COLL_SCHEDULE_ACTIVITIES].find({"project_number": project_number, "version_id": version_id}, {"_id": 0}).sort([("phase_id", 1), ("work_package_id", 1), ("planned_start_date", 1), ("activity_id", 1)])]


async def list_schedule_work_packages(db, project_number: str, *, version_id: str = "") -> List[Dict[str, Any]]:
    await ensure_project_schedule_foundation(db)
    query: Dict[str, Any] = {"project_number": project_number}
    if _clean(version_id):
        query["version_id"] = _clean(version_id)
    return [_sanitize(row) async for row in db[COLL_WORK_PACKAGES].find(query, {"_id": 0}).sort([("phase_id", 1), ("work_package_id", 1)])]


async def list_schedule_imports(db, project_number: str) -> List[Dict[str, Any]]:
    await ensure_project_schedule_foundation(db)
    return [_sanitize(row) async for row in db[COLL_SCHEDULE_IMPORTS].find({"project_number": project_number}, {"_id": 0}).sort([("imported_at", -1)])]


async def get_schedule_import_detail(db, project_number: str, import_id: str) -> Dict[str, Any]:
    await ensure_project_schedule_foundation(db)
    session = await db[COLL_SCHEDULE_IMPORTS].find_one({"project_number": project_number, "import_id": import_id}, {"_id": 0})
    if not session:
        raise LookupError("schedule_import_not_found")
    rows = [_sanitize(row) async for row in db[COLL_SCHEDULE_IMPORT_ROWS].find({"project_number": project_number, "import_id": import_id}, {"_id": 0}).sort([("row_number", 1)])]
    return {"session": _sanitize(session), "rows": rows, "count": len(rows)}


async def list_schedule_review_queue(db, *, project_number: str = "") -> List[Dict[str, Any]]:
    await ensure_project_schedule_foundation(db)
    query = {"project_number": project_number} if project_number else {}
    return [_sanitize(row) async for row in db[COLL_SCHEDULE_REVIEW].find(query, {"_id": 0}).sort([("priority", -1), ("updated_at", -1)]).limit(500)]


def _schedule_status_from_counts(total: int, approved: int, rejected: int, pending: int, activated: int) -> str:
    if activated == total and total > 0:
        return "activated"
    if total > 0 and approved + rejected == total and pending == 0:
        return "approved_ready"
    if approved > 0 or rejected > 0:
        return "partially_reviewed"
    return "review_required"


async def _refresh_import_counts(db, import_id: str) -> Dict[str, Any]:
    total = await db[COLL_SCHEDULE_IMPORT_ROWS].count_documents({"import_id": import_id})
    approved = await db[COLL_SCHEDULE_IMPORT_ROWS].count_documents({"import_id": import_id, "review_status": "approved"})
    rejected = await db[COLL_SCHEDULE_IMPORT_ROWS].count_documents({"import_id": import_id, "review_status": "rejected"})
    pending = await db[COLL_SCHEDULE_IMPORT_ROWS].count_documents({"import_id": import_id, "review_status": {"$in": ["pending_review", "review_required"]}})
    activated = await db[COLL_SCHEDULE_IMPORT_ROWS].count_documents({"import_id": import_id, "review_status": "activated"})
    status = _schedule_status_from_counts(total, approved, rejected, pending, activated)
    await db[COLL_SCHEDULE_IMPORTS].update_one({"import_id": import_id}, {"$set": {"total_rows": total, "approved_rows": approved, "rejected_rows": rejected, "needs_review_rows": pending, "activated_rows": activated, "status": status, "updated_at": _utcnow()}})
    session = await db[COLL_SCHEDULE_IMPORTS].find_one({"import_id": import_id}, {"_id": 0})
    return _sanitize(session or {})


async def create_schedule_import_session(
    db,
    project_number: str,
    *,
    filename: str,
    content_type: str,
    data: bytes,
    source_kind: str,
    target_version_kind: str,
    version_name: str,
    actor: Dict[str, Any],
) -> Dict[str, Any]:
    await ensure_project_schedule_foundation(db)
    job = await _load_job(db, project_number)
    source_kind = _status(source_kind or "csv", allowed=SOURCE_KINDS, default="csv")
    target_version_kind = _status(target_version_kind or "master_schedule", allowed=VERSION_KINDS, default="master_schedule")
    if not data:
        raise ValueError("schedule_import_file_required")
    rows, parser_warnings, parser_name, file_kind = _parse_import_file(filename, data, source_kind=source_kind)
    file_hash = _hash_bytes(data)
    duplicate = await db[COLL_SCHEDULE_IMPORTS].find_one({"project_number": project_number, "file_hash": file_hash}, {"_id": 0})
    if duplicate:
        detail = await get_schedule_import_detail(db, project_number, duplicate["import_id"])
        detail["duplicate_of"] = duplicate["import_id"]
        return detail
    work_types = await list_enterprise_work_types(db, include_archived=False)
    budget_lines = await _list_budget_lines_for_project(db, project_number)
    import_id = f"schedule-import:{project_number}:{uuid4().hex[:12]}"
    imported_at = _utcnow()
    row_docs = []
    for source_row in rows:
        normalized = _normalize_row(source_row.get("source_values") or {})
        suggestion = _build_row_suggestion(normalized, budget_lines, work_types)
        row_id = f"schedule-import-row:{import_id}:{source_row.get('row_number') or len(row_docs) + 1}"
        row_docs.append(
            {
                "row_id": row_id,
                "import_id": import_id,
                "project_number": project_number,
                "row_number": int(source_row.get("row_number") or (len(row_docs) + 1)),
                "sheet_name": _clean(source_row.get("sheet_name")),
                "page_number": source_row.get("page_number") or None,
                "source_values": _sanitize(source_row.get("source_values") or {}),
                "normalized": _sanitize(normalized),
                "suggestion": _sanitize(suggestion),
                "selected": _sanitize(
                    {
                        **normalized,
                        "budget_line_id": suggestion.get("budget_line_id") or "",
                        "customer_pay_item_number": suggestion.get("customer_pay_item_number") or normalized.get("customer_pay_item_number") or "",
                        "enterprise_work_type_id": suggestion.get("enterprise_work_type_id") or "",
                        "phase_id": suggestion.get("phase_id") or normalized.get("phase_id") or "",
                        "work_package_id": suggestion.get("work_package_id") or normalized.get("work_package_id") or "",
                        "project_cost_code": suggestion.get("project_cost_code") or normalized.get("project_cost_code") or "",
                    }
                ),
                "review_status": "review_required" if suggestion.get("warnings") else "pending_review",
                "created_at": imported_at,
                "created_by": _actor_label(actor),
                "updated_at": imported_at,
                "updated_by": _actor_label(actor),
            }
        )
    session = {
        "import_id": import_id,
        "project_number": project_number,
        "project_name": job.get("project_name") or job.get("name") or project_number,
        "status": "review_required",
        "source_kind": source_kind,
        "target_version_kind": target_version_kind,
        "version_name": _clean(version_name) or f"{job.get('project_name') or project_number} · Schedule {target_version_kind.replace('_', ' ').title()}",
        "filename": _clean_filename(filename),
        "content_type": _clean(content_type) or "application/octet-stream",
        "file_kind": file_kind,
        "file_hash": file_hash,
        "parser_name": parser_name,
        "parser_warnings": parser_warnings,
        "source_preservation": {
            "original_filename": _clean_filename(filename),
            "sha256": file_hash,
            "imported_at": imported_at,
            "imported_by": _actor_label(actor),
            "sample_rows": [_sanitize(row.get("source_values") or {}) for row in row_docs[:3]],
        },
        "advisory_engine": {"mode": "deterministic_governed_suggestions", "auto_approval": False},
        "extension_readiness": {
            "csv_runtime_certified": True,
            "excel_foundation_ready": True,
            "pdf_review_foundation_ready": True,
            "primavera_extension_ready": True,
            "ms_project_extension_ready": True,
        },
        "total_rows": len(row_docs),
        "approved_rows": 0,
        "rejected_rows": 0,
        "needs_review_rows": len(row_docs),
        "activated_rows": 0,
        "imported_at": imported_at,
        "imported_by": _actor_label(actor),
        "updated_at": imported_at,
        "updated_by": _actor_label(actor),
    }
    await db[COLL_SCHEDULE_IMPORTS].insert_one(session)
    if row_docs:
        await db[COLL_SCHEDULE_IMPORT_ROWS].insert_many(row_docs)
    await _upsert_review_item(
        db,
        {
            "review_id": f"schedule-review:import:{import_id}",
            "project_number": project_number,
            "status": "review_required",
            "priority": 95,
            "source_kind": "schedule_import_session",
            "source_record_id": import_id,
            "title": f"Schedule import review required for {project_number}",
            "reason": "Imported activities require governed mapping review before the master schedule can activate.",
            "confidence": "human_required",
            "provenance": {"source_kind": source_kind, "target_version_kind": target_version_kind, "filename": _clean_filename(filename)},
        },
    )
    await _write_audit(db, "schedule_import_staged", actor, "schedule_import", import_id, session)
    return await get_schedule_import_detail(db, project_number, import_id)


def _selected_payload_for_row(row: Dict[str, Any], payload: Dict[str, Any]) -> Dict[str, Any]:
    normalized = row.get("normalized") or {}
    suggestion = row.get("suggestion") or {}
    selected = deepcopy(row.get("selected") or {})
    selected.update(
        {
            "activity_id": _clean(payload.get("activity_id") or selected.get("activity_id") or normalized.get("activity_id")),
            "activity_name": _clean(payload.get("activity_name") or selected.get("activity_name") or normalized.get("activity_name")),
            "phase_id": _clean(payload.get("phase_id") or selected.get("phase_id") or suggestion.get("phase_id") or normalized.get("phase_id")),
            "work_package_id": _clean(payload.get("work_package_id") or selected.get("work_package_id") or suggestion.get("work_package_id") or normalized.get("work_package_id")),
            "budget_line_id": _clean(payload.get("budget_line_id") or selected.get("budget_line_id") or suggestion.get("budget_line_id")),
            "customer_pay_item_number": _clean(payload.get("customer_pay_item_number") or selected.get("customer_pay_item_number") or suggestion.get("customer_pay_item_number") or normalized.get("customer_pay_item_number")),
            "enterprise_work_type_id": _clean(payload.get("enterprise_work_type_id") or selected.get("enterprise_work_type_id") or suggestion.get("enterprise_work_type_id")),
            "project_cost_code": _clean(payload.get("project_cost_code") or selected.get("project_cost_code") or suggestion.get("project_cost_code") or normalized.get("project_cost_code")),
            "planned_start_date": _parse_date(payload.get("planned_start_date") or selected.get("planned_start_date") or normalized.get("planned_start_date")),
            "planned_finish_date": _parse_date(payload.get("planned_finish_date") or selected.get("planned_finish_date") or normalized.get("planned_finish_date")),
            "duration_days": max(1, _safe_int(payload.get("duration_days") if payload.get("duration_days") is not None else selected.get("duration_days", normalized.get("duration_days")), 1)),
            "predecessor_activity_ids": _coerce_list(payload.get("predecessor_activity_ids") if payload.get("predecessor_activity_ids") is not None else selected.get("predecessor_activity_ids", normalized.get("predecessor_activity_ids"))),
            "calendar_name": _clean(payload.get("calendar_name") or selected.get("calendar_name") or normalized.get("calendar_name") or "Default"),
            "status": _clean(payload.get("status") or selected.get("status") or normalized.get("status") or "not_started"),
            "percent_complete": max(0.0, min(100.0, _safe_float(payload.get("percent_complete") if payload.get("percent_complete") is not None else selected.get("percent_complete", normalized.get("percent_complete"))))),
            "owner": _clean(payload.get("owner") or selected.get("owner") or normalized.get("owner")),
            "priority": _normalize_priority(payload.get("priority") or selected.get("priority") or normalized.get("priority") or "normal"),
            "notes": _clean(payload.get("notes") or selected.get("notes") or normalized.get("notes")),
            "execution_strategy": _normalize_execution_strategy(payload.get("execution_strategy") or selected.get("execution_strategy") or normalized.get("execution_strategy") or "self_perform"),
            "planned_crew_ids": _normalize_resource_refs(payload.get("planned_crew_ids") if payload.get("planned_crew_ids") is not None else payload.get("planned_crew") if payload.get("planned_crew") is not None else selected.get("planned_crew_ids", normalized.get("planned_crew_ids")), kind="crew"),
            "planned_employee_ids": _normalize_resource_refs(payload.get("planned_employee_ids") if payload.get("planned_employee_ids") is not None else payload.get("planned_employees") if payload.get("planned_employees") is not None else selected.get("planned_employee_ids", normalized.get("planned_employee_ids")), kind="employee"),
            "planned_equipment_ids": _normalize_resource_refs(payload.get("planned_equipment_ids") if payload.get("planned_equipment_ids") is not None else payload.get("planned_equipment") if payload.get("planned_equipment") is not None else selected.get("planned_equipment_ids", normalized.get("planned_equipment_ids")), kind="equipment"),
            "planned_materials": _normalize_material_refs(payload.get("planned_materials") if payload.get("planned_materials") is not None else selected.get("planned_materials", normalized.get("planned_materials"))),
            "planned_vendor_refs": _normalize_vendor_refs(payload.get("planned_vendor_refs") if payload.get("planned_vendor_refs") is not None else payload.get("planned_vendors") if payload.get("planned_vendors") is not None else selected.get("planned_vendor_refs", normalized.get("planned_vendor_refs"))),
            "planned_subcontractor_refs": _normalize_subcontractor_refs(payload.get("planned_subcontractor_refs") if payload.get("planned_subcontractor_refs") is not None else payload.get("planned_subcontractors") if payload.get("planned_subcontractors") is not None else selected.get("planned_subcontractor_refs", normalized.get("planned_subcontractor_refs"))),
            "planned_production_quantity": _safe_float(payload.get("planned_production_quantity") if payload.get("planned_production_quantity") is not None else selected.get("planned_production_quantity", normalized.get("planned_production_quantity"))),
            "planned_hours": _safe_float(payload.get("planned_hours") if payload.get("planned_hours") is not None else selected.get("planned_hours", normalized.get("planned_hours"))),
            "planned_constraints": _constraint_tokens(payload.get("planned_constraints") if payload.get("planned_constraints") is not None else selected.get("planned_constraints", normalized.get("planned_constraints"))),
            "review_note": _clean(payload.get("review_note") or selected.get("review_note")),
        }
    )
    if not selected.get("planned_finish_date") and selected.get("planned_start_date"):
        try:
            start = date.fromisoformat(selected["planned_start_date"])
            selected["planned_finish_date"] = (start + timedelta(days=max(1, int(selected.get("duration_days") or 1)) - 1)).isoformat()
        except Exception:
            selected["planned_finish_date"] = ""
    return selected


async def review_schedule_import_row(db, project_number: str, import_id: str, row_id: str, payload: Dict[str, Any], *, actor: Dict[str, Any]) -> Dict[str, Any]:
    await ensure_project_schedule_foundation(db)
    session = await db[COLL_SCHEDULE_IMPORTS].find_one({"project_number": project_number, "import_id": import_id}, {"_id": 0})
    if not session:
        raise LookupError("schedule_import_not_found")
    row = await db[COLL_SCHEDULE_IMPORT_ROWS].find_one({"project_number": project_number, "import_id": import_id, "row_id": row_id}, {"_id": 0})
    if not row:
        raise LookupError("schedule_import_row_not_found")
    action = _clean(payload.get("action") or "approve").lower()
    if action not in {"approve", "reject", "needs_review"}:
        raise ValueError("schedule_row_action_invalid")
    selected = _selected_payload_for_row(row, payload)
    if action == "approve":
        required = [
            ("activity_id", selected.get("activity_id")),
            ("activity_name", selected.get("activity_name")),
            ("phase_id", selected.get("phase_id")),
            ("work_package_id", selected.get("work_package_id")),
            ("budget_line_id", selected.get("budget_line_id")),
            ("customer_pay_item_number", selected.get("customer_pay_item_number")),
            ("enterprise_work_type_id", selected.get("enterprise_work_type_id")),
            ("project_cost_code", selected.get("project_cost_code")),
            ("planned_start_date", selected.get("planned_start_date")),
        ]
        missing = [field for field, value in required if not _clean(value)]
        if missing:
            raise ValueError(f"schedule_row_missing_required:{','.join(missing)}")
    updated = deepcopy(row)
    updated["selected"] = _sanitize(selected)
    updated["review_status"] = "approved" if action == "approve" else "rejected" if action == "reject" else "review_required"
    updated["review_note"] = _clean(payload.get("review_note") or "")
    updated["reviewed_at"] = _utcnow()
    updated["reviewed_by"] = _actor_label(actor)
    updated["updated_at"] = updated["reviewed_at"]
    updated["updated_by"] = _actor_label(actor)
    await db[COLL_SCHEDULE_IMPORT_ROWS].replace_one({"row_id": row_id}, updated, upsert=True)
    review_id = f"schedule-review:row:{row_id}"
    if updated["review_status"] == "approved":
        await _mark_review_resolved(db, review_id, actor=actor, resolution_note="PM approved schedule row for activation.")
    else:
        await _upsert_review_item(
            db,
            {
                "review_id": review_id,
                "project_number": project_number,
                "status": "review_required",
                "priority": 90,
                "source_kind": "schedule_import_row",
                "source_record_id": row_id,
                "title": f"Schedule row {row.get('row_number')} needs governed review",
                "reason": _clean(payload.get("review_note")) or "PM held this activity row for additional review.",
                "confidence": "human_required",
                "provenance": {"import_id": import_id, "row_number": row.get("row_number"), "selected": _sanitize(selected)},
            },
        )
    await _refresh_import_counts(db, import_id)
    await _write_audit(db, "schedule_activity_reviewed", actor, "schedule_import_row", row_id, updated, before=row)
    return _sanitize(updated)


def _schedule_version_doc(project_number: str, project_name: str, payload: Dict[str, Any], *, actor: Dict[str, Any], source_import_id: str, baseline_version_id: str = "", parent_version_id: str = "") -> Dict[str, Any]:
    now = _utcnow()
    version_kind = _status(payload.get("version_kind") or "master_schedule", allowed=VERSION_KINDS, default="master_schedule")
    version_id = payload.get("version_id") or f"schedule-version:{project_number}:{_norm(version_kind)}:{uuid4().hex[:8]}"
    return {
        "version_id": version_id,
        "project_number": project_number,
        "project_name": project_name,
        "version_kind": version_kind,
        "version_name": _clean(payload.get("version_name") or f"{project_name} · {version_kind.replace('_', ' ').title()}"),
        "status": "active",
        "source_import_id": source_import_id,
        "baseline_version_id": _clean(baseline_version_id) or version_id,
        "parent_version_id": _clean(parent_version_id),
        "baseline_preserved": True,
        "created_at": now,
        "created_by": _actor_label(actor),
        "approved_at": now,
        "approved_by": _actor_label(actor),
        "activated_at": now,
        "activated_by": _actor_label(actor),
        "updated_at": now,
        "updated_by": _actor_label(actor),
        "counts": {
            "activity_count": 0,
            "work_package_count": 0,
            "constraint_refs": 0,
        },
    }


async def _budget_line_index(db, project_number: str) -> Dict[str, Dict[str, Any]]:
    rows = await _list_budget_lines_for_project(db, project_number)
    return {row.get("budget_line_id"): row for row in rows if row.get("budget_line_id")}


def _activity_doc(project_number: str, version_id: str, selected: Dict[str, Any], budget_line: Dict[str, Any], *, actor: Dict[str, Any], source_import_id: str, source_row_id: str, source_hash: str, baseline_version_id: str) -> Dict[str, Any]:
    now = _utcnow()
    return {
        "activity_id": _clean(selected.get("activity_id")),
        "project_number": project_number,
        "version_id": version_id,
        "baseline_version_id": baseline_version_id,
        "status": _clean(selected.get("status") or "not_started"),
        "activity_name": _clean(selected.get("activity_name")),
        "phase_id": _clean(selected.get("phase_id")),
        "work_package_id": _clean(selected.get("work_package_id")),
        "budget_line_id": _clean(selected.get("budget_line_id")),
        "customer_pay_item_number": _clean(selected.get("customer_pay_item_number") or budget_line.get("customer_pay_item_number")),
        "enterprise_work_type_id": _clean(selected.get("enterprise_work_type_id") or budget_line.get("enterprise_work_type_id")),
        "project_cost_code": _clean(selected.get("project_cost_code") or budget_line.get("project_cost_code")),
        "calendar_name": _clean(selected.get("calendar_name") or "Default"),
        "planned_start_date": _clean(selected.get("planned_start_date")),
        "planned_finish_date": _clean(selected.get("planned_finish_date")),
        "duration_days": max(1, _safe_int(selected.get("duration_days"), 1)),
        "predecessor_activity_ids": _coerce_list(selected.get("predecessor_activity_ids")),
        "owner": _clean(selected.get("owner")),
        "priority": _normalize_priority(selected.get("priority")),
        "percent_complete": max(0.0, min(100.0, _safe_float(selected.get("percent_complete")))),
        "notes": _clean(selected.get("notes")),
        "execution_strategy": _normalize_execution_strategy(selected.get("execution_strategy")),
        "planned_assignments": {
            "planned_crew_ids": _sanitize(selected.get("planned_crew_ids") or []),
            "planned_employee_ids": _sanitize(selected.get("planned_employee_ids") or []),
            "planned_equipment_ids": _sanitize(selected.get("planned_equipment_ids") or []),
            "planned_materials": _sanitize(selected.get("planned_materials") or []),
            "planned_vendor_refs": _sanitize(selected.get("planned_vendor_refs") or []),
            "planned_subcontractor_refs": _sanitize(selected.get("planned_subcontractor_refs") or []),
            "planned_production_quantity": _safe_float(selected.get("planned_production_quantity")),
            "planned_hours": _safe_float(selected.get("planned_hours")),
            "planned_constraints": _sanitize(selected.get("planned_constraints") or []),
        },
        "actual_links": {
            "work_block_ids": [],
            "daily_report_ids": [],
            "production_rows": 0,
        },
        "future_rollups": {
            "crew_cost": 0.0,
            "crew_productivity": 0.0,
            "crew_production": 0.0,
            "cost_per_unit": 0.0,
            "revenue_per_crew": 0.0,
            "margin": 0.0,
            "equipment_hours": 0.0,
            "idle_time": 0.0,
            "utilization": 0.0,
            "ownership_cost": 0.0,
            "rental_cost": 0.0,
            "material_delivered": 0.0,
            "material_installed": 0.0,
            "material_returned": 0.0,
            "material_waste": 0.0,
        },
        "source_import_id": source_import_id,
        "source_row_id": source_row_id,
        "source_document_hash": source_hash,
        "created_at": now,
        "created_by": _actor_label(actor),
        "updated_at": now,
        "updated_by": _actor_label(actor),
    }


def _assignment_projection(activity: Dict[str, Any]) -> Dict[str, Any]:
    planned = activity.get("planned_assignments") or {}
    notes = activity.get("notes") or ""
    crew_rows = planned.get("planned_crew_ids") or []
    planned_performer = activity.get("owner") or (crew_rows[0].get("label", "") if crew_rows else "")
    return normalize_job_assignment(
        {
            "code": activity.get("project_cost_code") or activity.get("activity_id"),
            "item_name": activity.get("activity_name") or activity.get("activity_id"),
            "unit": "LS",
            "authorized_quantity": planned.get("planned_production_quantity") or 1,
            "original_quantity": planned.get("planned_production_quantity") or 1,
            "forecast_quantity": planned.get("planned_production_quantity") or 1,
            "cpm_activity_id": activity.get("activity_id"),
            "cpm_activity_name": activity.get("activity_name"),
            "schedule_phase": activity.get("phase_id"),
            "schedule_start_date": activity.get("planned_start_date"),
            "duration_days": activity.get("duration_days") or 1,
            "predecessor_codes": activity.get("predecessor_activity_ids") or [],
            "planned_performer": planned_performer,
            "planned_equipment_units": [row.get("label") for row in planned.get("planned_equipment_ids") or [] if _clean(row.get("label"))],
            "resource_demand": {
                "labor_hours": planned.get("planned_hours") or 0,
                "required_materials": [row.get("description") for row in planned.get("planned_materials") or [] if _clean(row.get("description"))],
                "required_subcontractors": [row.get("subcontractor_name") for row in planned.get("planned_subcontractor_refs") or [] if _clean(row.get("subcontractor_name"))],
                "required_equipment_units": [row.get("label") for row in planned.get("planned_equipment_ids") or [] if _clean(row.get("label"))],
            },
            "notes": notes,
            "phase_id": activity.get("phase_id"),
            "work_package_id": activity.get("work_package_id"),
            "budget_line_id": activity.get("budget_line_id"),
            "customer_pay_item_number": activity.get("customer_pay_item_number"),
            "enterprise_work_type_id": activity.get("enterprise_work_type_id"),
            "project_cost_code": activity.get("project_cost_code"),
            "calendar_name": activity.get("calendar_name"),
            "schedule_status": activity.get("status"),
            "percent_complete": activity.get("percent_complete") or 0,
            "planned_crew_ids": [row.get("label") for row in planned.get("planned_crew_ids") or [] if _clean(row.get("label"))],
            "planned_employee_ids": [row.get("label") for row in planned.get("planned_employee_ids") or [] if _clean(row.get("label"))],
            "planned_materials": [row.get("description") for row in planned.get("planned_materials") or [] if _clean(row.get("description"))],
            "planned_vendor_refs": [row.get("vendor_name") for row in planned.get("planned_vendor_refs") or [] if _clean(row.get("vendor_name"))],
            "planned_subcontractor_refs": [row.get("subcontractor_name") for row in planned.get("planned_subcontractor_refs") or [] if _clean(row.get("subcontractor_name"))],
            "planned_constraints": [row.get("title") for row in planned.get("planned_constraints") or [] if _clean(row.get("title"))],
            "planned_production_quantity": activity.get("planned_assignments", {}).get("planned_production_quantity") or 0,
            "planned_hours": activity.get("planned_assignments", {}).get("planned_hours") or 0,
            "execution_strategy": activity.get("execution_strategy"),
        }
    )


def _work_package_doc(project_number: str, version_id: str, phase_id: str, work_package_id: str, activities: List[Dict[str, Any]], *, actor: Dict[str, Any]) -> Dict[str, Any]:
    now = _utcnow()
    budget_line_ids = sorted({row.get("budget_line_id") for row in activities if _clean(row.get("budget_line_id"))})
    cost_codes = sorted({row.get("project_cost_code") for row in activities if _clean(row.get("project_cost_code"))})
    customer_items = sorted({row.get("customer_pay_item_number") for row in activities if _clean(row.get("customer_pay_item_number"))})
    work_types = sorted({row.get("enterprise_work_type_id") for row in activities if _clean(row.get("enterprise_work_type_id"))})
    start_dates = sorted([row.get("planned_start_date") for row in activities if _clean(row.get("planned_start_date"))])
    finish_dates = sorted([row.get("planned_finish_date") for row in activities if _clean(row.get("planned_finish_date"))])
    crew_refs = [ref for row in activities for ref in (row.get("planned_assignments", {}).get("planned_crew_ids") or [])]
    equipment_refs = [ref for row in activities for ref in (row.get("planned_assignments", {}).get("planned_equipment_ids") or [])]
    material_refs = [ref for row in activities for ref in (row.get("planned_assignments", {}).get("planned_materials") or [])]
    vendor_refs = [ref for row in activities for ref in (row.get("planned_assignments", {}).get("planned_vendor_refs") or [])]
    subcontractor_refs = [ref for row in activities for ref in (row.get("planned_assignments", {}).get("planned_subcontractor_refs") or [])]
    constraint_refs = [ref for row in activities for ref in (row.get("planned_assignments", {}).get("planned_constraints") or [])]
    planned_hours = round(sum(_safe_float((row.get("planned_assignments") or {}).get("planned_hours")) for row in activities), 4)
    planned_production_quantity = round(sum(_safe_float((row.get("planned_assignments") or {}).get("planned_production_quantity")) for row in activities), 4)
    return {
        "work_package_id": work_package_id,
        "project_number": project_number,
        "version_id": version_id,
        "phase_id": phase_id,
        "title": work_package_id,
        "status": "active",
        "activity_count": len(activities),
        "activity_ids": [row.get("activity_id") for row in activities],
        "budget_line_ids": budget_line_ids,
        "cost_codes": cost_codes,
        "customer_pay_item_numbers": customer_items,
        "enterprise_work_type_ids": work_types,
        "planned_start_date": start_dates[0] if start_dates else "",
        "planned_finish_date": finish_dates[-1] if finish_dates else "",
        "planned_hours": planned_hours,
        "planned_production_quantity": planned_production_quantity,
        "planned_assignments": {
            "crew": _sanitize(crew_refs),
            "equipment": _sanitize(equipment_refs),
            "materials": _sanitize(material_refs),
            "vendors": _sanitize(vendor_refs),
            "subcontractors": _sanitize(subcontractor_refs),
            "constraints": _sanitize(constraint_refs),
        },
        "actual_links": {"daily_report_ids": [], "work_block_ids": []},
        "future_rollups": {"forecast_refs": [], "commitment_refs": [], "executive_rollups": []},
        "created_at": now,
        "created_by": _actor_label(actor),
        "updated_at": now,
        "updated_by": _actor_label(actor),
    }


async def _seed_lookahead_from_activities(db, project_number: str, activities: List[Dict[str, Any]], *, actor: Dict[str, Any]) -> Dict[str, Any]:
    existing = await get_project_lookahead(db, project_number)
    if existing.get("tasks") and any(task.get("activity_id") for task in existing.get("tasks") or []):
        return existing
    focus_tasks = []
    for activity in sorted(activities, key=lambda row: (row.get("planned_start_date") or "", row.get("activity_id") or ""))[:40]:
        planned = activity.get("planned_assignments") or {}
        focus_tasks.append(
            {
                "code": activity.get("project_cost_code") or activity.get("activity_id"),
                "activity_id": activity.get("activity_id"),
                "title": activity.get("activity_name"),
                "phase_id": activity.get("phase_id"),
                "work_package_id": activity.get("work_package_id"),
                "budget_line_id": activity.get("budget_line_id"),
                "customer_pay_item_number": activity.get("customer_pay_item_number"),
                "enterprise_work_type_id": activity.get("enterprise_work_type_id"),
                "planned_start": activity.get("planned_start_date"),
                "planned_finish": activity.get("planned_finish_date"),
                "planned_crews": _sanitize(planned.get("planned_crew_ids") or []),
                "planned_equipment": _sanitize(planned.get("planned_equipment_ids") or []),
                "planned_materials": _sanitize(planned.get("planned_materials") or []),
                "planned_vendors": _sanitize(planned.get("planned_vendor_refs") or []),
                "planned_subcontractors": _sanitize(planned.get("planned_subcontractor_refs") or []),
                "planned_constraints": _sanitize(planned.get("planned_constraints") or []),
                "responsible_party": activity.get("owner") or "PM",
                "notes": activity.get("notes") or "",
            }
        )
    payload = {
        **existing,
        "tasks": focus_tasks,
        "constraints": [constraint for row in focus_tasks for constraint in row.get("planned_constraints") or []],
        "comparison_note": "Lookahead stays a governed operational view of the active master schedule. Resequencing here does not overwrite the baseline schedule.",
    }
    return await save_project_lookahead(db, project_number, payload, actor=actor)


async def activate_schedule_import_session(db, project_number: str, import_id: str, *, actor: Dict[str, Any]) -> Dict[str, Any]:
    await ensure_project_schedule_foundation(db)
    session = await db[COLL_SCHEDULE_IMPORTS].find_one({"project_number": project_number, "import_id": import_id}, {"_id": 0})
    if not session:
        raise LookupError("schedule_import_not_found")
    rows = [row async for row in db[COLL_SCHEDULE_IMPORT_ROWS].find({"project_number": project_number, "import_id": import_id}, {"_id": 0}).sort([("row_number", 1)])]
    if not rows:
        raise ValueError("schedule_import_empty")
    pending = [row for row in rows if row.get("review_status") not in {"approved", "rejected"}]
    if pending:
        raise ValueError("schedule_import_review_incomplete")
    approved_rows = [row for row in rows if row.get("review_status") == "approved"]
    if not approved_rows:
        raise ValueError("schedule_import_has_no_approved_rows")
    budget_lines = await _budget_line_index(db, project_number)
    existing_versions = await list_schedule_versions(db, project_number)
    active_version = next((row for row in existing_versions if row.get("status") == "active"), None)
    baseline_version_id = active_version.get("baseline_version_id") if active_version else ""
    version = _schedule_version_doc(
        project_number,
        (_load := await _load_job(db, project_number)).get("project_name") or _load.get("name") or project_number,
        {"version_kind": session.get("target_version_kind"), "version_name": session.get("version_name")},
        actor=actor,
        source_import_id=import_id,
        baseline_version_id=baseline_version_id or "",
        parent_version_id=(active_version or {}).get("version_id") or "",
    )
    if active_version:
        await db[COLL_SCHEDULE_VERSIONS].update_many({"project_number": project_number, "status": "active"}, {"$set": {"status": "superseded", "updated_at": _utcnow(), "updated_by": _actor_label(actor)}})
    await db[COLL_SCHEDULE_VERSIONS].replace_one({"project_number": project_number, "version_id": version["version_id"]}, version, upsert=True)
    activity_docs = []
    work_package_groups: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for row in approved_rows:
        selected = row.get("selected") or {}
        budget_line = budget_lines.get(selected.get("budget_line_id"))
        if not budget_line:
            raise ValueError("schedule_budget_line_not_found")
        activity = _activity_doc(
            project_number,
            version["version_id"],
            selected,
            budget_line,
            actor=actor,
            source_import_id=import_id,
            source_row_id=row.get("row_id") or uuid4().hex,
            source_hash=session.get("file_hash") or "",
            baseline_version_id=version.get("baseline_version_id") or version["version_id"],
        )
        activity_docs.append(activity)
        work_package_groups[(activity.get("phase_id") or "", activity.get("work_package_id") or activity.get("activity_id") or "UNASSIGNED")].append(activity)
    if activity_docs:
        await db[COLL_SCHEDULE_ACTIVITIES].insert_many(activity_docs)
    work_package_docs = []
    for (phase_id, work_package_id), activities in work_package_groups.items():
        doc = _work_package_doc(project_number, version["version_id"], phase_id, work_package_id, activities, actor=actor)
        work_package_docs.append(doc)
        await db[COLL_WORK_PACKAGES].replace_one({"project_number": project_number, "version_id": version["version_id"], "work_package_id": work_package_id}, doc, upsert=True)
    assignments = [_assignment_projection(activity) for activity in activity_docs]
    await persist_project_assignments(db, project_number, assignments)
    await db.jobs_master.update_one(
        {"project_number": project_number},
        {
            "$set": {
                "oppc_schedule_spine": {
                    "active_version_id": version["version_id"],
                    "baseline_version_id": version.get("baseline_version_id") or version["version_id"],
                    "work_package_count": len(work_package_docs),
                    "activity_count": len(activity_docs),
                    "updated_at": _utcnow(),
                },
                "updated_at": _utcnow(),
            }
        },
        upsert=False,
    )
    version["counts"] = {"activity_count": len(activity_docs), "work_package_count": len(work_package_docs), "constraint_refs": sum(len((activity.get("planned_assignments") or {}).get("planned_constraints") or []) for activity in activity_docs)}
    version["updated_at"] = _utcnow()
    version["updated_by"] = _actor_label(actor)
    await db[COLL_SCHEDULE_VERSIONS].replace_one({"project_number": project_number, "version_id": version["version_id"]}, version, upsert=True)
    lookahead = await _seed_lookahead_from_activities(db, project_number, activity_docs, actor=actor)
    for row in approved_rows:
        await db[COLL_SCHEDULE_IMPORT_ROWS].update_one({"row_id": row["row_id"]}, {"$set": {"review_status": "activated", "activated_version_id": version["version_id"], "activated_at": _utcnow(), "updated_at": _utcnow(), "updated_by": _actor_label(actor)}})
        await _mark_review_resolved(db, f"schedule-review:row:{row['row_id']}", actor=actor, resolution_note="Schedule row activated into governed schedule version.")
    await db[COLL_SCHEDULE_IMPORTS].update_one({"import_id": import_id}, {"$set": {"status": "activated", "activated_version_id": version["version_id"], "activated_at": _utcnow(), "activated_by": _actor_label(actor), "updated_at": _utcnow(), "updated_by": _actor_label(actor)}})
    await _mark_review_resolved(db, f"schedule-review:import:{import_id}", actor=actor, resolution_note="Schedule import session activated after PM approval.")
    await _refresh_import_counts(db, import_id)
    await _write_audit(db, "schedule_version_activated", actor, "schedule_version", version["version_id"], version, metadata={"activity_count": len(activity_docs)})
    return {"version": _sanitize(version), "activity_count": len(activity_docs), "work_package_count": len(work_package_docs), "lookahead": _sanitize(lookahead)}


async def _active_schedule_version(db, project_number: str) -> Optional[Dict[str, Any]]:
    return await db[COLL_SCHEDULE_VERSIONS].find_one({"project_number": project_number, "status": "active"}, {"_id": 0}, sort=[("activated_at", -1)])


def _two_week_window_rows(activities: List[Dict[str, Any]], *, days: int) -> List[Dict[str, Any]]:
    anchor = datetime.now(timezone.utc).date()
    end = anchor + timedelta(days=max(days, 1))
    rows = []
    for row in activities:
        start = _parse_date(row.get("planned_start_date"))
        if start and anchor.isoformat() <= start <= end.isoformat():
            rows.append(row)
    return rows


async def get_schedule_spine_overview(db, project_number: str) -> Dict[str, Any]:
    await ensure_project_schedule_foundation(db)
    from services.project_schedule_actuals_spine import (  # noqa: PLC0415
        COLL_DAILY_WORK_PLANS,
        COLL_SCHEDULE_ACTUAL_CANDIDATES,
    )

    job, versions, imports, review_queue, budget_lines, lookahead = await asyncio.gather(
        _load_job(db, project_number),
        list_schedule_versions(db, project_number),
        list_schedule_imports(db, project_number),
        list_schedule_review_queue(db, project_number=project_number),
        _list_budget_lines_for_project(db, project_number),
        get_project_lookahead(db, project_number),
    )
    active_version = next((row for row in versions if row.get("status") == "active"), None)
    activities_task = list_schedule_activities(db, project_number, version_id=active_version["version_id"]) if active_version else asyncio.sleep(0, result=[])
    work_packages_task = list_schedule_work_packages(db, project_number, version_id=active_version["version_id"] if active_version else "")
    work_ledger_task = db.project_controls_work_ledger.find({"project_number": project_number}, {"_id": 0}).sort([("report_date", -1)]).limit(50).to_list(length=50)
    actual_candidate_count_task = db[COLL_SCHEDULE_ACTUAL_CANDIDATES].count_documents({"project_number": project_number})
    approved_actual_count_task = db[COLL_SCHEDULE_ACTUAL_CANDIDATES].count_documents({"project_number": project_number, "review_status": "approved"})
    daily_plan_count_task = db[COLL_DAILY_WORK_PLANS].count_documents({"project_number": project_number})
    activities, work_packages, work_ledger_rows, actual_candidate_count, approved_actual_count, daily_plan_count = await asyncio.gather(
        activities_task,
        work_packages_task,
        work_ledger_task,
        actual_candidate_count_task,
        approved_actual_count_task,
        daily_plan_count_task,
    )
    if active_version and activities:
        lookahead = await _seed_lookahead_from_activities(db, project_number, activities, actor={"email": "system", "role": "system"})
    work_ledger_rows = [_sanitize(row) for row in work_ledger_rows]
    actuals_summary = {
        "summary": {
            "candidates": int(actual_candidate_count or 0),
            "approved": int(approved_actual_count or 0),
            "review_required": max(int(actual_candidate_count or 0) - int(approved_actual_count or 0), 0),
            "daily_work_plans": int(daily_plan_count or 0),
        },
        "candidates": [],
        "daily_work_plans": [],
    }
    actual_chain = {
        "work_block_links": sum(1 for row in work_ledger_rows if _clean(row.get("schedule_activity_id"))),
        "daily_report_rows": len({row.get("source_report_id") for row in work_ledger_rows if row.get("source_report_id")}),
        "production_rows": sum(1 for row in work_ledger_rows if _to_float(row.get("installed_quantity"), 0.0) > 0),
        "candidate_rows": int(actual_candidate_count or 0),
        "approved_actuals": int(approved_actual_count or 0),
    }
    return {
        "project": {"project_number": project_number, "project_name": job.get("project_name") or job.get("name") or project_number},
        "authority_boundaries": {
            "project_identity": "jobs_master",
            "budget_truth": COLL_BUDGET_LINES,
            "schedule_versions": COLL_SCHEDULE_VERSIONS,
            "schedule_activities": COLL_SCHEDULE_ACTIVITIES,
            "work_packages": COLL_WORK_PACKAGES,
            "lookahead": "project_lookahead",
            "daily_field_actuals": "daily_reports",
            "operational_work_blocks": "project_controls_work_ledger",
            "constraints_truth": "operational_constraints",
            "ai_role": "advisory_only",
        },
        "counts": {
            "versions": len(versions),
            "activities": len(activities),
            "work_packages": len(work_packages),
            "imports": len(imports),
            "review_queue_open": sum(1 for row in review_queue if row.get("status") != "resolved"),
            "budget_lines": len(budget_lines),
            "actual_work_blocks": actual_chain["work_block_links"],
            "schedule_actual_candidates": int(actual_candidate_count or 0),
            "approved_schedule_actuals": int(approved_actual_count or 0),
            "daily_work_plans": int(daily_plan_count or 0),
        },
        "active_version": active_version,
        "versions": versions[:8],
        "activities": activities[:50],
        "work_packages": work_packages[:50],
        "imports": imports[:10],
        "review_queue": review_queue[:20],
        "budget_lines": budget_lines[:50],
        "lookahead": lookahead,
        "daily_work_plan": {},
        "lookahead_2w": _two_week_window_rows(activities, days=14),
        "lookahead_4w": _two_week_window_rows(activities, days=28),
        "actual_chain": actual_chain,
        "schedule_actuals": actuals_summary,
        "event_contracts": EVENT_CONTRACTS,
        "backfill": _sanitize(await db[COLL_SCHEDULE_RUNS].find_one({"run_type": "wp18c4_backfill"}, {"_id": 0}) or {"run_type": "wp18c4_backfill", "status": "pending_manual_run"}),
    }


async def get_admin_schedule_spine_overview(db, project_number: str = "") -> Dict[str, Any]:
    await ensure_project_schedule_foundation(db)
    from services.project_schedule_actuals_spine import get_admin_schedule_actuals_overview  # noqa: PLC0415

    query = {"project_number": project_number} if project_number else {}
    versions = [_sanitize(row) async for row in db[COLL_SCHEDULE_VERSIONS].find(query, {"_id": 0}).sort([("activated_at", -1)]).limit(100)]
    imports = [_sanitize(row) async for row in db[COLL_SCHEDULE_IMPORTS].find(query, {"_id": 0}).sort([("imported_at", -1)]).limit(100)]
    reviews = await list_schedule_review_queue(db, project_number=project_number)
    work_packages = [_sanitize(row) async for row in db[COLL_WORK_PACKAGES].find(query, {"_id": 0}).sort([("updated_at", -1)]).limit(100)]
    actuals = await get_admin_schedule_actuals_overview(db, project_number=project_number)
    return {
        "summary": {
            "projects_with_versions": len({row.get("project_number") for row in versions if row.get("project_number")}),
            "schedule_versions": await db[COLL_SCHEDULE_VERSIONS].count_documents(query),
            "schedule_activities": await db[COLL_SCHEDULE_ACTIVITIES].count_documents(query),
            "work_packages": await db[COLL_WORK_PACKAGES].count_documents(query),
            "imports": await db[COLL_SCHEDULE_IMPORTS].count_documents(query),
            "review_queue_open": sum(1 for row in reviews if row.get("status") != "resolved"),
            "schedule_actual_candidates": (actuals.get("summary") or {}).get("candidates") or 0,
            "approved_schedule_actuals": (actuals.get("summary") or {}).get("approved") or 0,
            "daily_work_plans": (actuals.get("summary") or {}).get("daily_work_plans") or 0,
        },
        "versions": versions,
        "imports": imports,
        "work_packages": work_packages,
        "review_queue": reviews[:100],
        "schedule_actuals": actuals,
        "backfill": _sanitize(await db[COLL_SCHEDULE_RUNS].find_one({"run_type": "wp18c4_backfill"}, {"_id": 0}) or {"run_type": "wp18c4_backfill", "status": "pending_manual_run"}),
        "event_contracts": EVENT_CONTRACTS,
    }


async def save_schedule_lookahead(db, project_number: str, payload: Dict[str, Any], *, actor: Dict[str, Any]) -> Dict[str, Any]:
    await ensure_project_schedule_foundation(db)
    row = await save_project_lookahead(db, project_number, payload, actor=actor)
    await _write_audit(db, "schedule_lookahead_saved", actor, "project_lookahead", row.get("lookahead_id") or f"lookahead:{project_number}:current", row)
    return row


async def record_schedule_distribution_event(db, *, project_number: str, actor: Dict[str, Any], export_kind: str, version_id: str = "", metadata: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    row = {
        "distribution_id": f"schedule-distribution:{project_number}:{uuid4().hex[:10]}",
        "project_number": project_number,
        "version_id": _clean(version_id),
        "export_kind": _clean(export_kind),
        "actor": _sanitize(actor or {}),
        "metadata": _sanitize(metadata or {}),
        "created_at": _utcnow(),
    }
    await db[COLL_SCHEDULE_DISTRIBUTION].insert_one(row)
    return _sanitize(row)


def _csv_payload(filename: str, header: List[str], rows: List[List[Any]]) -> Dict[str, Any]:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(header)
    writer.writerows(rows)
    return {"filename": filename, "content": buffer.getvalue()}


def _xlsx_payload(filename: str, header: List[str], rows: List[List[Any]]) -> Dict[str, Any]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.append(header)
    for row in rows:
        ws.append(row)
    output = io.BytesIO()
    wb.save(output)
    return {"filename": filename, "content": output.getvalue()}


def _assignment_export_rows(activities: List[Dict[str, Any]], *, assignment_key: str, id_key: str, label_key: str) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for activity in activities:
        planned = activity.get("planned_assignments") or {}
        refs = planned.get(assignment_key) or []
        for ref in refs:
            rows.append(
                [
                    activity.get("work_package_id"),
                    activity.get("activity_id"),
                    activity.get("activity_name"),
                    ref.get(id_key) or "",
                    ref.get(label_key) or "",
                    activity.get("planned_start_date"),
                    activity.get("planned_finish_date"),
                    planned.get("planned_hours") or 0,
                    planned.get("planned_production_quantity") or 0,
                    activity.get("status"),
                    activity.get("priority"),
                ]
            )
    return rows


def _material_export_rows(activities: List[Dict[str, Any]]) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for activity in activities:
        planned = activity.get("planned_assignments") or {}
        for ref in planned.get("planned_materials") or []:
            rows.append(
                [
                    activity.get("work_package_id"),
                    activity.get("activity_id"),
                    activity.get("activity_name"),
                    ref.get("material_id") or "",
                    ref.get("description") or "",
                    ref.get("quantity") or 0,
                    ref.get("unit") or "",
                    activity.get("planned_start_date"),
                    activity.get("planned_finish_date"),
                    activity.get("status"),
                ]
            )
    return rows


def _work_package_export_rows(activities: List[Dict[str, Any]]) -> List[List[Any]]:
    grouped: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for activity in activities:
        grouped[(activity.get("phase_id") or "", activity.get("work_package_id") or activity.get("activity_id") or "UNASSIGNED")].append(activity)
    rows: List[List[Any]] = []
    for (phase_id, work_package_id), package_rows in grouped.items():
        start_dates = sorted([row.get("planned_start_date") for row in package_rows if row.get("planned_start_date")])
        finish_dates = sorted([row.get("planned_finish_date") for row in package_rows if row.get("planned_finish_date")])
        planned_hours = sum(_safe_float((row.get("planned_assignments") or {}).get("planned_hours")) for row in package_rows)
        planned_qty = sum(_safe_float((row.get("planned_assignments") or {}).get("planned_production_quantity")) for row in package_rows)
        rows.append(
            [
                phase_id,
                work_package_id,
                len(package_rows),
                " | ".join(sorted({row.get("activity_id") for row in package_rows if row.get("activity_id")})),
                " | ".join(sorted({row.get("budget_line_id") for row in package_rows if row.get("budget_line_id")})),
                " | ".join(sorted({row.get("customer_pay_item_number") for row in package_rows if row.get("customer_pay_item_number")})),
                " | ".join(sorted({row.get("enterprise_work_type_id") for row in package_rows if row.get("enterprise_work_type_id")})),
                " | ".join(sorted({row.get("project_cost_code") for row in package_rows if row.get("project_cost_code")})),
                start_dates[0] if start_dates else "",
                finish_dates[-1] if finish_dates else "",
                planned_hours,
                planned_qty,
            ]
        )
    return rows


async def export_schedule_view(db, project_number: str, *, version_id: str, export_kind: str, actor: Dict[str, Any]) -> Dict[str, Any]:
    activities = await list_schedule_activities(db, project_number, version_id=version_id)
    if not activities:
        raise LookupError("schedule_export_not_found")
    kind = _clean(export_kind).lower() or "master_schedule_csv"
    if kind in {"two_week_csv", "two_week_xlsx"}:
        rows_source = _two_week_window_rows(activities, days=14)
        header = [
            "activity_id",
            "activity_name",
            "phase_id",
            "work_package_id",
            "budget_line_id",
            "customer_pay_item_number",
            "enterprise_work_type_id",
            "project_cost_code",
            "planned_start_date",
            "planned_finish_date",
            "duration_days",
            "owner",
            "priority",
            "status",
            "percent_complete",
        ]
        rows = [[row.get(key) for key in header] for row in rows_source]
    elif kind in {"four_week_csv", "four_week_xlsx"}:
        rows_source = _two_week_window_rows(activities, days=28)
        header = [
            "activity_id",
            "activity_name",
            "phase_id",
            "work_package_id",
            "budget_line_id",
            "customer_pay_item_number",
            "enterprise_work_type_id",
            "project_cost_code",
            "planned_start_date",
            "planned_finish_date",
            "duration_days",
            "owner",
            "priority",
            "status",
            "percent_complete",
        ]
        rows = [[row.get(key) for key in header] for row in rows_source]
    elif kind in {"crew_plan_csv", "crew_plan_xlsx"}:
        header = [
            "work_package_id",
            "activity_id",
            "activity_name",
            "crew_id",
            "crew_label",
            "planned_start_date",
            "planned_finish_date",
            "planned_hours",
            "planned_production_quantity",
            "status",
            "priority",
        ]
        rows = _assignment_export_rows(activities, assignment_key="planned_crew_ids", id_key="crew_id", label_key="label")
    elif kind in {"equipment_plan_csv", "equipment_plan_xlsx"}:
        header = [
            "work_package_id",
            "activity_id",
            "activity_name",
            "equipment_id",
            "equipment_label",
            "planned_start_date",
            "planned_finish_date",
            "planned_hours",
            "planned_production_quantity",
            "status",
            "priority",
        ]
        rows = _assignment_export_rows(activities, assignment_key="planned_equipment_ids", id_key="equipment_id", label_key="label")
    elif kind in {"material_plan_csv", "material_plan_xlsx"}:
        header = [
            "work_package_id",
            "activity_id",
            "activity_name",
            "material_id",
            "description",
            "quantity",
            "unit",
            "planned_start_date",
            "planned_finish_date",
            "status",
        ]
        rows = _material_export_rows(activities)
    elif kind in {"work_package_plan_csv", "work_package_plan_xlsx"}:
        header = [
            "phase_id",
            "work_package_id",
            "activity_count",
            "activity_ids",
            "budget_line_ids",
            "customer_pay_item_numbers",
            "enterprise_work_type_ids",
            "project_cost_codes",
            "planned_start_date",
            "planned_finish_date",
            "planned_hours",
            "planned_production_quantity",
        ]
        rows = _work_package_export_rows(activities)
    elif kind in {"forecast_schedule_csv", "forecast_schedule_xlsx"}:
        from services.project_schedule_actuals_spine import build_schedule_forecast_view  # noqa: PLC0415

        forecast = await build_schedule_forecast_view(db, project_number, version_id=version_id)
        header = [
            "activity_id",
            "activity_name",
            "work_package_id",
            "project_cost_code",
            "baseline_start_date",
            "baseline_finish_date",
            "current_start_date",
            "current_finish_date",
            "forecast_start_date",
            "forecast_finish_date",
            "forecast_status",
            "approved_percent_complete",
            "slip_days",
        ]
        rows = [[row.get(key) for key in header] for row in (forecast.get("rows") or [])]
    elif kind in {"schedule_actuals_csv", "schedule_actuals_xlsx"}:
        from services.project_schedule_actuals_spine import list_schedule_actual_candidates  # noqa: PLC0415

        candidates = await list_schedule_actual_candidates(db, project_number)
        header = [
            "candidate_id",
            "source_report_number",
            "report_date",
            "work_block_id",
            "resolved_activity_id",
            "resolved_activity_name",
            "review_status",
            "installed_quantity",
            "unit",
            "approved_percent_complete",
            "approved_installed_quantity",
            "approved_activity_id",
            "schedule_progress_status",
        ]
        rows = [
            [
                row.get("candidate_id"),
                row.get("source_report_number"),
                row.get("report_date"),
                row.get("work_block_id"),
                (row.get("activity_resolution") or {}).get("resolved_activity_id"),
                (row.get("activity_resolution") or {}).get("resolved_activity_name"),
                row.get("review_status"),
                (row.get("actual_facts") or {}).get("installed_quantity"),
                (row.get("actual_facts") or {}).get("unit"),
                (row.get("approved_actual") or {}).get("approved_percent_complete"),
                (row.get("approved_actual") or {}).get("approved_installed_quantity"),
                (row.get("approved_actual") or {}).get("activity_id"),
                (row.get("approved_actual") or {}).get("schedule_progress_status"),
            ]
            for row in candidates
        ]
    elif kind in {"daily_work_plan_csv", "daily_work_plan_xlsx"}:
        from services.project_schedule_actuals_spine import get_daily_work_plan  # noqa: PLC0415

        plan = await get_daily_work_plan(db, project_number)
        header = [
            "plan_item_id",
            "activity_id",
            "activity_name",
            "work_package_id",
            "budget_line_id",
            "customer_pay_item_number",
            "project_cost_code",
            "planned_quantity",
            "planned_hours",
            "actual_status",
            "approved_percent_complete",
            "daily_goal_note",
        ]
        rows = [[row.get(key) for key in header] for row in (plan.get("items") or [])]
    else:
        header = [
            "activity_id",
            "activity_name",
            "phase_id",
            "work_package_id",
            "budget_line_id",
            "customer_pay_item_number",
            "enterprise_work_type_id",
            "project_cost_code",
            "planned_start_date",
            "planned_finish_date",
            "duration_days",
            "owner",
            "priority",
            "status",
            "percent_complete",
        ]
        rows = [[row.get(key) for key in header] for row in activities]
    await record_schedule_distribution_event(db, project_number=project_number, actor=actor, export_kind=kind, version_id=version_id, metadata={"row_count": len(rows)})
    if kind.endswith("xlsx"):
        return _xlsx_payload(f"{project_number}_{kind}.xlsx", header, rows)
    return _csv_payload(f"{project_number}_{kind}.csv", header, rows)


async def queue_schedule_email_export(db, project_number: str, *, version_id: str, export_kind: str, recipients: List[str], actor: Dict[str, Any]) -> Dict[str, Any]:
    event = await record_schedule_distribution_event(
        db,
        project_number=project_number,
        actor=actor,
        export_kind="email_queue",
        version_id=version_id,
        metadata={"requested_export_kind": export_kind, "recipients": recipients, "delivery_status": "queued_review_only"},
    )
    await _upsert_review_item(
        db,
        {
            "review_id": f"schedule-review:email:{event['distribution_id']}",
            "project_number": project_number,
            "status": "review_required",
            "priority": 50,
            "source_kind": "schedule_email_export",
            "source_record_id": event["distribution_id"],
            "title": f"Schedule email export queued for {project_number}",
            "reason": "Email export architecture is queued through governance audit and not auto-sent from C4.",
            "confidence": "human_required",
            "provenance": event,
        },
    )
    return event


async def run_schedule_backfill(db, *, force: bool = False) -> Dict[str, Any]:
    await ensure_project_schedule_foundation(db)
    from services.project_schedule_actuals_spine import run_schedule_actuals_backfill  # noqa: PLC0415

    last_run = await db[COLL_SCHEDULE_RUNS].find_one({"run_type": "wp18c4_backfill"}, {"_id": 0})
    if last_run and not force:
        return _sanitize(last_run)
    foundation_reviews = 0
    assignments_seen = 0
    project_numbers = {_clean(row.get("project_number")) async for row in db.jobs_master.find({"project_number": {"$ne": ""}}, {"_id": 0, "project_number": 1})}
    for project_number in sorted(project_numbers):
        if not project_number:
            continue
        assignments = (await _load_job(db, project_number)).get("assigned_cost_codes") or []
        if assignments:
            assignments_seen += len(assignments)
        version_count = await db[COLL_SCHEDULE_VERSIONS].count_documents({"project_number": project_number})
        if assignments and version_count == 0:
            await _upsert_review_item(
                db,
                {
                    "review_id": f"schedule-review:foundation:{project_number}",
                    "project_number": project_number,
                    "status": "review_required",
                    "priority": 80,
                    "source_kind": "legacy_schedule_projection",
                    "source_record_id": project_number,
                    "title": f"Schedule spine review required for {project_number}",
                    "reason": "Legacy assigned cost-code schedule data exists, but no governed C4 schedule version has been activated yet.",
                    "confidence": "human_required",
                    "provenance": {"assignment_count": len(assignments)},
                },
            )
            foundation_reviews += 1
    actual_backfill = await run_schedule_actuals_backfill(db, force=force)
    report = {
        "run_type": "wp18c4_backfill",
        "run_id": f"wp18c4-backfill:{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}",
        "ran_at": _utcnow(),
        "force": force,
        "foundation_reviews_opened": foundation_reviews,
        "legacy_assignment_rows_observed": assignments_seen,
        "c5_actuals_backfill": actual_backfill,
        "status": "completed",
    }
    await db[COLL_SCHEDULE_RUNS].replace_one({"run_type": "wp18c4_backfill"}, report, upsert=True)
    return _sanitize(report)
