from __future__ import annotations

import csv
import hashlib
import io
import re
from copy import deepcopy
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple
from uuid import uuid4

import pdfplumber
from openpyxl import load_workbook

from services.project_controls_authority import (
    _actor_label,
    _clean,
    _fingerprint,
    _load_job,
    _norm,
    _sanitize,
    _status,
    _suggest_work_type,
    _to_float,
    _write_audit,
    ensure_project_controls_foundation,
    list_enterprise_work_types,
    list_project_mappings,
    list_project_pay_items,
    upsert_project_mapping,
    upsert_project_pay_item,
)


COLL_BUDGET_VERSIONS = "project_budget_versions"
COLL_BUDGET_LINES = "project_budget_lines"
COLL_BUDGET_IMPORTS = "project_budget_import_sessions"
COLL_BUDGET_IMPORT_ROWS = "project_budget_import_rows"
COLL_BUDGET_REVIEW = "project_budget_review_queue"
COLL_BUDGET_COMMITMENTS = "project_budget_commitment_candidates"
COLL_BUDGET_ACTUALS = "project_budget_actual_cost_candidates"
COLL_BUDGET_DISTRIBUTION = "project_budget_distribution_audit"
COLL_BUDGET_RUNS = "project_budget_runs"

VERSION_STAGES = [
    "bid",
    "awarded_contract",
    "original_approved_budget",
    "current_approved_budget",
    "pending_revision",
]

VERSION_STATUS = [
    "draft",
    "review_required",
    "approved",
    "active",
    "superseded",
    "archived",
]

LINE_KINDS = [
    "direct_cost",
    "allowance",
    "contingency",
    "management_reserve",
]

IMPORT_SOURCE_KINDS = [
    "schedule_of_values",
    "bid_tab",
    "pay_item_list",
    "engineer_bid_form",
    "csv",
    "excel",
    "pdf_review",
]

IMPORT_FILE_EXTENSIONS = {
    ".csv": "csv",
    ".xlsx": "excel",
    ".xlsm": "excel",
    ".xltx": "excel",
    ".xltm": "excel",
    ".pdf": "pdf_review",
}

BUDGET_EVENT_CONTRACTS = [
    {
        "event_key": "budget.import_staged",
        "producer": "project_budget_authority",
        "authority_owner": COLL_BUDGET_IMPORTS,
        "consumers": ["pm_project_budget", "admin_budget_governance", COLL_BUDGET_REVIEW],
        "idempotency_key": "project_number:import_id",
        "operator_visible_consequence": "Budget import is preserved for review and cannot activate silently.",
    },
    {
        "event_key": "budget.import_row_reviewed",
        "producer": "project_budget_authority",
        "authority_owner": COLL_BUDGET_IMPORT_ROWS,
        "consumers": ["pm_project_budget", "admin_budget_governance"],
        "idempotency_key": "import_id:row_id:reviewed_at",
        "operator_visible_consequence": "PM review overrides or accepts a suggestion before activation.",
    },
    {
        "event_key": "budget.version_activated",
        "producer": "project_budget_authority",
        "authority_owner": COLL_BUDGET_VERSIONS,
        "consumers": ["pm_project_budget", "admin_budget_governance", COLL_BUDGET_LINES],
        "idempotency_key": "project_number:version_id",
        "operator_visible_consequence": "Approved budget becomes active without overwriting historical budget truth.",
    },
    {
        "event_key": "budget.commitment_review_required",
        "producer": "project_budget_authority",
        "authority_owner": COLL_BUDGET_COMMITMENTS,
        "consumers": ["pm_project_budget", "admin_budget_governance", COLL_BUDGET_REVIEW],
        "idempotency_key": "project_number:source_po_id",
        "operator_visible_consequence": "Approved PO commitment must be linked through governed review, never guessed.",
    },
]

COLUMN_ALIASES = {
    "customer_pay_item_number": [
        "customer pay item",
        "customer pay item number",
        "pay item",
        "pay item number",
        "item no",
        "item number",
        "bid item",
        "sov item",
        "contract item",
        "cost code",
        "customer cost code",
        "payitem",
    ],
    "description": [
        "description",
        "item description",
        "work description",
        "scope",
        "pay item description",
        "bid item description",
    ],
    "quantity": ["qty", "quantity", "bid qty", "contract qty", "scheduled qty"],
    "unit": ["unit", "uom", "unit of measure"],
    "unit_price": ["unit price", "price", "bid unit price", "contract unit price", "rate"],
    "amount": ["amount", "value", "contract value", "line total", "extended amount", "total"],
    "project_cost_code": ["project cost code", "masci cost code", "internal cost code"],
    "phase_id": ["phase", "phase id"],
    "work_package_id": ["work package", "work package id"],
    "schedule_activity_id": ["schedule activity", "schedule activity id", "activity", "activity id"],
    "schedule_activity_name": ["schedule activity name", "activity name"],
    "line_kind": ["line kind", "budget kind", "category", "budget category"],
}


def _utcnow() -> str:
    return datetime.now(timezone.utc).isoformat()


def _safe_float(value: Any) -> float:
    if isinstance(value, str):
        value = value.replace(",", "").replace("$", "").strip()
    return round(_to_float(value, 0.0), 4)


def _clean_filename(filename: str) -> str:
    return "".join(ch if ch.isalnum() or ch in {"-", "_", "."} else "_" for ch in _clean(filename) or "upload")


def _canonical_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _clean(value).lower()).strip()


def _hash_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _extract_extension(filename: str) -> str:
    cleaned = _clean(filename)
    if "." not in cleaned:
        return ""
    return "." + cleaned.rsplit(".", 1)[-1].lower()


def _match_alias(headers: Dict[str, Any], logical_key: str) -> Any:
    aliases = COLUMN_ALIASES.get(logical_key) or []
    for alias in aliases:
        key = _canonical_key(alias)
        if key in headers and _clean(headers[key]) != "":
            return headers[key]
    return ""


def _line_kind_from_value(value: Any) -> str:
    raw = _canonical_key(str(value or ""))
    if "allowance" in raw:
        return "allowance"
    if "contingency" in raw:
        return "contingency"
    if "reserve" in raw:
        return "management_reserve"
    return "direct_cost"


def _normalize_source_row(raw_row: Dict[str, Any]) -> Dict[str, Any]:
    headers = {_canonical_key(key): value for key, value in (raw_row or {}).items()}
    customer_pay_item_number = _clean(_match_alias(headers, "customer_pay_item_number"))
    description = _clean(_match_alias(headers, "description"))
    quantity = _safe_float(_match_alias(headers, "quantity"))
    unit_price = _safe_float(_match_alias(headers, "unit_price"))
    amount_value = _match_alias(headers, "amount")
    amount = _safe_float(amount_value if _clean(amount_value) else quantity * unit_price)
    line_kind = _line_kind_from_value(_match_alias(headers, "line_kind") or description)
    normalized = {
        "customer_pay_item_number": customer_pay_item_number,
        "description": description,
        "quantity": quantity,
        "unit": _clean(_match_alias(headers, "unit")),
        "unit_price": unit_price,
        "budget_amount": amount,
        "project_cost_code": _clean(_match_alias(headers, "project_cost_code") or customer_pay_item_number),
        "phase_id": _clean(_match_alias(headers, "phase_id")),
        "work_package_id": _clean(_match_alias(headers, "work_package_id")),
        "schedule_activity_id": _clean(_match_alias(headers, "schedule_activity_id")),
        "schedule_activity_name": _clean(_match_alias(headers, "schedule_activity_name")),
        "line_kind": line_kind,
    }
    if not normalized["description"] and _clean(headers.get("raw text")):
        normalized["description"] = _clean(headers.get("raw text"))
    return normalized


def _pdf_line_to_row(line: str) -> Dict[str, Any]:
    text = _clean(line)
    if not text:
        return {}
    pattern = re.compile(
        r"^(?P<num>[A-Za-z0-9\-\.]+)\s+(?P<desc>.+?)\s+(?P<qty>\d[\d,]*(?:\.\d+)?)\s+(?P<unit>[A-Za-z/\-]+)\s+(?P<price>\$?\d[\d,]*(?:\.\d+)?)\s+(?P<amount>\$?\d[\d,]*(?:\.\d+)?)$"
    )
    match = pattern.match(text)
    if not match:
        return {"raw text": text, "description": text}
    return {
        "customer pay item": match.group("num"),
        "description": match.group("desc"),
        "qty": match.group("qty"),
        "unit": match.group("unit"),
        "unit price": match.group("price"),
        "amount": match.group("amount"),
        "raw text": text,
    }


def _parse_csv_bytes(data: bytes) -> Tuple[List[Dict[str, Any]], List[str], str]:
    warnings: List[str] = []
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
        warnings.append("CSV decoded with latin-1 fallback.")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for index, row in enumerate(reader, start=2):
        if not any(_clean(value) for value in (row or {}).values()):
            continue
        rows.append({"row_number": index, "source_values": _sanitize(row or {})})
    if not rows:
        warnings.append("No non-empty data rows were found in the CSV file.")
    return rows, warnings, "csv_dict_reader"


def _parse_excel_bytes(data: bytes) -> Tuple[List[Dict[str, Any]], List[str], str]:
    warnings: List[str] = []
    rows: List[Dict[str, Any]] = []
    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            continue
        headers = [str(cell or "").strip() for cell in values[0]]
        if not any(headers):
            headers = [f"column_{index + 1}" for index in range(len(values[0]))]
        for row_index, row_values in enumerate(values[1:], start=2):
            raw = {headers[index] if headers[index] else f"column_{index + 1}": row_values[index] for index in range(len(headers))}
            if not any(_clean(value) for value in raw.values()):
                continue
            rows.append(
                {
                    "row_number": row_index,
                    "sheet_name": sheet.title,
                    "source_values": _sanitize(raw),
                }
            )
    if not rows:
        warnings.append("No non-empty worksheet rows were found in the Excel file.")
    return rows, warnings, "openpyxl"


def _parse_pdf_bytes(data: bytes) -> Tuple[List[Dict[str, Any]], List[str], str]:
    warnings: List[str] = []
    rows: List[Dict[str, Any]] = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables() or []
            if tables:
                for table_index, table in enumerate(tables, start=1):
                    if not table:
                        continue
                    header = [str(cell or "").strip() for cell in (table[0] or [])]
                    if not any(header):
                        header = [f"column_{i + 1}" for i in range(len(table[0] or []))]
                    for row_number, values in enumerate(table[1:], start=2):
                        raw = {
                            header[i] if i < len(header) and header[i] else f"column_{i + 1}": values[i]
                            for i in range(len(values or []))
                        }
                        if not any(_clean(value) for value in raw.values()):
                            continue
                        rows.append(
                            {
                                "row_number": row_number,
                                "page_number": page_number,
                                "table_index": table_index,
                                "source_values": _sanitize(raw),
                            }
                        )
            else:
                text = page.extract_text() or ""
                line_rows = []
                for line_number, line in enumerate(text.splitlines(), start=1):
                    parsed = _pdf_line_to_row(line)
                    if not parsed:
                        continue
                    line_rows.append(
                        {
                            "row_number": line_number,
                            "page_number": page_number,
                            "source_values": _sanitize(parsed),
                        }
                    )
                if line_rows:
                    warnings.append(f"Page {page_number} required line-based PDF extraction and still needs human review.")
                    rows.extend(line_rows)
    if not rows:
        warnings.append("No reviewable table or line rows were extracted from the PDF.")
    return rows, warnings, "pdfplumber"


def _parse_import_file(filename: str, data: bytes) -> Tuple[List[Dict[str, Any]], List[str], str, str]:
    extension = _extract_extension(filename)
    file_kind = IMPORT_FILE_EXTENSIONS.get(extension)
    if not file_kind:
        raise ValueError("unsupported_budget_import_file")
    if file_kind == "csv":
        rows, warnings, parser = _parse_csv_bytes(data)
    elif file_kind == "excel":
        rows, warnings, parser = _parse_excel_bytes(data)
    else:
        rows, warnings, parser = _parse_pdf_bytes(data)
    return rows, warnings, parser, file_kind


def _tokenize(text: str) -> List[str]:
    return [token for token in re.split(r"[^a-z0-9]+", (text or "").lower()) if token]


def _suggest_existing_pay_item(normalized: Dict[str, Any], pay_items: List[Dict[str, Any]]) -> Dict[str, Any]:
    customer_number = _clean(normalized.get("customer_pay_item_number"))
    if customer_number:
        for pay_item in pay_items:
            if _clean(pay_item.get("customer_pay_item_number")).lower() == customer_number.lower():
                return {
                    "pay_item_id": pay_item.get("pay_item_id") or "",
                    "customer_pay_item_number": pay_item.get("customer_pay_item_number") or customer_number,
                    "confidence": "high",
                    "reason": "Exact customer pay-item number match found in project authority.",
                }
    desc_tokens = set(_tokenize(normalized.get("description") or ""))
    best: Optional[Dict[str, Any]] = None
    best_score = 0
    for pay_item in pay_items:
        item_tokens = set(_tokenize(f"{pay_item.get('customer_pay_item_number')} {pay_item.get('description') or ''}"))
        score = len(desc_tokens & item_tokens)
        if score > best_score:
            best = pay_item
            best_score = score
    if not best:
        return {
            "pay_item_id": "",
            "customer_pay_item_number": customer_number,
            "confidence": "review_required",
            "reason": "No governed project pay-item match was strong enough to reuse automatically.",
        }
    return {
        "pay_item_id": best.get("pay_item_id") or "",
        "customer_pay_item_number": best.get("customer_pay_item_number") or customer_number,
        "confidence": "medium" if best_score >= 2 else "review_required",
        "reason": "Description overlap suggested a reusable project pay item, but PM review remains required.",
    }


def _build_budget_row_suggestion(normalized: Dict[str, Any], pay_items: List[Dict[str, Any]], work_types: List[Dict[str, Any]]) -> Dict[str, Any]:
    pay_item = _suggest_existing_pay_item(normalized, pay_items)
    work_type = _suggest_work_type(
        {
            "customer_pay_item_number": normalized.get("customer_pay_item_number"),
            "description": normalized.get("description"),
        },
        work_types,
    )
    reasons = [pay_item.get("reason") or "", "Matched work-type keywords: " + ", ".join(work_type.get("matched_terms") or []) if work_type.get("matched_terms") else "No strong work-type keyword overlap was found."]
    warnings: List[str] = []
    if not _clean(normalized.get("customer_pay_item_number")):
        warnings.append("Customer pay-item number could not be resolved from source data.")
    if not _clean(normalized.get("description")):
        warnings.append("Description could not be resolved from source data.")
    if not work_type.get("primary_work_type_id"):
        warnings.append("Enterprise work type still needs PM review.")
    confidence_levels = [pay_item.get("confidence") or "review_required", work_type.get("confidence") or "review_required"]
    confidence = "high" if all(level == "high" for level in confidence_levels) else "medium" if "medium" in confidence_levels and "review_required" not in confidence_levels else "review_required"
    return {
        "customer_pay_item_id": pay_item.get("pay_item_id") or "",
        "customer_pay_item_number": pay_item.get("customer_pay_item_number") or _clean(normalized.get("customer_pay_item_number")),
        "enterprise_work_type_id": work_type.get("primary_work_type_id") or "",
        "project_cost_code": _clean(normalized.get("project_cost_code")),
        "phase_id": _clean(normalized.get("phase_id")),
        "work_package_id": _clean(normalized.get("work_package_id")),
        "schedule_activity_id": _clean(normalized.get("schedule_activity_id")),
        "schedule_activity_name": _clean(normalized.get("schedule_activity_name")),
        "confidence": confidence,
        "matched_terms": work_type.get("matched_terms") or [],
        "reasons": [reason for reason in reasons if reason],
        "warnings": warnings,
    }


def _selected_payload_for_row(row: Dict[str, Any], payload: Dict[str, Any]) -> Dict[str, Any]:
    normalized = row.get("normalized") or {}
    suggestion = row.get("suggestion") or {}
    selected = deepcopy(row.get("selected") or {})
    selected.update(
        {
            "customer_pay_item_id": _clean(payload.get("customer_pay_item_id") or selected.get("customer_pay_item_id") or suggestion.get("customer_pay_item_id")),
            "customer_pay_item_number": _clean(payload.get("customer_pay_item_number") or selected.get("customer_pay_item_number") or suggestion.get("customer_pay_item_number") or normalized.get("customer_pay_item_number")),
            "description": _clean(payload.get("description") or selected.get("description") or normalized.get("description")),
            "quantity": _safe_float(payload.get("quantity") if payload.get("quantity") is not None else selected.get("quantity", normalized.get("quantity"))),
            "unit": _clean(payload.get("unit") or selected.get("unit") or normalized.get("unit")),
            "unit_price": _safe_float(payload.get("unit_price") if payload.get("unit_price") is not None else selected.get("unit_price", normalized.get("unit_price"))),
            "budget_amount": _safe_float(payload.get("budget_amount") if payload.get("budget_amount") is not None else selected.get("budget_amount", normalized.get("budget_amount"))),
            "enterprise_work_type_id": _clean(payload.get("enterprise_work_type_id") or selected.get("enterprise_work_type_id") or suggestion.get("enterprise_work_type_id")),
            "project_cost_code": _clean(payload.get("project_cost_code") or selected.get("project_cost_code") or suggestion.get("project_cost_code") or normalized.get("project_cost_code")),
            "phase_id": _clean(payload.get("phase_id") or selected.get("phase_id") or suggestion.get("phase_id") or normalized.get("phase_id")),
            "work_package_id": _clean(payload.get("work_package_id") or selected.get("work_package_id") or suggestion.get("work_package_id") or normalized.get("work_package_id")),
            "schedule_activity_id": _clean(payload.get("schedule_activity_id") or selected.get("schedule_activity_id") or suggestion.get("schedule_activity_id") or normalized.get("schedule_activity_id")),
            "schedule_activity_name": _clean(payload.get("schedule_activity_name") or selected.get("schedule_activity_name") or suggestion.get("schedule_activity_name") or normalized.get("schedule_activity_name")),
            "line_kind": _line_kind_from_value(payload.get("line_kind") or selected.get("line_kind") or normalized.get("line_kind") or "direct_cost"),
            "review_note": _clean(payload.get("review_note") or selected.get("review_note")),
        }
    )
    if not selected["budget_amount"] and selected["quantity"] and selected["unit_price"]:
        selected["budget_amount"] = round(selected["quantity"] * selected["unit_price"], 4)
    return selected


def _review_status_from_counts(total: int, approved: int, rejected: int, needs_review: int, activated: int) -> str:
    if activated == total and total > 0:
        return "activated"
    if total > 0 and approved + rejected == total and needs_review == 0:
        return "approved_ready"
    if approved > 0 or rejected > 0:
        return "partially_reviewed"
    return "review_required"


async def _ensure_indexes(db) -> None:
    await db[COLL_BUDGET_VERSIONS].create_index([("project_number", 1), ("version_id", 1)], unique=True)
    await db[COLL_BUDGET_VERSIONS].create_index([("project_number", 1), ("status", 1), ("activated_at", -1)])
    await db[COLL_BUDGET_LINES].create_index([("project_number", 1), ("version_id", 1), ("budget_line_id", 1)], unique=True)
    await db[COLL_BUDGET_LINES].create_index([("project_number", 1), ("version_id", 1), ("customer_pay_item_number", 1)])
    await db[COLL_BUDGET_IMPORTS].create_index([("project_number", 1), ("import_id", 1)], unique=True)
    await db[COLL_BUDGET_IMPORTS].create_index([("project_number", 1), ("status", 1), ("imported_at", -1)])
    await db[COLL_BUDGET_IMPORT_ROWS].create_index([("import_id", 1), ("row_id", 1)], unique=True)
    await db[COLL_BUDGET_IMPORT_ROWS].create_index([("project_number", 1), ("review_status", 1), ("row_number", 1)])
    await db[COLL_BUDGET_REVIEW].create_index("review_id", unique=True)
    await db[COLL_BUDGET_REVIEW].create_index([("project_number", 1), ("status", 1), ("priority", -1), ("updated_at", -1)])
    await db[COLL_BUDGET_COMMITMENTS].create_index([("project_number", 1), ("source_po_id", 1)], unique=True)
    await db[COLL_BUDGET_ACTUALS].create_index([("project_number", 1), ("source_kind", 1), ("source_record_id", 1)], unique=True)
    await db[COLL_BUDGET_DISTRIBUTION].create_index([("project_number", 1), ("created_at", -1)])
    await db[COLL_BUDGET_RUNS].create_index([("run_type", 1)], unique=True)


async def _upsert_review_item(db, review: Dict[str, Any]) -> Dict[str, Any]:
    existing = await db[COLL_BUDGET_REVIEW].find_one({"review_id": review["review_id"]}, {"_id": 0})
    now = _utcnow()
    doc = {
        **(existing or {}),
        **review,
        "created_at": (existing or {}).get("created_at") or now,
        "updated_at": now,
    }
    await db[COLL_BUDGET_REVIEW].replace_one({"review_id": doc["review_id"]}, doc, upsert=True)
    return _sanitize(doc)


async def _mark_review_resolved(db, review_id: str, *, actor: Optional[Dict[str, Any]] = None, resolution_note: str = "") -> None:
    row = await db[COLL_BUDGET_REVIEW].find_one({"review_id": review_id}, {"_id": 0})
    if not row:
        return
    row["status"] = "resolved"
    row["resolution_note"] = _clean(resolution_note) or "Resolved by subsequent governed budget action."
    row["resolved_at"] = _utcnow()
    row["resolved_by"] = _actor_label(actor)
    row["updated_at"] = _utcnow()
    await db[COLL_BUDGET_REVIEW].replace_one({"review_id": review_id}, row, upsert=True)


async def _refresh_import_session_counts(db, import_id: str) -> Dict[str, Any]:
    total = await db[COLL_BUDGET_IMPORT_ROWS].count_documents({"import_id": import_id})
    approved = await db[COLL_BUDGET_IMPORT_ROWS].count_documents({"import_id": import_id, "review_status": "approved"})
    rejected = await db[COLL_BUDGET_IMPORT_ROWS].count_documents({"import_id": import_id, "review_status": "rejected"})
    needs_review = await db[COLL_BUDGET_IMPORT_ROWS].count_documents({"import_id": import_id, "review_status": {"$in": ["pending_review", "review_required"]}})
    activated = await db[COLL_BUDGET_IMPORT_ROWS].count_documents({"import_id": import_id, "review_status": "activated"})
    status = _review_status_from_counts(total, approved, rejected, needs_review, activated)
    counts = {
        "total_rows": total,
        "approved_rows": approved,
        "rejected_rows": rejected,
        "needs_review_rows": needs_review,
        "activated_rows": activated,
    }
    await db[COLL_BUDGET_IMPORTS].update_one({"import_id": import_id}, {"$set": {**counts, "status": status, "updated_at": _utcnow()}})
    session = await db[COLL_BUDGET_IMPORTS].find_one({"import_id": import_id}, {"_id": 0})
    return _sanitize(session or {**counts, "status": status})


async def _get_latest_active_version(db, project_number: str) -> Optional[Dict[str, Any]]:
    return await db[COLL_BUDGET_VERSIONS].find_one(
        {"project_number": project_number, "status": "active"},
        {"_id": 0},
        sort=[("activated_at", -1), ("created_at", -1)],
    )


def _allocation_amount(value: Any) -> float:
    return round(max(_safe_float(value), 0.0), 2)


def _normalize_link_allocations(raw_allocations: Any) -> List[Dict[str, Any]]:
    allocations: List[Dict[str, Any]] = []
    for row in raw_allocations or []:
        if not isinstance(row, dict):
            continue
        budget_line_id = _clean(row.get("budget_line_id"))
        amount = _allocation_amount(row.get("amount"))
        if not budget_line_id or amount <= 0:
            continue
        allocations.append({"budget_line_id": budget_line_id, "amount": amount})
    return allocations


def _budget_line_operator_label(line: Dict[str, Any]) -> str:
    parts = [
        _clean(line.get("customer_pay_item_number")),
        _clean(line.get("project_cost_code")),
        _clean(line.get("description")),
    ]
    label = " · ".join(part for part in parts if part)
    return label or _clean(line.get("budget_line_id")) or "Budget line"


async def _active_budget_line_index(db, project_number: str) -> Tuple[Optional[Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    active_version = await _get_latest_active_version(db, project_number)
    if not active_version:
        return None, {}
    lines = await list_project_budget_lines(db, project_number, version_id=active_version["version_id"])
    return active_version, {row["budget_line_id"]: row for row in lines if _clean(row.get("budget_line_id"))}


async def _refresh_project_review_lane(db, project_number: str, *, actor: Optional[Dict[str, Any]] = None) -> None:
    open_commitments = await db[COLL_BUDGET_COMMITMENTS].count_documents({"project_number": project_number, "review_status": {"$in": ["pending_review", "review_required"]}})
    open_actuals = await db[COLL_BUDGET_ACTUALS].count_documents({"project_number": project_number, "review_status": {"$in": ["pending_review", "review_required"]}})
    total_open = int(open_commitments or 0) + int(open_actuals or 0)
    review_id = f"budget-review:project:{project_number}"
    if total_open <= 0:
        await _mark_review_resolved(db, review_id, actor=actor, resolution_note="All commitment and actual-cost candidates are now linked or dispositioned.")
        return
    await _upsert_review_item(
        db,
        {
            "review_id": review_id,
            "project_number": project_number,
            "status": "open",
            "priority": 75,
            "title": "Budget trust-line linkage needs review",
            "description": "Approved commitments and receipt-based actual-cost candidates must be linked to governed budget lines before downstream EV metrics can publish at full confidence.",
            "owner_role": "pm",
            "category": "budget_linkage",
            "counts": {
                "open_commitment_candidates": int(open_commitments or 0),
                "open_actual_cost_candidates": int(open_actuals or 0),
            },
            "evidence": [
                COLL_BUDGET_COMMITMENTS,
                COLL_BUDGET_ACTUALS,
                COLL_BUDGET_LINES,
            ],
        },
    )


async def _recalculate_budget_financial_rollups(db, project_number: str, *, actor: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    active_version, line_index = await _active_budget_line_index(db, project_number)
    if not active_version or not line_index:
        await _refresh_project_review_lane(db, project_number, actor=actor)
        return {
            "project_number": project_number,
            "version_id": (active_version or {}).get("version_id") or "",
            "line_count": 0,
            "approved_commitment_candidates": 0,
            "approved_actual_cost_candidates": 0,
        }

    commitment_groups: Dict[str, List[Dict[str, Any]]] = {line_id: [] for line_id in line_index}
    actual_groups: Dict[str, List[Dict[str, Any]]] = {line_id: [] for line_id in line_index}

    approved_commitments = [
        _sanitize(row)
        async for row in db[COLL_BUDGET_COMMITMENTS].find(
            {"project_number": project_number, "review_status": "approved"},
            {"_id": 0},
        )
    ]
    approved_actuals = [
        _sanitize(row)
        async for row in db[COLL_BUDGET_ACTUALS].find(
            {"project_number": project_number, "review_status": "approved"},
            {"_id": 0},
        )
    ]

    for candidate in approved_commitments:
        allocations = _normalize_link_allocations(candidate.get("allocations"))
        for allocation in allocations:
            line = line_index.get(allocation["budget_line_id"])
            if not line:
                continue
            commitment_groups.setdefault(line["budget_line_id"], []).append(
                {
                    "candidate_id": candidate.get("candidate_id"),
                    "source_po_id": candidate.get("source_po_id"),
                    "po_number": candidate.get("po_number"),
                    "vendor": candidate.get("vendor"),
                    "amount": allocation["amount"],
                    "reviewed_at": candidate.get("reviewed_at"),
                    "reviewed_by": candidate.get("reviewed_by"),
                }
            )

    for candidate in approved_actuals:
        allocations = _normalize_link_allocations(candidate.get("allocations"))
        for allocation in allocations:
            line = line_index.get(allocation["budget_line_id"])
            if not line:
                continue
            actual_groups.setdefault(line["budget_line_id"], []).append(
                {
                    "candidate_id": candidate.get("candidate_id"),
                    "source_record_id": candidate.get("source_record_id"),
                    "source_kind": candidate.get("source_kind"),
                    "source_label": candidate.get("source_label"),
                    "amount": allocation["amount"],
                    "reviewed_at": candidate.get("reviewed_at"),
                    "reviewed_by": candidate.get("reviewed_by"),
                }
            )

    totals = {
        "budget_amount": 0.0,
        "commitment_amount": 0.0,
        "actual_cost_amount": 0.0,
        "forecast_amount": 0.0,
        "remaining_amount": 0.0,
        "revenue_amount": 0.0,
        "billing_amount": 0.0,
        "collections_amount": 0.0,
        "labor_cost_budget_amount": 0.0,
        "equipment_cost_budget_amount": 0.0,
        "material_cost_budget_amount": 0.0,
        "subcontract_cost_budget_amount": 0.0,
        "other_cost_budget_amount": 0.0,
    }

    now = _utcnow()
    for line_id, original in line_index.items():
        line = deepcopy(original)
        commitment_refs = commitment_groups.get(line_id, [])
        actual_refs = actual_groups.get(line_id, [])
        line["commitment_refs"] = _sanitize(commitment_refs)
        line["actual_cost_refs"] = _sanitize(actual_refs)
        line["commitment_amount"] = round(sum(_allocation_amount(row.get("amount")) for row in commitment_refs), 2)
        line["actual_cost_amount"] = round(sum(_allocation_amount(row.get("amount")) for row in actual_refs), 2)
        budget_amount = round(_safe_float(line.get("budget_amount")), 2)
        line["remaining_amount"] = round(max(budget_amount - line["actual_cost_amount"], 0.0), 2)
        line["forecast_amount"] = round(max(_safe_float(line.get("forecast_amount") or budget_amount), budget_amount), 2)
        line["updated_at"] = now
        if actor:
            line["updated_by"] = _actor_label(actor)
        await db[COLL_BUDGET_LINES].replace_one(
            {"project_number": project_number, "version_id": line.get("version_id"), "budget_line_id": line_id},
            line,
            upsert=True,
        )
        for key in totals:
            totals[key] = round(totals[key] + _safe_float(line.get(key)), 2)

    await db[COLL_BUDGET_VERSIONS].update_one(
        {"project_number": project_number, "version_id": active_version["version_id"]},
        {
            "$set": {
                "totals": totals,
                "updated_at": now,
                "updated_by": _actor_label(actor),
            }
        },
    )
    await _refresh_project_review_lane(db, project_number, actor=actor)
    return {
        "project_number": project_number,
        "version_id": active_version["version_id"],
        "line_count": len(line_index),
        "approved_commitment_candidates": len(approved_commitments),
        "approved_actual_cost_candidates": len(approved_actuals),
        "totals": totals,
    }


async def _review_budget_link_candidate(
    db,
    *,
    project_number: str,
    candidate_id: str,
    collection_name: str,
    actor: Dict[str, Any],
    action: str,
    allocations: Optional[List[Dict[str, Any]]] = None,
    review_note: str = "",
) -> Dict[str, Any]:
    await ensure_project_budget_foundation(db)
    candidate = await db[collection_name].find_one({"project_number": project_number, "candidate_id": candidate_id}, {"_id": 0})
    if not candidate:
        raise LookupError("budget_candidate_not_found")

    normalized_action = _clean(action).lower() or "review_required"
    if normalized_action not in {"approve", "reject", "review_required"}:
        raise ValueError("invalid_budget_link_action")

    active_version, line_index = await _active_budget_line_index(db, project_number)
    if normalized_action == "approve" and (not active_version or not line_index):
        raise ValueError("budget_link_requires_active_budget_version")

    normalized_allocations = _normalize_link_allocations(allocations)
    source_amount = _allocation_amount(
        candidate.get("commitment_amount")
        if collection_name == COLL_BUDGET_COMMITMENTS
        else candidate.get("candidate_amount")
    )
    if normalized_action == "approve":
        if not normalized_allocations:
            raise ValueError("budget_link_allocations_required")
        for allocation in normalized_allocations:
            if allocation["budget_line_id"] not in line_index:
                raise ValueError("budget_link_line_not_found")
        allocated_total = round(sum(_allocation_amount(row.get("amount")) for row in normalized_allocations), 2)
        if abs(allocated_total - source_amount) > 0.01:
            raise ValueError("budget_link_allocation_total_mismatch")
    else:
        normalized_allocations = []

    updated = deepcopy(candidate)
    updated["review_status"] = "approved" if normalized_action == "approve" else "rejected" if normalized_action == "reject" else "review_required"
    updated["reviewed_at"] = _utcnow()
    updated["reviewed_by"] = _actor_label(actor)
    updated["review_note"] = _clean(review_note)
    updated["allocations"] = [
        {
            "budget_line_id": row["budget_line_id"],
            "amount": row["amount"],
            "label": _budget_line_operator_label(line_index.get(row["budget_line_id"], {})),
        }
        for row in normalized_allocations
    ]
    updated["linked_amount"] = round(sum(_allocation_amount(row.get("amount")) for row in normalized_allocations), 2)
    updated["updated_at"] = _utcnow()
    updated["budget_line_id"] = normalized_allocations[0]["budget_line_id"] if len(normalized_allocations) == 1 else ""
    if normalized_action != "approve":
        updated["linked_amount"] = 0.0
        updated["budget_line_id"] = ""

    await db[collection_name].replace_one(
        {"project_number": project_number, "candidate_id": candidate_id},
        updated,
        upsert=True,
    )
    recalculated = await _recalculate_budget_financial_rollups(db, project_number, actor=actor)
    await _write_audit(
        db,
        "budget_link_candidate_reviewed",
        actor,
        "budget_link_candidate",
        candidate_id,
        updated,
        before=candidate,
        metadata={
            "collection_name": collection_name,
            "project_number": project_number,
            "action": normalized_action,
            "allocation_count": len(normalized_allocations),
            "recalculated_totals": recalculated.get("totals") or {},
        },
    )
    return _sanitize(updated)


async def review_budget_commitment_candidate(
    db,
    project_number: str,
    candidate_id: str,
    *,
    actor: Dict[str, Any],
    action: str,
    allocations: Optional[List[Dict[str, Any]]] = None,
    review_note: str = "",
) -> Dict[str, Any]:
    return await _review_budget_link_candidate(
        db,
        project_number=project_number,
        candidate_id=candidate_id,
        collection_name=COLL_BUDGET_COMMITMENTS,
        actor=actor,
        action=action,
        allocations=allocations,
        review_note=review_note,
    )


async def review_budget_actual_cost_candidate(
    db,
    project_number: str,
    candidate_id: str,
    *,
    actor: Dict[str, Any],
    action: str,
    allocations: Optional[List[Dict[str, Any]]] = None,
    review_note: str = "",
) -> Dict[str, Any]:
    return await _review_budget_link_candidate(
        db,
        project_number=project_number,
        candidate_id=candidate_id,
        collection_name=COLL_BUDGET_ACTUALS,
        actor=actor,
        action=action,
        allocations=allocations,
        review_note=review_note,
    )


def _line_financial_rollups(selected: Dict[str, Any]) -> Dict[str, Any]:
    budget_amount = _safe_float(selected.get("budget_amount"))
    return {
        "budget_amount": budget_amount,
        "commitment_amount": 0.0,
        "actual_cost_amount": 0.0,
        "forecast_amount": budget_amount,
        "remaining_amount": budget_amount,
        "revenue_amount": 0.0,
        "billing_amount": 0.0,
        "collections_amount": 0.0,
        "labor_cost_budget_amount": 0.0,
        "equipment_cost_budget_amount": 0.0,
        "material_cost_budget_amount": 0.0,
        "subcontract_cost_budget_amount": 0.0,
        "vendor_cost_budget_amount": 0.0,
        "production_quantity_rollup": _safe_float(selected.get("quantity")),
    }


def _budget_version_doc(project_number: str, project_name: str, payload: Dict[str, Any], *, parent_version_id: str = "", actor: Dict[str, Any], source_import_id: str = "") -> Dict[str, Any]:
    now = _utcnow()
    stage = _status(payload.get("stage") or "original_approved_budget", allowed=VERSION_STAGES, default="original_approved_budget")
    version_name = _clean(payload.get("version_name") or f"{project_name or project_number} · {stage.replace('_', ' ').title()}")
    version_id = payload.get("version_id") or f"budget-version:{project_number}:{_norm(stage)}:{uuid4().hex[:8]}"
    return {
        "version_id": version_id,
        "project_number": project_number,
        "project_name": project_name,
        "stage": stage,
        "version_name": version_name,
        "status": "active",
        "source_import_id": source_import_id,
        "parent_version_id": _clean(parent_version_id),
        "comparison_baseline_version_id": _clean(parent_version_id),
        "immutable_after_activation": stage == "original_approved_budget",
        "activation_guardrails": {
            "auto_activation": False,
            "pm_review_required": True,
            "pm_approval_required": True,
        },
        "trust_lines": {
            "customer_pay_items": "project_pay_item_registry",
            "enterprise_work_types": "enterprise_work_type_registry",
            "commitments": "po_requests",
            "actual_costs": "external_accounting_or_governed_receipt_review",
            "operational_work": "project_controls_work_ledger",
        },
        "totals": {
            "budget_amount": 0.0,
            "commitment_amount": 0.0,
            "actual_cost_amount": 0.0,
            "forecast_amount": 0.0,
            "remaining_amount": 0.0,
            "revenue_amount": 0.0,
            "billing_amount": 0.0,
            "collections_amount": 0.0,
            "allowance_amount": 0.0,
            "contingency_amount": 0.0,
            "management_reserve_amount": 0.0,
        },
        "created_at": now,
        "created_by": _actor_label(actor),
        "approved_at": now,
        "approved_by": _actor_label(actor),
        "activated_at": now,
        "activated_by": _actor_label(actor),
        "updated_at": now,
        "updated_by": _actor_label(actor),
        "notes": _clean(payload.get("notes")),
    }


def _budget_line_doc(project_number: str, version_id: str, selected: Dict[str, Any], *, pay_item_id: str, actor: Dict[str, Any], source_import_id: str, source_row_id: str, source_hash: str) -> Dict[str, Any]:
    now = _utcnow()
    budget_line_id = f"budget-line:{project_number}:{version_id}:{source_row_id}"
    rollups = _line_financial_rollups(selected)
    return {
        "budget_line_id": budget_line_id,
        "project_number": project_number,
        "version_id": version_id,
        "status": "active",
        "line_kind": _line_kind_from_value(selected.get("line_kind") or "direct_cost"),
        "customer_pay_item_id": _clean(pay_item_id),
        "customer_pay_item_number": _clean(selected.get("customer_pay_item_number")),
        "description": _clean(selected.get("description")),
        "enterprise_work_type_id": _clean(selected.get("enterprise_work_type_id")),
        "project_cost_code": _clean(selected.get("project_cost_code")),
        "phase_id": _clean(selected.get("phase_id")),
        "work_package_id": _clean(selected.get("work_package_id")),
        "schedule_activity_id": _clean(selected.get("schedule_activity_id")),
        "schedule_activity_name": _clean(selected.get("schedule_activity_name")),
        "daily_report_work_block_ids": [],
        "crew_ids": [],
        "employee_ids": [],
        "equipment_ids": [],
        "material_refs": [],
        "vendor_refs": [],
        "subcontractor_refs": [],
        "commitment_refs": [],
        "actual_cost_refs": [],
        "forecast_refs": [],
        "quantity": _safe_float(selected.get("quantity")),
        "unit": _clean(selected.get("unit")),
        "unit_budget_amount": _safe_float(selected.get("unit_price")),
        **rollups,
        "trust_lines": {
            "budget": COLL_BUDGET_LINES,
            "commitment": "po_requests",
            "actual_cost": "external_accounting_or_governed_receipt_review",
            "revenue": "customer_contract_or_future_billing_module",
            "billing": "future_billing_module",
            "collections": "future_collections_module",
        },
        "source_import_id": source_import_id,
        "source_row_id": source_row_id,
        "source_document_hash": source_hash,
        "source_lineage": {
            "preserved_import_review_required": True,
            "pm_approved_at": now,
            "pm_approved_by": _actor_label(actor),
        },
        "created_at": now,
        "created_by": _actor_label(actor),
        "updated_at": now,
        "updated_by": _actor_label(actor),
    }


async def ensure_project_budget_foundation(db) -> Dict[str, Any]:
    await _ensure_indexes(db)
    await ensure_project_controls_foundation(db)
    last_run = await db[COLL_BUDGET_RUNS].find_one({"run_type": "wp18c3_backfill"}, {"_id": 0})
    return {
        "ok": True,
        "backfill": _sanitize(last_run or {"run_type": "wp18c3_backfill", "status": "pending_manual_run"}),
        "event_contracts": BUDGET_EVENT_CONTRACTS,
    }


async def list_project_budget_versions(db, project_number: str) -> List[Dict[str, Any]]:
    await ensure_project_budget_foundation(db)
    rows = [
        _sanitize(row)
        async for row in db[COLL_BUDGET_VERSIONS].find({"project_number": project_number}, {"_id": 0}).sort([("activated_at", -1), ("created_at", -1)])
    ]
    return rows


async def list_project_budget_lines(db, project_number: str, *, version_id: str) -> List[Dict[str, Any]]:
    await ensure_project_budget_foundation(db)
    rows = [
        _sanitize(row)
        async for row in db[COLL_BUDGET_LINES].find({"project_number": project_number, "version_id": version_id}, {"_id": 0}).sort([("line_kind", 1), ("customer_pay_item_number", 1), ("description", 1)])
    ]
    return rows


async def list_budget_import_sessions(db, project_number: str) -> List[Dict[str, Any]]:
    await ensure_project_budget_foundation(db)
    return [
        _sanitize(row)
        async for row in db[COLL_BUDGET_IMPORTS].find({"project_number": project_number}, {"_id": 0}).sort([("imported_at", -1)])
    ]


async def get_budget_import_session_detail(db, project_number: str, import_id: str) -> Dict[str, Any]:
    await ensure_project_budget_foundation(db)
    session = await db[COLL_BUDGET_IMPORTS].find_one({"project_number": project_number, "import_id": import_id}, {"_id": 0})
    if not session:
        raise LookupError("budget_import_not_found")
    rows = [
        _sanitize(row)
        async for row in db[COLL_BUDGET_IMPORT_ROWS].find({"project_number": project_number, "import_id": import_id}, {"_id": 0}).sort([("row_number", 1), ("row_id", 1)])
    ]
    return {"session": _sanitize(session), "rows": rows, "count": len(rows)}


async def list_budget_review_queue(db, *, project_number: str = "") -> List[Dict[str, Any]]:
    await ensure_project_budget_foundation(db)
    query: Dict[str, Any] = {}
    if project_number:
        query["project_number"] = project_number
    rows = [
        _sanitize(row)
        async for row in db[COLL_BUDGET_REVIEW].find(query, {"_id": 0}).sort([("priority", -1), ("updated_at", -1)]).limit(500)
    ]
    return rows


async def create_budget_import_session(
    db,
    project_number: str,
    *,
    filename: str,
    content_type: str,
    data: bytes,
    source_kind: str,
    target_version_stage: str,
    version_name: str,
    actor: Dict[str, Any],
) -> Dict[str, Any]:
    await ensure_project_budget_foundation(db)
    job = await _load_job(db, project_number)
    source_kind = _status(source_kind or "csv", allowed=IMPORT_SOURCE_KINDS, default="csv")
    target_version_stage = _status(target_version_stage or "original_approved_budget", allowed=VERSION_STAGES, default="original_approved_budget")
    if not data:
        raise ValueError("budget_import_file_required")
    if len(data) > 15 * 1024 * 1024:
        raise ValueError("budget_import_file_too_large")
    rows, parser_warnings, parser_name, file_kind = _parse_import_file(filename, data)
    file_hash = _hash_bytes(data)
    duplicate = await db[COLL_BUDGET_IMPORTS].find_one({"project_number": project_number, "file_hash": file_hash}, {"_id": 0})
    if duplicate:
        detail = await get_budget_import_session_detail(db, project_number, duplicate["import_id"])
        detail["duplicate_of"] = duplicate["import_id"]
        return detail
    work_types = await list_enterprise_work_types(db, include_archived=False)
    pay_items = await list_project_pay_items(db, project_number)
    import_id = f"budget-import:{project_number}:{uuid4().hex[:12]}"
    imported_at = _utcnow()
    row_docs = []
    for source_row in rows:
        normalized = _normalize_source_row(source_row.get("source_values") or {})
        suggestion = _build_budget_row_suggestion(normalized, pay_items, work_types)
        row_id = f"budget-import-row:{import_id}:{source_row.get('row_number') or len(row_docs) + 1}"
        review_status = "review_required" if suggestion.get("warnings") else "pending_review"
        row_docs.append(
            {
                "row_id": row_id,
                "import_id": import_id,
                "project_number": project_number,
                "row_number": int(source_row.get("row_number") or (len(row_docs) + 1)),
                "sheet_name": _clean(source_row.get("sheet_name")),
                "page_number": source_row.get("page_number") or None,
                "table_index": source_row.get("table_index") or None,
                "source_values": _sanitize(source_row.get("source_values") or {}),
                "normalized": _sanitize(normalized),
                "suggestion": _sanitize(suggestion),
                "selected": _sanitize(
                    {
                        "customer_pay_item_id": suggestion.get("customer_pay_item_id") or "",
                        "customer_pay_item_number": suggestion.get("customer_pay_item_number") or normalized.get("customer_pay_item_number"),
                        "description": normalized.get("description"),
                        "quantity": normalized.get("quantity"),
                        "unit": normalized.get("unit"),
                        "unit_price": normalized.get("unit_price"),
                        "budget_amount": normalized.get("budget_amount"),
                        "enterprise_work_type_id": suggestion.get("enterprise_work_type_id") or "",
                        "project_cost_code": suggestion.get("project_cost_code") or normalized.get("project_cost_code"),
                        "phase_id": suggestion.get("phase_id") or normalized.get("phase_id"),
                        "work_package_id": suggestion.get("work_package_id") or normalized.get("work_package_id"),
                        "schedule_activity_id": suggestion.get("schedule_activity_id") or normalized.get("schedule_activity_id"),
                        "schedule_activity_name": suggestion.get("schedule_activity_name") or normalized.get("schedule_activity_name"),
                        "line_kind": normalized.get("line_kind") or "direct_cost",
                    }
                ),
                "review_status": review_status,
                "created_at": imported_at,
                "created_by": _actor_label(actor),
                "updated_at": imported_at,
                "updated_by": _actor_label(actor),
            }
        )
    session = {
        "import_id": import_id,
        "project_number": project_number,
        "project_name": job.get("project_name") or job.get("name") or project_number,
        "status": "review_required",
        "source_kind": source_kind,
        "target_version_stage": target_version_stage,
        "version_name": _clean(version_name) or f"{job.get('project_name') or project_number} · {target_version_stage.replace('_', ' ').title()}",
        "filename": _clean_filename(filename),
        "content_type": _clean(content_type) or "application/octet-stream",
        "file_kind": file_kind,
        "file_hash": file_hash,
        "file_extension": _extract_extension(filename),
        "parser_name": parser_name,
        "parser_warnings": parser_warnings,
        "advisory_engine": {
            "mode": "deterministic_governed_suggestions",
            "activation_policy": "pm_review_required",
            "auto_approval": False,
        },
        "source_preservation": {
            "original_filename": _clean_filename(filename),
            "sha256": file_hash,
            "imported_at": imported_at,
            "imported_by": _actor_label(actor),
            "sample_rows": [_sanitize(row.get("source_values") or {}) for row in row_docs[:3]],
        },
        "total_rows": len(row_docs),
        "approved_rows": 0,
        "rejected_rows": 0,
        "needs_review_rows": len(row_docs),
        "activated_rows": 0,
        "imported_at": imported_at,
        "imported_by": _actor_label(actor),
        "updated_at": imported_at,
        "updated_by": _actor_label(actor),
    }
    await db[COLL_BUDGET_IMPORTS].insert_one(session)
    if row_docs:
        await db[COLL_BUDGET_IMPORT_ROWS].insert_many(row_docs)
    if any(row.get("suggestion", {}).get("warnings") for row in row_docs):
        await _upsert_review_item(
            db,
            {
                "review_id": f"budget-review:import:{import_id}",
                "project_number": project_number,
                "status": "review_required",
                "priority": 95,
                "source_kind": "budget_import_session",
                "source_record_id": import_id,
                "title": f"Budget import review required for {project_number}",
                "reason": "One or more imported budget rows could not be resolved confidently. Preserved for PM review without silent activation.",
                "confidence": "human_required",
                "provenance": {
                    "filename": _clean_filename(filename),
                    "source_kind": source_kind,
                    "target_version_stage": target_version_stage,
                },
            },
        )
    await _write_audit(db, "budget_import_staged", actor, "budget_import", import_id, session)
    return await get_budget_import_session_detail(db, project_number, import_id)


async def review_budget_import_row(db, project_number: str, import_id: str, row_id: str, payload: Dict[str, Any], *, actor: Dict[str, Any]) -> Dict[str, Any]:
    await ensure_project_budget_foundation(db)
    session = await db[COLL_BUDGET_IMPORTS].find_one({"project_number": project_number, "import_id": import_id}, {"_id": 0})
    if not session:
        raise LookupError("budget_import_not_found")
    row = await db[COLL_BUDGET_IMPORT_ROWS].find_one({"project_number": project_number, "import_id": import_id, "row_id": row_id}, {"_id": 0})
    if not row:
        raise LookupError("budget_import_row_not_found")
    action = _clean(payload.get("action") or "approve").lower()
    if action not in {"approve", "reject", "needs_review"}:
        raise ValueError("budget_row_action_invalid")
    selected = _selected_payload_for_row(row, payload)
    if action == "approve":
        if not _clean(selected.get("customer_pay_item_number")):
            raise ValueError("customer_pay_item_number_required")
        if not _clean(selected.get("description")):
            raise ValueError("budget_line_description_required")
        if not _clean(selected.get("enterprise_work_type_id")):
            raise ValueError("enterprise_work_type_id_required")
    updated = deepcopy(row)
    updated["selected"] = _sanitize(selected)
    updated["review_status"] = "approved" if action == "approve" else "rejected" if action == "reject" else "review_required"
    updated["review_note"] = _clean(payload.get("review_note") or "")
    updated["reviewed_at"] = _utcnow()
    updated["reviewed_by"] = _actor_label(actor)
    updated["updated_at"] = updated["reviewed_at"]
    updated["updated_by"] = _actor_label(actor)
    await db[COLL_BUDGET_IMPORT_ROWS].replace_one({"row_id": row_id}, updated, upsert=True)
    review_id = f"budget-review:row:{row_id}"
    if updated["review_status"] == "approved":
        await _mark_review_resolved(db, review_id, actor=actor, resolution_note="PM approved import row for activation.")
    else:
        await _upsert_review_item(
            db,
            {
                "review_id": review_id,
                "project_number": project_number,
                "status": "review_required",
                "priority": 90,
                "source_kind": "budget_import_row",
                "source_record_id": row_id,
                "title": f"Budget import row {row.get('row_number')} needs governed review",
                "reason": _clean(payload.get("review_note")) or "PM marked the row for further review rather than activation.",
                "confidence": "human_required",
                "provenance": {
                    "import_id": import_id,
                    "row_number": row.get("row_number"),
                    "selected": _sanitize(selected),
                },
            },
        )
    session = await _refresh_import_session_counts(db, import_id)
    await _write_audit(db, "budget_import_row_reviewed", actor, "budget_import_row", row_id, updated, before=row, metadata={"import_id": import_id, "session_status": session.get("status")})
    return _sanitize(updated)


async def _sync_commitment_candidates_for_project(db, project_number: str, *, actor: Optional[Dict[str, Any]] = None) -> Dict[str, int]:
    created = 0
    reviewed = 0
    async for po in db.po_requests.find({"project_number": project_number, "status": {"$in": ["Approved", "Pending Receipt", "Overdue Receipt", "Receipt Uploaded", "Closed"]}}, {"_id": 0}):
        source_po_id = _clean(po.get("id"))
        if not source_po_id:
            continue
        amount = _safe_float(po.get("approved_amount") if po.get("approved_amount") is not None else po.get("estimated_amount"))
        existing = await db[COLL_BUDGET_COMMITMENTS].find_one({"project_number": project_number, "source_po_id": source_po_id}, {"_id": 0})
        candidate = {
            "candidate_id": f"budget-commitment:{project_number}:{source_po_id}",
            "project_number": project_number,
            "source_po_id": source_po_id,
            "po_number": _clean(po.get("po_number")),
            "vendor": _clean(po.get("vendor")),
            "description": _clean(po.get("description")),
            "commitment_amount": amount,
            "currency": "USD",
            "budget_line_id": _clean((existing or {}).get("budget_line_id")),
            "review_status": (existing or {}).get("review_status") or "review_required",
            "trust_line": "po_requests",
            "allocations": _sanitize((existing or {}).get("allocations") or []),
            "linked_amount": _safe_float((existing or {}).get("linked_amount")),
            "review_note": _clean((existing or {}).get("review_note")),
            "reviewed_at": (existing or {}).get("reviewed_at"),
            "reviewed_by": (existing or {}).get("reviewed_by"),
            "created_at": (existing or {}).get("created_at") or _utcnow(),
            "updated_at": _utcnow(),
        }
        if existing and candidate["review_status"] == "approved" and abs(_safe_float(candidate.get("linked_amount")) - amount) > 0.01:
            candidate["review_status"] = "review_required"
            candidate["review_note"] = _clean(candidate.get("review_note")) or "Source commitment amount changed after approval and requires a fresh governed allocation review."
        if existing:
            reviewed += 1
        else:
            created += 1
        await db[COLL_BUDGET_COMMITMENTS].replace_one({"project_number": project_number, "source_po_id": source_po_id}, candidate, upsert=True)
    unresolved = await db[COLL_BUDGET_COMMITMENTS].count_documents({"project_number": project_number, "review_status": "review_required"})
    if unresolved:
        await _upsert_review_item(
            db,
            {
                "review_id": f"budget-review:commitments:{project_number}",
                "project_number": project_number,
                "status": "review_required",
                "priority": 70,
                "source_kind": "po_commitment_candidates",
                "source_record_id": project_number,
                "title": f"Commitment linkage review required for {project_number}",
                "reason": "Approved or closed PO requests were preserved as commitments but were not guessed onto budget lines.",
                "confidence": "human_required",
                "provenance": {"review_required_count": unresolved},
            },
        )
    else:
        await _mark_review_resolved(db, f"budget-review:commitments:{project_number}", actor=actor, resolution_note="No unresolved commitment candidates remain.")
    return {"created": created, "review_required": unresolved, "touched": created + reviewed}


async def _sync_actual_cost_candidates_for_project(db, project_number: str, *, actor: Optional[Dict[str, Any]] = None) -> Dict[str, int]:
    created = 0
    async for po in db.po_requests.find({"project_number": project_number, "receipt_amount": {"$ne": None}}, {"_id": 0}):
        source_record_id = _clean(po.get("id"))
        if not source_record_id:
            continue
        existing = await db[COLL_BUDGET_ACTUALS].find_one({"project_number": project_number, "source_kind": "po_receipt", "source_record_id": source_record_id}, {"_id": 0})
        candidate = {
            "candidate_id": f"budget-actual:{project_number}:po-receipt:{source_record_id}",
            "project_number": project_number,
            "source_kind": "po_receipt",
            "source_record_id": source_record_id,
            "vendor": _clean(po.get("vendor")),
            "description": _clean(po.get("description")),
            "candidate_amount": _safe_float(po.get("receipt_amount")),
            "budget_line_id": _clean((existing or {}).get("budget_line_id")),
            "review_status": (existing or {}).get("review_status") or "review_required",
            "trust_line": "vendor_receipt_review_only_not_accounting_truth",
            "allocations": _sanitize((existing or {}).get("allocations") or []),
            "linked_amount": _safe_float((existing or {}).get("linked_amount")),
            "review_note": _clean((existing or {}).get("review_note")),
            "reviewed_at": (existing or {}).get("reviewed_at"),
            "reviewed_by": (existing or {}).get("reviewed_by"),
            "created_at": (existing or {}).get("created_at") or _utcnow(),
            "updated_at": _utcnow(),
        }
        if existing and candidate["review_status"] == "approved" and abs(_safe_float(candidate.get("linked_amount")) - _safe_float(po.get("receipt_amount"))) > 0.01:
            candidate["review_status"] = "review_required"
            candidate["review_note"] = _clean(candidate.get("review_note")) or "Receipt amount changed after approval and requires a fresh governed allocation review."
        if existing:
            pass
        else:
            created += 1
        await db[COLL_BUDGET_ACTUALS].replace_one(
            {"project_number": project_number, "source_kind": "po_receipt", "source_record_id": source_record_id},
            candidate,
            upsert=True,
        )
    unresolved = await db[COLL_BUDGET_ACTUALS].count_documents({"project_number": project_number, "review_status": "review_required"})
    if unresolved:
        await _upsert_review_item(
            db,
            {
                "review_id": f"budget-review:actuals:{project_number}",
                "project_number": project_number,
                "status": "review_required",
                "priority": 65,
                "source_kind": "actual_cost_candidates",
                "source_record_id": project_number,
                "title": f"Actual cost review required for {project_number}",
                "reason": "Candidate vendor receipt rows were preserved for review without treating them as accounting truth.",
                "confidence": "human_required",
                "provenance": {"review_required_count": unresolved},
            },
        )
    else:
        await _mark_review_resolved(db, f"budget-review:actuals:{project_number}", actor=actor, resolution_note="No unresolved actual-cost candidates remain.")
    return {"created": created, "review_required": unresolved}


async def activate_budget_import_session(db, project_number: str, import_id: str, *, actor: Dict[str, Any]) -> Dict[str, Any]:
    await ensure_project_budget_foundation(db)
    session = await db[COLL_BUDGET_IMPORTS].find_one({"project_number": project_number, "import_id": import_id}, {"_id": 0})
    if not session:
        raise LookupError("budget_import_not_found")
    rows = [
        row
        async for row in db[COLL_BUDGET_IMPORT_ROWS].find({"project_number": project_number, "import_id": import_id}, {"_id": 0}).sort([("row_number", 1)])
    ]
    if not rows:
        raise ValueError("budget_import_empty")
    pending = [row for row in rows if row.get("review_status") not in {"approved", "rejected"}]
    if pending:
        raise ValueError("budget_import_review_incomplete")
    approved_rows = [row for row in rows if row.get("review_status") == "approved"]
    if not approved_rows:
        raise ValueError("budget_import_has_no_approved_rows")
    current_active = await _get_latest_active_version(db, project_number)
    if session.get("target_version_stage") == "original_approved_budget":
        existing_original = await db[COLL_BUDGET_VERSIONS].find_one(
            {"project_number": project_number, "stage": "original_approved_budget", "status": {"$in": ["active", "superseded"]}},
            {"_id": 0},
        )
        if existing_original:
            raise ValueError("original_budget_already_locked")
    job = await _load_job(db, project_number)
    version = _budget_version_doc(
        project_number,
        job.get("project_name") or job.get("name") or project_number,
        {"stage": session.get("target_version_stage"), "version_name": session.get("version_name")},
        parent_version_id=(current_active or {}).get("version_id") or "",
        actor=actor,
        source_import_id=import_id,
    )
    if current_active and session.get("target_version_stage") == "current_approved_budget":
        await db[COLL_BUDGET_VERSIONS].update_many(
            {"project_number": project_number, "status": "active", "version_id": {"$ne": version["version_id"]}},
            {"$set": {"status": "superseded", "updated_at": _utcnow(), "updated_by": _actor_label(actor)}},
        )
    await db[COLL_BUDGET_VERSIONS].replace_one({"project_number": project_number, "version_id": version["version_id"]}, version, upsert=True)
    totals = deepcopy(version.get("totals") or {})
    line_docs = []
    work_types = await list_enterprise_work_types(db, include_archived=False)
    work_type_ids = {row.get("work_type_id") for row in work_types}
    for row in approved_rows:
        selected = row.get("selected") or {}
        if _clean(selected.get("enterprise_work_type_id")) and selected.get("enterprise_work_type_id") not in work_type_ids:
            raise ValueError("budget_import_selected_work_type_not_found")
        pay_item = await upsert_project_pay_item(
            db,
            project_number,
            {
                "customer_pay_item_number": selected.get("customer_pay_item_number"),
                "description": selected.get("description"),
                "unit": selected.get("unit"),
                "contract_quantity": selected.get("quantity"),
                "contract_unit_price": selected.get("unit_price"),
                "contract_value": selected.get("budget_amount"),
                "phase_id": selected.get("phase_id"),
                "work_package_id": selected.get("work_package_id"),
                "schedule_activity_id": selected.get("schedule_activity_id"),
                "schedule_activity_name": selected.get("schedule_activity_name"),
                "source": "budget_import_human_approved",
                "source_record": {"import_id": import_id, "row_id": row.get("row_id")},
                "provenance": {"owner": "project_pay_item_registry", "import_id": import_id, "pm_approved": True},
                "confidence": "human_confirmed",
                "status": "active",
            },
            actor=actor,
        )
        if _clean(selected.get("enterprise_work_type_id")):
            await upsert_project_mapping(
                db,
                project_number,
                {
                    "pay_item_id": pay_item.get("pay_item_id"),
                    "primary_work_type_id": selected.get("enterprise_work_type_id"),
                    "status": "approved",
                    "source": "budget_import_human_approved",
                    "confidence": "human_confirmed",
                    "explanation": _clean(selected.get("review_note")) or "PM approved mapping during governed budget activation.",
                },
                actor=actor,
            )
        line_doc = _budget_line_doc(
            project_number,
            version["version_id"],
            selected,
            pay_item_id=pay_item.get("pay_item_id") or "",
            actor=actor,
            source_import_id=import_id,
            source_row_id=row.get("row_id") or uuid4().hex,
            source_hash=session.get("file_hash") or "",
        )
        line_docs.append(line_doc)
        totals["budget_amount"] = round(_safe_float(totals.get("budget_amount")) + _safe_float(line_doc.get("budget_amount")), 4)
        totals["forecast_amount"] = round(_safe_float(totals.get("forecast_amount")) + _safe_float(line_doc.get("forecast_amount")), 4)
        totals["remaining_amount"] = round(_safe_float(totals.get("remaining_amount")) + _safe_float(line_doc.get("remaining_amount")), 4)
        if line_doc.get("line_kind") == "allowance":
            totals["allowance_amount"] = round(_safe_float(totals.get("allowance_amount")) + _safe_float(line_doc.get("budget_amount")), 4)
        elif line_doc.get("line_kind") == "contingency":
            totals["contingency_amount"] = round(_safe_float(totals.get("contingency_amount")) + _safe_float(line_doc.get("budget_amount")), 4)
        elif line_doc.get("line_kind") == "management_reserve":
            totals["management_reserve_amount"] = round(_safe_float(totals.get("management_reserve_amount")) + _safe_float(line_doc.get("budget_amount")), 4)
    if line_docs:
        await db[COLL_BUDGET_LINES].insert_many(line_docs)
    version["totals"] = totals
    version["updated_at"] = _utcnow()
    version["updated_by"] = _actor_label(actor)
    await db[COLL_BUDGET_VERSIONS].replace_one({"project_number": project_number, "version_id": version["version_id"]}, version, upsert=True)
    for row in approved_rows:
        await db[COLL_BUDGET_IMPORT_ROWS].update_one(
            {"row_id": row["row_id"]},
            {
                "$set": {
                    "review_status": "activated",
                    "activated_version_id": version["version_id"],
                    "activated_at": _utcnow(),
                    "updated_at": _utcnow(),
                    "updated_by": _actor_label(actor),
                }
            },
        )
        await _mark_review_resolved(db, f"budget-review:row:{row['row_id']}", actor=actor, resolution_note="Import row activated into governed budget version.")
    await db[COLL_BUDGET_IMPORTS].update_one(
        {"import_id": import_id},
        {
            "$set": {
                "status": "activated",
                "activated_version_id": version["version_id"],
                "activated_at": _utcnow(),
                "activated_by": _actor_label(actor),
                "updated_at": _utcnow(),
                "updated_by": _actor_label(actor),
            }
        },
    )
    await _mark_review_resolved(db, f"budget-review:import:{import_id}", actor=actor, resolution_note="Import session was activated after PM approval.")
    await _sync_commitment_candidates_for_project(db, project_number, actor=actor)
    await _sync_actual_cost_candidates_for_project(db, project_number, actor=actor)
    await _refresh_import_session_counts(db, import_id)
    await _write_audit(db, "budget_version_activated", actor, "budget_version", version["version_id"], version, metadata={"import_id": import_id, "line_count": len(line_docs)})
    return {"version": _sanitize(version), "line_count": len(line_docs)}


async def get_project_budget_overview(db, project_number: str) -> Dict[str, Any]:
    await ensure_project_budget_foundation(db)
    job = await _load_job(db, project_number)
    backfill = await db[COLL_BUDGET_RUNS].find_one({"run_type": "wp18c3_backfill"}, {"_id": 0})
    versions = await list_project_budget_versions(db, project_number)
    active_version = next((row for row in versions if row.get("status") == "active"), None)
    lines = await list_project_budget_lines(db, project_number, version_id=active_version["version_id"]) if active_version else []
    imports = await list_budget_import_sessions(db, project_number)
    await _sync_commitment_candidates_for_project(db, project_number)
    await _sync_actual_cost_candidates_for_project(db, project_number)
    review_queue = await list_budget_review_queue(db, project_number=project_number)
    commitment_candidates = [
        _sanitize(row)
        async for row in db[COLL_BUDGET_COMMITMENTS].find({"project_number": project_number}, {"_id": 0}).sort([("created_at", -1)]).limit(25)
    ]
    actual_candidates = [
        _sanitize(row)
        async for row in db[COLL_BUDGET_ACTUALS].find({"project_number": project_number}, {"_id": 0}).sort([("created_at", -1)]).limit(25)
    ]
    work_ledger_count = await db.project_controls_work_ledger.count_documents({"project_number": project_number})
    approved_pos = await db.po_requests.count_documents({"project_number": project_number, "status": {"$in": ["Approved", "Pending Receipt", "Overdue Receipt", "Receipt Uploaded", "Closed"]}})
    return {
        "project": {
            "project_number": project_number,
            "project_name": job.get("project_name") or job.get("name") or project_number,
            "pm_email": job.get("pm_email") or "",
            "co_pm_emails": job.get("co_pm_emails") or [],
        },
        "authority_boundaries": {
            "customer_pay_item_truth": "project_pay_item_registry",
            "enterprise_work_type_truth": "enterprise_work_type_registry",
            "budget_version_truth": COLL_BUDGET_VERSIONS,
            "budget_line_truth": COLL_BUDGET_LINES,
            "commitment_truth": "po_requests",
            "actual_cost_truth": "external_accounting_or_governed_receipt_review",
            "operational_work_truth": "project_controls_work_ledger",
            "ai_role": "advisory_only",
        },
        "counts": {
            "versions": len(versions),
            "active_lines": len(lines),
            "imports": len(imports),
            "review_queue_open": sum(1 for row in review_queue if row.get("status") != "resolved"),
            "commitment_candidates": len(commitment_candidates),
            "actual_cost_candidates": len(actual_candidates),
            "approved_po_requests": approved_pos,
            "work_ledger_rows": work_ledger_count,
        },
        "active_version": active_version,
        "latest_versions": versions[:8],
        "active_lines_preview": lines[:20],
        "imports": imports[:10],
        "review_queue": review_queue[:20],
        "commitment_candidates": commitment_candidates,
        "actual_cost_candidates": actual_candidates,
        "event_contracts": BUDGET_EVENT_CONTRACTS,
        "backfill": _sanitize(backfill or {"run_type": "wp18c3_backfill", "status": "pending_manual_run"}),
    }


async def get_admin_project_budget_overview(db, project_number: str = "") -> Dict[str, Any]:
    await ensure_project_budget_foundation(db)
    query = {"project_number": project_number} if project_number else {}
    backfill = await db[COLL_BUDGET_RUNS].find_one({"run_type": "wp18c3_backfill"}, {"_id": 0})
    versions = [_sanitize(row) async for row in db[COLL_BUDGET_VERSIONS].find(query, {"_id": 0}).sort([("activated_at", -1), ("created_at", -1)]).limit(100)]
    imports = [_sanitize(row) async for row in db[COLL_BUDGET_IMPORTS].find(query, {"_id": 0}).sort([("imported_at", -1)]).limit(100)]
    review_queue = await list_budget_review_queue(db, project_number=project_number)
    return {
        "summary": {
            "projects_with_versions": len({row.get("project_number") for row in versions if row.get("project_number")}),
            "budget_versions": await db[COLL_BUDGET_VERSIONS].count_documents(query),
            "budget_lines": await db[COLL_BUDGET_LINES].count_documents(query),
            "imports": await db[COLL_BUDGET_IMPORTS].count_documents(query),
            "review_queue_open": sum(1 for row in review_queue if row.get("status") != "resolved"),
            "commitment_candidates": await db[COLL_BUDGET_COMMITMENTS].count_documents(query),
            "actual_cost_candidates": await db[COLL_BUDGET_ACTUALS].count_documents(query),
        },
        "versions": versions,
        "imports": imports,
        "review_queue": review_queue[:100],
        "event_contracts": BUDGET_EVENT_CONTRACTS,
        "backfill": _sanitize(backfill or {"run_type": "wp18c3_backfill", "status": "pending_manual_run"}),
    }


async def run_project_budget_backfill(db, *, force: bool = False) -> Dict[str, Any]:
    await ensure_project_budget_foundation(db)
    last_run = await db[COLL_BUDGET_RUNS].find_one({"run_type": "wp18c3_backfill"}, {"_id": 0})
    if last_run and not force:
        return _sanitize(last_run)
    review_opened = 0
    commitment_candidates = 0
    actual_candidates = 0
    project_numbers = {
        _clean(row.get("project_number"))
        async for row in db.jobs_master.find({"project_number": {"$ne": ""}}, {"_id": 0, "project_number": 1})
    }
    project_numbers.update(_clean(row) for row in await db.po_requests.distinct("project_number", {"project_number": {"$ne": ""}}))
    project_numbers.update(_clean(row) for row in await db.project_pay_item_registry.distinct("project_number", {"project_number": {"$ne": ""}}))
    for project_number in sorted(project_numbers):
        project_number = _clean(project_number)
        if not project_number:
            continue
        job = await db.jobs_master.find_one({"project_number": project_number}, {"_id": 0, "project_name": 1, "project_number": 1}) or {"project_number": project_number}
        pay_item_count = await db.project_pay_item_registry.count_documents({"project_number": project_number})
        version_count = await db[COLL_BUDGET_VERSIONS].count_documents({"project_number": project_number})
        if pay_item_count > 0 and version_count == 0:
            await _upsert_review_item(
                db,
                {
                    "review_id": f"budget-review:foundation:{project_number}",
                    "project_number": project_number,
                    "status": "review_required",
                    "priority": 88,
                    "source_kind": "budget_foundation_missing",
                    "source_record_id": project_number,
                    "title": f"Budget hierarchy not yet activated for {project_number}",
                    "reason": "Project has governed pay-item authority but no governed budget version yet. Source records were preserved without fabrication.",
                    "confidence": "human_required",
                    "provenance": {
                        "project_name": job.get("project_name") or project_number,
                        "pay_item_count": pay_item_count,
                    },
                },
            )
            review_opened += 1
        sync_commitments = await _sync_commitment_candidates_for_project(db, project_number)
        sync_actuals = await _sync_actual_cost_candidates_for_project(db, project_number)
        commitment_candidates += sync_commitments.get("created", 0)
        actual_candidates += sync_actuals.get("created", 0)
    report = {
        "run_type": "wp18c3_backfill",
        "run_id": f"wp18c3-backfill:{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}",
        "ran_at": _utcnow(),
        "force": force,
        "foundation_reviews_opened": review_opened,
        "commitment_candidates_created": commitment_candidates,
        "actual_cost_candidates_created": actual_candidates,
        "status": "completed",
    }
    await db[COLL_BUDGET_RUNS].replace_one({"run_type": "wp18c3_backfill"}, report, upsert=True)
    return _sanitize(report)


def _csv_payload(filename: str, header: List[str], rows: List[List[Any]]) -> Dict[str, Any]:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(header)
    writer.writerows(rows)
    return {"filename": filename, "content": buffer.getvalue()}


async def record_budget_distribution_event(
    db,
    *,
    project_number: str,
    actor: Dict[str, Any],
    export_kind: str,
    version_id: str = "",
    metadata: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    await ensure_project_budget_foundation(db)
    row = {
        "distribution_id": f"budget-distribution:{project_number}:{uuid4().hex[:10]}",
        "project_number": project_number,
        "version_id": _clean(version_id),
        "export_kind": _clean(export_kind),
        "actor": _sanitize(actor or {}),
        "metadata": _sanitize(metadata or {}),
        "created_at": _utcnow(),
    }
    await db[COLL_BUDGET_DISTRIBUTION].insert_one(row)
    return _sanitize(row)


async def export_budget_version_rows(db, project_number: str, version_id: str, *, actor: Dict[str, Any]) -> Dict[str, Any]:
    lines = await list_project_budget_lines(db, project_number, version_id=version_id)
    if not lines:
        raise LookupError("budget_version_lines_not_found")
    await record_budget_distribution_event(
        db,
        project_number=project_number,
        actor=actor,
        export_kind="budget_lines_csv",
        version_id=version_id,
        metadata={"line_count": len(lines)},
    )
    rows = [
        [
            line.get("budget_line_id"),
            line.get("line_kind"),
            line.get("customer_pay_item_number"),
            line.get("description"),
            line.get("enterprise_work_type_id"),
            line.get("project_cost_code"),
            line.get("phase_id"),
            line.get("work_package_id"),
            line.get("schedule_activity_id"),
            line.get("quantity"),
            line.get("unit"),
            line.get("unit_budget_amount"),
            line.get("budget_amount"),
            line.get("commitment_amount"),
            line.get("actual_cost_amount"),
            line.get("forecast_amount"),
            line.get("remaining_amount"),
        ]
        for line in lines
    ]
    return _csv_payload(
        f"{project_number}_budget_{_norm(version_id)}.csv",
        [
            "budget_line_id",
            "line_kind",
            "customer_pay_item_number",
            "description",
            "enterprise_work_type_id",
            "project_cost_code",
            "phase_id",
            "work_package_id",
            "schedule_activity_id",
            "quantity",
            "unit",
            "unit_budget_amount",
            "budget_amount",
            "commitment_amount",
            "actual_cost_amount",
            "forecast_amount",
            "remaining_amount",
        ],
        rows,
    )


async def export_budget_version_comparison(
    db,
    project_number: str,
    *,
    left_version_id: str,
    right_version_id: str,
    actor: Dict[str, Any],
) -> Dict[str, Any]:
    left_rows = await list_project_budget_lines(db, project_number, version_id=left_version_id)
    right_rows = await list_project_budget_lines(db, project_number, version_id=right_version_id)
    if not left_rows and not right_rows:
        raise LookupError("budget_comparison_not_found")
    await record_budget_distribution_event(
        db,
        project_number=project_number,
        actor=actor,
        export_kind="budget_comparison_csv",
        metadata={"left_version_id": left_version_id, "right_version_id": right_version_id},
    )
    left_map = {f"{row.get('customer_pay_item_number')}::{row.get('description')}": row for row in left_rows}
    right_map = {f"{row.get('customer_pay_item_number')}::{row.get('description')}": row for row in right_rows}
    keys = sorted(set(left_map) | set(right_map))
    rows = []
    for key in keys:
        left = left_map.get(key) or {}
        right = right_map.get(key) or {}
        left_amount = _safe_float(left.get("budget_amount"))
        right_amount = _safe_float(right.get("budget_amount"))
        rows.append(
            [
                key.split("::", 1)[0],
                key.split("::", 1)[1] if "::" in key else key,
                left.get("line_kind") or right.get("line_kind") or "direct_cost",
                left_amount,
                right_amount,
                round(right_amount - left_amount, 4),
                left.get("enterprise_work_type_id") or "",
                right.get("enterprise_work_type_id") or "",
            ]
        )
    return _csv_payload(
        f"{project_number}_budget_compare_{_norm(left_version_id)}_vs_{_norm(right_version_id)}.csv",
        [
            "customer_pay_item_number",
            "description",
            "line_kind",
            "left_budget_amount",
            "right_budget_amount",
            "delta_budget_amount",
            "left_work_type_id",
            "right_work_type_id",
        ],
        rows,
    )
